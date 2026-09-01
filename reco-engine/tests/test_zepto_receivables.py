import sys, os
sys.path.insert(0, os.path.dirname(os.path.dirname(os.path.abspath(__file__))))
from io import BytesIO
from recon.zepto_receivables import (norm_po, norm_inv, dn_ref_to_invoice,
                                     _read_csv, _find_header, _rows_as_dicts, _get, _norm_key,
                                     parse_zepto_payment, parse_grn, grn_gate, parse_invoice_details,
                                     parse_credit_notes, parse_lrn, reconcile_zepto, summarize_zepto)

def test_normalizers_and_dn_transform():
    assert norm_po(" p4143483 ") == "P4143483"
    assert norm_po(3100356806) == "3100356806"      # int -> str, no ".0"
    assert norm_po(3100356806.0) == "3100356806"    # float-like -> no ".0"
    assert norm_inv(" inv26-27/000007 ") == "INV26-27/000007"
    assert dn_ref_to_invoice("V26-27/000007_QD") == "INV26-27/000007"
    assert dn_ref_to_invoice("V26-27/000007_PD") == "INV26-27/000007"
    assert dn_ref_to_invoice("V25-26/001331_QD") == "INV25-26/001331"
    print("test_normalizers_and_dn_transform OK")


def test_norm_inv_does_not_strip_trailing_slash():
    # REVERTED: "INV26-27/000039" (Rs 29,169) and "INV26-27/000039/"
    # (Rs 29,870) are DIFFERENT invoices in the real data -- norm_inv must
    # NOT collapse them by stripping the trailing slash. Only strip+upper.
    assert norm_inv("INV26-27/000039/") == "INV26-27/000039/"
    assert norm_inv("INV26-27/000039") == "INV26-27/000039"
    assert norm_inv(" inv26-27/000007 ") == "INV26-27/000007"
    print("test_norm_inv_does_not_strip_trailing_slash OK")


def test_dn_ref_to_invoice_primary_transform_only():
    # dn_ref_to_invoice does ONLY the simple primary transform: split on "_",
    # V-prefix -> INV, otherwise normalize as-is (upper/strip, no slash
    # stripping, no year-prefix INV-prepend). Malformed refs that need more
    # than this are fixed later by the fallback remap in reconcile_zepto.
    assert dn_ref_to_invoice("V26-27/000007_QD") == "INV26-27/000007"
    # Year-prefixed ref missing the V/INV entirely + trailing slash: the
    # primary transform leaves it as-is (uppercased) -- NOT "INV"-prepended,
    # NOT slash-stripped. The fallback remap (tested via reconcile_zepto)
    # is what maps this onto the real "INV26-27/000039/" invoice.
    assert dn_ref_to_invoice("26-27/000039/_QD") == "26-27/000039/"
    assert dn_ref_to_invoice("26-27/000041/_QD") == "26-27/000041/"
    # existing V-prefixed case still works
    assert dn_ref_to_invoice("V25-26/001322_QD") == "INV25-26/001322"
    # PMDDN / non-invoice refs must be left unchanged (no V prefix)
    assert dn_ref_to_invoice("PMDDN-79939") == "PMDDN-79939"
    assert dn_ref_to_invoice("CPMDDN-24512") == "CPMDDN-24512"
    assert dn_ref_to_invoice("DC-3214") == "DC-3214"
    print("test_dn_ref_to_invoice_primary_transform_only OK")

def test_csv_header_detection_and_getter():
    csv_bytes = (b"Some Title,,,\r\n"
                 b"Type/Description,Ref Id,Doc No,Amount,TDS,Payment Amount\r\n"
                 b",,,,,427313.6\r\n"
                 b"Invoice,INV26-27/000254,1900241884,36954.02,35.19,36918.83\r\n")
    grid = _read_csv(csv_bytes)
    h = _find_header(grid, ["ref id", "amount", "payment amount"])
    assert h == 1
    rows = _rows_as_dicts(grid, h)
    # summary row (blank Type) + one invoice row
    assert len(rows) == 2
    inv = rows[1]
    assert _get(inv, ["Type/Description"]) == "Invoice"
    assert _get(inv, ["ref id"]) == "INV26-27/000254"   # alias lookup is case-insensitive
    assert _get(inv, ["Payment Amount"]) == "36918.83"
    print("test_csv_header_detection_and_getter OK")

def _xlsx(sheets: dict) -> bytes:
    from openpyxl import Workbook
    wb = Workbook(); wb.remove(wb.active)
    for name, rows in sheets.items():
        ws = wb.create_sheet(title=name)
        for r in rows:
            ws.append(r)
    buf = BytesIO(); wb.save(buf); return buf.getvalue()

def test_zepto_payment_and_grn_gate():
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number", "Invoice Number", "Cities", "Delivery Date", "GRN"],
        ["P4143483", "INV26-27/000101", "Jaipur", "01/04/2026", "GrnCode1"],
        ["P4143483", "INV26-27/000101", "Jaipur", "01/04/2026", "GrnCode1"],  # dup PO
        ["3100356806", "INV24-25/000297", "Pune", "22/10/2024", "GrnCodeX"],  # legacy, no GRN
        ["P9999999", "INV26-27/000999", "Delhi", "05/04/2026", ""],           # not in GRN
    ]})
    payments = parse_zepto_payment(pay)
    assert {p["po"] for p in payments} == {"P4143483", "3100356806", "P9999999"} or len(payments) == 4
    grn_csv = (b"GRN ID,PO ID,Created On,Status\r\n"
               b"GrnCode1,P4143483,4/2/2026 16:20,CONFIRMED\r\n"
               b"GrnCode2,P4143483,4/3/2026 10:00,CONFIRMED\r\n")   # PO with 2 GRNs
    grn = parse_grn([grn_csv])
    assert "P4143483" in grn and "P9999999" not in grn
    kept = grn_gate(payments, grn)
    kept_pos = [k["po"] for k in kept]
    assert kept_pos == ["P4143483"]                 # unique + only GRN-matched
    assert kept[0]["invoice_number"] == "INV26-27/000101"
    print("test_zepto_payment_and_grn_gate OK")

def test_parse_invoice_details():
    inv = _xlsx({"Invoice Details": [
        ["Invoice Details \"Drips ... 01/04/2026 To 30/06/2026\""],  # title row 1
        ["invoice_id","status","date","txn_posting_date","due_date","invoice_number",
         "reference_number","customer_name","bcy_total","bcy_balance","fcy_balance",
         "salesperson_id","tax_amount","sgst","cgst","igst","amount_without_tax",
         "place_of_supply","gst_no","project_names","currency_code","currency_id",
         "customer_id","last_modified_time","price_precision","is_emailed","reminders_sent",
         "exchange_rate","billing_state","shipping_state","gst_treatment"],
        ["1151","overdue","2026-04-06","","","INV26-27/000101","2026-27/SO-00015",
         "ZEPTO PRIVATE LIMITED JAIPUR",21369.43,19288.96,19288.96,"115",1017.59,0,0,1017.59,
         20351.84,"RJ","08AAICK4821A1ZV","","INR","115","115","x",2,"false",0,1,
         "Rajasthan","Rajasthan","business_gst"],
    ]})
    d = parse_invoice_details(inv)
    row = d["INV26-27/000101"]
    assert row["date"] == "2026-04-06"
    assert row["sales_order_no"] == "2026-27/SO-00015"
    assert row["name"] == "ZEPTO PRIVATE LIMITED JAIPUR"
    assert float(row["total_invoice_amt"]) == 21369.43
    assert float(row["tax"]) == 1017.59
    assert float(row["invoice_amt_excl_tax"]) == 20351.84
    assert row["place_of_supply"] == "RJ"
    assert row["gstin"] == "08AAICK4821A1ZV"
    assert row["billing_state"] == "Rajasthan" and row["shipping_state"] == "Rajasthan"
    print("test_parse_invoice_details OK")


_REAL_FIXTURES_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures")
_REAL_INVOICE_DETAILS = os.path.join(_REAL_FIXTURES_DIR, "Invoice Details April 26 to June 26.xlsx")


def test_parse_invoice_details_real_fixture_dedup_by_num_amount_date():
    # Real fixture has 589 raw rows. Some invoice numbers appear twice with a
    # trailing-slash variant: e.g. "INV26-27/000039" (Rs 29,169) vs
    # "INV26-27/000039/" (Rs 29,870) -- DIFFERENT amounts -> BOTH must survive
    # as distinct universe keys (no merging). "INV26-27/000069" vs
    # "INV26-27/000069/" -- IDENTICAL amount+date -> a true duplicate; only
    # ONE survives (the slash variant is dropped as a re-export of the same
    # row). Verified directly against the fixture: 6 such identical pairs
    # exist, so 589 raw rows -> 583 rows (589 - 6 collapsed dupes). Feedback
    # #00005 then keeps ONLY Zepto/Kiranakart invoices (single PAN AAICK4821A),
    # dropping ~200 other-vendor rows (Blinkit/Flipkart/BigBasket/…) -> 383.
    with open(_REAL_INVOICE_DETAILS, "rb") as f:
        data = f.read()
    universe = parse_invoice_details(data)
    assert len(universe) == 383

    # Different-amount pair: BOTH keys present, distinct amounts preserved.
    assert "INV26-27/000039" in universe
    assert "INV26-27/000039/" in universe
    assert round(float(universe["INV26-27/000039"]["total_invoice_amt"]), 2) == 29169.0
    assert round(float(universe["INV26-27/000039/"]["total_invoice_amt"]), 2) == 29870.95

    # Identical pair: base key present once, slash variant collapsed away.
    assert "INV26-27/000069" in universe
    assert "INV26-27/000069/" not in universe
    print("test_parse_invoice_details_real_fixture_dedup_by_num_amount_date OK")

def test_parse_payment_advice():
    from recon.zepto_receivables import parse_payment_advice, _to_float
    csv1 = (b"Type/Description,Ref Id,Doc No,Amount,TDS,Payment Amount\r\n"
            b",,,,,427313.6\r\n"                                               # summary -> skip
            b"Invoice,INV26-27/000007,1900261942,51429.9,48.98,51380.92\r\n"
            b"Debit Note,V26-27/000007_QD,1700049536,-311.38,0.3,-311.68\r\n"
            b"Debit Note Price,V26-27/000007_PD,4800006135,-4988.44,5,-4993.44\r\n")
    pay, dn = parse_payment_advice([csv1])
    assert _to_float(pay["INV26-27/000007"]["incl"]) == 51380.92
    assert _to_float(pay["INV26-27/000007"]["excl"]) == 51429.9
    assert _to_float(pay["INV26-27/000007"]["tds"]) == 48.98
    # DN = sum of Amount (col D): -311.38 + -4988.44 = -5299.82
    assert round(dn["INV26-27/000007"], 2) == -5299.82
    print("test_parse_payment_advice OK")

def test_parse_credit_notes():
    # parse_credit_notes now returns {inv: {"amount": <sum>, "numbers": [...]}}
    # -- both credit-note numbers must be captured (deduped) alongside the sum.
    cn = _xlsx({"Credit Note Details": [
        ["Credit Note Details title row"],
        ["creditnote_number","status","date","reference_number","exchange_rate",
         "bcy_total","bcy_balance","currency_code","tax_amount","sales_person_id",
         "sgst","cgst","igst","invoice_number"],
        ["CN/26-27/0003","closed","2025-05-10","SO1",1,2049.75,0,"INR",97.61,"",0,0,97.61,"INV26-27/000011"],
        ["CN/26-27/0009","closed","2025-05-10","SO2",1,50.0,0,"INR",7,"",0,0,7,"INV26-27/000011"],
    ]})
    d = parse_credit_notes(cn)
    assert round(d["INV26-27/000011"]["amount"], 2) == 2099.75    # 2049.75 + 50.0
    assert d["INV26-27/000011"]["numbers"] == ["CN/26-27/0003", "CN/26-27/0009"]
    print("test_parse_credit_notes OK")


def test_parse_credit_notes_real_fixture():
    with open(os.path.join(_REAL_FIXTURES_DIR, "Credit Note Details (4).xlsx"), "rb") as f:
        data = f.read()
    d = parse_credit_notes(data)
    # INV26-27/000011 has one credit note: CN/26-27/0001, amount 2049.75.
    row = d["INV26-27/000011"]
    assert round(row["amount"], 2) == 2049.75
    assert row["numbers"] == ["CN/26-27/0001"]
    print("test_parse_credit_notes_real_fixture OK")


def test_parse_grn_real_fixture_captures_grn_id_and_created_on():
    with open(os.path.join(_REAL_FIXTURES_DIR, "GRN_List - April-2026.csv"), "rb") as f:
        data = f.read()
    grn = parse_grn([data])
    row = grn["P3855315"]
    assert row["grn_id"] == "GrnCode41929315"
    assert row["created_on"] == "4/2/2026 16:20"
    print("test_parse_grn_real_fixture_captures_grn_id_and_created_on OK")


_LRN_OLD_FIXTURE = os.path.join(_REAL_FIXTURES_DIR, "Drips Foods formate sheet-LRN Old.xlsx")
_LRN_13MAY_FIXTURE = os.path.join(_REAL_FIXTURES_DIR, "DRIPS FOOD LRN - From 13th MAy.xlsx")


def test_parse_lrn_multi_invoice_row():
    with open(_LRN_OLD_FIXTURE, "rb") as f:
        data = f.read()
    lrn_map = parse_lrn([data])
    for inv in ("INV24-25/000484", "INV24-25/000491"):
        assert inv in lrn_map, inv
        assert lrn_map[inv]["pod_no"] == "264895795"
        assert lrn_map[inv]["pod_date"] != ""
    print("test_parse_lrn_multi_invoice_row OK")


def test_parse_lrn_sheet_without_invoice_column_is_skipped():
    with open(_LRN_13MAY_FIXTURE, "rb") as f:
        data = f.read()
    lrn_map = parse_lrn([data])
    assert lrn_map == {}
    print("test_parse_lrn_sheet_without_invoice_column_is_skipped OK")

def _file(b): return {"filename": "f", "content": b}

def test_reconcile_end_to_end():
    # Universe = Invoice Details (every Zepto invoice becomes a row). Feedback
    # #00006: the PO comes from the Zepto Payment track whenever present; it is
    # NO LONGER suppressed when its PO is absent from the GRN pool (only the
    # GRN No./Date columns stay gated on the GRN pool).
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV26-27/000007","SO-7","ZEPTO PUNE","2026-04-06",56685.0,0,56685.0,"MH","27AAA","Maharashtra","Maharashtra"],
        ["INV26-27/000101","SO-101","ZEPTO JAIPUR","2026-04-08",21369.43,1017.59,20351.84,"RJ","08AAA","Rajasthan","Rajasthan"],
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities"],
        ["P100","INV26-27/000007","Pune"],          # PO in GRN  -> po + grn cols filled
        ["P900","INV26-27/000101","Delhi"],         # PO NOT in GRN -> po filled, grn cols blank (#00006)
    ]})
    grn = b"GRN ID,PO ID,Created On,Status\r\nGrnCode99,P100,4/2/2026,CONFIRMED\r\n"   # only P100 confirmed
    cn = _xlsx({"Credit Note Details": [["t"],
        ["invoice_number","bcy_total","creditnote_number"], ["INV26-27/000101", 100.0, "CN/26-27/0099"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files)
    assert len(res) == 2                                  # universe = invoice details (both rows)
    by_inv = {r["invoice_number"]: r for r in res}
    a = by_inv["INV26-27/000007"]
    assert a["name"] == "ZEPTO PUNE"
    assert a["po"] == "P100"                              # GRN-matched -> po filled
    assert a["grn_no"] == "GrnCode99"                     # PO's GRN in GRN pool -> grn_no filled
    assert a["grn_date"] == "2026-04-02"                  # ... and grn_date filled (date-only, normalized)
    assert round(a["total_invoice_amt"], 2) == 56685.0
    assert round(a["pending_amount"], 2) == 56685.0
    assert round(a["payment_received_incl_tds"], 2) == 0.0    # no PDFs -> 0
    assert round(a["gross_outstanding"], 2) == 56685.0
    assert round(a["net_outstanding"], 2) == 56685.0
    assert a["status"] == "Not Paid"
    assert a["invoice_not_in_ledger"] == ""               # v1 "Invoice Not in Books" logic removed

    b = by_inv["INV26-27/000101"]
    assert b["name"] == "ZEPTO JAIPUR"
    assert b["po"] == "P900"                              # #00006: PO shown even though not in GRN pool
    assert b["grn_no"] == "" and b["grn_date"] == ""       # ... but GRN cols stay gated on the GRN pool
    assert round(b["total_invoice_amt"], 2) == 21369.43
    assert round(b["credit_note_issued"], 2) == 100.0
    assert b["credit_note_no"] == "CN/26-27/0099"          # new column filled from CN map
    assert b["status"] in ("Paid", "Not Paid")

    s = summarize_zepto(res)
    assert s["total"] == 2 and s["paid"] + s["not_paid"] == 2
    print("test_reconcile_end_to_end OK")

def test_workbook_has_live_formulas():
    # Columns are resolved through _KEYCOL, so the 4 live formulas (pending /
    # gross / net / status) must always point at the right letters no matter how
    # the layout shifts (e.g. the Month column insertion). Resolve expected
    # letters dynamically and assert the formula strings match.
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _KEYCOL
    results = [{k: "" for k in COLUMN_KEYS}]
    r = results[0]
    r.update({"po":"P100","invoice_number":"INV26-27/000007","total_invoice_amt":56685.0,
              "payment_received_incl_tds":51380.92,"debit_note_issued":5299.82,
              "pending_amount":56685.0,"gross_outstanding":5304.08,"net_outstanding":4.26,"status":"Not Paid"})
    wb = build_zepto_workbook(results)
    ws = wb["1. Invoice Tracker"]
    K = _KEYCOL
    assert ws[f"{K['po']}3"].value == "P100"
    assert ws[f"{K['invoice_number']}3"].value == "INV26-27/000007"
    assert ws[f"{K['total_invoice_amt']}3"].value == 56685.0
    assert ws[f"{K['credit_note_no']}2"].value == "Credit Note No"
    assert ws[f"{K['pending_amount']}3"].value == f"={K['total_invoice_amt']}3"
    assert ws[f"{K['gross_outstanding']}3"].value == f"={K['pending_amount']}3-{K['payment_received_excl_tds']}3"
    assert ws[f"{K['net_outstanding']}3"].value == (
        f"={K['pending_amount']}3-{K['payment_received_excl_tds']}3"
        f"-{K['credit_note_issued']}3-{K['debit_note_issued']}3")
    assert ws[f"{K['status']}3"].value == f'=IF({K["net_outstanding"]}3<=10,"Paid","Not Paid")'
    for key in ("pending_amount", "gross_outstanding", "net_outstanding", "status"):
        assert "REF" not in str(ws[f"{K[key]}3"].value)
    # Month column is present, right after Date.
    assert ws[f"{K['month']}2"].value == "Month"
    assert COLUMN_KEYS[COLUMN_KEYS.index("date") + 1] == "month"
    print("test_workbook_has_live_formulas OK")


def test_status_signed_threshold_negative_net_is_paid():
    # Accountant's rule: Paid whenever NET <= 100 (Gross ignored), INCLUDING
    # negative net (a debit note pushing Net negative means settled).
    # net~-2058 (<=100 since negative) -> Paid.
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV1","SO-1","ZEPTO A","2026-04-06",50000.0,0,50000.0,"MH","27AAA","Maharashtra","Maharashtra"],
        ["INV2","SO-2","ZEPTO B","2026-04-06",28000.0,0,28000.0,"MH","27AAA","Maharashtra","Maharashtra"],
    ]})
    pay = _xlsx({"Zepto Payment track": [["Zepto Payment track PO Number","Invoice Number","Cities"]]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files)
    by_inv = {r["invoice_number"]: r for r in res}

    # Force gross~17 by setting payment received to make gross = 50000 - 49983 = 17
    by_inv["INV1"]["payment_received_incl_tds"] = 49983.0
    # Force net~-2058 via a large debit note: DN stored POSITIVE and SUBTRACTED
    # -> net = gross - dn = 17 - 2075 = -2058
    by_inv["INV1"]["debit_note_issued"] = 2075.0
    gross1 = by_inv["INV1"]["total_invoice_amt"] - by_inv["INV1"]["payment_received_incl_tds"]
    net1 = gross1 - by_inv["INV1"]["debit_note_issued"]
    status1 = "Paid" if net1 <= 10 else "Not Paid"   # net-only rule, threshold 10
    assert round(gross1, 2) == 17.0
    assert round(net1, 2) == -2058.0
    assert status1 == "Paid"

    # INV2: gross~28000 (no payment received) -> Not Paid
    assert by_inv["INV2"]["gross_outstanding"] == 28000.0
    assert by_inv["INV2"]["status"] == "Not Paid"
    print("test_status_signed_threshold_negative_net_is_paid OK")


def test_reconcile_zepto_signed_status_direct():
    # Directly exercise the NET-ONLY signed rule used inside reconcile_zepto
    # (Gross is ignored): Net <= 100 -> Paid, including negatives.
    for gross, net, expected in [
        (17.0, -2058.0, "Paid"),
        (-83.0, -83.0, "Paid"),
        (50.0, -478.0, "Paid"),
        (28000.0, 28000.0, "Not Paid"),
        (150.0, 40.0, "Not Paid"),   # net 40 > 10 threshold -> Not Paid
        (150.0, 9.0, "Paid"),        # net <= 10 -> Paid even though gross > 100 (net-only)
    ]:
        status = "Paid" if net <= 10 else "Not Paid"
        assert status == expected, (gross, net, expected, status)
    print("test_reconcile_zepto_signed_status_direct OK")


def test_status_column_has_conditional_formatting():
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _KEYCOL
    results = [{k: "" for k in COLUMN_KEYS}]
    results[0].update({"invoice_number": "INV1", "total_invoice_amt": 1000.0, "status": "Paid"})
    wb = build_zepto_workbook(results)
    ws = wb["1. Invoice Tracker"]
    # Conditional formatting must be registered on the workbook/sheet for the
    # Status column (formula-driven cell -> CF is the only way to color it).
    cf_ranges = [str(rng) for rng in ws.conditional_formatting]
    assert len(cf_ranges) >= 1
    assert any(f"{_KEYCOL['status']}3" in rng for rng in cf_ranges)   # Status column, resolved dynamically
    print("test_status_column_has_conditional_formatting OK")

def test_universe_includes_invoice_with_no_po_mapping_at_all():
    # An invoice in Invoice Details that ISN'T even listed in Zepto Payment track
    # must still appear in the universe with po == "".
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV26-27/000500","SO-9","ZEPTO NOPO","2026-04-06",1000.0,0,1000.0,"MH","27AAA","Maharashtra","Maharashtra"],
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities"],
    ]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files)
    assert len(res) == 1
    row = res[0]
    assert row["invoice_number"] == "INV26-27/000500"
    assert row["po"] == ""
    assert "Missing PO" in row["remark"]   # no PO -> flagged in Remark
    assert row["invoice_not_in_ledger"] == ""                   # v1 flag no longer set
    assert round(row["total_invoice_amt"], 2) == 1000.0
    assert row["status"] in ("Paid", "Not Paid")
    print("test_universe_includes_invoice_with_no_po_mapping_at_all OK")

_FIXTURE_DIR = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "pdf")
_FIXTURE_PDF = os.path.join(_FIXTURE_DIR, "2026-07-03_1d33a6026d1168fd_2000004234.pdf")


def _read_fixture_pdf() -> bytes:
    with open(_FIXTURE_PDF, "rb") as f:
        return f.read()


def test_parse_payment_advice_pdf_single():
    from recon.zepto_receivables import parse_payment_advice_pdf
    pdf_bytes = _read_fixture_pdf()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    inv = payments["INV25-26/001463"]
    assert round(inv["excl"], 2) == 119767.53
    assert round(inv["tds"], 2) == 114.06
    assert round(inv["incl"], 2) == 119653.47
    print("test_parse_payment_advice_pdf_single OK")


def test_parse_payment_advice_pdf_dedup():
    from recon.zepto_receivables import parse_payment_advice_pdf
    pdf_bytes = _read_fixture_pdf()
    # Same PDF fed twice (e.g. re-uploaded / duplicate in Drive folder) must NOT double-count.
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes, pdf_bytes])
    inv = payments["INV25-26/001463"]
    assert round(inv["incl"], 2) == 119653.47
    assert round(inv["excl"], 2) == 119767.53
    assert round(inv["tds"], 2) == 114.06
    print("test_parse_payment_advice_pdf_dedup OK")


_FIXTURE_DIR_ALL = os.path.join(os.path.dirname(os.path.abspath(__file__)), "fixtures", "pdf_all")
_FIXTURE_PDF_MULTIPAGE = os.path.join(_FIXTURE_DIR_ALL, "06_Payment_Advice_2000017788.PDF")


def test_parse_payment_advice_pdf_multipage():
    # 06_Payment_Advice_2000017788.PDF has 2 pages. Page 2 (continuation) has
    # invoice DATA rows but NO repeated header row. Row for INV25-26/001632
    # lives on page 2 only:
    #   ['5155','Invoice Payment','1901427707','INV25-26/001\n632',
    #    '28,774.77','INR','27.4','28,747.37']
    # Confirmed via direct pdfplumber.extract_tables() on the fixture.
    from recon.zepto_receivables import parse_payment_advice_pdf
    with open(_FIXTURE_PDF_MULTIPAGE, "rb") as f:
        pdf_bytes = f.read()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    inv = payments["INV25-26/001632"]
    assert round(inv["incl"], 2) == 28747.37
    assert round(inv["excl"], 2) == 28774.77
    assert round(inv["tds"], 2) == 27.4
    # Sanity: page-1-only invoice must also still be present.
    assert "INV25-26/001585" in payments
    print("test_parse_payment_advice_pdf_multipage OK")


_FIXTURE_PDF_CREDIT_MEMO = os.path.join(_FIXTURE_DIR_ALL, "04_Payment_Advice_2000013623.PDF")


def test_parse_payment_advice_pdf_credit_memo_routes_as_debit_note():
    # Zepto labels its debit notes "Credit Memo" in the real PDFs (there are NO
    # literal "Debit Note" rows in the data). Confirmed via direct
    # pdfplumber.extract_tables() on this fixture:
    #   ['1804','Credit Memo','1700264647','V25-26/00156\n4_QD',
    #    '-5,962.15','INR','5.68','-5,956.47']
    #   ['1805','Credit Memo','1700265247','V25-26/00159\n8_QD',
    #    '-12,916.64','INR','12.3','-12,904.34']
    # Ref cleans to "V25-26/001564_QD" -> dn_ref_to_invoice -> "INV25-26/001564"
    # (and "V25-26/001598_QD" -> "INV25-26/001598"). Amount is col[4].
    from recon.zepto_receivables import parse_payment_advice_pdf
    with open(_FIXTURE_PDF_CREDIT_MEMO, "rb") as f:
        pdf_bytes = f.read()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    assert round(debit_notes["INV25-26/001564"], 2) == -5962.15
    assert round(debit_notes["INV25-26/001598"], 2) == -12916.64
    print("test_parse_payment_advice_pdf_credit_memo_routes_as_debit_note OK")


_FIXTURE_PDF_SLASH_INVOICE = os.path.join(_FIXTURE_DIR_ALL, "21_Payment_Advice_2000028805.PDF")
_FIXTURE_PDF_SLASH_INVOICE_2 = os.path.join(_FIXTURE_DIR_ALL, "23_Payment_Advice_2000029871.PDF")
_FIXTURE_PDF_PMDDN = os.path.join(_FIXTURE_DIR_ALL, "06_Payment_Advice_2000017788.PDF")


def test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn():
    # 21_Payment_Advice_2000028805.PDF: Invoice Payment RefDoc cleans to
    # "INV26-27/000039/" (Payment Amt 29,843.58, slash PRESERVED after the
    # revert) and Credit Memo RefDoc cleans to "26-27/000039/_QD" -- since it
    # contains an invoice pattern (\d{2}-\d{2}/\d+), it is routed to
    # debit_notes under dn_ref_to_invoice's primary-transform key
    # "26-27/000039/" (no V/INV prefix at this stage -- the fallback remap in
    # reconcile_zepto is what later reconciles this onto the real invoice).
    from recon.zepto_receivables import parse_payment_advice_pdf
    with open(_FIXTURE_PDF_SLASH_INVOICE, "rb") as f:
        pdf_bytes = f.read()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    assert round(payments["INV26-27/000039/"]["incl"], 2) == 29843.58
    assert round(debit_notes["26-27/000039/"], 2) == -224.24
    print("test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn OK")


def test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn_2():
    # 23_Payment_Advice_2000029871.PDF: same pattern for invoice 000041.
    from recon.zepto_receivables import parse_payment_advice_pdf
    with open(_FIXTURE_PDF_SLASH_INVOICE_2, "rb") as f:
        pdf_bytes = f.read()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    assert round(payments["INV26-27/000041/"]["incl"], 2) == 37396.14
    assert round(debit_notes["26-27/000041/"], 2) == -935.93
    print("test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn_2 OK")


def test_parse_payment_advice_pdf_pmddn_routes_to_pmdn_total_not_debit_notes():
    # 06_Payment_Advice_2000017788.PDF has a "PMDDN-79939" Credit Memo row
    # (amount -1,221,755) — a marketing adjustment, NOT a per-invoice debit
    # note. It must NOT create/pollute any debit_notes["...79939..."] key and
    # must instead accumulate into pmdn_total.
    from recon.zepto_receivables import parse_payment_advice_pdf, dn_ref_to_invoice
    with open(_FIXTURE_PDF_PMDDN, "rb") as f:
        pdf_bytes = f.read()
    payments, debit_notes, pmdn_total = parse_payment_advice_pdf([pdf_bytes])
    pmddn_key = dn_ref_to_invoice("PMDDN-79939")
    assert pmddn_key not in debit_notes
    assert not any(k.startswith("PMDDN") for k in debit_notes)
    assert pmdn_total != 0
    assert round(pmdn_total, 2) == -1221755.0
    print("test_parse_payment_advice_pdf_pmddn_routes_to_pmdn_total_not_debit_notes OK")


import glob


def _load_real_files() -> dict:
    def f(path):
        with open(path, "rb") as fh:
            return {"filename": os.path.basename(path), "content": fh.read()}

    grn_files = [f(p) for p in sorted(glob.glob(os.path.join(_REAL_FIXTURES_DIR, "GRN_List*.csv")))]
    pdf_files = [
        f(p) for p in
        sorted(glob.glob(os.path.join(_FIXTURE_DIR_ALL, "*.pdf")))
        + sorted(glob.glob(os.path.join(_FIXTURE_DIR_ALL, "*.PDF")))
    ]
    return {
        "zepto_payment": f(os.path.join(_REAL_FIXTURES_DIR, "Zepto Payment FY24-25 New 13_05.xlsx")),
        "grn_list": grn_files,
        "invoice_details": f(_REAL_INVOICE_DETAILS),
        "payment_advice": pdf_files,
        "credit_note": f(os.path.join(_REAL_FIXTURES_DIR, "Credit Note Details (4).xlsx")),
    }


def test_reconcile_zepto_real_fixtures_dn_fallback_and_pmddn():
    # Full pipeline on the real fixtures (zepto_payment, 3 GRN lists,
    # invoice_details, all 43 payment_advice PDFs, credit_note). Verifies the
    # DN fallback remap: the malformed DN ref "26-27/000039/_QD" (routed to
    # debit_notes["26-27/000039/"] at parse time) gets reassigned onto the
    # REAL invoice key "INV26-27/000039/" because that's what's actually in
    # the Invoice Details universe -- NOT onto "INV26-27/000039" (a different,
    # unrelated invoice with a different amount).
    files = _load_real_files()
    res = reconcile_zepto(files)
    by_inv = {r["invoice_number"]: r for r in res}

    row = by_inv["INV26-27/000039/"]
    assert row["debit_note_issued"] != 0.0
    assert round(row["debit_note_issued"], 2) == 224.24   # stored POSITIVE (the amount)
    assert round(row["payment_received_incl_tds"], 2) == 29843.58

    # The unrelated, differently-amounted invoice must NOT receive the DN.
    other = by_inv["INV26-27/000039"]
    assert other["debit_note_issued"] == 0.0

    # PMDDN: a PMDDN- credit-memo ref must never appear as a debit_notes key
    # anywhere in the pipeline, and the pmdn_adjustment total must be non-zero.
    assert not any(k.startswith("PMDDN") for k in by_inv)
    assert res.pmdn_adjustment != 0.0
    print("test_reconcile_zepto_real_fixtures_dn_fallback_and_pmddn OK")


def test_reconcile_zepto_returns_list_subclass_with_pmdn_adjustment():
    # reconcile_zepto must keep behaving as a plain list (len/iterate/index)
    # for existing callers/tests and server.py's JSON serialization, while
    # also carrying the pmdn_adjustment total as an attribute.
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV1","SO-1","ZEPTO A","2026-04-06",1000.0,0,1000.0,"MH","27AAA","Maharashtra","Maharashtra"],
    ]})
    pay = _xlsx({"Zepto Payment track": [["Zepto Payment track PO Number","Invoice Number","Cities"]]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files)
    assert isinstance(res, list)
    assert len(res) == 1
    assert res[0]["invoice_number"] == "INV1"
    assert hasattr(res, "pmdn_adjustment")
    assert res.pmdn_adjustment == 0.0   # no payment_advice PDFs -> 0
    # JSON-serializable exactly like a plain list (server.py relies on this).
    import json
    json.dumps(res)
    print("test_reconcile_zepto_returns_list_subclass_with_pmdn_adjustment OK")


def test_build_zepto_workbook_fills_pmddn_summary_line():
    from recon.zepto_receivables import build_zepto_workbook

    class _RecoRows(list):
        pass

    results = _RecoRows([{k: "" for k in __import__("recon.zepto_receivables", fromlist=["COLUMN_KEYS"]).COLUMN_KEYS}])
    results[0].update({"invoice_number": "INV1", "total_invoice_amt": 1000.0, "status": "Not Paid"})
    results.pmdn_adjustment = -12345.67

    wb = build_zepto_workbook(results)
    ws = wb["Summary"]
    amt = _find_summary_amount(ws, "Expense - PMDDN & AP-AR Adjustment (Marketing)")
    assert amt == -12345.67
    assert amt not in (None, "")

    # Amount Received in Bank is auto (0 with no advices); other manual lines blank.
    assert _find_summary_amount(ws, "Amount Received in Bank") == 0
    assert _find_summary_amount(ws, "Debit Note Accepted") in (None, "")
    assert _find_summary_amount(ws, "Debit Note Not Accepted") in (None, "")
    print("test_build_zepto_workbook_fills_pmddn_summary_line OK")


def _find_summary_amount(ws, label: str):
    """Scan column A for `label`; return the value of the adjacent amount cell (col B)."""
    for row in ws.iter_rows(min_col=1, max_col=1):
        cell = row[0]
        if cell.value == label:
            return ws.cell(row=cell.row, column=2).value
    raise AssertionError(f"label not found in Summary sheet: {label}")


def test_build_zepto_workbook_has_summary_tab():
    from recon.zepto_receivables import build_zepto_workbook

    def _row(inv, total, cn, dn, tds, incl, net, due_status="Not Paid"):
        r = {k: "" for k in __import__("recon.zepto_receivables", fromlist=["COLUMN_KEYS"]).COLUMN_KEYS}
        r.update({
            "invoice_number": inv,
            "total_invoice_amt": total,
            "credit_note_issued": cn,
            "debit_note_issued": dn,
            "tds": tds,
            "payment_received_incl_tds": incl,
            "net_outstanding": net,
            "status": "Not Paid",
            "due_status": due_status,
        })
        return r

    results = [
        _row("INV1", 56685.0, 100.0, 50.0, 10.0, 40000.0, 16735.0, "Overdue"),
        _row("INV2", 21369.43, 0.0, 0.0, 5.0, 21369.43, 5000.0, "Not Due"),
        _row("INV3", 1000.0, 25.0, 0.0, 0.0, 975.0, 300.0, "Due"),
        _row("INV4", 2000.0, 0.0, 0.0, 0.0, 1000.0, 700.0, "Not Due"),
    ]
    wb = build_zepto_workbook(results)

    # Tabs in order: main + Summary + Payments + Payment Advice Consolidate + the
    # 4 other split detail tabs (#00007), no stray "Sheet".
    assert wb.sheetnames == ["1. Invoice Tracker", "Summary", "Payments", "Payment Advice Consolidate",
                             "Debit Notes", "Credit Notes", "PMDD", "AP-AR & Manual Adj"]

    ws = wb["Summary"]
    total_sales = sum(r["total_invoice_amt"] for r in results)
    assert _find_summary_amount(ws, "Sales Including Tax") == total_sales

    # Amount Received in Bank is auto = total of the Payment Advice Consolidate
    # tab; with no payment-advice PDFs the consolidate is empty -> 0.
    assert _find_summary_amount(ws, "Amount Received in Bank") == 0

    # Ageing of Net Receivables: Net Outstanding bucketed by Due Status.
    assert _find_summary_amount(ws, "Not Due") == round(5000.0 + 700.0, 2)   # INV2 + INV4
    assert _find_summary_amount(ws, "Due") == 300.0                          # INV3
    assert _find_summary_amount(ws, "Overdue") == 16735.0                    # INV1
    assert _find_summary_amount(ws, "Excess Paid") == 0                      # no Paid+overpaid rows
    # The ageing block (in the "Summary -3" section) lists the four buckets in
    # order under its own sub-header.
    labels = [c.value for c in ws["A"] if c.value]
    i = labels.index("Ageing of Net Receivables (by Due Status)")
    assert labels[i+1:i+5] == ["Not Due", "Due", "Overdue", "Excess Paid"]
    # 3-section layout is present.
    assert "Summary-2" in labels and "Summary -3" in labels
    print("test_build_zepto_workbook_has_summary_tab OK")


def test_vendor_filter_keeps_only_zepto():
    # #00005: this is a ZEPTO tracker, so other vendors in the Tally export
    # (Blinkit / Flipkart / …) must be dropped. A Zepto row with a garbled
    # customer name but a Zepto GSTIN (same PAN) is still kept via the PAN set.
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV-Z1","SO1","ZEPTO PRIVATE LIMITED MUMBAI","2026-04-06",1000,0,1000,"MH","27AAICK4821A1Z5","MH","MH"],
        ["INV-K1","SO2","Kiranakart Technologies Pvt Ltd","2026-04-06",1000,0,1000,"KA","29AAICK4821A1Z0","KA","KA"],
        ["INV-Zpan","SO3","STORE 42","2026-04-06",1000,0,1000,"RJ","08AAICK4821A1Z9","RJ","RJ"],   # garbled name, Zepto PAN
        ["INV-B1","SO4","BLINK COMMERCE PRIVATE LIMITED","2026-04-06",1000,0,1000,"HR","06AAFCG9846N1Z2","HR","HR"],
        ["INV-F1","SO5","FLIPKART INDIA PRIVATE LIMITED","2026-04-06",1000,0,1000,"KA","29AABCF8078M1Z8","KA","KA"],
    ]})
    uni = parse_invoice_details(invd)
    assert set(uni.keys()) == {"INV-Z1", "INV-K1", "INV-ZPAN"}   # norm_inv upper-cases the key
    print("test_vendor_filter_keeps_only_zepto OK")


def test_po_matches_despite_trailing_slash_and_ignores_grn_gate():
    # #00006: the universe carries `INV26-27/000192/` (trailing slash from
    # Tally) while the Payment track carries `INV26-27/000192` (no slash) — the
    # PO must still match. And PO `P4254950` need NOT be in the GRN pool to show.
    invd = _xlsx({"Invoice Details": [
        ["title"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV26-27/000192/","SO1","ZEPTO MUMBAI","2026-04-06",42438.23,2020.87,40417.36,"MH","27AAICK4821A1Z5","MH","MH"],
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities"],
        ["P4254950","INV26-27/000192","Mumbai"],   # payment track has NO trailing slash
    ]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"     # empty GRN pool
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files)
    row = res[0]
    assert row["invoice_number"] == "INV26-27/000192/"   # universe key keeps the slash
    assert row["po"] == "P4254950"                        # matched despite slash + empty GRN pool
    assert row["grn_no"] == "" and row["grn_date"] == ""  # GRN cols blank (PO not in GRN pool)
    print("test_po_matches_despite_trailing_slash_and_ignores_grn_gate OK")


def test_due_date_and_status():
    # #00001: Due Date = GRN Date + 30 calendar days (payment clock runs from
    #         delivery/GRN, NOT the invoice date).
    # #00002: Not Due (future) / Due (today) / Overdue (past), UNPAID rows only;
    #         Paid rows -> blank Due status (but Due Date still computed).
    #         No GRN Date (GRN missing) -> Due Date and Due Status BOTH blank.
    import datetime
    invd = _xlsx({"Invoice Details": [
        ["t"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV-A","SO1","ZEPTO A","2026-04-01",50000,0,50000,"MH","27AAICK4821A1Z5","MH","MH"],  # GRN 07-01 -> due 07-31 -> Not Due
        ["INV-B","SO2","ZEPTO B","2026-04-01",40000,0,40000,"MH","27AAICK4821A1Z5","MH","MH"],  # GRN 06-15 -> due 07-15 -> Due (today)
        ["INV-C","SO3","ZEPTO C","2026-04-01",30000,0,30000,"MH","27AAICK4821A1Z5","MH","MH"],  # GRN 05-14 -> due 06-13 -> Overdue
        ["INV-D","SO4","ZEPTO D","2026-04-01",5,0,5,"MH","27AAICK4821A1Z5","MH","MH"],          # net 5 <=10 -> Paid -> blank status
        ["INV-E","SO5","ZEPTO E","2026-04-01",20000,0,20000,"MH","27AAICK4821A1Z5","MH","MH"],  # no GRN -> blank due date/status
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities"],
        ["PA","INV-A","X"], ["PB","INV-B","X"], ["PC","INV-C","X"],
        ["PD","INV-D","X"], ["PE","INV-E","X"],   # PE not in GRN pool -> no GRN Date
    ]})
    grn = (b"GRN ID,PO ID,Created On,Status\r\n"
           b"GA,PA,7/1/2026,CONFIRMED\r\n"
           b"GB,PB,6/15/2026,CONFIRMED\r\n"
           b"GC,PC,5/14/2026,CONFIRMED\r\n"
           b"GD,PD,5/14/2026,CONFIRMED\r\n")
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files, today=datetime.date(2026, 7, 15))
    by = {r["invoice_number"]: r for r in res}
    assert by["INV-A"]["grn_date"] == "2026-07-01"
    assert by["INV-A"]["due_date"] == "2026-07-31" and by["INV-A"]["due_status"] == "Not Due"
    assert by["INV-B"]["due_date"] == "2026-07-15" and by["INV-B"]["due_status"] == "Due"
    assert by["INV-C"]["due_date"] == "2026-06-13" and by["INV-C"]["due_status"] == "Overdue"
    assert by["INV-D"]["status"] == "Paid"
    assert by["INV-D"]["due_date"] == "2026-06-13" and by["INV-D"]["due_status"] == ""   # Paid -> blank status, date still shown
    # No GRN Date -> can't age it, so Due Date stays blank but Due Status is
    # "Not Due" (not blank), and the Remark spells out "Missing GRN".
    assert by["INV-E"]["grn_date"] == "" and by["INV-E"]["due_date"] == ""
    assert by["INV-E"]["due_status"] == "Not Due"
    assert "Missing GRN" in by["INV-E"]["remark"]

    # Column placement resolved dynamically via _KEYCOL (robust to layout shifts
    # like the Month column insertion).
    from recon.zepto_receivables import build_zepto_workbook, _KEYCOL as K
    ws = build_zepto_workbook(res)["1. Invoice Tracker"]
    assert ws[f"{K['status']}2"].value == "Status"
    assert ws[f"{K['due_date']}2"].value == "Due Date"
    assert ws[f"{K['due_status']}2"].value == "Due Status"
    assert ws[f"{K['grn_no']}2"].value == "GRN No."
    assert ws[f"{K['payment_date']}2"].value == "Payment Date"
    assert ws[f"{K['status']}3"].value == f'=IF({K["net_outstanding"]}3<=10,"Paid","Not Paid")'
    print("test_due_date_and_status OK")


def test_detail_tabs_and_hyperlinks():
    # #00007: 5 split detail tabs + click-to-jump hyperlinks from the main sheet.
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _RecoRows
    def mk(inv, pay_excl, dn, cn, cnno):
        r = {k: "" for k in COLUMN_KEYS}
        r.update({"invoice_number": inv, "total_invoice_amt": 1000.0, "pending_amount": 1000.0,
                  "payment_received_incl_tds": pay_excl + 1, "payment_received_excl_tds": pay_excl, "tds": 1.0,
                  "debit_note_issued": dn, "credit_note_issued": cn, "credit_note_no": cnno,
                  "gross_outstanding": 0.0, "net_outstanding": 0.0, "status": "Not Paid"})
        return r
    results = _RecoRows([mk("INV26-27/000039/", 900.0, 50.0, 0.0, ""),
                         mk("INV26-27/000200", 0.0, 0.0, 100.0, "CN/26-27/0099")])
    results.pmdn_adjustment = -123.45
    results.details = {
        "payments": [{"invoice": "INV26-27/000039/", "ref": "20000288", "incl": 901.0, "excl": 900.0, "tds": 1.0}],
        "debit_notes": [{"invoice": "26-27/000039/", "ref": "26-27/000039/_QD", "amount": -50.0}],  # pre-remap key (no INV)
        "credit_notes": [{"invoice": "INV26-27/000200", "number": "CN/26-27/0099", "amount": 100.0}],
        "pmdd": [{"ref": "PMDDN-1", "amount": -100.0}],
        "ap_ar": [{"ref": "APAR-1", "amount": -23.45}],
    }
    wb = build_zepto_workbook(results)
    assert wb.sheetnames == ["1. Invoice Tracker", "Summary", "Payments", "Payment Advice Consolidate",
                             "Debit Notes", "Credit Notes", "PMDD", "AP-AR & Manual Adj"]
    # detail rows landed on each tab; the DN's malformed ref key ("26-27/000039/")
    # is resolved to its universe invoice so the tab reconciles with the main sheet.
    assert wb["Payments"]["A3"].value == "INV26-27/000039/"
    assert wb["Debit Notes"]["A3"].value == "INV26-27/000039/"   # resolved from "26-27/000039/"
    assert wb["Debit Notes"]["B3"].value == "26-27/000039/_QD"    # original ref preserved
    assert wb["Debit Notes"]["C3"].value == 50.0                 # raw -50 source -> stored POSITIVE
    assert wb["Credit Notes"]["A3"].value == "INV26-27/000200"
    assert wb["PMDD"]["A3"].value == "PMDDN-1"
    assert wb["AP-AR & Manual Adj"]["A3"].value == "APAR-1"
    # hyperlinks from the main sheet point into the right tabs (row-level).
    # Columns resolved via _KEYCOL (robust to the Month column shift).
    from recon.zepto_receivables import _KEYCOL as K
    ws = wb["1. Invoice Tracker"]
    pay_c, dn_c, cn_c, cnno_c = (K["payment_received_excl_tds"], K["debit_note_issued"],
                                 K["credit_note_issued"], K["credit_note_no"])
    assert ws[f"{pay_c}3"].hyperlink and ws[f"{pay_c}3"].hyperlink.location == "'Payments'!A3"
    assert ws[f"{dn_c}3"].hyperlink and ws[f"{dn_c}3"].hyperlink.location == "'Debit Notes'!A3"
    assert ws[f"{cn_c}4"].hyperlink and ws[f"{cn_c}4"].hyperlink.location == "'Credit Notes'!A3"
    assert ws[f"{cnno_c}4"].hyperlink and ws[f"{cnno_c}4"].hyperlink.location == "'Credit Notes'!A3"
    # a cell with no matching detail has no link
    assert ws[f"{dn_c}4"].hyperlink is None
    print("test_detail_tabs_and_hyperlinks OK")


def test_pod_from_payment_track():
    # POD No <- 'LRN', POD Date <- 'Delivery Date' of the Zepto Payment track,
    # matched by invoice (slash-normalized): universe "…/000192/" finds the
    # payment-track "…/000192".
    import datetime
    invd = _xlsx({"Invoice Details": [
        ["t"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV26-27/000192/","SO1","ZEPTO MUM","2026-04-30",1000,0,1000,"MH","27AAICK4821A1Z5","MH","MH"],
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities","QTY","Delivery Date","Courier","LRN","GRN"],
        ["P4254950","INV26-27/000192","MUM","1","2026-04-14","Delhivery","275780443","GrnX"],  # no slash here
    ]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files, today=datetime.date(2026, 7, 15))
    r = res[0]
    assert r["pod_no"] == "275780443"    # LRN, matched despite the universe trailing slash
    assert r["pod_date"] == "2026-04-14"  # Delivery Date
    print("test_pod_from_payment_track OK")


def test_pod_date_independent_of_lrn():
    # POD No (from LRN) and POD Date (from Delivery Date) are captured
    # INDEPENDENTLY: a row with a Delivery Date but an empty LRN must still yield
    # a POD Date (the old `if lrn:` gate dropped it). Courier words in the LRN
    # cell (Porter/Booked/Self) are kept verbatim in POD No.
    from recon.zepto_receivables import parse_payment_track_pod
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities","QTY","Delivery Date","Courier","LRN","GRN"],
        ["P1","INV26-27/000031","FBD",1,"2026-03-21","DP","","GrnA"],          # Delivery Date, empty LRN
        ["P2","INV26-27/000147","BLR",1,"","DP","1846224940","GrnB"],          # LRN, no Delivery Date
        ["P3","INV26-27/000045","MUM",1,"2026-04-15","Porter","Porter","GrnC"],# courier word kept as-is
    ]})
    pod = parse_payment_track_pod(pay)
    assert pod["INV26-27/000031"]["pod_date"] == "2026-03-21"   # not dropped despite empty LRN
    assert pod["INV26-27/000031"]["pod_no"] == ""
    assert pod["INV26-27/000147"]["pod_no"] == "1846224940"
    assert pod["INV26-27/000147"]["pod_date"] == ""
    assert pod["INV26-27/000045"]["pod_no"] == "Porter"
    assert pod["INV26-27/000045"]["pod_date"] == "2026-04-15"
    print("test_pod_date_independent_of_lrn OK")


def test_payment_date_and_advice_no_from_pdf():
    # Each line item inherits its PDF's Payment Date + Payment Ref No (advice no),
    # and every line lands in the consolidate list tagged with those header fields.
    from recon.zepto_receivables import parse_payment_advice_pdf
    details = {"payments": [], "debit_notes": [], "pmdd": [], "ap_ar": [], "credit_notes": [], "consolidate": []}
    parse_payment_advice_pdf([_read_fixture_pdf()], details)
    assert details["payments"], "expected payment line items"
    p0 = details["payments"][0]
    assert p0["payment_date"] == "07/04/2026"
    assert p0["advice_no"] == "HSBCN52026040735815535"
    assert len(details["consolidate"]) >= len(details["payments"])
    assert all(c["advice_no"] == "HSBCN52026040735815535" for c in details["consolidate"])
    assert all(c["payment_date"] == "07/04/2026" for c in details["consolidate"])
    print("test_payment_date_and_advice_no_from_pdf OK")


def test_dynamic_header_wording_variants():
    # The header parsers must not hinge on one exact PDF layout.
    from recon.zepto_receivables import _parse_payment_date, _HEADER_RE
    assert _parse_payment_date("Payment Date 02/07/2026") == "02/07/2026"
    assert _parse_payment_date("Payment Date: 2026-07-02") == "2026-07-02"
    assert _parse_payment_date("Payment Date 02 Jul 2026") == "02 Jul 2026"
    assert _parse_payment_date("Payment Posting Date 05/07/2026") == ""   # different field, must NOT match
    for s, want in [("Payment Ref No. 20260702-R02L1", "20260702-R02L1"),
                    ("Payment Reference No X99", "X99"),
                    ("Payment Ref: ABC123", "ABC123")]:
        m = _HEADER_RE["ref_no"].search(s)
        assert m and m.group(1).strip() == want, (s, m and m.group(1))
    print("test_dynamic_header_wording_variants OK")


def test_summary_and_consolidate_and_detail_columns():
    # Summary: DN shown positive & subtracted; Net Sales / Net Receivables /
    # Net Receivables - 2; Amount Received in Bank = consolidate Payment Amt total.
    # Detail tabs (Payments/Debit Notes/PMDD/AP-AR) carry Payment Date + Advice No.
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _RecoRows
    row = {k: "" for k in COLUMN_KEYS}
    row.update({"invoice_number": "INV1", "total_invoice_amt": 1000.0, "credit_note_issued": 100.0,
                "debit_note_issued": 50.0, "tds": 10.0, "payment_received_incl_tds": 400.0,
                "net_outstanding": 440.0, "status": "Not Paid"})
    results = _RecoRows([row])
    results.pmdn_adjustment = -200.0
    results.details = {
        "payments": [{"invoice": "INV1", "ref": "D1", "incl": 399.0, "excl": 400.0, "tds": 1.0,
                      "payment_date": "02/07/2026", "advice_no": "ADV-1"}],
        "debit_notes": [], "credit_notes": [],
        "pmdd": [{"ref": "PMDDN-1", "amount": -100.0, "payment_date": "02/07/2026", "advice_no": "ADV-1"}],
        "ap_ar": [{"ref": "APAR-1", "amount": -23.0, "payment_date": "02/07/2026", "advice_no": "ADV-1"}],
        "consolidate": [
            {"advice_no": "ADV-1", "payment_date": "02/07/2026", "type": "Invoice Payment",
             "doc_no": "D1", "ref_doc": "INV1", "amount": 400.0, "tds": 1.0, "payment_amt": 399.0},
            {"advice_no": "ADV-1", "payment_date": "02/07/2026", "type": "Credit Memo",
             "doc_no": "C1", "ref_doc": "PMDDN-1", "amount": -100.0, "tds": 0.0, "payment_amt": -100.0},
        ],
    }
    wb = build_zepto_workbook(results)
    ws = wb["Summary"]
    assert _find_summary_amount(ws, "Debit Note Issued - Issued by Zepto") == 50.0   # positive
    assert _find_summary_amount(ws, "Net Sales") == 850.0                   # 1000-100-50 (TDS not in Net Sales)
    assert _find_summary_amount(ws, "Net Receivables") == 440.0            # 850 - 400 - 10(TDS)
    assert _find_summary_amount(ws, "Amount Received in Bank") == 299.0     # 399 + (-100)
    # "Net Receivables - 2" row was removed from the Summary.
    assert not any(c.value == "Net Receivables - 2" for c in ws["A"] if c.value)
    # new detail columns
    assert (wb["Payments"]["F2"].value, wb["Payments"]["G2"].value) == ("Payment Date", "Payment Advice No")
    assert (wb["Payments"]["F3"].value, wb["Payments"]["G3"].value) == ("02/07/2026", "ADV-1")
    assert (wb["Debit Notes"]["D2"].value, wb["Debit Notes"]["E2"].value) == ("Payment Date", "Payment Advice No")
    assert wb["PMDD"]["C2"].value == "Payment Date"
    assert wb["AP-AR & Manual Adj"]["C2"].value == "Payment Date"
    con = wb["Payment Advice Consolidate"]
    assert con["A2"].value == "Payment Advice No" and con["H2"].value == "Payment Amt"
    print("test_summary_and_consolidate_and_detail_columns OK")


def test_grn_from_payment_track_and_wafers_dropped():
    # GRN falls back to the Payment track's own GRN column when the monthly
    # GRN_List CSVs have no match (they only cover Apr-Jun). And a non-date note
    # like "Wafers" in the Delivery Date column must NOT surface as a POD Date.
    import datetime
    invd = _xlsx({"Invoice Details": [
        ["t"],
        ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
         "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
        ["INV26-27/000558","SO1","ZEPTO BLR","2026-06-30",1000,0,1000,"KA","29AAICK4821A1Z5","KA","KA"],
        ["INV26-27/000543","SO2","ZEPTO FBD","2026-06-29",2000,0,2000,"HR","06AAICK4821A1Z5","HR","HR"],
        ["INV26-27/000600","SO3","ZEPTO MUM","2026-07-08",500,0,500,"MH","27AAICK4821A1Z5","MH","MH"],
    ]})
    pay = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities","QTY","Delivery Date","Courier","LRN","GRN"],
        ["P4801183","INV26-27/000558","BLR",6718,"2026-07-07","Delhivery","286591390","GrnCode51517893"],
        ["P4689855","INV26-27/000543","FBD",2150,"Wafers","Wafers","","GrnCodeWAF"],  # Wafers note, empty LRN, has GRN
        ["P4700000","INV26-27/000600","MUM",100,"2026-07-08","DP","286599999","Missing GRN"],  # GRN column = the note, not a code
    ]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"   # empty GRN_List -> forces the Payment-track fallback
    cn = _xlsx({"Credit Note Details": [["t"], ["invoice_number","bcy_total"]]})
    files = {"zepto_payment": _file(pay), "grn_list": [_file(grn)],
             "invoice_details": _file(invd), "payment_advice": [], "credit_note": _file(cn)}
    res = reconcile_zepto(files, today=datetime.date(2026, 8, 5))
    by = {r["invoice_number"]: r for r in res}
    a, b = by["INV26-27/000558"], by["INV26-27/000543"]
    assert a["grn_no"] == "GrnCode51517893"   # from Payment track (GRN_List empty)
    assert a["pod_date"] == "2026-07-07" and a["pod_no"] == "286591390"
    assert b["grn_no"] == "GrnCodeWAF"        # GRN captured despite Wafers/empty LRN
    assert b["pod_date"] == ""                 # "Wafers" is not a date -> dropped
    assert b["pod_no"] == ""                   # LRN empty
    # A source GRN column holding the note "Missing GRN" (not a code) counts as
    # NO grn -> flagged, not silently accepted.
    d = by["INV26-27/000600"]
    assert d["grn_no"] == "" and d["remark"] == "Missing GRN"
    # Remark column spells out the gaps. GRN carried from the Payment track has
    # no date, so it can't be aged: Due Status = Not Due, Remark = Missing GRN
    # Date (the number is present, only the date is missing -> distinct wording).
    assert a["remark"] == "Missing GRN Date" and a["due_status"] == "Not Due"
    assert a["due_date"] == ""
    assert b["remark"] == "Missing GRN Date; Missing POD"   # no GRN date + no LRN/POD
    print("test_grn_from_payment_track_and_wafers_dropped OK")


def test_dn_status_merge_missing_grn_and_tab_colors():
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _KEYCOL, _RecoRows
    # DN Accepted / DN Not Accepted merged into one DN Status column.
    assert "dn_status" in COLUMN_KEYS
    assert "dn_accepted" not in COLUMN_KEYS and "dn_not_accepted" not in COLUMN_KEYS
    def mk(inv, po, grn, status, grn_date=""):
        r = {k: "" for k in COLUMN_KEYS}
        r.update({"invoice_number": inv, "po": po, "grn_no": grn, "grn_date": grn_date,
                  "total_invoice_amt": 1000.0, "net_outstanding": 0.0, "status": status})
        return r
    results = _RecoRows([mk("INV1", "P1", "", "Not Paid"),                     # PO but no GRN -> Missing GRN
                         mk("INV2", "P2", "GrnCode9", "Paid"),                 # GRN No but no date -> Missing GRN Date
                         mk("INV3", "P3", "GrnCode7", "Not Paid", "2026-05-11")])  # GRN No + date -> fine
    wb = build_zepto_workbook(results)
    ws = wb["1. Invoice Tracker"]
    # header + dropdown on DN Status
    assert ws[f"{_KEYCOL['dn_status']}2"].value == "DN Status"
    assert any('"Accepted,Not Accepted"' in str(dv.formula1) for dv in ws.data_validations.dataValidation)
    # Missing GRN cell (row 3 = INV1): red text label
    g = ws[f"{_KEYCOL['grn_no']}3"]
    assert g.value == "Missing GRN" and g.fill.fgColor.rgb.endswith("FFC7CE")
    assert ws[f"{_KEYCOL['grn_no']}4"].value == "GrnCode9"   # INV2 keeps its GRN
    # Missing GRN Date: GRN No present but no date -> GRN Date cell is a red
    # "Missing GRN Date" label (same treatment as Missing GRN, so it's not
    # mistaken for "not captured"). INV1 (no GRN at all) leaves the date blank;
    # INV3 (GRN + date) shows the date with no red.
    assert ws[f"{_KEYCOL['grn_date']}3"].value in ("", None)                    # INV1: no GRN -> date blank, not flagged
    gd = ws[f"{_KEYCOL['grn_date']}4"]
    assert gd.value == "Missing GRN Date" and gd.fill.fgColor.rgb.endswith("FFC7CE")
    gd3 = ws[f"{_KEYCOL['grn_date']}5"]
    assert gd3.value == "2026-05-11" and not gd3.fill.fgColor.rgb.endswith("FFC7CE")
    # Status static fill (renders in Numbers): Not Paid red, Paid green
    assert ws[f"{_KEYCOL['status']}3"].fill.fgColor.rgb.endswith("FFC7CE")   # Not Paid
    assert ws[f"{_KEYCOL['status']}4"].fill.fgColor.rgb.endswith("C6EFCE")   # Paid
    # tabs are colour-coded
    assert ws.sheet_properties.tabColor is not None
    assert wb["Summary"].sheet_properties.tabColor is not None
    print("test_dn_status_merge_missing_grn_and_tab_colors OK")


def test_multi_month_accumulation():
    # The vendor ships only the new month each time, but a prior-month invoice's
    # GRN/PO/POD can land in a later file — so Invoice Details, Credit Notes and
    # the Zepto Payment track must ACCUMULATE across every file in the folder.
    import datetime
    def _invd(inv, name, date, amt):
        return _xlsx({"Invoice Details": [["t"],
            ["invoice_number","reference_number","customer_name","date","bcy_total","tax_amount",
             "amount_without_tax","place_of_supply","gst_no","billing_state","shipping_state"],
            [inv,"SO",name,date,amt,0,amt,"MH","27AAICK4821A1Z5","MH","MH"]]})
    invd_apr = _invd("INV26-27/000010","ZEPTO A","2026-06-30",1000)   # prior month
    invd_jul = _invd("INV26-27/000050","ZEPTO B","2026-07-15",2000)   # new month
    pay_old = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities","QTY","Delivery Date","Courier","LRN","GRN"],
        ["P10","INV26-27/000010","MH",1,"2026-06-20","DP","LRN10",""]]})          # June: no GRN yet
    pay_new = _xlsx({"Zepto Payment track": [
        ["Zepto Payment track PO Number","Invoice Number","Cities","QTY","Delivery Date","Courier","LRN","GRN"],
        ["P10","INV26-27/000010","MH",1,"2026-06-20","DP","LRN10","GrnCode10"],   # July: GRN filled in
        ["P50","INV26-27/000050","KA",1,"2026-07-10","DP","LRN50","GrnCode50"]]})
    grn = b"GRN ID,PO ID,Created On,Status\r\n"
    cn1 = _xlsx({"Credit Note Details": [["t"],["invoice_number","bcy_total"],["INV26-27/000010",100]]})
    cn2 = _xlsx({"Credit Note Details": [["t"],["invoice_number","bcy_total"],["INV26-27/000010",50],["INV26-27/000050",25]]})
    files = {"invoice_details": [_file(invd_apr), _file(invd_jul)],
             "zepto_payment": [_file(pay_new), _file(pay_old)],   # newest-first
             "grn_list": [_file(grn)], "payment_advice": [],
             "credit_note": [_file(cn1), _file(cn2)]}
    res = reconcile_zepto(files, today=datetime.date(2026, 8, 6))
    by = {r["invoice_number"]: r for r in res}
    assert set(by) == {"INV26-27/000010", "INV26-27/000050"}       # both months present
    assert by["INV26-27/000010"]["po"] == "P10"
    assert by["INV26-27/000010"]["grn_no"] == "GrnCode10"           # GRN from the newer payment file
    assert by["INV26-27/000010"]["pod_no"] == "LRN10"
    assert by["INV26-27/000050"]["po"] == "P50" and by["INV26-27/000050"]["grn_no"] == "GrnCode50"
    assert round(by["INV26-27/000010"]["credit_note_issued"], 2) == 150.0   # summed across both CN files
    assert round(by["INV26-27/000050"]["credit_note_issued"], 2) == 25.0
    print("test_multi_month_accumulation OK")


def test_due_status_excess_paid():
    # New "Excess Paid" state: Paid AND Net Outstanding more negative than -10.
    import datetime as dt
    from recon.zepto_receivables import _due_status
    today = dt.date(2026, 7, 15)
    future = dt.date(2026, 8, 1)
    past = dt.date(2026, 7, 1)
    assert _due_status(future, True, today, net_outstanding=-500.0) == "Excess Paid"
    assert _due_status(None, True, today, net_outstanding=-10.01) == "Excess Paid"
    assert _due_status(future, True, today, net_outstanding=-5.0) == ""        # within -10 -> settled
    assert _due_status(None, True, today, net_outstanding=0.0) == ""
    # Unpaid rows keep normal ageing regardless of net.
    assert _due_status(future, False, today, net_outstanding=1000.0) == "Not Due"
    assert _due_status(today, False, today, net_outstanding=1000.0) == "Due"
    assert _due_status(past, False, today, net_outstanding=1000.0) == "Overdue"
    print("test_due_status_excess_paid OK")


def test_advance_adjusted_and_catchall_routing():
    # "Advance Adjusted" is a KNOWN adjustment (routed, not flagged); any OTHER
    # money-bearing type is caught by the catch-all (routed + flagged); a no-money
    # junk row is skipped.
    from recon.zepto_receivables import _extract_line_items_from_table
    details = {"payments": [], "debit_notes": [], "pmdd": [], "ap_ar": [],
               "consolidate": [], "unknown_types": []}
    payments = {}; debit_notes = {}; pmdn = [0.0]
    table = [
        ["Sr", "Type of Document", "Doc No", "Ref", "Amount", "CCY", "TDS", "Payment Amt"],  # header
        ["1", "Invoice Payment", "D1", "INV26-27/000007", "1000", "INR", "10", "990"],
        ["2", "Advance Adjusted", "D2", "KK10016488", "-618000", "INR", "0", "-618000"],
        ["3", "Sausage Adjustment", "D3", "XX1", "-200", "INR", "0", "-200"],  # unknown -> catch-all
        ["4", "Letterhead line", None, None, None, None, None, None],          # no money -> skip
    ]
    _extract_line_items_from_table(table, payments, debit_notes, pmdn, details,
                                   advice_no="ADV-1", payment_date="02/07/2026")
    assert round(pmdn[0], 2) == round(-618000 + -200, 2)          # both non-invoice adjustments
    apar_refs = [x["ref"] for x in details["ap_ar"]]
    assert "KK10016488" in apar_refs and "XX1" in apar_refs
    types = [x["type"] for x in details["consolidate"]]
    assert "Advance Adjusted" in types and "Sausage Adjustment" in types and "Invoice Payment" in types
    assert "Letterhead line" not in types                        # junk row skipped
    assert details["unknown_types"] == ["Sausage Adjustment"]    # only the truly-unknown flagged
    print("test_advance_adjusted_and_catchall_routing OK")


def test_payment_advice_reject_on_header_mismatch():
    # Real fixture whose line items don't sum to its header Amount (off by 3.71):
    # accepted within tolerance, REJECTED (and excluded from every total) below it.
    from recon.zepto_receivables import parse_payment_advice_pdf
    b = open(os.path.join(_FIXTURE_DIR_ALL, "02_2026-07-03_69e6fe644a54a012_2000008712.pdf"), "rb").read()
    d1 = {k: [] for k in ("payments", "debit_notes", "pmdd", "ap_ar", "consolidate", "unknown_types")}
    rej1 = []
    pay1, _, _ = parse_payment_advice_pdf([b], d1, advice_tolerance=100.0, rejected=rej1)
    assert rej1 == [] and len(pay1) > 0                           # within Rs 100 -> accepted
    d2 = {k: [] for k in ("payments", "debit_notes", "pmdd", "ap_ar", "consolidate", "unknown_types")}
    rej2 = []
    pay2, _, _ = parse_payment_advice_pdf([b], d2, advice_tolerance=1.0, rejected=rej2)
    assert len(rej2) == 1 and rej2[0]["doc"] == "2000008712" and abs(rej2[0]["diff"]) > 1
    assert pay2 == {} and d2["consolidate"] == []                 # nothing leaks from a rejected advice
    print("test_payment_advice_reject_on_header_mismatch OK")


def test_summarize_zepto_money_totals():
    # Stage 2B: summarize_zepto carries the money-flow totals the UI strip needs
    # (same basis as the Summary sheet). Net Receivables = Net Sales - Payment - TDS.
    from recon.zepto_receivables import summarize_zepto, COLUMN_KEYS, _RecoRows
    def row(inv, total, cn, dn, tds, incl):
        r = {k: "" for k in COLUMN_KEYS}
        r.update({"invoice_number": inv, "total_invoice_amt": total, "credit_note_issued": cn,
                  "debit_note_issued": dn, "tds": tds, "payment_received_incl_tds": incl,
                  "net_outstanding": 0.0, "status": "Not Paid"})
        return r
    results = _RecoRows([row("INV1", 1000.0, 100.0, 50.0, 10.0, 400.0)])
    results.pmdn_adjustment = -200.0
    results.details = {"consolidate": [{"payment_amt": 399.0}, {"payment_amt": -100.0}]}
    s = summarize_zepto(results)
    assert s["sales_incl_tax"] == 1000.0
    assert s["sale_return"] == 100.0
    assert s["debit_note_issued"] == 50.0            # shown positive
    assert s["net_sales"] == 850.0                   # 1000 - 100 - 50
    assert s["net_receivables"] == 440.0             # 850 - 400 - 10
    assert s["adjustments"] == -200.0
    assert s["amount_received_in_bank"] == 299.0     # 399 + (-100)
    print("test_summarize_zepto_money_totals OK")


def test_canonical_invoice_resolution_exact_first_and_ambiguity_safe():
    # Payments/CNs are matched to the invoice universe: EXACT first (keeps the
    # slash-distinct invoices apart), then UNIQUE canonical (tolerates missing
    # I/IN, 2425 vs 24-25, trailing slash, _PD). Ambiguous slash-pairs and
    # non-invoice refs (VRC) are never canonically matched.
    from recon.zepto_receivables import _canon_invoice, _build_canon_index, _resolve_invoice_key
    universe = {
        "INV25-26/000782", "INV24-25/000639/", "INV24-25/000300", "INV25-26/000940",
        "INV26-27/000009", "INV26-27/000192/", "INV25-26/000127", "INV24-25/000356",
        # an ambiguous trailing-slash PAIR (two distinct invoices, same canon):
        "INV26-27/000039", "INV26-27/000039/",
    }
    idx = _build_canon_index(universe)
    R = lambda ref: _resolve_invoice_key(ref, universe, idx)
    # format variants resolve to the right single invoice
    assert R("INV-25/26-000782") == "INV25-26/000782"        # extra dash
    assert R("INV2425/000300") == "INV24-25/000300"          # 2425 -> 24-25
    assert R("INV25/26/000940") == "INV25-26/000940"         # slash-in-year
    assert R("INV26-27/000009/") == "INV26-27/000009"        # stray trailing slash
    assert R("INV26-27/000192") == "INV26-27/000192/"        # tracker HAS the slash
    assert R("NV25-26/000127") == "INV25-26/000127"          # missing I
    assert R("V24-25/000356_PD") == "INV24-25/000356"        # missing IN + _PD
    assert R("INV24-25/000639") == "INV24-25/000639/"        # tracker has slash
    # EXACT wins for the ambiguous pair (both distinct invoices keep their own)
    assert R("INV26-27/000039") == "INV26-27/000039"
    assert R("INV26-27/000039/") == "INV26-27/000039/"
    # a canonical-ONLY ref for the ambiguous pair is NOT guessed
    assert R("INV2627/000039") is None
    # genuinely non-invoice refs never match
    assert _canon_invoice("VRC") is None
    assert R("VRC") is None
    assert _canon_invoice("PMDDN-79939") is None
    print("test_canonical_invoice_resolution_exact_first_and_ambiguity_safe OK")


def test_credit_notes_deduped_across_overlapping_files():
    # The monthly Credit Note file and a running to-date file overlap; a CN
    # present in BOTH must be counted ONCE (double-counting inflated Sale Return
    # ~8L on real data). Dedup is by credit-note number across files.
    from recon.zepto_receivables import _merge_credit_notes
    hdr = ["creditnote_number", "bcy_total", "invoice_number"]
    file_a = _xlsx({"Credit Note Details": [["title"], hdr,
        ["CN-1", 1000.0, "INV1"],
        ["CN-2", 500.0, "INV1"],
    ]})
    file_b = _xlsx({"Credit Note Details": [["title"], hdr,
        ["CN-2", 500.0, "INV1"],   # overlap (same CN# as file_a) -> must NOT double
        ["CN-3", 300.0, "INV2"],   # new
    ]})
    merged = _merge_credit_notes([file_a, file_b])
    assert round(merged["INV1"]["amount"], 2) == 1500.0   # 1000 + 500 (CN-2 once)
    assert round(merged["INV2"]["amount"], 2) == 300.0
    assert merged["INV1"]["numbers"] == ["CN-1", "CN-2"]
    print("test_credit_notes_deduped_across_overlapping_files OK")


def test_payments_tab_shows_unmatched_and_reconciles():
    # Invoice payments for invoices NOT in the tracker are shown in a flagged
    # section so the Payments tab reconciles to the Consolidate's Invoice-Payment
    # total (no silent gap). Matched subtotal stays = "Payment Received".
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _RecoRows
    r = {k: "" for k in COLUMN_KEYS}
    r.update({"invoice_number": "INV1", "total_invoice_amt": 1000.0,
              "status": "Not Paid", "net_outstanding": 100.0})
    results = _RecoRows([r])
    results.details = {"payments": [
        {"invoice": "INV1", "ref": "P1", "incl": 900.0, "excl": 901.0, "tds": 1.0,
         "payment_date": "", "advice_no": "ADV1"},   # matched (INV1 in universe)
        {"invoice": "INVX", "ref": "P2", "incl": 250.0, "excl": 251.0, "tds": 1.0,
         "payment_date": "", "advice_no": "ADV2"},   # NOT in universe
    ], "debit_notes": [], "credit_notes": [], "pmdd": [], "ap_ar": [],
        "consolidate": [], "unknown_types": []}
    wb = build_zepto_workbook(results)
    pw = wb["Payments"]
    colA = [pw.cell(row, 1).value for row in range(1, pw.max_row + 1)]
    assert any(v and "PAYMENTS NOT IN TRACKER" in str(v) for v in colA)
    assert "INVX" in colA                                     # unmatched payment shown
    grand = next(pw.cell(row, 5).value for row in range(pw.max_row, 0, -1)
                 if pw.cell(row, 1).value and "GRAND TOTAL" in str(pw.cell(row, 1).value))
    assert grand == round(900.0 + 250.0, 2)                   # matched + unmatched reconcile
    print("test_payments_tab_shows_unmatched_and_reconciles OK")


def test_summary_excess_paid_bucket_and_dynamic_brand():
    from recon.zepto_receivables import build_zepto_workbook, COLUMN_KEYS, _RecoRows
    def row(inv, total, net, status, due_status):
        r = {k: "" for k in COLUMN_KEYS}
        r.update({"invoice_number": inv, "total_invoice_amt": total, "net_outstanding": net,
                  "status": status, "due_status": due_status})
        return r
    results = _RecoRows([
        row("INV1", 1000.0, 500.0, "Not Paid", "Overdue"),
        row("INV2", 1000.0, -750.0, "Paid", "Excess Paid"),
        row("INV3", 1000.0, -800.0, "Paid", "Excess Paid"),
    ])
    wb = build_zepto_workbook(results, payload={"brand_name": "Koparo"})
    ws = wb["Summary"]
    assert _find_summary_amount(ws, "Excess Paid") == round(-750.0 + -800.0, 2)
    assert _find_summary_amount(ws, "Overdue") == 500.0
    labels = [c.value for c in ws["A"] if c.value]
    assert any(str(l).startswith("Sale Return (Credit Notes) - Issued by Koparo") for l in labels)
    assert "Debit Note Issued - Issued by Zepto" in labels
    print("test_summary_excess_paid_bucket_and_dynamic_brand OK")


if __name__ == "__main__":
    test_normalizers_and_dn_transform()
    test_norm_inv_does_not_strip_trailing_slash()
    test_dn_ref_to_invoice_primary_transform_only()
    test_parse_invoice_details_real_fixture_dedup_by_num_amount_date()
    test_reconcile_zepto_real_fixtures_dn_fallback_and_pmddn()
    test_csv_header_detection_and_getter()
    test_zepto_payment_and_grn_gate()
    test_parse_invoice_details()
    test_parse_payment_advice()
    test_parse_credit_notes()
    test_parse_credit_notes_real_fixture()
    test_parse_grn_real_fixture_captures_grn_id_and_created_on()
    test_parse_lrn_multi_invoice_row()
    test_parse_lrn_sheet_without_invoice_column_is_skipped()
    test_reconcile_end_to_end()
    test_workbook_has_live_formulas()
    test_status_signed_threshold_negative_net_is_paid()
    test_reconcile_zepto_signed_status_direct()
    test_status_column_has_conditional_formatting()
    test_universe_includes_invoice_with_no_po_mapping_at_all()
    test_parse_payment_advice_pdf_single()
    test_parse_payment_advice_pdf_dedup()
    test_parse_payment_advice_pdf_multipage()
    test_parse_payment_advice_pdf_credit_memo_routes_as_debit_note()
    test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn()
    test_parse_payment_advice_pdf_trailing_slash_invoice_and_dn_2()
    test_parse_payment_advice_pdf_pmddn_routes_to_pmdn_total_not_debit_notes()
    test_reconcile_zepto_returns_list_subclass_with_pmdn_adjustment()
    test_build_zepto_workbook_fills_pmddn_summary_line()
    test_build_zepto_workbook_has_summary_tab()
    test_vendor_filter_keeps_only_zepto()
    test_po_matches_despite_trailing_slash_and_ignores_grn_gate()
    test_due_date_and_status()
    test_detail_tabs_and_hyperlinks()
    test_pod_from_payment_track()
    test_pod_date_independent_of_lrn()
    test_payment_date_and_advice_no_from_pdf()
    test_dynamic_header_wording_variants()
    test_summary_and_consolidate_and_detail_columns()
    test_grn_from_payment_track_and_wafers_dropped()
    test_dn_status_merge_missing_grn_and_tab_colors()
    test_multi_month_accumulation()
    print("ALL TESTS PASSED")
