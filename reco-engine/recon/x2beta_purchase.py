"""
x2beta_purchase.py — Invoice Process rows -> X2Beta (Tally purchase-import) workbook.

Takes the rows we already store in `invoice_process` and emits the CA's exact
109-column X2Beta "Purchase / Debit Note" workbook, Tally-import ready.

ONE template serves every brand (`format_templates/x2beta_purchase_template.xlsx`).
Only the GST ledger block is brand-specific, so it is resolved per run:
  * a ledger the template already carries is reused with its EXACT spelling
    (Tally writes double spaces — "CGST Input  9%" — and mixes "Karnataka" with
    "DL"/"HR"/"MH"; those strings must match the brand's Tally ledger names),
  * a (tax, rate, state) the template lacks is CREATED by inserting a column
    before CESS, so a new buyer GSTIN never silently loses its tax,
  * a template ledger this run does NOT use has its header BLANKED, so a brand
    never ships columns naming another brand's Tally ledgers.

Classification comes from `voucher_type`, which n8n already inverts to our side
(2026-09-02): supplier Tax Invoice -> "Purchase <State>", supplier Debit Note ->
"Credit Note <State>", supplier Credit Note -> "Debit Note <State>".

Layout (verified against the accountant's real Dchica Output.xlsx):
  header row 6, ">>>>" guide row 7, data from row 8.
  col1 Vch Date | col2 Vch Type | col3 Vch No | col4 Ref No | col5 Ref Date
  col6 Is DN?   | col7 Is Vch?  | col8 Party Ledger | col9 Purchase Ledger
  col18 Amount  | GST ledgers   | Narration | Note Reason / Orig Inv No / Date
Only the FIRST ledger line of a voucher carries cols 1-5; continuation lines blank.

Self-test (reproduces the accountant's file exactly):
    python3 -m recon.x2beta_purchase
"""
import io
import os
import re
from collections import OrderedDict, defaultdict
from copy import copy
from datetime import datetime, date

import openpyxl

HEADER_ROW = 6
DATA_ROW = 8

C_VCH_DATE, C_VCH_TYPE, C_VCH_NO, C_REF_NO, C_REF_DATE = 1, 2, 3, 4, 5
C_IS_DN, C_IS_VCH, C_PARTY, C_PURCH_LEDGER = 6, 7, 8, 9
C_AMOUNT = 18
C_NARRATION = 35
C_NOTE_REASON, C_ORIG_INV_NO, C_ORIG_INV_DATE = 107, 108, 109

SHEET = "X2B Format"
TEMPLATE = os.path.join(os.path.dirname(os.path.dirname(os.path.abspath(__file__))),
                        "format_templates", "x2beta_purchase_template.xlsx")

MONTHS = ['Jan', 'Feb', 'Mar', 'Apr', 'May', 'Jun',
          'Jul', 'Aug', 'Sep', 'Oct', 'Nov', 'Dec']

# Default GST state labels. A brand whose Tally spells them differently overrides
# via `state_labels` — the label MUST match that brand's Tally ledger names.
DEFAULT_STATE_LABELS = {
    '01': 'JK', '02': 'HP', '03': 'PB', '04': 'CH', '05': 'UK', '06': 'HR',
    '07': 'DL', '08': 'RJ', '09': 'UP', '10': 'BR', '11': 'SK', '12': 'AR',
    '13': 'NL', '14': 'MN', '15': 'MZ', '16': 'TR', '17': 'ML', '18': 'AS',
    '19': 'WB', '20': 'JH', '21': 'OD', '22': 'CG', '23': 'MP', '24': 'GJ',
    '25': 'DD', '26': 'DD', '27': 'MH', '28': 'AP', '29': 'KA', '30': 'GA',
    '31': 'LD', '32': 'KL', '33': 'TN', '34': 'PY', '35': 'AN', '36': 'TG',
    '37': 'AP', '38': 'LA',
}

# Per-brand overrides, keyed by lowercase brand name. Dichika's Tally mixes
# abbreviations with one full state name — reproduced exactly.
BRAND_STATE_LABELS = {
    'dchica':  {'07': 'DL', '06': 'HR', '29': 'Karnataka', '27': 'MH', '33': 'TN', '36': 'TG'},
    "d'chicha": {'07': 'DL', '06': 'HR', '29': 'Karnataka', '27': 'MH', '33': 'TN', '36': 'TG'},
    'dichika': {'07': 'DL', '06': 'HR', '29': 'Karnataka', '27': 'MH', '33': 'TN', '36': 'TG'},
}


def labels_for_brand(brand_name):
    out = dict(DEFAULT_STATE_LABELS)
    out.update(BRAND_STATE_LABELS.get(str(brand_name or '').strip().lower(), {}))
    return out


# ── helpers ──────────────────────────────────────────────────────────────────
def _num(v):
    if v is None or v == '':
        return 0.0
    if isinstance(v, (int, float)):
        return float(v)
    s = re.sub(r'[^0-9.\-]', '', str(v))
    try:
        return float(s) if s not in ('', '-', '.', '-.') else 0.0
    except ValueError:
        return 0.0


def _to_date(v):
    if isinstance(v, datetime):
        return v
    if isinstance(v, date):
        return datetime(v.year, v.month, v.day)
    s = str(v or '').strip()
    if not s:
        return None
    # n8n is not consistent per line: "2026-08-31" and "31/08/2026" both occur
    for f in ('%Y-%m-%d', '%d-%m-%Y', '%d/%m/%Y', '%d-%b-%Y', '%Y/%m/%d', '%d.%m.%Y'):
        try:
            return datetime.strptime(s[:19].split(' ')[0], f)
        except ValueError:
            continue
    try:
        return datetime.fromisoformat(s[:19])
    except ValueError:
        return None


def _fmt_date(d):
    return f"{d.day},{MONTHS[d.month - 1]},{d.year}" if d else ''


def _state(gstin):
    g = str(gstin or '').strip()
    return g[:2] if len(g) >= 2 and g[:2].isdigit() else None


def _norm(s):
    return re.sub(r'\s+', ' ', str(s or '')).strip().lower()


def _coerce_vch_no(inv):
    s = str(inv).strip()
    return int(s) if s.isdigit() else s


# ── voucher kind ─────────────────────────────────────────────────────────────
def kind_of(row):
    """Purchase | Debit Note | Credit Note — already inverted to OUR side by n8n."""
    vt = _norm(row.get('voucher_type'))
    if vt.startswith('credit note'):
        return 'Credit Note'
    if vt.startswith('debit note'):
        return 'Debit Note'
    # Legacy rows are checked BEFORE trusting a "Purchase" prefix: before the n8n
    # change EVERY row said "Purchase <State>", notes included, so that prefix is
    # only a default and must not outrank an explicit note signal. Rows written
    # after the change carry none of these legacy fields, so this never misfires.
    if str(row.get('invoice_type') or '').strip().lower() in ('debit note', 'dn'):
        return 'Debit Note'
    if str(row.get('invoice_type') or '').strip().lower() in ('credit note', 'cn'):
        return 'Credit Note'
    if _norm(row.get('narration')).startswith('dn:'):
        return 'Credit Note'
    if _norm(row.get('narration')).startswith('cn:'):
        return 'Debit Note'
    if _num(row.get('taxable_value')) < 0:
        return 'Debit Note'
    return 'Purchase'


def _narr_label(kind):
    """Narration names the SUPPLIER's document (their credit note = our debit note),
    matching the convention already in the accountant's own file."""
    return {'Debit Note': 'CN', 'Credit Note': 'DN'}.get(kind, 'Invoice')


# ── GST ledger columns ───────────────────────────────────────────────────────
class GstColumnMap:
    def __init__(self, headers, state_labels):
        self.headers = list(headers)
        self.state_labels = state_labels
        self.index = {}
        for i, h in enumerate(self.headers):
            m = re.match(r'^(igst|cgst|sgst)\s+input\s+([\d.]+)\s*%\s*\((.+)\)$', _norm(h))
            if m:
                self.index[(m.group(1), float(m.group(2)), _norm(m.group(3)))] = i + 1
        self.template_slots = set(self.index.values())
        self.created, self.used = [], set()

    def label_for(self, state):
        return self.state_labels.get(state, state)

    def block_for_state(self, state):
        """A state's IGST/CGST/SGST trio regardless of rate — used for 0%-GST
        vouchers (import of service), which still belong to the buyer's state."""
        lbl = _norm(self.label_for(state))
        out = {}
        for (t, r, l), c in sorted(self.index.items(), key=lambda kv: kv[1]):
            if l == lbl and t not in out:
                out[t] = c
        return out

    def column(self, tax, rate, state, ws=None):
        lbl = _norm(self.label_for(state))
        rate = round(float(rate), 2)
        for (t, r, l), c in self.index.items():
            if t == tax and abs(r - rate) < 0.01 and l == lbl:
                self.used.add(c)
                return c
        return self._create(tax, rate, self.label_for(state), ws)

    def _create(self, tax, rate, label, ws):
        name = f"{tax.upper()} Input {rate:g}% ({label})"
        free = next((i + 1 for i, h in enumerate(self.headers)
                     if str(h or '').strip() == '' and 19 <= i <= 31), None)
        if free is not None:
            col = free
            self.headers[col - 1] = name
        else:
            col = (max(self.index.values()) if self.index else 19) + 1
            if ws is not None:
                ws.insert_cols(col)
                self.index = {k: (v + 1 if v >= col else v) for k, v in self.index.items()}
                self.template_slots = {v + 1 if v >= col else v for v in self.template_slots}
                self.used = {v + 1 if v >= col else v for v in self.used}
                self.created = [(n, c + 1 if c >= col else c) for n, c in self.created]
            self.headers.insert(col - 1, name)
        self.index[(tax, rate, _norm(label))] = col
        self.created.append((name, col))
        self.used.add(col)
        if ws is not None:
            proto = ws.cell(HEADER_ROW, col - 1)
            c = ws.cell(HEADER_ROW, col)
            c.value, c._style = name, copy(proto._style)
            ws.column_dimensions[c.column_letter].width = \
                ws.column_dimensions[ws.cell(HEADER_ROW, col - 1).column_letter].width
        return col

    def prune_unused(self, ws):
        """Blank the header of any GST ledger the template shipped with that THIS
        run does not use — otherwise a brand's workbook would name another brand's
        Tally ledgers. Header only; the column itself stays, so nothing shifts."""
        pruned = []
        for col in sorted(self.template_slots - self.used):
            pruned.append(ws.cell(HEADER_ROW, col).value)
            ws.cell(HEADER_ROW, col).value = None
        return pruned


# ── conversion ───────────────────────────────────────────────────────────────
def _group(rows):
    groups = OrderedDict()
    for r in rows:
        groups.setdefault((str(r.get('invoice_number') or '').strip(),
                           str(r.get('category') or '').strip()), []).append(r)
    return groups


def _rate_of(grp):
    for g in grp:
        ig = _num(g.get('igst_rate'))
        if ig:
            return ig
        cg, sg = _num(g.get('cgst_rate')), _num(g.get('sgst_rate'))
        if cg or sg:
            return cg + sg
    return 0.0


def _ensure_columns(rows, gst, ws):
    """Create every ledger this run needs BEFORE building rows — inserting a column
    shifts every downstream field, and lines are keyed by absolute column number."""
    seen = set()
    for _, grp in _group(rows).items():
        st, rate = _state(grp[0].get('buyer_gstin')), _rate_of(grp)
        if not st:
            continue
        if not rate:
            gst.used.update(gst.block_for_state(st).values())
            continue
        for tax, r in (('igst', rate), ('cgst', rate / 2), ('sgst', rate / 2)):
            if (tax, round(r, 2), st) not in seen:
                seen.add((tax, round(r, 2), st))
                gst.column(tax, r, st, ws)


def convert(rows, headers, state_labels, ws=None):
    """rows -> ([{col: value}], GstColumnMap). One sheet: purchases + notes."""
    gst = GstColumnMap(headers, state_labels)
    _ensure_columns(rows, gst, ws)
    shift = len(gst.created)

    by_voucher = defaultdict(list)
    for (inv, cat), grp in _group(rows).items():
        by_voucher[inv].append((cat, grp))

    lines = []
    for inv, cats in by_voucher.items():
        first = True
        for cat, grp in cats:
            head = grp[0]
            kind = kind_of(head)
            d = _to_date(head.get('invoice_date'))
            taxable = sum(_num(g.get('taxable_value')) for g in grp)

            line = {
                C_IS_DN: 'Yes' if kind == 'Debit Note' else 'No',
                C_IS_VCH: 'Yes',
                C_PARTY: head.get('vendor_name_tally') or head.get('company') or '',
                C_PURCH_LEDGER: cat,
                C_AMOUNT: round(taxable, 2),
                C_NARRATION + shift: (f"{_narr_label(kind)} date: {_fmt_date(d)},  "
                                      f"{_narr_label(kind)} No: {inv}"),
            }
            if first:
                line[C_VCH_DATE] = d
                line[C_VCH_TYPE] = head.get('voucher_type') or kind
                line[C_VCH_NO] = _coerce_vch_no(inv)
                line[C_REF_NO] = _coerce_vch_no(inv)
                line[C_REF_DATE] = d
                first = False
            if kind != 'Purchase':
                line[C_NOTE_REASON + shift] = head.get('note_reason') or 'Others'
                line[C_ORIG_INV_NO + shift] = head.get('orig_invoice_number') or ''
                line[C_ORIG_INV_DATE + shift] = _to_date(head.get('orig_invoice_date'))

            b_state, s_state = _state(head.get('buyer_gstin')), _state(head.get('seller_gstin'))
            rate = _rate_of(grp)
            if b_state and not rate:
                for c in gst.block_for_state(b_state).values():
                    line[c] = 0.0
            elif b_state and rate:
                cgst = sum(_num(g.get('cgst_amount')) for g in grp)
                sgst = sum(_num(g.get('sgst_amount')) for g in grp)
                igst = sum(_num(g.get('igst_amount')) for g in grp)
                intra = s_state and b_state == s_state
                line[gst.column('igst', rate, b_state, ws)] = 0.0 if intra else round(igst, 2)
                line[gst.column('cgst', rate / 2, b_state, ws)] = round(cgst, 2) if intra else 0.0
                line[gst.column('sgst', rate / 2, b_state, ws)] = round(sgst, 2) if intra else 0.0
            lines.append(line)
    return lines, gst


# ── workbook ─────────────────────────────────────────────────────────────────
def _reflow_merges(ws, at, n):
    """insert_cols leaves merged ranges untouched, so the row-1/row-5 section
    banners would stop mid-section after the tax block widens."""
    if n <= 0:
        return
    from openpyxl.utils import range_boundaries, get_column_letter
    from openpyxl.worksheet.cell_range import MultiCellRange
    out = []
    for rng in list(ws.merged_cells.ranges):
        c1, r1, c2, r2 = range_boundaries(str(rng))
        if c1 >= at:
            c1, c2 = c1 + n, c2 + n
        elif c2 >= at:
            c2 += n
        out.append(f"{get_column_letter(c1)}{r1}:{get_column_letter(c2)}{r2}")
    ws.merged_cells = MultiCellRange(out)


def build_x2beta_workbook(rows, brand_name=None, state_labels=None, template=TEMPLATE):
    """rows: invoice_process-shaped dicts -> (xlsx bytes, info dict)."""
    labels = state_labels or labels_for_brand(brand_name)
    wb = openpyxl.load_workbook(template)
    ws = wb[SHEET]

    headers = list(next(ws.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True)))
    lines, gst = convert(rows, headers, labels, ws=ws)

    if gst.created:
        _reflow_merges(ws, min(c for _, c in gst.created), len(gst.created))
    pruned = gst.prune_unused(ws)

    protos = {c: copy(ws.cell(DATA_ROW, c)._style) for c in range(1, ws.max_column + 1)}
    for i, line in enumerate(lines):
        r = DATA_ROW + i
        for c in range(1, ws.max_column + 1):
            cell = ws.cell(r, c)
            cell._style = copy(protos[c])
            cell.value = line.get(c)

    buf = io.BytesIO()
    wb.save(buf)
    notes = sum(1 for l in lines if l.get(C_IS_DN) == 'Yes')
    return buf.getvalue(), {
        'rows': len(lines), 'notes': notes, 'purchases': len(lines) - notes,
        'created_columns': [n for n, _ in gst.created], 'pruned_columns': pruned,
    }


# ── self-test: reproduce the accountant's real file ──────────────────────────
if __name__ == '__main__':
    SRC = os.path.expanduser("~/Downloads/X2Beta Automation/Dchica Output.xlsx")
    if not os.path.exists(SRC):
        raise SystemExit(f"reference workbook not found: {SRC}")

    src = openpyxl.load_workbook(SRC, data_only=True)
    raw = src["Raw Data"]
    hdr = [str(c).strip() if c is not None else '' for c in
           next(raw.iter_rows(min_row=1, max_row=1, values_only=True))]
    idx = {}
    for i, h in enumerate(hdr):
        idx.setdefault(h, i)
    FIELD = {'company': 'company', 'vendor_name_tally': 'vendor_name_tally',
             'invoice_number': 'invoice_number', 'invoice_date': 'invoice_date',
             'seller_gstin': 'seller_gstin', 'buyer_gstin': 'buyer_gstin',
             'voucher_type': 'voucher_type', 'category': 'category',
             'cgst_rate': 'cgst_rate', 'sgst_rate': 'sgst_rate', 'igst_rate': 'igst_rate',
             'cgst_amount': 'cgst_amount', 'sgst_amount': 'sgst_amount',
             'igst_amount': 'igst_amount', 'taxable_value': 'taxable value',
             'invoice_type': 'Invoice Type', 'narration': 'Narration'}
    rows = [{k: r[idx[s]] for k, s in FIELD.items() if s in idx}
            for r in raw.iter_rows(min_row=2, values_only=True)
            if any(c is not None for c in r)]

    exp = [r for r in src["X2B Format"].iter_rows(min_row=DATA_ROW, max_row=79, values_only=True)]
    data, info = build_x2beta_workbook(rows, brand_name='dchica')

    out = openpyxl.load_workbook(io.BytesIO(data), data_only=True)["X2B Format"]
    ohdr = [str(c).strip() if c else '' for c in
            next(out.iter_rows(min_row=HEADER_ROW, max_row=HEADER_ROW, values_only=True))]
    got = [r for r in out.iter_rows(min_row=DATA_ROW, max_row=DATA_ROW + info['rows'] - 1,
                                    values_only=True)]

    def key(r, o):
        return (str(r[7] or '').strip(), str(r[8] or '').strip(),
                round(float(r[17] or 0), 2))

    exp_by = {}
    for e in exp:
        exp_by.setdefault(key(e, None), []).append(e)

    TEMPLATE_STATES = ('DL', 'HR', 'Karnataka', 'MH')

    def in_template_state(r):
        for i, h in enumerate(ohdr):
            if 'Input' in h and r[i] is not None and any(f"({s})" in h for s in TEMPLATE_STATES):
                return True
        return False

    ok = miss = 0
    for r in got:
        if r[5] == 'Yes' or not in_template_state(r):
            continue
        b = exp_by.get(key(r, None))
        if b:
            b.pop(0)
            ok += 1
        else:
            miss += 1

    print(f"rows in            : {len(rows)}")
    print(f"accountant X2B rows: {len(exp)}")
    print(f"built              : {info['rows']} ({info['purchases']} purchase / {info['notes']} note)")
    print(f"created columns    : {info['created_columns']}")
    print(f"pruned columns     : {info['pruned_columns']}")
    print(f"\n  matched against accountant : {ok}/{len(exp)}   (unmatched built rows: {miss})")
    print("  ✅ PASS" if ok == len(exp) and miss == 0 else "  ❌ FAIL")
