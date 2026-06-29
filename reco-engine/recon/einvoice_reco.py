"""
E-Invoice Reconciliation — GST Portal E-Invoice Register vs Books (Sales + Credit Note).

Outward-supply (sales) mirror of gstr_2b_books.py. Automates the manual SOP:
  Part A  Books  — read the combined Sales + Credit Note sheet, Taxable = "Value" column,
                   dynamically sum the Output IGST/CGST/SGST ledger heads, pivot to invoice level.
  Part B  Portal — merge the E-Invoice Register "b2b, sez, de" + "cdnr" sheets
                   (Credit/Debit notes negated), pivot to invoice level.
  Part C  Reco   — match by invoice number, lay side by side, subtract → differences.

ALL parsing/matching utilities are REUSED from gstr_2b_books.py — nothing there is modified.
Output workbook mirrors the accountant's working files tab-for-tab:
  Books - Combined | Books - Pivot | Books - After Pivot |
  E-Invoice - Combined | E-Invoice - Pivot | E-Invoice - After Pivot | RECO
"""

from __future__ import annotations

from io import BytesIO
from typing import Any

import pandas as pd

from .core import (
    NormalizedInvoice,
    MatchResult,
    normalize_doc_no,
    parse_date,
    round_money,
    summarize,
)
from .gstr_2b_books import (
    _ensure_xlsx,
    _norm,
    _is_tax_col,
    _find_value_col,
    reconcile_by_invoice_no,
)


# ── Generic sheet reader ───────────────────────────────────────────────────────

def _read_sheet_dicts(data: bytes, sheet_keys: list[str],
                      header_keywords: list[str]) -> list[dict[str, Any]]:
    """Find the first sheet whose name contains any of sheet_keys (case-insensitive),
    locate its header row by scanning the first 15 rows for >=2 header_keywords,
    and return the data rows as list[dict] keyed by the (uniquified) header names."""
    try:
        xl = pd.ExcelFile(BytesIO(data))
    except Exception:
        return []

    target = None
    for sn in xl.sheet_names:
        low = sn.strip().lower()
        if any(k in low for k in sheet_keys):
            target = sn
            break
    if target is None:
        return []

    raw = xl.parse(target, header=None, dtype=object)
    if len(raw) < 2:
        return []

    hdr_idx = -1
    for i in range(min(15, len(raw))):
        cells = [str(x).strip().lower() for x in raw.iloc[i]]
        hits = sum(any(kw in c for c in cells) for kw in header_keywords)
        if hits >= 2:
            hdr_idx = i
            break
    if hdr_idx == -1:
        return []

    raw_headers = [str(v).strip() if str(v) != "nan" else f"_col{i}"
                   for i, v in enumerate(raw.iloc[hdr_idx])]
    seen: dict[str, int] = {}
    headers = []
    for h in raw_headers:
        c = seen.get(h, 0)
        seen[h] = c + 1
        headers.append(f"{h}_{c}" if c else h)

    rows = raw.iloc[hdr_idx + 1:].copy()
    rows.columns = headers
    rows = rows.dropna(how="all")
    return rows.fillna("").to_dict(orient="records")


def _col(row: dict[str, Any], *aliases: str) -> Any:
    """Return the first cell whose normalized header matches/contains an alias."""
    norm_row = {_norm(k): v for k, v in row.items()}
    # exact normalized match first
    for a in aliases:
        na = _norm(a)
        if na in norm_row:
            return norm_row[na]
    # contains fallback
    for a in aliases:
        na = _norm(a)
        for k, v in norm_row.items():
            if na in k:
                return v
    return ""


def _sum_tax(row: dict[str, Any], token: str) -> float:
    """Tax amount for a tax type. If the sheet carries an explicit summary column
    (a bare 'IGST'/'CGST'/'SGST' header — as the accountant appends per SOP A-3),
    use it directly. Otherwise dynamically sum the per-rate ledger heads
    ('Output IGST 18% DL', …). This avoids double-counting when BOTH exist."""
    norm_row = {_norm(k): v for k, v in row.items()}
    if token in norm_row:                      # exact summary column present
        return round_money(norm_row[token])
    total = 0.0
    for k, v in row.items():
        kl = str(k).lower()
        if token == "sgst":
            if "sgst" in kl or "utgst" in kl:
                total += round_money(v)
        elif token in kl:
            total += round_money(v)
    return round(total, 2)


# ── Books parser (Sales + Credit Note, pre-combined) ───────────────────────────

def parse_einvoice_books(data: bytes) -> tuple[list[dict[str, Any]], list[NormalizedInvoice]]:
    """Read the combined Sales + Credit Note workbook. Taxable = 'Value' column;
    IGST/CGST/SGST = dynamic sum of the Output GST ledger heads. Credit Notes are
    already negative in the Combined sheet. Returns (combined_rows, records)."""
    data = _ensure_xlsx(data)

    # Prefer a pre-merged "Combined" sheet; fall back to Sales Reg + Credit Note.
    rows = _read_sheet_dicts(data, ["combined"], ["date", "particulars", "voucher", "value"])
    if not rows:
        rows = []
        for sk, sign in (["sales"], 1.0), (["credit"], -1.0):
            sub = _read_sheet_dicts(data, sk, ["date", "particulars", "voucher", "value"])
            for r in sub:
                r["__sign"] = sign
            rows.extend(sub)

    # Full raw rows for the "Books - Combined" tab — every uploaded data row, all
    # original columns, exactly as given (internal helper keys stripped).
    raw_rows: list[dict[str, Any]] = [
        {str(k): v for k, v in row.items() if not str(k).startswith("__")} for row in rows
    ]

    records: list[NormalizedInvoice] = []
    idx = 0
    for row in rows:
        date_raw = _col(row, "Date")
        doc_date = parse_date(date_raw)
        if not doc_date:
            continue  # skip title / summary / blank rows

        doc_no = str(_col(row, "Voucher No.", "Voucher No", "Invoice Number") or "").strip()
        name = str(_col(row, "Particulars", "Party Name") or "").strip()
        if not doc_no and not name:
            continue

        sign = float(row.get("__sign", 1.0))  # 1.0 if Combined already-negated; -1 for CN fallback
        value_col = _find_value_col(row)
        taxable = round_money(row.get(value_col)) if value_col else 0.0
        igst = _sum_tax(row, "igst")
        cgst = _sum_tax(row, "cgst")
        sgst = _sum_tax(row, "sgst")
        taxable, igst, cgst, sgst = (round(x * sign, 2) for x in (taxable, igst, cgst, sgst))

        doc_type = "CRN" if taxable < 0 or "credit" in str(_col(row, "Voucher Type")).lower() else "INV"

        idx += 1
        rec = NormalizedInvoice(
            source="Books",
            row_id=f"Books-{idx}",
            supplier_name=name,
            doc_type=doc_type,
            doc_no=doc_no,
            normalized_doc_no=normalize_doc_no(doc_no),
            doc_date=doc_date,
            taxable_value=taxable,
            igst=igst, cgst=cgst, sgst=sgst,
            raw={str(k): v for k, v in row.items() if not str(k).startswith("__")},
        )
        records.append(rec)
    return raw_rows, records


# ── Portal E-Invoice Register parser (b2b/sez/de + cdnr) ───────────────────────

def _portal_date(raw: Any) -> str:
    """Normalise an e-invoice portal date to ISO YYYY-MM-DD so it lines up with the
    books side (which already parses to ISO). The portal writes '03-Sep-2025',
    a format parse_date() does not handle — leaving it unnormalised made every
    matched row falsely flag an 'Invoice Date Mismatch'."""
    from datetime import datetime
    s = str(raw or "").strip()
    if not s:
        return ""
    for fmt in ("%d-%b-%Y", "%d-%b-%y", "%d-%B-%Y"):
        try:
            return datetime.strptime(s, fmt).strftime("%Y-%m-%d")
        except ValueError:
            pass
    return parse_date(s) or s


def parse_einvoice_register(data: bytes) -> tuple[list[dict[str, Any]], list[NormalizedInvoice]]:
    """Merge the 'b2b, sez, de' and 'cdnr' sheets of the E-Invoice Register.
    CDNR (credit/debit notes) amounts are negated. Returns (combined_rows, records)."""
    data = _ensure_xlsx(data)
    kw = ["invoice number", "note number", "taxable value", "receiver name"]
    b2b = _read_sheet_dicts(data, ["b2b"], kw)
    cdnr = _read_sheet_dicts(data, ["cdnr"], kw)

    # For the full-column "E-Invoice - Combined" tab: align CDNR's note columns to the
    # b2b invoice columns and negate CDNR amounts, so b2b + cdnr stack into one table.
    _CDNR_RENAME = {"note number": "Invoice number", "note date": "Invoice date",
                    "note value": "Invoice value"}
    _NEG_TOKENS = ("taxable value", "integrated tax", "central tax", "state/ut tax",
                   "cess", "invoice value", "note value")

    def _maybe_neg(v):
        s = str(v).replace(",", "").replace("₹", "").strip()
        try:
            return -float(s)
        except (ValueError, TypeError):
            return v

    raw_rows: list[dict[str, Any]] = []
    records: list[NormalizedInvoice] = []
    idx = 0
    for rows, sign, dtype, src in ((b2b, 1.0, "INV", "B2B"), (cdnr, -1.0, "CRN", "CDNR")):
        for row in rows:
            doc_no = str(_col(row, "Invoice number", "Note Number") or "").strip()
            date_raw = _col(row, "Invoice date", "Note Date")
            doc_date = _portal_date(date_raw)
            name = str(_col(row, "Receiver Name") or "").strip()
            gstin = str(_col(row, "GSTIN/UIN of Recipient", "GSTIN/UIN", "GSTIN") or "").strip().upper()
            if not doc_no and not name:
                continue

            taxable = round(round_money(_col(row, "Taxable Value")) * sign, 2)
            igst = round(round_money(_col(row, "Integrated Tax")) * sign, 2)
            cgst = round(round_money(_col(row, "Central Tax")) * sign, 2)
            sgst = round(round_money(_col(row, "State/UT Tax")) * sign, 2)

            idx += 1
            rec = NormalizedInvoice(
                source="E-Invoice",
                row_id=f"EINV-{idx}",
                supplier_gstin=gstin,
                supplier_name=name,
                doc_type=dtype,
                doc_no=doc_no,
                normalized_doc_no=normalize_doc_no(doc_no),
                doc_date=doc_date,
                taxable_value=taxable,
                igst=igst, cgst=cgst, sgst=sgst,
                sheet_name="cdnr" if dtype == "CRN" else "b2b",
                raw={str(k): v for k, v in row.items()},
            )
            records.append(rec)

            # full-column display row: Source first, original cols (CDNR renamed + negated)
            disp: dict[str, Any] = {"Source": src}
            for k, v in row.items():
                nk = _norm(k)
                key = _CDNR_RENAME.get(nk, k)
                if sign < 0 and any(t in nk for t in _NEG_TOKENS):
                    v = _maybe_neg(v)
                disp[key] = v
            raw_rows.append(disp)
    return raw_rows, records


# ── Pivot (group by invoice → sum the 4 amounts) ───────────────────────────────

def _pivot(records: list[NormalizedInvoice]) -> list[NormalizedInvoice]:
    """Aggregate records to invoice level. Key = normalized doc no + name + date,
    summing taxable/igst/cgst/sgst. Mirrors the accountant's pivot step."""
    buckets: dict[tuple, NormalizedInvoice] = {}
    order: list[tuple] = []
    for r in records:
        key = (r.normalized_doc_no, _norm(r.supplier_name), str(r.doc_date))
        if key not in buckets:
            buckets[key] = NormalizedInvoice(
                source=r.source, row_id=r.row_id,
                supplier_gstin=r.supplier_gstin, supplier_name=r.supplier_name,
                doc_type=r.doc_type, doc_no=r.doc_no,
                normalized_doc_no=r.normalized_doc_no, doc_date=r.doc_date,
                taxable_value=0.0, igst=0.0, cgst=0.0, sgst=0.0,
                sheet_name=r.sheet_name,
            )
            order.append(key)
        agg = buckets[key]
        agg.taxable_value = round(agg.taxable_value + r.taxable_value, 2)
        agg.igst = round(agg.igst + r.igst, 2)
        agg.cgst = round(agg.cgst + r.cgst, 2)
        agg.sgst = round(agg.sgst + r.sgst, 2)
        if not agg.supplier_gstin and r.supplier_gstin:
            agg.supplier_gstin = r.supplier_gstin
    return [buckets[k] for k in order]


# ── Top-level orchestrator ─────────────────────────────────────────────────────

def reconcile_einvoice_top(einvoice_data: bytes, books_data: bytes,
                           tolerance: float = 1.0):
    """Parse both inputs, pivot each to invoice level, reconcile by invoice number.
    Returns a dict bundle with everything the workbook builder needs."""
    einv_raw_rows, einv_records = parse_einvoice_register(einvoice_data)
    books_raw_rows, books_records = parse_einvoice_books(books_data)

    einv_pivot = _pivot(einv_records)
    books_pivot = _pivot(books_records)

    # Portal = left ("gstr2b" slot), Books = right ("purchase" slot)
    results = reconcile_by_invoice_no(einv_pivot, books_pivot, tolerance=tolerance)

    # Display-only: relabel each result's remark into E-Invoice wording (the base
    # engine writes "2B"-worded strings). Does NOT touch matching / category / amounts —
    # only the suggested_action label shown in the dashboard & DB.
    for r in results:
        remark, _ = _einvoice_category(r)
        r.suggested_action = remark

    return {
        "einv_raw_rows": einv_raw_rows,
        "books_raw_rows": books_raw_rows,
        "einv_pivot": einv_pivot,
        "books_pivot": books_pivot,
        "results": results,
    }


# ── Workbook builder ───────────────────────────────────────────────────────────

def _einvoice_category(result: MatchResult) -> tuple[str, str]:
    """Return (remark, group) for a result, in E-Invoice terminology (not 2B)."""
    g, b = result.gstr2b, result.purchase   # g = E-Invoice side, b = Books side
    if g and not b:
        return "Showing in E-Invoice but Not in Books", "einv_only"
    if b and not g:
        return "Showing in Books but Not in E-Invoice", "books_only"
    # both present
    if result.category in ("Amount Mismatch", "Partially Matched"):
        return result.category, "mismatch"
    return "Matched", "matched"


def build_einvoice_workbook(payload: dict) -> Any:
    """Build the 7-sheet output workbook from a reconcile_einvoice_top() bundle."""
    import openpyxl
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    thin = Side(style="thin", color="FFD0D0D0")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    center = Alignment(horizontal="center", vertical="center")
    right = Alignment(horizontal="right", vertical="center")
    hdr_font = Font(bold=True, color="FFFFFFFF", size=10)
    sec_font = Font(bold=True, color="FFFFFFFF", size=11)
    num_fmt = "#,##0.00"

    FILL_HDR = PatternFill("solid", fgColor="1F3864")
    FILL_EINV = PatternFill("solid", fgColor="2E5A88")
    FILL_BOOKS = PatternFill("solid", fgColor="7A5230")
    FILL_DIFF = PatternFill("solid", fgColor="4A4A4A")
    FILL_MATCH = PatternFill("solid", fgColor="C8E6C9")
    FILL_MISM = PatternFill("solid", fgColor="FFE0B2")
    FILL_EONLY = PatternFill("solid", fgColor="FFF9C4")
    FILL_BONLY = PatternFill("solid", fgColor="F8CBAD")

    wb = openpyxl.Workbook()
    wb.remove(wb.active)

    def _simple_sheet(title: str, headers: list[str], rows: list[dict], keys: list[str]):
        ws = wb.create_sheet(title=title[:31])
        for c, h in enumerate(headers, 1):
            cell = ws.cell(row=1, column=c, value=h)
            cell.font = hdr_font; cell.fill = FILL_HDR; cell.alignment = center; cell.border = border
        for ri, row in enumerate(rows, 2):
            for c, k in enumerate(keys, 1):
                v = row.get(k, "")
                cell = ws.cell(row=ri, column=c, value=v)
                cell.border = border
                if isinstance(v, (int, float)):
                    cell.number_format = num_fmt; cell.alignment = right
        widths = [13, 34, 20, 15, 13, 13, 13]
        for c in range(1, len(headers) + 1):
            ws.column_dimensions[get_column_letter(c)].width = widths[c - 1] if c <= len(widths) else 15
        ws.freeze_panes = "A2"
        return ws

    def _full_sheet(title: str, rows: list[dict]):
        """Write rows with ALL their columns (union of keys, first-seen order) —
        used for the input Combined tabs so nothing is dropped."""
        ws = wb.create_sheet(title=title[:31])
        cols: list[str] = []
        seen: set = set()
        for r in rows:
            for k in r.keys():
                if k not in seen:
                    seen.add(k); cols.append(k)
        if not cols:
            cols = ["(no data)"]
        for c, h in enumerate(cols, 1):
            cell = ws.cell(row=1, column=c, value=str(h))
            cell.font = hdr_font; cell.fill = FILL_HDR; cell.alignment = center; cell.border = border
        for ri, row in enumerate(rows, 2):
            for c, k in enumerate(cols, 1):
                v = row.get(k, "")
                cell = ws.cell(row=ri, column=c, value=v)
                cell.border = border
                if isinstance(v, (int, float)):
                    cell.number_format = num_fmt; cell.alignment = right
        for c, k in enumerate(cols, 1):
            maxlen = len(str(k))
            for row in rows[:200]:
                l = len(str(row.get(k, "")))
                if l > maxlen:
                    maxlen = l
            ws.column_dimensions[get_column_letter(c)].width = min(maxlen + 2, 45)
        ws.freeze_panes = "A2"
        return ws

    def _pivot_rows(records, name_label):
        out = []
        for r in records:
            out.append({
                name_label: r.supplier_name, "Invoice number": r.doc_no, "Date": r.doc_date,
                "Taxable": r.taxable_value, "CGST": r.cgst, "SGST": r.sgst, "IGST": r.igst,
            })
        return out

    # 1-3. Books — Combined shows ALL original input columns
    _full_sheet("Books - Combined", payload["books_raw_rows"])
    books_pivot_rows = _pivot_rows(payload["books_pivot"], "Particulars")
    _simple_sheet("Books - Pivot",
                  ["Particulars", "Voucher No.", "Date", "Taxable", "CGST", "SGST", "IGST"],
                  books_pivot_rows, ["Particulars", "Voucher No.", "Date", "Taxable", "CGST", "SGST", "IGST"])
    _simple_sheet("Books - After Pivot",
                  ["Date", "Particulars", "Voucher No.", "Taxable", "CGST", "SGST", "IGST"],
                  [{"Date": r["Date"], **r} for r in books_pivot_rows],
                  ["Date", "Particulars", "Voucher No.", "Taxable", "CGST", "SGST", "IGST"])

    # 4-6. E-Invoice — Combined shows ALL original portal columns (b2b + cdnr merged)
    _full_sheet("E-Invoice - Combined", payload["einv_raw_rows"])
    einv_pivot_rows = _pivot_rows(payload["einv_pivot"], "Receiver Name")
    _simple_sheet("E-Invoice - Pivot",
                  ["Receiver Name", "Invoice number", "Date", "Taxable", "CGST", "SGST", "IGST"],
                  einv_pivot_rows, ["Receiver Name", "Invoice number", "Date", "Taxable", "CGST", "SGST", "IGST"])
    _simple_sheet("E-Invoice - After Pivot",
                  ["Receiver Name", "Invoice number", "Date", "Taxable", "CGST", "SGST", "IGST"],
                  einv_pivot_rows, ["Receiver Name", "Invoice number", "Date", "Taxable", "CGST", "SGST", "IGST"])

    # 7. RECO
    ws = wb.create_sheet(title="RECO")
    # Row 1: section headers
    sections = [("As Per E-Invoice", 1, 7, FILL_EINV), ("As Per Books Data", 9, 15, FILL_BOOKS),
                ("Difference", 17, 20, FILL_DIFF)]
    for label, c1, c2, fill in sections:
        ws.merge_cells(start_row=1, start_column=c1, end_row=1, end_column=c2)
        cell = ws.cell(row=1, column=c1, value=label)
        cell.font = sec_font; cell.fill = fill; cell.alignment = center
    headers = ["Receiver Name", "Invoice No.", "Date", "Taxable", "CGST", "SGST", "IGST", "",
               "Particulars", "Voucher No.", "Date", "Taxable", "CGST", "SGST", "IGST", "",
               "Taxable", "CGST", "SGST", "IGST", "", "Remark"]
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=2, column=c, value=h)
        cell.font = hdr_font; cell.alignment = center; cell.border = border
        if 1 <= c <= 7: cell.fill = FILL_EINV
        elif 9 <= c <= 15: cell.fill = FILL_BOOKS
        elif 17 <= c <= 20: cell.fill = FILL_DIFF
        elif c == 22: cell.fill = FILL_HDR

    fill_map = {"matched": FILL_MATCH, "mismatch": FILL_MISM,
                "einv_only": FILL_EONLY, "books_only": FILL_BONLY}
    r = 3
    for res in payload["results"]:
        g, b = res.gstr2b, res.purchase
        remark, group = _einvoice_category(res)
        gt = g.taxable_value if g else 0.0; bt = b.taxable_value if b else 0.0
        gc = g.cgst if g else 0.0; bc = b.cgst if b else 0.0
        gs = g.sgst if g else 0.0; bs = b.sgst if b else 0.0
        gi = g.igst if g else 0.0; bi = b.igst if b else 0.0
        rowvals = [
            g.supplier_name if g else "", g.doc_no if g else "", g.doc_date if g else "",
            gt, gc, gs, gi, "",
            b.supplier_name if b else "", b.doc_no if b else "", b.doc_date if b else "",
            bt, bc, bs, bi, "",
            round(gt - bt, 2), round(gc - bc, 2), round(gs - bs, 2), round(gi - bi, 2), "",
            remark,
        ]
        for c, v in enumerate(rowvals, 1):
            cell = ws.cell(row=r, column=c, value=v)
            cell.border = border
            if isinstance(v, (int, float)):
                cell.number_format = num_fmt; cell.alignment = right
            if c in (8, 16, 21):
                continue
            cell.fill = fill_map.get(group, PatternFill())
        r += 1

    widths = {1: 30, 2: 18, 3: 13, 9: 30, 10: 18, 11: 13, 22: 34}
    for c in range(1, 23):
        ws.column_dimensions[get_column_letter(c)].width = widths.get(c, 12)
    ws.freeze_panes = "A3"

    return wb
