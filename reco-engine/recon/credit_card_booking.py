"""
credit_card_booking.py — Credit Card Statement → Tally booking entries.

Produces a two-tab workbook:

  1. "Credit Card" — EVERY extracted entry (Dr and Cr), rendered under the
     statement's OWN column headers (Date / Transaction Details / Merchant
     Category / Amount (Rs.) — whatever that issuer actually prints).
  2. "Working"     — the booked entries. Credit-side rows are excluded (a
     PAYMENT RECEIVED clears the card from the bank statement, so booking it
     here would double-count). Every remaining row is:
         Debit  = the mapped counterparty ledger
         Credit = "<Bank> Credit Card"

PDF extraction is delegated wholesale to `pdf_bank_extractor.extract_bank_statement`
rather than rebuilt. That inherits, for free:
  • the iLovePDF OCR fallback for no-text-layer PDFs (these card statements have
    their fonts flattened to vector outlines — 0 chars, thousands of curves),
  • dynamic column detection + the format-template cache (`format_learn`), so a
    layout is learned once and every later run of BOTH tools reuses it at $0,
  • date-less continuation-line folding — the second line of a wrapped narration,
    and the "10/04/25 20.00 USD" forex line, attach to the parent transaction
    instead of becoming phantom rows. A date printed INSIDE a narration is
    therefore never mistaken for the transaction date.

Ledger mapping mirrors the Universal Bank Statement classifier's layer order
(`new-backend/scripts/classify.py`):

    L0  card rules      — bank charges / excluded payment rows (deterministic)
    L1  learned DB      — per-brand merchant directory, seeded from booked history
    L2  COA fuzzy       — thefuzz against the brand's ledger_master
    L3  Claude          — GenSpark `claude-haiku-4-5`, choosing ONLY from real
                          COA candidates, may abstain
    L4  Suspense        — never guessed

Nothing here mutates the PDF → Bank Statement agent's behaviour; it is called
read-only.
"""
import io
import os
import re
import json
import logging
from datetime import datetime, date

logger = logging.getLogger(__name__)

# The card itself is one side of every entry. Resolved from the statement's
# detected bank name, then snapped to the brand's real COA spelling.
_CARD_LEDGER_SUFFIX = "Credit Card"

# ── Deterministic card rules (L0) ────────────────────────────────────────────
# Issuer-levied charges: these rows carry no merchant and no usable Merchant
# Category (the converter leaves junk there — "1951699391", "0", blank), so the
# narration is the only signal. Matched before any merchant logic.
_BANK_CHARGE_RE = re.compile(
    r"FOREIGN\s*CURRENCY\s*MARKUP"
    r"|DCC\s*TRANSACTION\s*FEE"
    r"|GOODS\s*&?\s*SERVICES?\s*TAX"
    r"|\bGST\s*@"
    r"|LATE\s*PAYMENT\s*(FEE|CHARGE)"
    r"|FINANCE\s*CHARGE"
    r"|INTEREST\s*CHARGE"
    r"|(ANNUAL|JOINING|RENEWAL|OVERLIMIT|CASH\s*ADVANCE)\s*FEE"
    r"|SURCHARGE",
    re.I,
)
_BANK_CHARGE_LEDGER = "Bank charges"

# Card bill payments. Booked from the BANK statement side, so booking them here
# too would double-count the payment. They still appear on the "Credit Card" tab.
_PAYMENT_RE = re.compile(
    r"PAYMENT\s*RECEIVED|PAYMENT\s*-?\s*THANK\s*YOU|AUTO\s*DEBIT\s*PAYMENT",
    re.I,
)

# ── Merchant key extraction ─────────────────────────────────────────────────
# Acquirer / gateway prefixes glued onto the merchant name. The same forms the
# bank classifier's _acquirer_star_merchant handles: RAZ*, PAY*, PYU*, Easebuzz*.
# OCR renders the star as a curly quote often enough that both are accepted.
_GATEWAY_RE = re.compile(
    r"^(RAZ|PAY|PYU|EASEBUZZ|PAYU|BILLDESK|CCAVENUE|INSTAMOJO|CASHFREE)\s*[\*“”‘’\"']\s*",
    re.I,
)

# The "- Ref No: MT2601..." tail carries no merchant identity.
_REF_TAIL_RE = re.compile(r"\s*[-–—]?\s*Ref\s*No\.?\s*:?\s*\S+", re.I)

# Merchant Category phrases appended to the narration by some extractions.
# Matched anywhere (they can repeat, and OCR mangles them: 'Utiity Services',
# 'Transpodation Sewices', 'Retail Outiet Services').
_TRAILING_CATEGORY_RE = re.compile(
    r"\b(?:[A-Za-z]+\s+){0,2}"
    r"(?:Servi[cs]es?|Sewi[cs]es?|Stores?|Outlets?|Outiet|Utilities)\b",
    re.I,
)

# A forex continuation line folded into the narration: "15/01/26 129.00 USD".
_FOREX_LINE_RE = re.compile(
    r"\d{1,2}[/-]\d{1,2}[/-]\d{2,4}\s+[\d,]+\.?\d*\s+[A-Z]{3}\b", re.I
)

# City / country / legal-form tokens that identify a location, not a merchant.
_LOCATION_TOKENS = {
    "IND", "INDIA", "IN", "USA", "US", "GBR", "SGP", "CA", "NY", "DEL", "MAH",
    "KAR", "HAR", "TN", "UP", "WB", "GJ", "MUMBAI", "MUMBA", "BANGALORE",
    "BENGALURU", "GURGAON", "GURUGRAM", "DELHI", "NEW", "NOIDA", "CHENNAI",
    "HYDERABAD", "PUNE", "KOLKATA", "AHMEDABAD", "JAIPUR", "SAN", "FRANCISCO",
    "MATEO", "JOSE", "IRVINE", "SEATTLE", "LONDON", "SINGAPORE", "DUBLIN",
    "DUB", "KAUNAS", "BRUSSELS", "YORK", "SOUTH", "WEST", "EAST", "NORTH",
    "SOUTHWESTDELH", "URBANKARNATAK", "WWW", "HTTPS", "HTTP", "COM",
}
_LEGAL_TOKENS = {
    "PVT", "PRIVATE", "LTD", "LIMITED", "LLP", "INC", "LLC", "CO", "CORP",
    "PTE", "PLC", "GMBH", "SA", "BV", "L", "S", "D", "DE",
}
_NOISE_TOKENS = _LOCATION_TOKENS | _LEGAL_TOKENS


def merchant_key(narration: str) -> str:
    """Reduce a card narration to a stable merchant identity key.

        'RAZ*Shopify Commerce S Mumbai IND - Ref No: MT2603...'  ->  'SHOPIFY COMMERCE'
        'M/S. BVC TRADEPORT Mumbai IND - Ref No: MT2601...'      ->  'BVC TRADEPORT'

    Deliberately lossy and deterministic: the same merchant must collapse to the
    same key across months even as the reference number and city suffix change.
    OCR damage inside the merchant name ('FIipkart' for 'Flipkart') is NOT
    repaired here — the learned directory is seeded from the very OCR output the
    accountant booked, so the damaged spelling is itself a known key, and the
    fuzzy layer catches the rest.
    """
    s = str(narration or "")
    s = _REF_TAIL_RE.sub(" ", s)
    s = _FOREX_LINE_RE.sub(" ", s)
    # Strip the Merchant Category. Depending on how a statement is extracted it
    # may or may not arrive glued onto the narration, and leaving it in changes
    # the key for the SAME merchant ('SHIPROCKET' vs
    # 'SHIPROCKET TRANSPORTATION SERVICES'), which silently breaks every
    # directory hit learned from the other extraction path.
    s = _TRAILING_CATEGORY_RE.sub(" ", s)
    s = re.sub(r"\s+", " ", s).strip()
    s = _GATEWAY_RE.sub("", s)
    s = re.sub(r"^M/?S\.?\s+", "", s, flags=re.I)          # "M/S. BVC TRADEPORT"
    s = s.upper()
    s = re.sub(r"[^A-Z0-9 ]+", " ", s)
    toks = [t for t in s.split() if t]
    out = []
    for t in toks:
        if t in _NOISE_TOKENS or t.isdigit():
            continue
        out.append(t)
        if len(out) >= 3:                                   # first 3 distinctive tokens
            break
    return " ".join(out)


def _norm_ledger(s) -> str:
    """Case/space-insensitive ledger comparison key."""
    return re.sub(r"\s+", " ", str(s or "").strip()).lower()


# Directory key types, most specific first. Mirrors bank_payee_directory's
# key_type column (phone / vpa / payee / exact) — a card statement has no phone
# or VPA, so the identity ladder is built from the merchant string instead.
# Every one of these is AUTO-LEARNED from booked history; `keyword` is the extra
# type reserved for hand-added rules.
KEY_TYPES = ("exact", "merchant", "merchant2", "merchant1")
# The merchant-identity types, which share one value namespace when reading.
_MERCHANT_KEY_TYPES = ("merchant", "merchant2", "merchant1")


def extract_keys(narration: str) -> list[tuple[str, str]]:
    """Auto-learn every lookup key a narration supports, most specific first.

    Card statements truncate the merchant to the printed column width, so the
    SAME merchant yields different full keys across months:
        'SHOPFLO SOUTH WEST DE'  -> 'SHOPFLO DE'   vs  'SHOPFLO'
        'BHARTI AIRTEL LIMITED'  -> 'BHARTI AIRTEL LIMI' vs '... LIMIT'
        'AMAZON PAY INDIA PRIVA' -> 'AMAZON PAY PRIVA'   vs 'AMAZON PAY'
    Storing the leading-2 and leading-1 token forms alongside the full key
    collapses those variants onto one identity. `merchant1` is only ever ACCEPTED
    when unambiguous for the brand (see CardClassifier._directory_lookup) — a lone
    generic token must never silently win a merchant it doesn't own.

    Returns [(key_type, key_value), ...]. Writers store ALL of them; the reader
    tries them in this order and takes the first hit.
    """
    out = []
    exact = _REF_TAIL_RE.sub("", str(narration or "")).strip().upper()
    exact = re.sub(r"\s+", " ", exact)
    if exact:
        out.append(("exact", exact))

    full = merchant_key(narration)
    if not full:
        return out
    out.append(("merchant", full))
    toks = full.split()
    seen = {full}
    for n, kt in ((2, "merchant2"), (1, "merchant1")):
        if len(toks) > n:
            cand = " ".join(toks[:n])
            if cand not in seen:
                seen.add(cand)
                out.append((kt, cand))
    return out


# ── Amount / date parsing ───────────────────────────────────────────────────
_AMT_RE = re.compile(r"(-?[\d,]+\.?\d*)")


def parse_amount(value):
    """('1,200.00 Dr') -> (1200.0, 'Dr').  Returns (None, None) when unparseable.

    Direction comes from an explicit Dr/Cr word when present, else the sign.
    """
    if value is None:
        return None, None
    if isinstance(value, (int, float)):
        v = float(value)
        return abs(v), ("Cr" if v < 0 else "Dr")
    s = str(value).strip()
    if not s:
        return None, None
    side = None
    if re.search(r"\bCR\b|\(CR\)", s, re.I):
        side = "Cr"
    elif re.search(r"\bDR\b|\(DR\)", s, re.I):
        side = "Dr"
    # OCR reads a decimal POINT as a comma often enough to matter: on the Feb-26
    # statement '144,00' and '184,00' were the only two amounts the 600 dpi pass
    # got wrong, and they account for the entire 328.00 gap against the
    # statement's own total. A trailing ',dd' with no '.' anywhere is a decimal
    # comma; Indian grouping ('1,14,019.74') always keeps its '.', so this can
    # never misread a thousands separator.
    body = s.replace("₹", "")
    if "." not in body:
        body = re.sub(r",(\d{2})(?!\d)", r".\1", body)

    m = _AMT_RE.search(body)
    if not m:
        return None, None
    try:
        v = float(m.group(1).replace(",", ""))
    except ValueError:
        return None, None
    if side is None:
        side = "Cr" if v < 0 else "Dr"
    return abs(v), side


_DATE_FORMATS = (
    "%d/%m/%Y", "%d-%m-%Y", "%d/%m/%y", "%d-%m-%y",
    "%Y-%m-%d", "%d.%m.%Y", "%d.%m.%y", "%d %b %Y", "%d-%b-%Y", "%d-%b-%y",
)


def parse_date(value):
    """Parse a statement date cell to a `date`, or None.

    Only ever applied to the DATE column. A date appearing inside a narration
    (the '10/04/25 20.00 USD' forex line) is never routed here — those lines
    arrive already folded into their parent transaction by the extractor.
    """
    if value is None or value == "":
        return None
    if isinstance(value, datetime):
        return value.date()
    if isinstance(value, date):
        return value
    s = str(value).strip()
    if not s:
        return None
    s = re.sub(r"\s+00:00:00$", "", s)
    for fmt in _DATE_FORMATS:
        try:
            return datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    # OCR eats the separators in a date and leaves a digit in their place:
    # '02/02/2026' comes back as '0210212026'. On the Feb-26 statement EVERY
    # February row was printed this way, so all 66 of them failed the date check
    # and the month silently disappeared from the output.
    digits = re.sub(r"\D", "", s)
    if len(digits) == 10 and digits[2] in "01" and digits[5] in "01":
        try:
            return date(int(digits[6:10]), int(digits[3:5]), int(digits[0:2]))
        except ValueError:
            pass

    m = re.search(r"(\d{1,2})[/.\-](\d{1,2})[/.\-](\d{2,4})", s)
    if m:
        d, mo, y = (int(x) for x in m.groups())
        y = y + 2000 if y < 100 else y
        try:
            return date(y, mo, d)
        except ValueError:
            return None
    return None


# ── Classifier (mirrors the Universal Bank Statement layer order) ───────────
class CardClassifier:
    """Resolve a card narration to a COA ledger through ordered layers.

    Layer order matches `new-backend/scripts/classify.py`: the learned directory
    outranks everything derivable from the narration, because the correct ledger
    for a merchant is a booking POLICY decision (BSESR → 'Warehouse Electricity
    Charges' for one brand, 'Best Electrical' for another) that no amount of
    string matching can infer.
    """

    # thefuzz score below which a COA match is not trusted on its own.
    FUZZY_ACCEPT = 88
    # Band handed to the LLM to arbitrate rather than accepted or dropped outright.
    FUZZY_ARBITRATE = 70

    def __init__(self, coa=None, directory=None, card_ledger="", llm=None):
        self.coa = [str(c).strip() for c in (coa or []) if str(c or "").strip()]
        self._coa_by_norm = {}
        for name in self.coa:
            self._coa_by_norm.setdefault(_norm_ledger(name), name)

        # directory rows: {'key_type': 'merchant', 'key_value': 'BVC TRADEPORT',
        #                  'ledger': '...'}  — brand-scoped by the caller.
        # Entries whose ledger no longer exists in the COA are dropped, exactly
        # as the bank classifier validates its directory against the master.
        self.directory = {}                       # (key_type, key_value) -> ledger
        self._dropped_directory = []
        for row in (directory or []):
            key = str(row.get("key_value") or "").strip().upper()
            led = str(row.get("ledger") or "").strip()
            if not key or not led:
                continue
            kt = str(row.get("key_type") or "merchant").strip().lower()
            resolved = self.resolve_coa(led)
            if resolved is None and self.coa:
                self._dropped_directory.append((kt, key, led))
                continue
            self.directory.setdefault((kt, key), resolved or led)

        # Value-keyed view of the merchant-family types (see _directory_lookup).
        self._merchant_by_value = {}
        for (kt, kv), led in self.directory.items():
            if kt in _MERCHANT_KEY_TYPES:
                self._merchant_by_value.setdefault(kv, led)

        # merchant1 keys owned by more than one ledger — never accepted.
        self._ambiguous_short = self._find_ambiguous_short_keys()

        self.card_ledger = self.resolve_coa(card_ledger) or card_ledger
        self.llm = llm
        self.llm_calls = 0

    # -- COA ---------------------------------------------------------------
    def resolve_coa(self, name):
        """Snap a ledger literal to the COA's exact spelling, or None.

        This is what makes the output import into Tally cleanly: the historical
        working files say 'Bank Charges' while the COA says 'Bank charges', and
        only the COA spelling is a real ledger.
        """
        if not name:
            return None
        hit = self._coa_by_norm.get(_norm_ledger(name))
        if hit:
            return hit
        return None if self.coa else str(name).strip()

    def _find_ambiguous_short_keys(self):
        """merchant1 keys the brand has booked to more than one ledger.

        'AMAZON' is safe for a brand that only ever books Amazon Pay, and unsafe
        for one that splits Amazon Ads / Amazon Storage Fee / Amazon Pay. So this
        is decided per brand from that brand's own history, never hardcoded.
        """
        seen, bad = {}, set()
        for (kt, key), led in self.directory.items():
            head = key.split()[0] if key.split() else ""
            if not head:
                continue
            if head in seen and _norm_ledger(seen[head]) != _norm_ledger(led):
                bad.add(head)
            seen.setdefault(head, led)
        return bad

    # -- Layers ------------------------------------------------------------
    def _card_rules(self, narration):
        """L0 — issuer charges. Deterministic, outranks the directory.

        Anchored to the START of the narration on purpose. A charge row always
        BEGINS with the charge name, whereas a merchant row can end up carrying
        one: on the Apr-25 OCR a 'DCC TRANSACTION FEE' line folded into the
        FRESHWORKS row above it, and an unanchored match booked a real vendor's
        ₹3,597 to Bank charges.
        """
        if _BANK_CHARGE_RE.match((narration or "").lstrip()):
            led = self.resolve_coa(_BANK_CHARGE_LEDGER) or _BANK_CHARGE_LEDGER
            return {"ledger": led, "layer": "Card Rule", "confidence": "High",
                    "rule": "Bank charge"}
        return None

    def _directory_lookup(self, narration):
        """L1 — per-brand auto-learned keyword directory, most specific key first."""
        for kt, key in extract_keys(narration):
            if kt == "merchant1" and key in self._ambiguous_short:
                continue                      # ambiguous for THIS brand — don't guess
            led = self.directory.get((kt, key))
            if led is None and kt in _MERCHANT_KEY_TYPES:
                # The merchant types are ONE namespace for reading. The same value
                # is stored as `merchant` when it was a whole key and looked up as
                # `merchant1` when it is now a prefix, so demanding an identical
                # key_type made every such hit miss (measured: SHIPROCKET, BSESR
                # and HELIUM10 all fell to Suspense despite being in the
                # directory).
                led = self._merchant_by_value.get(key)
            if led:
                return {"ledger": led, "layer": "Learned DB", "confidence": "High",
                        "rule": f"{kt}[{key}]"}
        # A hand-added keyword rule matches anywhere in the narration, unlike the
        # auto-learned identity keys which are anchored to the merchant.
        up = str(narration or "").upper()
        for (kt, key), led in self.directory.items():
            if kt == "keyword" and key in up:
                return {"ledger": led, "layer": "Learned DB", "confidence": "High",
                        "rule": f"keyword[{key}]"}
        return None

    def _coa_fuzzy(self, narration):
        """L2 — fuzzy match the merchant identity against real COA ledgers.

        Raw token_set_ratio is not safe on a chart of accounts: a COA contains
        grouped/sub ledgers ('13. Storage Fee Amazon') and promotional buckets
        ('Business Promotion') that share tokens with a merchant name without
        being that merchant. Two guards fix the failures this actually produced:

          • ANCHOR — the merchant's leading token must appear in the ledger, and
            a ledger that STARTS with it is preferred over one that merely
            contains it. 'AMAZON' then reaches 'Amazon Pay …', not
            '13. Storage Fee Amazon'.
          • MARGIN — a win over the runner-up that is too narrow means the COA
            holds several equally plausible ledgers; that is a judgement call, so
            it is handed to the LLM/Suspense rather than guessed.
        """
        if not self.coa:
            return None, 0
        query = merchant_key(narration)
        if not query or len(query) < 3:
            return None, 0
        try:
            from thefuzz import process, fuzz
        except ImportError:
            logger.warning("thefuzz unavailable — COA fuzzy layer disabled")
            return None, 0

        cands = process.extract(query, self.coa, scorer=fuzz.token_set_ratio, limit=12)
        if not cands:
            return None, 0

        head = query.split()[0].lower()

        def anchored(name):
            low = re.sub(r"[^a-z0-9 ]+", " ", name.lower())
            toks = low.split()
            if not toks:
                return 0
            if toks[0].startswith(head) or head.startswith(toks[0]):
                return 2                       # ledger leads with the merchant
            return 1 if any(t.startswith(head) or head.startswith(t) for t in toks) else 0

        ranked = sorted(
            ((name, score, anchored(name)) for name, score in cands),
            key=lambda r: (r[2], r[1]), reverse=True,
        )
        name, score, anc = ranked[0]
        if anc == 0:
            # Nothing in the COA actually carries this merchant's leading token.
            return None, score
        # Narrow win among equally-anchored ledgers. Two cases look identical to a
        # score comparison but are not:
        #   • 'BigFoot Retail Solutions Private Limited' vs the same name + ' B2B'
        #     — ONE vendor with a sub-ledger. The base ledger is the right answer.
        #   • two unrelated vendors that happen to score alike — a real judgement
        #     call, which belongs to the LLM or the reviewer, not to a tiebreak.
        # Treating the first case as ambiguous cost 9 rows: the correct ledger
        # scored exactly at the accept threshold, the tie clamped it one below,
        # and the row fell to Suspense.
        peers = [r for r in ranked[1:] if r[2] == anc and _norm_ledger(r[0]) != _norm_ledger(name)]
        if peers and score - peers[0][1] < 5:
            best_n = _norm_ledger(name)
            family = [r for r in [ranked[0]] + peers
                      if _norm_ledger(r[0]).startswith(best_n) or best_n.startswith(_norm_ledger(r[0]))]
            if len(family) == len(peers) + 1:
                # All tied names extend one another — same vendor. Take the most
                # general (shortest) and keep the score.
                shortest = min(family, key=lambda r: len(r[0]))
                return shortest[0], score
            return name, min(score, self.FUZZY_ACCEPT - 1)
        return name, score

    def classify_all(self, narrations, contexts=None):
        """Classify a whole statement: deterministic pass, then ONE batched LLM pass.

        Splitting it this way is what keeps L3 affordable. Rows the deterministic
        layers already settled never reach the model, and the rows that do are
        de-duplicated by merchant key first — a statement with 60 Amazon lines and
        3 unknown vendors asks about 3 merchants, not 63 rows.
        """
        results = [self.classify(n) for n in narrations]
        if not self.llm:
            return results

        # Unique unresolved merchants only — 60 Amazon lines and 3 unknown
        # vendors means the model is asked about 3 merchants, not 63 rows.
        contexts = contexts or [{}] * len(narrations)
        pending, by_key = [], {}
        for idx, (narr, res) in enumerate(zip(narrations, results)):
            if res["layer"] != "Suspense":
                continue
            key = merchant_key(narr) or str(narr)
            if key not in by_key:
                ctx = contexts[idx] if idx < len(contexts) else {}
                by_key[key] = len(pending)
                pending.append({
                    "narration": narr,
                    "category": (ctx or {}).get("category") or "",
                    "amount": (ctx or {}).get("amount"),
                    # OCR soup carries no merchant, but the category and amount
                    # survive — hand those over rather than dropping the row, and
                    # tell the model to abstain rather than guess.
                    "garbled": _is_garbled(narr),
                })

        if not pending:
            return results

        try:
            self.llm_calls += 1
            answers = self.llm(pending) or {}
        except Exception as e:                       # never fail a run on the LLM
            logger.warning("Card LLM layer failed: %s: %s", type(e).__name__, str(e)[:160])
            return results

        for i, (narr, res) in enumerate(zip(narrations, results)):
            if res["layer"] != "Suspense":
                continue
            slot = by_key.get(merchant_key(narr) or str(narr))
            if slot is None:
                continue
            led = answers.get(slot)
            resolved = self.resolve_coa(led) if led else None
            if resolved:
                results[i] = {"ledger": resolved, "layer": "Claude",
                              "confidence": "Medium", "rule": "LLM arbitration"}
        return results

    def classify(self, narration):
        narration = str(narration or "").strip()
        if not narration:
            return {"ledger": self.resolve_coa("Suspense") or "Suspense",
                    "layer": "Suspense", "confidence": "Low", "rule": "empty narration"}

        hit = self._card_rules(narration)
        if hit:
            return hit

        hit = self._directory_lookup(narration)
        if hit:
            return hit

        name, score = self._coa_fuzzy(narration)
        if name and score >= self.FUZZY_ACCEPT:
            return {"ledger": name, "layer": "COA Fuzzy", "confidence": "High",
                    "rule": f"fuzzy {score}"}

        # Anything weaker is left Suspense here so classify_all() can hand the
        # whole set to Claude in one call. A fuzzy guess in the arbitration band
        # is recorded as a hint, not applied.
        return {"ledger": self.resolve_coa("Suspense") or "Suspense",
                "layer": "Suspense", "confidence": "Low",
                "rule": f"no match (best fuzzy {name!r} @ {score})" if name else "no match"}

    def _llm_candidates(self, narration, k=15):
        """Top-k REAL COA ledgers for the model to choose among."""
        if not self.coa:
            return []
        query = merchant_key(narration)
        try:
            from thefuzz import process, fuzz
        except ImportError:
            return []
        got = process.extract(query, self.coa, scorer=fuzz.token_set_ratio, limit=k)
        return [g[0] for g in got]


# ── L3 transport — Claude via GenSpark's OpenAI-compatible proxy ────────────
# Raw urllib, no SDK, consistent with ilovepdf_ocr.py and the bank classifier.
# Keys come from env or new-backend/.env, same fallback chain as the rest of the
# engine. Cost is O(distinct unresolved merchants) — resolved rows never reach
# here, and identical narrations are asked once — NOT O(statement rows).
_GSK_DEFAULT_BASE = "https://www.genspark.ai/api/llm_proxy/v1"
_GSK_MODEL = "claude-haiku-4-5"
_LLM_UA = "curl/8.7.1"      # urllib's default User-Agent trips Cloudflare on GenSpark


def _env(*names):
    for n in names:
        v = os.environ.get(n)
        if v:
            return v.strip()
    here = os.path.dirname(os.path.abspath(__file__))
    for cand in (os.path.join(here, "..", "..", "new-backend", ".env"),
                 os.path.join(here, "..", "..", "..", "new-backend", ".env")):
        try:
            with open(cand, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    for n in names:
                        if line.startswith(n + "="):
                            val = line.split("=", 1)[1].strip()
                            if val:
                                return val
        except Exception:
            continue
    return None


_LLM_INSTRUCTIONS = (
    "You map Indian credit-card merchant narrations to ledgers from the company's "
    "Tally chart of accounts, given in full below.\n"
    "Rules:\n"
    "1. Answer ONLY with a ledger copied EXACTLY, character for character, from the "
    "chart of accounts below.\n"
    "2. If no ledger is clearly the right counterparty, answer ABSTAIN. An abstention "
    "is correct and safe; a wrong ledger corrupts the books.\n"
    "3. Narrations are OCR output from a card statement and may be misspelled "
    "('FIipkart' = 'Flipkart', 'Transpodation' = 'Transportation') or truncated to the "
    "printed column width ('BHARTI AIRTEL LIMIT', 'ETIME OFFICE SOFTE').\n"
    "4. Prefer the specific vendor ledger over a generic expense bucket.\n"
    "5. Indian statutory payees map to their statutory ledger, not a vendor "
    "(e.g. a CBDT payment is a tax ledger, not a supplier).\n"
    "Return ONLY a JSON array: "
    "[{\"i\": <index>, \"ledger\": \"<exact ledger or ABSTAIN>\"}]"
)


def _build_cached_block(coa):
    """The byte-stable prefix: instructions + the WHOLE chart of accounts, sorted.

    Sorted order matters — prompt caching only hits when the prefix is byte-identical
    across calls, so the ledger list must not depend on DB row order.
    """
    ledgers = "\n".join(sorted({str(c).strip() for c in coa if str(c or "").strip()}))
    return f"{_LLM_INSTRUCTIONS}\n\nCHART OF ACCOUNTS:\n{ledgers}"


def _llm_post(url, headers, payload, timeout=90):
    import ssl
    import urllib.request
    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        ctx = ssl.create_default_context()
    req = urllib.request.Request(
        url, data=json.dumps(payload).encode("utf-8"), method="POST",
        headers={"Content-Type": "application/json", "User-Agent": _LLM_UA, **headers},
    )
    with urllib.request.urlopen(req, timeout=timeout, context=ctx) as resp:
        return json.loads(resp.read().decode("utf-8"))


def make_llm_batch_resolver(coa, model=None, timeout=90, max_tokens=4096):
    """Return resolve(narrations) -> {index: ledger}, or None if no credential.

    The chart of accounts is sent ONCE, in its own `cache_control` block, and every
    unresolved narration for the run is answered in that SAME call. Two consequences:

      * COST — the COA (Koparo: ~2,780 ledgers) is not re-sent per row. It is a cached
        prefix, so a run pays one cache write and the rows themselves are a few hundred
        tokens. This mirrors `anthropic_classify_batch` in the bank classifier, which
        was written after per-row calls cost 2.14M input tokens on one statement.
      * ACCURACY — the model chooses from the REAL, COMPLETE ledger list rather than a
        fuzzy-shortlist. That is what lets it reach a ledger fuzzy matching would never
        surface, e.g. narration 'PAY*CBDT GURGAON HAR' -> 'TDS on Contract (94C)'.

    Returns {} rather than raising on failure — L3 is an accuracy layer, never a
    dependency: a dead key must degrade the run to Suspense, not break it.
    """
    key = _env("GSK_API_KEY")
    base = _env("GSK_BASE_URL") or _GSK_DEFAULT_BASE
    url, headers, native = f"{base.rstrip('/')}/chat/completions", {}, False
    if key:
        headers["Authorization"] = f"Bearer {key}"
    else:                                        # fall back to Anthropic native
        key = _env("ANTHROPIC_API_KEY")
        if not key:
            return None
        url = "https://api.anthropic.com/v1/messages"
        headers = {"x-api-key": key, "anthropic-version": "2023-06-01"}
        native = True

    chosen = model or _env("CC_LLM_MODEL") or _GSK_MODEL
    cached_block = _build_cached_block(coa)
    coa_norm = {_norm_ledger(c): str(c).strip() for c in coa if str(c or "").strip()}
    stats = {"calls": 0, "usage": []}

    def resolve(narrations):
        if not narrations:
            return {}
        lines = []
        for i, item in enumerate(narrations):
            if isinstance(item, dict):
                bits = [f"{i}. NARRATION: {item.get('narration','')}"]
                if item.get("category"):
                    bits.append(f"   MERCHANT CATEGORY: {item['category']}")
                if item.get("amount") is not None:
                    bits.append(f"   AMOUNT: {item['amount']}")
                if item.get("garbled"):
                    bits.append("   NOTE: the merchant name is OCR-DAMAGED and may be "
                                "unreadable. Use the merchant category as a hint. If you "
                                "cannot identify the merchant with confidence, ABSTAIN.")
                lines.append("\n".join(bits))
            else:
                lines.append(f"{i}. NARRATION: {item}")
        ask = "Map each narration to a ledger.\n\n" + "\n\n".join(lines)
        try:
            stats["calls"] += 1
            if native:
                data = _llm_post(url, headers, {
                    "model": chosen, "max_tokens": max_tokens,
                    "messages": [{"role": "user", "content": [
                        {"type": "text", "text": cached_block,
                         "cache_control": {"type": "ephemeral"}},
                        {"type": "text", "text": ask},
                    ]}],
                }, timeout)
                text = next((b.get("text", "") for b in data.get("content", [])
                             if b.get("type") == "text"), "")
            else:
                data = _llm_post(url, headers, {
                    "model": chosen, "max_tokens": max_tokens, "temperature": 0,
                    "messages": [
                        {"role": "system", "content": [
                            {"type": "text", "text": cached_block,
                             "cache_control": {"type": "ephemeral"}}]},
                        {"role": "user", "content": ask},
                    ],
                }, timeout)
                text = ((data.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
            stats["usage"].append(data.get("usage") or {})
        except Exception as e:
            logger.warning("Card LLM call failed: %s: %s", type(e).__name__, str(e)[:200])
            return {}

        m = re.search(r"\[.*\]", text or "", re.S)
        if not m:
            return {}
        try:
            parsed = json.loads(m.group(0))
        except Exception:
            return {}
        out = {}
        for row in parsed if isinstance(parsed, list) else []:
            try:
                i = int(row.get("i"))
            except (TypeError, ValueError, AttributeError):
                continue
            led = str(row.get("ledger") or "").strip()
            if not led or led.upper() == "ABSTAIN" or not (0 <= i < len(narrations)):
                continue
            # Containment guard: the model may ONLY return a real COA ledger.
            # Anything else is hallucinated and is dropped to Suspense.
            real = coa_norm.get(_norm_ledger(led))
            if real:
                out[i] = real
        return out

    resolve.stats = stats
    return resolve


# ── Statement chrome ────────────────────────────────────────────────────────
# A card statement PDF is not just a table: it opens with an Overview / Statement
# Summary block (Previous Balance, Total Amount Due, Credit Limit, cardholder
# name and address) and closes with a footer after "End of the Statement".
# Neither is a transaction. This matters more than it looks: a footer line lands
# directly after the last transaction and, being date-less, would otherwise be
# FOLDED INTO that transaction's narration by the continuation-line logic and
# corrupt its merchant key.
_END_OF_STATEMENT_RE = re.compile(
    r"end\s*of\s*(the\s*)?statement|"
    r"this\s+is\s+a\s+computer\s+generated\s+statement",
    re.I,
)
_CHROME_RE = re.compile(
    r"previous\s*balance|total\s*amount\s*due|minimum\s*amount\s*due|"
    r"statement\s*(period|date|summary)|payment\s*due\s*date|"
    r"credit\s*limit|available\s*(credit|cash)|cash\s*limit|"
    r"opening\s*balance|closing\s*balance|"
    r"your\s*name\s*&?\s*address|registered\s*(mobile|email)|"
    r"reward\s*points?|points?\s*(earned|balance)|"
    r"yes\s*touch|phonebanking|sms\s*[\"“']?help|customer\s*care|"
    r"program\s*administrator|overview",
    re.I,
)


# Footer text can arrive INSIDE a narration rather than on its own line: the
# column extractor attaches words to a transaction by position, so a footer
# printed level with the last row lands in that row's narration cell. Dropping
# chrome lines therefore isn't enough — the narration itself has to be cut at
# the first footer marker. Left uncut this both looks wrong in the Working tab
# and poisons the `exact` key learned from that row.
# Markers are chosen to be impossible inside a real merchant name.
_FOOTER_MARKER_RE = re.compile(
    r"<\s*CUSTID\s*>"
    r"|\bSMS\s*[\"“'']?\s*Help"
    r"|\bYES\s*TOUCH\b"
    r"|\bPhone\s*Banking\b"
    r"|\bCustomer\s*Care\b"
    r"|\bToll\s*Free\b"
    r"|\bPage\s*\d+\s*[o0]f\s*\d+"
    r"|\b[LU]\d{5}[A-Z]{2}\d{4}[A-Z]{3}\d{6}\b"          # CIN
    r"|@[\w.]+\.bank\.[a-z]{2,3}\b",
    re.I,
)


def _cut_footer(narration: str) -> str:
    """Truncate a narration at the first statement-footer marker."""
    m = _FOOTER_MARKER_RE.search(str(narration or ""))
    if not m:
        return narration
    # Trim the dangling connector word the footer was joined on ("… Services space").
    return re.sub(r"\s+\S{1,6}$", "", narration[:m.start()].rstrip()).strip() or narration[:m.start()].strip()


def _is_chrome(narration: str) -> bool:
    """A statement-furniture line rather than a transaction."""
    n = str(narration or "").strip()
    if not n:
        return False
    return bool(_CHROME_RE.search(n)) and not re.search(r"Ref\s*No", n, re.I)


def _strip_statement_chrome(rows: list) -> list:
    """Drop the opening summary block and everything after the end marker."""
    out = []
    for row in rows:
        narr = row.get("narration") or ""
        if _END_OF_STATEMENT_RE.search(narr):
            break                                  # nothing below this is a txn
        if _is_chrome(narr):
            continue
        cleaned = _cut_footer(narr)
        if cleaned != narr:
            row["narration"] = cleaned
            if isinstance(row.get("cells"), dict) and "narration" in row["cells"]:
                row["cells"]["narration"] = cleaned
        out.append(row)
    return out


# ── Input: Excel path ───────────────────────────────────────────────────────
# Header-driven, never sheet-name or row-index driven. Real files in the wild use
# sheets called 'Table 1', 'Table 2', 'Table 11', 'Credit Card', 'Sheet1',
# 'Sheet2' and 'Working', with the header on row 1 or row 2 and blank spacer
# columns between the real ones.
_HDR_DATE = ("date", "txn date", "transaction date", "tran date")
_HDR_NARR = ("transaction details", "description", "narration", "particulars",
             "transaction description", "details")
_HDR_AMT = ("amount (rs.)", "amount(rs.)", "amount rs", "amount", "amt")
_HDR_CAT = ("merchant category", "category", "merchant category code")


def _hdr_match(cell, wanted):
    c = re.sub(r"\s+", " ", str(cell or "").strip().lower())
    return any(c == w for w in wanted)


def _read_excel_row(raw, rows, i_date, i_narr, i_amt, i_cat):
    """Parse one worksheet row into `rows`. Returns False at end-of-statement."""
    def at(i):
        return raw[i] if i is not None and i < len(raw) else None

    d_raw, narr, amt_raw = at(i_date), at(i_narr), at(i_amt)
    narr = re.sub(r"\s+", " ", str(narr or "")).strip()
    d = parse_date(d_raw)
    amount, side = parse_amount(amt_raw)

    # PDF→Excel converters emit RAGGED rows: the same sheet can carry the
    # amount one column left of where the header sits (measured on the real
    # Feb-26 file — header at col 12, half the rows at col 11). Trusting the
    # header index alone silently produced amount=None on those rows. So when
    # the header position yields nothing, scan for the value instead.
    if amount is None:
        for j in range(len(raw) - 1, -1, -1):
            if j == i_narr:
                continue
            v, s = parse_amount(raw[j])
            if v is not None and (isinstance(raw[j], (int, float))
                                  or re.search(r"\d", str(raw[j] or ""))):
                amount, side, amt_raw = v, s, raw[j]
                break

    if _END_OF_STATEMENT_RE.search(narr):
        return False                           # footer follows — stop this table

    # A date-less line with no amount is a wrapped narration / forex line —
    # fold it into the row above instead of emitting a phantom transaction.
    # Statement furniture must NEVER be folded: it would silently corrupt the
    # merchant key of the last real transaction.
    if d is None and amount is None:
        if narr and rows and not _is_chrome(narr):
            rows[-1]["narration"] = (rows[-1]["narration"] + " " + narr).strip()
            rows[-1]["cells"]["narration"] = rows[-1]["narration"]
        return True
    if not narr and amount is None:
        return True
    if _is_chrome(narr):
        return True

    cells = {"date": d, "narration": narr, "amount": amt_raw}
    if i_cat is not None:
        cat = at(i_cat)
        if cat in (None, ""):
            # Same ragged-row problem as the amount: the Merchant Category
            # drifts a column too. Take the last text cell before the amount
            # that isn't the narration.
            for j in range(len(raw) - 1, -1, -1):
                if j == i_narr:
                    continue
                v = raw[j]
                if isinstance(v, str) and v.strip() and parse_amount(v)[0] is None:
                    cat = v.strip()
                    break
        cells["category"] = cat
    rows.append({"cells": cells, "date": d, "narration": narr,
                 "amount": amount, "side": side})

    return True


def extract_from_excel(content: bytes) -> dict:
    """Parse a converted credit-card statement workbook into the common shape.

    Returns the same contract as the PDF path so downstream code is identical:
        {columns: [{key, header}], rows: [{cells, date, narration, amount, side}]}
    """
    import openpyxl

    wb = openpyxl.load_workbook(io.BytesIO(content), data_only=True)

    # Collect EVERY transaction table in the workbook, not just the best one.
    # PDF→Excel converters split a long statement across several sheets
    # ('Table 1' + 'Table 2'), and reading only one silently drops the rest —
    # on the real May-25 and Jun-25 files that lost more than half the
    # statement, which the control-total gate then correctly refused.
    tables = []
    for ws in wb.worksheets:
        for r in range(1, min(ws.max_row, 15) + 1):
            cells = [c.value for c in ws[r]]
            i_date = next((i for i, c in enumerate(cells) if _hdr_match(c, _HDR_DATE)), None)
            i_narr = next((i for i, c in enumerate(cells) if _hdr_match(c, _HDR_NARR)), None)
            i_amt = next((i for i, c in enumerate(cells) if _hdr_match(c, _HDR_AMT)), None)
            if i_date is None or i_narr is None or i_amt is None:
                continue
            i_cat = next((i for i, c in enumerate(cells) if _hdr_match(c, _HDR_CAT)), None)
            booked = any(_hdr_match(c, ("debit", "credit")) for c in cells)
            tables.append({"ws": ws, "hrow": r, "i_date": i_date, "i_narr": i_narr,
                           "i_amt": i_amt, "i_cat": i_cat, "booked": booked})
            break                      # one header per sheet

    if not tables:
        return {"columns": [], "rows": [], "error":
                "No credit-card transaction table found — expected columns "
                "Date, Transaction Details and Amount."}

    # Prefer the raw extraction sheets, but only while they are actually
    # complete. On the real May-25 file the converter captured just 11 rows as a
    # table while the accountant's own sheet held all 40, so preferring "raw"
    # unconditionally threw away three quarters of the statement.
    # Taking rows from a booked sheet inherits nothing: only date, narration and
    # amount are read, and the ledger is re-derived from scratch either way.
    raw = [t for t in tables if not t["booked"]]
    booked_tables = [t for t in tables if t["booked"]]

    def _row_span(ts):
        return sum(max(0, t["ws"].max_row - t["hrow"]) for t in ts)

    chosen = raw or booked_tables
    if raw and booked_tables and _row_span(booked_tables) > _row_span(raw):
        chosen = booked_tables
    best = chosen[0]

    ws, hrow = best["ws"], best["hrow"]
    i_date, i_narr, i_amt, i_cat = best["i_date"], best["i_narr"], best["i_amt"], best["i_cat"]
    headers = [c.value for c in ws[hrow]]
    columns = [{"key": "date", "header": str(headers[i_date] or "Date")},
               {"key": "narration", "header": str(headers[i_narr] or "Transaction Details")}]
    if i_cat is not None:
        columns.append({"key": "category",
                        "header": str(headers[i_cat] or "Merchant Category")})
    columns.append({"key": "amount", "header": str(headers[i_amt] or "Amount (Rs.)")})

    rows = []
    for tbl in chosen:
        t_ws, t_hrow = tbl["ws"], tbl["hrow"]
        i_date, i_narr, i_amt, i_cat = tbl["i_date"], tbl["i_narr"], tbl["i_amt"], tbl["i_cat"]
        for raw in t_ws.iter_rows(min_row=t_hrow + 1, values_only=True):
            if not _read_excel_row(raw, rows, i_date, i_narr, i_amt, i_cat):
                break

    # The issuer name usually sits on a COVER sheet, not the transaction sheet,
    # so sample text from the whole workbook for bank detection.
    hint = []
    for sheet in wb.worksheets:
        for row_cells in sheet.iter_rows(min_row=1, max_row=14, values_only=True):
            for v in row_cells:
                if isinstance(v, str) and v.strip():
                    hint.append(v.strip())
        if len(hint) > 400:
            break

    return {"columns": columns, "rows": _strip_statement_chrome(rows),
            "source_sheet": ws.title, "header_row": hrow,
            "bank_hint": " | ".join(hint[:120]),
            "statement_text": " | ".join(hint)}


# Canonical card-statement column names. The tab must show the statement's OWN
# headers, but OCR on a flattened-vector PDF returns things like
# "your outstanding Statement ‘ Date ‘" — that is OCR damage, not the real name.
# When a mangled header still contains a recognisable card-statement column name,
# show the clean form; otherwise leave whatever the statement actually printed.
_CANON_HEADERS = (
    (re.compile(r"transaction\s*details", re.I), "Transaction Details"),
    (re.compile(r"merchant\s*category", re.I), "Merchant Category"),
    (re.compile(r"amount", re.I), "Amount (Rs.)"),
    (re.compile(r"\bdate\b", re.I), "Date"),
    (re.compile(r"description|particulars|narration", re.I), "Transaction Details"),
)
_TAG_FALLBACK = {"date": "Date", "narr": "Transaction Details", "amount": "Amount (Rs.)"}


def _clean_header(header: str, tag: str = None) -> str:
    h = re.sub(r"\s+", " ", str(header or "")).strip()
    for rx, canon in _CANON_HEADERS:
        if rx.search(h):
            return canon
    if tag in _TAG_FALLBACK:
        return _TAG_FALLBACK[tag]
    return h or (tag or "Column").title()


# Words that are statement furniture rather than a merchant identity. A narration
# built only from these (e.g. 'Ref No: Transportation Services') carries no
# merchant and means the extraction picked the wrong column.
_GENERIC_NARR_TOKENS = {
    "SERVICES", "SERVICE", "SEWICES", "STORES", "STORE", "OUTLET", "OUTLETS",
    "OUTIET", "RETAIL", "MISCELLANEOUS", "TRANSPORTATION", "TRANSPODATION",
    "BUSINESS", "UTILITY", "UTIITY", "UTILITIES", "GOVERNMENT", "CARD",
    "STATEMENT", "REF", "NO", "CATEGORY", "MERCHANT", "DETAILS", "AMOUNT",
}


def _narration_quality(rows) -> int:
    """How many rows carry a REAL merchant name.

    Used to choose between the column-geometry and line-based extractions. A
    plain 'has some words' test is not enough: a mis-picked column is full of
    'Transportation Services' / 'Ref No:' text that looks word-like but names no
    merchant, which is exactly how the wrong column won on the Apr-25 OCR.
    """
    good = 0
    for r in rows:
        key = merchant_key(r.get("narration") or "")
        if not key or _is_garbled(r.get("narration") or ""):
            continue
        if any(t not in _GENERIC_NARR_TOKENS for t in key.split()):
            good += 1
    return good


_MAX_OCR_BYTES = 40_000_000      # matches the PDF → Bank agent's own cap
_MAX_OCR_PAGES = 50


def _has_text_layer(content: bytes, password: str = "") -> bool:
    import pdfplumber
    try:
        kw = {"password": password} if password else {}
        with pdfplumber.open(io.BytesIO(content), **kw) as pdf:
            for pg in pdf.pages[:8]:
                if (pg.extract_text() or "").strip():
                    return True
            return False
    except Exception:
        return True                  # unreadable — let the extractor report it


def _ocr_once(content: bytes, password: str = ""):
    """OCR a no-text-layer statement. iLovePDF first, Tesseract as fallback.

    Both engines fail differently and neither dominates, measured on the Feb-26
    statement (102 transactions):

        iLovePDF   53 rows extracted, 2 unmapped  — loses rows, reads merchants well
        Tesseract  99 rows extracted, 29 unmapped — finds rows, mangles merchants

    iLovePDF leads because the merchant name is what this agent exists to map:
    a row whose merchant is unreadable becomes manual work every month, and
    garbled text is deliberately barred from the learned directory, so those
    corrections never accumulate. Rows that iLovePDF drops are caught by the
    control-total gate rather than posted, which is the safe failure.

    Tesseract takes over whenever iLovePDF is unavailable or returns nothing —
    it needs `tesseract` and `pdftoppm` on the box, so it is not always there.
    """
    if len(content) > _MAX_OCR_BYTES:
        logger.warning("Card PDF too large to OCR (%.1f MB)", len(content) / 1e6)
        return None

    try:
        from recon import ilovepdf_ocr
        out = ilovepdf_ocr.ocr_pdf(content, filename="card-statement.pdf")
        if out:
            return out
        logger.warning("iLovePDF returned nothing — trying Tesseract.")
    except Exception as e:
        logger.warning("iLovePDF OCR failed (%s: %s) — trying Tesseract.",
                       type(e).__name__, str(e)[:160])

    try:
        from recon import tesseract_ocr
        if tesseract_ocr.available():
            return tesseract_ocr.ocr_pdf(content, password=password)
        logger.warning("Tesseract not installed — no OCR fallback available.")
    except Exception as e:
        logger.warning("Tesseract OCR failed: %s: %s", type(e).__name__, str(e)[:160])
    return None


def extract_from_pdf(content: bytes, password: str = "") -> dict:
    """Parse a card statement PDF by delegating to the PDF → Bank Statement engine.

    Called READ-ONLY: this adds no behaviour to that agent and changes none of it.
    Everything it already solves — OCR fallback, dynamic columns, the format
    template cache, continuation-line folding — is inherited as-is.
    """
    from recon.pdf_bank_extractor import extract_bank_statement

    # OCR at most ONCE. extract_bank_statement will OCR internally but does not
    # hand the searchable copy back, and the line-based fallback needs that same
    # text — so when there is no text layer we do the OCR here, keep the bytes,
    # and tell the extractor not to repeat it. Two OCR calls per statement would
    # otherwise be billed for every scanned upload.
    work = content
    if not _has_text_layer(content, password):
        ocr_bytes = _ocr_once(content, password)
        if ocr_bytes:
            work = ocr_bytes

    data = extract_bank_statement(work, password=password, _allow_ocr=(work is content))
    if data.get("error"):
        return {"columns": [], "rows": [], "error": data["error"]}

    cols = data.get("columns") or []
    tag_of = {c["key"]: c.get("tag") for c in cols}
    narr_key = next((c["key"] for c in cols if c.get("tag") == "narr"), None)
    txns = data.get("transactions") or []

    # OCR mangles the header row on these flattened-vector statements, and a
    # mangled header can mis-tag the narration column ('Transaction Details'
    # arrived as 'balance. Details Transaction Details' and was tagged `balance`,
    # leaving every narration blank and every row Suspense). Repair it HERE
    # rather than in the shared extractor: pick the column carrying the most
    # word-like text as the narration. Read-only, local to this agent.
    def _texty(key):
        total = 0
        for t in txns[:60]:
            v = (t.get("cells") or {}).get(key)
            if isinstance(v, str):
                total += len(re.findall(r"[A-Za-z]{3,}", v))
        return total

    if not narr_key or not any(str((t.get("cells") or {}).get(narr_key, "")).strip()
                               for t in txns):
        date_key = next((c["key"] for c in cols if c.get("tag") == "date"), None)
        skip = {date_key} | {c["key"] for c in cols if c.get("tag") == "amount"}
        scored = [(_texty(c["key"]), c["key"]) for c in cols if c["key"] not in skip]
        scored.sort(reverse=True)
        if scored and scored[0][0] > 0:
            narr_key = scored[0][1]
            tag_of[narr_key] = "narr"
            logger.info("Card: recovered narration column %r from mangled headers", narr_key)

    columns = [{"key": c["key"],
                "header": _clean_header(c.get("header") or c["key"], tag_of.get(c["key"]))}
               for c in cols]

    rows = []
    for t in txns:
        cells = dict(t.get("cells") or {})
        debit = float(t.get("debit") or 0)
        credit = float(t.get("credit") or 0)
        amount = debit if debit else credit
        side = "Cr" if credit and not debit else "Dr"
        narration = (cells.get(narr_key) if narr_key else "") or t.get("description") or ""
        rows.append({"cells": cells, "date": parse_date(t.get("date")),
                     "narration": re.sub(r"\s+", " ", str(narration)).strip(),
                     "amount": amount, "side": side})

    # ESCALATION — if column-geometry extraction produced no usable narrations,
    # re-parse the SAME (already-OCR'd) bytes line-wise. Card statements are one
    # transaction per line, so this recovers what a bad header row destroyed.
    usable = _narration_quality(rows)
    # Also re-parse when the column path fails to reach the statement's own
    # total: good-looking narrations say nothing about whether rows are MISSING,
    # and a silently short extraction is the failure that matters most.
    _stated = parse_statement_totals(_pdf_text(work, password)).get("stated_debits")
    _col_dr = round(sum(r["amount"] or 0 for r in rows if r.get("side") != "Cr"), 2)
    _short = bool(_stated and abs(_col_dr - _stated) > 2.0)
    if usable < max(1, int(len(rows) * 0.8)) or _short:
        line_data = parse_card_lines(work, password)
        line_rows = line_data.get("rows") or []
        line_usable = _narration_quality(line_rows)
        logger.info("Card: column path gave %d/%d usable narrations, line path %d/%d",
                    usable, len(rows), line_usable, len(line_rows))
        # Choose between the two extractions by which one COMES CLOSER TO THE
        # STATEMENT'S OWN TOTAL, when the statement prints it. Narration quality
        # was the wrong criterion on its own: on Feb-26 the column path produced
        # cleaner narrations and therefore won, while quietly holding 30 fewer
        # transactions than the line path.
        # Union rather than choose: each pass sees rows the other misses.
        united = _union_extractions(rows, line_rows)
        if len(united) >= max(len(rows), len(line_rows)):
            return {"columns": line_data.get("columns") or columns,
                    "rows": _strip_statement_chrome(united),
                    "statement_text": _pdf_text(work, password),
                    "bank_name": data.get("bank_name") or "",
                    "account_no": data.get("account_no") or "",
                    "period_from": data.get("period_from") or "",
                    "period_to": data.get("period_to") or "",
                    "validation": data.get("validation") or {},
                    "extract_mode": "union"}

        if line_usable > usable or _narration_quality(line_rows) > 0:
            merged = _merge_extractions(rows, line_rows)
            return {**line_data,
                    "statement_text": _pdf_text(work, password),
                    "rows": merged,
                    "bank_name": data.get("bank_name") or "",
                    "account_no": data.get("account_no") or "",
                    "period_from": data.get("period_from") or "",
                    "period_to": data.get("period_to") or "",
                    "validation": data.get("validation") or {},
                    "extract_mode": "line"}

    return {"columns": columns, "rows": _strip_statement_chrome(rows),
            "extract_mode": "column",
            "statement_text": _pdf_text(work, password),
            "bank_name": data.get("bank_name") or "",
            "account_no": data.get("account_no") or "",
            "period_from": data.get("period_from") or "",
            "period_to": data.get("period_to") or "",
            "validation": data.get("validation") or {},
            "_tags": tag_of}


# ── Output workbook ─────────────────────────────────────────────────────────
# Palette follows the app's own brand tokens (#0748EE primary) so a downloaded
# workbook looks like it came from Colonel, not from a default openpyxl sheet.
_C_HEADER_BG = "0748EE"
_C_HEADER_FG = "FFFFFF"
_C_BAND = "F6F8FC"
_C_BORDER = "D8E1F0"
_C_SUSPENSE_BG = "FEF3C7"
_C_SUSPENSE_FG = "92400E"
_C_CR_FG = "047857"
_C_META_FG = "475569"
_FMT_INR = '#,##,##0.00'          # Indian grouping: 1,02,124.00
_FMT_DATE = 'dd-mm-yyyy'


def _style_header(ws, row, ncols):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    fill = PatternFill("solid", fgColor=_C_HEADER_BG)
    font = Font(bold=True, color=_C_HEADER_FG, size=10, name="Calibri")
    align = Alignment(vertical="center", horizontal="left", wrap_text=True)
    thin = Side(style="thin", color=_C_HEADER_BG)
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill, cell.font, cell.alignment = fill, font, align
        cell.border = Border(bottom=thin)
    ws.row_dimensions[row].height = 26
    ws.freeze_panes = ws.cell(row=row + 1, column=1)


def _autosize(ws, widths):
    from openpyxl.utils import get_column_letter
    for i, w in enumerate(widths, start=1):
        ws.column_dimensions[get_column_letter(i)].width = w


def _write_meta(ws, meta, ncols):
    """Statement identity block above the table. Returns the header row index."""
    from openpyxl.styles import Font, Alignment
    title = ws.cell(row=1, column=1, value=meta.get("title") or "Credit Card Statement")
    title.font = Font(bold=True, size=14, color="0F172A", name="Calibri")
    ws.row_dimensions[1].height = 22

    bits = []
    for label, key in (("Bank", "bank_name"), ("Card", "account_no"),
                       ("Period", "period"), ("Ledger", "card_ledger")):
        val = meta.get(key)
        if val:
            bits.append(f"{label}: {val}")
    sub = ws.cell(row=2, column=1, value="   •   ".join(bits))
    sub.font = Font(size=9, color=_C_META_FG, name="Calibri")
    sub.alignment = Alignment(vertical="center")
    return 4


def _finish_table(ws, first_row, last_row, ncols, band=True):
    from openpyxl.styles import PatternFill, Border, Side
    thin = Side(style="thin", color=_C_BORDER)
    band_fill = PatternFill("solid", fgColor=_C_BAND)
    for r in range(first_row, last_row + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(row=r, column=c)
            cell.border = Border(bottom=thin)
            if band and (r - first_row) % 2 == 1 and cell.fill.fgColor.rgb in (None, "00000000"):
                cell.fill = band_fill
    ws.auto_filter.ref = f"A{first_row - 1}:{ws.cell(row=first_row - 1, column=ncols).coordinate}"


def build_workbook(result: dict) -> bytes:
    """Two-tab workbook: 'Credit Card' (everything) + 'Working' (booked entries)."""
    from openpyxl import Workbook
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side

    wb = Workbook()
    meta = result.get("meta") or {}

    # ── Tab 1 — Credit Card: every entry, under the statement's OWN headers ──
    ws = wb.active
    ws.title = "Credit Card"
    columns = result.get("columns") or []
    headers = [c["header"] for c in columns]
    hrow = _write_meta(ws, {**meta, "title": "Credit Card Statement — Extracted"},
                       len(headers))
    for i, h in enumerate(headers, start=1):
        ws.cell(row=hrow, column=i, value=h)
    _style_header(ws, hrow, len(headers))

    r = hrow + 1
    for row in result.get("all_rows") or []:
        cells = row.get("cells") or {}
        for i, col in enumerate(columns, start=1):
            v = cells.get(col["key"])
            cell = ws.cell(row=r, column=i)
            if isinstance(v, (datetime, date)):
                cell.value, cell.number_format = v, _FMT_DATE
            elif isinstance(v, (int, float)):
                cell.value, cell.number_format = v, _FMT_INR
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = None if v is None else str(v)
                cell.alignment = Alignment(vertical="top", wrap_text=(col["key"] != "date"))
            if row.get("side") == "Cr":
                cell.font = Font(color=_C_CR_FG, size=10)
        r += 1
    last = r - 1
    if last >= hrow + 1:
        _finish_table(ws, hrow + 1, last, len(headers))
    # Width is chosen from the HEADER, not the column key. On the PDF path the
    # keys are positional ('c0', 'c1', …), so keying off them silently gave the
    # Transaction Details column the 16-char default and wrapped every narration
    # into a tall vertical stack.
    def _width_for(header: str) -> int:
        h = str(header or "").lower()
        if "detail" in h or "description" in h or "narration" in h or "particular" in h:
            return 62
        if "date" in h:
            return 12
        if "categor" in h:
            return 26
        if "amount" in h or "debit" in h or "credit" in h or "balance" in h:
            return 16
        return 18

    _autosize(ws, [_width_for(c["header"]) for c in columns])

    # ── Check Point ─────────────────────────────────────────────────────────
    # The statement's own arithmetic, shown as working rather than as a verdict.
    # An accountant can compare two numbers and read one difference; that is what
    # earns trust, not a badge that says "verified".
    v = result.get("verification") or {}
    tot_r = last + 2
    ws.cell(row=tot_r, column=1, value="Check Point").font = Font(bold=True, size=11, color="0F172A")
    tot_r += 1

    lines = [
        ("Opening balance", v.get("previous_balance")),
        (f"Total debits (Dr) — {sum(1 for x in (result.get('all_rows') or []) if x.get('side') != 'Cr')} rows",
         v.get("total_debits")),
        (f"Total credits (Cr) — {sum(1 for x in (result.get('all_rows') or []) if x.get('side') == 'Cr')} rows",
         v.get("total_credits")),
        ("Computed closing", v.get("computed_closing")),
        ("Statement says", v.get("total_amount_due")),
        ("Difference", v.get("difference")),
    ]
    for i, (label, val) in enumerate(lines):
        is_last = (i == len(lines) - 1)
        lc = ws.cell(row=tot_r + i, column=1, value=label)
        lc.font = Font(bold=is_last, size=10, color="475569" if not is_last else "0F172A")
        vc = ws.cell(row=tot_r + i, column=2, value=val)
        vc.font = Font(bold=is_last, size=10)
        vc.number_format = _FMT_INR
        if is_last:
            vc.border = Border(top=Side(style="thin", color=_C_BORDER))
            lc.border = Border(top=Side(style="thin", color=_C_BORDER))

    status_row = tot_r + len(lines) + 1
    status = v.get("status")
    text, bg, fg = {
        "verified": ("VERIFIED — every rupee on the statement is accounted for.",
                     "ECFDF5", "047857"),
        "mismatch": (f"NOT VERIFIED — {v.get('reason', '')} No booking entries were produced.",
                     "FEF2F2", "991B1B"),
    }.get(status, ("COULD NOT VERIFY — the statement's summary block was not readable, "
                   "so completeness has not been proven. Check the entries against the PDF.",
                   "FFFBEB", "92400E"))
    sc = ws.cell(row=status_row, column=1, value=text)
    sc.font = Font(bold=True, size=10, color=fg)
    sc.fill = PatternFill("solid", fgColor=bg)
    sc.alignment = Alignment(wrap_text=True, vertical="center")
    ws.merge_cells(start_row=status_row, start_column=1,
                   end_row=status_row, end_column=max(2, len(headers)))
    ws.row_dimensions[status_row].height = 30

    # ── Tab 2 — Working: booked entries only ────────────────────────────────
    ws2 = wb.create_sheet("Working")
    W_HEADERS = ["Date", "Transaction Details", "Merchant Category", "Amount (Rs.)",
                 "Voucher Type", "Debit", "Credit", "Amount", "Mapped By"]
    hrow2 = _write_meta(ws2, {**meta, "title": "Working — Booking Entries"},
                        len(W_HEADERS))
    for i, h in enumerate(W_HEADERS, start=1):
        ws2.cell(row=hrow2, column=i, value=h)
    _style_header(ws2, hrow2, len(W_HEADERS))

    susp_fill = PatternFill("solid", fgColor=_C_SUSPENSE_BG)
    r = hrow2 + 1
    booked = result.get("working_rows") or []
    for row in booked:
        vals = [row.get("date"), row.get("narration"), row.get("category"),
                row.get("amount_text"), row.get("voucher_type"), row.get("debit"),
                row.get("credit"), row.get("amount"), row.get("layer")]
        for i, v in enumerate(vals, start=1):
            cell = ws2.cell(row=r, column=i)
            if isinstance(v, (datetime, date)):
                cell.value, cell.number_format = v, _FMT_DATE
            elif i == 8 and isinstance(v, (int, float)):
                cell.value, cell.number_format = v, _FMT_INR
                cell.alignment = Alignment(horizontal="right")
            else:
                cell.value = None if v is None else str(v)
                cell.alignment = Alignment(vertical="top", wrap_text=(i == 2))
            cell.font = Font(size=10)
        if row.get("is_suspense"):
            for i in range(1, len(W_HEADERS) + 1):
                c = ws2.cell(row=r, column=i)
                c.fill = susp_fill
                c.font = Font(size=10, color=_C_SUSPENSE_FG, bold=(i == 6))
        r += 1
    last2 = r - 1
    if last2 >= hrow2 + 1:
        _finish_table(ws2, hrow2 + 1, last2, len(W_HEADERS))
    _autosize(ws2, [12, 58, 22, 15, 16, 40, 26, 14, 14])

    if result.get("blocked"):
        msg = ws2.cell(row=hrow2 + 1, column=1,
                       value="No booking entries were produced — the statement did not "
                             "reconcile to its own Total Amount Due. See Check Point on the "
                             "Credit Card tab. Do not post from this file.")
        msg.font = Font(bold=True, size=10, color="991B1B")
        msg.alignment = Alignment(wrap_text=True, vertical="center")
        ws2.merge_cells(start_row=hrow2 + 1, start_column=1,
                        end_row=hrow2 + 1, end_column=len(W_HEADERS))
        ws2.row_dimensions[hrow2 + 1].height = 34

    tot2 = last2 + 2
    summary = result.get("summary") or {}
    ws2.cell(row=tot2, column=1, value="Total").font = Font(bold=True, size=10)
    tc = ws2.cell(row=tot2, column=8, value=sum(x.get("amount") or 0 for x in booked))
    tc.font, tc.number_format = Font(bold=True, size=10), _FMT_INR
    lines = [
        ("Booked entries", summary.get("booked", 0)),
        ("Unmapped (Suspense)", summary.get("suspense", 0)),
        ("Credit rows excluded", summary.get("excluded_credits", 0)),
        ("Card payments excluded", summary.get("excluded_payments", 0)),
    ]
    for i, (label, val) in enumerate(lines, start=2):
        ws2.cell(row=tot2 + i, column=1, value=label).font = Font(size=9, color=_C_META_FG)
        ws2.cell(row=tot2 + i, column=2, value=val).font = Font(size=9, bold=True)

    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


# ── Entry point ─────────────────────────────────────────────────────────────
def _card_ledger_for(bank_name: str) -> str:
    bank = re.sub(r"\s+", " ", str(bank_name or "").strip())
    return f"{bank} {_CARD_LEDGER_SUFFIX}".strip() if bank else _CARD_LEDGER_SUFFIX


def _detect_bank_name(filename: str, data: dict) -> str:
    """Best-effort issuer name from the filename or the statement's own text.

    Reuses the PDF → Bank agent's bank-name patterns so the two agents agree on
    spelling ('Yes Bank', not 'YES BANK'). Returns "" when unsure — an empty
    result is handled by the caller, and is far safer than naming the wrong card.
    """
    try:
        from recon.pdf_bank_extractor import _BANK_NAME_PATTERNS
    except Exception:
        return ""
    haystacks = [str(filename or ""), str(data.get("bank_hint") or "")]
    for row in (data.get("rows") or [])[:40]:
        haystacks.append(str(row.get("narration") or ""))
    for col in (data.get("columns") or []):
        haystacks.append(str(col.get("header") or ""))
    blob = " ".join(haystacks).upper()
    for rx, name in _BANK_NAME_PATTERNS:
        if re.search(rx, blob):
            return name
    return ""


def run(content: bytes, filename: str = "", password: str = "", coa=None,
        directory=None, card_ledger: str = "", voucher_type: str = "",
        use_llm: bool = True, allow_unverified: bool = False) -> dict:
    """Book one credit card statement.

    `coa` and `directory` are supplied by the caller (Node reads them from the
    brand's ledger_master / cc_merchant_directory under RLS) — this module never
    touches the database itself, exactly as the other reco agents work.
    """
    is_pdf = str(filename or "").lower().endswith(".pdf") or content[:4] == b"%PDF"
    data = extract_from_pdf(content, password) if is_pdf else extract_from_excel(content)
    if data.get("error"):
        return {"error": data["error"]}
    rows = data.get("rows") or []
    if not rows:
        return {"error": "No transactions found in the statement."}

    coa = list(coa or [])
    bank_name = data.get("bank_name") or ""
    if not bank_name:
        bank_name = _detect_bank_name(filename, data)
    # An explicit card_ledger from the caller always wins: a brand can hold several
    # cards ('Yes Bank Credit Card', 'HDFC Credit Card 4021', …) and only the user
    # knows which statement this is. The derived name is a fallback, not a guess
    # we impose — if it does not resolve to a real COA ledger the run still
    # proceeds and `card_ledger_resolved` tells the UI to ask.
    resolved_card = card_ledger or (_card_ledger_for(bank_name) if bank_name else "")

    llm = make_llm_batch_resolver(coa) if (use_llm and coa) else None
    clf = CardClassifier(coa=coa, directory=directory,
                         card_ledger=resolved_card, llm=llm)

    # Split before classifying: only rows that will actually be booked need a
    # ledger, so excluded rows never cost an LLM token.
    working_src, excluded_credits, excluded_payments = [], 0, 0
    for row in rows:
        if _PAYMENT_RE.match((row.get("narration") or "").lstrip()):
            excluded_payments += 1
            continue
        if row.get("side") == "Cr":
            excluded_credits += 1
            continue
        working_src.append(row)

    def _ctx(row):
        cells = row.get("cells") or {}
        cat = next((cells[k] for k in ("category", "merchant_category")
                    if k in cells and cells[k] not in (None, "")), "")
        return {"category": str(cat or ""), "amount": row.get("amount")}

    # ── Control-total gate ────────────────────────────────────────────────
    # Run BEFORE classification: if the statement doesn't reconcile there is no
    # point paying for an LLM call, and more importantly the run must not
    # produce booking entries. A short extraction that still yields tidy
    # vouchers is the dangerous failure — it looks finished.
    totals = parse_statement_totals(data.get("statement_text") or "")
    verification = verify_statement(rows, totals)
    if verification["status"] == "mismatch" and not allow_unverified:
        logger.warning("Card: statement did NOT reconcile — %s", verification["reason"])
        result = {
            "columns": data.get("columns") or [],
            "all_rows": rows,
            "working_rows": [],          # deliberately empty: nothing may be posted
            "verification": verification,
            "blocked": True,
            "summary": {"extracted": len(rows), "booked": 0, "suspense": 0,
                        "zero_amount": 0, "excluded_credits": excluded_credits,
                        "excluded_payments": excluded_payments, "llm_calls": 0,
                        "by_layer": {}},
            "meta": {"bank_name": bank_name, "account_no": data.get("account_no") or "",
                     "period": " to ".join(x for x in (totals.get("period_from"),
                                                       totals.get("period_to")) if x),
                     "card_ledger": clf.card_ledger, "card_ledger_resolved": False,
                     "extract_mode": data.get("extract_mode") or ""},
            "learned_keys": [],
            "dropped_directory": [],
        }
        result["excel"] = build_workbook(result)
        return result

    outcomes = clf.classify_all([r["narration"] for r in working_src],
                                contexts=[_ctx(r) for r in working_src])

    suspense_name = clf.resolve_coa("Suspense") or "Suspense"
    working_rows, learned, suspense = [], {}, 0
    for row, res in zip(working_src, outcomes):
        is_susp = _norm_ledger(res["ledger"]) == _norm_ledger(suspense_name)
        suspense += 1 if is_susp else 0
        cells = row.get("cells") or {}
        working_rows.append({
            "date": row.get("date"),
            "narration": row.get("narration"),
            "category": next((cells[k] for k in ("category", "merchant_category")
                              if k in cells and cells[k] not in (None, "")), ""),
            "amount_text": f"{row['amount']:,.2f} Dr" if row.get("amount") is not None else "",
            "voucher_type": voucher_type or "",
            "debit": res["ledger"],
            "credit": clf.card_ledger,
            "amount": row.get("amount"),
            "layer": res["layer"],
            "confidence": res.get("confidence"),
            "rule": res.get("rule"),
            "is_suspense": is_susp,
            "no_amount": not row.get("amount"),
        })
        # Auto-learn: anything the run resolved to a REAL ledger becomes a
        # directory candidate keyed every way the narration supports. Suspense
        # and issuer-charge rows teach nothing.
        if not is_susp and res["layer"] not in ("Card Rule",):
            for kt, key in extract_keys(row["narration"]):
                learned.setdefault((kt, key), res["ledger"])

    # Log what the LLM layer actually cost. The chart of accounts is a cached
    # prefix, so a healthy run shows cache_read ≈ the whole COA and cache_write
    # near zero — if cache_write ever equals cache_read the prefix stopped being
    # byte-stable and the COA is being re-billed on every call.
    if llm is not None and getattr(llm, "stats", None):
        st = llm.stats
        read = sum((u.get("cache_read_input_tokens") or 0) for u in st["usage"])
        write = sum((u.get("cache_creation_input_tokens") or 0) for u in st["usage"])
        fresh = sum((u.get("prompt_tokens") or u.get("input_tokens") or 0)
                    for u in st["usage"]) - read - write
        logger.info("Card LLM: %d call(s)  cache_read=%d  cache_write=%d  uncached=%d",
                    st["calls"], read, write, max(0, fresh))

    zero_amount = sum(1 for w in working_rows if not w.get("amount"))
    if zero_amount:
        logger.warning("Card: %d booked row(s) have no amount — the statement's value was "
                       "lost in extraction and they must not be posted as-is", zero_amount)

    summary = {
        "extracted": len(rows),
        "booked": len(working_rows),
        "suspense": suspense,
        "zero_amount": zero_amount,
        "excluded_credits": excluded_credits,
        "excluded_payments": excluded_payments,
        "llm_calls": getattr(llm, "stats", {}).get("calls", 0) if llm else 0,
        "by_layer": {},
    }
    for w in working_rows:
        summary["by_layer"][w["layer"]] = summary["by_layer"].get(w["layer"], 0) + 1

    period = " to ".join(x for x in (data.get("period_from") or totals.get("period_from"),
                                     data.get("period_to") or totals.get("period_to")) if x)
    result = {
        "columns": data.get("columns") or [],
        "all_rows": rows,
        "working_rows": working_rows,
        "verification": verification,
        "blocked": False,
        "summary": summary,
        "meta": {"bank_name": bank_name, "account_no": data.get("account_no") or "",
                 "period": period, "card_ledger": clf.card_ledger,
                 # False => the derived name is NOT a real ledger for this brand;
                 # the UI must ask the user which card this statement belongs to.
                 "card_ledger_resolved": bool(
                     (card_ledger or bank_name) and clf.resolve_coa(clf.card_ledger)),
                 "extract_mode": data.get("extract_mode") or ("excel" if not is_pdf else "column")},
        # [(key_type, key_value, ledger)] for the caller to upsert per brand.
        "learned_keys": [{"key_type": kt, "key_value": kv, "ledger": led}
                         for (kt, kv), led in sorted(learned.items())],
        "dropped_directory": clf._dropped_directory,
    }
    result["excel"] = build_workbook(result)
    return result


# ── Line-based card parser (fallback) ───────────────────────────────────────
# Column-geometry extraction is the primary path, but it depends on locating a
# clean header row. On an OCR'd card statement the header often arrives polluted
# with bleed text from the marketing block above it ("your outstanding balance.
# Details Transaction Details"), which both mis-tags the narration column and
# puts the column x-ranges in the wrong place — leaving every narration blank.
#
# A card statement is, however, strictly ONE TRANSACTION PER LINE:
#     <date>  <merchant narration>  [<merchant category>]  <amount> Dr|Cr
# so parsing the text line-wise sidesteps geometry entirely. Measured on the
# Apr-25 OCR: 0 usable narrations column-wise vs 28 transactions / 23 clean
# narrations line-wise.
_LINE_RE = re.compile(
    r"^(?P<date>\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{10})\s+"
    r"(?P<body>.*?)\s+"
    r"(?P<amt>[\d,]+\.\d{2})\s*(?P<side>Dr|Cr)?\s*$",
    re.I,
)
# A transaction line whose DATE was mangled by OCR ('2000312025' for
# '29/03/2025'). Anchored on the strong signal instead: an amount + Dr/Cr at the
# end of the line, preceded by a leading token that is mostly digits.
_BROKEN_DATE_LINE_RE = re.compile(
    r"^(?P<date>[\d/\-]{6,12})\s+"
    r"(?P<body>.*?)\s+"
    r"(?P<amt>[\d,]+\.\d{2})\s*(?P<side>Dr|Cr)?\s*$",
    re.I,
)

# Merchant Category is a short Title Case service phrase printed after the
# narration. OCR mangles it ('Transpodation Sewices', 'Utiity Services'), so it
# is matched loosely by its shape and its trailing noun.
_CATEGORY_RE = re.compile(
    r"\s+((?:[A-Z][A-Za-z]{2,}\s+){0,2}"
    r"(?:Services?|Sewices?|Stores?|Outlets?|Utilities))\s*$"
)
# A narration OCR'd into glyph soup. Real merchant names are mostly letters,
# digits, spaces and a few separators; anything denser than this is unusable and
# must go to Suspense rather than poison the learned directory.
_GARBLE_RE = re.compile(r"[^A-Za-z0-9 ,./:*&'\-()#]")


def _is_garbled(text: str) -> bool:
    s = str(text or "")
    return len(_GARBLE_RE.findall(s)) > 3


def parse_card_lines(pdf_bytes: bytes, password: str = "") -> dict:
    """Line-wise parse of a card statement. Returns the common extract shape."""
    import pdfplumber

    rows = []
    open_kwargs = {"password": password} if password else {}
    try:
        pdf = pdfplumber.open(io.BytesIO(pdf_bytes), **open_kwargs)
    except Exception as e:
        logger.warning("Line parser could not open PDF: %s", type(e).__name__)
        return {"columns": [], "rows": []}

    with pdf:
        for page in pdf.pages:
            text = page.extract_text() or ""
            try:
                page.flush_cache()
            except Exception:
                pass
            for raw_line in text.split("\n"):
                line = re.sub(r"\s+", " ", raw_line).strip()
                if not line:
                    continue
                if _END_OF_STATEMENT_RE.search(line):
                    return _finish_lines(rows)
                m = _LINE_RE.match(line)
                broken_date = False
                if not m:
                    # OCR mangles separators, so a real transaction's date can
                    # arrive as '2000312025' instead of '29/03/2025'. Such a line
                    # still ENDS in an amount + Dr/Cr, and treating it as a
                    # continuation would fold a whole transaction into the row
                    # above — silently merging two merchants and losing a row.
                    m = _BROKEN_DATE_LINE_RE.match(line)
                    broken_date = bool(m)
                if not m:
                    # Genuine date-less continuation: a wrapped merchant name, or
                    # the "21/03/25 200.00 USD" forex line.
                    if rows and not _is_chrome(line) and not _looks_like_header(line):
                        prev = rows[-1]
                        prev["narration"] = f"{prev['narration']} {line}".strip()
                        prev["cells"]["narration"] = prev["narration"]
                    continue

                body = (m.group("body") or "").strip()
                category = ""
                cm = _CATEGORY_RE.search(body)
                if cm:
                    category = cm.group(1).strip()
                    body = body[: cm.start()].strip()

                amount, side = parse_amount(f"{m.group('amt')} {m.group('side') or 'Dr'}")
                d = parse_date(m.group("date"))
                if d is None and broken_date:
                    # Keep the row; carry the last good date so it still lands in
                    # the right period. Losing the row entirely would be worse.
                    d = rows[-1]["date"] if rows else None
                if _is_chrome(body):
                    continue
                rows.append({
                    "cells": {"date": d, "narration": body,
                              "category": category, "amount": line[-24:]},
                    "date": d, "narration": body,
                    "amount": amount, "side": side or "Dr",
                })
    return _finish_lines(rows)


_HEADER_WORDS_RE = re.compile(
    r"transaction\s*details|merchant\s*category|amount\s*\(rs|credit\s*card\s*statement",
    re.I,
)


def _looks_like_header(line: str) -> bool:
    return bool(_HEADER_WORDS_RE.search(line or ""))


def _finish_lines(rows):
    for r in rows:
        r["cells"]["amount"] = (f"{r['amount']:,.2f} {r['side']}"
                                if r.get("amount") is not None else "")
    return {
        "columns": [{"key": "date", "header": "Date"},
                    {"key": "narration", "header": "Transaction Details"},
                    {"key": "category", "header": "Merchant Category"},
                    {"key": "amount", "header": "Amount (Rs.)"}],
        "rows": _strip_statement_chrome(rows),
    }


def _merge_extractions(col_rows, line_rows):
    """Combine the column-geometry and line-based extractions of one statement.

    They fail in opposite directions on OCR'd card statements:

      * COLUMN path reads each cell from its own x-range, so a merchant wrapped
        across two visual lines comes out intact — but OCR damage to an amount
        ('3.00 Dr' -> '3.000r') can push it outside the amount column's x-range
        and the value is lost as 0.00.
      * LINE path reads the whole text line, so the trailing amount is reliable —
        but a wrapped merchant cell has its two lines overlaid by extract_text()
        into glyph soup.

    So: take the extraction with the more complete AMOUNTS as the base, then
    upgrade any garbled narration from the other where a row matches on date and
    amount. Neither is authoritative on its own.
    """
    def amount_completeness(rows):
        return sum(1 for r in rows if (r.get("amount") or 0) > 0)

    if not col_rows:
        return line_rows
    if not line_rows:
        return col_rows

    if amount_completeness(line_rows) >= amount_completeness(col_rows):
        base, other = line_rows, col_rows
    else:
        base, other = col_rows, line_rows

    index, by_amount = {}, {}
    for r in other:
        amt = r.get("amount")
        if amt is None:
            continue
        key = round(float(amt), 2)
        index.setdefault((r.get("date"), key), []).append(r)
        by_amount.setdefault(key, []).append(r)

    # Recover amounts the base extraction lost. A row that survived with its
    # narration but a 0.00 amount is the dangerous case: it books a zero-value
    # voucher that LOOKS complete. Match the other extraction on date + merchant
    # identity (there is no amount to match on, by definition).
    by_merchant = {}
    for r in other:
        amt = r.get("amount")
        if amt:
            by_merchant.setdefault((r.get("date"), merchant_key(r.get("narration") or "")), []).append(r)

    recovered = 0
    for row in base:
        if row.get("amount"):
            continue
        cands = by_merchant.get((row.get("date"), merchant_key(row.get("narration") or ""))) or []
        if len(cands) == 1:
            row["amount"] = cands[0]["amount"]
            row["side"] = cands[0].get("side") or row.get("side")
            recovered += 1
    if recovered:
        logger.info("Card: recovered %d amount(s) lost by the primary extraction", recovered)

    upgraded = 0
    for row in base:
        amt = row.get("amount")
        if amt is None:
            continue
        key = round(float(amt), 2)
        cands = index.get((row.get("date"), key)) or []
        if not cands:
            # OCR also damages DATES ('29/03/2025' -> '2000312025'), so the two
            # extractions can disagree on a row's date while agreeing on its
            # amount. Fall back to amount alone, but ONLY when that amount is
            # unique in the statement — otherwise two same-value transactions
            # could swap narrations.
            same = by_amount.get(key) or []
            if len(same) == 1:
                cands = same
        if not cands:
            continue
        cur = row.get("narration") or ""
        cur_bad = _is_garbled(cur) or _narration_quality([row]) == 0
        if not cur_bad:
            continue
        for c in cands:
            if _narration_quality([c]) > 0 and not _is_garbled(c.get("narration") or ""):
                row["narration"] = c["narration"]
                row.setdefault("cells", {})["narration"] = c["narration"]
                upgraded += 1
                break
    if upgraded:
        logger.info("Card: upgraded %d garbled narrations from the other extraction",
                    upgraded)
    return base


# ── Control total: the statement's own arithmetic ───────────────────────────
# Every card statement prints a summary block whose numbers must satisfy
#
#     Previous Balance + Total Debits − Total Credits = Total Amount Due
#
# Reproducing the issuer's stated Total Amount Due from the rows we extracted is
# the only real proof we captured the whole statement. Row counts prove nothing —
# a page can go missing and the remaining rows still look perfectly well-formed.
# Measured: the Feb-26 PDF booked ₹8,59,980.94 of ₹11,78,855.58 with 49 rows
# silently dropped, and every extracted row looked correct.
_AMT = r"(?:Rs\.?|₹|INR)?\s*([\d,]+\.\d{2})"

_PREV_BAL_RE = re.compile(r"previous\s*balance\s*:?\s*" + _AMT + r"\s*(Dr|Cr)?", re.I)
_OPEN_BAL_RE = re.compile(r"opening\s*balance\s*:?\s*" + _AMT + r"\s*(Dr|Cr)?", re.I)
# "Minimum Amount Due" must never satisfy this — hence the explicit `total`.
_TOTAL_DUE_RE = re.compile(r"total\s*amount\s*due\s*:?\s*" + _AMT, re.I)
_PERIOD_RE = re.compile(
    r"statement\s*period\s*:?\s*(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})\s*(?:to|-)\s*"
    r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4})", re.I)


def _amount_after(label_re, text, window=160, require_side=None):
    """First amount appearing within `window` chars after a label.

    OCR flattens the statement's two-column summary box, so a label and its value
    routinely end up on different lines with unrelated text between them:

        Previous Balance :
        Statement Period: Credit Limit: Rs. 3,50,161.34 Dr

    Requiring the value to sit immediately after the label therefore finds
    nothing on exactly the scanned statements that most need checking.
    """
    # Try EVERY occurrence of the label, not just the first. Card statements
    # repeat these phrases in their terms-and-conditions boilerplate ("...the
    # Total Amount Due / Minimum Amount Due as may be applicable..."), and on the
    # Tesseract read that boilerplate comes BEFORE the summary box — so taking
    # the first match found no figure at all and reported the statement
    # unverifiable.
    for m in label_re.finditer(text):
        seg = text[m.end(): m.end() + window]
        got = _first_amount(seg, require_side)
        if got[0] is not None:
            return got
    return None, None


def _first_amount(seg, require_side=None):
    if require_side:
        # The period totals always carry a Dr/Cr marker. Requiring it skips the
        # OCR wreckage around them — on a real Feb-26 scan the first "amount"
        # after the purchases label was 17.90, salvaged out of the mangled
        # credit-limit figure "RS, 17.90,80672", while the true 11,78,855.58 Dr
        # sat a few tokens further along.
        a = re.search(_AMT + r"\s*(" + require_side + r")\b", seg, re.I)
    else:
        a = re.search(_AMT + r"\s*(Dr|Cr)?", seg)
    if not a:
        return None, None
    return float(a.group(1).replace(",", "")), (a.group(2) or "").title()


_PREV_BAL_LBL = re.compile(r"previous\s*balance\s*:?", re.I)
_OPEN_BAL_LBL = re.compile(r"opening\s*balance\s*:?", re.I)
_TOTAL_DUE_LBL = re.compile(r"total\s*amount\s*due\s*:?", re.I)
# The statement also prints its OWN period totals. When present these are the
# strongest evidence available: they are the issuer's own sum of the very rows
# we are trying to reproduce, so a mismatch localises the failure to extraction
# rather than to the balance arithmetic.
_PURCHASES_LBL = re.compile(
    r"current\s*purchases?\b[^:]{0,60}?(?:charges)?\s*:?", re.I)
_PAYMENTS_LBL = re.compile(r"payments?\s*&?\s*credi[^:]{0,20}recei[^:]{0,10}:?", re.I)


def parse_statement_totals(text: str) -> dict:
    """Pull the summary block's control figures out of statement text.

    Returns keys that are None when absent — a missing figure means "cannot
    verify", which must be reported as such and never as "verified".
    """
    t = re.sub(r"[ \t]+", " ", str(text or ""))

    prev, prev_side = _amount_after(_PREV_BAL_LBL, t)
    if prev is None:
        prev, prev_side = _amount_after(_OPEN_BAL_LBL, t)
    if prev is not None and prev_side.lower() == "cr":
        prev = -prev

    due, _ = _amount_after(_TOTAL_DUE_LBL, t)
    stated_dr, _ = _amount_after(_PURCHASES_LBL, t, window=220, require_side="Dr")
    stated_cr, _ = _amount_after(_PAYMENTS_LBL, t, window=220, require_side="Cr")

    m = _PERIOD_RE.search(t)
    return {
        "previous_balance": prev,
        "total_amount_due": due,
        "stated_debits": stated_dr,
        "stated_credits": stated_cr,
        "period_from": m.group(1) if m else None,
        "period_to": m.group(2) if m else None,
    }


def verify_statement(rows: list, totals: dict, tolerance: float = 2.0) -> dict:
    """Check the extracted rows against the statement's own control total."""
    debits = round(sum(r["amount"] or 0 for r in rows if r.get("side") != "Cr"), 2)
    credits = round(sum(r["amount"] or 0 for r in rows if r.get("side") == "Cr"), 2)

    prev = totals.get("previous_balance")
    due = totals.get("total_amount_due")
    if prev is None or due is None:
        return {
            "status": "unverifiable",
            "reason": ("The statement's summary block (Previous Balance / Total Amount Due) "
                       "could not be read, so completeness cannot be proven."),
            "total_debits": debits, "total_credits": credits,
            "previous_balance": prev, "total_amount_due": due,
            "computed_closing": None, "difference": None,
        }

    # Prefer the statement's own printed period totals when available — they
    # measure the same thing we extracted, so a gap names the missing value
    # directly instead of inferring it from balances.
    stated_dr = totals.get("stated_debits")
    if stated_dr is not None:
        gap = round(debits - stated_dr, 2)
        if abs(gap) > tolerance:
            return {
                "status": "mismatch",
                "reason": (f"The statement's own purchases total is {stated_dr:,.2f} but only "
                           f"{debits:,.2f} was extracted — {abs(gap):,.2f} of transactions is "
                           f"missing or misread."),
                "total_debits": debits, "total_credits": credits,
                "previous_balance": prev, "total_amount_due": due,
                "stated_debits": stated_dr, "stated_credits": totals.get("stated_credits"),
                "computed_closing": round(prev + debits - credits, 2),
                "difference": gap,
            }

    computed = round(prev + debits - credits, 2)
    diff = round(computed - due, 2)
    ok = abs(diff) <= tolerance
    return {
        "status": "verified" if ok else "mismatch",
        "reason": "" if ok else (
            f"Extracted rows total {computed:,.2f} but the statement says {due:,.2f} — "
            f"a difference of {abs(diff):,.2f}. Transactions are missing or misread."),
        "total_debits": debits, "total_credits": credits,
        "previous_balance": prev, "total_amount_due": due,
        "stated_debits": totals.get("stated_debits"),
        "stated_credits": totals.get("stated_credits"),
        "computed_closing": computed, "difference": diff,
    }


# ── Union of the two extractions ────────────────────────────────────────────
# The column and line passes each see transactions the other misses, so picking
# a winner always loses rows. Measured on the Feb-26 OCR: column found 53 with
# the larger amounts, line found 72 including the whole of February, and the
# statement actually holds 102. Neither is a subset of the other.
#
# The reference number is the join key: it is printed on every row, it is unique
# per transaction, and it survives OCR far better than the merchant name or the
# date (both of which are routinely mangled).
_REF_ID_RE = re.compile(r"Ref\s*No\.?\s*:?\s*([A-Z0-9]{8,})", re.I)


def _ref_of(narration):
    m = _REF_ID_RE.search(str(narration or ""))
    return m.group(1).upper() if m else None


def _better_row(a, b):
    """Pick the more trustworthy of two views of the SAME transaction."""
    if a is None:
        return b
    if b is None:
        return a
    # An amount is the thing that must not be lost — a row without one cannot be
    # posted at all, so it always loses to a row that has one.
    if bool(a.get("amount")) != bool(b.get("amount")):
        return a if a.get("amount") else b
    # Then prefer the readable narration; OCR soup teaches the directory nothing.
    qa, qb = _narration_quality([a]), _narration_quality([b])
    if qa != qb:
        return a if qa > qb else b
    if _is_garbled(a.get("narration") or "") != _is_garbled(b.get("narration") or ""):
        return b if _is_garbled(a.get("narration") or "") else a
    # Finally prefer the row that actually parsed a date.
    if (a.get("date") is None) != (b.get("date") is None):
        return a if a.get("date") is not None else b
    return a


def _union_extractions(col_rows, line_rows):
    """All transactions either pass found, de-duplicated on reference number."""
    if not col_rows:
        return line_rows or []
    if not line_rows:
        return col_rows

    merged = {}
    order = []
    unkeyed = []
    for row in list(col_rows) + list(line_rows):
        ref = _ref_of(row.get("narration"))
        if not ref:
            # OCR sometimes destroys the reference on one pass but not the other,
            # so a ref-less row may still be a transaction we already have. Match
            # it against everything merged so far on date + amount before
            # admitting it — skipping this double-counted ₹1,58,191 on Feb-26.
            amt = round(float(row.get("amount") or 0), 2)
            twin = next((k for k in order
                         if merged[k].get("date") == row.get("date")
                         and round(float(merged[k].get("amount") or 0), 2) == amt
                         and amt), None)
            if twin is not None:
                merged[twin] = _better_row(merged[twin], row)
                continue
            key = ("noref", row.get("date"), amt,
                   merchant_key(row.get("narration") or ""))
            if key in merged:
                merged[key] = _better_row(merged[key], row)
                continue
            merged[key] = row
            order.append(key)
            unkeyed.append(key)
            continue
        if ref in merged:
            merged[ref] = _better_row(merged[ref], row)
        else:
            merged[ref] = row
            order.append(ref)

    out = [merged[k] for k in order]
    out.sort(key=lambda r: (r.get("date") or date.min))
    logger.info("Card: union of extractions — column %d + line %d -> %d unique "
                "transactions (%d had no reference number)",
                len(col_rows), len(line_rows), len(out), len(unkeyed))
    return out
