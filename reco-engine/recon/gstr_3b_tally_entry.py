"""
GSTR-3B Tally Entry Generator
Parses a GSTR-3B file (PDF or Excel) and generates Tally journal entries per SOP.

SOP (GSTR-3B_SOP.docx):
  Part 1 — ITC Transfer to Credit Ledger
    Debit  : Credit Ledger [Tax] [StateAbbr]  ← Row 4(C) Net ITC Available
    Credit : Input [Tax] [StateAbbr]          ← Row 4(4) Inward ISD + Row 4(5) All Other ITC

  Part 2 — Setting Off Output Liability
    Debit  : Output [Tax] [StateAbbr]         ← Row 6.1(A) total output tax (ITC + Cash)
    Credit : Credit Ledger [Tax] [StateAbbr]  ← Row 6.1(A) Tax paid through ITC
    Credit : Cash Ledger [StateAbbr]          ← Row 6.1(A) Tax paid in Cash (if > 0)

  Part 3 — RCM (only if Row 6.1(B) has values)
    Credit : RCM [Tax] [StateAbbr]            ← Row 6.1(B) Reverse Charge amounts
"""

from __future__ import annotations

import re
from io import BytesIO
from typing import Any

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import Alignment, Border, Font, PatternFill, Side

from recon.gstr_2b_books import GST_STATE_CODES, _GSTIN_RE, _ensure_xlsx


# ---------------------------------------------------------------------------
# State abbreviation map  (vehicle-registration / RTO codes used in Tally)
# ---------------------------------------------------------------------------

GST_STATE_SHORT: dict[str, str] = {
    "01": "JK",  "02": "HP",  "03": "PB",  "04": "CH",
    "05": "UK",  "06": "HR",  "07": "DL",  "08": "RJ",
    "09": "UP",  "10": "BR",  "11": "SK",  "12": "AR",
    "13": "NL",  "14": "MN",  "15": "MZ",  "16": "TR",
    "17": "ML",  "18": "AS",  "19": "WB",  "20": "JH",
    "21": "OD",  "22": "CT",  "23": "MP",  "24": "GJ",
    "25": "DD",  "26": "DN",  "27": "MH",  "28": "AP",
    "29": "KA",  "30": "GA",  "31": "LD",  "32": "KL",
    "33": "TN",  "34": "PY",  "35": "AN",  "36": "TS",
    "37": "AP",  "38": "LA",  "97": "OT",  "99": "CJ",
}


# ---------------------------------------------------------------------------
# Helpers
# ---------------------------------------------------------------------------

_MONTHS = [
    "January", "February", "March", "April", "May", "June",
    "July", "August", "September", "October", "November", "December",
]
_PERIOD_RE = re.compile(
    r'(?:' + '|'.join(_MONTHS) + r')\s+\d{4}', re.IGNORECASE
)
_PERIOD_MM_YYYY = re.compile(r'\b(0[1-9]|1[0-2])[/-](\d{4})\b')
_GSTIN_LOOSE   = re.compile(r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]')


def _parse_num(val: Any) -> float:
    """Parse Indian-formatted number string to float. Returns 0.0 on failure."""
    if val is None:
        return 0.0
    s = str(val).strip()
    # pdfplumber splits numbers across visual lines: '4521481.\n00' → '4521481.00'
    s = s.replace('\n', '').replace('\r', '')
    # PDF watermark letters prepended to values: 'F4521481.00' → '4521481.00'
    s = re.sub(r'^[A-Za-z]{1,4}', '', s).strip()
    s = s.replace(',', '').replace('₹', '')
    if not s or s in ('-', '--', 'nan', 'NaN', 'None', 'N/A', ''):
        return 0.0
    try:
        return float(s)
    except (ValueError, TypeError):
        return 0.0


def _is_numeric(val: Any) -> bool:
    """Return True only if val can be parsed as a float (not just any non-empty string)."""
    s = str(val or '').strip()
    s = s.replace('\n', '').replace('\r', '')
    s = re.sub(r'^[A-Za-z]{1,4}', '', s).strip()
    s = s.replace(',', '').replace('₹', '')
    if not s or s in ('-', '--', 'nan', 'NaN', 'None', 'N/A', ''):
        return False
    try:
        float(s)
        return True
    except (ValueError, TypeError):
        return False


def _empty_tax() -> dict:
    return {'igst': 0.0, 'cgst': 0.0, 'sgst': 0.0}


def _nums_from_text(text: str) -> list[float]:
    """Extract all numbers from a text string."""
    tokens = re.findall(r'[\d,]+\.?\d*', text)
    out = []
    for t in tokens:
        try:
            out.append(float(t.replace(',', '')))
        except ValueError:
            pass
    return out


def _state_info(gstin: str) -> tuple[str, str]:
    """Return (full_name, abbreviation) from a GSTIN's first 2 digits."""
    code = gstin[:2] if len(gstin) >= 2 else ''
    full  = GST_STATE_CODES.get(code, code)
    short = GST_STATE_SHORT.get(code, code)
    return full, short


# ---------------------------------------------------------------------------
# PDF Parser
# ---------------------------------------------------------------------------

def _collect_nums_from_rows(all_rows: list[list[str]], start: int, count: int) -> list[float]:
    """Collect strictly-numeric values from all_rows[start : start+count]."""
    nums: list[float] = []
    end = min(start + count, len(all_rows))
    for ri in range(start, end):
        for cell in all_rows[ri]:
            if _is_numeric(cell):
                nums.append(_parse_num(cell))
    return nums


def _extract_61_section(nums: list[float]) -> tuple[dict, dict, dict]:
    """
    GSTR-3B 6.1 table comes in TWO PDF layouts:
      Layout A (columns-first): [TaxIGST,TaxCGST,TaxSGST,TaxCess, ITCIGST,ITCCGST,ITCSGST,ITCCess, CashIGST,CashCGST,CashSGST,...]
      Layout B (rows-first):    [IGST(Tax,ITC,Cash), CGST(Tax,ITC,Cash), SGST(Tax,ITC,Cash), Cess(Tax,ITC,Cash)]

    Validates using Tax = ITC + Cash (±1 tolerance). Returns (tax_amt, itc, cash).
    """
    TOL = 1.0

    def valid(t, i, c):
        return all(abs(t[k] - (i[k] + c[k])) <= TOL for k in ('igst', 'cgst', 'sgst'))

    if len(nums) >= 11:
        # Layout A
        ta = {'igst': nums[0], 'cgst': nums[1],  'sgst': nums[2]}
        ia = {'igst': nums[4], 'cgst': nums[5],  'sgst': nums[6]}
        ca = {'igst': nums[8], 'cgst': nums[9],  'sgst': nums[10]}
        if valid(ta, ia, ca):
            return ta, ia, ca

        # Layout B
        tb = {'igst': nums[0], 'cgst': nums[3], 'sgst': nums[6]}
        ib = {'igst': nums[1], 'cgst': nums[4], 'sgst': nums[7]}
        cb = {'igst': nums[2], 'cgst': nums[5], 'sgst': nums[8]}
        if valid(tb, ib, cb):
            return tb, ib, cb

    if len(nums) >= 6:
        t = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        i = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        c = {'igst': nums[3], 'cgst': nums[4], 'sgst': nums[5]}
        if valid(t, i, c):
            return t, i, c
        # Try: Tax in first 3, Cash in next 3 with ITC=0
        i2 = _empty_tax()
        if valid(t, i2, c):
            return t, i2, c

    if len(nums) >= 3:
        t = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
        return t, t.copy(), _empty_tax()

    return _empty_tax(), _empty_tax(), _empty_tax()


def _parse_gstr3b_pdf(data: bytes) -> dict:
    try:
        import pdfplumber
    except ImportError:
        raise ValueError(
            "PDF parsing requires pdfplumber. "
            "Install with: pip3 install pdfplumber"
        )

    result = {
        'gstin': '', 'state': '', 'state_short': '', 'period': '',
        'row_4c':       _empty_tax(),
        'row_4_45':     _empty_tax(),
        'row_61a_tax':  _empty_tax(),  # Tax Amount (output liability) per tax type
        'row_61a_itc':  _empty_tax(),
        'row_61a_cash': _empty_tax(),
        'row_61b':      _empty_tax(),  # RCM Tax Amount (= Cash paid for RCM)
    }

    all_rows: list[list[str]] = []
    full_text = ''

    with pdfplumber.open(BytesIO(data)) as pdf:
        for page in pdf.pages:
            page_text = page.extract_text() or ''
            full_text += page_text + '\n'
            for table in page.extract_tables() or []:
                for row in table:
                    clean = [str(c or '').strip() for c in row]
                    if any(c for c in clean):
                        all_rows.append(clean)

    # ---- GSTIN ----
    gstin_match = _GSTIN_LOOSE.search(full_text)
    if gstin_match:
        g = gstin_match.group()
        result['gstin'] = g
        result['state'], result['state_short'] = _state_info(g)

    # ---- Period ----
    pm = _PERIOD_RE.search(full_text)
    if pm:
        result['period'] = pm.group().title()
    else:
        mm = _PERIOD_MM_YYYY.search(full_text)
        if mm:
            result['period'] = f"{_MONTHS[int(mm.group(1)) - 1]} {mm.group(2)}"

    # ---- Parse table rows ----
    row_4_4 = _empty_tax()
    row_4_5 = _empty_tax()

    n_rows = len(all_rows)

    for i, row in enumerate(all_rows):
        label = ' '.join(row).lower()

        # Row 4(C) – Net ITC Available
        # Values may be on this row or the next; use a 2-row window
        if re.search(r'net\s+itc\s+available|4\s*\(c\)', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3 and nums[0] + nums[1] + nums[2] > 0:
                result['row_4c'] = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}

        # Row 4(A)(4) – Inward supplies from ISD
        # Use ^ so '16(4)' in other rows doesn't match; keep ISD aliases unanchored
        if re.search(r'^\(4\)|^\(iv\)|inward.*isd|from\s+isd', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3:
                row_4_4 = {'igst': row_4_4['igst'] + nums[0],
                            'cgst': row_4_4['cgst'] + nums[1],
                            'sgst': row_4_4['sgst'] + nums[2]}

        # Row 4(A)(5) – All other ITC
        # Use ^ so '9(5)' in 6.1(B) label and '17(5)' in rule refs don't match
        if re.search(r'^\(5\)|^\(v\)|all\s+other\s+itc', label):
            nums = _collect_nums_from_rows(all_rows, i, 2)
            if len(nums) >= 3:
                row_4_5 = {'igst': row_4_5['igst'] + nums[0],
                            'cgst': row_4_5['cgst'] + nums[1],
                            'sgst': row_4_5['sgst'] + nums[2]}

        # Row 6.1(A) – "(A) Other than reverse charge" label row.
        # The NEXT 3 rows are the IGST / CGST / SGST data rows respectively.
        # Each data row layout (11 cols):
        #   0:TaxType  1:TaxPayable  2:Adj  3:NetTaxPayable  4:ITC-IGST  5:ITC-CGST
        #   6:ITC-SGST  7:ITC-Cess  8:CashPaid  9:InterestCash  10:LateFee
        if re.search(r'other\s+than\s+reverse\s+charge', label):
            def _gc(r, col):
                try: return _parse_num(r[col])
                except IndexError: return 0.0
            igst_r = all_rows[i + 1] if i + 1 < n_rows else []
            cgst_r = all_rows[i + 2] if i + 2 < n_rows else []
            sgst_r = all_rows[i + 3] if i + 3 < n_rows else []
            igst_tax  = _gc(igst_r, 1)
            cgst_tax  = _gc(cgst_r, 1)
            sgst_tax  = _gc(sgst_r, 1)
            igst_cash = _gc(igst_r, 8)
            cgst_cash = _gc(cgst_r, 8)
            sgst_cash = _gc(sgst_r, 8)
            result['row_61a_tax']  = {'igst': igst_tax,  'cgst': cgst_tax,  'sgst': sgst_tax}
            result['row_61a_itc']  = {'igst': max(0.0, igst_tax - igst_cash),
                                       'cgst': max(0.0, cgst_tax - cgst_cash),
                                       'sgst': max(0.0, sgst_tax - sgst_cash)}
            result['row_61a_cash'] = {'igst': igst_cash, 'cgst': cgst_cash, 'sgst': sgst_cash}

        # Row 6.1(B) – "(B) Reverse charge …" label row. Same column structure.
        # RCM: Tax Payable (col 1) = Cash paid (col 8); ITC = 0 for RCM.
        if re.search(r'reverse\s+charge', label) and 'other than' not in label:
            def _gc2(r, col):
                try: return _parse_num(r[col])
                except IndexError: return 0.0
            igst_r2 = all_rows[i + 1] if i + 1 < n_rows else []
            cgst_r2 = all_rows[i + 2] if i + 2 < n_rows else []
            sgst_r2 = all_rows[i + 3] if i + 3 < n_rows else []
            result['row_61b'] = {
                'igst': _gc2(igst_r2, 1),
                'cgst': _gc2(cgst_r2, 1),
                'sgst': _gc2(sgst_r2, 1),
            }

    # Fallback: if table extraction gave nothing for 4(C), try raw text
    if not any(result['row_4c'].values()):
        result['row_4c'] = _extract_row_from_text(full_text, r'net\s+itc\s+available')

    result['row_4_45'] = {
        'igst': row_4_4['igst'] + row_4_5['igst'],
        'cgst': row_4_4['cgst'] + row_4_5['cgst'],
        'sgst': row_4_4['sgst'] + row_4_5['sgst'],
    }

    return result


def _extract_row_from_text(text: str, pattern: str) -> dict:
    """Regex fallback: find a line matching pattern, return first 3 nums as igst/cgst/sgst."""
    for line in text.split('\n'):
        if re.search(pattern, line, re.IGNORECASE):
            nums = [n for n in _nums_from_text(line) if n >= 0]
            if len(nums) >= 3:
                return {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}
    return _empty_tax()


# ---------------------------------------------------------------------------
# Excel Parser
# ---------------------------------------------------------------------------

def _parse_gstr3b_excel(data: bytes, filename: str) -> dict:
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if data[:8] == b'\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1' or suffix == 'xls':
        data = _ensure_xlsx(data)

    result = {
        'gstin': '', 'state': '', 'state_short': '', 'period': '',
        'row_4c':       _empty_tax(),
        'row_4_45':     _empty_tax(),
        'row_61a_tax':  _empty_tax(),
        'row_61a_itc':  _empty_tax(),
        'row_61a_cash': _empty_tax(),
        'row_61b':      _empty_tax(),
    }

    sheets = pd.read_excel(BytesIO(data), sheet_name=None, dtype=str, header=None)

    all_text_parts: list[str] = []
    for _sname, df in sheets.items():
        df = df.fillna('')
        for _, row in df.iterrows():
            for cell in row:
                s = str(cell).strip()
                if s:
                    all_text_parts.append(s)

    all_text = ' '.join(all_text_parts)

    for part in all_text_parts:
        upper = part.upper()
        if _GSTIN_RE.match(upper):
            result['gstin'] = upper
            result['state'], result['state_short'] = _state_info(upper)
            break

    pm = _PERIOD_RE.search(all_text)
    if pm:
        result['period'] = pm.group().title()
    else:
        mm = _PERIOD_MM_YYYY.search(all_text)
        if mm:
            result['period'] = f"{_MONTHS[int(mm.group(1)) - 1]} {mm.group(2)}"

    row_4_4 = _empty_tax()
    row_4_5 = _empty_tax()

    for _sname, df in sheets.items():
        df = df.fillna('')
        rows_raw = df.values.tolist()

        # Detect IGST/CGST/SGST column indices from the header row
        igst_col = cgst_col = sgst_col = None
        for raw_row in rows_raw:
            cells_lower = [str(c).strip().lower() for c in raw_row]
            if any('igst' in c or 'integrated' in c for c in cells_lower):
                for ci, c in enumerate(cells_lower):
                    if ('igst' in c or 'integrated' in c) and igst_col is None:
                        igst_col = ci
                    elif ('cgst' in c or 'central tax' in c) and cgst_col is None:
                        cgst_col = ci
                    elif ('sgst' in c or 'utgst' in c or ('state' in c and 'tax' in c)) and sgst_col is None:
                        sgst_col = ci
                break

        igst_col = igst_col if igst_col is not None else 1
        cgst_col = cgst_col if cgst_col is not None else 2
        sgst_col = sgst_col if sgst_col is not None else 3

        def get(row_data, col):
            try:
                return _parse_num(row_data[col])
            except IndexError:
                return 0.0

        n_rows = len(rows_raw)

        for ri, raw_row in enumerate(rows_raw):
            label = ' '.join(str(c) for c in raw_row).lower()

            def _row_nums(start_ri, look=2):
                """Collect strictly-numeric values from rows[start_ri : start_ri+look]."""
                out = []
                for lr in rows_raw[start_ri:start_ri + look]:
                    out.extend(_parse_num(c) for c in lr if _is_numeric(c))
                return out

            # Row 4(C) — use _is_numeric so text cells ("Net ITC Available") are skipped
            if re.search(r'net\s+itc\s+available|4\s*\(c\)', label):
                nums = _row_nums(ri)
                if len(nums) >= 3 and nums[0] + nums[1] + nums[2] > 0:
                    result['row_4c'] = {'igst': nums[0], 'cgst': nums[1], 'sgst': nums[2]}

            # Row 4(A)(4) – ISD
            if re.search(r'\(4\)|\(iv\)|inward.*isd|from\s+isd', label):
                nums = _row_nums(ri)
                if len(nums) >= 3:
                    row_4_4['igst'] += nums[0]
                    row_4_4['cgst'] += nums[1]
                    row_4_4['sgst'] += nums[2]

            # Row 4(A)(5) – All other ITC
            if re.search(r'\(5\)|\(v\)|all\s+other\s+itc', label):
                nums = _row_nums(ri)
                if len(nums) >= 3:
                    row_4_5['igst'] += nums[0]
                    row_4_5['cgst'] += nums[1]
                    row_4_5['sgst'] += nums[2]

            # Row 6.1(A) – Other than reverse charge
            if re.search(r'other\s+than\s+reverse\s+charge|6\.1.*\(a\)', label):
                nums = []
                for look in rows_raw[ri:ri+4]:
                    nums.extend(_parse_num(c) for c in look if _is_numeric(c))
                tax, itc, cash = _extract_61_section(nums)
                result['row_61a_tax']  = tax
                result['row_61a_itc']  = itc
                result['row_61a_cash'] = cash

            # Row 6.1(B) – Reverse charge (RCM)
            if (re.search(r'6\.1.*\(b\)', label) or
                    (re.search(r'reverse\s+charge', label) and 'other than' not in label)):
                nums = []
                for look in rows_raw[ri:ri+4]:
                    nums.extend(_parse_num(c) for c in look if _is_numeric(c))
                tax, _itc, _cash = _extract_61_section(nums)
                result['row_61b'] = tax

    result['row_4_45'] = {
        'igst': row_4_4['igst'] + row_4_5['igst'],
        'cgst': row_4_4['cgst'] + row_4_5['cgst'],
        'sgst': row_4_4['sgst'] + row_4_5['sgst'],
    }

    return result


# ---------------------------------------------------------------------------
# Main dispatcher
# ---------------------------------------------------------------------------

def parse_gstr3b(filename: str, data: bytes) -> dict:
    """Parse a GSTR-3B file (PDF or Excel) and return structured extracted data."""
    suffix = filename.lower().rsplit('.', 1)[-1] if '.' in filename else ''
    if suffix == 'pdf':
        return _parse_gstr3b_pdf(data)
    return _parse_gstr3b_excel(data, filename)


# ---------------------------------------------------------------------------
# Tally entry row builder
# ---------------------------------------------------------------------------

def build_tally_entries(parsed: dict) -> list[dict]:
    """
    Generate two Tally journal entries matching actual practice (from screenshots):

    Journal 1 — ITC Availed (Part 1 of SOP, combined with RCM credit):
      Dr Credit Ledger IGST/CGST/SGST  ← Row 4(C) Net ITC Available
      Cr Input IGST/CGST/SGST          ← Row 4(4+5) regular ITC
      Cr RCM IGST/CGST/SGST            ← Row 6.1(B) RCM amounts (if any)

    Journal 2 — ITC Utilized (Part 2 of SOP):
      Dr Output IGST/CGST/SGST         ← Row 6.1(A) Tax Amount per tax type
      Cr Credit Ledger IGST            ← TOTAL ITC used (sum of all tax types)
      Cr Electronic Cash Ledger        ← Total cash paid (if any, for non-RCM)

    Journal 3 — RCM Tax Payment (only if row_61b has values):
      Dr Output IGST/CGST/SGST         ← Row 6.1(B) tax payable per type
      Cr Electronic Cash Ledger        ← Total RCM cash paid
    """
    st        = parsed.get('state_short', parsed.get('state', ''))
    r4c       = parsed['row_4c']
    r4_45     = parsed['row_4_45']
    r61a_tax  = parsed.get('row_61a_tax', _empty_tax())
    r61a_itc  = parsed['row_61a_itc']
    r61a_cash = parsed['row_61a_cash']
    r61b      = parsed['row_61b']

    entries: list[dict] = []

    def section(title: str):
        entries.append({'_type': 'section', 'sno': '', 'particulars': title, 'debit': '', 'credit': ''})

    def col_header():
        entries.append({'_type': 'header', 'sno': 'S.No', 'particulars': 'Particulars',
                        'debit': 'Debit (₹)', 'credit': 'Credit (₹)'})

    def row(sno: str, particular: str, debit='', credit=''):
        entries.append({'_type': 'data', 'sno': sno, 'particulars': particular,
                        'debit': debit, 'credit': credit})

    def blank():
        entries.append({'_type': 'blank', 'sno': '', 'particulars': '', 'debit': '', 'credit': ''})

    TAX = [('IGST', 'igst'), ('CGST', 'cgst'), ('SGST', 'sgst')]
    has_rcm = any(r61b[k] for k in ('igst', 'cgst', 'sgst'))

    # ── Journal 1: ITC Availed ─────────────────────────────────────────────
    # Debit side: Credit Ledger for each tax type (= Net ITC Available, Row 4C)
    section('Journal 1 — ITC Availed (Narration: ITC Availed)')
    col_header()
    sno = 1
    for tax_name, tax_key in TAX:
        row(str(sno), f'Credit Ledger {tax_name} {st}', debit=r4c[tax_key])
        sno += 1
    # Credit side: Input Ledger (regular ITC from Row 4(4+5))
    for tax_name, tax_key in TAX:
        row('', f'Input {tax_name} {st}', credit=r4_45[tax_key])
    # Credit side: RCM Ledger (ITC from RCM paid, Row 6.1(B))
    if has_rcm:
        for tax_name, tax_key in TAX:
            amt = r61b[tax_key]
            if amt:
                row('', f'RCM {tax_name} {st}', credit=amt)

    blank()

    # ── Journal 2: ITC Utilized ────────────────────────────────────────────
    # Debit side: Output Ledger per tax type (Tax Amount from Row 6.1(A))
    # Credit side: Credit Ledger IGST for total ITC used (IGST ITC covers all taxes)
    section('Journal 2 — ITC Utilized (Narration: ITC Utilization)')
    col_header()

    # Use Tax Amount (ITC + Cash); fall back to ITC if tax not extracted
    def tax_amt(key):
        t = r61a_tax.get(key, 0.0)
        if t:
            return t
        return r61a_itc.get(key, 0.0) + r61a_cash.get(key, 0.0)

    sno = 1
    for tax_name, tax_key in TAX:
        row(str(sno), f'Output {tax_name} {st}', debit=tax_amt(tax_key))
        sno += 1

    # Total ITC used across all tax types → credit to IGST Credit Ledger
    total_itc  = sum(r61a_itc[k]  for k in ('igst', 'cgst', 'sgst'))
    total_cash = sum(r61a_cash[k] for k in ('igst', 'cgst', 'sgst'))
    row('', f'Credit Ledger IGST {st}', credit=total_itc)
    if total_cash:
        row('', f'Electronic Cash Ledger {st}', credit=total_cash)

    blank()

    # ── Journal 3: RCM Tax Payment ────────────────────────────────────────
    # 6.1(B) Reverse Charge: tax payable = cash paid (no ITC offset for RCM).
    # Dr Output IGST/CGST/SGST per tax type, Cr Electronic Cash Ledger = total.
    if has_rcm:
        section('Journal 3 — RCM Tax Payment (Narration: RCM)')
        col_header()
        sno = 1
        for tax_name, tax_key in TAX:
            amt = r61b[tax_key]
            if amt:
                row(str(sno), f'Output {tax_name} {st}', debit=amt)
                sno += 1
        total_rcm = sum(r61b[k] for k in ('igst', 'cgst', 'sgst'))
        row('', f'Electronic Cash Ledger {st}', credit=total_rcm)

    return entries


# ---------------------------------------------------------------------------
# Excel Workbook Builder
# ---------------------------------------------------------------------------

_NAV_FILL   = PatternFill('solid', fgColor='123C69')
_SECT_FILL  = PatternFill('solid', fgColor='FF6600')
_HDR_FILL   = PatternFill('solid', fgColor='D6E4F7')
_WHITE_FONT = Font(bold=True, color='FFFFFF', name='Calibri', size=11)
_SECT_FONT  = Font(bold=True, color='FFFFFF', name='Calibri', size=10)
_HDR_FONT   = Font(bold=True, color='1E3A5F', name='Calibri', size=10)
_DATA_FONT  = Font(name='Calibri', size=10)
_ZERO_FONT  = Font(name='Calibri', size=10, color='888888')
_NUM_FMT    = '#,##0.00'

_THIN   = Side(style='thin', color='CCCCCC')
_BORDER = Border(left=_THIN, right=_THIN, top=_THIN, bottom=_THIN)


def build_gstr3b_tally_workbook(parsed: dict, entries: list[dict]) -> Workbook:
    wb = Workbook()
    ws = wb.active
    ws.title = 'Tally Entries'

    gstin       = parsed.get('gstin',       'N/A')
    state_full  = parsed.get('state',       'N/A')  # full name for header display
    state_short = parsed.get('state_short', 'N/A')  # abbreviation shown alongside
    period      = parsed.get('period',      'N/A')

    ws.column_dimensions['A'].width = 7
    ws.column_dimensions['B'].width = 44
    ws.column_dimensions['C'].width = 20
    ws.column_dimensions['D'].width = 20

    # Row 1: info header — show both full name and abbreviation
    ws.merge_cells('A1:D1')
    ws.row_dimensions[1].height = 28
    c = ws['A1']
    c.value = f'GSTIN: {gstin}    |    State: {state_full} ({state_short})    |    Period: {period}'
    c.font  = _WHITE_FONT
    c.fill  = _NAV_FILL
    c.alignment = Alignment(horizontal='center', vertical='center')

    ws.append(['', '', '', ''])
    ws.row_dimensions[2].height = 6

    current_row = 3

    for entry in entries:
        rtype = entry.get('_type', 'data')

        if rtype == 'blank':
            ws.append(['', '', '', ''])
            ws.row_dimensions[current_row].height = 8
            current_row += 1
            continue

        if rtype == 'section':
            ws.merge_cells(f'A{current_row}:D{current_row}')
            c = ws.cell(row=current_row, column=1, value=entry['particulars'])
            c.font  = _SECT_FONT
            c.fill  = _SECT_FILL
            c.alignment = Alignment(horizontal='left', vertical='center', indent=1)
            ws.row_dimensions[current_row].height = 22
            current_row += 1
            continue

        if rtype == 'header':
            for col, (key, align) in enumerate([
                ('sno',         Alignment(horizontal='center', vertical='center')),
                ('particulars', Alignment(horizontal='left',   vertical='center')),
                ('debit',       Alignment(horizontal='center', vertical='center')),
                ('credit',      Alignment(horizontal='center', vertical='center')),
            ], start=1):
                c = ws.cell(row=current_row, column=col, value=entry[key])
                c.font      = _HDR_FONT
                c.fill      = _HDR_FILL
                c.border    = _BORDER
                c.alignment = align
            ws.row_dimensions[current_row].height = 18
            current_row += 1
            continue

        # Data row
        debit_raw  = entry.get('debit', '')
        credit_raw = entry.get('credit', '')
        debit_num  = debit_raw  if isinstance(debit_raw,  (int, float)) else None
        credit_num = credit_raw if isinstance(credit_raw, (int, float)) else None

        is_zero = (
            (debit_num  is not None and debit_num  == 0.0) or
            (credit_num is not None and credit_num == 0.0)
        )
        font = _ZERO_FONT if is_zero else _DATA_FONT

        cells_spec = [
            (1, entry.get('sno', ''),         Alignment(horizontal='center', vertical='center')),
            (2, entry.get('particulars', ''), Alignment(horizontal='left',   vertical='center', indent=1)),
            (3, debit_num  if debit_num  is not None else '', Alignment(horizontal='right', vertical='center')),
            (4, credit_num if credit_num is not None else '', Alignment(horizontal='right', vertical='center')),
        ]

        for col, val, align in cells_spec:
            c = ws.cell(row=current_row, column=col, value=val)
            c.font      = font
            c.border    = _BORDER
            c.alignment = align
            if col in (3, 4) and isinstance(val, (int, float)):
                c.number_format = _NUM_FMT

        ws.row_dimensions[current_row].height = 16
        current_row += 1

    ws.freeze_panes = 'A2'
    return wb
