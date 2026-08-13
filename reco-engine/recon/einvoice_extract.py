"""E-Invoice Extractor — deterministic GST e-invoice PDF -> 3-sheet register.

Standalone-testable. pdfplumber primary; tesseract fallback only for scans.
"""
import re, io

# India GST state code map (POS name -> 2-digit code)
STATE_CODE = {
    "JAMMU AND KASHMIR": "01", "HIMACHAL PRADESH": "02", "PUNJAB": "03", "CHANDIGARH": "04",
    "UTTARAKHAND": "05", "HARYANA": "06", "DELHI": "07", "RAJASTHAN": "08", "UTTAR PRADESH": "09",
    "BIHAR": "10", "SIKKIM": "11", "ARUNACHAL PRADESH": "12", "NAGALAND": "13", "MANIPUR": "14",
    "MIZORAM": "15", "TRIPURA": "16", "MEGHALAYA": "17", "ASSAM": "18", "WEST BENGAL": "19",
    "JHARKHAND": "20", "ODISHA": "21", "CHHATTISGARH": "22", "MADHYA PRADESH": "23", "GUJARAT": "24",
    "DAMAN AND DIU": "25", "DADRA AND NAGAR HAVELI": "26", "MAHARASHTRA": "27", "KARNATAKA": "29",
    "GOA": "30", "LAKSHADWEEP": "31", "KERALA": "32", "TAMIL NADU": "33", "PUDUCHERRY": "34",
    "ANDAMAN AND NICOBAR ISLANDS": "35", "TELANGANA": "36", "ANDHRA PRADESH": "37", "LADAKH": "38",
}

def _num(s):
    try:
        return float(str(s).replace(",", "").strip())
    except Exception:
        return 0.0

def _parse_item_line(s):
    """Format-AGNOSTIC parse of ONE line-item row. Instead of a rigid positional
    regex, it anchors on the few things every GST e-invoice guarantees:
      • a leading SlNo, • a 6–8 digit HSN, • a "GST + Cess" rate group,
      • the identity  total ≈ taxable × (1 + rate%).
    Column separators (spaces, '|', '+', extra State-Cess sub-columns) can't fool
    it, so it reads Click2Shop, Zepto, and other billing layouts unchanged.
    Returns None if the line is not a line-item row (a 3% math-guard rejects
    header/description/summary lines)."""
    t = re.sub(r"\|", " ", s)
    t = re.sub(r"\s+", " ", t).strip()
    m = re.match(r"^(\d+)\s+(.+)$", t)
    if not m:
        return None
    slno = int(m.group(1))
    rest = m.group(2)
    hm = re.search(r"(?<!\d)(\d{6,8})(?!\d)", rest)          # HSN / SAC
    if not hm:
        return None
    desc = rest[:hm.start()].strip()
    hsn = hm.group(1)
    tail = rest[hm.end():].strip()
    qm = re.match(r"(\d+)\s+([A-Za-z]+)\s+(.+)$", tail)      # Qty  Unit  <numbers>
    if not qm:
        return None
    qty = _num(qm.group(1))
    unit = qm.group(2).upper()
    nums = qm.group(3)
    tg = re.search(r"([\d.]+)\s*\+\s*([\d.]+)", nums)        # GST + Cess
    if not tg:
        return None
    gst = _num(tg.group(1))
    cess = _num(tg.group(2))
    before = [_num(x) for x in re.findall(r"[\d,]+\.?\d*", nums[:tg.start()])]
    after = [_num(x) for x in re.findall(r"[\d,]+\.?\d*", nums[tg.end():])]
    if not before or not after:
        return None
    taxable = before[-1]                                     # number just before the tax group
    rate = before[0]                                         # unit price
    disc = before[1] if len(before) >= 3 else 0.0
    expected = taxable * (1 + (gst + cess) / 100)
    total = min(after, key=lambda x: abs(x - expected))     # the printed total, disambiguated by the math
    if taxable <= 0 or abs(total - expected) > max(2.0, expected * 0.03):
        return None                                          # not a real line-item row
    return {
        "slno": slno, "desc": desc, "hsn": hsn, "qty": qty, "unit": unit,
        "rate": rate, "discount": disc, "taxable": taxable,
        "gst_rate": gst, "cess_rate": cess, "other": 0.0, "total": round(total, 2),
    }

def _source_bytes(source):
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    try:
        with open(source, "rb") as fh:
            return fh.read()
    except Exception:
        return b""

def _extract_text_bytes(pdf_bytes):
    import pdfplumber
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)

def _pdf_text(source):
    """Digital text first (pdfplumber). If the PDF is a scan/flattened image
    (almost no text), fall back through OCR: tesseract (local, free, private) →
    iLovePDF (external, env-gated) as a last resort."""
    pdf_bytes = _source_bytes(source)
    text = _extract_text_bytes(pdf_bytes) if pdf_bytes else ""

    if len(text.strip()) < 40 and pdf_bytes:
        # 1) tesseract — OCR the scan into a searchable PDF, then re-extract text
        try:
            import tesseract_ocr
            if tesseract_ocr.available():
                ocr_pdf = tesseract_ocr.ocr_pdf(pdf_bytes)   # returns searchable-PDF bytes
                if ocr_pdf:
                    text = _extract_text_bytes(ocr_pdf)
        except Exception:
            pass
        # 2) iLovePDF — external OCR, only if a key is configured (privacy/cost trade-off)
        if len(text.strip()) < 40:
            try:
                alt = _ilovepdf_ocr_text(pdf_bytes)
                if alt and len(alt.strip()) >= 40:
                    text = alt
            except Exception:
                pass
    return text

def _ilovepdf_ocr_text(pdf_bytes):
    """Optional last-resort OCR via iLovePDF. Disabled unless ILOVEPDF_PUBLIC_KEY
    is set — it sends the document to a third party (cost + data leaves the box),
    so tesseract above is preferred. Wire the signed task flow here when a key is
    provided; returns None when unavailable."""
    import os
    if not os.environ.get("ILOVEPDF_PUBLIC_KEY"):
        return None
    return None  # TODO: implement iLovePDF auth→start→upload→ocr→download when a key is supplied

def _grab(text, label, stop=None):
    """Value after 'label :' up to the next label token or end of line."""
    m = re.search(re.escape(label) + r"\s*:?\s*(.+)", text)
    return m.group(1).strip() if m else ""

class NotAnEInvoice(ValueError):
    """Raised when a PDF is not a GST e-invoice (no IRN / Ack No.)."""


def is_einvoice(text):
    """A GST e-invoice ALWAYS carries an IRN (long hex) + Ack No. — anything
    else (normal invoice, bank statement, random PDF) fails this check."""
    has_irn = bool(re.search(r"IRN\s*:\s*[0-9a-fA-F]{16,}", text or ""))
    has_ack = bool(re.search(r"Ack\s*No\.?\s*:", text or "", re.I))
    return has_irn and has_ack


def parse_einvoice(source):
    """Parse ONE e-invoice PDF -> {header:{...}, line_items:[{...}]}.
    Raises NotAnEInvoice if the PDF is not a GST e-invoice."""
    text = _pdf_text(source)
    if not is_einvoice(text):
        raise NotAnEInvoice("This is not an E-Invoice PDF")
    lines = [ln.rstrip() for ln in text.splitlines()]
    joined = "\n".join(lines)
    H = {}

    # IRN wraps to the next line: 'IRN : <hex40> Ack No. : .. Ack Date : ..' then '<hex continuation>'
    m = re.search(r"IRN\s*:\s*([0-9a-fA-F]+)\s+Ack No\.?\s*:\s*(\S+)\s+Ack Date\s*:\s*([0-9:\-\s]+?)(?:\n|$)", joined)
    irn = ""
    if m:
        irn = m.group(1)
        H["ack_no"] = m.group(2).strip()
        H["ack_date"] = m.group(3).strip()
        # continuation: the following line that is pure hex
        idx = joined[:m.end()].count("\n")
        if idx + 1 < len(lines):
            nxt = lines[idx + 1].strip()
            if re.fullmatch(r"[0-9a-fA-F]{4,}", nxt):
                irn += nxt
    H["irn"] = irn

    def one(label, pat=r"(.+)"):
        mm = re.search(re.escape(label) + r"\s*:?\s*" + pat, joined)
        return mm.group(1).strip() if mm else ""

    H["invoice_no"]   = one("Document No.", r"(\S+)")
    H["invoice_date"] = one("Document Date", r"([0-9\-\/]+)")
    H["invoice_type"] = one("Document Type", r"([A-Za-z ]+?)\s+Document Date")
    H["supply_type"]  = one("Supply type Code", r"(\S+)")
    H["pos"]          = one("Place of Supply", r"([A-Za-z& ]+?)\s*$").upper() or \
                        (re.search(r"Place of Supply\s*:?\s*([A-Za-z& ]+)", joined) or [None, ""])[1].strip().upper()
    H["state_code"]   = STATE_CODE.get(H["pos"], "")
    # RCM: e-invoices print reverse charge; default N
    rc = re.search(r"Reverse Charge\s*:?\s*(Yes|No|Y|N)", joined, re.I)
    H["rcm"] = "Y" if (rc and rc.group(1).lower().startswith("y")) else "N"
    ewb = re.search(r"E-?Way Bill(?:\s*No\.?)?\s*:?\s*(\S+)", joined, re.I)
    H["eway"] = ewb.group(1).strip() if ewb else ""
    veh = re.search(r"Vehicle(?:\s*No\.?)?\s*:?\s*(\S+)", joined, re.I)
    H["vehicle"] = veh.group(1).strip() if veh else ""

    # Party details — Supplier / Recipient GSTINs (first two GSTIN matches after '3.Party Details')
    gstins = re.findall(r"GSTIN\s*:\s*([0-9A-Z]{15})", joined)
    H["supplier_gstin"] = gstins[0] if gstins else ""
    H["recipient_gstin"] = gstins[1] if len(gstins) > 1 else ""
    # supplier name = the line right under the top GSTIN or after 'Supplier :'
    top = re.search(r"^([0-9A-Z]{15})\s*\n(.+)", joined)
    H["supplier_name"] = top.group(2).strip() if top else ""

    # ---- Line items ----
    # Gate on the "Details of Goods" header when present, but if a layout doesn't
    # have that exact phrase, parse the whole doc — the math-guard in
    # _parse_item_line keeps header/summary lines from being mistaken for items.
    items, cur = [], None
    has_marker = any("Details of Goods" in l for l in lines)
    started = not has_marker
    for ln in lines:
        if has_marker and "Details of Goods" in ln:
            started = True
            continue
        if not started:
            continue
        parsed = _parse_item_line(ln)
        if parsed:
            if cur:
                items.append(cur)
            cur = parsed
        elif cur is not None:
            # continuation of description (skip the '0.00 + 0' state-cess line)
            s = ln.strip()
            if s and not re.fullmatch(r"[\d.]+\s*\+\s*[\d.]+", s):
                cur["desc"] += " " + s
    if cur:
        items.append(cur)

    # clean descriptions: collapse spaces + drop trailing '|'
    for it in items:
        it["desc"] = re.sub(r"\s*\|\s*", " ", it["desc"])
        it["desc"] = re.sub(r"\s+", " ", it["desc"]).strip()

    return {"header": H, "line_items": items}


def compute_tax(header, it):
    """Split GST into IGST vs CGST/SGST. Tax is the invoice's ACTUAL printed tax
    (line total - taxable - other charges), matching the e-invoice exactly."""
    sup_state = (header.get("supplier_gstin", "") or "")[:2]
    pos_state = header.get("state_code", "")
    rate = it["gst_rate"]
    taxable = it["taxable"]
    line_tax = round(it["total"] - taxable - it["other"], 2)      # actual printed tax
    cess_amt = round(taxable * it["cess_rate"] / 100, 2)
    gst_tax = round(line_tax - cess_amt, 2)
    interstate = bool(sup_state and pos_state and sup_state != pos_state)
    if interstate:
        igst_r, igst_a = rate, gst_tax
        cgst_r = cgst_a = sgst_r = sgst_a = 0
    else:
        igst_r = igst_a = 0
        cgst_r = sgst_r = rate / 2
        cgst_a = sgst_a = round(gst_tax / 2, 2)
    total_tax = round(igst_a + cgst_a + sgst_a + cess_amt, 2)
    round_off = round(it["total"] - (taxable + total_tax + it["other"]), 2)
    return dict(cgst_r=cgst_r, cgst_a=cgst_a, sgst_r=sgst_r, sgst_a=sgst_a,
                igst_r=igst_r, igst_a=igst_a, cess_a=cess_amt,
                total_tax=total_tax, round_off=round_off)


# ── 39-column register rows + the two pivots ─────────────────────────────────
REGISTER_COLS = [
    "S.No", "Invoice No.", "Invoice Date", "IRN", "Ack No.", "Ack Date",
    "Supplier Name", "Supplier GSTIN", "Supplier Address", "Recipient Name",
    "Recipient GSTIN/UIN", "Recipient Address", "Place of Supply (POS)", "State Code",
    "Invoice Type", "Reverse Charge (Y/N)", "E-Way Bill No.", "Vehicle No.", "SlNo",
    "Item Description", "HSN/SAC Code", "Quantity", "Unit", "Rate", "Taxable Value",
    "Discount", "CGST Rate (%)", "CGST Amount", "SGST Rate (%)", "SGST Amount",
    "IGST Rate (%)", "IGST Amount", "Cess Rate (%)", "Cess Amount", "Total Tax",
    "Other Charges", "Invoice Value (Line)", "TAX RATE ", "Round Off",
]

def build_register_rows(parsed_invoices):
    rows, sno = [], 0
    for r in parsed_invoices:
        h = r["header"]
        for it in r["line_items"]:
            sno += 1
            tx = compute_tax(h, it)
            rows.append({
                "S.No": sno, "Invoice No.": h.get("invoice_no", ""),
                "Invoice Date": h.get("invoice_date", ""), "IRN": h.get("irn", ""),
                "Ack No.": h.get("ack_no", ""), "Ack Date": h.get("ack_date", ""),
                "Supplier Name": h.get("supplier_name", ""), "Supplier GSTIN": h.get("supplier_gstin", ""),
                "Supplier Address": h.get("supplier_address", ""), "Recipient Name": h.get("recipient_name", ""),
                "Recipient GSTIN/UIN": h.get("recipient_gstin", ""), "Recipient Address": h.get("recipient_address", ""),
                "Place of Supply (POS)": h.get("pos", ""), "State Code": h.get("state_code", ""),
                "Invoice Type": h.get("invoice_type", ""), "Reverse Charge (Y/N)": h.get("rcm", "N"),
                "E-Way Bill No.": h.get("eway", ""), "Vehicle No.": h.get("vehicle", ""),
                "SlNo": it["slno"], "Item Description": it["desc"], "HSN/SAC Code": it["hsn"],
                "Quantity": it["qty"], "Unit": it["unit"], "Rate": it["rate"], "Taxable Value": it["taxable"],
                "Discount": it["discount"], "CGST Rate (%)": tx["cgst_r"], "CGST Amount": tx["cgst_a"],
                "SGST Rate (%)": tx["sgst_r"], "SGST Amount": tx["sgst_a"], "IGST Rate (%)": tx["igst_r"],
                "IGST Amount": tx["igst_a"], "Cess Rate (%)": it["cess_rate"], "Cess Amount": tx["cess_a"],
                "Total Tax": tx["total_tax"], "Other Charges": it["other"],
                "Invoice Value (Line)": it["total"], "TAX RATE ": it["gst_rate"], "Round Off": tx["round_off"],
            })
    return rows

def _pivot(rows, keys, sums):
    agg = {}
    order = []
    for r in rows:
        k = tuple(r[c] for c in keys)
        if k not in agg:
            agg[k] = {s: 0.0 for s in sums}
            order.append(k)
        for s in sums:
            agg[k][s] += _num(r[s])
    return agg, order


# ── styling palette (matches the app's blue theme) ───────────────────────────
_NAVY   = "0F172A"; _BLUE = "0748EE"; _BLUELT = "E8EFFE"; _ZEBRA = "F8FAFC"
_GRID   = "E2E8F0"; _MUTED = "64748B"
_MONEY  = '#,##0.00'; _INT = '#,##0'; _PCT = '0.00'
# which register columns get which number format
_FMT = {
    "Quantity": _INT,
    "Rate": _MONEY, "Taxable Value": _MONEY, "Discount": _MONEY, "CGST Amount": _MONEY,
    "SGST Amount": _MONEY, "IGST Amount": _MONEY, "Cess Amount": _MONEY, "Total Tax": _MONEY,
    "Other Charges": _MONEY, "Invoice Value (Line)": _MONEY, "Round Off": _MONEY,
    "CGST Rate (%)": _PCT, "SGST Rate (%)": _PCT, "IGST Rate (%)": _PCT, "Cess Rate (%)": _PCT,
    "TAX RATE ": _PCT,
}

def _style_table(ws, header_row, ncols, first_data_row, last_data_row, headers,
                 total_row=None, fmt_by_name=None, title_row=None):
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    thin = Side(style="thin", color=_GRID)
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    hdr_fill = PatternFill("solid", fgColor=_BLUE)
    zebra = PatternFill("solid", fgColor=_ZEBRA)
    tot_fill = PatternFill("solid", fgColor=_BLUELT)
    tot_font = Font(bold=True, color=_NAVY)
    fmt_by_name = fmt_by_name or {}

    if title_row:
        ws.merge_cells(start_row=title_row, start_column=1, end_row=title_row, end_column=ncols)
        tc = ws.cell(title_row, 1)
        tc.font = Font(bold=True, size=13, color=_NAVY); tc.alignment = Alignment(vertical="center")
        ws.row_dimensions[title_row].height = 24

    # header
    for c in range(1, ncols + 1):
        cell = ws.cell(header_row, c)
        cell.font = hdr_font; cell.fill = hdr_fill; cell.border = border
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[header_row].height = 30

    # data cells: zebra + borders + number formats
    for r in range(first_data_row, (last_data_row or first_data_row - 1) + 1):
        for c in range(1, ncols + 1):
            cell = ws.cell(r, c); cell.border = border
            nm = headers[c - 1]
            if nm in fmt_by_name:
                cell.number_format = fmt_by_name[nm]; cell.alignment = Alignment(horizontal="right")
        if (r - first_data_row) % 2 == 1:
            for c in range(1, ncols + 1):
                if not ws.cell(r, c).fill or ws.cell(r, c).fill.fgColor.rgb in (None, "00000000"):
                    ws.cell(r, c).fill = zebra

    # grand-total row highlight
    if total_row:
        for c in range(1, ncols + 1):
            cell = ws.cell(total_row, c)
            cell.fill = tot_fill; cell.font = tot_font; cell.border = border
            nm = headers[c - 1]
            if nm in fmt_by_name:
                cell.number_format = fmt_by_name[nm]

    # column widths from content (capped)
    for c in range(1, ncols + 1):
        width = len(str(headers[c - 1])) + 2
        for r in range(first_data_row, (last_data_row or first_data_row - 1) + 1):
            v = ws.cell(r, c).value
            if v is not None:
                width = max(width, min(len(str(v)) + 2, 55))
        ws.column_dimensions[get_column_letter(c)].width = min(max(width, 9), 55)

    ws.freeze_panes = ws.cell(first_data_row, 1)
    ws.auto_filter.ref = f"A{header_row}:{get_column_letter(ncols)}{max(last_data_row or header_row, header_row)}"
    ws.sheet_view.showGridLines = False


def build_workbook(parsed_invoices):
    """Build the styled 3-sheet GST e-Invoice Register workbook (openpyxl)."""
    from openpyxl import Workbook
    rows = build_register_rows(parsed_invoices)
    supplier = (parsed_invoices[0]["header"].get("supplier_name") if parsed_invoices else "") or ""
    wb = Workbook()

    # ---- Sheet 1: Invoice Details GSTR (39 cols) ----
    ws = wb.active
    ws.title = "Invoice Details GSTR"
    ws.cell(1, 1, f"{supplier} — GST e-Invoice Register (Sheet 1: Invoice Details)")
    for c, name in enumerate(REGISTER_COLS, 1):
        ws.cell(3, c, name)
    for i, r in enumerate(rows, 4):
        for c, name in enumerate(REGISTER_COLS, 1):
            ws.cell(i, c, r.get(name, ""))
    _style_table(ws, header_row=3, ncols=len(REGISTER_COLS), first_data_row=4,
                 last_data_row=3 + len(rows), headers=REGISTER_COLS, fmt_by_name=_FMT, title_row=1)

    # ---- Sheet 2: GSTR HSN ----
    hsn_agg, hsn_order = _pivot(rows, ["HSN/SAC Code", "Unit", "TAX RATE "],
                               ["Quantity", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"])
    wh = wb.create_sheet("GSTR HSN")
    hcols = ["HSN/SAC Code", "Unit", "TAX RATE ", "Sum of Quantity", "Sum of Taxable Value",
             "Sum of IGST Amount", "Sum of CGST Amount", "Sum of SGST Amount"]
    hfmt = {"Sum of Quantity": _INT, "TAX RATE ": _PCT, "Sum of Taxable Value": _MONEY,
            "Sum of IGST Amount": _MONEY, "Sum of CGST Amount": _MONEY, "Sum of SGST Amount": _MONEY}
    for c, name in enumerate(hcols, 1):
        wh.cell(1, c, name)
    gt = {k: 0.0 for k in ["Quantity", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"]}
    for i, k in enumerate(hsn_order, 2):
        v = hsn_agg[k]
        wh.cell(i, 1, k[0]); wh.cell(i, 2, k[1]); wh.cell(i, 3, k[2])
        for c, s in enumerate(["Quantity", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"], 4):
            wh.cell(i, c, round(v[s], 2)); gt[s] += v[s]
    gr = len(hsn_order) + 2
    wh.cell(gr, 1, "Grand Total")
    for c, s in enumerate(["Quantity", "Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"], 4):
        wh.cell(gr, c, round(gt[s], 2))
    _style_table(wh, header_row=1, ncols=8, first_data_row=2, last_data_row=len(hsn_order) + 1,
                 headers=hcols, total_row=gr, fmt_by_name=hfmt)

    # ---- Sheet 3: GSTR B2B ----
    b2b_agg, b2b_order = _pivot(rows, ["Invoice No.", "Invoice Date", "Recipient GSTIN/UIN", "TAX RATE "],
                               ["Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"])
    wb2 = wb.create_sheet("GSTR B2B")
    bcols = ["Invoice No.", "Invoice Date", "Recipient GSTIN/UIN", "TAX RATE ", "Sum of Taxable Value",
             "Sum of IGST Amount", "Sum of CGST Amount", "Sum of SGST Amount"]
    bfmt = {"TAX RATE ": _PCT, "Sum of Taxable Value": _MONEY, "Sum of IGST Amount": _MONEY,
            "Sum of CGST Amount": _MONEY, "Sum of SGST Amount": _MONEY}
    for c, name in enumerate(bcols, 1):
        wb2.cell(1, c, name)
    gt2 = {k: 0.0 for k in ["Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"]}
    for i, k in enumerate(b2b_order, 2):
        v = b2b_agg[k]
        for c, val in enumerate(k, 1):
            wb2.cell(i, c, val)
        for c, s in enumerate(["Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"], 5):
            wb2.cell(i, c, round(v[s], 2)); gt2[s] += v[s]
    gr2 = len(b2b_order) + 2
    wb2.cell(gr2, 1, "Grand Total")
    for c, s in enumerate(["Taxable Value", "IGST Amount", "CGST Amount", "SGST Amount"], 5):
        wb2.cell(gr2, c, round(gt2[s], 2))
    _style_table(wb2, header_row=1, ncols=8, first_data_row=2, last_data_row=len(b2b_order) + 1,
                 headers=bcols, total_row=gr2, fmt_by_name=bfmt)

    return wb


def _pivot_simple(rows, keys, sums):
    agg, _ = _pivot(rows, keys, sums)
    return agg


if __name__ == "__main__":
    import sys, glob, os
    folder = sys.argv[1] if len(sys.argv) > 1 else os.path.expanduser("~/Downloads/E-Invoice Extraction")
    pdfs = [p for p in glob.glob(os.path.join(folder, "*.pdf")) if "16 files" not in os.path.basename(p)]
    grand_taxable = grand_igst = 0.0
    total_items = 0
    for p in sorted(pdfs):
        r = parse_einvoice(p)
        h = r["header"]
        n = len(r["line_items"])
        total_items += n
        sub_t = sub_i = 0
        for it in r["line_items"]:
            tx = compute_tax(h, it)
            sub_t += it["taxable"]; sub_i += tx["igst_a"]
        grand_taxable += sub_t; grand_igst += sub_i
        print(f"{os.path.basename(p)[:40]:42} inv={h.get('invoice_no'):16} irn={h.get('irn','')[:12]}.. "
              f"pos={h.get('pos'):8} sup={h.get('supplier_gstin','')[:2]} rec={h.get('recipient_gstin','')[:2]} "
              f"items={n:3} taxable={sub_t:12.2f} igst={sub_i:10.2f}")
    print(f"\nTOTAL items={total_items}  taxable={grand_taxable:.2f}  igst={grand_igst:.2f}")
    print("Fixture single-file check: FBA(026)=133467.40/6673.41  FAB(032)=89827.85/4491.42")
