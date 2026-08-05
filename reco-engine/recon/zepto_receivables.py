"""Zepto Receivables reconciliation engine.

Pipeline: GRN gate (PO match) -> per-invoice enrichment from Invoice Details,
Payment Advice, Credit Note -> `1. Invoice Tracker` sheet with live formulas.
"""
from __future__ import annotations

import csv
import io
import re
from io import BytesIO
from typing import Any

import pandas as pd
from .gstr_2b_books import _ensure_xlsx   # OLE2 .xls -> .xlsx passthrough


def norm_po(v: Any) -> str:
    if v is None:
        return ""
    s = str(v).strip()
    if s.endswith(".0") and s[:-2].isdigit():   # pandas float-ified ints
        s = s[:-2]
    return s.upper()


def norm_inv(v: Any) -> str:
    if v is None:
        return ""
    return str(v).strip().upper()


def dn_ref_to_invoice(ref: str) -> str:
    """`V26-27/000007_QD` -> `INV26-27/000007` (drop `_...`, `V`->`INV`).

    Simple primary transform only. Malformed refs that are missing the
    V/INV prefix (e.g. `26-27/000039/_QD`) or that carry a mismatched
    trailing slash are NOT special-cased here — that reconciliation against
    the real invoice universe happens in the fallback remap inside
    `reconcile_zepto` (see `_dn_fallback_remap`). Non-invoice refs
    (PMDDN/CPMDDN/DC/etc.) are returned unchanged (normalized) so callers can
    route them separately.
    """
    base = str(ref or "").split("_")[0].strip()
    if base[:1].upper() == "V":
        base = "INV" + base[1:]
    return norm_inv(base)


def _norm_key(s: Any) -> str:
    return "".join(str(s or "").lower().split())


def _read_csv(data: bytes) -> list[list]:
    text = data.decode("utf-8-sig", errors="replace")
    return [row for row in csv.reader(io.StringIO(text))]


def _read_sheet(data: bytes, sheet: Any = 0) -> list[list]:
    data = _ensure_xlsx(data)
    df = pd.read_excel(BytesIO(data), sheet_name=sheet, header=None, dtype=object)
    grid = []
    for _, r in df.iterrows():
        grid.append(["" if (v is None or (isinstance(v, float) and pd.isna(v))) else v for v in r.tolist()])
    return grid


def _has_sheet(data: bytes, name: str) -> bool:
    try:
        xls = pd.ExcelFile(BytesIO(_ensure_xlsx(data)))
        return name in xls.sheet_names
    except Exception:
        return False


def _clean(v: Any) -> str:
    s = "" if v is None else str(v).strip()
    return "" if s.lower() == "nan" else s


def _find_header(grid: list[list], tokens: list[str], scan: int = 25) -> int:
    want = [_norm_key(t) for t in tokens]
    for i, row in enumerate(grid[:scan]):
        keys = {_norm_key(c) for c in row}
        if all(any(w in k for k in keys) for w in want):
            return i
    raise ValueError(f"Header row not found for tokens {tokens}")


def _rows_as_dicts(grid: list[list], header_idx: int) -> list[dict]:
    headers = [_clean(c) for c in grid[header_idx]]
    out = []
    for row in grid[header_idx + 1:]:
        if all(_clean(c) == "" for c in row):
            continue
        d = {}
        for j, h in enumerate(headers):
            if h:
                d[h] = row[j] if j < len(row) else ""
        out.append(d)
    return out


def _get(row: dict, aliases: list[str]) -> str:
    nk = {_norm_key(k): v for k, v in row.items()}
    for a in aliases:
        k = _norm_key(a)
        if k in nk:
            val = _clean(nk[k])
            if val != "":
                return val
    return ""


def parse_zepto_payment(data: bytes) -> list[dict]:
    try:
        grid = _read_sheet(data, "Zepto Payment track")
    except Exception:
        grid = _read_sheet(data, 0)
    h = _find_header(grid, ["po number", "invoice number"])
    rows = _rows_as_dicts(grid, h)
    out = []
    # Extract the actual column names from the header row for exact matching
    header_row = grid[h]
    po_col = None
    inv_col = None
    for col in header_row:
        nk = _norm_key(col)
        if "ponumber" in nk:
            po_col = col
        if "invoicenumber" in nk:
            inv_col = col

    for r in rows:
        po = ""
        inv = ""
        if po_col:
            po = norm_po(_get(r, [po_col]))
        if inv_col:
            inv = norm_inv(_get(r, [inv_col]))
        if not po:
            continue
        out.append({"po": po, "invoice_number": inv})
    return out


def parse_grn(datas: list[bytes]) -> dict[str, dict]:
    """First-win per PO. Each value carries both the GRN ID (for the
    'GRN No.' column) and the Created On date (for the 'GRN Date' column)."""
    pool: dict[str, dict] = {}
    for data in datas:
        grid = _read_csv(data)
        h = _find_header(grid, ["po id"])
        for r in _rows_as_dicts(grid, h):
            po = norm_po(_get(r, ["PO ID", "PO Id"]))
            if po and po not in pool:
                pool[po] = {
                    "grn_id": _get(r, ["GRN ID", "GRN Id", "GRN Code"]),
                    "created_on": _get(r, ["Created On", "Created on"]),
                }
    return pool


def grn_gate(payments: list[dict], grn: dict[str, dict]) -> list[dict]:
    kept, seen = [], set()
    for p in payments:
        po = p["po"]
        if po in grn and po not in seen:
            seen.add(po)
            kept.append(p)
    return kept


def _pan_from_gstin(gstin: Any) -> str:
    """PAN (entity id) = chars 3-12 of a 15-char GSTIN (first 2 chars are the
    state code). All of a company's state GSTINs share one PAN, so PAN cleanly
    identifies the vendor regardless of ship-to state."""
    s = str(gstin or "").strip().upper()
    return s[2:12] if len(s) >= 12 else ""


def _name_is_zepto(name: Any) -> bool:
    """Zepto's legal name is Kiranakart Technologies — older Tally rows carry
    'Kiranakart'/'kiranakat' instead of 'Zepto'."""
    n = str(name or "").lower()
    return "zepto" in n or "kiranak" in n


def parse_invoice_details(data: bytes) -> dict[str, dict]:
    """Build the invoice universe, keyed by the ORIGINAL invoice number
    (trailing slash preserved) so genuinely distinct invoices that merely
    share a base number — e.g. `INV26-27/000039` (Rs 29,169) vs
    `INV26-27/000039/` (Rs 29,870) — both survive as separate rows.

    A row is dropped as a TRUE duplicate only when a prior row already
    matched on (base invoice number with trailing slash stripped, amount,
    date) — this collapses exact re-exports like `INV26-27/000069` /
    `INV26-27/000069/` (identical amount + date) into one row.

    Feedback #00005: this is a ZEPTO receivables tracker, so only Zepto
    (a.k.a. Kiranakart) invoices are kept. The Tally export also contains other
    vendors (Blinkit / Flipkart / BigBasket / …) which, if left in, corrupt the
    totals and the Summary tab. A row is Zepto if its customer name matches
    zepto/kiranak, OR its GSTIN-PAN belongs to the set of PANs seen on
    name-matched Zepto rows (catches a mis-named row that still has a Zepto GSTIN).
    """
    grid = _read_sheet(data, "Invoice Details") if _has_sheet(data, "Invoice Details") else _read_sheet(data, 0)
    h = _find_header(grid, ["invoice_number", "customer_name", "bcy_total"])
    records: list[tuple[str, dict]] = []
    seen: set[tuple[str, float, str]] = set()
    for r in _rows_as_dicts(grid, h):
        inv = norm_inv(_get(r, ["invoice_number"]))
        if not inv:
            continue
        amount = _to_float(_get(r, ["bcy_total"]))
        date = _get(r, ["date"])
        dedup_key = (inv.rstrip("/"), amount, date)
        if dedup_key in seen:
            continue   # exact duplicate (same base#, amount, date) — skip
        seen.add(dedup_key)
        records.append((inv, {
            "date": date,
            "sales_order_no": _get(r, ["reference_number"]),
            "name": _get(r, ["customer_name"]),
            "total_invoice_amt": _get(r, ["bcy_total"]),
            "tax": _get(r, ["tax_amount"]),
            "invoice_amt_excl_tax": _get(r, ["amount_without_tax"]),
            "place_of_supply": _get(r, ["place_of_supply"]),
            "gstin": _get(r, ["gst_no", "gstin"]),
            "billing_state": _get(r, ["billing_state"]),
            "shipping_state": _get(r, ["shipping_state"]),
        }))
    # Zepto-only filter (#00005): PAN set from clearly Zepto-named rows, then keep
    # any row that is name-matched OR shares a Zepto PAN.
    zepto_pans = {
        _pan_from_gstin(f["gstin"]) for _, f in records
        if _name_is_zepto(f["name"]) and _pan_from_gstin(f["gstin"])
    }
    out: dict[str, dict] = {}
    for inv, f in records:
        if _name_is_zepto(f["name"]) or _pan_from_gstin(f["gstin"]) in zepto_pans:
            out[inv] = f
    return out


def _to_float(v: Any) -> float:
    s = str(v or "").replace(",", "").replace("₹", "").strip()
    if s in ("", "-", "nan", "None"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _parse_date(v: Any):
    """Parse an invoice date cell into a datetime.date, or None. Handles ISO
    (`2026-04-08`, with or without a trailing ` 00:00:00` time) and the common
    DD-MM-YYYY / DD-MM-YYYY-with-slash / MM-DD-YYYY exports."""
    import datetime as _dt
    s = str(v or "").strip()
    if not s or s.lower() in ("nan", "nat", "none"):
        return None
    s = s.split(" ")[0].split("T")[0]   # drop any time component
    for fmt in ("%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%m/%d/%Y", "%Y/%m/%d"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _due_date(inv_date):
    """#00001: Due Date = invoice date + 30 calendar days."""
    import datetime as _dt
    return inv_date + _dt.timedelta(days=30) if inv_date else None


def _due_status(due, is_paid: bool, today) -> str:
    """#00002: Not Due (due in future) / Due (due today) / Overdue (due passed),
    for UNPAID invoices only. Paid rows and rows with no due date -> blank."""
    if is_paid or due is None:
        return ""
    if due > today:
        return "Not Due"
    if due == today:
        return "Due"
    return "Overdue"


def parse_payment_advice(datas: list[bytes]) -> tuple[dict, dict]:
    payments: dict[str, dict] = {}
    debit_notes: dict[str, float] = {}
    for data in datas:
        grid = _read_csv(data)
        h = _find_header(grid, ["ref id", "amount", "payment amount"])
        for r in _rows_as_dicts(grid, h):
            typ = _get(r, ["Type/Description", "Type"]).lower()
            ref = _get(r, ["Ref Id", "Ref ID"])
            if not typ:                       # summary/total row
                continue
            if typ.startswith("invoice"):
                inv = norm_inv(ref)
                if not inv:
                    continue
                acc = payments.setdefault(inv, {"incl": 0.0, "excl": 0.0, "tds": 0.0})
                acc["incl"] += _to_float(_get(r, ["Payment Amount"]))
                acc["excl"] += _to_float(_get(r, ["Amount"]))
                acc["tds"] += _to_float(_get(r, ["TDS"]))
            elif typ.startswith("debit note"):
                inv = dn_ref_to_invoice(ref)
                if not inv:
                    continue
                debit_notes[inv] = debit_notes.get(inv, 0.0) + _to_float(_get(r, ["Amount"]))
    return payments, debit_notes


# Dynamic header matching — tolerant of wording/spacing/punctuation variants
# ("Payment Ref No." / "Payment Reference No" / "Payment Ref: X" / "Payment
# Ref. Number X"; "Payment Doc" / "Payment Document No."). Case-insensitive so
# it never hinges on one exact PDF layout.
_HEADER_RE = {
    "ref_no": re.compile(r"Payment\s+Ref(?:erence)?(?:\s+No\.?|\s+Number|\s+#)?\.?\s*[:\-]?\s*([^\n]+)", re.IGNORECASE),
    "doc": re.compile(r"Payment\s+Doc(?:ument)?(?:\s+No\.?|\s+Number)?\.?\s*[:\-]?\s*([^\n]+)", re.IGNORECASE),
}

# fallback line-item regex for when table extraction yields nothing usable, e.g.:
# "1428 Invoice Payment 1901333067 119,767.53 INR 114.06 119,653.47"
_LINE_ITEM_RE = re.compile(
    r"^\s*\d+\s+(Invoice Payment|Invoice|Debit Note(?:\s+\w+)?)\s+(\S+)\s+"
    r"([\d,]+\.?\d*)\s+([A-Z]{3})\s+([\d,]+\.?\d*)\s+([\d,]+\.?\d*)\s*$"
)


def _parse_header_fields(text: str) -> tuple[str, str, str]:
    ref_m = _HEADER_RE["ref_no"].search(text)
    doc_m = _HEADER_RE["doc"].search(text)
    # "Amount" header line looks like "Amount 381,403.44" and precedes the
    # amount-in-words line; the table's "Payment Amt." column has a different label.
    amt_m = re.search(r"^Amount\s+([\d,]+\.\d{2}|[\d,]+)\s*$", text, re.MULTILINE)
    ref_no = ref_m.group(1).strip() if ref_m else ""
    doc = doc_m.group(1).strip() if doc_m else ""
    amount = amt_m.group(1).strip() if amt_m else ""
    return ref_no, doc, amount


# "Payment Date 02/07/2026" in the PDF header. Dynamic: case-insensitive, and
# the two words "Payment Date" must be ADJACENT so it never matches "Payment
# Posting Date" (which is a different field). The value is captured for several
# common date shapes — DD/MM/YYYY, DD-MM-YYYY, YYYY-MM-DD, or "02 Jul 2026" —
# so a change in the PDF's date format doesn't silently drop the value. Every
# line item in a given PDF inherits this one date (and the Payment Ref No.).
_DATE_TOKEN = r"(\d{1,2}[/-]\d{1,2}[/-]\d{2,4}|\d{4}[/-]\d{1,2}[/-]\d{1,2}|\d{1,2}[ /-][A-Za-z]{3,9}[ /-]\d{2,4})"
_PAYMENT_DATE_RE = re.compile(r"Payment\s+Date\s*[:\-]?\s*" + _DATE_TOKEN, re.IGNORECASE)


def _parse_payment_date(text: str) -> str:
    m = _PAYMENT_DATE_RE.search(text or "")
    return m.group(1).strip() if m else ""


def _clean_ref_doc(cell: Any) -> str:
    return re.sub(r"\s+", "", str(cell or ""))


_INVOICE_KEY_RE = re.compile(r"^INV\d{2}-\d{2}/\d+$")
# Matches ANY ref that carries an invoice-shaped year/number pattern, even
# malformed ones missing the V/INV prefix or carrying an odd trailing slash,
# e.g. "V25-26/001564_QD", "26-27/000039/_QD", "INV26-27/000039_QD". This is
# intentionally looser than _INVOICE_KEY_RE — it decides ROUTING (is this an
# invoice-linked debit note at all?), while the later fallback remap in
# `reconcile_zepto` fixes up the exact key against the real invoice universe.
_INVOICE_PATTERN_RE = re.compile(r"\d{2}-\d{2}/\d+")


def _apply_line_item(payments: dict, debit_notes: dict, pmdn_box: list, typ: str, ref_doc: str,
                      amount: float, tds: float, payment_amt: float, details: dict | None = None,
                      advice_no: str = "", payment_date: str = "", doc_no: str = "") -> None:
    typ_l = typ.strip().lower()
    # Consolidate: EVERY recognized line item from every payment-advice PDF lands
    # here verbatim (invoice payments, credit memos/debit notes, PMDD, AP-AR),
    # tagged with its PDF's Payment Advice No + Payment Date. The sum of the
    # `payment_amt` column feeds the Summary's "Amount Received in Bank".
    if details is not None and "consolidate" in details:
        details["consolidate"].append({
            "advice_no": advice_no, "payment_date": payment_date, "type": typ,
            "doc_no": doc_no, "ref_doc": ref_doc,
            "amount": amount, "tds": tds, "payment_amt": payment_amt,
        })
    if typ_l.startswith("invoice"):
        inv = norm_inv(ref_doc)
        if not inv:
            return
        acc = payments.setdefault(inv, {"incl": 0.0, "excl": 0.0, "tds": 0.0})
        acc["incl"] += payment_amt
        acc["excl"] += amount
        acc["tds"] += tds
        if details is not None:
            details["payments"].append({"invoice": inv, "ref": ref_doc, "incl": payment_amt,
                                        "excl": amount, "tds": tds,
                                        "payment_date": payment_date, "advice_no": advice_no})
    elif typ_l.startswith("debit note") or typ_l.startswith("credit memo"):
        # Zepto's real PDFs label debit notes "Credit Memo" (there are no
        # literal "Debit Note" rows in the data); route both the same way.
        # Not every "Credit Memo" row is a per-invoice debit note though —
        # PMDDN/CPMDDN/DC-... refs and similar are marketing/adjustment
        # entries, not tied to a real invoice. Route by whether the RAW ref
        # carries an invoice-shaped pattern (\d{2}-\d{2}/\d+) — this catches
        # malformed refs missing the V/INV prefix (e.g. "26-27/000039/_QD")
        # that would otherwise be misfiled as PMDDN just because
        # dn_ref_to_invoice's simple primary transform didn't fully resolve
        # them. Exact-key reconciliation against the real invoice universe
        # happens later via the fallback remap in `reconcile_zepto`.
        if _INVOICE_PATTERN_RE.search(ref_doc):
            inv = dn_ref_to_invoice(ref_doc)
            if not inv:
                return
            debit_notes[inv] = debit_notes.get(inv, 0.0) + amount
            if details is not None:
                details["debit_notes"].append({"invoice": inv, "ref": ref_doc, "amount": amount,
                                               "payment_date": payment_date, "advice_no": advice_no})
        else:
            pmdn_box[0] += amount
            if details is not None:
                details["pmdd"].append({"ref": ref_doc, "amount": amount,
                                        "payment_date": payment_date, "advice_no": advice_no})
    elif typ_l.startswith("ap-ar"):
        pmdn_box[0] += amount
        if details is not None:
            details["ap_ar"].append({"ref": ref_doc, "amount": amount,
                                     "payment_date": payment_date, "advice_no": advice_no})


def _is_data_row_type(typ: str) -> bool:
    """True for recognized document types we route (invoice / debit note /
    credit memo — Zepto's label for debit notes — / AP-AR adjustment)."""
    t = typ.strip().lower()
    return (t.startswith("invoice") or t.startswith("debit note")
            or t.startswith("credit memo") or t.startswith("ap-ar"))


def _extract_line_items_from_table(table: list[list], payments: dict, debit_notes: dict,
                                    pmdn_box: list, details: dict | None = None,
                                    advice_no: str = "", payment_date: str = "") -> bool:
    """Shape-based row extraction — does NOT require a header row on the page.

    A row is a DATA row iff it has >= 8 cells and col[1] (Type of Document) is
    a recognized type (invoice / debit note / credit memo / ap-ar
    adjustment), matched case-insensitively. This works uniformly on page 1
    (which also has header/title rows to skip) and continuation pages (which
    have data rows only, no header at all).
    """
    found_any = False
    for row in table:
        if len(row) < 8:
            continue
        typ = _clean(row[1])
        if not typ or _norm_key(typ) == _norm_key("Type of Document"):
            continue
        if not _is_data_row_type(typ):
            # Unrecognized types or letterhead/junk rows — skip silently,
            # never crash.
            continue
        doc_no = _clean(row[2]) if len(row) > 2 else ""
        ref_doc = _clean_ref_doc(row[3])
        amount = _to_float(row[4])
        tds = _to_float(row[6])
        payment_amt = _to_float(row[7])
        _apply_line_item(payments, debit_notes, pmdn_box, typ, ref_doc, amount, tds, payment_amt,
                         details, advice_no, payment_date, doc_no)
        found_any = True
    return found_any


def _extract_line_items_from_text(text: str, payments: dict, debit_notes: dict, pmdn_box: list,
                                   filename: str = None, details: dict | None = None,
                                   advice_no: str = "", payment_date: str = "") -> bool:
    # Last-resort fallback only: the shape-based table extraction in
    # `_extract_line_items_from_table` handles both header and header-less
    # (multi-page continuation) pages on its own. This text-regex path is not
    # relied upon for normal PDFs and must NEVER raise for the multi-page
    # case — it simply reports whether it found anything so the caller can
    # decide, at the whole-PDF level, whether parsing genuinely failed.
    found_any = False
    for line in text.splitlines():
        m = _LINE_ITEM_RE.match(line)
        if not m:
            continue
        typ, ref_doc_raw, amount_s, _ccy, tds_s, payment_amt_s = m.groups()
        if not _is_data_row_type(typ):
            continue
        ref_doc = _clean_ref_doc(ref_doc_raw)
        _apply_line_item(payments, debit_notes, pmdn_box, typ, ref_doc,
                          _to_float(amount_s), _to_float(tds_s), _to_float(payment_amt_s), details,
                          advice_no, payment_date)
        found_any = True
    return found_any


def parse_payment_advice_pdf(pdf_bytes_list: list[bytes], details: dict | None = None) -> tuple[dict, dict, float]:
    """Parse Zepto PDF payment-advice files with PDF-level header dedup.

    Returns (payments, debit_notes, pmdn_total). `pmdn_total` accumulates
    "Credit Memo" rows whose ref doesn't resolve to a real invoice number
    (PMDDN/CPMDDN/DC-... marketing adjustments) plus all "AP-AR Adjustment"
    rows — these are NOT per-invoice debit notes.

    A whole PDF is skipped if its header triple (Payment Ref No., Payment Doc,
    Amount) was already seen — this prevents double-counting duplicate uploads.

    #00007: if `details` is provided (a dict with lists `payments`,
    `debit_notes`, `pmdd`, `ap_ar`), each recognized line item is ALSO appended
    to it (for the split detail tabs) — purely additive; the returned sums are
    unchanged and dedup still applies (skipped PDFs contribute no detail rows).
    """
    import pdfplumber

    payments: dict[str, dict] = {}
    debit_notes: dict[str, float] = {}
    pmdn_box = [0.0]
    seen: set[tuple[str, str, str]] = set()

    for pdf_bytes in pdf_bytes_list:
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            if not pdf.pages:
                continue
            first_text = pdf.pages[0].extract_text() or ""
            ref_no, doc, amount = _parse_header_fields(first_text)
            payment_date = _parse_payment_date(first_text)
            triple = (ref_no, doc, amount)
            if ref_no and doc and triple in seen:
                continue   # duplicate PDF — skip entirely
            if ref_no and doc:
                seen.add(triple)

            # Shape-based extraction across ALL pages — header row is only
            # ever present (if at all) on page 1; continuation pages have
            # data rows with no header. A PDF is only unparseable if NOT ONE
            # recognized data row was found anywhere in it.
            pdf_found_any = False
            for page in pdf.pages:
                for table in page.extract_tables():
                    if _extract_line_items_from_table(table, payments, debit_notes, pmdn_box, details,
                                                      ref_no, payment_date):
                        pdf_found_any = True
                if not pdf_found_any:
                    text = page.extract_text() or ""
                    if _extract_line_items_from_text(text, payments, debit_notes, pmdn_box, None, details,
                                                     ref_no, payment_date):
                        pdf_found_any = True

            if not pdf_found_any:
                raise ValueError(
                    "Payment-advice PDF has no extractable data rows; genuinely "
                    "unparseable file. Payment Doc: " + (doc or "<unknown>")
                )

    return payments, debit_notes, pmdn_box[0]


_CN_NUMBER_ALIASES = ["creditnote_number", "credit_note_number", "creditnote number", "Credit Note#", "credit note#"]


def parse_credit_notes(data: bytes, details: dict | None = None) -> dict[str, dict]:
    """Per invoice: sum of `bcy_total` across all its credit notes, plus the
    list of (deduped, non-empty) credit-note numbers that made up that sum.

    #00007: if `details` is provided, each individual credit-note row is also
    appended to `details["credit_notes"]` (invoice, number, amount) for the
    Credit Notes detail tab — additive; the returned aggregation is unchanged.
    """
    grid = _read_sheet(data, "Credit Note Details") if _has_sheet(data, "Credit Note Details") else _read_sheet(data, 0)
    h = _find_header(grid, ["invoice_number", "bcy_total"])
    out: dict[str, dict] = {}
    for r in _rows_as_dicts(grid, h):
        inv = norm_inv(_get(r, ["invoice_number"]))
        if not inv:
            continue
        amount = _to_float(_get(r, ["bcy_total"]))
        number = _get(r, _CN_NUMBER_ALIASES)
        acc = out.setdefault(inv, {"amount": 0.0, "numbers": []})
        acc["amount"] += amount
        if number and number not in acc["numbers"]:
            acc["numbers"].append(number)
        if details is not None:
            details["credit_notes"].append({"invoice": inv, "number": number, "amount": amount})
    return out


_LRN_INVOICE_ALIASES = ["invoice numbers drips foods", "invoice_number", "invoice number", "invoice numbers"]
_LRN_ID_ALIASES = ["LRN/ AWB", "LRN/AWB", "LRN"]
_LRN_DATE_ALIASES = ["Delivered Date to Destination", "Delivered Date", "Deliver Date"]


def _raw_get(row: dict, aliases: list[str]) -> Any:
    """Like `_get` but returns the RAW cell value (no str() coercion) so
    datetime/Timestamp objects survive for date formatting."""
    nk = {_norm_key(k): v for k, v in row.items()}
    for a in aliases:
        k = _norm_key(a)
        if k in nk:
            v = nk[k]
            if v is not None and v != "":
                return v
    return None


def _fmt_lrn_date(v: Any) -> str:
    if v is None:
        return ""
    if hasattr(v, "strftime"):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            pass
    s = str(v).strip()
    return "" if s.lower() in ("nan", "nat", "none") else s


def parse_lrn(datas: list[bytes]) -> dict[str, dict]:
    """Build {norm_inv: {"pod_no": <LRN/AWB>, "pod_date": <delivered date str>}}
    (first-win per invoice) from one or more LRN/POD tracking sheets.

    A sheet with NO invoice-number column (aliases in `_LRN_INVOICE_ALIASES`)
    is skipped entirely -- e.g. the courier's own tracking sheet that has no
    way to link rows back to an invoice.

    A single invoice cell may list MULTIPLE invoices separated by whitespace
    (spaces or newlines), e.g. "INV24-25/000484 INV24-25/000491" -- the same
    LRN/AWB + delivered date is applied to every invoice token in that row.
    """
    out: dict[str, dict] = {}
    for data in datas:
        try:
            grid = _read_sheet(data, 0)
        except Exception:
            continue
        try:
            h = _find_header(grid, ["lrn"])
        except ValueError:
            continue

        header_keys = {_norm_key(c) for c in grid[h]}
        inv_alias_keys = {_norm_key(a) for a in _LRN_INVOICE_ALIASES}
        if not (header_keys & inv_alias_keys):
            continue   # no invoice-number column on this sheet -> skip entirely

        for r in _rows_as_dicts(grid, h):
            inv_cell = _get(r, _LRN_INVOICE_ALIASES)
            if not inv_cell:
                continue
            pod_no = _get(r, _LRN_ID_ALIASES)
            pod_date = _fmt_lrn_date(_raw_get(r, _LRN_DATE_ALIASES))
            for tok in re.split(r"\s+", inv_cell.strip()):
                if not tok or "INV" not in tok.upper():
                    continue
                inv = norm_inv(tok)
                if inv and inv not in out:
                    out[inv] = {"pod_no": pod_no, "pod_date": pod_date}
    return out


def parse_payment_track_pod(data: bytes) -> dict[str, dict]:
    """POD from the Zepto Payment track sheet: POD No <- `LRN` column, POD Date
    <- `Delivery Date` column, keyed by invoice number with any trailing slash
    stripped (first row that carries an LRN wins per invoice). This is the
    primary POD source; the Drips LRN sheet (`parse_lrn`) is the fallback."""
    try:
        grid = _read_sheet(data, "Zepto Payment track")
    except Exception:
        try:
            grid = _read_sheet(data, 0)
        except Exception:
            return {}
    try:
        h = _find_header(grid, ["po number", "invoice number"])
    except ValueError:
        return {}
    out: dict[str, dict] = {}
    for r in _rows_as_dicts(grid, h):
        inv = norm_inv(_get(r, ["Invoice Number"])).rstrip("/")
        if not inv:
            continue
        # POD No and POD Date are captured INDEPENDENTLY: a row can carry a
        # Delivery Date with an empty LRN (or vice-versa). The old code gated the
        # whole row on `if lrn:`, silently dropping the Delivery Date whenever the
        # LRN cell was blank. First non-empty value per field wins per invoice.
        # LRN is kept verbatim (courier words like Porter/Booked/Self stay as-is).
        lrn = _get(r, ["LRN"])
        pod_date = _fmt_lrn_date(_raw_get(r, ["Delivery Date"]))
        if not lrn and not pod_date:
            continue
        cur = out.setdefault(inv, {"pod_no": "", "pod_date": ""})
        if lrn and not cur["pod_no"]:
            cur["pod_no"] = lrn
        if pod_date and not cur["pod_date"]:
            cur["pod_date"] = pod_date
    return out


COLUMN_KEYS = [
    "po","date","invoice_number","sales_order_no","name","total_invoice_amt","tax",
    "invoice_amt_excl_tax","place_of_supply","gstin","billing_state","shipping_state",
    "pending_amount","payment_received_incl_tds","payment_received_excl_tds","tds",
    "debit_note_issued","dn_accepted","dn_not_accepted","credit_note_issued","credit_note_no",
    "gross_outstanding","net_outstanding","status","due_date","due_status","grn_no","grn_date",
    "invoice_not_in_ledger","pod_no","pod_date","payment_date",
]


def _one(files: dict, field: str) -> bytes | None:
    v = files.get(field)
    if v is None:
        return None
    if isinstance(v, list):
        return v[0]["content"] if v else None
    return v["content"]


def _many(files: dict, field: str) -> list[bytes]:
    v = files.get(field)
    if v is None:
        return []
    items = v if isinstance(v, list) else [v]
    return [it["content"] for it in items if it.get("content")]


class _RecoRows(list):
    """A plain-list subclass so `reconcile_zepto` can carry `pmdn_adjustment`
    alongside the row data without breaking existing callers/tests (len,
    iteration, indexing) or server.py's JSON serialization of the result."""
    pass


def _dn_fallback_remap(debit_notes: dict[str, float], invoice_universe: set[str]) -> dict[str, float]:
    """Fix up malformed debit-note keys by matching candidate variants
    against the REAL invoice universe (from Invoice Details).

    For every `debit_notes` key not already present in `invoice_universe`,
    build candidates by stripping a leading `INV`/`V` (if any) to get a
    `core`, then trying both WITH and WITHOUT a trailing slash, always with
    an `INV` prefix, in this priority order (first match wins):
        1. "INV"+core                    (preserves the ref's own slash-ness)
        2. "INV"+core.rstrip("/")+"/"
        3. "INV"+core.rstrip("/")
        4. "INV"+core+"/"
    The DN amount is reassigned (merge-added) onto the winning key. If
    nothing matches, the original key is left untouched (unmatched —
    harmless).
    """
    out: dict[str, float] = {}
    for key, amount in debit_notes.items():
        if key in invoice_universe:
            out[key] = out.get(key, 0.0) + amount
            continue

        core = key
        if core[:3].upper() == "INV":
            core = core[3:]
        elif core[:1].upper() == "V":
            core = core[1:]

        ordered_candidates = [
            "INV" + core,
            "INV" + core.rstrip("/") + "/",
            "INV" + core.rstrip("/"),
            "INV" + core + "/",
        ]
        candidates = list(dict.fromkeys(ordered_candidates))   # dedupe, keep order
        match = next((c for c in candidates if c in invoice_universe), None)
        if match:
            out[match] = out.get(match, 0.0) + amount
        else:
            out[key] = out.get(key, 0.0) + amount   # unmatched — leave as-is
    return out


def reconcile_zepto(files: dict, today=None) -> list[dict]:
    import datetime as _dt
    if today is None:
        today = _dt.date.today()
    # #00007: collect line-item detail for the split tabs (Payments / Debit
    # Notes / Credit Notes / PMDD / AP-AR). Additive — sums below are unchanged.
    details = {"payments": [], "debit_notes": [], "pmdd": [], "ap_ar": [], "credit_notes": [],
               "consolidate": []}
    invoice_details = parse_invoice_details(_one(files, "invoice_details") or b"")
    payments_raw = parse_zepto_payment(_one(files, "zepto_payment") or b"")
    grn = parse_grn(_many(files, "grn_list"))
    pay_map, dn_map, pmdn_total = parse_payment_advice_pdf(_many(files, "payment_advice"), details)
    cn_map = parse_credit_notes(_one(files, "credit_note") or b"", details)
    lrn_map = parse_lrn(_many(files, "lrn"))
    pod_track = parse_payment_track_pod(_one(files, "zepto_payment") or b"")   # POD from Payment track LRN/Delivery Date

    dn_map = _dn_fallback_remap(dn_map, set(invoice_details.keys()))

    # invoice -> PO map from the Zepto Payment track (first PO wins per invoice).
    # Feedback #00006: match on the invoice number with any trailing slash
    # stripped, because the Payment track carries `INV26-27/000192` while the
    # Tally universe may carry `INV26-27/000192/` — the same invoice must still
    # find its PO despite that stray slash.
    inv_to_po: dict[str, str] = {}
    for p in payments_raw:
        inv = p.get("invoice_number", "").rstrip("/")
        po = p.get("po", "")
        if inv and po and inv not in inv_to_po:
            inv_to_po[inv] = po

    results: list[dict] = []
    for inv, det in invoice_details.items():
        row = {key: "" for key in COLUMN_KEYS}
        row["invoice_number"] = inv

        # #00006: show the PO whenever the Payment track has one for this
        # invoice — do NOT suppress it just because the PO is absent from the
        # GRN list. The GRN No./Date columns stay gated on the GRN list.
        po = inv_to_po.get(inv.rstrip("/"), "")
        row["po"] = po
        if po and po in grn:
            row["grn_no"] = grn[po]["grn_id"]
            row["grn_date"] = grn[po]["created_on"]

        for f in ("date","sales_order_no","name","place_of_supply","gstin","billing_state","shipping_state"):
            row[f] = det[f]
        row["total_invoice_amt"] = _to_float(det["total_invoice_amt"])
        row["tax"] = _to_float(det["tax"])
        row["invoice_amt_excl_tax"] = _to_float(det["invoice_amt_excl_tax"])

        pay = pay_map.get(inv, {})
        row["payment_received_incl_tds"] = round(pay.get("incl", 0.0), 2)
        row["payment_received_excl_tds"] = round(pay.get("excl", 0.0), 2)
        row["tds"] = round(pay.get("tds", 0.0), 2)
        row["debit_note_issued"] = round(dn_map.get(inv, 0.0), 2)

        cn = cn_map.get(inv)
        row["credit_note_issued"] = round(cn["amount"], 2) if cn else 0.0
        row["credit_note_no"] = ", ".join(cn["numbers"]) if cn else ""

        # POD: primary = Zepto Payment track (LRN + Delivery Date, slash-
        # normalized match); fallback = Drips LRN sheet.
        lr = pod_track.get(inv.rstrip("/")) or lrn_map.get(inv)
        row["pod_no"] = lr["pod_no"] if lr else ""
        row["pod_date"] = lr["pod_date"] if lr else ""

        pending = _to_float(row["total_invoice_amt"])
        row["pending_amount"] = round(pending, 2)
        # Feedback #00004 (corrected formulas):
        #   Gross Receivable = Pending - Payment Received (excl. TDS)
        #   Net Receivable   = Pending - Payment Received (excl. TDS) - Credit Note - Debit Note
        gross = pending - row["payment_received_excl_tds"]
        net = gross - row["credit_note_issued"] - row["debit_note_issued"]
        row["gross_outstanding"] = round(gross, 2)
        row["net_outstanding"] = round(net, 2)
        # Status is driven by NET OUTSTANDING ONLY (Gross is ignored). Signed
        # threshold (NOT abs()): Net <= 10 -> Paid, INCLUDING negative net
        # (nothing left to collect / overpaid) which the accountant treats as
        # settled too.
        row["status"] = "Paid" if net <= 10 else "Not Paid"

        # #00001/#00002: Due Date = invoice date + 30 days; Due status (Not Due /
        # Due / Overdue) vs `today`, for UNPAID rows only (Paid -> blank).
        inv_dt = _parse_date(det.get("date"))
        due = _due_date(inv_dt)
        row["due_date"] = due.isoformat() if due else ""
        row["due_status"] = _due_status(due, row["status"] == "Paid", today)
        results.append(row)

    out = _RecoRows(results)
    out.pmdn_adjustment = round(pmdn_total, 2)
    out.details = details   # #00007: line-item detail for the split tabs
    return out


def summarize_zepto(results: list[dict]) -> dict:
    paid = sum(1 for r in results if r["status"] == "Paid")
    not_paid = sum(1 for r in results if r["status"] == "Not Paid")
    return {
        "total": len(results),
        "paid": paid,
        "not_paid": not_paid,
        "not_in_invoice_details": 0,
    }


# Excel column letters, 1-based, matching COLUMN_KEYS order (A..AD).
# "PO" was prepended in Task 2 (Invoice-Details universe + PO column); the
# live formula strings below (pending/gross/net/status) are re-pointed to
# these shifted columns in Task 3. "Credit Note No" was inserted after
# "Credit Note Issued" in the 5-column task, shifting everything after it
# one letter to the right (U..AC -> V..AD).
_LETTERS = ["A","B","C","D","E","F","G","H","I","J","K","L","M","N","O","P","Q","R","S","T","U","V","W","X","Y","Z","AA","AB","AC","AD","AE","AF"]
_HEADERS = ["PO","Date","Invoice_number","Sales Order No.","Name","Total Invoice Amt","Tax",
    "Invoice Amt (Excl. Tax)","Place of Supply","GSTIN","Billing State","shipping_state",
    "Pending Amount","Payment Received (Including TDS)","Payment Received (Excluding TDS)","TDS",
    "Debit Note Issued","DN Accepted","DN Not Accepted","Credit Note Issued","Credit Note No",
    "Gross Outstanding Amt","Net Outstanding Amt","Status","Due Date","Due Status","GRN No.","GRN Date",
    "Invoice Not Available in Zepto Ledger","POD No","POD Date","Payment Date"]
_FORMULA_COLS = {"pending_amount","gross_outstanding","net_outstanding","status"}
_MONEY_KEYS = {"total_invoice_amt","tax","invoice_amt_excl_tax","pending_amount",
    "payment_received_incl_tds","payment_received_excl_tds","tds","debit_note_issued",
    "credit_note_issued","gross_outstanding","net_outstanding"}
# Text (non-money) columns that must NOT get the "#,##0.00" number format even
# though they sit among money columns in the sheet.
_TEXT_KEYS = {"credit_note_no", "grn_no", "pod_no", "pod_date", "grn_date"}

# Per-column widths for "1. Invoice Tracker" — wider for names/refs/GSTIN,
# tighter for money/date/status columns. Keyed by COLUMN_KEYS.
_COLUMN_WIDTHS = {
    "po": 14, "date": 12, "invoice_number": 20, "sales_order_no": 20, "name": 30,
    "total_invoice_amt": 15, "tax": 12, "invoice_amt_excl_tax": 16,
    "place_of_supply": 12, "gstin": 18, "billing_state": 14, "shipping_state": 14,
    "pending_amount": 14, "payment_received_incl_tds": 16, "payment_received_excl_tds": 16,
    "tds": 11, "debit_note_issued": 14, "dn_accepted": 12, "dn_not_accepted": 12,
    "credit_note_issued": 14, "credit_note_no": 16, "gross_outstanding": 15, "net_outstanding": 15,
    "status": 12, "due_date": 12, "due_status": 12, "grn_no": 14, "grn_date": 12, "invoice_not_in_ledger": 16,
    "pod_no": 12, "pod_date": 12, "payment_date": 12,
}

# Debit Note display format: value stays NEGATIVE (so `Net = M-O-T-Q` is
# unchanged) but the minus sign is hidden on screen — "positive;negative;zero"
# with the negative section written WITHOUT a leading "-".
_DN_DISPLAY_FMT = "#,##0.00;#,##0.00;0.00"

# Paid/Not-Paid conditional formatting palette (finance-report standard).
_PAID_FILL_HEX = "C6EFCE"
_PAID_FONT_HEX = "006100"
_NOTPAID_FILL_HEX = "FFC7CE"
_NOTPAID_FONT_HEX = "9C0006"


_KEY_TOTAL_LABELS = {"Net Sales", "Net Receivables", "Net Receivables - 2"}


def build_summary_sheet(wb, results: list[dict]):
    """Add a `Summary` worksheet: auto-computed totals + blank-for-manual lines."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side

    ws = wb.create_sheet(title="Summary")
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)

    title_font = Font(bold=True, size=13, color="123C69")
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    label_font = Font(bold=False, size=10)
    key_total_font = Font(bold=True, size=10, color="123C69")
    manual_fill = PatternFill("solid", fgColor="F2F2F2")
    money_fmt = "#,##0.00"

    ws["A1"] = "Zepto Receivable Summary"
    ws["A1"].font = title_font
    ws["A1"].alignment = Alignment(horizontal="left", vertical="center")
    ws.merge_cells("A1:C1")
    ws.row_dimensions[1].height = 22

    ws.append([])   # row 2 blank spacer
    ws.append(["Particular", "Amount", "Remarks"])
    for col in (1, 2, 3):
        c = ws.cell(row=3, column=col)
        c.font = hdr_font
        c.fill = hdr_fill
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        c.border = border
    ws.row_dimensions[3].height = 20

    sales_incl_tax = sum(_to_float(r.get("total_invoice_amt", 0)) for r in results)
    sale_return = sum(_to_float(r.get("credit_note_issued", 0)) for r in results)
    debit_note_issued = sum(_to_float(r.get("debit_note_issued", 0)) for r in results)
    tds_deducted = sum(_to_float(r.get("tds", 0)) for r in results)
    payment_received = sum(_to_float(r.get("payment_received_incl_tds", 0)) for r in results)
    pmdn = getattr(results, "pmdn_adjustment", 0.0)

    # Debit Note is stored NEGATIVE on the tracker (so Net = M-O-T-Q holds); the
    # Summary shows it POSITIVE and subtracts it (matches the CA's reference).
    dn_positive = abs(debit_note_issued)
    net_sales = sales_incl_tax - sale_return - dn_positive - tds_deducted
    net_receivables = net_sales - payment_received

    # "Amount Received in Bank" = total of the Payment Advice Consolidate tab
    # (sum of the per-line Payment Amt across every payment-advice PDF = the
    # actual net cash credited).
    details = getattr(results, "details", None) or {}
    amount_received_in_bank = round(
        sum(_to_float(x.get("payment_amt", 0)) for x in details.get("consolidate", [])), 2)

    net_receivables_2 = net_receivables + pmdn

    # (label, value_or_None, remark, kind). kind: "key" = headline total (bold),
    # "manual" = grey informational/manual row, "" = plain auto row.
    rows_spec = [
        ("Sales Including Tax", round(sales_incl_tax, 2), "", ""),
        ("Sale Return (Credit Notes)", round(sale_return, 2), "", ""),
        ("Debit Note Issued", round(dn_positive, 2), "Shown positive; deducted in Net Sales", ""),
        ("TDS Deducted", round(tds_deducted, 2), "", ""),
        ("Net Sales", round(net_sales, 2), "Sales Incl. Tax - Sale Return - Debit Note - TDS", "key"),
        ("Payment Received (Including TDS)", round(payment_received, 2), "", ""),
        ("Net Receivables", round(net_receivables, 2), "Net Sales - Payment Received (Incl. TDS)", "key"),
        ("Debit Note Accepted", None, "Manual entry", "manual"),
        ("Debit Note Not Accepted", None, "Manual entry", "manual"),
        ("Amount Received in Bank", amount_received_in_bank, "Total of Payment Advice Consolidate tab", "manual"),
        ("Expense - PMDDN & AP-AR Adjustment (Marketing)", round(pmdn, 2), "Auto-computed from Payment Advice PDFs", "manual"),
        ("Previous Year Marketing Exp. Invoices", None, "Manual entry", "manual"),
        ("Net Receivables - 2", round(net_receivables_2, 2), "Net Receivables + Expense (PMDDN & AP-AR)", "key"),
    ]

    r = 4
    for label, value, remark, kind in rows_spec:
        row_font = key_total_font if kind == "key" else label_font
        ws.cell(row=r, column=1, value=label).font = row_font
        amt_cell = ws.cell(row=r, column=2, value=value)
        amt_cell.number_format = money_fmt
        amt_cell.font = row_font
        amt_cell.alignment = Alignment(horizontal="right")
        ws.cell(row=r, column=3, value=remark).font = Font(italic=True, size=9, color="595959")
        for col in (1, 2, 3):
            cell = ws.cell(row=r, column=col)
            cell.border = border
            if kind == "manual":
                cell.fill = manual_fill
        r += 1

    ws.column_dimensions["A"].width = 45
    ws.column_dimensions["B"].width = 20
    ws.column_dimensions["C"].width = 45
    ws.freeze_panes = "A4"
    return ws


_DETAIL_LINK_FONT_HEX = "0563C1"


def _canon_inv(k: str) -> str:
    """Normalize an invoice-ish key for hyperlink matching: upper/trim, and add
    the `INV` prefix if it's missing but the year/number pattern is present
    (the DN fallback remap can leave detail keys like `26-27/000039/` while the
    main sheet carries `INV26-27/000039/`)."""
    k = norm_inv(k)
    if k and not k.startswith("INV") and re.match(r"\d{2}-\d{2}/", k):
        k = "INV" + k
    return k


def _resolve_to_universe(inv: str, universe: set) -> str | None:
    """Map a (possibly malformed) debit-note ref key onto the matching invoice
    in the universe, mirroring `_dn_fallback_remap`'s candidate order. Returns
    the universe key or None. Used to filter the Debit Notes detail tab to the
    same rows the main sheet counted."""
    inv = norm_inv(inv)
    if inv in universe:
        return inv
    core = inv
    if core[:3] == "INV":
        core = core[3:]
    elif core[:1] == "V":
        core = core[1:]
    for cand in ("INV" + core, "INV" + core.rstrip("/") + "/",
                 "INV" + core.rstrip("/"), "INV" + core + "/"):
        if cand in universe:
            return cand
    return None


def _build_detail_sheets(wb, results):
    """#00007: create the 5 split detail tabs (Payments / Debit Notes / Credit
    Notes / PMDD / AP-AR & Manual Adj) from the line-item detail collected in
    `reconcile_zepto`. Returns row-index maps for the invoice-linked tabs
    ({inv: first_row}) so the main sheet can hyperlink into them."""
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter

    details = getattr(results, "details", None) or {}
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    title_font = Font(bold=True, size=13, color="123C69")
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    total_font = Font(bold=True, size=10, color="123C69")
    money_fmt = "#,##0.00"

    def make_sheet(title, headers, rows, money_cols, key_col, widths, hide_minus_cols=frozenset()):
        ws = wb.create_sheet(title=title)
        ncol = len(headers)
        c = ws.cell(row=1, column=1, value=title)
        c.font = title_font; c.alignment = Alignment(horizontal="left", vertical="center")
        ws.merge_cells(start_row=1, start_column=1, end_row=1, end_column=ncol)
        ws.row_dimensions[1].height = 20
        for j, htext in enumerate(headers, start=1):
            hc = ws.cell(row=2, column=j, value=htext)
            hc.font = hdr_font; hc.fill = hdr_fill; hc.border = border
            hc.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        ws.row_dimensions[2].height = 20
        rowmap, totals, r = {}, {mc: 0.0 for mc in money_cols}, 3
        for item in rows:
            for j in range(1, ncol + 1):
                val = item[j - 1]
                cell = ws.cell(row=r, column=j, value=val)
                cell.border = border
                if j in money_cols:
                    cell.number_format = _DN_DISPLAY_FMT if j in hide_minus_cols else money_fmt
                    cell.alignment = Alignment(horizontal="right")
                    totals[j] += _to_float(val)
            if key_col is not None:
                k = _canon_inv(item[key_col - 1])
                if k and k not in rowmap:
                    rowmap[k] = r
            r += 1
        if rows:
            tc = ws.cell(row=r, column=1, value="TOTAL"); tc.font = total_font; tc.border = border
            for mc in money_cols:
                cell = ws.cell(row=r, column=mc, value=round(totals[mc], 2))
                cell.number_format = _DN_DISPLAY_FMT if mc in hide_minus_cols else money_fmt
                cell.font = total_font
                cell.border = border; cell.alignment = Alignment(horizontal="right")
        else:
            ws.cell(row=3, column=1, value="(no records)").font = Font(italic=True, color="808080")
        for j, w in enumerate(widths, start=1):
            ws.column_dimensions[get_column_letter(j)].width = w
        ws.freeze_panes = "A3"
        return rowmap

    # Filter the invoice-linked tabs to the tracker's universe so their totals
    # reconcile with the main sheet (the source files also carry prior-year /
    # out-of-range line items that never appear as a main-sheet row).
    universe = {r.get("invoice_number", "") for r in results}

    pay = sorted((p for p in details.get("payments", []) if p["invoice"] in universe),
                 key=lambda x: x["invoice"])
    pay_rows = [[p["invoice"], p["ref"], round(p["excl"], 2), round(p["tds"], 2), round(p["incl"], 2),
                 p.get("payment_date", ""), p.get("advice_no", "")] for p in pay]

    # Debit notes: resolve each (possibly malformed) ref key onto its universe
    # invoice exactly as the main-sheet remap did; drop those that don't resolve.
    dn = []
    for d in details.get("debit_notes", []):
        u = _resolve_to_universe(d["invoice"], universe)
        if u:
            dn.append((u, d["ref"], d["amount"], d.get("payment_date", ""), d.get("advice_no", "")))
    dn.sort(key=lambda x: x[0])
    dn_rows = [[u, ref, round(amt, 2), pdate, adv] for (u, ref, amt, pdate, adv) in dn]

    cn = sorted((c for c in details.get("credit_notes", []) if c["invoice"] in universe),
                key=lambda x: x["invoice"])
    cn_rows = [[c["invoice"], c.get("number", ""), round(c["amount"], 2)] for c in cn]
    pmdd_rows = [[p["ref"], round(p["amount"], 2), p.get("payment_date", ""), p.get("advice_no", "")]
                 for p in details.get("pmdd", [])]
    apar_rows = [[a["ref"], round(a["amount"], 2), a.get("payment_date", ""), a.get("advice_no", "")]
                 for a in details.get("ap_ar", [])]

    # Payment Advice Consolidate: every line item from every payment-advice PDF,
    # verbatim, tagged with its Payment Advice No + Payment Date. Not filtered to
    # the universe — this is the raw consolidation, and its Payment Amt total is
    # the "Amount Received in Bank" on the Summary.
    con_rows = [[c.get("advice_no", ""), c.get("payment_date", ""), c.get("type", ""),
                 c.get("doc_no", ""), c.get("ref_doc", ""),
                 round(_to_float(c.get("amount", 0)), 2), round(_to_float(c.get("tds", 0)), 2),
                 round(_to_float(c.get("payment_amt", 0)), 2)]
                for c in details.get("consolidate", [])]

    pay_date_adv_hdr = ["Payment Date", "Payment Advice No"]
    pay_map = make_sheet("Payments",
                         ["Invoice", "Ref / Payment Doc", "Amount (Excl. TDS)", "TDS", "Amount (Incl. TDS)"] + pay_date_adv_hdr,
                         pay_rows, {3, 4, 5}, key_col=1, widths=[22, 22, 18, 12, 18, 14, 20])
    # Payment Advice Consolidate sits right after Payments (matches the CA layout).
    make_sheet("Payment Advice Consolidate",
               ["Payment Advice No", "Payment Date", "Type of Document", "Doc No", "Ref Doc",
                "Amount", "TDS", "Payment Amt"],
               con_rows, {6, 7, 8}, key_col=None, widths=[20, 14, 18, 16, 22, 16, 12, 16])
    dn_map = make_sheet("Debit Notes", ["Invoice", "Ref", "Amount"] + pay_date_adv_hdr, dn_rows, {3}, key_col=1,
                        widths=[22, 26, 16, 14, 20], hide_minus_cols={3})   # show DN amount without the minus sign
    cn_map = make_sheet("Credit Notes", ["Invoice", "Credit Note No", "Amount"], cn_rows, {3}, key_col=1, widths=[22, 22, 16])
    make_sheet("PMDD", ["Ref / Description", "Amount"] + pay_date_adv_hdr, pmdd_rows, {2}, key_col=None, widths=[42, 18, 14, 20])
    make_sheet("AP-AR & Manual Adj", ["Ref / Description", "Amount"] + pay_date_adv_hdr, apar_rows, {2}, key_col=None, widths=[42, 18, 14, 20])
    return {"payments": pay_map, "debit_notes": dn_map, "credit_notes": cn_map}


def build_zepto_workbook(results: list[dict], payload: dict | None = None):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.formatting.rule import CellIsRule
    wb = Workbook()
    ws = wb.active
    ws.title = "1. Invoice Tracker"
    thin = Side(style="thin", color="D9D9D9")
    border = Border(left=thin, right=thin, top=thin, bottom=thin)
    hdr_font = Font(bold=True, size=10, color="FFFFFF")
    hdr_fill = PatternFill("solid", fgColor="1F3864")
    group_font = Font(bold=True, size=9, color="123C69")
    group_fill = PatternFill("solid", fgColor="D9E2F3")
    zebra_fill = PatternFill("solid", fgColor="F5F7FA")

    # Row 1: source-group labels (merged)
    # PO-first layout: PO(A) + From Tally(B-L) | Pending(M, formula) |
    # Payment Advice(N-P) | From Zepto Ledger/Payment Advice(Q-U, now
    # includes Credit Note No) | Computed(V-X) | From Zepto Dashboard(Y-AA) |
    # From Courier(AB-AD). Shifted one letter right of the old layout
    # (U..AC) by the "Credit Note No" column insertion in the 5-col task.
    ws.append([""] * len(_HEADERS))
    # #00001/#00002: "Aging" group (Due Date + Due Status) inserted right after
    # the Computed block (Status = X); everything after it shifts +2 letters.
    groups = [("From Tally","B","L"),("Payment Advice (Zepto Portal)","N","P"),
              ("From Zepto Ledger / Payment Advice","Q","U"),("Computed","V","X"),
              ("Aging (Invoice Date + 30d)","Y","Z"),
              ("From Zepto Dashboard","AA","AC"),("From Courier (Delhivery)","AD","AF")]
    for label, c1, c2 in groups:
        ws.merge_cells(f"{c1}1:{c2}1")
        cell = ws[f"{c1}1"]; cell.value = label
        cell.font = group_font
        cell.fill = group_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 18

    # Row 2: column headers
    ws.append(_HEADERS)
    for i, _ in enumerate(_HEADERS, start=1):
        c = ws.cell(row=2, column=i)
        c.font = hdr_font; c.fill = hdr_fill; c.border = border
        c.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[2].height = 30

    # Data rows from row 3
    for idx, row in enumerate(results):
        r = idx + 3
        row_fill = zebra_fill if idx % 2 == 1 else None
        for col_i, key in enumerate(COLUMN_KEYS, start=1):
            letter = _LETTERS[col_i - 1]
            if key == "pending_amount":
                val = f"=F{r}"
            elif key == "gross_outstanding":
                # #00004: Pending - Payment Received (excl. TDS)
                val = f"=M{r}-O{r}"
            elif key == "net_outstanding":
                # #00004: Gross - Credit Note - Debit Note
                val = f"=M{r}-O{r}-T{r}-Q{r}"
            elif key == "status":
                if row.get("invoice_not_in_ledger"):
                    val = row.get("status", "")
                else:
                    # Status from NET OUTSTANDING (col W) only — Gross (V) ignored.
                    val = f'=IF(W{r}<=10,"Paid","Not Paid")'
            else:
                val = row.get(key, "")
                if val == "" and key in _MONEY_KEYS:
                    val = 0
            c = ws.cell(row=r, column=col_i, value=val)
            c.border = border
            if key not in _TEXT_KEYS and (key in _MONEY_KEYS or key in _FORMULA_COLS - {"status"}):
                # Debit Note keeps its negative VALUE (formula -Q unchanged) but
                # displays without the minus sign.
                c.number_format = _DN_DISPLAY_FMT if key == "debit_note_issued" else "#,##0.00"
                c.alignment = Alignment(horizontal="right")
            elif key == "status":
                c.alignment = Alignment(horizontal="center")
            if row_fill is not None:
                c.fill = row_fill

    last_row = len(results) + 2   # last data row (header is row 2, data starts row 3)

    # Column widths — wider for names/refs/GSTIN, tighter for money/status.
    for col_i, key in enumerate(COLUMN_KEYS, start=1):
        ws.column_dimensions[_LETTERS[col_i - 1]].width = _COLUMN_WIDTHS.get(key, 16)

    # Conditional formatting on the Status column (formula-driven cell, so
    # color must be applied via CF rules rather than a static fill).
    if last_row >= 3:
        status_range = f"X3:X{last_row}"
        paid_rule = CellIsRule(
            operator="equal", formula=['"Paid"'],
            fill=PatternFill("solid", fgColor=_PAID_FILL_HEX),
            font=Font(color=_PAID_FONT_HEX, bold=True),
        )
        not_paid_rule = CellIsRule(
            operator="equal", formula=['"Not Paid"'],
            fill=PatternFill("solid", fgColor=_NOTPAID_FILL_HEX),
            font=Font(color=_NOTPAID_FONT_HEX, bold=True),
        )
        ws.conditional_formatting.add(status_range, paid_rule)
        ws.conditional_formatting.add(status_range, not_paid_rule)

    # Freeze header rows (1-2) + PO/Date/Invoice/Sales-Order columns (A-D)
    # so key identifiers stay visible when scrolling right/down.
    ws.freeze_panes = "E3"

    build_summary_sheet(wb, results)

    # #00007: split detail tabs + click-to-jump hyperlinks from the main sheet.
    maps = _build_detail_sheets(wb, results)
    from openpyxl.worksheet.hyperlink import Hyperlink
    link_font = Font(color=_DETAIL_LINK_FONT_HEX, underline="single")

    def _link(row_i, col_i, sheet_title, target_row):
        cell = ws.cell(row=row_i, column=col_i)
        cell.hyperlink = Hyperlink(ref=cell.coordinate, location=f"'{sheet_title}'!A{target_row}")
        cell.font = link_font   # tint to signal it's clickable (number format/alignment kept)

    for idx, row in enumerate(results):
        r = idx + 3
        inv = _canon_inv(row.get("invoice_number", ""))
        if inv in maps["payments"]:
            tr = maps["payments"][inv]
            _link(r, 14, "Payments", tr)      # N  Payment Received (Incl TDS)
            _link(r, 15, "Payments", tr)      # O  Payment Received (Excl TDS)
        if inv in maps["debit_notes"]:
            _link(r, 17, "Debit Notes", maps["debit_notes"][inv])   # Q  Debit Note Issued
        if inv in maps["credit_notes"]:
            tr = maps["credit_notes"][inv]
            _link(r, 20, "Credit Notes", tr)  # T  Credit Note Issued
            _link(r, 21, "Credit Notes", tr)  # U  Credit Note No
    return wb
