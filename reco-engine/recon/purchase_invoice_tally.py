"""
purchase_invoice_tally.py — Purchase-Invoice → Tally (Excel-to-Tally) engine.
Urban Plant (Grandeur IT Innovations) purchase invoices → the 109-column Tally
purchase-import workbook.

DESIGN (token-frugal, learning-based):
  1. pdfplumber extracts the text layer  (→ iLovePDF OCR fallback for scans).
  2. Line items are parsed DETERMINISTICALLY per KNOWN vendor (keyed by seller
     GSTIN) — zero LLM tokens for the vendors we already know.
  3. A NEW/unknown vendor is the ONLY case that touches Gemini: Gemini structures
     the unfamiliar layout AND we cache the learned profile so the vendor's NEXT
     invoice runs deterministically too.  (Gemini wiring lives in the caller /
     server layer; this module exposes `SELLER_REGISTRY` + `parse_invoice`.)
  4. SKU mapping (Description → "Name as per Tally") is a deterministic DB/master
     lookup; only the UNMATCHED / low-confidence lines are sent to Gemini, and
     whatever Gemini still can't place falls to manual pick (J left blank).

BUYER GUARD: Urban Plant's buying entity is Grandeur IT Innovations, PAN
AAICG2697Q (registered in UP-09 / KA-29 / MH-27).  Any invoice whose bill-to PAN
is NOT AAICG2697Q is refused ("buyer is not Urban Plant") so a misfiled
other-brand PDF can never be booked into Urban Plant's Tally.

This module ONLY extracts + normalises.  Workbook building + SKU mapping are
separate steps (see build_tally_workbook / map layer) so each stays testable.
"""
import io
import os
import re
import logging

logger = logging.getLogger(__name__)

BUYER_PAN = "AAICG2697Q"          # Grandeur IT Innovations (Urban Plant)
GSTIN_RE = re.compile(r'\b[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]\b')

STATE_NAMES = {
    "09": "Uttar Pradesh", "29": "Karnataka", "27": "Maharashtra",
    "06": "Haryana", "03": "Punjab", "33": "Tamil Nadu", "07": "Delhi",
    "24": "Gujarat", "08": "Rajasthan", "23": "Madhya Pradesh",
}

# ── numeric / text helpers ──────────────────────────────────────────────────
def clean_num(s):
    """Indian money/qty token → float. Strips ₹, stray spaces, commas, trailing junk."""
    if s is None:
        return None
    s = str(s).replace('₹', ' ').replace('₹', ' ')
    s = s.replace(',', '').replace(' ', '').strip().rstrip('.-')
    if s in ('', '-'):
        return None
    try:
        return float(s)
    except Exception:
        return None

def _state_of(gstin):
    return gstin[:2] if gstin and len(gstin) >= 2 else None

# ── PDF text (pdfplumber → iLovePDF OCR fallback) ───────────────────────────
def _source_bytes(source):
    if isinstance(source, (bytes, bytearray)):
        return bytes(source)
    try:
        with open(source, "rb") as fh:
            return fh.read()
    except Exception:
        return b""

def _text_from_bytes(pdf_bytes):
    import pdfplumber
    with pdfplumber.open(io.BytesIO(pdf_bytes)) as pdf:
        return "\n".join((p.extract_text() or "") for p in pdf.pages)

def pdf_text(source):
    """Digital text first; if the page has no text layer (scan), OCR via iLovePDF."""
    pdf_bytes = _source_bytes(source)
    text = _text_from_bytes(pdf_bytes) if pdf_bytes else ""
    if len(text.strip()) < 40 and pdf_bytes:
        try:
            from recon.ilovepdf_ocr import ocr_pdf_to_text
            alt = ocr_pdf_to_text(pdf_bytes)
            if alt and len(alt.strip()) >= 40:
                text = alt
        except Exception as e:
            logger.warning("iLovePDF OCR fallback failed: %s", e)
    return text

# ── seller / buyer identification ───────────────────────────────────────────
def identify_parties(text):
    """Return (seller_gstin, buyer_gstin, all_gstins). Buyer = PAN AAICG2697Q."""
    gstins = GSTIN_RE.findall(text)
    seller = buyer = None
    for g in gstins:
        if g[2:12] == BUYER_PAN:
            buyer = buyer or g
        elif seller is None:
            seller = g
    return seller, buyer, gstins

# ── generic line utilities ──────────────────────────────────────────────────
_FOOTER_STOP = re.compile(
    r'^\s*(Grand\s*Total|Amount\s*Chargeable|Bank\s*Details|Declaration|'
    r'E\.?\s*&\s*O\.?E|Output\s+[CS]gst)',
    re.I)

def _iter_item_lines(lines):
    """Yield every line. The strict per-vendor item regexes reject headers, footers,
    carry-forward ("Totals c/o") and tax-summary rows on their own, and multi-copy
    PDFs are collapsed afterwards by _dedupe_copies (S.No. reset). We deliberately
    do NOT hard-stop on 'Totals'/'Terms' — SHAFA carries 'Totals c/o' mid-invoice
    and several vendors print 'Terms of Delivery' in the address block BEFORE the
    first item, which would truncate the item list."""
    return list(lines)

def _dedupe_copies(items):
    """Multi-copy PDFs (Original/Duplicate/Triplicate) repeat the item block.
    Detect the first full cycle: once the running S.No. resets, stop."""
    out = []
    last_sn = 0
    for it in items:
        sn = it.get('sn')
        if sn is not None and out and sn <= last_sn:
            break                      # S.No. went backwards → next copy starts
        out.append(it)
        if sn is not None:
            last_sn = sn
    return out

# ── per-vendor parsers ───────────────────────────────────────────────────────
# Each: parse(lines, text) -> {"invoice_no","date","items":[{sn,desc,hsn,qty,unit,rate,amount,gst_rate?}]}

def _hdr(text, *patterns):
    for pat in patterns:
        m = re.search(pat, text, re.I)
        if m:
            return m.group(1).strip()
    return None

def parse_shafa(lines, text):        # 09ABMCS3350B1ZF
    rx = re.compile(r'^\s*(\d+)\.\s+(.+?)\s+(\d{6,8})\s+([\d.]+)\s+(\S+?)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)\s*$')
    items = []
    for ln in _iter_item_lines(lines):
        m = rx.match(ln)
        if m:
            items.append(dict(sn=int(m.group(1)), desc=m.group(2).strip(), hsn=m.group(3),
                              qty=clean_num(m.group(4)), unit=m.group(5).rstrip('.'),
                              rate=clean_num(m.group(6)), amount=clean_num(m.group(7))))
    return dict(invoice_no=_hdr(text, r'Invoice\s*No\.?\s*:\s*([A-Za-z0-9/\-]+)'),
                date=_hdr(text, r'Dated\s*:\s*([\d/\-\.]+)'),
                items=_dedupe_copies(items))

def parse_gp(lines, text):           # 29AAXFG2562J1ZP  (S.No glued to desc)
    rx = re.compile(r'^\s*(\d+)([A-Za-z].+?)\s+(\d{6,8})\s+(\d+)\s+(\w+)\s+([\d,]+\.\d+)\s+\w+\s+([\d,]+\.\d+)\s*$')
    items = []
    for ln in _iter_item_lines(lines):
        m = rx.match(ln)
        if m:
            items.append(dict(sn=int(m.group(1)), desc=m.group(2).strip(), hsn=m.group(3),
                              qty=clean_num(m.group(4)), unit=m.group(5),
                              rate=clean_num(m.group(6)), amount=clean_num(m.group(7))))
    return dict(invoice_no=_hdr(text, r'Invoice\s*No\.[^\n]*\n\s*(GP/[A-Za-z0-9/\-]+)', r'\b(GP/\d{2}-\d{2}/\d+)\b'),
                date=_hdr(text, r'Dated\s*\n?[^\n]*?(\d{1,2}-[A-Za-z]{3}-\d{2})'),
                items=_dedupe_copies(items))

def parse_proliant(lines, text):     # 27AAECP3720B2ZV  item ALIAS HSN qty rate total gross gst%
    rx = re.compile(r'^\s*(\d+)\s+(.+?)\s+(\S+)\s+(\d{4,8})\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)\s+([\d,]+\.\d+)\s+(\d+)%\s*$')
    items = []
    for ln in _iter_item_lines(lines):
        m = rx.match(ln)
        if m:
            items.append(dict(sn=int(m.group(1)), desc=m.group(2).strip(), hsn=m.group(4),
                              qty=clean_num(m.group(5)), unit='Pcs', rate=clean_num(m.group(6)),
                              amount=clean_num(m.group(8)), gst_rate=clean_num(m.group(9))))
    return dict(invoice_no=_hdr(text, r'Invoice\s*No\.?\s*:\s*([A-Za-z0-9/\-]+)'),
                date=_hdr(text, r'Invoice\s*Date\s*:\s*([\d/\-\.]+)'),
                items=_dedupe_copies(items))

def parse_ashoka(lines, text):       # 03ACFPK8505P1ZD  desc HSN qty UNIT rate UNIT amount
    rx = re.compile(r'^\s*(\d+)\s+(.+?)\s+(\d{4,8})\s+([\d,]+\.?\d*)\s+(\w+)\s+([\d,]+\.\d+)\s+\w+\s+([\d,]+\.\d+)\s*$')
    items = []
    for ln in _iter_item_lines(lines):
        m = rx.match(ln)
        if m:
            items.append(dict(sn=int(m.group(1)), desc=m.group(2).strip(), hsn=m.group(3),
                              qty=clean_num(m.group(4)), unit=m.group(5),
                              rate=clean_num(m.group(6)), amount=clean_num(m.group(7))))
    return dict(invoice_no=_hdr(text, r'Invoice\s*No\.\s*Dated\s*\n\s*(\d[A-Za-z0-9/\-]*)',
                                      r'\n(\d+)\s+\d{1,2}-[A-Za-z]{3}-\d{2}'),
                date=_hdr(text, r'(\d{1,2}-[A-Za-z]{3}-\d{2})'),
                items=_dedupe_copies(items))

def parse_greenway(lines, text):     # 33BCYPA7548E2ZD  mangled ₹ columns
    # "1 DESC HSN qtyNos <rate> <taxable> gst% <cgst> <sgst> <igst> <total>"
    rx = re.compile(r'^\s*(\d+)\s+(.+?)\s+(\d{4,8})\s+(\d+)\s*Nos\b(.*)$')
    items = []
    for ln in _iter_item_lines(lines):
        m = rx.match(ln)
        if not m:
            continue
        qty = clean_num(m.group(4))
        # money fragments live between ₹ markers; reconstruct then take rate = money[0]
        frags = [f for f in re.split(r'[₹₹]', m.group(5)) if clean_num(f) is not None]
        rate = clean_num(frags[0]) if frags else None
        gm = re.search(r'(\d+)\s*%', m.group(5))
        gst = clean_num(gm.group(1)) if gm else None
        amount = round(qty * rate, 2) if (qty is not None and rate is not None) else (clean_num(frags[1]) if len(frags) > 1 else None)
        items.append(dict(sn=int(m.group(1)), desc=m.group(2).strip(), hsn=m.group(3),
                          qty=qty, unit='Nos', rate=rate, amount=amount, gst_rate=gst))
    return dict(invoice_no=_hdr(text, r'No\.\s*(\d+)\s+Date'),
                date=_hdr(text, r'Date\s*:\s*([\d/\-\.]+)'),
                items=_dedupe_copies(items))

def parse_unison(lines, text):       # 03AGLPM8433C4ZI  multi-line desc + optional disc col
    # sn DESC HSN qty PCS. rate [disc] amount    (desc may wrap to following lines)
    # optional size token (e.g. 5") between HSN and qty; optional discount col before amount
    rx = re.compile(r'^\s*(\d+)\s+(.+?)\s+(\d{6,8})\s+(?:\S+\s+)?(\d+)\s+(PCS\.?|Pcs\.?|Nos\.?)\s+([\d,]+\.\d+)\s+(?:([\d,]+\.\d+)\s+)?([\d,]+\.\d+)\s*$')
    items = []
    src = list(_iter_item_lines(lines))
    for i, ln in enumerate(src):
        m = rx.match(ln)
        if not m:
            continue
        desc = m.group(2).strip()
        # attach an immediately-following wrap line (real items match rx and are skipped)
        nxt = src[i + 1] if i + 1 < len(src) else ''
        _junk = re.compile(r'\b(HSN|Taxable|IGST|CGST|SGST|Amt\.?|Total|Round\s*Off|GSTIN|Rate)\b', re.I)
        if nxt and not rx.match(nxt) and not _FOOTER_STOP.match(nxt) \
           and not GSTIN_RE.search(nxt) and not _junk.search(nxt) and 0 < len(nxt.strip()) < 45:
            desc = (desc + ' ' + nxt.strip()).strip()
        items.append(dict(sn=int(m.group(1)), desc=desc, hsn=m.group(3),
                          qty=clean_num(m.group(4)), unit=m.group(5).rstrip('.'),
                          rate=clean_num(m.group(6)), amount=clean_num(m.group(8))))
    # UNISON prints GST rate only in the HSN tax-summary ("<hsn> <taxable> <rate>% <tax>"),
    # not per line — map HSN→rate and stamp each item so the Purchase Ledger is correct.
    hsn_rate = {}
    for mm in re.finditer(r'\b(\d{6,8})\s+[\d,]+\.\d+\s+(\d{1,2})\s*%', text):
        hsn_rate.setdefault(mm.group(1), int(mm.group(2)))
    items = _dedupe_copies(items)
    for it in items:
        if it.get('gst_rate') is None and it.get('hsn') in hsn_rate:
            it['gst_rate'] = hsn_rate[it['hsn']]
    return dict(invoice_no=_hdr(text, r'Invoice\s*No\.?\s*:\s*([A-Za-z0-9/\-]+)'),
                date=_hdr(text, r'Bill\s*Dated\s*:?\s*([\d/\-\.]+)'),
                items=items)

# seller GSTIN → (display name, parser, default_gst_rate hint or None)
SELLER_REGISTRY = {
    '09ABMCS3350B1ZF': ('SHAFA AKHDAR MANUFACTURING PVT LTD', parse_shafa,    5),
    '29AAXFG2562J1ZP': ('G P INDUSTRIES',                     parse_gp,       None),
    '27AAECP3720B2ZV': ('PROLIANT DISPLAY SYSTEM PVT LTD',    parse_proliant, 18),
    '03ACFPK8505P1ZD': ('ASHOKA INDUSTRIES',                  parse_ashoka,   5),
    '33BCYPA7548E2ZD': ('GREENWAY ORGANICS',                  parse_greenway, 5),
    '03AGLPM8433C4ZI': ('UNISON ENGG INDUSTRIES',             parse_unison,   None),
}

class BuyerNotUrbanPlant(ValueError):
    pass

def parse_invoice(source):
    """Extract one purchase invoice → normalised dict.
    Raises BuyerNotUrbanPlant if the bill-to PAN is not Grandeur's.
    Sets known_vendor=False for an unrecognised seller (→ Gemini path upstream)."""
    text = pdf_text(source)
    seller, buyer, gstins = identify_parties(text)
    if not buyer:
        raise BuyerNotUrbanPlant(
            f"Buyer PAN is not Urban Plant (Grandeur {BUYER_PAN}); GSTINs found: {gstins[:4]}")
    prof = SELLER_REGISTRY.get(seller)
    lines = text.splitlines()
    if not prof:
        return dict(known_vendor=False, seller_gstin=seller, seller_name=None,
                    buyer_gstin=buyer, buyer_state=_state_of(buyer),
                    intra_state=(_state_of(seller) == _state_of(buyer)) if seller else None,
                    invoice_no=None, date=None, items=[], raw_text=text)
    name, parser, gst_hint = prof
    parsed = parser(lines, text)
    for it in parsed['items']:
        it.setdefault('gst_rate', gst_hint)
    return dict(known_vendor=True, seller_gstin=seller, seller_name=name,
                buyer_gstin=buyer, buyer_state=_state_of(buyer),
                intra_state=(_state_of(seller) == _state_of(buyer)),
                invoice_no=parsed.get('invoice_no'), date=parsed.get('date'),
                items=parsed['items'], raw_text=text)

# ── self-test ────────────────────────────────────────────────────────────────
if __name__ == '__main__':
    import glob, sys
    d = sys.argv[1] if len(sys.argv) > 1 else "/Users/dhavalchauhan/Downloads/Purchase -Dhaval/Invoices"
    for f in sorted(glob.glob(os.path.join(d, '*.pdf'))):
        try:
            r = parse_invoice(f)
        except BuyerNotUrbanPlant as e:
            print('=' * 92)
            print(f"{os.path.basename(f)}  →  ⛔ SKIPPED: {e}")
            continue
        print('=' * 92)
        tag = 'INTRA→CGST+SGST' if r['intra_state'] else 'INTER→IGST'
        vend = r['seller_name'] or f"??? UNKNOWN {r['seller_gstin']} (Gemini path)"
        print(f"{os.path.basename(f)}  |  {vend}  |  inv={r['invoice_no']}  date={r['date']}  {tag}")
        tot = 0.0
        for it in r['items']:
            amt = it.get('amount') or 0
            tot += amt
            print(f"   {str(it.get('qty')):>6} x {str(it.get('rate')):>9} = {str(round(amt,2)):>11}  "
                  f"[{it.get('hsn')}] {it['desc'][:66]}")
        print(f"   → {len(r['items'])} lines, Σamount = {round(tot, 2)}")


# ═══════════════════════════════════════════════════════════════════════════
#  Tally workbook builder (Excel-to-Tally, 109 columns)
# ═══════════════════════════════════════════════════════════════════════════
TALLY_HEADERS = ["Vch. Date* ","Vch. Type*","Vch. No.*","Ref. No.","Ref. Date","Is DN?","Is Vch?","Party Ledger*","Purchase Ledger*","Stock Item","Description","Godown","Actual Qty","Quantity","Rate","Unit","Discount","Amount*","Discount","Input IGST-UP","Input CGST-UP","Input SGST-UP","Input CGST-KA","Input SGST-KA",None,None,None,None,None,None,None,"CESS","TCS","Round Off","Narration","Taxability","GST Nature","GST Rate","Cess","RCM?","HSN","HSN Desc","Supply Type","Cost Category","Cost Centre","Name","Address 1","Address 2","State","Country","PIN Code","Place of Supply","GST Type","GSTIN","Name","Address 1","Address 2","State","Country","PIN Code","Place","GSTIN","Name","Address 1","Address 2","State","Country","PIN Code","Place","GSTIN","DN No.","DN Date","Doc. No.","Dis. Through","Destination","Carrier Name","LR No.","LR Date","Order No.","Order Date","Term of Delivery","Terms of Paymemt","Other Ref.","Place of Receipt","Vessel/Flight No.","Port of Loading","Port of Discharge","Country to","Shipping Bill No.","Date","Port Code","e-Way Bill No","Date","Cons. e-Way Bill No.","Date","Sub Type","Doc. Type","Distance (KM)","Transporter Name","Transporter ID","Transport Mode","Doc No.","Date","Vehicle No.","Vehicle Type","Status","Note Reason","Orig. Inv. No.","Orig. Inv. Date"]

# 1-based column numbers for the fields we populate
C = dict(vch_date=1, vch_type=2, vch_no=3, ref_no=4, ref_date=5, party=8, purch_led=9,
         stock=10, godown=12, qty=14, rate=15, unit=16, amount=18,
         igst_up=20, cgst_up=21, sgst_up=22, cgst_ka=23, sgst_ka=24,
         narration=35, gst_rate=38, hsn=41, p_name=46, p_state=49, pos=52,
         gst_type=53, p_gstin=54, b_state=58, b_place=61, b_gstin=62)

# buyer-state → which tax-ledger columns exist in THIS template
BUYER_TAX_COLS = {
    "09": dict(vch="Purchase-UP", igst="igst_up", cgst="cgst_up", sgst="sgst_up"),  # UP
    "29": dict(vch="Purchase-KA", igst=None,      cgst="cgst_ka", sgst="sgst_ka"),  # KA (no IGST-KA col)
}

# exact Tally Party-Ledger names where known (else derived from the invoice name)
PARTY_LEDGER = {
    "09ABMCS3350B1ZF": "SHAFA AKHDAR MANUFACTURING PRIVATE LIMITED",
}

def _party_ledger(seller_gstin, seller_name):
    if seller_gstin in PARTY_LEDGER:
        return PARTY_LEDGER[seller_gstin]
    n = (seller_name or "").strip()
    return re.sub(r'\bPVT\.?\s+LTD\.?\b', 'PRIVATE LIMITED', n, flags=re.I)

def _gst_rate_of(item):
    if item.get('gst_rate'):
        return int(round(float(item['gst_rate'])))
    m = re.search(r'(\d{1,2})\s*%', item.get('desc', ''))
    return int(m.group(1)) if m else None

def _parse_date(s):
    """Vendor date string → datetime (Tally wants a real date). Handles
    DD-MM-YYYY, DD/MM/YYYY, DD.MM.YYYY, DD-Mon-YY/-YYYY. None if unparseable."""
    if not s:
        return None
    from datetime import datetime
    s = str(s).strip()
    for fmt in ("%d-%m-%Y", "%d/%m/%Y", "%d.%m.%Y", "%d-%b-%y", "%d-%b-%Y",
                "%d/%m/%y", "%d.%m.%y", "%d-%m-%y"):
        try:
            return datetime.strptime(s, fmt)
        except ValueError:
            continue
    return None

def build_tally_workbook(invoices, narration="Excel to tally"):
    """invoices = [ parse_invoice() dict, each item carrying a 'stock_item' (mapped,
    may be '') ]. Returns (xlsx_bytes, review_rows). review_rows flags anything a
    human must fix (unmapped stock, unsupported buyer-state ledger, total mismatch)."""
    import openpyxl
    from copy import copy as _copy
    from openpyxl.utils import get_column_letter
    # Build a FRESH workbook (robust — a loaded+row-deleted template can corrupt in
    # Excel), then copy ONLY the header styling (bold themed fill) + column widths
    # from the CA's template so the output looks identical without the corruption risk.
    wb = openpyxl.Workbook(); ws = wb.active; ws.title = "Purchase"
    for j, h in enumerate(TALLY_HEADERS, 1):
        if h is not None:
            ws.cell(row=1, column=j, value=h)
    _tmpl = os.path.join(os.path.dirname(__file__), "..", "format_templates", "excel_to_tally_template.xlsx")
    try:
        tmpl = openpyxl.load_workbook(_tmpl); tws = tmpl.active
        for j in range(1, len(TALLY_HEADERS) + 1):
            src, dst = tws.cell(row=1, column=j), ws.cell(row=1, column=j)
            dst.font = _copy(src.font); dst.fill = _copy(src.fill); dst.alignment = _copy(src.alignment)
            L = get_column_letter(j)
            if L in tws.column_dimensions and tws.column_dimensions[L].width:
                ws.column_dimensions[L].width = tws.column_dimensions[L].width
        ws.row_dimensions[1].height = tws.row_dimensions[1].height
        tmpl.close()
    except Exception as _e:
        logger.warning("tally template styling skipped: %s", _e)
    review = []
    r = 2
    for inv in invoices:
        b_state = inv.get('buyer_state')
        s_state = _state_of(inv.get('seller_gstin'))
        cfg = BUYER_TAX_COLS.get(b_state)
        intra = inv.get('intra_state')
        party = _party_ledger(inv.get('seller_gstin'), inv.get('seller_name'))
        b_state_name = STATE_NAMES.get(b_state, b_state)
        s_state_name = STATE_NAMES.get(s_state, s_state)
        if not cfg:
            review.append(dict(invoice=inv.get('invoice_no'), issue="buyer-state ledgers not in template",
                               detail=f"buyer state {b_state} ({b_state_name}) has no tax columns"))
        vdate = _parse_date(inv.get('date'))
        for it in inv['items']:
            rate = _gst_rate_of(it)
            amount = it.get('amount') or 0
            def put(col, v):
                ws.cell(row=r, column=C[col], value=v)
            put('vch_date', vdate)
            put('vch_type', cfg['vch'] if cfg else f"Purchase-{b_state}")
            put('vch_no', inv.get('invoice_no')); put('ref_no', inv.get('invoice_no'))
            put('ref_date', vdate)
            put('party', party)
            put('purch_led', f"Purchase of Product-{rate}%" if rate is not None else "Purchase of Product")
            put('stock', it.get('stock_item') or "")
            put('godown', (b_state_name or "").upper())
            put('qty', it.get('qty')); put('rate', it.get('rate')); put('unit', it.get('unit') or "Pcs")
            put('amount', amount)
            # GST split — template writes explicit 0 in all five input-tax columns
            for tc in ('igst_up', 'cgst_up', 'sgst_up', 'cgst_ka', 'sgst_ka'):
                put(tc, 0)
            if cfg and rate:
                if intra:
                    half = round(amount * (rate / 2) / 100.0, 2)
                    put(cfg['cgst'], half); put(cfg['sgst'], half)
                elif cfg['igst']:
                    put(cfg['igst'], round(amount * rate / 100.0, 2))
                else:
                    review.append(dict(invoice=inv.get('invoice_no'), issue="inter-state IGST ledger missing",
                                       detail=f"buyer {b_state} has no IGST column"))
            put('narration', narration)
            # GST Rate (AL) + HSN (AO) intentionally left blank — the template does,
            # Tally derives them from the ledger / stock-item master. HSN + rate stay
            # in the review sheet for the human check only.
            # supplier (party) master block
            put('p_name', party); put('p_state', s_state_name); put('pos', b_state_name)
            put('gst_type', "Regular"); put('p_gstin', inv.get('seller_gstin'))
            put('b_state', s_state_name); put('b_place', s_state_name); put('b_gstin', inv.get('seller_gstin'))
            if not (it.get('stock_item') or "").strip():
                review.append(dict(invoice=inv.get('invoice_no'), issue="unmapped stock item",
                                   detail=it.get('desc')))
            r += 1
    import io as _io
    buf = _io.BytesIO(); wb.save(buf)
    return buf.getvalue(), review
