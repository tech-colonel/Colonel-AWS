"""GSTR-1 vs Books Reconciliation Engine (v2 — Accountant Workflow).

Steps (in order):
  0. GSTR-1 Pivot   (optional — only when gstr1_pdf uploaded)
     Validates OCTA Excel monthly totals against GST-portal PDF.
  1. GSTR-1 vs GSTR-3B monthly comparison  → Section 1 of GST Reco sheet
  2. Books All Sales vs GSTR-1 All monthly → Section 2
  3. Books B2B vs GSTR-1 B2B monthly       → Section 3
  4. Books B2C vs GSTR-1 B2C monthly       → Section 4
  5. B2B Reco  — invoice-level, all Tally columns pass-through + Remark
  6. B2C Reco  — State + Rate annual aggregation

Input files:
  tally_sales  (required)  Tally Sales Register .xlsx
  gstr1_octa   (required)  OCTA download with Final GSTR-1 + GSTR3B + GSTR2B sheets
  gstr1_pdf    (optional)  GST-portal GSTR-1 PDF  → enables Step 0
  credit_note  (optional)  Separate Credit Note Register .xlsx

Both input files are accepted in raw or pre-worked form:
  * Sales Register — a hand-worked register carries `Total Sales / Total IGST /
    Total CGST / Total SGST` (plus Month, Catogary, States). A raw Tally export
    carries ledger-wise columns instead, and a trailing "Grand Total" row; the
    totals are derived from the ledgers and appended so they stay visible in the
    output, the footer row is dropped, and category/state/rate are inferred.
  * OCTA GSTR-1 — some exports fold B2CS rows into the GSTR-1 sheet, others split
    them into a separate summary sheet; both are read.
"""

from __future__ import annotations

import logging
import re
from collections import Counter, defaultdict
from datetime import datetime
from io import BytesIO

import pandas as pd
from openpyxl import load_workbook

from .core import normalize_doc_no, round_money
from .parsers import normalize_header

logger = logging.getLogger(__name__)

# ---------------------------------------------------------------------------
# Constants
# ---------------------------------------------------------------------------

_FY_MONTHS = [
    "April", "May", "June", "July", "August", "September",
    "October", "November", "December", "January", "February", "March",
]

_SHORT_TO_FULL = {
    "apr": "April", "may": "May", "jun": "June", "jul": "July",
    "aug": "August", "sep": "September", "oct": "October",
    "nov": "November", "dec": "December", "jan": "January",
    "feb": "February", "mar": "March",
}

# Keywords that flag an Amazon consolidated Tally ledger row
_AMAZON_TALLY_KEYWORDS = (
    "amazon b2b", "amazon inter", "amazon intra",
    "amazon seller", "amazon b2c", "amazon ecommerce",
)

# Regex to detect Amazon consolidated invoice numbers (AMZ-INTER-04, AMZ-INTRA-01, etc.)
_AMZ_INV_RE = re.compile(r"^AMZ[-_](INTER|INTRA)", re.IGNORECASE)

# Standard GST state codes (used to line up Books states with GSTR-1 Place of Supply)
_GST_STATE_CODES = {
    "01": "Jammu And Kashmir", "02": "Himachal Pradesh", "03": "Punjab",
    "04": "Chandigarh", "05": "Uttarakhand", "06": "Haryana", "07": "Delhi",
    "08": "Rajasthan", "09": "Uttar Pradesh", "10": "Bihar", "11": "Sikkim",
    "12": "Arunachal Pradesh", "13": "Nagaland", "14": "Manipur", "15": "Mizoram",
    "16": "Tripura", "17": "Meghalaya", "18": "Assam", "19": "West Bengal",
    "20": "Jharkhand", "21": "Odisha", "22": "Chhattisgarh", "23": "Madhya Pradesh",
    "24": "Gujarat", "25": "Daman And Diu",
    "26": "Dadra And Nagar Haveli And Daman And Diu", "27": "Maharashtra",
    "28": "Andhra Pradesh", "29": "Karnataka", "30": "Goa", "31": "Lakshadweep",
    "32": "Kerala", "33": "Tamil Nadu", "34": "Puducherry",
    "35": "Andaman And Nicobar Islands", "36": "Telangana",
    "37": "Andhra Pradesh", "38": "Ladakh", "97": "Other Territory",
}

# Standard notified GST rates — used to snap a derived rate onto a real slab
_STD_GST_RATES = (0.0, 0.1, 0.25, 1.0, 1.5, 3.0, 5.0, 6.0, 7.5, 12.0, 18.0, 28.0)

# "B2B" / "B2C" appearing as a standalone token in a Tally ledger/party name
_B2X_TOKEN_RE = re.compile(r"(?<![a-z0-9])b2([bc])(?![a-z0-9])", re.IGNORECASE)

# Tally Sales Register headers that are never a sales/tax ledger amount
_TALLY_NON_LEDGER_HEADERS = {
    "date", "month", "catogary", "category", "particulars", "buyer", "party",
    "party name", "ledger name", "states", "state", "voucher type", "voucher no",
    "voucher ref no", "ref no", "narration", "quantity", "qty", "value",
    "gross total", "round off", "place of supply", "rate", "gst rate", "tax rate",
    "hsn", "hsn code", "remark", "gstin", "gstin/uin", "buyer gstin", "gst no",
    "invoice no", "doc no", "bill no", "invoice date", "voucher date",
    "total sales", "total igst", "total cgst", "total sgst", "total cess",
}

# Footer / summary rows Tally appends to an export — never real transactions
_TALLY_TOTAL_LABELS = {
    "grand total", "total", "sub total", "subtotal", "sub-total",
    "opening balance", "closing balance", "carried over", "brought forward",
}

# How many amount-agreeing invoices it takes before a books-series → GSTR-1-series
# renaming is trusted enough to match on the trailing serial alone
_SERIES_MAP_MIN_HITS = 3

# Derived Books total columns, in the exact names an accountant hand-adds
_TALLY_DERIVED_COLS = {
    "taxable": "Total Sales",
    "igst":    "Total IGST",
    "cgst":    "Total CGST",
    "sgst":    "Total SGST",
}


# ---------------------------------------------------------------------------
# Month helpers
# ---------------------------------------------------------------------------

def _norm_month(val) -> str:
    """Normalise any month representation → full name like 'April'."""
    if val is None:
        return "Unknown"
    # Handle pandas NaT
    try:
        import pandas as _pd
        if _pd.isnull(val):
            return "Unknown"
    except Exception:
        pass
    if isinstance(val, datetime):
        try:
            return val.strftime("%B")
        except Exception:
            return "Unknown"
    text = str(val).strip()
    if not text or text.lower() in ("nat", "nan", "none", ""):
        return "Unknown"
    # Datetime-like string
    try:
        return datetime.fromisoformat(text[:10]).strftime("%B")
    except Exception:
        pass
    # Short abbreviation: "Apr 2025" or just "Apr"
    short = text[:3].lower()
    return _SHORT_TO_FULL.get(short, text.title())


def _month_sort_key(m: str) -> int:
    try:
        return _FY_MONTHS.index(m)
    except ValueError:
        return 99


def _sorted_months(months: set) -> list[str]:
    return sorted(months, key=_month_sort_key)


def _zero_amounts() -> dict:
    return {"taxable": 0.0, "igst": 0.0, "cgst": 0.0, "sgst": 0.0}


def _add_amounts(acc: dict, taxable=0, igst=0, cgst=0, sgst=0, sign=1) -> None:
    acc["taxable"] += sign * _f(taxable)
    acc["igst"]    += sign * _f(igst)
    acc["cgst"]    += sign * _f(cgst)
    acc["sgst"]    += sign * _f(sgst)


def _f(v) -> float:
    """Safe float conversion, treating NaN/None as 0."""
    try:
        x = float(v)
        return 0.0 if (x != x) else x  # NaN check
    except (TypeError, ValueError):
        return 0.0


def _norm_inv(v) -> str:
    """Normalise invoice number for matching — strip non-alphanumeric, uppercase."""
    return re.sub(r"[^A-Z0-9]", "", str(v or "").upper())


def _inv_serial(v) -> str:
    """
    Trailing running number of a document number — 'LI/INV/25-26/39' → '39'.

    Tally's voucher series and the series actually filed in GSTR-1 are often
    different prefixes over the same running number ('LI/SV/25-26/39' in the books
    vs 'LI/INV/25-26/39' on the portal). The serial is the only part the two share,
    so it is kept as a *fallback* key for invoice matching.

    Leading zeros are preserved so a credit-note serial ('0001') can never collide
    with an invoice serial ('1').
    """
    m = re.search(r"(\d+)\s*$", str(v or "").strip())
    return m.group(1) if m else ""


def _inv_prefix(v) -> str:
    """Everything before the trailing serial — 'LI/INV/25-26/39' → 'LIINV2526'."""
    norm, serial = _norm_inv(v), _inv_serial(v)
    return norm[:-len(serial)] if serial and norm.endswith(serial) else norm


def _classify_tally_ledger(header) -> str | None:
    """Classify a Sales Register column → 'taxable' | 'igst' | 'cgst' | 'sgst' | None."""
    n = normalize_header(header)
    if not n or n in _TALLY_NON_LEDGER_HEADERS:
        return None
    if re.fullmatch(r"col_?\d+", n):                 # unnamed placeholder column
        return None
    if "inward" in n or n.startswith("input"):       # purchase-side ledger parked in a sales register
        return None
    if "cess" in n:                                  # not aggregated by this agent
        return None
    if "igst" in n:
        return "igst"
    if "cgst" in n:
        return "cgst"
    if "sgst" in n:
        return "sgst"
    return "taxable"


def _drop_tally_total_rows(df: pd.DataFrame) -> pd.DataFrame:
    """
    Remove Tally's own footer rows ("Grand Total", "Closing Balance", …).

    A raw export ends with a Grand Total line carrying the sum of every ledger
    column. Left in, it doubles the Books turnover. A hand-worked register has
    already had it deleted, so this is a no-op there.
    """
    if df.empty:
        return df

    label_col = _find_col(df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
    if label_col is None:
        return df

    labels = df[label_col].map(lambda v: re.sub(r"\s+", " ", str(v or "")).strip().lower())
    is_total = labels.isin(_TALLY_TOTAL_LABELS)
    if not is_total.any():
        return df

    logger.info("Tally: dropped %d footer/total row(s): %s", int(is_total.sum()),
                ", ".join(sorted(set(labels[is_total]))))
    return df[~is_total].reset_index(drop=True)


def _derive_tally_totals(df: pd.DataFrame) -> pd.DataFrame:
    """
    Add the `Total Sales / Total IGST / Total CGST / Total SGST` columns when the
    register does not already carry them.

    A hand-worked register has these four columns; a raw Tally export instead has
    ledger-wise columns (`Sales @18%`, `Interstate Sales @5%`, `OUTPUT IGST@18%`,
    `Amazon B2C Sales`, …). Without them every Books amount reads as 0 and every
    invoice reconciles as a difference. The derived columns are appended at the end
    so they stay visible in the output Sales Register / B2B Reco sheets.

    Columns that already exist are never touched — a pre-worked register is
    returned byte-for-byte unchanged.
    """
    if df.empty:
        return df

    existing = {normalize_header(c) for c in df.columns}
    missing = {kind: name for kind, name in _TALLY_DERIVED_COLS.items()
               if normalize_header(name) not in existing}
    if not missing:
        return df

    positions: dict[str, list[int]] = {k: [] for k in _TALLY_DERIVED_COLS}
    for i, col in enumerate(df.columns):
        kind = _classify_tally_ledger(col)
        if kind:
            positions[kind].append(i)

    if not any(positions.values()):
        logger.warning("Tally: no ledger columns recognised — Books totals stay 0")
        return df

    for kind, name in missing.items():
        cols = positions[kind]
        if cols:
            columns = [[_f(v) for v in df.iloc[:, i].tolist()] for i in cols]
            df[name] = [round_money(sum(vals)) for vals in zip(*columns)]
        else:
            df[name] = 0.0
        logger.info("Tally: derived %r from %d ledger column(s): %s", name, len(cols),
                    ", ".join(str(df.columns[i]) for i in cols[:6]) or "none")
    return df


def _category_from_text(text) -> str | None:
    """Read an explicit B2B / B2C marker out of a Tally ledger or party name."""
    m = _B2X_TOKEN_RE.search(str(text or ""))
    if not m:
        return None
    return "B2B" if m.group(1).lower() == "b" else "B2C"


def _state_from_row(row, state_col, gstin_col, part_col) -> str:
    """
    Resolve a Books row to a state name comparable with GSTR-1 Place of Supply.

    Order: explicit States column → buyer GSTIN state code → a state code or name
    embedded in the ledger name (e.g. "B2C Interstate-01-J&K", "BLINKIT B2C-GUJARAT").
    """
    if state_col:
        raw = str(_col_val(row, state_col, "")).strip()
        if raw:
            return _extract_pos_state(raw)

    if gstin_col:
        gstin = re.sub(r"[^A-Z0-9]", "", str(_col_val(row, gstin_col, "")).upper())
        if len(gstin) == 15 and gstin[:2] in _GST_STATE_CODES:
            return _GST_STATE_CODES[gstin[:2]]

    if part_col:
        text = str(_col_val(row, part_col, ""))
        m = re.search(r"(?<!\d)(\d{2})(?!\d)", text)
        if m and m.group(1) in _GST_STATE_CODES:
            return _GST_STATE_CODES[m.group(1)]
        upper = text.upper()
        for name in _GST_STATE_CODES.values():
            if name.upper() in upper:
                return name

    return "Unknown"


def _derive_rate(taxable, igst, cgst, sgst) -> float:
    """Back out the GST rate from the amounts when the register has no Rate column."""
    base = _f(taxable)
    if not base:
        return 0.0
    pct = abs((_f(igst) + _f(cgst) + _f(sgst)) / base) * 100.0
    nearest = min(_STD_GST_RATES, key=lambda r: abs(r - pct))
    return nearest if abs(nearest - pct) <= 0.6 else round(pct, 2)


def _is_amazon_tally_row(particulars: str, inv_no: str) -> bool:
    p = particulars.lower()
    if any(kw in p for kw in _AMAZON_TALLY_KEYWORDS):
        return True
    if _AMZ_INV_RE.match(str(inv_no or "").strip()):
        return True
    return False


# ---------------------------------------------------------------------------
# OCTA Excel reader
# ---------------------------------------------------------------------------

def read_octa_excel(file_info: dict) -> tuple[pd.DataFrame, pd.DataFrame, pd.DataFrame]:
    """Auto-detect and return (gstr1_df, gstr3b_df, gstr2b_df) from the OCTA Excel."""
    data = file_info["content"]
    wb = load_workbook(BytesIO(data), data_only=True, read_only=True)

    gstr1_df = pd.DataFrame()
    gstr3b_df = pd.DataFrame()
    gstr2b_df = pd.DataFrame()
    b2c_summary_df = pd.DataFrame()

    for sheet in wb.worksheets:
        rows = list(sheet.iter_rows(values_only=True))
        if not rows:
            continue

        # Scan up to first 15 rows to find the real header row
        # (OCTA sheets often have summary numbers in rows 1-4 before the header)
        header_row = None
        for i, r in enumerate(rows[:15]):
            nh = [normalize_header(v) for v in r if v]
            # GSTR-1 header: "doc no" + "customer gstin"
            if gstr1_df.empty and _has_cols(nh, ["doc no", "customer gstin"]):
                header_row = i
                break
            # B2C summary header: "summary type" + "taxable value", no invoice-level "doc no"
            if b2c_summary_df.empty and _is_b2c_summary_header(nh):
                header_row = i
                break
            # GSTR-2B header: "supplier gstin" + ("itc eligible" or "gstr1 filing period")
            if gstr2b_df.empty and _has_cols(nh, ["supplier gstin"]) and any("itc" in h or "filing" in h for h in nh):
                header_row = i
                break
            # GSTR-3B header: has "section" column
            if gstr3b_df.empty and any("section" in h for h in nh) and any("apr" in h or "may" in h or "apr 2" in h for h in nh):
                header_row = i
                break

        if header_row is None:
            continue

        headers = [str(v or "").strip() for v in rows[header_row]]
        norm_headers = [normalize_header(h) for h in headers]

        # Detect Final GSTR-1: has "Doc No" and "Customer GSTIN"
        if gstr1_df.empty and _has_cols(norm_headers, ["doc no", "customer gstin"]):
            data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                         if any(v is not None and str(v).strip() for v in r)]
            gstr1_df = pd.DataFrame(data_rows)
            logger.info("OCTA GSTR-1 sheet detected: '%s' (%d rows)", sheet.title, len(gstr1_df))
            continue

        # Detect a separate B2C summary sheet: "Summary Type" + "Taxable Value", no "Doc No".
        # Some OCTA exports fold B2CS rows into the GSTR-1 sheet; others split them out here.
        if b2c_summary_df.empty and _is_b2c_summary_header(norm_headers):
            data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                         if any(v is not None and str(v).strip() for v in r)]
            b2c_summary_df = _prepare_b2c_summary(pd.DataFrame(data_rows), sheet.title)
            continue

        # Detect GSTR3B: has "Section" column and month columns
        if gstr3b_df.empty and any("section" in h for h in norm_headers):
            all_text = " ".join(str(v) for row in rows for v in row if v)
            if "3.1.a" in all_text.lower():
                data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                             if any(v is not None and str(v).strip() for v in r)]
                gstr3b_df = pd.DataFrame(data_rows)
                logger.info("OCTA GSTR-3B sheet detected: '%s' (%d rows)", sheet.title, len(gstr3b_df))
                continue

        # Detect GSTR2B: has "Supplier GSTIN" + ITC columns
        if gstr2b_df.empty and _has_cols(norm_headers, ["supplier gstin"]) and any("itc" in h or "filing" in h for h in norm_headers):
            data_rows = [dict(zip(headers, r)) for r in rows[header_row + 1:]
                         if any(v is not None and str(v).strip() for v in r)]
            gstr2b_df = pd.DataFrame(data_rows)
            logger.info("OCTA GSTR-2B sheet detected: '%s' (%d rows)", sheet.title, len(gstr2b_df))

    # Fold a split-out B2C summary sheet back into GSTR-1 so the B2C reco has data
    if not b2c_summary_df.empty:
        gstr1_df = (pd.concat([gstr1_df, b2c_summary_df], ignore_index=True)
                    if not gstr1_df.empty else b2c_summary_df)

    return gstr1_df, gstr3b_df, gstr2b_df


def _prepare_b2c_summary(df: pd.DataFrame, sheet_title: str) -> pd.DataFrame:
    """
    Keep the B2C rows of an OCTA summary sheet and map them onto the GSTR-1 schema.

    Anything whose Summary Type mentions B2C counts (B2CS, "B2CS Sales",
    amendments, …). Other summary types — TCS/e-commerce Section 52 disclosures,
    exempt/nil schedules — are left out: they restate sales already reported
    elsewhere and would double-count.
    """
    if df.empty:
        return df

    type_col = _find_col(df, ["Summary Type", "Type", "Doc Type"])
    if type_col is None:
        return pd.DataFrame()

    is_b2c = df[type_col].astype(str).str.contains("b2c", case=False, na=False)
    kept, dropped = df[is_b2c].copy(), df[~is_b2c]

    if not dropped.empty:
        logger.info("OCTA B2C summary '%s': skipped %d non-B2C row(s) — %s",
                    sheet_title, len(dropped),
                    ", ".join(sorted(dropped[type_col].astype(str).unique())[:6]))
    if kept.empty:
        return pd.DataFrame()

    renames = {type_col: "Doc Type"}
    taxable_col = _find_col(kept, ["Taxable Value", "Item Taxable Value"])
    if taxable_col and normalize_header(taxable_col) != "item taxable value":
        renames[taxable_col] = "Item Taxable Value"
    kept = kept.rename(columns=renames)

    # B2C rows carry no counterparty GSTIN — make that explicit for the B2B/B2C split
    if _find_col(kept, ["Customer GSTIN", "GSTIN of Recipient"]) is None:
        kept["Customer GSTIN"] = ""

    logger.info("OCTA B2C summary '%s': merged %d B2C row(s) into GSTR-1",
                sheet_title, len(kept))
    return kept


def _has_cols(norm_headers: list[str], required: list[str]) -> bool:
    joined = " ".join(norm_headers)
    return all(req in joined for req in required)


def _is_b2c_summary_header(norm_headers: list[str]) -> bool:
    """
    True for a rate-wise B2CS summary sheet, false for the HSN summary.

    A portal export carries both, and they look alike — 'Summary Type' + 'Taxable
    Value' with no 'Doc No'. But the HSN sheet restates *every* supply, B2B and
    B2C alike, so folding it into GSTR-1 would roughly double the turnover. The
    HSN column is what tells them apart.
    """
    if not _has_cols(norm_headers, ["summary type", "taxable value"]):
        return False
    if _has_cols(norm_headers, ["doc no"]):
        return False
    return not any("hsn" in h for h in norm_headers)


# ---------------------------------------------------------------------------
# Tally Sales Register reader (raw — preserves all columns)
# ---------------------------------------------------------------------------

def read_tally_sales_raw(file_info: dict) -> pd.DataFrame:
    """Read Tally Sales Register, auto-detect header row, return full DataFrame."""
    data = file_info["content"]
    filename = file_info.get("filename", "")
    suffix = filename.lower().rsplit(".", 1)[-1] if "." in filename else ""

    if suffix in {"xlsx", "xlsm", "xls"}:
        wb = load_workbook(BytesIO(data), data_only=True, read_only=True)
        # Use the sheet that looks most like a sales register
        best_sheet = None
        best_score = 0
        for sheet in wb.worksheets:
            rows = list(sheet.iter_rows(values_only=True))
            for i, row in enumerate(rows[:15]):
                nh = [normalize_header(v) for v in row if v]
                score = sum(1 for t in ["date", "voucher no", "total sales", "particulars", "category"] if any(t in h for h in nh))
                if score > best_score:
                    best_score = score
                    best_sheet = (sheet, rows, i)
        if best_sheet is None:
            raise ValueError("No recognisable Sales Register sheet found in Tally file.")
        sheet, rows, hdr_idx = best_sheet
        headers = [str(v or f"Col_{i}").strip() for i, v in enumerate(rows[hdr_idx])]
        data_rows = []
        for row in rows[hdr_idx + 1:]:
            if not any(v is not None and str(v).strip() for v in row):
                continue
            data_rows.append({headers[i]: (row[i] if i < len(row) else None) for i in range(len(headers))})
        df = pd.DataFrame(data_rows)
    else:
        df = pd.read_csv(BytesIO(data), encoding="utf-8-sig")

    df = _derive_tally_totals(_drop_tally_total_rows(df))
    logger.info("Tally Sales Register: %d rows, %d cols", len(df), len(df.columns))
    return df


def read_credit_note_raw(file_info: dict) -> pd.DataFrame:
    """Read Credit Note Register as raw DataFrame (same structure as Tally)."""
    return read_tally_sales_raw(file_info)


# Amount columns to sign-flip alongside the ledger columns when folding in credit notes
_CN_FLIP_EXTRA = {"gross total", "value"}


def merge_credit_notes(tally_df: pd.DataFrame, cn_df: pd.DataFrame | None) -> pd.DataFrame:
    """
    Fold a separate Credit Note Register into the Sales Register for reconciliation.

    Tally exports credit notes two ways:
      * inside the Sales Register, already negative (`Credit Note` voucher types), or
      * as a standalone Credit Note Register, where the amounts are positive.

    GSTR-1 always reports credit notes as negatives, so a standalone register has to
    be sign-flipped before it can net off sales — otherwise Books stays overstated by
    the full credit-note value and every GSTR-1 credit note reads "Not in Books".

    The direction is detected from the data, not assumed, and rows already present in
    the Sales Register are skipped so nothing is counted twice.
    """
    if cn_df is None or cn_df.empty or tally_df.empty:
        return tally_df

    inv_col_t = _find_col(tally_df, ["Voucher No.", "Voucher No", "Invoice No", "Doc No", "Bill No"])
    inv_col_c = _find_col(cn_df,    ["Voucher No.", "Voucher No", "Invoice No", "Doc No", "Bill No"])
    vt_col    = _find_col(tally_df, ["Voucher Type"])

    # Credit notes the Sales Register already carries — used to avoid double counting
    embedded_invs: set[str] = set()
    if vt_col:
        for _, row in tally_df.iterrows():
            if "credit note" in str(_col_val(row, vt_col, "")).lower():
                norm = _norm_inv(_col_val(row, inv_col_t, "")) if inv_col_t else ""
                if norm:
                    embedded_invs.add(norm)
    if embedded_invs:
        logger.info("Credit notes: Sales Register already holds %d credit note(s)", len(embedded_invs))

    # Detect the register's sign convention from its own amounts
    ts_col = _find_col(cn_df, ["Total Sales", "Taxable Value", "Taxable Amount"])
    values = [_f(v) for v in cn_df[ts_col]] if ts_col is not None else []
    positives = sum(1 for v in values if v > 0)
    negatives = sum(1 for v in values if v < 0)
    sign = -1.0 if positives >= negatives else 1.0

    amount_keys = {normalize_header(n) for n in _TALLY_DERIVED_COLS.values()} | _CN_FLIP_EXTRA
    cn_cols = {normalize_header(c): c for c in cn_df.columns}

    merged_rows: list[dict] = []
    skipped = 0
    for _, row in cn_df.iterrows():
        norm = _norm_inv(_col_val(row, inv_col_c, "")) if inv_col_c else ""
        if norm and norm in embedded_invs:
            skipped += 1
            continue
        out: dict = {}
        for col in tally_df.columns:
            key = normalize_header(col)
            src = cn_cols.get(key)
            if src is None:
                out[col] = None
            elif key in amount_keys or _classify_tally_ledger(col):
                out[col] = round_money(sign * _f(row[src]))
            else:
                out[col] = row[src]
        merged_rows.append(out)

    if not merged_rows:
        logger.info("Credit notes: nothing to merge (%d already in the Sales Register)", skipped)
        return tally_df

    logger.info("Credit notes: merged %d row(s) with sign %+d%s",
                len(merged_rows), int(sign),
                f", skipped {skipped} already in the Sales Register" if skipped else "")
    return pd.concat([tally_df, pd.DataFrame(merged_rows)], ignore_index=True)


# ---------------------------------------------------------------------------
# Column detection helpers for DataFrames
# ---------------------------------------------------------------------------

def _find_col(df: pd.DataFrame, candidates: list[str]) -> str | None:
    """Return the first DataFrame column that matches any candidate (normalized)."""
    norm_cols = {normalize_header(c): c for c in df.columns}
    for cand in candidates:
        key = normalize_header(cand)
        if key in norm_cols:
            return norm_cols[key]
    return None


def _col_val(row, col: str | None, default=None):
    if col is None or col not in row.index:
        return default
    v = row[col]
    if v is None or (isinstance(v, float) and v != v):
        return default
    return v


# ---------------------------------------------------------------------------
# GSTR-3B monthly totals extractor
# ---------------------------------------------------------------------------

def extract_gstr3b_monthly(gstr3b_df: pd.DataFrame) -> dict[str, dict]:
    """
    Extract monthly taxable/IGST/CGST/SGST from GSTR-3B DataFrame.
    Sums 3.1.A + 3.1.B (zero-rated) for taxable and IGST.
    Returns {month_name: {taxable, igst, cgst, sgst}}.
    """
    if gstr3b_df.empty:
        return {}

    # Identify the "Section" and "Type" columns
    section_col = _find_col(gstr3b_df, ["Section", "Section Name"])
    type_col    = _find_col(gstr3b_df, ["Type", "Row Type"])

    if section_col is None:
        logger.warning("GSTR-3B: no Section column found")
        return {}

    # Month columns: everything after "Type" that looks like a month
    skip_cols = {section_col, type_col} if type_col else {section_col}
    month_cols: dict[str, str] = {}  # month_name → column_name
    for col in gstr3b_df.columns:
        if col in skip_cols:
            continue
        m = _norm_month(col)
        if m in _FY_MONTHS:
            month_cols[m] = col

    if not month_cols:
        logger.warning("GSTR-3B: no month columns detected")
        return {}

    result: dict[str, dict] = {m: _zero_amounts() for m in month_cols}

    for _, row in gstr3b_df.iterrows():
        section = str(_col_val(row, section_col, "")).strip()
        rtype   = str(_col_val(row, type_col, "")).strip().lower() if type_col else ""

        is_3_1a = "3.1.a" in section.lower() or (section.startswith("3.1.A") if section else False)
        is_3_1b = "3.1.b" in section.lower() or (section.startswith("3.1.B") if section else False)

        if not (is_3_1a or is_3_1b):
            continue

        is_supply = "supply value" in rtype or "supply" in rtype
        is_igst   = "integrated" in rtype or "igst" in rtype
        is_cgst   = "central" in rtype or "cgst" in rtype
        is_sgst   = ("state" in rtype and "supply" not in rtype) or "sgst" in rtype

        for month, col in month_cols.items():
            val = _f(_col_val(row, col, 0))
            if val == 0:
                continue
            if is_supply:
                result[month]["taxable"] += val
            elif is_igst:
                result[month]["igst"] += val
            elif is_cgst:
                result[month]["cgst"] += val
            elif is_sgst:
                result[month]["sgst"] += val

    return result


# ---------------------------------------------------------------------------
# GSTR-1 monthly aggregator (from OCTA Final GSTR-1)
# ---------------------------------------------------------------------------

def gstr1_month_basis(gstr1_df: pd.DataFrame):
    """
    Return `(month_of_row, basis_name)` — how a GSTR-1 row is placed in a month.

    Tax Period is the filing period, and for a **monthly** filer it is also the
    month the invoice belongs to, so it stays the default and monthly filers keep
    their existing numbers exactly.

    A **quarterly (QRMP/IFF)** filer breaks that: the whole quarter is filed under
    one period, so months the business actually sold in report zero against Books
    and the monthly sections read as pure difference. When that pattern is present
    — some month carries invoice-dated turnover but no filed turnover at all — the
    invoice's own Doc Date is used instead, which is what Books is keyed on.

    Rows with no Doc Date (B2CS summary rows carry only a period) always fall back
    to Tax Period, so nothing is dropped.
    """
    period_col = _find_col(gstr1_df, ["Tax Period", "Tax period", "Period", "Return Period"])
    date_col   = _find_col(gstr1_df, ["Doc Date", "Invoice Date", "Document Date"])

    def by_period(row) -> str:
        return _norm_month(_col_val(row, period_col)) if period_col else "Unknown"

    if period_col is None or date_col is None:
        return by_period, "Tax Period"

    taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value", "Taxable Amount", "Taxable"])
    igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax", "IGST Amount"])
    cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax", "CGST Amount"])
    sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax", "SGST Amount"])

    filed: set[str] = set()
    dated: set[str] = set()
    for _, row in gstr1_df.iterrows():
        amount = (_f(_col_val(row, taxable_col, 0)) + _f(_col_val(row, igst_col, 0))
                  + _f(_col_val(row, cgst_col, 0)) + _f(_col_val(row, sgst_col, 0)))
        if amount == 0:
            continue
        mp = _norm_month(_col_val(row, period_col))
        md = _norm_month(_col_val(row, date_col))
        if mp in _FY_MONTHS:
            filed.add(mp)
        if md in _FY_MONTHS:
            dated.add(md)

    gaps = sorted(dated - filed, key=_month_sort_key)
    if not gaps:
        return by_period, "Tax Period"

    def by_doc_date(row) -> str:
        month = _norm_month(_col_val(row, date_col))
        return month if month in _FY_MONTHS else by_period(row)

    logger.info("GSTR-1: filed for %d month(s), invoiced across %d — quarterly/IFF filing "
                "detected (no filing in %s); placing rows by Doc Date instead of Tax Period",
                len(filed), len(dated), ", ".join(gaps))
    return by_doc_date, "Doc Date"


def aggregate_gstr1_monthly(gstr1_df: pd.DataFrame) -> dict[str, dict]:
    """Aggregate Final GSTR-1 sheet by month (see `gstr1_month_basis`)."""
    if gstr1_df.empty:
        return {}

    period_col  = _find_col(gstr1_df, ["Tax Period", "Tax period", "Period", "Return Period"])
    taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value", "Taxable Amount", "Taxable"])
    igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax", "IGST Amount"])
    cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax", "CGST Amount"])
    sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax", "SGST Amount"])

    if period_col is None:
        logger.warning("GSTR-1: Tax Period column not found")
        return {}

    month_of, _basis = gstr1_month_basis(gstr1_df)

    result: dict[str, dict] = defaultdict(_zero_amounts)
    for _, row in gstr1_df.iterrows():
        month = month_of(row)
        if month not in _FY_MONTHS:
            continue
        _add_amounts(
            result[month],
            taxable=_col_val(row, taxable_col, 0),
            igst=_col_val(row, igst_col, 0),
            cgst=_col_val(row, cgst_col, 0),
            sgst=_col_val(row, sgst_col, 0),
        )
    return dict(result)


# ---------------------------------------------------------------------------
# Books monthly aggregator (from Tally Sales Register)
# ---------------------------------------------------------------------------

def aggregate_books_monthly(tally_df: pd.DataFrame, category: str | None = None,
                            g1_b2b_invs: set | None = None) -> dict[str, dict]:
    """
    Aggregate Tally by month.
    category: None=all rows, 'B2B'=only B2B rows, 'B2C'=only B2C rows.
    """
    if tally_df.empty:
        return {}

    inv_col     = _find_col(tally_df, ["Voucher No.", "Voucher No", "Invoice No", "Doc No", "Bill No"])
    ref_col     = _find_col(tally_df, ["Voucher Ref. No.", "Voucher Ref No", "Ref No"])
    inv_cols    = tuple(c for c in (inv_col, ref_col) if c)
    month_col   = _find_col(tally_df, ["Month", "Mth"])
    date_col    = _find_col(tally_df, ["Date", "Invoice Date", "Voucher Date"])
    cat_col     = _find_col(tally_df, ["Category", "Catogary", "Cat", "Type"])
    part_col    = _find_col(tally_df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN", "Buyer GSTIN", "GST No"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value", "Taxable Amount", "Net Amount"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST", "Integrated Tax"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST", "Central Tax"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST", "State Tax"])

    result: dict[str, dict] = defaultdict(_zero_amounts)

    for _, row in tally_df.iterrows():
        # Determine month
        month = _norm_month(_col_val(row, month_col) or _col_val(row, date_col))
        if month not in _FY_MONTHS:
            continue

        # Skip header-like rows
        taxable = _f(_col_val(row, taxable_col, 0))
        igst_v  = _f(_col_val(row, igst_col, 0))
        cgst_v  = _f(_col_val(row, cgst_col, 0))
        sgst_v  = _f(_col_val(row, sgst_col, 0))
        if taxable == 0 and igst_v == 0 and cgst_v == 0 and sgst_v == 0:
            continue

        # Category filter
        if category is not None:
            row_cat = _infer_category(row, cat_col, gstin_col, part_col, inv_cols, g1_b2b_invs)
            if row_cat != category:
                continue

        _add_amounts(result[month], taxable=taxable, igst=igst_v, cgst=cgst_v, sgst=sgst_v)

    return dict(result)


def _gstr1_b2b_invoice_set(gstr1_df: pd.DataFrame) -> set[str]:
    """Normalised invoice numbers that GSTR-1 reported against a customer GSTIN."""
    if gstr1_df is None or gstr1_df.empty:
        return set()

    inv_col   = _find_col(gstr1_df, ["Doc No", "Invoice Number", "Invoice No"])
    gstin_col = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient", "Buyer GSTIN"])
    if inv_col is None:
        return set()

    invoices: set[str] = set()
    for _, row in gstr1_df.iterrows():
        gstin = str(_col_val(row, gstin_col, "")).strip()
        if not gstin or gstin.upper() in ("N/A", "NA"):
            continue
        norm = _norm_inv(_col_val(row, inv_col, ""))
        if norm:
            invoices.add(norm)
    return invoices


def _infer_category(row, cat_col: str | None, gstin_col: str | None,
                    part_col: str | None = None,
                    inv_cols: tuple = (), g1_b2b_invs: set | None = None) -> str:
    """Return 'B2B' or 'B2C' for a Tally row."""
    if cat_col:
        cat = str(_col_val(row, cat_col, "")).strip().upper()
        if "B2B" in cat:
            return "B2B"
        if "B2C" in cat:
            return "B2C"
    # GSTR-1 is the filed authority. An invoice reported there against a customer
    # GSTIN is B2B in the books too, even when the Tally row carries no GSTIN —
    # common for marketplace party ledgers (KiranaKart/Zepto, Blinkit, …).
    if g1_b2b_invs:
        for col in inv_cols:
            norm = _norm_inv(_col_val(row, col, ""))
            if norm and norm in g1_b2b_invs:
                return "B2B"
    # An explicit marker in the ledger name beats GSTIN presence: B2C bucket ledgers
    # ("B2C-(KAR TO KAR)") carry the seller's own GSTIN, and Amazon B2B consolidated
    # rows carry none at all.
    if part_col:
        cat = _category_from_text(_col_val(row, part_col, ""))
        if cat:
            return cat
    # Fall back to GSTIN presence
    if gstin_col:
        gstin = str(_col_val(row, gstin_col, "")).strip()
        if gstin and gstin.upper() not in ("#N/A", "N/A", "NA", "NONE", ""):
            if len(re.sub(r"[^A-Z0-9]", "", gstin.upper())) == 15:
                return "B2B"
    return "B2C"


# ---------------------------------------------------------------------------
# Monthly comparison table builder (for GST Reco sections)
# ---------------------------------------------------------------------------

def build_monthly_comparison(
    left: dict[str, dict],
    right: dict[str, dict],
    left_label: str,
    right_label: str,
) -> list[dict]:
    """
    Build list of monthly comparison rows.
    Each row: {month, left_taxable/igst/cgst/sgst, right_taxable/igst/cgst/sgst, diff_*}
    """
    all_months = set(left) | set(right)
    rows = []
    for month in _sorted_months(all_months):
        lv = left.get(month, _zero_amounts())
        rv = right.get(month, _zero_amounts())
        rows.append({
            "month": month,
            f"{left_label}_taxable": round_money(lv["taxable"]),
            f"{left_label}_igst":    round_money(lv["igst"]),
            f"{left_label}_cgst":    round_money(lv["cgst"]),
            f"{left_label}_sgst":    round_money(lv["sgst"]),
            f"{right_label}_taxable": round_money(rv["taxable"]),
            f"{right_label}_igst":    round_money(rv["igst"]),
            f"{right_label}_cgst":    round_money(rv["cgst"]),
            f"{right_label}_sgst":    round_money(rv["sgst"]),
            "diff_taxable": round_money(lv["taxable"] - rv["taxable"]),
            "diff_igst":    round_money(lv["igst"]    - rv["igst"]),
            "diff_cgst":    round_money(lv["cgst"]    - rv["cgst"]),
            "diff_sgst":    round_money(lv["sgst"]    - rv["sgst"]),
        })
    # Totals row
    if rows:
        tot = {
            "month": "Total",
            f"{left_label}_taxable":  sum(r[f"{left_label}_taxable"]  for r in rows),
            f"{left_label}_igst":     sum(r[f"{left_label}_igst"]     for r in rows),
            f"{left_label}_cgst":     sum(r[f"{left_label}_cgst"]     for r in rows),
            f"{left_label}_sgst":     sum(r[f"{left_label}_sgst"]     for r in rows),
            f"{right_label}_taxable": sum(r[f"{right_label}_taxable"] for r in rows),
            f"{right_label}_igst":    sum(r[f"{right_label}_igst"]    for r in rows),
            f"{right_label}_cgst":    sum(r[f"{right_label}_cgst"]    for r in rows),
            f"{right_label}_sgst":    sum(r[f"{right_label}_sgst"]    for r in rows),
            "diff_taxable": sum(r["diff_taxable"] for r in rows),
            "diff_igst":    sum(r["diff_igst"]    for r in rows),
            "diff_cgst":    sum(r["diff_cgst"]    for r in rows),
            "diff_sgst":    sum(r["diff_sgst"]    for r in rows),
        }
        rows.append(tot)
    return rows


# ---------------------------------------------------------------------------
# B2B Reconciliation (new — invoice-level, all Tally columns pass-through)
# ---------------------------------------------------------------------------

def reconcile_b2b_new(
    tally_df: pd.DataFrame,
    gstr1_df: pd.DataFrame,
    tolerance: float = 1.0,
) -> list[dict]:
    """
    Match Tally B2B rows against GSTR-1 by invoice number.
    Returns list of row dicts: all original Tally columns + GSTR-1 matched cols + Diff + Remark.
    Unmatched GSTR-1 rows are appended at the bottom.
    """
    if tally_df.empty:
        return []

    # Detect key Tally columns
    inv_col     = _find_col(tally_df, ["Voucher No.", "Voucher No", "Voucher Ref. No.", "Invoice No", "Doc No", "Bill No"])
    ref_col     = _find_col(tally_df, ["Voucher Ref. No.", "Voucher Ref No", "Ref No"])
    cat_col     = _find_col(tally_df, ["Category", "Catogary"])
    part_col    = _find_col(tally_df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN", "Buyer GSTIN"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value", "Taxable Amount"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST"])

    # Build GSTR-1 index by normalised invoice number → row dict
    g1_inv_col     = _find_col(gstr1_df, ["Doc No", "Invoice Number", "Invoice No"])
    g1_gstin_col   = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient", "Buyer GSTIN"])
    g1_taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value", "Taxable"])
    g1_igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax"])
    g1_cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax"])
    g1_sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax"])

    # Index GSTR-1 rows
    gstr1_by_inv: dict[str, list[int]] = defaultdict(list)
    # Secondary index on the trailing serial only — used when the two sides number
    # the same invoice under different series prefixes (see `_inv_serial`).
    gstr1_by_serial: dict[str, list[int]] = defaultdict(list)
    gstr1_rows = []
    if not gstr1_df.empty:
        for idx, row in gstr1_df.iterrows():
            # Skip B2C rows (no Customer GSTIN)
            gstin = str(_col_val(row, g1_gstin_col, "")).strip()
            if not gstin or gstin.upper() in ("", "N/A", "NA"):
                continue
            inv = str(_col_val(row, g1_inv_col, "")).strip()
            norm = _norm_inv(inv)
            g1_row = {
                "inv_no":  inv,
                "gstin":   gstin,
                "taxable": _f(_col_val(row, g1_taxable_col, 0)),
                "igst":    _f(_col_val(row, g1_igst_col, 0)),
                "cgst":    _f(_col_val(row, g1_cgst_col, 0)),
                "sgst":    _f(_col_val(row, g1_sgst_col, 0)),
                "_idx":    len(gstr1_rows),
            }
            gstr1_rows.append(g1_row)
            if norm:
                gstr1_by_inv[norm].append(len(gstr1_rows) - 1)
            serial = _inv_serial(inv)
            if serial:
                gstr1_by_serial[serial].append(len(gstr1_rows) - 1)

    def _serial_pool(books_gstin: str, doc_no) -> list[int]:
        """
        Unmatched GSTR-1 rows sharing this document's trailing serial.

        A bare serial is a weak key, so it only counts when it is unambiguous: the
        counterparty GSTIN has to agree, and when the books row carries no GSTIN the
        serial must point at exactly one unmatched GSTR-1 row.
        """
        serial = _inv_serial(doc_no)
        if not serial:
            return []
        pool = [j for j in gstr1_by_serial.get(serial, []) if j not in matched_gstr1]
        if not pool:
            return []
        key = re.sub(r"[^A-Z0-9]", "", str(books_gstin).upper())
        if len(key) == 15:
            return [j for j in pool
                    if re.sub(r"[^A-Z0-9]", "", gstr1_rows[j]["gstin"].upper()) == key]
        return pool if len(pool) == 1 else []

    g1_b2b_invs = _gstr1_b2b_invoice_set(gstr1_df)
    inv_cols = tuple(c for c in (inv_col, ref_col) if c)

    matched_gstr1: set[int] = set()
    result_rows: list[dict] = []
    # Rows the exact-number pass could not place, retried in a second pass below
    leftovers: list[tuple] = []

    for _, tally_row in tally_df.iterrows():
        # Get key fields
        inv_no   = str(_col_val(tally_row, inv_col, "")).strip()
        ref_no   = str(_col_val(tally_row, ref_col, "")).strip() if ref_col else ""
        part     = str(_col_val(tally_row, part_col, "")).strip()
        row_gstin = str(_col_val(tally_row, gstin_col, "")).strip() if gstin_col else ""
        taxable  = _f(_col_val(tally_row, taxable_col, 0))
        igst_v   = _f(_col_val(tally_row, igst_col, 0))
        cgst_v   = _f(_col_val(tally_row, cgst_col, 0))
        sgst_v   = _f(_col_val(tally_row, sgst_col, 0))

        # Skip empty / header-like rows
        if not inv_no and not part and taxable == 0:
            continue

        # Check category
        row_cat = _infer_category(tally_row, cat_col, gstin_col, part_col, inv_cols, g1_b2b_invs)
        if row_cat != "B2B":
            continue  # B2C rows handled separately

        # Build output row: all original Tally columns
        out = {col: _json_safe_val(tally_row[col]) for col in tally_df.columns}

        # Amazon detection
        effective_inv = inv_no or ref_no
        if _is_amazon_tally_row(part, effective_inv):
            # Try to find a GSTR-1 match anyway (for reference)
            norm = _norm_inv(effective_inv)
            candidates = [j for j in gstr1_by_inv.get(norm, []) if j not in matched_gstr1]
            if candidates:
                g1 = gstr1_rows[candidates[0]]
                out.update({
                    "_gstr1_inv_no": g1["inv_no"],
                    "_gstr1_gstin":  g1["gstin"],
                    "_gstr1_taxable": g1["taxable"],
                    "_gstr1_igst":   g1["igst"],
                    "_gstr1_cgst":   g1["cgst"],
                    "_gstr1_sgst":   g1["sgst"],
                    "_diff_taxable": round_money(taxable - g1["taxable"]),
                    "_diff_igst":    round_money(igst_v  - g1["igst"]),
                    "_diff_cgst":    round_money(cgst_v  - g1["cgst"]),
                    "_diff_sgst":    round_money(sgst_v  - g1["sgst"]),
                })
                matched_gstr1.add(candidates[0])
            else:
                _set_empty_gstr1(out)
            out["_remark"] = "Amazon Entry As per Tally"
            result_rows.append(out)
            continue

        # Regular B2B matching
        norm = _norm_inv(effective_inv)
        candidates = [j for j in gstr1_by_inv.get(norm, []) if j not in matched_gstr1]

        if candidates:
            # Pick best match by closest taxable value
            best_j = min(candidates, key=lambda j: abs(taxable - gstr1_rows[j]["taxable"]))
            g1 = gstr1_rows[best_j]
            dt = round_money(taxable - g1["taxable"])
            di = round_money(igst_v  - g1["igst"])
            dc = round_money(cgst_v  - g1["cgst"])
            ds = round_money(sgst_v  - g1["sgst"])
            is_diff = any(abs(d) > tolerance for d in (dt, di, dc, ds))
            out.update({
                "_gstr1_inv_no":  g1["inv_no"],
                "_gstr1_gstin":   g1["gstin"],
                "_gstr1_taxable": g1["taxable"],
                "_gstr1_igst":    g1["igst"],
                "_gstr1_cgst":    g1["cgst"],
                "_gstr1_sgst":    g1["sgst"],
                "_diff_taxable":  dt,
                "_diff_igst":     di,
                "_diff_cgst":     dc,
                "_diff_sgst":     ds,
                "_remark":        "Diff" if is_diff else "Match",
            })
            matched_gstr1.add(best_j)
        else:
            _set_empty_gstr1(out)
            out["_remark"] = "Not in GSTR-1"
            leftovers.append((len(result_rows), row_gstin, inv_no, ref_no,
                              taxable, igst_v, cgst_v, sgst_v))

        result_rows.append(out)

    # Second pass — only over rows the exact-number pass could not place.
    # Running it afterwards (rather than inline) means a weaker key can never claim
    # a GSTR-1 row that some later invoice matches exactly: every match the first
    # pass made stands, and this can only turn "Not in GSTR-1" into a match.
    # A serial on its own is not evidence of anything — 'BLR4-T-19' and 'DEL4-T-19'
    # are different branches' invoices. It only becomes trustworthy when the file
    # shows the two sides systematically renaming one series into another: count the
    # serial hits whose amounts also agree, per (books series → GSTR-1 series) pair,
    # and trust only the pairs that recur. A one-off coincidence never qualifies.
    series_hits: Counter = Counter()
    for _pos, row_gstin, inv_no, ref_no, taxable, *_ in leftovers:
        for doc_no in (inv_no, ref_no):
            for j in _serial_pool(row_gstin, doc_no):
                if abs(taxable - gstr1_rows[j]["taxable"]) <= tolerance:
                    series_hits[(_inv_prefix(doc_no), _inv_prefix(gstr1_rows[j]["inv_no"]))] += 1
    trusted_series = {pair for pair, hits in series_hits.items() if hits >= _SERIES_MAP_MIN_HITS}
    for (books_series, g1_series) in sorted(trusted_series):
        logger.info("B2B Reco: books series %r maps to GSTR-1 series %r (%d corroborating "
                    "invoice(s)) — serial matching enabled for it",
                    books_series, g1_series, series_hits[(books_series, g1_series)])

    def _serial_candidates(books_gstin: str, *doc_nos) -> list[int]:
        for doc_no in doc_nos:
            cands = [j for j in _serial_pool(books_gstin, doc_no)
                     if (_inv_prefix(doc_no), _inv_prefix(gstr1_rows[j]["inv_no"])) in trusted_series]
            if cands:
                return cands
        return []

    recovered = 0
    for pos, row_gstin, inv_no, ref_no, taxable, igst_v, cgst_v, sgst_v in leftovers:
        candidates = []
        if ref_no:
            candidates = [j for j in gstr1_by_inv.get(_norm_inv(ref_no), []) if j not in matched_gstr1]
        if not candidates:
            candidates = _serial_candidates(row_gstin, inv_no, ref_no)
        if not candidates:
            continue
        best_j = min(candidates, key=lambda j: abs(taxable - gstr1_rows[j]["taxable"]))
        g1 = gstr1_rows[best_j]
        dt = round_money(taxable - g1["taxable"])
        di = round_money(igst_v  - g1["igst"])
        dc = round_money(cgst_v  - g1["cgst"])
        ds = round_money(sgst_v  - g1["sgst"])
        result_rows[pos].update({
            "_gstr1_inv_no":  g1["inv_no"],
            "_gstr1_gstin":   g1["gstin"],
            "_gstr1_taxable": g1["taxable"],
            "_gstr1_igst":    g1["igst"],
            "_gstr1_cgst":    g1["cgst"],
            "_gstr1_sgst":    g1["sgst"],
            "_diff_taxable":  dt,
            "_diff_igst":     di,
            "_diff_cgst":     dc,
            "_diff_sgst":     ds,
            "_remark":        "Diff" if any(abs(d) > tolerance for d in (dt, di, dc, ds)) else "Match",
        })
        matched_gstr1.add(best_j)
        recovered += 1
    if recovered:
        logger.info("B2B Reco: %d row(s) matched on reference/serial number after the "
                    "exact-number pass (books and GSTR-1 use different voucher series)", recovered)

    # Append unmatched GSTR-1 rows at the bottom
    for j, g1 in enumerate(gstr1_rows):
        if j in matched_gstr1:
            continue
        out: dict = {col: None for col in tally_df.columns}
        # Fill in the GSTR-1 columns
        is_amazon = any(kw in g1["inv_no"].lower() for kw in ("amz", "amazon"))
        out.update({
            "_gstr1_inv_no":  g1["inv_no"],
            "_gstr1_gstin":   g1["gstin"],
            "_gstr1_taxable": g1["taxable"],
            "_gstr1_igst":    g1["igst"],
            "_gstr1_cgst":    g1["cgst"],
            "_gstr1_sgst":    g1["sgst"],
            "_diff_taxable": -g1["taxable"],
            "_diff_igst":    -g1["igst"],
            "_diff_cgst":    -g1["cgst"],
            "_diff_sgst":    -g1["sgst"],
            "_remark": "Amazon Entry as per GSTR-1" if is_amazon else "Not in Books",
        })
        result_rows.append(out)

    logger.info("B2B Reco: %d rows (%d GSTR-1 unmatched)", len(result_rows), len(gstr1_rows) - len(matched_gstr1))
    return result_rows


def _set_empty_gstr1(out: dict) -> None:
    out.update({
        "_gstr1_inv_no": "", "_gstr1_gstin": "",
        "_gstr1_taxable": None, "_gstr1_igst": None,
        "_gstr1_cgst": None, "_gstr1_sgst": None,
        "_diff_taxable": None, "_diff_igst": None,
        "_diff_cgst": None, "_diff_sgst": None,
    })


def _json_safe_val(v):
    """Convert a value to JSON-serialisable form."""
    if v is None:
        return None
    # Handle pandas NaT / NaN
    try:
        import pandas as _pd
        if _pd.isnull(v):
            return None
    except Exception:
        pass
    if isinstance(v, bool):
        return v
    if isinstance(v, int):
        return v
    if isinstance(v, float):
        return None if v != v else v  # NaN → None
    if isinstance(v, datetime):
        try:
            return v.strftime("%Y-%m-%d")
        except Exception:
            return None
    # pandas Timestamp
    try:
        return v.strftime("%Y-%m-%d")
    except AttributeError:
        pass
    if isinstance(v, str):
        return v
    try:
        f = float(v)
        return None if f != f else f
    except (TypeError, ValueError):
        return str(v)


# ---------------------------------------------------------------------------
# B2C Reconciliation (state + rate annual aggregation)
# ---------------------------------------------------------------------------

def reconcile_b2c_new(
    tally_df: pd.DataFrame,
    gstr1_df: pd.DataFrame,
    tolerance: float = 1.0,
) -> list[dict]:
    """State + rate annual aggregation: GSTR-1 B2C vs Books B2C."""
    # Books B2C aggregation
    cat_col     = _find_col(tally_df, ["Category", "Catogary"])
    part_col    = _find_col(tally_df, ["Particulars", "Party Name", "Buyer", "Ledger Name"])
    gstin_col   = _find_col(tally_df, ["GSTIN", "GSTIN/UIN"])
    state_col   = _find_col(tally_df, ["States", "Place of Supply", "State"])
    rate_col    = _find_col(tally_df, ["Rate", "Tax Rate", "GST Rate"])
    taxable_col = _find_col(tally_df, ["Total Sales", "Taxable Value"])
    igst_col    = _find_col(tally_df, ["Total IGST", "IGST"])
    cgst_col    = _find_col(tally_df, ["Total CGST", "CGST"])
    sgst_col    = _find_col(tally_df, ["Total SGST", "SGST"])

    inv_col     = _find_col(tally_df, ["Voucher No.", "Voucher No", "Invoice No", "Doc No", "Bill No"])
    ref_col     = _find_col(tally_df, ["Voucher Ref. No.", "Voucher Ref No", "Ref No"])
    inv_cols    = tuple(c for c in (inv_col, ref_col) if c)
    g1_b2b_invs = _gstr1_b2b_invoice_set(gstr1_df)

    books_agg: dict[tuple, dict] = defaultdict(_zero_amounts)

    for _, row in tally_df.iterrows():
        cat = _infer_category(row, cat_col, gstin_col, part_col, inv_cols, g1_b2b_invs)
        if cat != "B2C":
            continue
        taxable = _f(_col_val(row, taxable_col, 0))
        igst_v  = _f(_col_val(row, igst_col, 0))
        cgst_v  = _f(_col_val(row, cgst_col, 0))
        sgst_v  = _f(_col_val(row, sgst_col, 0))
        if taxable == 0 and igst_v == 0 and cgst_v == 0 and sgst_v == 0:
            continue
        # A raw register has no States/Rate column — fall back to the GSTIN state code,
        # the ledger name, and the rate implied by the amounts.
        state = _state_from_row(row, state_col, gstin_col, part_col)
        rate  = _f(_col_val(row, rate_col, 0)) if rate_col else 0.0
        if not rate:
            rate = _derive_rate(taxable, igst_v, cgst_v, sgst_v)
        key   = (state, rate)
        _add_amounts(books_agg[key], taxable=taxable, igst=igst_v, cgst=cgst_v, sgst=sgst_v)

    # GSTR-1 B2C aggregation
    g1_gstin_col   = _find_col(gstr1_df, ["Customer GSTIN", "GSTIN of Recipient"])
    g1_state_col   = _find_col(gstr1_df, ["Place of Supply", "State"])
    g1_rate_col    = _find_col(gstr1_df, ["GST Rate", "Rate", "Tax Rate"])
    g1_taxable_col = _find_col(gstr1_df, ["Item Taxable Value", "Taxable Value"])
    g1_igst_col    = _find_col(gstr1_df, ["IGST", "Integrated Tax"])
    g1_cgst_col    = _find_col(gstr1_df, ["CGST", "Central Tax"])
    g1_sgst_col    = _find_col(gstr1_df, ["SGST", "State Tax"])

    gstr1_agg: dict[tuple, dict] = defaultdict(_zero_amounts)

    if not gstr1_df.empty:
        for _, row in gstr1_df.iterrows():
            gstin = str(_col_val(row, g1_gstin_col, "")).strip()
            if gstin and len(re.sub(r"[^A-Z0-9]", "", gstin.upper())) == 15:
                continue  # B2B row — skip
            state = _extract_pos_state(str(_col_val(row, g1_state_col, "")).strip())
            g_taxable = _f(_col_val(row, g1_taxable_col, 0))
            g_igst    = _f(_col_val(row, g1_igst_col, 0))
            g_cgst    = _f(_col_val(row, g1_cgst_col, 0))
            g_sgst    = _f(_col_val(row, g1_sgst_col, 0))
            rate  = _f(_col_val(row, g1_rate_col, 0))
            if not rate:
                rate = _derive_rate(g_taxable, g_igst, g_cgst, g_sgst)
            key   = (state, rate)
            _add_amounts(
                gstr1_agg[key],
                taxable=g_taxable, igst=g_igst, cgst=g_cgst, sgst=g_sgst,
            )

    all_keys = sorted(set(books_agg) | set(gstr1_agg), key=lambda k: (k[0], k[1]))
    rows = []
    for state, rate in all_keys:
        bv = books_agg.get((state, rate), _zero_amounts())
        gv = gstr1_agg.get((state, rate), _zero_amounts())
        rows.append({
            "state":           state,
            "rate":            rate,
            "gstr1_taxable":   round_money(gv["taxable"]),
            "gstr1_igst":      round_money(gv["igst"]),
            "gstr1_cgst":      round_money(gv["cgst"]),
            "gstr1_sgst":      round_money(gv["sgst"]),
            "books_taxable":   round_money(bv["taxable"]),
            "books_igst":      round_money(bv["igst"]),
            "books_cgst":      round_money(bv["cgst"]),
            "books_sgst":      round_money(bv["sgst"]),
            "diff_taxable":    round_money(gv["taxable"] - bv["taxable"]),
            "diff_igst":       round_money(gv["igst"]    - bv["igst"]),
            "diff_cgst":       round_money(gv["cgst"]    - bv["cgst"]),
            "diff_sgst":       round_money(gv["sgst"]    - bv["sgst"]),
        })
    logger.info("B2C Reco: %d state+rate rows", len(rows))
    return rows


def _extract_pos_state(pos: str) -> str:
    """
    Strip a state code prefix like '07-Delhi' → 'Delhi'.

    When the code is a known GST state code the canonical name is used, so that
    Books states derived from a GSTIN spell out identically to GSTR-1's Place of
    Supply and the two sides actually group together.
    """
    if "-" in pos:
        parts = pos.split("-", 1)
        code = parts[0].strip()
        if code.isdigit():
            return _GST_STATE_CODES.get(code.zfill(2), parts[1].strip().title())
    return pos.strip().title()


# ---------------------------------------------------------------------------
# PDF parser for GSTR-1 portal PDF
# ---------------------------------------------------------------------------

def parse_gstr1_pdf_monthly(pdf_bytes: bytes) -> dict[str, dict] | None:
    """
    Extract monthly totals from GST-portal GSTR-1 PDF.
    Returns {month: {taxable, igst, cgst, sgst}} or None on failure.
    """
    try:
        import pdfplumber
    except ImportError:
        logger.warning("pdfplumber not installed — PDF parsing skipped")
        return None

    try:
        result: dict[str, dict] = {}
        with pdfplumber.open(BytesIO(pdf_bytes)) as pdf:
            for page in pdf.pages:
                tables = page.extract_tables()
                for table in tables:
                    if not table or len(table) < 2:
                        continue
                    header = [str(c or "").lower().strip() for c in table[0]]
                    # Look for Month | Taxable | IGST | CGST | SGST columns
                    month_idx   = _col_idx(header, ["month", "period", "return period"])
                    taxable_idx = _col_idx(header, ["taxable", "taxable value", "taxable amount"])
                    igst_idx    = _col_idx(header, ["igst", "integrated"])
                    cgst_idx    = _col_idx(header, ["cgst", "central"])
                    sgst_idx    = _col_idx(header, ["sgst", "state"])
                    if month_idx is None or (taxable_idx is None and igst_idx is None):
                        continue
                    for data_row in table[1:]:
                        if not data_row or len(data_row) <= (month_idx or 0):
                            continue
                        month = _norm_month(data_row[month_idx])
                        if month not in _FY_MONTHS:
                            continue
                        result[month] = {
                            "taxable": _parse_pdf_num(data_row, taxable_idx),
                            "igst":    _parse_pdf_num(data_row, igst_idx),
                            "cgst":    _parse_pdf_num(data_row, cgst_idx),
                            "sgst":    _parse_pdf_num(data_row, sgst_idx),
                        }
        return result if result else None
    except Exception as e:
        logger.warning("PDF parsing failed: %s", e)
        return None


def _col_idx(header: list[str], candidates: list[str]) -> int | None:
    for i, h in enumerate(header):
        if any(c in h for c in candidates):
            return i
    return None


def _parse_pdf_num(row: list, idx: int | None) -> float:
    if idx is None or idx >= len(row):
        return 0.0
    return _f(str(row[idx] or "").replace(",", "").strip())


# ---------------------------------------------------------------------------
# DataFrame → JSON-safe records helper
# ---------------------------------------------------------------------------

def df_to_records(df: pd.DataFrame) -> list[dict]:
    """Convert DataFrame to list of JSON-safe dicts."""
    records = []
    for _, row in df.iterrows():
        records.append({col: _json_safe_val(row[col]) for col in df.columns})
    return records


# ---------------------------------------------------------------------------
# Summary builder (for frontend display)
# ---------------------------------------------------------------------------

def build_summary(b2b_rows: list[dict], b2c_rows: list[dict]) -> dict:
    remark_counts: dict[str, int] = defaultdict(int)
    for r in b2b_rows:
        remark_counts[r.get("_remark", "Unknown")] += 1
    return {
        "B2B rows": len(b2b_rows),
        "B2C rows": len(b2c_rows),
        **{f"B2B: {k}": v for k, v in remark_counts.items()},
    }
