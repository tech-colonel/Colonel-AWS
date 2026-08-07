"""
pdf_bank_extractor.py — Converts ANY Indian bank statement PDF (digital text
layer) to structured Excel, reproducing the statement's own columns.

Approach (deterministic — the same path the best open-source parsers use):
  1. Find the transaction-table header (multi-line aware) by keyword scoring.
  2. Detect column boundaries by X-projection of the DATA rows — column
     separators are the vertical whitespace gaps that persist across rows.
     This is layout-agnostic: it discovers serial-#, Value-Date, MODE, etc.
     as their own columns instead of forcing a fixed schema.
  3. Semantically tag columns (date / narration / ref / debit / credit /
     balance) so we can build rows and the running-balance check points.
  4. Assemble rows anchor-first: a line with a valid date in the primary date
     column is a transaction anchor; date-less lines attach to the nearest
     anchor by Y (handles narration that wraps ABOVE and BELOW the amount line,
     e.g. ICICI).

Output Excel = every source column, verbatim + Check Point 1 / Check Point 2 +
a Statement Summary block. Verified against the statement's own printed totals
and/or running-balance continuity (the "Golden Rule").
"""
from __future__ import annotations

import logging
import re
from io import BytesIO
from typing import Optional

from recon import format_learn

logger = logging.getLogger(__name__)

# Tier 3 (full-PDF Claude extraction) is the expensive last resort — it sends the
# whole statement to the LLM. Cap it to small PDFs so a large statement can never
# blow up tokens/cost; big statements must be handled by Tier 1/2 (deterministic +
# learned template), never by full re-extraction.
_MAX_LLM_EXTRACT_PAGES = 5

# OCR fallback caps — OCR (rasterise + re-parse a big image PDF) is memory-heavy and
# the production reco engine runs on a small (2 GB) shared box; an OOM there would
# crash ALL agents. So only OCR scans within these limits; larger scans get a clear
# "too large for the server" message instead of risking the engine.
_MAX_OCR_BYTES = 40_000_000     # ~40 MB input
_MAX_OCR_PAGES = 50

# ──────────────────────────────────────────────────────────────────
# Column header keyword lists (lowercase substring match)
# ──────────────────────────────────────────────────────────────────
_DATE_KW    = ["txn date", "tran date", "transaction date", "value date", "posting date", "date"]
_NARR_KW    = ["narration", "description", "particulars", "transaction details", "details", "remarks"]
_REF_KW     = ["chq", "ref no", "cheque no", "reference", "instrument", "ref."]
# NOTE: never put a bare "dr"/"cr" here — it substring-matches "DesCRiption",
# "WithDRawal" etc. Standalone Dr/Cr flag columns are handled in _tag_for().
_DEBIT_KW   = ["withdrawal amt", "withdrawals", "withdrawal", "debit (inr)", "debit amt",
               "(dr)", "(dr.)", "dr amount", "debit", "withdrawl", "paid out"]
_CREDIT_KW  = ["deposit amt", "deposits", "deposit", "credit (inr)", "credit amt",
               "(cr)", "(cr.)", "cr amount", "credit", "paid in"]
_BALANCE_KW = ["closing balance", "running balance", "balance (inr)", "balance"]
# Credit-card / single-amount statements: one signed "Amount" column + a DR/CR
# flag column (no Withdrawal/Deposit split, no running balance).
_AMOUNT_KW  = ["transaction amount", "txn amount", "amount (inr)", "amount"]
_DRCR_KW    = ["dr/cr", "cr/dr", "dr / cr", "cr / dr", "debit/credit", "credit/debit"]
_TOTAL_SET  = {"total", "grand total", "totals", "total:", "subtotal", "closing balance", "statement summary"}
# End-of-transactions / footer phrases. A line matching one is SKIPPED (not a
# transaction) — but we do NOT permanently stop, because several of these recur
# in per-page headers/footers BEFORE the statement ends (e.g. a bank's
# "Registered Office" address printed at the foot of every page). Kept
# conservative: only phrases that are footer/legend boilerplate, never a txn row.
_END_MARKERS = ("commonly used narrations", "end of statement",
                "this is a system generated", "unless the constituent",
                "registered office", "please examine the statement",
                "computer generated statement", "abbreviations used",
                "the limits and effective available", "details on our products")
# Per-page footer like "Page 2 of 30" — skip it (else it bleeds into the last column).
_PAGE_FOOTER_RE = re.compile(r'^page\s+\d+\s+of\s+\d+\b')


def _clean_text(s: str) -> str:
    """Normalise odd PDF glyphs in extracted text: replacement char (�), non-breaking
    / zero-width spaces → normal space; collapse whitespace."""
    if not s:
        return s
    s = s.replace('�', ' ').replace('\xa0', ' ').replace('​', '').replace('\t', ' ')
    return ' '.join(s.split())

# Real reference number: at least 6 consecutive digits
_REAL_REF_RE = re.compile(r'\d{6,}')

# Tag detection order — first match wins. Unique tags (debit/credit/balance/narr/ref)
# are claimed once; date may repeat (Transaction Date + Value Date).
_TAG_ORDER = [
    ('drcr',    _DRCR_KW),
    ('balance', _BALANCE_KW),
    ('credit',  _CREDIT_KW),
    ('debit',   _DEBIT_KW),
    ('amount',  _AMOUNT_KW),
    ('ref',     _REF_KW),
    ('narr',    _NARR_KW),
    ('date',    _DATE_KW),
]
_UNIQUE_TAGS = {'balance', 'credit', 'debit', 'amount', 'drcr', 'ref', 'narr'}
_NUMERIC_TAGS = {'debit', 'credit', 'balance', 'amount'}


def _match(cell: str, keywords: list[str]) -> bool:
    c = (cell or "").lower().strip()
    return any(k in c for k in keywords)


def _tag_for(text: str) -> Optional[str]:
    t = (text or "").lower().strip()
    if not t:
        return None
    # Standalone Dr/Cr flag column (exact token only — never substring)
    if t in ('dr', 'dr.', 'cr', 'cr.'):
        return 'debit' if t[0] == 'd' else 'credit'
    for tag, kws in _TAG_ORDER:
        if any(k in t for k in kws):
            return tag
    return None


# ──────────────────────────────────────────────────────────────────
# Amount parsing — handles ₹, commas, Dr/Cr suffix, (parentheses), trailing −
# ──────────────────────────────────────────────────────────────────
_AMT_RE = re.compile(r'-?\d[\d,]*\.?\d*')


def _fix_ocr_number(s: str) -> str:
    """OCR frequently misreads a thousands COMMA as a period, producing a number with
    two or more dots (e.g. '15.487.70' for 15,487.70; '1.23.456.78' for 1,23,456.78).
    When a token has ≥2 dots, keep the LAST as the decimal point and drop the rest
    (they were thousands separators). Single-dot (normal) numbers are untouched, so
    this only ever helps OCR'd scans and never changes a clean digital number."""
    if s.count('.') >= 2:
        head, _, tail = s.rpartition('.')
        return head.replace('.', '') + '.' + tail
    return s


def _parse_amount(value) -> float:
    """Parse a currency cell to float. Returns 0.0 when the cell is not a number.
    Understands 1,234.00 / ₹1,234 / (1,234.00) neg / 1,234.00Dr / 1,234.00 Cr / 1234.00-."""
    if value is None:
        return 0.0
    if isinstance(value, (int, float)):
        return float(value)
    s = str(value).strip()
    if not s or s.lower() in ("-", "—", "nil", "n/a", "na", "--", "b/f", "c/f"):
        return 0.0
    neg = False
    low = s.lower()
    # Parenthesised negative
    if s.startswith("(") and s.endswith(")"):
        neg = True
    # Trailing Dr = withdrawal is caller's concern; here Dr/Cr only flags sign for balance cells
    if low.endswith("dr") or low.endswith("dr.") or low.endswith("-"):
        neg = True
    s = re.sub(r"[₹,\s]", "", s)
    s = _fix_ocr_number(s)   # collapse OCR multi-dot thousands (15.487.70 → 15487.70)
    m = _AMT_RE.search(s.replace("(", "").replace(")", ""))
    if not m:
        return 0.0
    try:
        v = float(m.group(0).replace(",", ""))
    except ValueError:
        return 0.0
    return -abs(v) if neg else v


_MONTHS = {'jan': 1, 'feb': 2, 'mar': 3, 'apr': 4, 'may': 5, 'jun': 6,
           'jul': 7, 'aug': 8, 'sep': 9, 'oct': 10, 'nov': 11, 'dec': 12}


def _parse_date(s: str) -> str:
    """Return a normalised DD/MM/YYYY if `s` is a single date token, else ''.
    Rejects strings with trailing junk (two dates, date+ref) so they aren't
    mistaken for a transaction anchor."""
    s = (s or "").strip()
    m = re.match(r"^(\d{1,2})[/\-\.](\d{1,2})[/\-\.](\d{2,4})$", s)
    if m:
        d, mo, y = m.group(1), m.group(2), m.group(3)
        if len(y) == 2:
            y = "20" + y
        return f"{int(d):02d}/{int(mo):02d}/{y}"
    m = re.match(r"^(\d{1,2})[\s\-]+([A-Za-z]{3,9})[\s\-]+(\d{2,4})$", s)
    if m:
        mon = _MONTHS.get(m.group(2)[:3].lower())
        if mon:
            y = m.group(3)
            if len(y) == 2:
                y = "20" + y
            return f"{int(m.group(1)):02d}/{mon:02d}/{y}"
    return ""


_LEADING_DATE_RE = re.compile(
    r'^\s*(\d{1,2}[/\-.]\d{1,2}[/\-.]\d{2,4}|\d{1,2}[\s\-][A-Za-z]{3,9}[\s\-]\d{2,4})')


def _leading_date(s: str) -> tuple[str, str]:
    """If the cell STARTS with a date, return (original_date_token, trailing_text).
    Tight columns often bleed the first narration word into the date cell
    (e.g. "02/04/25 APBS C-"); this recovers the date (kept in its original
    format for display) and the bled text so the row is still recognised."""
    m = _LEADING_DATE_RE.match(s or "")
    if not m or not _parse_date(m.group(1)):
        return "", ""
    return m.group(1), s[m.end():].strip()


# ──────────────────────────────────────────────────────────────────
# Line grouping
# ──────────────────────────────────────────────────────────────────
# OCR renderings of table ruling (vertical) lines. On scanned statements a standalone
# one of these merges adjacent header columns (Axis "Tran Date"|"Chq No" → loses the date
# tag) or sticks to an amount ("737166.42]"). Dropped in line grouping + tolerated in amount
# recognition. NEVER stripped from narration content (only whole-token / leading-trailing).
_OCR_SEP_CHARS = "|│┃¦‖∣ǀ︱❘][‹›"
_OCR_SEP_SET = set(_OCR_SEP_CHARS)


def _group_lines(words: list, tol: float = 3.5) -> list[dict]:
    """Group extracted words into visual lines by Y (top). Returns [{y, words[]}] sorted top→bottom."""
    # Drop standalone ruling-line glyphs (a token made ONLY of separator chars) so they
    # don't bridge column gaps on scanned/OCR'd statements. A narration token that merely
    # contains a '|' among other chars is kept (set difference is non-empty).
    words = [w for w in words
             if (w.get('text') or '').strip() and (set(w['text'].strip()) - _OCR_SEP_SET)]
    lines: list[dict] = []
    for w in sorted(words, key=lambda w: (w['top'], w['x0'])):
        placed = False
        for ln in reversed(lines[-6:]):
            if abs(w['top'] - ln['y']) <= tol:
                ln['words'].append(w)
                ln['y'] = (ln['y'] * ln['n'] + w['top']) / (ln['n'] + 1)
                ln['n'] += 1
                placed = True
                break
        if not placed:
            lines.append({'y': w['top'], 'n': 1, 'words': [w]})
    for ln in lines:
        ln['words'].sort(key=lambda w: w['x0'])
    return sorted(lines, key=lambda ln: ln['y'])


# ──────────────────────────────────────────────────────────────────
# Header band detection (multi-line aware)
# ──────────────────────────────────────────────────────────────────
def _distinct_tags(words: list) -> set:
    tags = set()
    for w in words:
        t = _tag_for(w['text'])
        if t:
            tags.add(t)
    return tags


# Long lowercase words that ARE legitimate column labels. Anything else long and
# lowercase-initial is running text, not a header cell.
_HEADER_WORD_WHITELIST = {
    'particulars', 'balance', 'deposit', 'deposits', 'withdrawal', 'withdrawals',
    'description', 'narration', 'transaction', 'transactions', 'details', 'amount',
    'cheque', 'reference', 'remarks', 'closing', 'opening', 'credit', 'debit',
    'value', 'branch', 'instrument', 'particular', 'chq/ref', 'nos.',
}


def _is_prose_line(line: dict) -> bool:
    """Running text rather than a row of column labels.

    Detected by a long lowercase-initial word that is not a known column label —
    'outstanding', 'available', 'statement period' prose and so on. Kept
    deliberately narrow so genuinely wrapped headers ('Running' / 'Balance',
    ICICI's 3-line header) are never rejected.
    """
    for w in line.get('words', []):
        t = (w.get('text') or '').strip().strip('.,:;')
        if len(t) >= 6 and t[:1].islower() and t.lower() not in _HEADER_WORD_WHITELIST:
            return True
    return False


def _find_header_band(lines: list[dict]) -> Optional[dict]:
    """Choose the transaction-table header. Tries each line and each pair of
    consecutive lines (wrapped headers like 'Running / Balance'); picks the band
    with the most distinct semantic tags. Requires a date column plus a balance
    or (debit/credit) column so summary tables above don't win."""
    best = None
    for i in range(len(lines)):
        for span in (1, 2, 3):
            if i + span > len(lines):
                continue
            # span 2-3 is for headers wrapped across multiple lines (e.g. "Running"/"Balance",
            # or ICICI's 3-line "Withdrawal|Deposit|Balance" / "S No.|Cheque|Remarks" / "Date|Amount(INR)").
            # Never merge a DATA row into the header — a real date value in the extra
            # line means it's a transaction row, not a wrapped header.
            if span >= 2 and any(_parse_date(w['text'])
                                 for ln in lines[i + 1:i + span] for w in ln['words']):
                continue
            # Never merge PROSE into the header. Statements print marketing text
            # directly above the table ("...to know your outstanding balance.
            # Details"), and a stray word like "balance." adds a `balance` tag,
            # so the polluted band outscores the real header and every column
            # below it is mis-tagged. Measured on a Yes Bank credit-card
            # statement: 'Transaction Details' was tagged `balance`, so every
            # narration was run through _parse_amount() and destroyed.
            # Checks EVERY line in the band, including the first: the prose sits
            # ABOVE the real header, so a band starting on the prose line would
            # slip through a check that only looked at the trailing lines.
            if span >= 2 and any(_is_prose_line(ln) for ln in lines[i:i + span]):
                continue
            band_words = [w for ln in lines[i:i + span] for w in ln['words']]
            # header lines are short; a data line with 30 words isn't a header
            if len(band_words) > 22:
                continue
            tags = _distinct_tags(band_words)
            has_money = bool(tags & {'balance', 'debit', 'credit', 'amount'})
            # Require date + a description/narration column + a money column. Real
            # transaction tables always have a narration column; summary/totals
            # tables (e.g. a credit-card "Cards Debits Credits Date" block on page 1)
            # do not — so this skips them and finds the true transaction header,
            # which on credit-card statements sits on a later page.
            if not (tags >= {'date', 'narr'}) or not has_money:
                continue
            score = len(tags)
            key = (score, -span, -len(band_words))
            if best is None or key > best['key']:
                y0 = min(w['top'] for w in band_words)
                y1 = max(w['bottom'] for w in band_words)
                best = {'key': key, 'idx': i, 'span': span,
                        'words': band_words, 'y0': y0, 'y1': y1, 'tags': tags}
    return best


# ──────────────────────────────────────────────────────────────────
# Column detection — HYBRID: header-gap clustering + data-projection split
# ──────────────────────────────────────────────────────────────────
_HDR_GAP = 8.0     # header words > this far apart start a new column
_MIN_GAP = 5       # data whitespace gap (pt) that marks a sub-column split


def _data_bands(data_lines, lo: float, hi: float, page_w: int) -> list[list[int]]:
    """Whitespace-separated occupancy bands within x-range [lo, hi), from the
    DATA rows only (used to split tight amount columns without description
    interference — only words whose CENTER falls in the range count)."""
    occ = [0] * (page_w + 2)
    for ln in data_lines[:200]:
        for w in ln['words']:
            cx = (w['x0'] + w['x1']) / 2
            if lo <= cx < hi:
                a = max(int(lo), int(w['x0']))
                b = min(int(hi), int(w['x1']))
                for x in range(a, b + 1):
                    occ[x] += 1
    bands: list[list[int]] = []
    i, W = int(lo), int(hi)
    while i < W:
        if occ[i] > 0:
            j = i
            while j < W and occ[j] > 0:
                j += 1
            if bands and i - bands[-1][1] - 1 < _MIN_GAP:
                bands[-1][1] = j - 1
            else:
                bands.append([i, j - 1])
            i = j
        else:
            i += 1
    return bands


_AMOUNT_TOKEN_RE = re.compile(r'^\(?-?[\d,]*\d\.\d{1,2}\)?(?:\s*\(?[dc]r\.?\)?)?$', re.I)


def _looks_amount(text: str) -> bool:
    """A right-aligned money value: digits with a 1–2 dp decimal (₹4,601.31),
    optional (), optional Dr/Cr. Deliberately excludes long ref numbers (no
    decimal) so we don't mistake them for amounts."""
    return bool(_AMOUNT_TOKEN_RE.match(
        (text or "").strip().replace('₹', '').strip(_OCR_SEP_CHARS + " ")))


def _right_edge_columns(data_lines, lo: float, hi: float, tol: float = 6.0) -> list[float]:
    """Split an amount region by clustering the RIGHT edges (x1) of decimal
    money tokens. Right-aligned columns each share a right edge, so this cleanly
    separates Withdrawals/Deposits/Balance even when they never co-occur on a
    row (occupancy projection can't). Returns representative x per column."""
    xs = []
    for ln in data_lines[:400]:
        for w in ln['words']:
            cx = (w['x0'] + w['x1']) / 2
            if lo <= cx < hi and _looks_amount(w['text']):
                xs.append(w['x1'])
    if len(xs) < 3:
        return []
    xs.sort()
    clusters = [[xs[0]]]
    for x in xs[1:]:
        if x - clusters[-1][-1] <= tol:
            clusters[-1].append(x)
        else:
            clusters.append([x])
    support = max(2, len(xs) // 25)
    return [sum(c) / len(c) for c in clusters if len(c) >= support]


def _split_region(data_lines, lo, hi, page_w, hwords) -> Optional[list[dict]]:
    """Try to split a header region into sub-columns. Prefer right-edge amount
    clustering (robust for money columns); fall back to whitespace bands. Accept
    a split only if every sub-column gets ≥1 header word and #subs ≤ #headers —
    so wide text columns never fragment into phantom empty columns."""
    candidates: list[list[tuple]] = []

    reps = sorted(_right_edge_columns(data_lines, lo, hi))
    if len(reps) >= 2:
        ranges = []
        for k, r in enumerate(reps):
            sl = lo if k == 0 else (reps[k - 1] + r) / 2.0
            sr = hi if k == len(reps) - 1 else (r + reps[k + 1]) / 2.0
            ranges.append((sl, sr))
        candidates.append(ranges)

    bands = _data_bands(data_lines, lo, hi, page_w)
    if len(bands) >= 2:
        ranges = []
        for k, (s, e) in enumerate(bands):
            sl = lo if k == 0 else (bands[k - 1][1] + s) / 2.0
            sr = hi if k == len(bands) - 1 else (e + bands[k + 1][0]) / 2.0
            ranges.append((sl, sr))
        candidates.append(ranges)

    best = None
    for cand in candidates:
        if not (2 <= len(cand) <= len(hwords)):
            continue
        subs, ok = [], True
        for (sl, sr) in cand:
            hw = [w for w in hwords if sl <= (w['x0'] + w['x1']) / 2 < sr]
            if not hw:
                ok = False
                break
            subs.append({'x0': sl, 'x1': sr, 'words': hw})
        if ok and (best is None or len(subs) > len(best)):
            best = subs
    return best


def _text_boundary(data_lines: list, lo: float, hi: float, n: int = 300) -> Optional[float]:
    """For two adjacent LEFT-aligned text columns, find the vertical whitespace
    corridor in the DATA between them (within [lo, hi]) and return its midpoint.
    Returns None when there's no clear corridor — i.e. the right column is (nearly)
    empty, so the caller keeps the left column extended to the next header. This
    handles both a right column whose data starts LEFT of its header (Union Bank
    Remarks) and an empty right column that must not steal overflow (IDFC ref)."""
    spans = []
    for ln in data_lines[:n]:
        for w in ln['words']:
            if w['x1'] > lo and w['x0'] < hi:
                spans.append((max(w['x0'], lo), min(w['x1'], hi)))
    if len(spans) < 2:
        return None
    spans.sort()
    merged = [list(spans[0])]
    for a, b in spans[1:]:
        if a <= merged[-1][1] + 0.5:
            merged[-1][1] = max(merged[-1][1], b)
        else:
            merged.append([a, b])
    if len(merged) < 2:
        return None
    gap, mid = max(((a2 - b1, (b1 + a2) / 2.0) for (_, b1), (a2, _) in zip(merged, merged[1:])),
                   key=lambda t: t[0])
    return mid if gap >= 3.0 else None


def _col_data_tokens(col: dict, data_lines: list, sample: int = 60) -> list[list[str]]:
    """Non-empty per-line token lists that fall inside this column's X-range."""
    out = []
    for ln in data_lines[:400]:
        toks = [w['text'] for w in ln['words']
                if col['x0'] <= (w['x0'] + w['x1']) / 2 < col['x1'] and (w['text'] or '').strip()]
        if toks:
            out.append(toks)
        if len(out) >= sample:
            break
    return out


def _is_dateish_column(col: dict, data_lines: list) -> bool:
    """A column whose DATA is mostly dates (a secondary 'Value Dt' column) — kept
    OUT of the narration merge so a running value-date never pollutes the narration."""
    rows = _col_data_tokens(col, data_lines)
    if len(rows) < 4:
        return False
    hit = sum(1 for toks in rows if _leading_date(" ".join(toks))[0] or _parse_date(" ".join(toks)))
    return hit / len(rows) > 0.5


def _is_numberish_column(col: dict, data_lines: list) -> bool:
    """A column whose DATA is mostly money-shaped — a defensive guard so a
    mis-tagged amount column is never swallowed into the narration merge (which
    would drop money and break balance verification)."""
    rows = _col_data_tokens(col, data_lines)
    if len(rows) < 4:
        return False
    hit = sum(1 for toks in rows if all(_looks_amount(t) for t in toks))
    return hit / len(rows) > 0.5


def _merge_text_columns(columns: list[dict], data_lines: list[dict]) -> list[dict]:
    """Collapse each run of ≥2 consecutive DESCRIPTIVE columns into a single
    narration column. Header-clustering often sees a wrapped narration as two
    adjacent text columns (ICICI 'Cheque Number' + 'Transaction Remarks', Axis
    'Chq No' + 'Particulars'); binning words by X then splits one narration into
    both, fragmenting/scrambling it. Date and numeric columns break the run and
    are never merged; a date-like or amount-like text column is also excluded so
    a value-date or a mis-tagged amount can never be swallowed (amounts/balance,
    hence verification, are untouched)."""
    TEXT = {'narr', 'ref', 'other'}

    def _mergeable(c: dict) -> bool:
        return (c['tag'] in TEXT
                and not c.get('is_primary_date')
                and not _is_dateish_column(c, data_lines)
                and not _is_numberish_column(c, data_lines))

    out: list[dict] = []
    i, n = 0, len(columns)
    while i < n:
        if _mergeable(columns[i]):
            j = i + 1
            while j < n and _mergeable(columns[j]):
                j += 1
            run = columns[i:j]
            if len(run) > 1:
                narr_member = next((m for m in run if m['tag'] == 'narr'), None)
                out.append({
                    'header': (narr_member or run[0])['header'],
                    'tag': 'narr',
                    'x0': min(m['x0'] for m in run),
                    'x1': max(m['x1'] for m in run),
                    'is_primary_date': False,
                })
            else:
                out.append(columns[i])
            i = j
        else:
            out.append(columns[i])
            i += 1
    for k, c in enumerate(out):
        c['key'] = f"c{k}"
    return out


def _merge_symbol_columns(columns: list[dict]) -> list[dict]:
    """Fold a trailing symbol/punctuation-only header fragment into the preceding
    NUMERIC column. Credit-card statements print the amount header as "Amount(in ₹)";
    the ₹ glyph renders as a stray token so header-clustering leaves a phantom ")" /
    "` )" column that then captures the right-aligned amounts. Extending the amount
    column over it (out to the page edge) recovers those values."""
    out: list[dict] = []
    for c in columns:
        prev = out[-1] if out else None
        label = c.get('header') or ''
        is_symbol = not re.search(r'[A-Za-z0-9]', label)
        if prev is not None and is_symbol and prev['tag'] in _NUMERIC_TAGS:
            prev['x1'] = max(prev['x1'], c['x1'])
            continue
        out.append(c)
    for k, c in enumerate(out):
        c['key'] = f"c{k}"
    return out


def _detect_columns(page, header: dict, data_lines: list[dict]) -> Optional[list[dict]]:
    """Cluster header words into columns by X-gap (cleanly separates text
    columns from right-aligned amount columns), then split any cluster whose
    DATA shows multiple whitespace-separated bands (e.g. tight Withdrawals/
    Deposits/Balance headers that sit only a few points apart)."""
    page_w = int(page.width) + 2
    hdr = sorted(header['words'], key=lambda w: w['x0'])
    if not hdr:
        return None

    # 1) Header-gap clusters
    clusters: list[dict] = []
    for w in hdr:
        if clusters and w['x0'] - clusters[-1]['x1'] <= _HDR_GAP:
            clusters[-1]['words'].append(w)
            clusters[-1]['x1'] = max(clusters[-1]['x1'], w['x1'])
        else:
            clusters.append({'words': [w], 'x0': w['x0'], 'x1': w['x1']})

    # 2) Cluster x-boundaries. Default = midpoint between neighbours (ends fill the
    #    page). BUT text columns are LEFT-aligned: a long description overflows past
    #    the header midpoint into the visual gap before the next column. When a text
    #    column is followed by ANOTHER text column, give the left column all the space
    #    up to the next column's header left-edge, so the narration is not split into
    #    the (often empty) adjacent text column — e.g. IDFC "Particulars" overflowing
    #    into an empty "Cheq./Ref.No". Any boundary touching an amount column keeps the
    #    midpoint (amounts are right-aligned and further guarded by right_x downstream).
    def _cluster_tag(cl: dict) -> str:
        lbl = " ".join(w['text'] for w in sorted(cl['words'], key=lambda w: (w['top'], w['x0'])))
        return _tag_for(lbl) or 'other'
    ctags = [_cluster_tag(cl) for cl in clusters]

    def _boundary(i: int) -> float:
        cl, nxt = clusters[i], clusters[i + 1]
        if ctags[i] in _NUMERIC_TAGS or ctags[i + 1] in _NUMERIC_TAGS:
            return (cl['x1'] + nxt['x0']) / 2.0           # amount involved → header midpoint
        # text → text: split at the DATA whitespace corridor between the two columns;
        # if none (right column empty), the left column owns the gap up to next header.
        b = _text_boundary(data_lines, float(cl['x0']), float(nxt['x1']))
        return b if b is not None else float(nxt['x0'])

    regions: list[tuple[float, float, dict]] = []
    for i, cl in enumerate(clusters):
        left = 0.0 if i == 0 else _boundary(i - 1)
        right = float(page_w) if i == len(clusters) - 1 else _boundary(i)
        regions.append((left, right, cl))

    # 3) Within each region, split ONLY when the data shows as many bands as the
    #    cluster has header words AND every band gets a header word. This splits
    #    tight amount headers (Withdrawals/Deposits/Balance) apart without
    #    fragmenting a wide text column into phantom empty columns.
    raw_cols: list[dict] = []
    for left, right, cl in regions:
        subs = _split_region(data_lines, left, right, page_w, cl['words'])
        if subs:
            raw_cols.extend(subs)
        else:
            raw_cols.append({'x0': left, 'x1': right, 'words': cl['words']})

    # 4) Label + tag each column (multi-line header words joined top→bottom)
    columns: list[dict] = []
    claimed: set = set()
    for k, rc in enumerate(raw_cols):
        hwords = sorted(rc['words'], key=lambda w: (w['top'], w['x0']))
        label = " ".join(w['text'] for w in hwords).strip()
        tag = _tag_for(label)
        if tag in _UNIQUE_TAGS and tag in claimed:
            tag = 'other'
        if tag in _UNIQUE_TAGS:
            claimed.add(tag)
        columns.append({
            'key': f"c{k}",
            'header': label or f"Column {k + 1}",
            'tag': tag or 'other',
            'x0': rc['x0'], 'x1': rc['x1'],
        })

    # Recover an amount column split by a rendered currency glyph (credit cards:
    # "Amount(in ₹)" → phantom ")" column) before anything else uses the columns.
    columns = _merge_symbol_columns(columns)

    # Collapse a narration that header-clustering split across adjacent text
    # columns (e.g. ICICI Cheque Number + Transaction Remarks) into one column.
    columns = _merge_text_columns(columns, data_lines)

    tags = {c['tag'] for c in columns}
    if 'date' not in tags or not (tags & _NUMERIC_TAGS):
        return None

    date_cols = [c for c in columns if c['tag'] == 'date']
    for c in columns:
        c['is_primary_date'] = (c is date_cols[0]) if date_cols else False

    # Right-edge X of each amount column: real money values are right-aligned and
    # share this edge; description numbers that leak into the column sit to the
    # left of it and are rejected at cell-assignment time.
    import statistics
    for c in columns:
        c['right_x'] = None
        if c['tag'] in _NUMERIC_TAGS:
            x1s = sorted(w['x1'] for ln in data_lines[:400] for w in ln['words']
                         if c['x0'] <= (w['x0'] + w['x1']) / 2 < c['x1'] and _looks_amount(w['text']))
            if len(x1s) >= 3:
                # Real values in an amount column share ONE right edge. A secondary
                # left sub-column can hide inside the same detected column (e.g. the
                # credit-card "International amount" 0.00 sitting beside the ₹ Amount);
                # cluster the right-edges and lock onto the RIGHTMOST cluster so the
                # true right-aligned amount wins and the 0.00 sub-column is ignored.
                clusters = [[x1s[0]]]
                for x in x1s[1:]:
                    if x - clusters[-1][-1] <= 15:
                        clusters[-1].append(x)
                    else:
                        clusters.append([x])
                c['right_x'] = statistics.median(clusters[-1])   # rightmost cluster

    logger.info("Detected %d columns: %s", len(columns),
                [(c['header'], c['tag']) for c in columns])
    return columns


_RIGHT_TOL = 12.0  # pt: how close an amount's right edge must be to the column's


def _assign_cells(line_words: list, columns: list[dict]) -> dict:
    """Bin a line's words into columns by center X. Returns {col_key: text}.
    For amount columns, a word is only kept if it is right-aligned to the
    column's amount edge — this rejects description text that overflows into
    the amount column (a common cause of phantom debits/credits)."""
    cells = {c['key']: '' for c in columns}
    for w in sorted(line_words, key=lambda w: w['x0']):
        cx = (w['x0'] + w['x1']) / 2
        for c in columns:
            if c['x0'] <= cx < c['x1']:
                if c['tag'] in _NUMERIC_TAGS:
                    # An amount cell accepts only money-shaped, right-aligned
                    # tokens — never a date/ref fragment overflowing from the
                    # description column.
                    if not _looks_amount(w['text']):
                        break
                    rx = c.get('right_x')
                    if rx is not None and abs(w['x1'] - rx) > _RIGHT_TOL:
                        break
                cur = cells[c['key']]
                cells[c['key']] = (cur + ' ' + w['text']) if cur else w['text']
                break
    return cells


# ──────────────────────────────────────────────────────────────────
# Anchor-based transaction assembly
# ──────────────────────────────────────────────────────────────────
def _primary_date_key(columns: list[dict]) -> Optional[str]:
    for c in columns:
        if c.get('is_primary_date'):
            return c['key']
    return None


def _tag_key(columns: list[dict], tag: str) -> Optional[str]:
    for c in columns:
        if c['tag'] == tag:
            return c['key']
    return None


def _is_header_like(cells: dict, columns: list[dict]) -> bool:
    # A real data row is never a repeated header: bail out if the date column holds a
    # real date, or any amount/balance cell holds a money value. (Guards against a
    # narration like "CARDLESS DEPOSIT" tripping the header-keyword count below.)
    dk = _primary_date_key(columns)
    if dk and _leading_date(cells.get(dk, ''))[0]:
        return False
    for c in columns:
        if c['tag'] in _NUMERIC_TAGS and _looks_amount((cells.get(c['key'], '') or '').strip()):
            return False
    hits = 0
    for c in columns:
        if c['tag'] != 'other' and _tag_for(cells.get(c['key'], '')):
            hits += 1
    return hits >= 3


def _build_transactions(all_lines: list[dict], columns: list[dict]) -> list[dict]:
    """all_lines: [{y, cells}] across the whole statement (in reading order).
    A line whose primary-date cell parses as a date is an anchor; date-less
    lines attach to the nearest anchor by Y (per page-block, carried across)."""
    date_key = _primary_date_key(columns)
    text_keys = [c['key'] for c in columns if c['tag'] not in _NUMERIC_TAGS and c['key'] != date_key]
    debit_key = _tag_key(columns, 'debit')
    credit_key = _tag_key(columns, 'credit')
    bal_key = _tag_key(columns, 'balance')
    amount_key = _tag_key(columns, 'amount')   # single signed-amount column (credit cards)
    drcr_key = _tag_key(columns, 'drcr')        # DR/CR direction flag

    def _amounts(cells):
        """Return (debit, credit) for a row. Prefers explicit debit/credit columns;
        else derives them from a single Amount column + a DR/CR flag."""
        d = max(0.0, _parse_amount(cells.get(debit_key))) if debit_key else 0.0
        c = max(0.0, _parse_amount(cells.get(credit_key))) if credit_key else 0.0
        if amount_key and not (debit_key or credit_key):
            raw = str(cells.get(amount_key, '') or '')
            amt = abs(_parse_amount(raw))
            flag = (cells.get(drcr_key, '') if drcr_key else '').strip().upper()
            if not flag:
                # No separate DR/CR column — direction is embedded in the amount cell
                # itself, e.g. Union Bank "228.00(Dr)" / "1500.00(Cr)".
                low = raw.lower()
                if '(cr)' in low or low.rstrip().endswith('cr'):
                    flag = 'C'
                elif '(dr)' in low or low.rstrip().endswith('dr'):
                    flag = 'D'
            if not flag:
                # Still no direction — use the SIGN of the amount. Credit-card
                # statements have one signed Amount column: positive = purchase
                # (debit / money owed), negative = payment or reversal (credit).
                # Only negatives change behaviour here (positives already fell to
                # debit below), so positive-only single-amount banks are unaffected.
                signed = _parse_amount(raw)
                if signed < 0:
                    flag = 'C'
                elif signed > 0:
                    flag = 'D'
            if flag.startswith('C'):        # CR → money in (payment/refund) = credit
                c = amt
            else:                            # DR / blank → money out (purchase) = debit
                d = amt
        return d, c

    # Typical line height → cap how far a date-less line may attach to an anchor
    # (a real wrapped narration is a few lines away; a footer block is far below).
    ys = sorted(ln['y'] for ln in all_lines)
    gaps = [b - a for a, b in zip(ys, ys[1:]) if 0 < b - a < 40]
    import statistics
    line_h = statistics.median(gaps) if gaps else 12.0
    max_attach = max(30.0, 5 * line_h)

    anchors: list[dict] = []
    pending: list[dict] = []   # date-less lines waiting to attach to nearest anchor

    def flush_pending():
        """Attach each pending (date-less) line to its nearest anchor by Y,
        but only within `max_attach` — footer/legend text far below the last
        transaction is dropped rather than glued onto its description."""
        if not pending:
            return
        for pl in pending:
            if not anchors:
                break
            # A wrapped-narration line belongs to the transaction it sits UNDER
            # (the anchor above). Only attach to the anchor BELOW when that one is
            # CLEARLY closer — i.e. a transaction-type label printed just above its
            # own amount row (ICICI "Credit trxn"/"iDirect trxn"). On a tie (a middle
            # line of a 3+ line wrap, equidistant between two rows) the line stays
            # with the row above. This stops continuation text leaking downward.
            above = [a for a in anchors if a['y'] <= pl['y']]
            below = [a for a in anchors if a['y'] > pl['y']]
            na = max(above, key=lambda a: a['y']) if above else None
            nb = min(below, key=lambda a: a['y']) if below else None
            da = (pl['y'] - na['y']) if na else float('inf')
            db = (nb['y'] - pl['y']) if nb else float('inf')
            if nb is not None and db + 0.4 * line_h < da:
                near = nb
            else:
                near = na if na is not None else nb
            if near is None or abs(near['y'] - pl['y']) > max_attach:
                continue
            for k in text_keys:
                txt = pl['cells'].get(k, '').strip()
                if txt:
                    near['frag'].setdefault(k, []).append((pl['y'], txt))
            # Some layouts (e.g. Bank of Baroda) print the running BALANCE on the
            # tran-date anchor line but the WITHDRAWAL/DEPOSIT amount on the SEPARATE
            # value-date continuation line. Amounts land in numeric columns, which the
            # text merge above ignores — so adopt a continuation line's debit/credit
            # (and balance) when the anchor is still missing it. Amount cells are
            # already money-shaped + right-aligned (guarded in _assign_cells), so a
            # wrapped narration token never leaks in here.
            if near['debit'] == 0.0 and near['credit'] == 0.0:
                pd_, pc_ = _amounts(pl['cells'])
                if pd_ or pc_:
                    near['debit'], near['credit'] = pd_, pc_
                    # Reflect into raw_cells too — the OUTPUT columns are rendered from
                    # raw_cells, so without this the amount verifies but shows blank.
                    if pd_ and debit_key:
                        near['raw_cells'][debit_key] = pl['cells'].get(debit_key, '')
                    if pc_ and credit_key:
                        near['raw_cells'][credit_key] = pl['cells'].get(credit_key, '')
                    if amount_key and not (debit_key or credit_key):
                        near['raw_cells'][amount_key] = pl['cells'].get(amount_key, '')
                        if drcr_key and pl['cells'].get(drcr_key, '').strip():
                            near['raw_cells'][drcr_key] = pl['cells'].get(drcr_key, '')
            if not near.get('has_balance') and bal_key and pl['cells'].get(bal_key, '').strip():
                near['balance'] = _parse_amount(pl['cells'].get(bal_key))
                near['has_balance'] = True
                near['raw_cells'][bal_key] = pl['cells'].get(bal_key, '')
        pending.clear()

    for ln in all_lines:
        cells = ln['cells']
        # Skip repeated headers
        if _is_header_like(cells, columns):
            continue
        # Totals / summary / footer / legend lines are not transactions — skip just
        # this line. Do NOT stop permanently: these phrases recur in per-page
        # headers/footers (e.g. a "Registered Office" line at the foot of each page).
        joined = " ".join(v.lower().strip() for v in cells.values() if v.strip())
        if any(t == joined or joined.startswith(t + " ") for t in _TOTAL_SET) \
                or any(m in joined for m in _END_MARKERS) \
                or _PAGE_FOOTER_RE.match(joined):
            continue

        date_raw = (cells.get(date_key, '') if date_key else '').strip()
        dtok, drest = _leading_date(date_raw)
        if dtok:
            # New anchor
            frag: dict = {}
            for k in text_keys:
                t = cells.get(k, '').strip()
                if t:
                    frag[k] = [(ln['y'], t)]
            # Recover narration text that bled into a too-wide date cell.
            if drest:
                nk = _tag_key(columns, 'narr')
                if nk:
                    frag.setdefault(nk, []).insert(0, (ln['y'] - 0.001, drest))
            anchors.append({
                'y': ln['y'],
                'date': dtok,
                # Debit/credit are never negative — from explicit columns or an
                # Amount+DR/CR split. Balance may be negative (overdraft).
                'debit': _amounts(cells)[0],
                'credit': _amounts(cells)[1],
                'balance': _parse_amount(cells.get(bal_key)) if bal_key else 0.0,
                'has_balance': bool(bal_key and cells.get(bal_key, '').strip()),
                'raw_cells': dict(cells),
                'frag': frag,
            })
            # Flush AFTER appending so a pending line that sits just ABOVE this new
            # anchor (e.g. ICICI's "Credit trxn"/"iDirect trxn" transaction-type
            # label, printed above the amount row) attaches to the nearer/correct
            # transaction instead of the previous one.
            flush_pending()
        else:
            # Date-less line: either continuation text, or a stray line.
            if any(cells.get(k, '').strip() for k in text_keys):
                pending.append(ln)
    flush_pending()

    # Derive missing debit/credit from balance movement (single-amount layouts)
    if (debit_key is None) ^ (credit_key is None) and bal_key is not None:
        prev = None
        for a in anchors:
            if prev is not None and a['has_balance']:
                delta = round(a['balance'] - prev, 2)
                if delta >= 0:
                    a['credit'] = delta
                else:
                    a['debit'] = -delta
            prev = a['balance'] if a['has_balance'] else prev

    # Finalise cells: join text fragments in reading order; numbers as floats
    txns: list[dict] = []
    for a in anchors:
        cells_out: dict = {}
        for c in columns:
            k = c['key']
            if c['tag'] == 'amount':
                # A single amount column may embed the direction ("1500.00(Cr)").
                # Preserve that text verbatim so Cr/Dr isn't lost; keep a plain amount
                # numeric (e.g. credit-card statements) so totals/format still work.
                raw = str(a['raw_cells'].get(k, '') or '').strip()
                # Preserve the SIGN for a single Amount column (credit-card layouts:
                # negative = payment/reversal). Clamping to ≥0 would blank out every
                # credit in the rendered sheet even though debit/credit are derived
                # correctly. Cells carrying a Dr/Cr word keep their text verbatim.
                cells_out[k] = _clean_text(raw) if re.search(r'[A-Za-z]', raw) else _parse_amount(raw)
            elif c['tag'] in _NUMERIC_TAGS:
                amt = _parse_amount(a['raw_cells'].get(k))
                cells_out[k] = amt if c['tag'] == 'balance' else max(0.0, amt)
            elif k == date_key:
                cells_out[k] = a['date']   # clean date token (bled narration already recovered)
            else:
                frags = sorted(a['frag'].get(k, []), key=lambda t: t[0])
                cells_out[k] = _clean_text(" ".join(t for _, t in frags))
        txns.append({
            'cells': cells_out,
            'date': a['date'],
            'debit': a['debit'],
            'credit': a['credit'],
            'balance': a['balance'],
            'description': cells_out.get(_tag_key(columns, 'narr'), ''),
            'ref_no': cells_out.get(_tag_key(columns, 'ref'), ''),
        })
    return txns


# ──────────────────────────────────────────────────────────────────
# Bank detection (IFSC-first — reliable across every Indian bank)
# ──────────────────────────────────────────────────────────────────
_IFSC_BANK = {
    'HDFC': 'HDFC Bank', 'ICIC': 'ICICI Bank', 'SBIN': 'State Bank of India',
    'KKBK': 'Kotak Mahindra Bank', 'UTIB': 'Axis Bank', 'INDB': 'IndusInd Bank',
    'YESB': 'Yes Bank', 'PUNB': 'Punjab National Bank', 'CNRB': 'Canara Bank',
    'BARB': 'Bank of Baroda', 'UBIN': 'Union Bank of India', 'IDFB': 'IDFC First Bank',
    'IBKL': 'IDBI Bank', 'FDRL': 'Federal Bank', 'RATN': 'RBL Bank', 'BDBL': 'Bandhan Bank',
    'AUBL': 'AU Small Finance Bank', 'IDIB': 'Indian Bank', 'CBIN': 'Central Bank of India',
    'MAHB': 'Bank of Maharashtra', 'PSIB': 'Punjab & Sind Bank', 'KARB': 'Karnataka Bank',
    'CIUB': 'City Union Bank', 'SIBL': 'South Indian Bank', 'TMBL': 'Tamilnad Mercantile Bank',
    'DBSS': 'DBS Bank', 'HSBC': 'HSBC Bank', 'CITI': 'Citibank', 'SCBL': 'Standard Chartered',
    'KVBL': 'Karur Vysya Bank', 'DLXB': 'Dhanlaxmi Bank', 'JAKA': 'J&K Bank',
    'UCBA': 'UCO Bank', 'IOBA': 'Indian Overseas Bank', 'MSNU': 'Equitas Small Finance Bank',
}
_BANK_NAME_PATTERNS = [
    (r'STATE BANK OF INDIA', 'State Bank of India'), (r'HDFC BANK', 'HDFC Bank'),
    (r'ICICI BANK', 'ICICI Bank'), (r'KOTAK MAHINDRA', 'Kotak Mahindra Bank'),
    (r'KOTAK BANK', 'Kotak Mahindra Bank'), (r'AXIS BANK', 'Axis Bank'),
    (r'INDUSIND', 'IndusInd Bank'), (r'YES BANK', 'Yes Bank'),
    (r'PUNJAB NATIONAL BANK', 'Punjab National Bank'), (r'CANARA BANK', 'Canara Bank'),
    (r'BANK OF BARODA', 'Bank of Baroda'), (r'UNION BANK', 'Union Bank of India'),
    (r'IDFC FIRST', 'IDFC First Bank'), (r'IDBI BANK', 'IDBI Bank'),
    (r'FEDERAL BANK', 'Federal Bank'), (r'RBL BANK', 'RBL Bank'),
    (r'BANDHAN', 'Bandhan Bank'), (r'AU SMALL FINANCE', 'AU Small Finance Bank'),
    (r'INDIAN BANK', 'Indian Bank'), (r'CENTRAL BANK OF INDIA', 'Central Bank of India'),
    (r'BANK OF MAHARASHTRA', 'Bank of Maharashtra'), (r'KARNATAKA BANK', 'Karnataka Bank'),
    (r'SOUTH INDIAN BANK', 'South Indian Bank'), (r'STANDARD CHARTERED', 'Standard Chartered'),
]
_IFSC_RE = re.compile(r'\b([A-Z]{4})0[A-Z0-9]{6}\b')


def _detect_bank(header_text: str, full_text: str) -> str:
    for m in _IFSC_RE.finditer(header_text or ''):
        if m.group(1) in _IFSC_BANK:
            return _IFSC_BANK[m.group(1)]
    hu = (header_text or '').upper()
    for pat, name in _BANK_NAME_PATTERNS:
        if re.search(pat, hu):
            return name
    codes = [m.group(1) for m in _IFSC_RE.finditer(full_text or '') if m.group(1) in _IFSC_BANK]
    if codes:
        from collections import Counter
        return _IFSC_BANK[Counter(codes).most_common(1)[0][0]]
    fu = (full_text or '').upper()
    for pat, name in _BANK_NAME_PATTERNS:
        if re.search(pat, fu):
            return name
    return ''


# ──────────────────────────────────────────────────────────────────
# Metadata
# ──────────────────────────────────────────────────────────────────
def _extract_metadata(pdf, header_text: str, all_text: str) -> dict:
    meta = {
        "bank_name": "", "account_no": "", "account_name": "",
        "period_from": "", "period_to": "",
        "opening_balance": None, "pdf_total_debit": None,
        "pdf_total_credit": None, "closing_balance": None,
        "dr_count": None, "cr_count": None,
    }
    try:
        meta["bank_name"] = _detect_bank(header_text, all_text)

        m = re.search(r"Account\s*(?:No|Number)[.:\s]*([0-9\sXx]{6,25})", all_text, re.IGNORECASE)
        if m:
            meta["account_no"] = re.sub(r"[^0-9X]", "", m.group(1))[-4:]

        m = re.search(
            r"(?:Statement\s+From|Period[:\s]*From|From)\s*[:\s]*"
            r"(\d{1,2}[/\-\.][A-Za-z0-9]{1,4}[/\-\.]\d{2,4})"
            r"\s*(?:[Tt]o|-)\s*[:\s]*(\d{1,2}[/\-\.][A-Za-z0-9]{1,4}[/\-\.]\d{2,4})",
            all_text, re.IGNORECASE,
        )
        if m:
            meta["period_from"], meta["period_to"] = m.group(1), m.group(2)

        for pat in [
            r"M[/.]?S[./]?\s+([A-Z][A-Z\s&\-\.]{5,80}?)(?:\n|\r|$)",
            r"(?:Account\s+Name|Name)\s*[:\s]+([A-Z][A-Z\s&\-\.]{5,60}?)(?:\n|\r|$)",
        ]:
            m2 = re.search(pat, all_text)
            if m2:
                name = re.sub(r"\s+", " ", m2.group(1)).strip().rstrip(".,")
                if len(name) >= 4:
                    meta["account_name"] = name[:80]
                    break

        _A = r"[\d,]+\.\d{2}"
        _N = r"\d+"
        sm = re.search(
            r"(" + _A + r")\s+(" + _N + r")\s+(" + _N + r")\s+(" + _A + r")\s+(" + _A + r")\s+(" + _A + r")",
            all_text,
        )
        if sm:
            def _f(s): return float(s.replace(",", ""))
            meta["opening_balance"]  = _f(sm.group(1))
            meta["dr_count"]         = int(sm.group(2))
            meta["cr_count"]         = int(sm.group(3))
            meta["pdf_total_debit"]  = _f(sm.group(4))
            meta["pdf_total_credit"] = _f(sm.group(5))
            meta["closing_balance"]  = _f(sm.group(6))
    except Exception as e:
        logger.warning("Metadata error: %s", e)
    return meta


# ──────────────────────────────────────────────────────────────────
# Core extraction
# ──────────────────────────────────────────────────────────────────
def _finalize(transactions: list[dict], columns: list[dict], meta: dict) -> dict:
    """Compute validation (PDF totals + balance continuity + Golden-Rule net) and
    assemble the public result dict. Shared by the deterministic and LLM paths."""
    computed_debit  = round(sum(t["debit"]  for t in transactions), 2)
    computed_credit = round(sum(t["credit"] for t in transactions), 2)
    pdf_d = meta.get("pdf_total_debit")
    pdf_c = meta.get("pdf_total_credit")

    totals_found = pdf_d is not None and pdf_c is not None
    debit_match  = totals_found and abs(computed_debit  - pdf_d) < 2.0
    credit_match = totals_found and abs(computed_credit - pdf_c) < 2.0
    pdf_totals_ok = totals_found and debit_match and credit_match

    has_bal = _tag_key(columns, 'balance') is not None
    checked = mismatches = 0
    prev_bal = None
    if has_bal:
        for t in transactions:
            b = t.get("balance")
            if prev_bal is not None:
                checked += 1
                expected = round(prev_bal + t["credit"] - t["debit"], 2)
                if abs(expected - b) > 1.0:
                    mismatches += 1
            prev_bal = b
    continuity_ok = has_bal and checked >= max(1, len(transactions) // 2) and mismatches == 0

    # Reverse-chronological statements (newest row first — e.g. Bank of Baroda):
    # forward continuity fails, but each row's balance equals the NEXT (older) row's
    # balance plus this row's credit minus debit. Check that when forward fails.
    rev_checked = rev_mismatch = 0
    if has_bal and not continuity_ok and len(transactions) >= 2:
        for i in range(len(transactions) - 1):
            b_i = transactions[i].get("balance")
            b_next = transactions[i + 1].get("balance")
            if b_i is None or b_next is None:
                continue
            rev_checked += 1
            expected_next = round(b_i - transactions[i]["credit"] + transactions[i]["debit"], 2)
            if abs(expected_next - b_next) > 1.0:
                rev_mismatch += 1
    reverse_ok = has_bal and rev_checked > 0 and rev_checked >= max(1, (len(transactions) - 1) // 2) and rev_mismatch == 0

    opening_balance = meta.get("opening_balance")
    if opening_balance is None and transactions and has_bal:
        first = transactions[0]
        opening_balance = round(first["balance"] - (first["credit"] - first["debit"]), 2)
    closing_balance = meta.get("closing_balance")
    if closing_balance is None and transactions and has_bal:
        closing_balance = transactions[-1]["balance"]

    net_ok = (opening_balance is not None and closing_balance is not None and
              abs(round(opening_balance + computed_credit - computed_debit, 2) - closing_balance) < 2.0)

    balance_reconciled = bool(continuity_ok or net_ok or reverse_ok)
    # Credit-card / single-amount statements have no running balance to reconcile;
    # each row's direction is explicit (DR/CR flag or an Amount column), so the
    # parse is unambiguous — accept it when rows were extracted.
    has_amount = _tag_key(columns, 'amount') is not None or _tag_key(columns, 'drcr') is not None
    drcr_ok = has_amount and not has_bal and len(transactions) > 0
    verified = pdf_totals_ok or balance_reconciled or drcr_ok
    verify_method = ("pdf_totals" if pdf_totals_ok else
                     ("balance" if balance_reconciled else
                      ("drcr_flag" if drcr_ok else "none")))

    validation = {
        "pdf_total_debit":       pdf_d,
        "pdf_total_credit":      pdf_c,
        "computed_total_debit":  computed_debit,
        "computed_total_credit": computed_credit,
        "debit_match":           debit_match,
        "credit_match":          credit_match,
        "verified":              verified,
        "verify_method":         verify_method,
        "balance_reconciled":    balance_reconciled,
        "balance_rows_checked":  checked,
        "balance_mismatches":    mismatches,
        "totals_found_in_pdf":   totals_found,
        "opening_balance":       opening_balance,
        "closing_balance":       closing_balance,
        "dr_count":              meta.get("dr_count"),
        "cr_count":              meta.get("cr_count"),
    }

    preview = [{
        "date": t["date"], "description": t["description"], "ref_no": t["ref_no"],
        "debit": t["debit"], "credit": t["credit"], "balance": t["balance"],
    } for t in transactions[:10]]

    return {
        "bank_name":         meta.get("bank_name", ""),
        "account_no":        meta.get("account_no", ""),
        "account_name":      meta.get("account_name", ""),
        "period_from":       meta.get("period_from", ""),
        "period_to":         meta.get("period_to", ""),
        "columns":           [{"key": c["key"], "header": c["header"], "tag": c["tag"]} for c in columns],
        "_columns_full":     columns,
        "transaction_count": len(transactions),
        "transactions":      transactions,
        "validation":        validation,
        "preview_rows":      preview,
    }


# ──────────────────────────────────────────────────────────────────
# Claude (Anthropic) fallback — fires ONLY when deterministic extraction
# fails (no table detected, zero rows, or balance doesn't reconcile). Rare
# path. Raw HTTPS (urllib) — no SDK dependency. Key from env or new-backend/.env.
# ──────────────────────────────────────────────────────────────────
_LLM_COLS = [("c0", "Txn Date", "date"), ("c1", "Description", "narr"),
             ("c2", "Chq./Ref.No.", "ref"), ("c3", "Debit", "debit"),
             ("c4", "Credit", "credit"), ("c5", "Balance", "balance")]


def _anthropic_key() -> Optional[str]:
    import os
    k = os.environ.get("ANTHROPIC_API_KEY")
    if k:
        return k.strip()
    here = os.path.dirname(os.path.abspath(__file__))
    for cand in (os.path.join(here, "..", "..", "new-backend", ".env"),
                 os.path.join(here, "..", "..", "..", "new-backend", ".env")):
        try:
            with open(cand, encoding="utf-8") as fh:
                for line in fh:
                    line = line.strip()
                    if line.startswith("ANTHROPIC_API_KEY="):
                        return line.split("=", 1)[1].strip()
        except Exception:
            continue
    return None


_EXTRACT_TOOL = {
    "name": "emit_statement",
    "description": "Return the statement's transaction table: the ORIGINAL column headers (verbatim, "
                   "never renamed) with each column's semantic tag, and every transaction row as an "
                   "array of cell values in the SAME order as the headers.",
    "input_schema": {
        "type": "object",
        "properties": {
            "columns": {
                "type": "array",
                "items": {
                    "type": "object",
                    "properties": {
                        "header": {"type": "string", "description": "column header exactly as printed"},
                        "tag": {"type": "string", "enum": sorted(format_learn._VALID_TAGS)},
                    },
                    "required": ["header", "tag"],
                },
            },
            "rows": {
                "type": "array",
                "items": {"type": "array", "items": {"type": "string"}},
            },
        },
        "required": ["columns", "rows"],
    },
}


def _llm_extract(pdf_bytes: bytes, all_text: str, meta: dict) -> Optional[dict]:
    """Tier 3 — full Claude extraction (small PDFs only). Claude returns the
    statement's ORIGINAL column headers + tags + rows; we build the output from
    those real headers so names are preserved (never forced to Debit/Credit)."""
    key = _anthropic_key()
    if not key or not (all_text or "").strip():
        return None
    import os
    import ssl
    import urllib.request
    import json as _json

    model = os.environ.get("ANTHROPIC_MODEL", "claude-haiku-4-5")
    MAXCHARS = 60000
    truncated = len(all_text) > MAXCHARS
    text = all_text[:MAXCHARS]
    prompt = (
        "Extract the transaction table from this bank / credit-card statement.\n"
        "Return, via the tool: (1) the ORIGINAL column headers exactly as printed — do NOT rename "
        "them (keep 'Money out', 'Paid', 'Withdrawal', etc. as-is); (2) each column's semantic tag "
        "(date, narr=description, ref, debit=money out, credit=money in, amount=single signed amount, "
        "drcr=DR/CR flag, balance, other); (3) every transaction row as an array of cell values in the "
        "SAME order as the headers (empty string for blank cells). Skip header rows, sub-totals, and "
        "opening/closing summary lines that are not real transactions.\n\n"
        + ("[NOTE: statement text was truncated to fit]\n\n" if truncated else "")
        + "STATEMENT TEXT:\n" + text
    )
    body = _json.dumps({
        "model": model, "max_tokens": 16000,
        "tools": [_EXTRACT_TOOL], "tool_choice": {"type": "tool", "name": "emit_statement"},
        "messages": [{"role": "user", "content": prompt}],
    }).encode("utf-8")
    try:
        import certifi
        _ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        _ctx = ssl.create_default_context()
    try:
        req = urllib.request.Request("https://api.anthropic.com/v1/messages", data=body, headers={
            "content-type": "application/json", "x-api-key": key, "anthropic-version": "2023-06-01",
        })
        with urllib.request.urlopen(req, timeout=120, context=_ctx) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
    except Exception as e:
        logger.warning("Claude fallback call failed: %s: %s", type(e).__name__, str(e)[:160])
        return None
    if data.get("stop_reason") == "refusal":
        return None
    inp = None
    for b in data.get("content", []):
        if b.get("type") == "tool_use" and b.get("name") == "emit_statement":
            inp = b.get("input") or {}
            break
    if not inp:
        return None
    raw_cols = inp.get("columns") or []
    rows = inp.get("rows") or []
    if not raw_cols or not rows:
        return None

    # Build columns from the ORIGINAL headers + tags (unique tags claimed once).
    columns = []
    claimed = set()
    for i, rc in enumerate(raw_cols):
        tag = rc.get("tag") if rc.get("tag") in _APPLY_TAGS else "other"
        if tag in _UNIQUE_TAGS and tag in claimed:
            tag = "other"
        if tag in _UNIQUE_TAGS:
            claimed.add(tag)
        columns.append({"key": f"c{i}", "header": (rc.get("header") or f"Column {i + 1}").strip(),
                        "tag": tag, "x0": 0.0, "x1": 0.0, "is_primary_date": False, "right_x": None})
    date_cols = [c for c in columns if c["tag"] == "date"]
    if date_cols:
        date_cols[0]["is_primary_date"] = True

    def kof(t):
        return next((c["key"] for c in columns if c["tag"] == t), None)
    dk, ck, ak, fk = kof("debit"), kof("credit"), kof("amount"), kof("drcr")
    bk, dtk, nk, rk = kof("balance"), _primary_date_key(columns), kof("narr"), kof("ref")

    txns = []
    for row in rows:
        vals = row if isinstance(row, list) else []
        if not any(str(v).strip() for v in vals):
            continue
        cells = {}
        for i, c in enumerate(columns):
            v = vals[i] if i < len(vals) else ""
            cells[c["key"]] = _parse_amount(v) if c["tag"] in _NUMERIC_TAGS else _clean_text(str(v or ""))
        d = max(0.0, cells.get(dk, 0.0)) if dk else 0.0
        cr = max(0.0, cells.get(ck, 0.0)) if ck else 0.0
        if ak and not (dk or ck):
            amt = abs(cells.get(ak, 0.0))
            flag = str(cells.get(fk, "")).strip().upper() if fk else ""
            cr, d = (amt, 0.0) if flag.startswith("C") else (0.0, amt)
        txns.append({
            "cells": cells,
            "date": str(cells.get(dtk, "")) if dtk else "",
            "debit": d, "credit": cr,
            "balance": cells.get(bk, 0.0) if bk else 0.0,
            "description": str(cells.get(nk, "")) if nk else "",
            "ref_no": str(cells.get(rk, "")) if rk else "",
        })
    if not txns:
        return None
    res = _finalize(txns, columns, meta)
    v = res["validation"]
    v["verify_method"] = (v["verify_method"] + "+llm") if v["verified"] else "llm"
    v["llm_used"] = True
    # Internal hints so the caller can cache this format (Tier-3 → next-time Tier 1).
    res["_llm_columns"] = [{"header": c["header"], "tag": c["tag"]} for c in columns]
    res["_amount_mode"] = _infer_amount_mode(columns)
    res["_primary_date_header"] = next((c["header"] for c in columns if c.get("is_primary_date")), "")
    logger.info("Claude fallback extracted %d txns, %d cols (verified=%s)",
                len(txns), len(columns), v["verified"])
    return res


# ──────────────────────────────────────────────────────────────────
# Format-learning integration — template cache + one-shot LLM re-tag
# ──────────────────────────────────────────────────────────────────
_APPLY_TAGS = {'date', 'narr', 'ref', 'debit', 'credit', 'amount', 'drcr', 'balance', 'other'}


def _headers_of(columns: list) -> list:
    return [c.get('header', '') for c in columns]


def _infer_amount_mode(columns: list) -> str:
    tags = {c['tag'] for c in columns}
    if 'debit' in tags or 'credit' in tags:
        return 'debit_credit'
    if 'amount' in tags and 'drcr' in tags:
        return 'amount_drcr_flag'
    if 'amount' in tags:
        return 'amount_only'
    return 'debit_credit'


def _apply_template_tags(columns: list, tmpl: dict) -> None:
    """Override deterministic column tags with a cached/learned template (matched
    by normalised header). Geometry/x-ranges are untouched — only meaning changes."""
    tmap = {format_learn.norm(c.get('header')): c.get('tag') for c in tmpl.get('columns', [])}
    pdh = format_learn.norm(tmpl.get('primary_date_header', ''))
    claimed: set = set()
    for c in columns:
        t = tmap.get(format_learn.norm(c.get('header')))
        if t in _APPLY_TAGS:
            if t in _UNIQUE_TAGS and t in claimed:
                c['tag'] = 'other'
            else:
                c['tag'] = t
                if t in _UNIQUE_TAGS:
                    claimed.add(t)
    date_cols = [c for c in columns if c['tag'] == 'date']
    for c in columns:
        c['is_primary_date'] = bool(pdh and format_learn.norm(c.get('header')) == pdh)
    if date_cols and not any(c.get('is_primary_date') for c in columns):
        date_cols[0]['is_primary_date'] = True


def _sample_rows_for_llm(all_lines: list, columns: list, n: int = 12) -> list:
    keys = [c['key'] for c in columns]
    out = []
    for ln in all_lines:
        vals = [ln['cells'].get(k, '') for k in keys]
        if sum(1 for v in vals if str(v).strip()) >= 2:
            out.append(vals)
        if len(out) >= n:
            break
    return out


def _save_success_template(sig: str, bank: str, columns: list) -> None:
    format_learn.save_template(sig, {
        "signature": sig, "bank": bank, "source": "deterministic",
        "amount_mode": _infer_amount_mode(columns),
        "primary_date_header": next((c['header'] for c in columns if c.get('is_primary_date')), ""),
        "columns": [{"header": c['header'], "tag": c['tag']} for c in columns],
    })


# ──────────────────────────────────────────────────────────────────
# Statement gate — this agent processes ONLY bank & credit-card
# statements. Payment advices, invoices, POs, payslips, remittance
# advices etc. are rejected up front (no LLM call) so we never
# "entertain" any other document type. Real statements of ANY format
# still pass easily (a statement keyword OR a genuine multi-row
# date+amount transaction table is enough).
# ──────────────────────────────────────────────────────────────────
_STMT_POS_KW = (
    "account statement", "statement of account", "bank statement", "e-statement",
    "credit card statement", "card statement", "statement period", "statement date",
    "statement of transactions", "transaction details", "account summary",
    "ifsc", "micr", "opening balance", "closing balance", "available balance",
    "withdrawal", "deposit", "credit limit", "minimum amount due", "payment due date",
    "total amount due", "total dues", "account number", "account no", "a/c no",
)
_STMT_NEG_KW = (
    "payment advice", "remittance advice", "settlement advice", "remittance information",
    "tax invoice", "proforma invoice", "commercial invoice", "purchase order",
    "delivery challan", "goods receipt", "e-way bill", "eway bill",
    "salary slip", "payslip", "pay slip", "amount in words", "payment ref no",
    "payment reference", "type of document", "payment amt", "payment doc", "self billing",
)
_STMT_DATE_RE = re.compile(
    r'\b(\d{1,2}[/-][A-Za-z0-9]{2,3}[/-]\d{2,4}|\d{1,2}\s+[A-Za-z]{3,9}\.?\s+\d{2,4}|\d{4}-\d{2}-\d{2})\b')
_STMT_AMT_RE = re.compile(r'\d[\d,]*\.\d{2}\b')


def _statement_signals(text: str):
    tl = (text or "").lower()
    pos = sum(1 for k in _STMT_POS_KW if k in tl)
    neg = sum(1 for k in _STMT_NEG_KW if k in tl)
    rows = 0
    for ln in (text or "").splitlines():
        if _STMT_DATE_RE.search(ln) and _STMT_AMT_RE.search(ln):
            rows += 1
    return pos, neg, rows


def _looks_like_statement(text: str) -> bool:
    """True only for bank / credit-card statements. Generous to real statements
    (any statement keyword, or a genuine ≥6-row date+amount transaction table,
    passes), strict against other document types (payment advice / invoice / PO)."""
    pos, neg, rows = _statement_signals(text)
    if rows >= 6 and neg == 0:          # a real transaction table, no competing doc marker
        return True
    return pos >= 1 and pos >= neg       # else require an explicit statement indicator


def _reject_scanned(too_large: bool = False) -> dict:
    """Scanned / image-only PDF (no text layer) where OCR was unavailable, failed, or
    the scan was too large to OCR safely on the server."""
    msg = ("This scanned PDF is too large for the server to OCR right now (limit ~40 MB / 50 pages). "
           "Please upload a smaller / lower-resolution scan, or a text-based digital statement."
           if too_large else
           "This looks like a scanned / image-only PDF (no text layer) and automatic OCR could not "
           "read it. Please upload a text-based (digital) statement, or a clearer scan.")
    return {
        "bank_name": "", "account_no": "", "account_name": "",
        "period_from": "", "period_to": "", "columns": [],
        "transaction_count": 0, "transactions": [],
        "validation": {"verified": False, "verify_method": "scanned", "error": msg},
        "preview_rows": [],
    }


def _reject_unreadable() -> dict:
    """Graceful message for an encrypted / password-protected / corrupt PDF that
    pdfplumber cannot open — never a 500. (iLovePDF auto-unlock / a UI password
    field will handle these once wired; until then the accountant gets a clear ask.)"""
    return {
        "bank_name": "", "account_no": "", "account_name": "",
        "period_from": "", "period_to": "", "columns": [],
        "transaction_count": 0, "transactions": [],
        "validation": {"verified": False, "verify_method": "unreadable",
                       "error": "This PDF could not be opened — it appears to be password-protected "
                                "or corrupted. Please remove the password / unlock the PDF and upload it again."},
        "preview_rows": [],
    }


def _reject_not_statement() -> dict:
    """Graceful rejection for non-statement uploads (never a 500)."""
    return {
        "bank_name": "", "account_no": "", "account_name": "",
        "period_from": "", "period_to": "", "columns": [],
        "transaction_count": 0, "transactions": [],
        "validation": {"verified": False, "verify_method": "rejected",
                       "error": "This file is not a bank or credit-card statement. The "
                                "PDF → Bank Statement agent only processes bank and credit-card "
                                "statements (any format). Please upload a bank or credit-card statement."},
        "preview_rows": [],
    }


def _cache_llm_template(sig, bank: str, res: dict) -> None:
    """Strip the internal `_llm_*` hints off an LLM result and — when a
    deterministic layout signature is available and the extraction VERIFIED —
    persist the column mapping as a template. Next time the same format lands,
    Tier 1.5 applies these tags and the deterministic path handles it for free."""
    cols = res.pop("_llm_columns", None)
    amode = res.pop("_amount_mode", None)
    pdh = res.pop("_primary_date_header", "")
    if not sig or not cols or not res.get("validation", {}).get("verified"):
        return
    ok = format_learn.save_template(sig, {
        "signature": sig, "bank": bank, "source": "llm-fallback",
        "amount_mode": amode, "primary_date_header": pdh, "columns": cols,
    })
    if ok:
        logger.info("Cached Tier-3 format for reuse (sig=%s, %d cols)", sig, len(cols))


def _llm_extract_safe(pdf_bytes: bytes, all_text: str, meta: dict) -> Optional[dict]:
    """Tier-3 must NEVER crash the request. `_llm_extract` guards only the HTTP
    call; the response-parsing / row-building / _finalize code after it is not.
    A valid 200 with an unexpected columns/rows shape would otherwise raise and
    surface as a 500 to the accountant. Any failure here → None → the caller
    falls back to a graceful 0-row result instead of an error."""
    try:
        return _llm_extract(pdf_bytes, all_text, meta)
    except Exception as e:
        logger.warning("Tier-3 _llm_extract crashed, degrading gracefully: %s: %s",
                       type(e).__name__, str(e)[:160])
        return None


def _infer_columns_headerless(raw_lines: list, page_w: float) -> Optional[list[dict]]:
    """Infer a column layout for a statement that prints NO column-header row.
    Uses the DATA geometry: rows that start with a date are anchors; right-aligned
    money tokens cluster into the amount columns (rightmost = balance); the left
    date token(s) and the wide middle text become Date/Value Date/Narration.
    Returns columns (same shape as _detect_columns) or None. The caller accepts the
    result only if the running balance reconciles, so a wrong guess is discarded."""
    import statistics
    anchors = []
    for ln in raw_lines:
        ws = sorted(ln['words'], key=lambda w: w['x0'])
        if ws and _leading_date(ws[0]['text'])[0]:
            anchors.append(ws)
    if len(anchors) < 5:
        return None
    rights = sorted(_right_edge_columns([{'words': ws} for ws in anchors], 0.0, page_w, tol=6.0))
    if len(rights) < 2:
        return None
    amt_left = max(0.0, 2 * rights[0] - rights[1])   # left edge of the amount block

    # Leading date column(s): first token is a date; an optional 2nd date = value date.
    d1x1 = statistics.median([ws[0]['x1'] for ws in anchors])
    d2 = [ws[1] for ws in anchors if len(ws) > 1 and _leading_date(ws[1]['text'])[0] and ws[1]['x1'] < amt_left]
    has_vdate = len(d2) >= 0.6 * len(anchors)

    cols: list[dict] = []
    def add(header, tag, x0, x1, primary=False, right_x=None):
        cols.append({'key': f'c{len(cols)}', 'header': header, 'tag': tag,
                     'x0': float(x0), 'x1': float(x1), 'is_primary_date': primary, 'right_x': right_x})

    if has_vdate:
        d2x0 = statistics.median([w['x0'] for w in d2]); d2x1 = statistics.median([w['x1'] for w in d2])
        date_x1 = (d1x1 + d2x0) / 2
        add('Date', 'date', 0.0, date_x1, primary=True)
        add('Value Date', 'date', date_x1, d2x1 + 3)
        narr_x0 = d2x1 + 3
    else:
        date_x1 = d1x1 + 3
        add('Date', 'date', 0.0, date_x1, primary=True)
        narr_x0 = date_x1
    if narr_x0 >= amt_left - 5:
        return None                                   # no room for a description column
    add('Narration', 'narr', narr_x0, amt_left)

    # Amount columns: rightmost = balance; 3 → Withdrawal/Deposit/Balance; 2 → Amount/Balance.
    n = len(rights)
    tags = ['other'] * n
    tags[-1] = 'balance'
    hdrs = {'balance': 'Balance'}
    if n >= 3:
        tags[-3], tags[-2] = 'debit', 'credit'
        hdrs.update({'debit': 'Withdrawal', 'credit': 'Deposit'})
    elif n == 2:
        tags[-2] = 'debit'
        hdrs['debit'] = 'Amount'
    for i, R in enumerate(rights):
        x0 = amt_left if i == 0 else (rights[i - 1] + R) / 2
        x1 = page_w if i == n - 1 else (R + rights[i + 1]) / 2
        add(hdrs.get(tags[i], f'Amount {i + 1}'), tags[i], x0, x1, right_x=R)
    return cols


_RUPEE = "₹"


def _rb_is_money(w) -> bool:
    return _looks_amount((w.get('text') or '').replace(_RUPEE, '').strip(_OCR_SEP_CHARS + " "))


def _rb_find_date(words):
    for w in words[:5]:
        if _parse_date(w['text']):
            return w
    return None


def _extract_rupee_balance_format(raw_lines: list[dict], meta: dict,
                                  columns: Optional[list[dict]] = None) -> Optional[dict]:
    """Fallback for the 'rupee-balance line' layout (e.g. Bank of India): the running
    balance is printed on a SEPARATE ₹-prefixed line just BELOW each dated anchor, and
    the transaction amount is a single money token whose column X drifts row-to-row, so
    the geometry pipeline mis-buckets it. Re-derive line-by-line:
      • amount   = the anchor line's (non-₹) money token
      • balance  = the following ₹-line's money token
      • direction= /DR/·/CR/ (or DR/CR/CREDIT/DEBIT) in the narration, else balance delta
    Only accepted if it VERIFIES (balance continuity / reverse-chrono / net) — a wrong
    guess can't self-verify, so this can never silently corrupt output. Returns None when
    the layout doesn't match or doesn't verify."""
    # ── classify lines ──
    L = []
    for ln in raw_lines:
        ws = [w for w in ln['words'] if (w.get('text') or '').strip()]
        if not ws:
            continue
        L.append({
            'y': ln['y'], 'ws': sorted(ws, key=lambda w: w['x0']),
            'date': _rb_find_date(ws),
            'monies': [w for w in ws if _rb_is_money(w)],
            'rupee': any(_RUPEE in w['text'] for w in ws),
        })
    anchors = [x for x in L if x['date'] and x['monies']]
    bal_lines = [x for x in L if x['rupee'] and x['monies'] and not x['date'] and len(x['ws']) <= 3]
    # Signature guard: needs enough anchors AND a ₹-balance line under a good share of them.
    if len(anchors) < 5 or len(bal_lines) < len(anchors) * 0.3:
        return None

    import re as _re
    dr_re = _re.compile(r'(?:/DR/|\bDR\b|DEBIT|WITHDRAW)', _re.I)
    cr_re = _re.compile(r'(?:/CR/|\bCR\b|CREDIT|INW)', _re.I)

    def serial_of(ws, date_w):
        for w in ws:
            if w['x1'] <= date_w['x0'] and _re.fullmatch(r'\d{1,5}', w['text']):
                return w['text']
        return ''

    txns_raw = []
    n = len(L)
    i = 0
    while i < n:
        cur = L[i]
        if not (cur['date'] and cur['monies']):
            i += 1
            continue
        date_w = cur['date']
        amt_tokens = [w for w in cur['monies'] if _RUPEE not in w['text']]
        amount = _parse_amount(amt_tokens[-1]['text']) if amt_tokens else 0.0
        amt_w = amt_tokens[-1] if amt_tokens else None
        narr = [w['text'] for w in cur['ws']
                if w is not date_w and w is not amt_w and w['x0'] >= date_w['x0'] and not _rb_is_money(w)]
        serial = serial_of(cur['ws'], date_w)
        balance = None
        j = i + 1
        while j < n and not (L[j]['date'] and L[j]['monies']):
            nxt = L[j]
            if balance is None and nxt['rupee'] and nxt['monies']:
                balance = _parse_amount(nxt['monies'][-1]['text'])
            else:
                narr += [w['text'] for w in nxt['ws'] if _RUPEE not in w['text'] and not _rb_is_money(w)]
            j += 1
        txns_raw.append({'serial': serial, 'date_raw': date_w['text'],
                         'date': _parse_date(date_w['text']) or date_w['text'],
                         'amount': amount, 'balance': balance,
                         'narr': _clean_text(" ".join(narr)).strip()})
        i = j

    if len(txns_raw) < 5:
        return None

    # ── direction: balance MOVEMENT is ground truth (reverse-chrono: this row is newer
    # than the next). Where the running balance is available, derive debit/credit from
    # the delta and self-correct the amount to |delta| (guards a mis-read amount token).
    # Fall back to the /DR//CR/ narration flag only when there's no balance to compare. ──
    for k, t in enumerate(txns_raw):
        d = ''
        if t['balance'] is not None:
            nb = next((txns_raw[m]['balance'] for m in range(k + 1, len(txns_raw))
                       if txns_raw[m]['balance'] is not None), None)
            if nb is not None:
                delta = round(t['balance'] - nb, 2)
                d = 'C' if delta >= 0 else 'D'
                if abs(abs(delta) - t['amount']) > 1.0:   # trust the balance-derived amount
                    t['amount'] = abs(delta)
        if not d:
            d = 'C' if cr_re.search(t['narr']) else ('D' if dr_re.search(t['narr']) else 'D')
        t['dir'] = d

    cols = columns or [
        {'key': 'c0', 'header': 'Sr No', 'tag': 'other'},
        {'key': 'c1', 'header': 'Date', 'tag': 'date'},
        {'key': 'c2', 'header': 'Remarks', 'tag': 'narr'},
        {'key': 'c3', 'header': 'Debit', 'tag': 'debit'},
        {'key': 'c4', 'header': 'Credit', 'tag': 'credit'},
        {'key': 'c5', 'header': 'Balance', 'tag': 'balance'},
    ]
    sk = _tag_key(cols, 'other') or 'c0'
    dk = _tag_key(cols, 'date') or 'c1'
    nk = _tag_key(cols, 'narr') or 'c2'
    debk = _tag_key(cols, 'debit') or 'c3'
    crk = _tag_key(cols, 'credit') or 'c4'
    bk = _tag_key(cols, 'balance') or 'c5'

    txns = []
    for t in txns_raw:
        debit = t['amount'] if t['dir'] == 'D' else 0.0
        credit = t['amount'] if t['dir'] == 'C' else 0.0
        bal = t['balance'] if t['balance'] is not None else 0.0
        txns.append({
            'cells': {sk: t['serial'], dk: t['date_raw'], nk: t['narr'],
                      debk: debit, crk: credit, bk: bal},
            'date': t['date'], 'debit': debit, 'credit': credit, 'balance': bal,
            'description': t['narr'], 'ref_no': '',
        })

    res = _finalize(txns, cols, meta)
    if not res['validation'].get('verified'):
        return None
    res['validation']['verify_method'] += '+rupeebal'
    return res


def extract_bank_statement(pdf_bytes: bytes, password: str = "", _allow_ocr: bool = True) -> dict:
    import pdfplumber

    columns: Optional[list[dict]] = None
    all_lines: list[dict] = []       # {y, cells} in reading order (page-offset applied)
    header_found_page = None

    # Guarded open — an encrypted / password-protected / corrupt PDF must degrade
    # to a clear message, never a 500. `password` (from the UI field, when provided)
    # is passed straight to pdfplumber so open-password statements can be read.
    open_kwargs = {"password": password} if password else {}
    try:
        _pdf = pdfplumber.open(BytesIO(pdf_bytes), **open_kwargs)
    except Exception as e:
        logger.warning("PDF open failed (encrypted/corrupt?): %s: %s", type(e).__name__, str(e)[:160])
        return _reject_unreadable()
    with _pdf as pdf:
        n_pages = len(pdf.pages)
        # Extract text page-by-page and flush each page's cache — bounds memory on large
        # / image-heavy (OCR'd) PDFs so pdfplumber never holds all pages at once (a 30-page
        # OCR'd scan otherwise balloons to >1 GB and OOMs the small shared box).
        _parts = []
        for pg in pdf.pages:
            _parts.append(pg.extract_text() or "")
            try:
                pg.flush_cache()
            except Exception:
                pass
        all_text = "\n".join(_parts)
        # No-text-layer PDF (scanned, or fonts flattened to vector outlines): pdfplumber
        # sees no words. OCR it ONCE via iLovePDF (fallback only — gated to this case so
        # normal statements never incur OCR cost) and re-extract on the searchable copy.
        if _allow_ocr and not all_text.strip() and \
                sum(len(pg.extract_words()) for pg in pdf.pages[:8]) == 0:
            # Protect the small shared box: OCR (rasterise + re-parse) is memory-heavy,
            # so skip very large scans rather than risk OOM-ing the engine for all agents.
            if len(pdf_bytes) > _MAX_OCR_BYTES or n_pages > _MAX_OCR_PAGES:
                logger.warning("Scanned PDF too large for server OCR (%.1f MB, %d pages) — skipping.",
                               len(pdf_bytes) / 1e6, n_pages)
                return _reject_scanned(too_large=True)
            from recon import ilovepdf_ocr
            logger.info("No text layer in %d-page PDF — attempting iLovePDF OCR fallback.", n_pages)
            ocr_bytes = ilovepdf_ocr.ocr_pdf(pdf_bytes, filename="statement.pdf")
            if ocr_bytes:
                res = extract_bank_statement(ocr_bytes, password="", _allow_ocr=False)
                ocr_bytes = None
                import gc
                gc.collect()
                return res
            return _reject_scanned()
        p0 = pdf.pages[0]
        header_text = " ".join(w['text'] for w in p0.extract_words()
                               if w['top'] < p0.height * 0.40)
        meta = _extract_metadata(pdf, header_text, all_text)

        y_offset = 0.0
        all_raw_lines: list[dict] = []   # every positioned line — for the headerless fallback
        page_w = 0.0
        for page_num, page in enumerate(pdf.pages):
            page_w = max(page_w, float(page.width))
            words = page.extract_words(x_tolerance=1.5, y_tolerance=3, keep_blank_chars=False)
            if not words:
                y_offset += float(page.height)
                continue
            lines = _group_lines(words)
            for ln in lines:
                all_raw_lines.append({'y': y_offset + ln['y'], 'words': ln['words']})

            if columns is None:
                header = _find_header_band(lines)
                if header is None:
                    y_offset += float(page.height)
                    continue
                data_lines = [ln for ln in lines if ln['y'] > header['y1'] + 0.5]
                columns = _detect_columns(page, header, data_lines)
                if columns is None:
                    y_offset += float(page.height)
                    continue
                header_found_page = page_num
                start_lines = data_lines
            else:
                start_lines = lines

            for ln in start_lines:
                all_lines.append({
                    'y': y_offset + ln['y'],
                    'cells': _assign_cells(ln['words'], columns),
                })
            y_offset += float(page.height)
            try:
                page.flush_cache()      # release this page before the next (memory-bounded)
            except Exception:
                pass

    # Headerless-table fallback: some statements (e.g. SBI) print NO column-header row,
    # so _find_header_band never fires. Infer the columns from the DATA geometry and
    # accept ONLY if the running balance reconciles — a wrong guess can't self-verify,
    # so this can't silently corrupt output.
    if columns is None:
        inferred = _infer_columns_headerless(all_raw_lines, page_w)
        if inferred:
            hl_lines = [{'y': ln['y'], 'cells': _assign_cells(ln['words'], inferred)} for ln in all_raw_lines]
            hl_txns = _build_transactions(hl_lines, inferred)
            hl_res = _finalize(hl_txns, inferred, meta)
            if hl_txns and hl_res["validation"].get("verified"):
                logger.info("Headerless inference verified: %d txns, %d cols.", len(hl_txns), len(inferred))
                hl_res["validation"]["verify_method"] += "+headerless"
                return hl_res
            logger.info("Headerless inference did not verify (%d txns) — falling through.",
                        len(hl_txns) if hl_txns else 0)

    if columns is not None:
        bank = meta.get("bank_name", "")
        sig = format_learn.layout_signature(bank, _headers_of(columns))
        cached = format_learn.load_template(sig)
        if cached:
            _apply_template_tags(columns, cached)   # Tier 1.5 — reuse a learned format ($0)

        transactions = _build_transactions(all_lines, columns)
        result = _finalize(transactions, columns, meta)
        logger.info("Extracted %d txns (%d cols) verified=%s (sig=%s cached=%s)",
                    len(transactions), len(columns), result["validation"]["verified"], sig, bool(cached))
        if transactions and result["validation"].get("verified"):
            if not cached:                          # auto-learn a template from a good run
                _save_success_template(sig, bank, columns)
            return result

        # Deterministic geometry didn't verify. Try the 'rupee-balance line' layout
        # (balance on a separate ₹-line, drifting amount column) before any LLM/OCR.
        rb = _extract_rupee_balance_format(all_raw_lines, meta, columns)
        if rb:
            logger.info("Rupee-balance line format verified: %d txns", rb["transaction_count"])
            return rb

        # A table was detected but did not verify. Before spending any LLM call,
        # confirm this is actually a statement (not an invoice / advice that merely
        # has a table). Verified statements never reach here, so this can't reject one.
        if not _looks_like_statement(all_text):
            logger.info("Rejected: detected a table but document is not a bank/CC statement.")
            return _reject_not_statement()

        # Tier 2 — TRY HARD to LEARN the format: one Haiku call re-tags the detected
        # columns from a tiny sample; we re-extract deterministically on a COPY (so the
        # fallback result is never corrupted) and, on success, CACHE the template so next
        # time Tier 1 handles this format for free. Skipped only if we already learned
        # this exact format before (avoids re-calling Haiku for a format that won't verify).
        import copy as _copy
        already_learned = bool(cached and cached.get("source") in ("llm", "llm-fallback", "learned"))
        key = _anthropic_key()
        if key and not already_learned:
            try:
                learned = format_learn.learn_template(
                    bank, _headers_of(columns), _sample_rows_for_llm(all_lines, columns), key)
                if learned:
                    cols2 = _copy.deepcopy(columns)
                    _apply_template_tags(cols2, learned)
                    tx2 = _build_transactions(all_lines, cols2)
                    res2 = _finalize(tx2, cols2, meta)
                    if tx2 and res2["validation"].get("verified"):
                        learned["signature"] = sig
                        format_learn.save_template(sig, learned)
                        res2["validation"]["verify_method"] += "+learned"
                        logger.info("Format learned + cached (sig=%s): %d txns", sig, len(tx2))
                        return res2
            except Exception as e:
                # Never let format-learning crash the request — fall through to Tier 3.
                logger.warning("Tier-2 format-learning failed, falling through: %s: %s",
                               type(e).__name__, str(e)[:160])

        # Tier 3 — full Claude extraction, last resort. Only for small PDFs
        # (≤ _MAX_LLM_EXTRACT_PAGES) so a large statement never blows up tokens/cost.
        # On success we also cache the format so future runs skip straight to Tier 1.
        if n_pages <= _MAX_LLM_EXTRACT_PAGES:
            alt = _llm_extract_safe(pdf_bytes, all_text, meta)
            if alt and (not transactions or alt["validation"].get("verified")):
                _cache_llm_template(sig, bank, alt)
                return alt
        else:
            logger.info("Skipping Tier-3 full extraction: %d pages > %d limit.",
                        n_pages, _MAX_LLM_EXTRACT_PAGES)
        return result

    # No transaction table detected deterministically. First reject anything that
    # isn't a bank/credit-card statement (payment advice, invoice, PO …) — no LLM call.
    if not _looks_like_statement(all_text):
        logger.info("Rejected: not a bank/credit-card statement (no table + no statement signals).")
        return _reject_not_statement()

    # Looks like a statement but the geometry defeated deterministic detection → Tier 3
    # (small PDFs only). No deterministic signature exists here, so it can't be cached
    # into the Tier-1 path — we just strip the internal hints off the result.
    if n_pages <= _MAX_LLM_EXTRACT_PAGES:
        alt = _llm_extract_safe(pdf_bytes, all_text, meta)
        if alt:
            _cache_llm_template(None, "", alt)   # strips internal _llm_* hints (no sig → no cache)
            return alt
        logger.warning("Looks like a statement but no transaction table could be extracted.")
    else:
        logger.warning("No table detected and %d pages > %d — Tier-3 skipped.",
                       n_pages, _MAX_LLM_EXTRACT_PAGES)
    return {
        "bank_name": "", "account_no": "", "account_name": "",
        "period_from": "", "period_to": "", "columns": [],
        "transaction_count": 0, "transactions": [],
        "validation": {"verified": False, "verify_method": "none",
                       "error": "Could not detect a transaction table in the PDF."},
        "preview_rows": [],
    }


# ──────────────────────────────────────────────────────────────────
# Excel builder — reproduces source columns + Check Point 1 / 2 + Summary
# ──────────────────────────────────────────────────────────────────
def build_pdf_bank_excel(data: dict) -> bytes:
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill, Alignment, Border, Side
    from openpyxl.utils import get_column_letter
    from openpyxl.formatting.rule import CellIsRule

    columns = data.get("_columns_full") or []
    txns = data.get("transactions", [])

    wb = Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    thin = Side(border_style="thin", color="D0D5DD")
    BORDER = Border(left=thin, right=thin, top=thin, bottom=thin)
    HDR_FILL = PatternFill("solid", fgColor="0748EE")
    HDR_FONT = Font(bold=True, color="FFFFFF", size=10, name="Calibri")
    DAT_FONT = Font(size=10, name="Calibri")
    ALT_FILL = PatternFill("solid", fgColor="EEF3FF")
    CENTER = Alignment(horizontal="center", vertical="center")
    LEFT = Alignment(horizontal="left", vertical="center", wrap_text=False)
    RIGHT = Alignment(horizontal="right", vertical="center")

    if not columns:
        ws["A1"] = "No transaction table detected in this PDF."
        buf = BytesIO(); wb.save(buf); return buf.getvalue()

    # Column layout: source columns, then Check Point 1 / Check Point 2
    src_headers = [c["header"] for c in columns]
    HEADERS = src_headers + ["Check Point 1", "Check Point 2"]
    ncol = len(HEADERS)
    cp1_idx = len(src_headers) + 1   # 1-based
    cp2_idx = cp1_idx + 1

    # Locate tagged columns (1-based excel indices)
    def col_index(tag):
        for i, c in enumerate(columns):
            if c["tag"] == tag:
                return i + 1
        return None
    debit_i = col_index("debit")
    credit_i = col_index("credit")
    bal_i = col_index("balance")
    amount_i = col_index("amount")

    def _hdr(i):   # original PDF header of column i (1-based) — output never renames columns
        return columns[i - 1]["header"] if i and 0 < i <= len(columns) else ""

    numeric_keys = {c["key"] for c in columns if c["tag"] in _NUMERIC_TAGS}

    # Header row
    for i, h in enumerate(HEADERS, start=1):
        c = ws.cell(row=1, column=i, value=h)
        c.fill, c.font, c.alignment, c.border = HDR_FILL, HDR_FONT, CENTER, BORDER
        letter = get_column_letter(i)
        # width heuristic by tag
        tag = columns[i - 1]["tag"] if i - 1 < len(columns) else "cp"
        ws.column_dimensions[letter].width = 60 if tag == "narr" else (14 if tag == "date" else 18)
    ws.row_dimensions[1].height = 22

    # Data rows
    for r, txn in enumerate(txns, start=2):
        alt = (r % 2 == 0)
        cells = txn.get("cells", {})
        for i, c in enumerate(columns, start=1):
            val = cells.get(c["key"], "")
            is_num = c["key"] in numeric_keys
            if is_num and (val == 0.0 or val == 0):
                val = None
            cell = ws.cell(row=r, column=i, value=val)
            cell.font, cell.border = DAT_FONT, BORDER
            if alt:
                cell.fill = ALT_FILL
            if is_num:
                cell.alignment = RIGHT
                if isinstance(val, (int, float)):
                    cell.number_format = '#,##0.00'
            else:
                cell.alignment = LEFT

        # Check Point 1: prev balance + credit − debit  (live formula when columns exist)
        cp1 = None
        if r > 2 and bal_i and (debit_i or credit_i):
            bal_col = get_column_letter(bal_i)
            cr = f"+{get_column_letter(credit_i)}{r}" if credit_i else ""
            dr = f"-{get_column_letter(debit_i)}{r}" if debit_i else ""
            cp1 = f"={bal_col}{r - 1}{cr}{dr}"
        cc = ws.cell(row=r, column=cp1_idx, value=cp1)
        cc.font, cc.border, cc.alignment = DAT_FONT, BORDER, RIGHT
        if alt: cc.fill = ALT_FILL
        if cp1: cc.number_format = '#,##0.00'

        # Check Point 2: stated balance − Check Point 1  (should be 0)
        cp2 = None
        if r > 2 and bal_i:
            bal_col = get_column_letter(bal_i)
            cp2 = f"={bal_col}{r}-{get_column_letter(cp1_idx)}{r}"
        c2 = ws.cell(row=r, column=cp2_idx, value=cp2)
        c2.font, c2.border, c2.alignment = DAT_FONT, BORDER, RIGHT
        if alt: c2.fill = ALT_FILL
        if cp2: c2.number_format = '#,##0.00'
        ws.row_dimensions[r].height = 15

    n = len(txns)
    last_row = n + 1

    # Conditional formatting on Check Point 2: green = 0, red ≠ 0
    if n > 1 and bal_i:
        rng = f"{get_column_letter(cp2_idx)}3:{get_column_letter(cp2_idx)}{last_row}"
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="between", formula=["-0.01", "0.01"],
            fill=PatternFill("solid", fgColor="C6EFCE")))
        ws.conditional_formatting.add(rng, CellIsRule(
            operator="notBetween", formula=["-0.01", "0.01"],
            fill=PatternFill("solid", fgColor="FFC7CE")))

    # ── Statement Summary block ──────────────────────────────────
    val = data.get("validation", {})
    summ = last_row + 2
    LBL_FILL = PatternFill("solid", fgColor="2C3E50")
    LBL_FONT = Font(bold=True, color="FFFFFF", size=9, name="Calibri")
    VAL_FONT = Font(bold=True, size=10, name="Calibri")
    VAL_FILL = PatternFill("solid", fgColor="EBF5FB")

    def put(row, col, value, fill, font, align, numfmt=None):
        if col < 1:
            return
        c = ws.cell(row=row, column=col, value=value)
        c.fill, c.font, c.alignment, c.border = fill, font, align, BORDER
        if numfmt:
            c.number_format = numfmt

    ws.cell(row=summ, column=1, value="STATEMENT SUMMARY").fill = LBL_FILL
    ws.cell(row=summ, column=1).font = LBL_FONT
    ws.cell(row=summ, column=1).border = BORDER
    # Summary labels use the ORIGINAL PDF column headers (never renamed to Debit/Credit).
    if debit_i:
        put(summ, debit_i, f"Total {_hdr(debit_i)}", LBL_FILL, LBL_FONT, CENTER)
        put(summ + 1, debit_i, f"=SUM({get_column_letter(debit_i)}2:{get_column_letter(debit_i)}{last_row})",
            VAL_FILL, VAL_FONT, RIGHT, '#,##0.00')
    if credit_i:
        put(summ, credit_i, f"Total {_hdr(credit_i)}", LBL_FILL, LBL_FONT, CENTER)
        put(summ + 1, credit_i, f"=SUM({get_column_letter(credit_i)}2:{get_column_letter(credit_i)}{last_row})",
            VAL_FILL, VAL_FONT, RIGHT, '#,##0.00')
    if amount_i and not (debit_i or credit_i):
        amt_key = columns[amount_i - 1]["key"]
        amount_is_text = any(isinstance(t.get("cells", {}).get(amt_key), str) and t["cells"][amt_key].strip()
                             for t in txns)
        put(summ, amount_i, f"Total {_hdr(amount_i)}", LBL_FILL, LBL_FONT, CENTER)
        if amount_is_text:
            # Direction embedded in the amount text (Cr/Dr) — a raw SUM mixes deposits
            # and withdrawals, so show the parsed Deposit(Cr) / Withdrawal(Dr) totals.
            put(summ + 1, amount_i,
                f"Cr {val.get('computed_total_credit', 0):,.2f}  |  Dr {val.get('computed_total_debit', 0):,.2f}",
                VAL_FILL, VAL_FONT, RIGHT)
        else:
            put(summ + 1, amount_i, f"=SUM({get_column_letter(amount_i)}2:{get_column_letter(amount_i)}{last_row})",
                VAL_FILL, VAL_FONT, RIGHT, '#,##0.00')
    if bal_i:
        put(summ, bal_i, f"Closing {_hdr(bal_i)}", LBL_FILL, LBL_FONT, CENTER)
        put(summ + 1, bal_i, val.get("closing_balance"), VAL_FILL, VAL_FONT, RIGHT, '#,##0.00')
    ws.cell(row=summ + 1, column=1, value=f"Opening: {val.get('opening_balance')}").font = VAL_FONT

    # PDF-stated totals comparison (if the statement printed them)
    if val.get("pdf_total_debit") is not None:
        cmp = summ + 2
        CMP_FONT = Font(italic=True, size=9, color="555555", name="Calibri")
        CMP_FILL = PatternFill("solid", fgColor="F8F9FA")
        put(cmp, 1, "As per PDF", CMP_FILL, CMP_FONT, LEFT)
        if debit_i:
            put(cmp, debit_i, val.get("pdf_total_debit"), CMP_FILL, CMP_FONT, RIGHT, '#,##0.00')
        if credit_i:
            put(cmp, credit_i, val.get("pdf_total_credit"), CMP_FILL, CMP_FONT, RIGHT, '#,##0.00')

    ws.freeze_panes = "A2"
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
