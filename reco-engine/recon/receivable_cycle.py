"""Receivable Cycle reconciliation engine.

Fully self-contained — deliberately does NOT import from any other recon/*.py
module (gstr_2b_books, zepto_receivables, etc.). This is a different business
process (order -> invoice -> courier COD remittance -> return/SRN tracking),
not a GST reconciliation, and must be free to evolve without touching or
being touched by the GST/bank/zepto engines.

Pipeline
--------
Base table: "Combine Tally GST report" (one row per invoice line item).
Enriched with:
  - Sales Order Combine  -> channel Order ID + delivery status/time/tracking
  - Delhivery / Ekart / Xpressbees COD settlement exports -> remittance
    amount + settlement status, matched on AWB/tracking number, bucketed
    into a FY24-25 / FY25-26 column pair per the fiscal year of the
    settlement row (Apr-Mar Indian FY)
  - Combined SRN (Unicommerce sales-return / credit-note) report -> return
    tracking, matched on the original invoice number (fallback: AWB),
    bucketed into FY24-25 / FY25-26 columns using the FY embedded in the
    SRN/credit-note's own number (e.g. "SRN/25-26/0001")

Produces two sheet families:
  - "Main Sheet"      -> every Tally GST row, enriched
  - "COD main sheet"  -> Tally GST rows with Payment Method == COD, plus one
                         sub-sheet per courier (Delivery/Ekart/Xpressbees/
                         DTDC/Self shipping), each further enriched with
                         that courier's settlement-match columns. DTDC and
                         Self shipping have no settlement export to join
                         against, so they carry only the Sales-Order-Combine
                         derived columns.
"""
from __future__ import annotations

import datetime as _dt
import re
from io import BytesIO
from typing import Any

import pandas as pd

# --------------------------------------------------------------------------
# Generic helpers (deliberately re-implemented here, not imported, to keep
# this engine fully independent of the other recon/*.py modules)
# --------------------------------------------------------------------------


def _clean(v: Any) -> str:
    if v is None:
        return ""
    if isinstance(v, float) and pd.isna(v):
        return ""
    s = str(v).strip()
    return "" if s.lower() in ("nan", "nat", "none") else s


def _to_float(v: Any) -> float:
    s = _clean(v).replace(",", "").replace("₹", "")
    if s in ("", "-"):
        return 0.0
    try:
        return float(s)
    except ValueError:
        return 0.0


def _norm_header(s: Any) -> str:
    return re.sub(r"\s+", " ", str(s or "")).strip().lower()


def _norm_code(v: Any) -> str:
    """Normalize an AWB / order / invoice code for matching: uppercase,
    whitespace-stripped, and pandas float-ified integers (e.g. AWB numbers
    read as `20407610551725.00`) collapsed back to their integer string."""
    s = _clean(v)
    if not s:
        return ""
    if s.endswith(".0") and s[:-2].replace(".", "", 1).lstrip("-").isdigit():
        s = s[:-2]
    elif re.fullmatch(r"-?\d+\.0+", s):
        s = s.split(".")[0]
    return re.sub(r"\s+", "", s).upper()


def _get(row: dict, *aliases: str) -> Any:
    nk = {_norm_header(k): v for k, v in row.items()}
    for alias in aliases:
        key = _norm_header(alias)
        if key in nk:
            v = nk[key]
            if _clean(v) != "":
                return v
    return ""


def _parse_date(v: Any) -> _dt.date | None:
    if v is None or v == "":
        return None
    if isinstance(v, _dt.datetime):
        return v.date()
    if isinstance(v, _dt.date):
        return v
    if hasattr(v, "to_pydatetime"):   # pandas Timestamp
        try:
            return v.to_pydatetime().date()
        except Exception:
            pass
    if isinstance(v, (int, float)) and not isinstance(v, bool):
        # Excel serial date (1900 date system, with the well-known 1900
        # leap-year bug baked into Excel/pandas' own epoch handling).
        try:
            return (_dt.date(1899, 12, 30) + _dt.timedelta(days=float(v)))
        except Exception:
            return None
    raw = _clean(v)
    if not raw:
        return None
    # ISO-8601 datetime with a "T" separator and/or timezone offset (e.g. Ekart
    # settlement exports' "2025-02-21T17:06:22+05:30") — must be tried on the
    # full string before the space-split below, which would otherwise leave
    # the trailing "T17:06:22+05:30" attached and fail every %Y-%m-%d attempt.
    try:
        return _dt.datetime.fromisoformat(raw.replace("Z", "+00:00")).date()
    except ValueError:
        pass
    s = raw.split(" ")[0]
    if not s:
        return None
    for fmt in ("%m/%d/%y", "%m/%d/%Y", "%Y-%m-%d", "%d/%m/%Y", "%d-%m-%Y", "%m-%d-%y"):
        try:
            return _dt.datetime.strptime(s, fmt).date()
        except ValueError:
            continue
    return None


def _fiscal_year_label(d: _dt.date | None) -> str:
    """Indian FY (Apr-Mar) as e.g. "24-25". Empty string if date unknown."""
    if d is None:
        return ""
    start_year = d.year if d.month >= 4 else d.year - 1
    return f"{start_year % 100:02d}-{(start_year + 1) % 100:02d}"


def _detect_header_row(raw: pd.DataFrame, max_scan: int = 10) -> int:
    """Real-world courier settlement exports routinely carry a leading
    "totals" row (a handful of stray numbers, no labels) above the actual
    header. Pick the row with the most non-blank string cells in the first
    few rows — the header row always wins that count against a totals row
    or a data row."""
    best_row, best_score = 0, -1
    for i in range(min(max_scan, len(raw))):
        score = sum(1 for v in raw.iloc[i] if isinstance(v, str) and v.strip())
        if score > best_score:
            best_row, best_score = i, score
    return best_row


def _mangle_duplicate_columns(columns) -> list:
    """Reproduce pandas' own duplicate-column suffixing (`X`, `X.1`, `X.2`, …)
    since we assign `.columns` by hand below instead of letting a header-row
    read do it for us."""
    seen: dict[str, int] = {}
    out = []
    for col in columns:
        c = _clean(col)
        if c in seen:
            seen[c] += 1
            out.append(f"{c}.{seen[c]}")
        else:
            seen[c] = 0
            out.append(c)
    return out


def _frame_from_headerless(raw: pd.DataFrame) -> pd.DataFrame:
    if raw.empty:
        return raw
    header_row = _detect_header_row(raw)
    df = raw.iloc[header_row + 1:].reset_index(drop=True)
    df.columns = _mangle_duplicate_columns(raw.iloc[header_row])
    return df


def _read_table(data: bytes) -> pd.DataFrame:
    """Reads every sheet of an Excel workbook — courier settlement exports
    routinely split FY-wise data across sheets (e.g. "FY 24-25" +
    "Apr-25 to till Date") and every sheet must be combined, not just the
    first — or the single table of a CSV. Each sheet's header row is
    detected independently before the sheets are stacked."""
    if not data:
        return pd.DataFrame()
    result = None
    for reader, kwargs in (
        (pd.read_excel, {"sheet_name": None}),
        (pd.read_excel, {"sheet_name": None, "engine": "xlrd"}),
        (pd.read_csv, {"encoding": "utf-8-sig"}),
        (pd.read_csv, {"encoding": "latin-1"}),
    ):
        try:
            result = reader(BytesIO(data), dtype=object, header=None, **kwargs)
            break
        except Exception:
            continue
    if result is None:
        raise ValueError("Unrecognized file format — expected .xlsx/.xls/.csv")

    if isinstance(result, dict):
        frames = [f for f in (_frame_from_headerless(sheet) for sheet in result.values()) if not f.empty]
        return pd.concat(frames, ignore_index=True) if frames else pd.DataFrame()
    return _frame_from_headerless(result)


def _rows(df: pd.DataFrame) -> list[dict]:
    if df.empty:
        return []
    df = df.rename(columns=lambda c: _clean(c))
    out = []
    for record in df.to_dict(orient="records"):
        out.append({k: ("" if (v is None or (isinstance(v, float) and pd.isna(v))) else v)
                    for k, v in record.items()})
    return out


# --------------------------------------------------------------------------
# Column contracts (verbatim column sets/order for each output sheet)
# --------------------------------------------------------------------------

# "Channel entry" .. "IRN" — shared verbatim between the Combine Tally GST
# report and every output sheet (Main Sheet + all COD sheets).
TAIL_COLUMNS = [
    "Channel entry", "Channel Ledger", "Product Name", "Product SKU Code", "Qty",
    "Unit Price", "Currency", "conversion rate", "Total", "Customer Name",
    "Shipping Address Name", "Shipping Address State", "Shipping Address Country",
    "Shipping Provider", "AWB num", "Sales", "Sales Ledger", "CGST", "CGST Rate",
    "SGST", "SGST Rate", "IGST", "IGST Rate", "UTGST", "UTGST Rate", "CESS",
    "CESS Rate", "Other charges", "Other charges Ledger", "Other charges1",
    "Other charges Ledger1", "Service tax", "ST Ledger", "Discount Ledger",
    "Discount Amount", "IMEI", "Godown", "Dispatch Date/Cancellation Date",
    "Narration", "Entity", "Voucher Type Name", "TIN NO", "Original Invoice Date",
    "Original Sale No", "Channel Invoice Created", "Channel State", "Customer GSTIN",
    "Channel_Party GSTIN", "Billing Party Code", "Tax Verification",
    "GST Registration Type", "TCS Amount", "Adjustment In Selling Price",
    "Adjustment In Discount", "Other charges Ledger2", "Other charges2",
    "Store Credit", "Prepaid Amount", "Tax On Other Charges",
    "Tax On Other Charges1", "Tax On Other Charges2", "IRN",
]

FRONT_MAIN = ["Date", "Sale Order Number", "Invoice number", "Sale Order Number (Duplicate)",
              "Remitence", "Srn", "Remark", "Order ID"]
MAIN_SHEET_COLUMNS = (
    FRONT_MAIN + TAIL_COLUMNS
    + ["Payment Method", "Courier States",
       "SRN from (Unnicommerce Refund sheet 24-25)", "SRN from (Unnicommerce Refund sheet 25-26)",
       "2", "3", "4"]
)

FRONT_COD = ["Date", "Sale Order Number", "Invoice number", "Sale Order Number (Duplicate)", "Order ID"]
COD_MAIN_COLUMNS = FRONT_COD + TAIL_COLUMNS + ["Payment Method"]

DELIVERY_COLUMNS = (
    FRONT_COD
    + ["Order ID of 24-25 from delhivary setlment report", "Delivery Status from sales order report",
       "Delivery Time from Sales order report", "AWB no in Settlmnt report of Fy 24-25",
       "AWB no in Settlmnt report of Fy 25-26"]
    + TAIL_COLUMNS + ["Payment Method", "SRN Status"]
)
EKART_COLUMNS = (
    FRONT_COD
    + ["Delivery Status from sales order report", "Delivery Time from Sales order report",
       "Tracking ID from Sales Order report", "Tracking ID in Ekart Sttlmt reports Fy 24-25",
       "Tracking ID in Ekart Sttlmt reports Fy 25-26"]
    + TAIL_COLUMNS + ["Payment Method", "SRN Status"]
)
XPRESSBEES_COLUMNS = (
    FRONT_COD
    + ["Delivery Status from sales order report", "Delivery Time from Sales order report",
       "Shipping Id ( AWB) in Xpressbees Sttlmt reports of Fy 24-25",
       "Order ID in Xpressbees Sttlmt reports after Mar 25"]
    + TAIL_COLUMNS + ["Payment Method", "SRN Status"]
)
DTDC_COLUMNS = (
    FRONT_COD
    + ["Delivery Status from sales order report", "Delivery Time from Sales order report"]
    + TAIL_COLUMNS + ["Payment Method", "SRN Status"]
)
SELF_SHIP_COLUMNS = FRONT_COD + TAIL_COLUMNS + ["Payment Method"]

# Couriers the Receivable Amount calculation covers, and — per courier — the
# settlement-match columns that mean "not yet settled" when ALL of them are
# blank on a row. DTDC has no settlement export at all (see module docstring),
# so every DTDC COD row is unconditionally pending.
RECEIVABLE_COURIERS = ["Delivery", "Ekart", "Xpressbees", "DTDC"]
_PENDING_MATCH_COLUMNS = {
    "Delivery": ["AWB no in Settlmnt report of Fy 24-25", "AWB no in Settlmnt report of Fy 25-26"],
    "Ekart": ["Tracking ID in Ekart Sttlmt reports Fy 24-25", "Tracking ID in Ekart Sttlmt reports Fy 25-26"],
    "Xpressbees": ["Shipping Id ( AWB) in Xpressbees Sttlmt reports of Fy 24-25",
                   "Order ID in Xpressbees Sttlmt reports after Mar 25"],
    "DTDC": [],
}

COD_SHEET_ORDER = ["COD main sheet", "Delivery", "Ekart", "Xpressbees", "DTDC", "Self shipping"]
_SHEET_COLUMNS = {
    "COD main sheet": COD_MAIN_COLUMNS,
    "Delivery": DELIVERY_COLUMNS,
    "Ekart": EKART_COLUMNS,
    "Xpressbees": XPRESSBEES_COLUMNS,
    "DTDC": DTDC_COLUMNS,
    "Self shipping": SELF_SHIP_COLUMNS,
}


# --------------------------------------------------------------------------
# Input parsers
# --------------------------------------------------------------------------


def parse_tally_gst(data: bytes) -> list[dict]:
    rows = _rows(_read_table(data))
    out = []
    for row in rows:
        invoice_no = _clean(_get(row, "Invoice number"))
        sale_order = _clean(_get(row, "Sale Order Number"))
        if not invoice_no and not sale_order:
            continue
        row["_awb"] = _norm_code(_get(row, "AWB num"))
        row["_invoice_key"] = _norm_code(invoice_no)
        row["_payment_method"] = _clean(_get(row, "Payment Method")).upper()
        row["_shipping_provider"] = _clean(_get(row, "Shipping Provider")).upper()
        out.append(row)
    return out


def parse_sales_order(data: bytes) -> dict[str, dict[str, dict]]:
    """Item-level file — one row per SKU, first-win per key. Indexed by AWB
    / courier tracking number (matches ~99% of real Tally rows) with the
    channel invoice code as a fallback (~98%) for rows with no AWB yet.
    NOT indexed by "Sale Order Code": that internal Unicommerce sequence
    number does not correlate with Tally's "Sale Order Number" at all
    (different ID schemes — verified against real exports), so it is
    useless as a join key here."""
    by_awb: dict[str, dict] = {}
    by_invoice: dict[str, dict] = {}
    for row in _rows(_read_table(data)):
        entry = {
            "order_id": _clean(_get(row, "Display Order Code")),
            "status": _clean(_get(row, "Sale Order Item Status", "On Hold")),
            "delivery_time": _clean(_get(row, "Delivery Time", "Dispatch Date")),
            "tracking_number": _clean(_get(row, "Tracking Number")),
        }
        awb_key = _norm_code(_get(row, "Tracking Number"))
        if awb_key and awb_key not in by_awb:
            by_awb[awb_key] = entry
        invoice_key = _norm_code(_get(row, "Invoice Code"))
        if invoice_key and invoice_key not in by_invoice:
            by_invoice[invoice_key] = entry
    return {"by_awb": by_awb, "by_invoice": by_invoice}


def _lookup_sales_order(sales_order_idx: dict[str, dict[str, dict]], awb: str, invoice_key: str) -> dict:
    if awb:
        entry = sales_order_idx["by_awb"].get(awb)
        if entry is not None:
            return entry
    if invoice_key:
        entry = sales_order_idx["by_invoice"].get(invoice_key)
        if entry is not None:
            return entry
    return {}


def _parse_courier_settlement(datas: list[bytes], awb_aliases: list[str], date_aliases: list[str],
                               amount_aliases: list[str], status_aliases: list[str],
                               order_id_aliases: list[str]) -> dict[str, dict]:
    out: dict[str, dict] = {}
    for data in datas:
        for row in _rows(_read_table(data)):
            awb = _norm_code(_get(row, *awb_aliases))
            if not awb:
                continue
            fy = _fiscal_year_label(_parse_date(_get(row, *date_aliases)))
            out[awb] = {
                "fy": fy,
                "remit_amount": _to_float(_get(row, *amount_aliases)),
                "status": _clean(_get(row, *status_aliases)) if status_aliases else "",
                "order_id": _clean(_get(row, *order_id_aliases)) if order_id_aliases else "",
            }
    return out


def parse_delhivery(datas: list[bytes]) -> dict[str, dict]:
    return _parse_courier_settlement(
        datas,
        awb_aliases=["waybill_num", "waybill_num.1"],
        date_aliases=["pickup_date", "status_date"],
        amount_aliases=["payable", "cod_amount", "cod"],
        status_aliases=["status"],
        order_id_aliases=["order_id"],
    )


def parse_ekart(datas: list[bytes]) -> dict[str, dict]:
    return _parse_courier_settlement(
        datas,
        awb_aliases=["TRACKING_ID", "SHIPMENT_ID"],
        date_aliases=["DELIVERY_DATE", "date"],
        amount_aliases=["COD_AMOUNT"],
        status_aliases=[],   # Ekart settlement export carries no per-row status column
        order_id_aliases=["ORDER_ID"],
    )


def parse_xpressbees(datas: list[bytes]) -> dict[str, dict]:
    return _parse_courier_settlement(
        datas,
        awb_aliases=["Shipping Id", "POID"],
        date_aliases=["Delivery Date", "Shipping Date", "Transaction Date", "date"],
        amount_aliases=["Net Payment"],
        status_aliases=["Shipment Status"],
        order_id_aliases=["POID"],
    )


_SRN_FY_RE = re.compile(r"(\d{2}-\d{2})")


def parse_srn(datas: list[bytes]) -> dict[tuple[str, str], dict]:
    """Keyed by ("invoice", <original invoice#>) and ("awb", <awb>) so the
    Main Sheet lookup can try invoice-number first, AWB as a fallback.
    Also carries the return's own date + amount (Receivable Amount calc —
    the SRN/credit-note's own value, not the original order's) — the date
    aliases mirror the Combined SRN report being itself a Tally-style
    credit-note register, so "Date" is expected; the others are a defensive
    fallback in case the export names it differently."""
    out: dict[tuple[str, str], dict] = {}
    for data in datas:
        for row in _rows(_read_table(data)):
            srn_number = _clean(_get(row, "Invoice number"))
            if not srn_number:
                continue
            fy_match = _SRN_FY_RE.search(srn_number)
            record = {
                "srn_number": srn_number,
                "fy": fy_match.group(1) if fy_match else "",
                "date": _parse_date(_get(row, "Date", "Return Date", "Credit Note Date", "Invoice Date")),
                "amount": _to_float(_get(row, "Total")),
            }
            orig_invoice = _norm_code(_get(row, "Original Invoice No", "Original Invoice No.1"))
            awb = _norm_code(_get(row, "AWB num"))
            if orig_invoice:
                out.setdefault(("invoice", orig_invoice), record)
            if awb:
                out.setdefault(("awb", awb), record)
    return out


def _in_period(d: _dt.date | None, period: dict | None) -> bool:
    """True if `d` falls within `period`'s [start_month/year, end_month/year]
    (inclusive, Receivable Cycle's own selected run period)."""
    if d is None or not period:
        return False
    start = period["start_year"] * 12 + (period["start_month"] - 1)
    end = period["end_year"] * 12 + (period["end_month"] - 1)
    v = d.year * 12 + (d.month - 1)
    return start <= v <= end


# --------------------------------------------------------------------------
# Sheet builders
# --------------------------------------------------------------------------


def _compute_remark(payment_method: str, remitence: Any, srn_rec: dict | None) -> str:
    if srn_rec:
        return "Returned / Refunded (SRN matched)"
    if payment_method == "PREPAID":
        return "Prepaid - no COD action needed"
    if payment_method == "COD":
        return "COD remitted" if remitence not in ("", None) else "COD remittance pending"
    return ""


def build_main_sheet(tally_rows: list[dict], sales_order_idx: dict, delhivery_idx: dict,
                      ekart_idx: dict, xpressbees_idx: dict, srn_idx: dict) -> list[dict]:
    out = []
    for row in tally_rows:
        awb = row["_awb"]
        order_entry = _lookup_sales_order(sales_order_idx, awb, row["_invoice_key"])

        settlement = None
        if awb:
            for source in (delhivery_idx, ekart_idx, xpressbees_idx):
                if awb in source:
                    settlement = source[awb]
                    break

        srn_rec = srn_idx.get(("invoice", row["_invoice_key"])) if row["_invoice_key"] else None
        if srn_rec is None and awb:
            srn_rec = srn_idx.get(("awb", awb))

        remitence = settlement["remit_amount"] if settlement else ""
        courier_state = (settlement.get("status") if settlement else "") or order_entry.get("status", "")
        srn_value = srn_rec["srn_number"] if srn_rec else ""

        values = {col: row.get(col, "") for col in TAIL_COLUMNS}
        values.update({
            "Date": row.get("Date", ""),
            "Sale Order Number": row.get("Sale Order Number", ""),
            "Invoice number": row.get("Invoice number", ""),
            "Sale Order Number (Duplicate)": row.get("Sale Order Number.1", ""),
            "Remitence": remitence,
            "Srn": srn_value,
            "Remark": _compute_remark(row["_payment_method"], remitence, srn_rec),
            "Order ID": order_entry.get("order_id", ""),
            "Payment Method": row.get("Payment Method", ""),
            "Courier States": courier_state,
            "SRN from (Unnicommerce Refund sheet 24-25)": srn_value if srn_rec and srn_rec.get("fy") == "24-25" else "",
            "SRN from (Unnicommerce Refund sheet 25-26)": srn_value if srn_rec and srn_rec.get("fy") == "25-26" else "",
            "2": "", "3": "", "4": "",
        })
        # Re-key into MAIN_SHEET_COLUMNS' own order: `values` above is built
        # TAIL_COLUMNS-first for convenience, which left the dict's own key
        # order mismatched against the workbook's actual column order (values
        # are written into the workbook by explicit column list either way,
        # so this never affected the .xlsx — but the web "View" derives table
        # columns from a row's own key order, so it was showing everything
        # out of sequence).
        out.append({col: values.get(col, "") for col in MAIN_SHEET_COLUMNS})
    return out


def _courier_bucket(shipping_provider_upper: str) -> str:
    if "DELHIVERY" in shipping_provider_upper:
        return "Delivery"
    if "EKART" in shipping_provider_upper:
        return "Ekart"
    if "XPRESSB" in shipping_provider_upper:   # covers XPRESSBEES / the XPRESSBESS typo seen in real exports
        return "Xpressbees"
    if "DTDC" in shipping_provider_upper:
        return "DTDC"
    if "SELF" in shipping_provider_upper:
        return "Self shipping"
    return "Other COD"


def build_cod_sheets(tally_rows: list[dict], sales_order_idx: dict, delhivery_idx: dict,
                      ekart_idx: dict, xpressbees_idx: dict, srn_idx: dict,
                      period: dict | None = None
                      ) -> tuple[dict[str, list[dict]], dict[str, list[str]], dict]:
    buckets: dict[str, list[dict]] = {name: [] for name in COD_SHEET_ORDER}
    columns: dict[str, list[str]] = dict(_SHEET_COLUMNS)
    receivable: dict[str, dict] = {
        name: {"pending_rows": 0, "pending_amount": 0.0, "srn_rows": 0, "srn_deduction": 0.0}
        for name in RECEIVABLE_COURIERS
    }

    for row in tally_rows:
        if row["_payment_method"] != "COD":
            continue
        awb = row["_awb"]
        order_entry = _lookup_sales_order(sales_order_idx, awb, row["_invoice_key"])

        base = {col: row.get(col, "") for col in TAIL_COLUMNS + ["Payment Method"]}
        base.update({
            "Date": row.get("Date", ""),
            "Sale Order Number": row.get("Sale Order Number", ""),
            "Invoice number": row.get("Invoice number", ""),
            "Sale Order Number (Duplicate)": row.get("Sale Order Number.1", ""),
            "Order ID": order_entry.get("order_id", ""),
        })
        # Re-key into COD_MAIN_COLUMNS' own order — see the matching comment in
        # build_main_sheet: `base` is built TAIL_COLUMNS-first for convenience,
        # which otherwise leaves a row's own dict-key order mismatched against
        # the workbook's actual column order.
        buckets["COD main sheet"].append({col: base.get(col, "") for col in COD_MAIN_COLUMNS})

        # Same SRN match used by the Main Sheet's "Srn" column (invoice# first,
        # AWB fallback) — surfaced here as "SRN Status" on the courier sheets.
        srn_rec = srn_idx.get(("invoice", row["_invoice_key"])) if row["_invoice_key"] else None
        if srn_rec is None and awb:
            srn_rec = srn_idx.get(("awb", awb))

        bucket = _courier_bucket(row["_shipping_provider"])
        entry = dict(base)
        entry["SRN Status"] = srn_rec["srn_number"] if srn_rec else ""
        if bucket == "Delivery":
            match = delhivery_idx.get(awb) if awb else None
            entry["Delivery Status from sales order report"] = order_entry.get("status", "")
            entry["Delivery Time from Sales order report"] = order_entry.get("delivery_time", "")
            entry["Order ID of 24-25 from delhivary setlment report"] = (
                match.get("order_id", "") if match and match.get("fy") == "24-25" else "")
            entry["AWB no in Settlmnt report of Fy 24-25"] = awb if match and match.get("fy") == "24-25" else ""
            entry["AWB no in Settlmnt report of Fy 25-26"] = awb if match and match.get("fy") == "25-26" else ""
        elif bucket == "Ekart":
            match = ekart_idx.get(awb) if awb else None
            entry["Delivery Status from sales order report"] = order_entry.get("status", "")
            entry["Delivery Time from Sales order report"] = order_entry.get("delivery_time", "")
            entry["Tracking ID from Sales Order report"] = order_entry.get("tracking_number", "")
            entry["Tracking ID in Ekart Sttlmt reports Fy 24-25"] = awb if match and match.get("fy") == "24-25" else ""
            entry["Tracking ID in Ekart Sttlmt reports Fy 25-26"] = awb if match and match.get("fy") == "25-26" else ""
        elif bucket == "Xpressbees":
            match = xpressbees_idx.get(awb) if awb else None
            entry["Delivery Status from sales order report"] = order_entry.get("status", "")
            entry["Delivery Time from Sales order report"] = order_entry.get("delivery_time", "")
            # Xpressbees switched its settlement-report id scheme after Mar-25 — pre-cutover
            # (FY24-25) matches populate the "Shipping Id" column, post-cutover (FY25-26)
            # matches populate the "after Mar 25" Order ID column.
            entry["Shipping Id ( AWB) in Xpressbees Sttlmt reports of Fy 24-25"] = (
                awb if match and match.get("fy") == "24-25" else "")
            entry["Order ID in Xpressbees Sttlmt reports after Mar 25"] = (
                awb if match and match.get("fy") == "25-26" else "")
        elif bucket == "DTDC":
            entry["Delivery Status from sales order report"] = order_entry.get("status", "")
            entry["Delivery Time from Sales order report"] = order_entry.get("delivery_time", "")
        elif bucket == "Self shipping":
            pass
        else:
            columns.setdefault(bucket, COD_MAIN_COLUMNS)

        # Re-key into this bucket's own column order (same reasoning as above).
        bucket_columns = columns.get(bucket, COD_MAIN_COLUMNS)
        buckets.setdefault(bucket, []).append({col: entry.get(col, "") for col in bucket_columns})

        # Receivable Amount: for the 4 covered couriers, sum the Total of rows
        # not yet settled (their FY settlement-match columns are all blank);
        # if that same still-pending row was ALSO returned (SRN-matched) within
        # the run's selected period, its SRN amount is deducted — it was never
        # actually receivable. A row already settled by the courier doesn't
        # affect this even if later SRN-matched (its remittance is a separate
        # concern from this pending pool).
        if bucket in RECEIVABLE_COURIERS:
            pending_cols = _PENDING_MATCH_COLUMNS[bucket]
            is_pending = all(not entry.get(c) for c in pending_cols) if pending_cols else True
            if is_pending:
                receivable[bucket]["pending_rows"] += 1
                receivable[bucket]["pending_amount"] += _to_float(entry.get("Total"))
                if srn_rec and _in_period(srn_rec.get("date"), period):
                    receivable[bucket]["srn_rows"] += 1
                    receivable[bucket]["srn_deduction"] += srn_rec.get("amount", 0.0)

    for r in receivable.values():
        r["receivable_amount"] = round(r["pending_amount"] - r["srn_deduction"], 2)
        r["pending_amount"] = round(r["pending_amount"], 2)
        r["srn_deduction"] = round(r["srn_deduction"], 2)

    receivable_summary = {
        "period": period,
        "couriers": receivable,
        "total_pending_amount": round(sum(r["pending_amount"] for r in receivable.values()), 2),
        "total_srn_deduction": round(sum(r["srn_deduction"] for r in receivable.values()), 2),
        "total_receivable_amount": round(sum(r["receivable_amount"] for r in receivable.values()), 2),
    }

    return buckets, columns, receivable_summary


# --------------------------------------------------------------------------
# Top-level entry point + workbook writer
# --------------------------------------------------------------------------


def reconcile_receivable_cycle(files: dict, period: dict | None = None) -> dict:
    """`files` keys: tally_gst (bytes, required), sales_order (bytes,
    required), delhivery/ekart/xpressbees/srn (list[bytes], optional).
    `period`: {"start_month","start_year","end_month","end_year"} (all ints)
    — the run's selected period, used only to decide which SRN/return rows
    count against the Receivable Amount deduction. None disables the
    deduction (every courier's receivable_amount == its pending_amount)."""
    tally_rows = parse_tally_gst(files.get("tally_gst") or b"")
    if not tally_rows:
        raise ValueError("Combine Tally GST report has no usable rows (check the file has a "
                          "'Sale Order Number' or 'Invoice number' column).")

    sales_order_idx = (parse_sales_order(files.get("sales_order"))
                        if files.get("sales_order") else {"by_awb": {}, "by_invoice": {}})
    delhivery_idx = parse_delhivery(files.get("delhivery") or [])
    ekart_idx = parse_ekart(files.get("ekart") or [])
    xpressbees_idx = parse_xpressbees(files.get("xpressbees") or [])
    srn_idx = parse_srn(files.get("srn") or [])

    main_sheet = build_main_sheet(tally_rows, sales_order_idx, delhivery_idx, ekart_idx, xpressbees_idx, srn_idx)
    cod_sheets, cod_columns, receivable_summary = build_cod_sheets(
        tally_rows, sales_order_idx, delhivery_idx, ekart_idx, xpressbees_idx, srn_idx, period)

    summary = {
        "tally_rows": len(tally_rows),
        "prepaid_rows": sum(1 for r in tally_rows if r["_payment_method"] == "PREPAID"),
        "cod_rows": sum(1 for r in tally_rows if r["_payment_method"] == "COD"),
        "srn_matched": sum(1 for r in main_sheet if r.get("Srn")),
        "cod_remitted": sum(1 for r in main_sheet if r.get("Remark") == "COD remitted"),
        "cod_remittance_pending": sum(1 for r in main_sheet if r.get("Remark") == "COD remittance pending"),
    }
    for name, rows in cod_sheets.items():
        summary[f"{name.lower().replace(' ', '_')}_rows"] = len(rows)

    return {
        "main_sheet": main_sheet,
        "cod_sheets": cod_sheets,
        "cod_columns": cod_columns,
        "summary": summary,
        "receivable_summary": receivable_summary,
    }


def _write_sheet(ws, columns: list[str], rows: list[dict]) -> None:
    from openpyxl.styles import Alignment, Font, PatternFill

    ws.append(columns)
    for cell in ws[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = PatternFill("solid", fgColor="1F3864")
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
    ws.row_dimensions[1].height = 26
    for row in rows:
        ws.append([row.get(col, "") for col in columns])
    for i, col in enumerate(columns, start=1):
        ws.column_dimensions[ws.cell(row=1, column=i).column_letter].width = max(12, min(30, len(col) + 2))
    ws.freeze_panes = "A2"


def build_receivable_cycle_workbook(result: dict):
    from openpyxl import Workbook

    wb = Workbook()
    ws = wb.active
    ws.title = "Main Sheet"
    _write_sheet(ws, MAIN_SHEET_COLUMNS, result["main_sheet"])

    cod_sheets = result["cod_sheets"]
    cod_columns = result["cod_columns"]
    ordered_names = list(COD_SHEET_ORDER) + [n for n in cod_sheets if n not in COD_SHEET_ORDER]
    for name in ordered_names:
        rows = cod_sheets.get(name)
        if rows is None or (name not in COD_SHEET_ORDER and not rows):
            continue
        sheet = wb.create_sheet(title=name[:31])
        _write_sheet(sheet, cod_columns.get(name, COD_MAIN_COLUMNS), rows)

    return wb
