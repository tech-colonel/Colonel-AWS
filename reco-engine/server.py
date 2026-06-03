from __future__ import annotations

import logging
logging.basicConfig(level=logging.INFO, format="%(name)s %(levelname)s: %(message)s")

from email.parser import BytesParser
from email.policy import default
from http.server import BaseHTTPRequestHandler, ThreadingHTTPServer
import json
import os
from pathlib import Path
import sys
import re
import threading
from urllib.parse import parse_qs, urlparse
from uuid import uuid4

from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill

ROOT = Path(__file__).resolve().parent
STATIC = ROOT / "static"
if str(ROOT) not in sys.path:
    sys.path.insert(0, str(ROOT))

from recon.core import MatchResult, reconcile, summarize
from recon.gstr_2a_2b_books import read_three_way_uploads, reconcile_three_way, summarize_three_way
from recon.gstr_3b_vs_2b import read_3b_2b_uploads, reconcile_3b_vs_2b, summarize_3b_vs_2b, build_month_pivot
from recon.gstr_1_vs_books import (
    read_gstr1_vs_books_uploads, reconcile_monthly_summary,
    reconcile_b2b, reconcile_b2c, summarize_gstr1_reco,
)
from recon.parsers import read_upload, read_excel_rows, normalize_rows
from recon.bank_reco import process_bank_statement
from recon.gstr_2b_books import reconcile_gstr2b_vs_books
from recon.gstr_2b_books_multistate import (
    reconcile_gstr2b_vs_books_multistate,
    build_gstr2b_books_multistate_workbook,
)
from recon.gstr_3b_tally_entry import (
    parse_gstr3b,
    build_tally_entries,
    build_gstr3b_tally_workbook,
)


JOBS: dict[str, dict] = {}

# Limit simultaneous reconciliation jobs to prevent OOM under heavy load.
# Each job can hold large DataFrames in memory; 8 concurrent runs is safe
# for a server with 8+ GB RAM. Override via MAX_CONCURRENT_RECO env var.
_MAX_RECO = int(os.environ.get("MAX_CONCURRENT_RECO", "8"))
_RECO_SEMAPHORE = threading.Semaphore(_MAX_RECO)


class ReconciliationHandler(BaseHTTPRequestHandler):
    server_version = "CARecon/0.1"

    def do_GET(self) -> None:
        parsed = urlparse(self.path)
        path = parsed.path
        if path == "/":
            self.serve_static("index.html")
            return
        if path.startswith("/static/"):
            self.serve_static(path.removeprefix("/static/"))
            return
        if path.startswith("/api/jobs/") and path.endswith("/export.xlsx"):
            job_id = path.split("/")[3]
            self.export_job(job_id)
            return
        if path.startswith("/api/jobs/"):
            job_id = path.split("/")[3]
            self.write_json(JOBS.get(job_id) or {"error": "Job not found"}, 200 if job_id in JOBS else 404)
            return
        self.write_json({"error": "Not found"}, 404)

    def do_POST(self) -> None:
        parsed = urlparse(self.path)
        if parsed.path != "/api/reconcile":
            self.write_json({"error": "Not found"}, 404)
            return
        # Reject immediately if the server is at capacity — avoids OOM
        if not _RECO_SEMAPHORE.acquire(blocking=False):
            self.write_json(
                {"error": "Server busy. Too many reconciliations running. Retry in 30 seconds."},
                503,
            )
            return
        try:
            fields, files = self.read_multipart()
            tolerance = float(fields.get("tolerance", "1") or 1)
            reco_type = fields.get("reco_type", "gst_2b_purchase")

            if reco_type == "gstr_3b_vs_2b":
                gstr2b_records, gstr3b_records = read_3b_2b_uploads(files)
                results = reconcile_3b_vs_2b(gstr2b_records, gstr3b_records, tolerance=tolerance)
                pivot = build_month_pivot(results)
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize_3b_vs_2b(results),
                    "pivot": pivot,
                    "counts": {
                        "gstr2b_records": len(gstr2b_records),
                        "gstr3b_records": len(gstr3b_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            if reco_type == "gstr_2a_2b_books":
                gstr2a_records, gstr2b_records, books_records = read_three_way_uploads(files)
                results = reconcile_three_way(gstr2a_records, gstr2b_records, books_records, tolerance=tolerance)
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize_three_way(results),
                    "counts": {
                        "gstr2a_records": len(gstr2a_records),
                        "gstr2b_records": len(gstr2b_records),
                        "books_records": len(books_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            if reco_type == "gstr_1_vs_books":
                books, gstr1_b2b, gstr1_b2c = read_gstr1_vs_books_uploads(files)
                monthly = reconcile_monthly_summary(books, gstr1_b2b, gstr1_b2c, tolerance=tolerance)
                b2b_results = reconcile_b2b(books, gstr1_b2b, tolerance=tolerance)
                b2c_results = reconcile_b2c(books, gstr1_b2c, tolerance=tolerance)
                all_results = [r.as_dict() for r in b2b_results] + [r.as_dict() for r in b2c_results]
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize_gstr1_reco(b2b_results, b2c_results, monthly),
                    "monthly_summary": [m.as_dict() for m in monthly],
                    "counts": {
                        "books_records": len(books),
                        "gstr1_b2b_records": len(gstr1_b2b),
                        "gstr1_b2c_records": len(gstr1_b2c),
                        "result_rows": len(all_results),
                    },
                    "results": all_results,
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            if reco_type == "bank_reco":
                bank_file = files.get("bank_statement")
                if not bank_file:
                    self.write_json({"error": "Upload Bank Statement file."}, 400)
                    return
                
                payload = process_bank_statement(bank_file["content"])
                JOBS[payload["job_id"]] = payload
                self.write_json(payload)
                return

            # GSTR-2B vs Books (Purchase Register + Debit Note Register)
            if reco_type == "gstr_2b_books":
                gstr2b_file = files.get("gstr2b")
                purchase_file = files.get("purchase")
                debit_file = files.get("debit")
                if not gstr2b_file or not purchase_file or not debit_file:
                    self.write_json({"error": "Upload GSTR-2B, Purchase Register, and Debit Note Register files."}, 400)
                    return
                from recon.gstr_2b_books import _ensure_xlsx
                gstr2b_bytes = _ensure_xlsx(gstr2b_file["content"])
                purchase_bytes = _ensure_xlsx(purchase_file["content"])
                debit_bytes = _ensure_xlsx(debit_file["content"])
                gstr2b_records, books_records, results = reconcile_gstr2b_vs_books(
                    gstr2b_bytes,
                    purchase_bytes,
                    debit_bytes,
                    tolerance=tolerance,
                )
                job_id = uuid4().hex
                import base64
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": summarize(results),
                    "counts": {
                        "gstr2b_records": len(gstr2b_records),
                        "books_records": len(books_records),
                        "result_rows": len(results),
                    },
                    "results": [result.as_dict() for result in results],
                    "_gstr2b_b64": base64.b64encode(gstr2b_bytes).decode("utf-8"),
                    "_purchase_b64": base64.b64encode(purchase_bytes).decode("utf-8"),
                    "_debit_b64": base64.b64encode(debit_bytes).decode("utf-8"),
                }
                JOBS[job_id] = payload
                self.write_json(payload)
                return

            # GSTR-2B vs Books — Multi-State (N files per input type)
            if reco_type == "gstr_2b_books_multistate":
                def _file_list(name):
                    """Normalise single dict or list-of-dicts into list of content bytes."""
                    val = files.get(name)
                    if val is None:
                        return []
                    items = val if isinstance(val, list) else [val]
                    return [item["content"] for item in items if item.get("content")]

                def _file_items(name):
                    """Return raw list of file dicts (keeps filename alongside content)."""
                    val = files.get(name)
                    if val is None:
                        return []
                    items = val if isinstance(val, list) else [val]
                    return [item for item in items if item.get("content")]

                _GSTIN_IN_FILENAME = re.compile(
                    r'[0-9]{2}[A-Z]{5}[0-9]{4}[A-Z][0-9A-Z]Z[0-9A-Z]',
                    re.IGNORECASE,
                )

                def _entity_gstin_from_filename(fn: str) -> str:
                    """Extract entity GSTIN from GSTN portal filename (e.g. 102024_29AAECF7751Q1ZS_GSTR2B_...)."""
                    m = _GSTIN_IN_FILENAME.search(fn.upper())
                    return m.group(0).upper() if m else ""

                gstr2b_items  = _file_items("gstr2b")
                gstr2b_list   = [it["content"] for it in gstr2b_items]
                entity_gstins = [_entity_gstin_from_filename(it.get("filename", "")) for it in gstr2b_items]

                purchase_list = _file_list("purchase")
                debit_list    = _file_list("debit")

                if not gstr2b_list or not purchase_list:
                    self.write_json({"error": "Upload at least one GSTR-2B and one Purchase Register file."}, 400)
                    return

                gstr2b_recs, books_recs, results = reconcile_gstr2b_vs_books_multistate(
                    gstr2b_list, purchase_list, debit_list or [b""] * len(purchase_list),
                    tolerance=tolerance,
                    entity_gstins=entity_gstins,
                )

                import base64
                job_id  = uuid4().hex
                payload = {
                    "job_id":    job_id,
                    "reco_type": reco_type,
                    "summary":   summarize(results),
                    "counts": {
                        "gstr2b_records": len(gstr2b_recs),
                        "books_records":  len(books_recs),
                        "result_rows":    len(results),
                        "file_count":     len(gstr2b_list),
                    },
                    "results": [result.as_dict() for result in results],
                    # First file of each type (for base workbook source sheets — state 1)
                    "_gstr2b_b64":   base64.b64encode(gstr2b_list[0]).decode("utf-8"),
                    "_purchase_b64": base64.b64encode(purchase_list[0]).decode("utf-8"),
                    "_debit_b64":    base64.b64encode(debit_list[0]).decode("utf-8") if debit_list else "",
                    # All state files (for adding per-state source sheets to the workbook)
                    "_all_gstr2b_b64":   [base64.b64encode(f).decode("utf-8") for f in gstr2b_list],
                    "_all_purchase_b64": [base64.b64encode(f).decode("utf-8") for f in purchase_list if f],
                    "_all_debit_b64":    [base64.b64encode(f).decode("utf-8") for f in debit_list   if f],
                    # Stash MatchResult objects so export_job can rebuild Remark 3
                    "_results_obj":  results,
                }
                JOBS[job_id] = payload
                self.write_json({k: v for k, v in payload.items() if not k.startswith("_")})
                return

            if reco_type == "gstr_3b_tally_entry":
                gstr3b_file = files.get("gstr3b")
                if not gstr3b_file:
                    self.write_json({"error": "Upload GSTR-3B file."}, 400)
                    return
                parsed = parse_gstr3b(gstr3b_file["filename"], gstr3b_file["content"])
                entries = build_tally_entries(parsed)
                job_id = uuid4().hex
                payload = {
                    "job_id": job_id,
                    "reco_type": reco_type,
                    "summary": {
                        "gstin": parsed["gstin"],
                        "state": parsed["state"],
                        "period": parsed["period"],
                    },
                    "counts": {"entry_rows": len(entries)},
                    "results": entries,
                    "_parsed": parsed,
                }
                JOBS[job_id] = payload
                self.write_json({k: v for k, v in payload.items() if not k.startswith("_")})
                return

            # Default two-file reconciliation (gst_2b_purchase)
            gstr2b_file = files.get("gstr2b")
            purchase_file = files.get("purchase")
            if not gstr2b_file or not purchase_file:
                self.write_json({"error": "Upload both GSTR-2B and Purchase Register files."}, 400)
                return

            gstr2b_records = read_upload(gstr2b_file["filename"], gstr2b_file["content"], "GSTR-2B")
            purchase_records = read_upload(purchase_file["filename"], purchase_file["content"], "Purchase Register")
            results = reconcile(gstr2b_records, purchase_records, tolerance=tolerance)

            job_id = uuid4().hex
            payload = {
                "job_id": job_id,
                "reco_type": reco_type,
                "summary": summarize(results),
                "counts": {
                    "gstr2b_records": len(gstr2b_records),
                    "purchase_records": len(purchase_records),
                    "result_rows": len(results),
                },
                "results": [result.as_dict() for result in results],
            }
            JOBS[job_id] = payload
            self.write_json(payload)
        except Exception as exc:
            self.write_json({"error": str(exc)}, 500)
        finally:
            _RECO_SEMAPHORE.release()

    def read_multipart(self) -> tuple[dict[str, str], dict[str, dict]]:
        content_type = self.headers.get("Content-Type", "")
        content_length = int(self.headers.get("Content-Length", "0"))
        body = self.rfile.read(content_length)
        logging.info("Multipart: Content-Type=%s, body length=%d", content_type, len(body))
        message = BytesParser(policy=default).parsebytes(
            f"Content-Type: {content_type}\r\nMIME-Version: 1.0\r\n\r\n".encode() + body
        )
        fields: dict[str, str] = {}
        files: dict[str, dict] = {}
        for part in message.iter_parts():
            disposition = part.get_content_disposition()
            name = part.get_param("name", header="content-disposition")
            filename = part.get_filename()
            content = part.get_payload(decode=True) or b""
            logging.info("  Part: disposition=%s, name=%s, filename=%s, size=%d", disposition, name, filename, len(content))
            if disposition != "form-data":
                continue
            if filename:
                entry = {"filename": filename, "content": content}
                if name in files:
                    existing = files[name]
                    if not isinstance(existing, list):
                        files[name] = [existing]
                    files[name].append(entry)
                else:
                    files[name] = entry
            else:
                fields[name] = content.decode("utf-8", errors="replace")
        def _log_file(v):
            if isinstance(v, list):
                return [{'fn': i['filename'], 'sz': len(i['content'])} for i in v]
            return {'fn': v['filename'], 'sz': len(v['content'])}
        logging.info("Parsed fields: %s, files: %s", list(fields.keys()), {k: _log_file(v) for k, v in files.items()})
        return fields, files

    def serve_static(self, relative_path: str) -> None:
        target = (STATIC / relative_path).resolve()
        if not str(target).startswith(str(STATIC.resolve())) or not target.exists():
            self.write_json({"error": "Not found"}, 404)
            return
        content_type = "text/plain"
        if target.suffix == ".html":
            content_type = "text/html; charset=utf-8"
        elif target.suffix == ".css":
            content_type = "text/css; charset=utf-8"
        elif target.suffix == ".js":
            content_type = "application/javascript; charset=utf-8"
        data = target.read_bytes()
        self.send_response(200)
        self.send_header("Content-Type", content_type)
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def write_json(self, payload: dict, status: int = 200) -> None:
        data = json.dumps(payload, indent=2).encode("utf-8")
        self.send_response(status)
        self.send_header("Content-Type", "application/json; charset=utf-8")
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def export_job(self, job_id: str) -> None:
        payload = JOBS.get(job_id)
        if not payload:
            self.write_json({"error": "Job not found"}, 404)
            return
        workbook = build_workbook(
            payload["results"],
            payload["summary"],
            payload["counts"],
            payload.get("reco_type", "gst_2b_purchase"),
            pivot=payload.get("pivot"),
            payload=payload,
        )
        from io import BytesIO

        buffer = BytesIO()
        workbook.save(buffer)
        data = buffer.getvalue()
        self.send_response(200)
        reco_type = payload.get("reco_type", "gst_2b_purchase")
        if reco_type == "bank_reco":
            filename_prefix = "bank_statement"
        elif reco_type == "gstr_1_vs_books":
            filename_prefix = "gstr1_vs_books"
        elif reco_type == "gstr_2b_books_multistate":
            filename_prefix = "2b_vs_books_multistate"
        elif reco_type == "gstr_3b_tally_entry":
            filename_prefix = "gstr3b_tally_entry"
        else:
            filename_prefix = "reconciliation"
        self.send_header("Content-Type", "application/vnd.openxmlformats-officedocument.spreadsheetml.sheet")
        self.send_header("Content-Disposition", f'attachment; filename="{filename_prefix}-{job_id}.xlsx"')
        self.send_header("Content-Length", str(len(data)))
        self.end_headers()
        self.wfile.write(data)

    def log_message(self, format: str, *args) -> None:
        print(f"{self.address_string()} - {format % args}")


def build_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int], reco_type: str = "gst_2b_purchase", pivot: list[dict] | None = None, payload: dict | None = None) -> Workbook:
    if reco_type == "gstr_2b_books":
        from recon.gstr_2b_books import build_gstr2b_books_workbook
        return build_gstr2b_books_workbook(results, payload=payload)
    if reco_type == "gstr_2b_books_multistate":
        # Use MatchResult objects (with suggested_action_3) if available in the job payload
        result_objs = (payload or {}).get("_results_obj") or results
        return build_gstr2b_books_multistate_workbook(result_objs, payload=payload)
    if reco_type == "gstr_2a_2b_books":
        return build_three_way_workbook(results, summary, counts)
    if reco_type == "gstr_3b_vs_2b":
        return build_3b_vs_2b_workbook(results, summary, counts, pivot or [])
    if reco_type == "bank_reco":
        return build_bank_reco_workbook(results, summary, counts)
    if reco_type == "gstr_1_vs_books":
        p = payload or {}
        return build_gstr1_workbook(
            p.get("results", results),
            p.get("monthly_summary", []),
            summary,
            counts,
        )
    if reco_type == "gstr_3b_tally_entry":
        return build_gstr3b_tally_workbook((payload or {}).get("_parsed", {}), results)

    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["GSTR-2B records", counts["gstr2b_records"]])
    summary_sheet.append(["Purchase records", counts["purchase_records"]])
    summary_sheet.append(["Result rows", counts["result_rows"]])
    summary_sheet.append([])
    summary_sheet.append(["Category", "Count"])
    for category, count in summary.items():
        summary_sheet.append([category, count])
    style_header(summary_sheet)

    categories = sorted({result["category"] for result in results})
    for category in categories:
        sheet = workbook.create_sheet(safe_sheet_title(category))
        write_result_sheet(sheet, [result for result in results if result["category"] == category])
    return workbook


def safe_sheet_title(title: str) -> str:
    cleaned = re.sub(r"[\[\]\*:/\\?]", "-", title).strip()
    return (cleaned or "Sheet")[:31]


def write_result_sheet(sheet, rows: list[dict]) -> None:
    headers = [
        "Category",
        "Confidence",
        "2B GSTIN",
        "PR GSTIN",
        "2B Doc No",
        "PR Doc No",
        "2B Date",
        "PR Date",
        "2B Taxable",
        "PR Taxable",
        "2B Tax",
        "PR Tax",
        "Mismatches",
        "Suggested Action",
        "Explanation",
    ]
    sheet.append(headers)
    for result in rows:
        gstr2b = result.get("gstr2b") or {}
        purchase = result.get("purchase") or {}
        sheet.append(
            [
                result["category"],
                result["confidence"],
                gstr2b.get("supplier_gstin", ""),
                purchase.get("supplier_gstin", ""),
                gstr2b.get("doc_no", ""),
                purchase.get("doc_no", ""),
                gstr2b.get("doc_date", ""),
                purchase.get("doc_date", ""),
                gstr2b.get("taxable_value", ""),
                purchase.get("taxable_value", ""),
                gstr2b.get("total_tax", ""),
                purchase.get("total_tax", ""),
                ", ".join(result.get("mismatch_fields", [])),
                result["suggested_action"],
                result["explanation"],
            ]
        )
    style_header(sheet)


def build_three_way_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int]) -> Workbook:
    workbook = Workbook()
    summary_sheet = workbook.active
    summary_sheet.title = "Summary"
    summary_sheet.append(["Metric", "Value"])
    summary_sheet.append(["GSTR-2A records", counts.get("gstr2a_records", 0)])
    summary_sheet.append(["GSTR-2B records", counts.get("gstr2b_records", 0)])
    summary_sheet.append(["Books records", counts.get("books_records", 0)])
    summary_sheet.append(["Result rows", counts.get("result_rows", 0)])
    summary_sheet.append([])
    summary_sheet.append(["Category", "Count", "Amount"])
    for category, count in summary.items():
        amount = sum(record_amount(result) for result in results if result["category"] == category)
        summary_sheet.append([category, count, amount])
    style_header(summary_sheet)

    line_sheet = workbook.create_sheet("Invoice-level Matching")
    write_three_way_sheet(line_sheet, results)

    pan_sheet = workbook.create_sheet("PAN Summary")
    pan_totals: dict[str, dict[str, float]] = {}
    for result in results:
        pan = first_value(result, "pan") or "Unknown"
        bucket = pan_totals.setdefault(pan, {"2A": 0.0, "2B": 0.0, "Books": 0.0})
        if result.get("gstr2a"):
            bucket["2A"] += result["gstr2a"].get("doc_value", 0) or 0
        if result.get("gstr2b"):
            bucket["2B"] += result["gstr2b"].get("doc_value", 0) or 0
        if result.get("books"):
            bucket["Books"] += result["books"].get("doc_value", 0) or 0
    pan_sheet.append(["PAN", "As per 2A", "As per 2B", "As per Books", "2A - Books", "2B - Books"])
    for pan, totals in sorted(pan_totals.items()):
        pan_sheet.append([pan, totals["2A"], totals["2B"], totals["Books"], totals["2A"] - totals["Books"], totals["2B"] - totals["Books"]])
    style_header(pan_sheet)

    for category in sorted({result["category"] for result in results}):
        sheet = workbook.create_sheet(safe_sheet_title(category))
        write_three_way_sheet(sheet, [result for result in results if result["category"] == category])
    return workbook


def write_three_way_sheet(sheet, rows: list[dict]) -> None:
    headers = [
        "Category",
        "Confidence",
        "PAN",
        "Supplier",
        "Doc No",
        "2A Value",
        "2B Value",
        "Books Value",
        "Difference",
        "2A Month",
        "2B Month",
        "Books Month",
        "Mismatches",
        "Suggested Action",
        "Explanation",
    ]
    sheet.append(headers)
    for result in rows:
        value_2a = (result.get("gstr2a") or {}).get("doc_value", 0) or 0
        value_2b = (result.get("gstr2b") or {}).get("doc_value", 0) or 0
        value_books = (result.get("books") or {}).get("doc_value", 0) or 0
        portal_value = value_2b if result.get("gstr2b") else value_2a
        sheet.append(
            [
                result["category"],
                result["confidence"],
                first_value(result, "pan"),
                first_value(result, "supplier_name"),
                first_value(result, "doc_no"),
                value_2a,
                value_2b,
                value_books,
                portal_value - value_books,
                (result.get("gstr2a") or {}).get("month", ""),
                (result.get("gstr2b") or {}).get("month", ""),
                (result.get("books") or {}).get("month", ""),
                ", ".join(result.get("mismatch_fields", [])),
                result["suggested_action"],
                result["explanation"],
            ]
        )
    style_header(sheet)


def first_value(result: dict, field: str) -> str:
    for key in ("gstr2a", "gstr2b", "books"):
        record = result.get(key) or {}
        if record.get(field):
            return record[field]
    return ""


def record_amount(result: dict) -> float:
    for key in ("books", "gstr2b", "gstr2a"):
        record = result.get(key)
        if record:
            return record.get("doc_value", 0) or 0
    return 0.0


def style_header(sheet) -> None:
    fill = PatternFill("solid", fgColor="123C69")
    for cell in sheet[1]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill
    for column_cells in sheet.columns:
        max_length = max((len(str(cell.value or "")) for cell in column_cells), default=8)
        sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 48)


def build_3b_vs_2b_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int], pivot: list[dict]) -> Workbook:
    """Build the Excel workbook matching the accountant's expected format.

    Pro-accountant sign convention:
      B2B:  2B row = POSITIVE values, 3B mirror row = NEGATIVE (negate 2B)
      CDNR: 2B row = NEGATIVE values (credit notes reduce ITC), 3B mirror row = POSITIVE

    The 3B mirror row ALWAYS uses 2B values (negated) — NOT the raw 3B PR values.
    This ensures every matched pair nets to EXACTLY zero.

    Output ordering:
      1. Matched CDNRs (sorted by date)
      2. Not Claimed in 3B (CDNR entries)
      3. Matched B2B (sorted by date)
      4. Not Claimed in 3B (B2B entries)
      5. Not in 2B entries
    """
    workbook = Workbook()

    # --- Primary sheet: 2B vs 3B (matches accountant's expected output) ---
    ws = workbook.active
    ws.title = "2B vs 3B"
    headers = [
        "Month", "Date", "Invoice No", "Party Name",
        "Value", "CGST", "SGST", "IGST",
        "Source", "Remarks", "Remarks 2 (Detail Explanation)", "Taxable",
        "Effect in Books", "Effect in 9-9C",
    ]
    ws.append(headers)

    def _is_cdnr(r: dict) -> bool:
        """Check if a result originates from a CDNR source."""
        rec_2b = r.get("rec_2b")
        if rec_2b and "CDNR" in (rec_2b.get("source", "") or "").upper():
            return True
        return False

    def _sort_key(r: dict) -> tuple:
        """Sort key: CDNRs first, then B2B; within each group, sort by date."""
        is_cdnr = _is_cdnr(r)
        category = r.get("category", "")
        # Category order: Matched/Amount Mismatch first, then Not Claimed, then Not in 2B
        cat_order = 0 if category in ("Matched", "Amount Mismatch") else 1 if category == "Not Claimed in 3B" else 2
        rec = r.get("rec_2b") or r.get("rec_3b") or {}
        date_str = rec.get("date", "") or ""
        return (0 if is_cdnr else 1, cat_order, date_str)

    sorted_results = sorted(results, key=_sort_key)

    for r in sorted_results:
        category = r.get("category", "")
        rec_2b = r.get("rec_2b")
        rec_3b = r.get("rec_3b")
        is_cdnr = _is_cdnr(r)

        if category in ("Matched", "Amount Mismatch"):
            if rec_2b:
                # Get raw 2B values from portal (always positive in portal)
                value_2b = rec_2b.get("value", 0) or 0
                cgst_2b = rec_2b.get("cgst", 0) or 0
                sgst_2b = rec_2b.get("sgst", 0) or 0
                igst_2b = rec_2b.get("igst", 0) or 0
                taxable_2b = rec_2b.get("taxable", 0) or 0

                if is_cdnr:
                    # CDNR: Credit notes REDUCE ITC → 2B row is NEGATIVE
                    sign_2b = -1
                else:
                    # B2B: Regular invoices → 2B row is POSITIVE
                    sign_2b = 1

                # Row 1: 2B side
                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    sign_2b * abs(value_2b),
                    sign_2b * abs(cgst_2b) if cgst_2b else 0,
                    sign_2b * abs(sgst_2b) if sgst_2b else 0,
                    sign_2b * abs(igst_2b) if igst_2b else 0,
                    rec_2b.get("source", "2B"),
                    "Matched",
                    None,
                    sign_2b * abs(taxable_2b),
                    None,
                    None,
                ])

                # Row 2: 3B Working mirror row (NEGATE the 2B row → nets to zero)
                # Use 2B values with opposite sign, and 2B party name for consistency
                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    -sign_2b * abs(value_2b),
                    -sign_2b * abs(cgst_2b) if cgst_2b else 0,
                    -sign_2b * abs(sgst_2b) if sgst_2b else 0,
                    -sign_2b * abs(igst_2b) if igst_2b else 0,
                    "3B Working",
                    "Matched",
                    None,
                    -sign_2b * abs(taxable_2b),
                    None,
                    None,
                ])

        elif category == "Not in 2B":
            # Single row — 3B Working entry not in 2B
            if rec_3b:
                ws.append([
                    rec_3b.get("month", ""),
                    rec_3b.get("date", ""),
                    rec_3b.get("invoice_no", ""),
                    rec_3b.get("party_name", ""),
                    rec_3b.get("value", 0),
                    rec_3b.get("cgst", 0) or 0,
                    rec_3b.get("sgst", 0) or 0,
                    rec_3b.get("igst", 0) or 0,
                    "3B Working",
                    "Not in 2B",
                    rec_3b.get("remarks2", "") or "ITC Reversal",
                    rec_3b.get("taxable", 0),
                    "No effect",
                    "No effect",
                ])

        elif category == "Not Claimed in 3B":
            # Single row — 2B entry not claimed in 3B
            if rec_2b:
                value_2b = rec_2b.get("value", 0) or 0
                cgst_2b = rec_2b.get("cgst", 0) or 0
                sgst_2b = rec_2b.get("sgst", 0) or 0
                igst_2b = rec_2b.get("igst", 0) or 0
                taxable_2b = rec_2b.get("taxable", 0) or 0

                ws.append([
                    rec_2b.get("month", ""),
                    rec_2b.get("date", ""),
                    rec_2b.get("invoice_no", ""),
                    rec_2b.get("party_name", ""),
                    value_2b,
                    cgst_2b if cgst_2b else 0,
                    sgst_2b if sgst_2b else 0,
                    igst_2b if igst_2b else 0,
                    rec_2b.get("source", "2B"),
                    "Not claimed in 3B",
                    rec_2b.get("remarks2", ""),
                    taxable_2b,
                    "No effect",
                    "No effect",
                ])

    style_header(ws)

    # --- Summary sheet ---
    summary_ws = workbook.create_sheet("Summary")
    summary_ws.append(["GSTR-3B vs GSTR-2B Reconciliation"])
    summary_ws.append([])
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["GSTR-2B records", counts.get("gstr2b_records", 0)])
    summary_ws.append(["GSTR-3B Working records", counts.get("gstr3b_records", 0)])
    summary_ws.append(["Total result rows", counts.get("result_rows", 0)])
    summary_ws.append([])
    summary_ws.append(["Category", "Count"])
    for cat, cnt in summary.items():
        summary_ws.append([cat, cnt])
    style_header_row(summary_ws, 3)
    style_header_row(summary_ws, 8)

    # --- Month Pivot sheet ---
    pv = workbook.create_sheet("Month Pivot")
    pv.append([
        "Month",
        "2B IGST", "2B CGST", "2B SGST",
        "3B IGST", "3B CGST", "3B SGST",
        "IGST Diff", "CGST Diff", "SGST Diff",
        "2B Value", "3B Value", "Value Diff",
    ])
    for row in pivot:
        pv.append([
            row["month"],
            row["igst_2b"], row["cgst_2b"], row["sgst_2b"],
            row["igst_3b"], row["cgst_3b"], row["sgst_3b"],
            row["igst_diff"], row["cgst_diff"], row["sgst_diff"],
            row["value_2b"], row["value_3b"], row["value_diff"],
        ])
    style_header(pv)

    return workbook


def style_header_row(sheet, row_num: int) -> None:
    """Style a specific row as a header (used for multi-section summary sheets)."""
    fill = PatternFill("solid", fgColor="123C69")
    for cell in sheet[row_num]:
        cell.font = Font(bold=True, color="FFFFFF")
        cell.fill = fill


def build_gstr1_workbook(
    results: list[dict],
    monthly_summary: list[dict],
    summary: dict[str, int],
    counts: dict[str, int],
) -> Workbook:
    workbook = Workbook()

    # --- Sheet 1: Monthly Summary ---
    ws_monthly = workbook.active
    ws_monthly.title = "Monthly Summary"
    monthly_headers = [
        "Month", "Status",
        "Books Taxable", "GSTR-1 Taxable", "Diff Taxable",
        "Books CGST", "GSTR-1 CGST", "Diff CGST",
        "Books SGST", "GSTR-1 SGST", "Diff SGST",
        "Books IGST", "GSTR-1 IGST", "Diff IGST",
    ]
    ws_monthly.append(monthly_headers)
    for row in monthly_summary:
        ws_monthly.append([
            row.get("month", ""),
            row.get("status", ""),
            row.get("books_taxable", 0), row.get("gstr1_taxable", 0), row.get("diff_taxable", 0),
            row.get("books_cgst", 0), row.get("gstr1_cgst", 0), row.get("diff_cgst", 0),
            row.get("books_sgst", 0), row.get("gstr1_sgst", 0), row.get("diff_sgst", 0),
            row.get("books_igst", 0), row.get("gstr1_igst", 0), row.get("diff_igst", 0),
        ])
    style_header(ws_monthly)

    # --- Sheet 2: B2B Reconciliation ---
    b2b_rows = [r for r in results if r.get("reco_level") == "B2B"]
    ws_b2b = workbook.create_sheet("B2B Reconciliation")
    b2b_headers = [
        "Category", "Confidence",
        "Books Invoice No", "Books Date", "Books Party", "Books GSTIN",
        "GSTR-1 Invoice No", "GSTR-1 Date", "GSTR-1 Party", "GSTR-1 GSTIN",
        "Books Taxable", "GSTR-1 Taxable", "Diff Taxable",
        "Books CGST", "GSTR-1 CGST",
        "Books SGST", "GSTR-1 SGST",
        "Books IGST", "GSTR-1 IGST",
        "Mismatches", "Suggested Action",
    ]
    ws_b2b.append(b2b_headers)
    for r in b2b_rows:
        bk = r.get("books") or {}
        g1 = r.get("gstr1") or {}
        ws_b2b.append([
            r.get("category", ""), r.get("confidence", ""),
            bk.get("doc_no", ""), bk.get("doc_date", ""), bk.get("party_name", ""), bk.get("buyer_gstin", ""),
            g1.get("doc_no", ""), g1.get("doc_date", ""), g1.get("party_name", ""), g1.get("buyer_gstin", ""),
            bk.get("taxable_value", 0), g1.get("taxable_value", 0),
            round(float(bk.get("taxable_value") or 0) - float(g1.get("taxable_value") or 0), 2),
            bk.get("cgst", 0), g1.get("cgst", 0),
            bk.get("sgst", 0), g1.get("sgst", 0),
            bk.get("igst", 0), g1.get("igst", 0),
            ", ".join(r.get("mismatch_fields", [])),
            r.get("suggested_action", ""),
        ])
    style_header(ws_b2b)

    # --- Sheet 3: B2C Reconciliation ---
    b2c_rows = [r for r in results if r.get("reco_level") == "B2C"]
    ws_b2c = workbook.create_sheet("B2C Reconciliation")
    b2c_headers = [
        "Category", "Place of Supply", "Tax Rate (%)",
        "Books Taxable", "GSTR-1 Taxable", "Diff Taxable",
        "Books CGST", "GSTR-1 CGST", "Diff CGST",
        "Books SGST", "GSTR-1 SGST", "Diff SGST",
        "Books IGST", "GSTR-1 IGST", "Diff IGST",
        "Suggested Action",
    ]
    ws_b2c.append(b2c_headers)
    for r in b2c_rows:
        ws_b2c.append([
            r.get("category", ""), r.get("place_of_supply", ""), r.get("tax_rate", ""),
            r.get("books_taxable", 0), r.get("gstr1_taxable", 0), r.get("diff_taxable", 0),
            r.get("books_cgst", 0), r.get("gstr1_cgst", 0), r.get("diff_cgst", 0),
            r.get("books_sgst", 0), r.get("gstr1_sgst", 0), r.get("diff_sgst", 0),
            r.get("books_igst", 0), r.get("gstr1_igst", 0), r.get("diff_igst", 0),
            r.get("suggested_action", ""),
        ])
    style_header(ws_b2c)

    # --- Sheet 4: Difference Report (exceptions only) ---
    ws_diff = workbook.create_sheet("Difference Report")
    diff_headers = ["Level", "Category", "Invoice No / State", "Party / Rate", "Difference", "Suggested Action"]
    ws_diff.append(diff_headers)
    matched_cats = {"Matched", "B2B: Matched", "B2C: Matched"}
    for r in results:
        if r.get("category") in matched_cats or r.get("category", "").endswith(": Matched"):
            continue
        level = r.get("reco_level", "")
        if level == "B2B":
            bk = r.get("books") or r.get("gstr1") or {}
            inv = bk.get("doc_no", "")
            party = bk.get("party_name", "")
            diff = round(
                float((r.get("books") or {}).get("taxable_value") or 0)
                - float((r.get("gstr1") or {}).get("taxable_value") or 0), 2
            )
        else:
            inv = r.get("place_of_supply", "")
            party = f"Rate {r.get('tax_rate', '')}%"
            diff = r.get("diff_taxable", 0)
        ws_diff.append([
            level, r.get("category", ""), inv, party, diff, r.get("suggested_action", ""),
        ])
    style_header(ws_diff)

    # --- Sheet 5: Summary ---
    ws_sum = workbook.create_sheet("Summary")
    ws_sum.append(["GSTR-1 vs Books Reconciliation"])
    ws_sum.append([])
    ws_sum.append(["Metric", "Value"])
    ws_sum.append(["Books records", counts.get("books_records", 0)])
    ws_sum.append(["GSTR-1 B2B records", counts.get("gstr1_b2b_records", 0)])
    ws_sum.append(["GSTR-1 B2C records", counts.get("gstr1_b2c_records", 0)])
    ws_sum.append([])
    ws_sum.append(["Category", "Count"])
    for cat, cnt in summary.items():
        ws_sum.append([cat, cnt])
    style_header_row(ws_sum, 3)
    style_header_row(ws_sum, 8)

    return workbook


def main() -> None:
    query = parse_qs(urlparse("?" + "port=8765").query)
    port = int(query.get("port", ["8765"])[0])
    server = ThreadingHTTPServer(("0.0.0.0", port), ReconciliationHandler)
    print(f"CA Reconciliation Tool running at http://127.0.0.1:{port}")
    server.serve_forever()


def build_bank_reco_workbook(results: list[dict], summary: dict[str, int], counts: dict[str, int]) -> Workbook:
    workbook = Workbook()
    
    ws = workbook.active
    ws.title = "OD acc Working"
    headers = ["Txn Date", "Description", "Debit", "Credit", "Balance", "Type ", "Ledger name"]
    ws.append(headers)
    
    for r in results:
        ws.append([
            r.get("txn_date", ""),
            r.get("original_description", ""),
            r.get("debit") if r.get("debit") else None,
            r.get("credit") if r.get("credit") else None,
            r.get("balance", 0),
            r.get("predicted_type", ""),
            r.get("predicted_ledger", "")
        ])
    style_header(ws)
    
    summary_ws = workbook.create_sheet("Summary")
    summary_ws.append(["Bank Statement Classification"])
    summary_ws.append([])
    summary_ws.append(["Metric", "Value"])
    summary_ws.append(["Total Ledgers", counts.get("master_ledgers", 0)])
    summary_ws.append(["Total Bank Rows", counts.get("bank_rows", 0)])
    summary_ws.append([])
    summary_ws.append(["Confidence Level", "Count"])
    for cat, cnt in summary.items():
        summary_ws.append([cat, cnt])
    style_header_row(summary_ws, 3)
    style_header_row(summary_ws, 7)
    
    return workbook

if __name__ == "__main__":
    main()
