#!/usr/bin/env python3
"""
build_cc_seed.py — mine a merchant→ledger seed directory out of a folder of
already-booked credit card working files.

Input : a folder of month sub-folders, each holding the accountant's workbook.
        The transaction table is found by HEADER, never by sheet name or row
        index, so this works on the real-world mess (sheets called 'Table 1',
        'Table 11', 'Credit Card', 'Sheet2', 'Working'; header on row 1 or 2).
Output: JSON [{key_value, ledger, months, source_narration}] ready to load into
        cc_merchant_directory for a brand.

Conflicts (one merchant booked to two ledgers across months — BSESR was
'Best Electrical' early, then 'Warehouse Electricity Charges') resolve to the
MOST RECENT month's ledger, on the reasoning that the newer booking reflects
current policy. Every conflict is reported so the decision is visible.

Rows that teach nothing are skipped: 'Suspense' (an unmapped row, not a
learning), card bill payments, and issuer charges (handled by the L0 card rule).

Usage:
    python3 tools/build_cc_seed.py "<folder>" -o <out.json>
"""
import os
import re
import sys
import json
import glob
import argparse
from collections import defaultdict

sys.path.insert(0, os.path.join(os.path.dirname(os.path.abspath(__file__)), ".."))
from recon.credit_card_booking import (            # noqa: E402
    merchant_key, extract_keys, _BANK_CHARGE_RE, _PAYMENT_RE, _norm_ledger,
)

# "6. Sep 25 Booked" -> 6 . The numeric prefix is the accountant's own booking
# order, which is what "most recent" must mean here (Apr 25 = 1 ... Mar 26 = 12).
_ORDER_RE = re.compile(r"^\s*(\d+)\s*[.\-]")


def month_order(folder_name: str) -> int:
    m = _ORDER_RE.match(folder_name)
    return int(m.group(1)) if m else 999


def find_table(ws, want=("debit", "credit")):
    """Locate the header row by content. Returns (header_row_idx, {name: col_idx})."""
    for r in range(1, min(ws.max_row, 12) + 1):
        cells = [str(c.value or "").strip().lower() for c in ws[r]]
        if all(any(w == c for c in cells) for w in want):
            idx = {}
            for i, c in enumerate(cells):
                if c and c not in idx:
                    idx[c] = i
            return r, idx
    return None, None


def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("folder")
    ap.add_argument("-o", "--out", default="cc_seed.json")
    args = ap.parse_args()

    import openpyxl

    # key -> [(month_order, ledger, narration)]
    observed = defaultdict(list)
    files = sorted(glob.glob(os.path.join(args.folder, "*", "*.xlsx")))
    if not files:
        print(f"No .xlsx found under {args.folder}", file=sys.stderr)
        return 1

    scanned = booked = 0
    for path in files:
        order = month_order(os.path.basename(os.path.dirname(path)))
        try:
            wb = openpyxl.load_workbook(path, data_only=True)
        except Exception as e:
            print(f"  ! skip {os.path.basename(path)}: {type(e).__name__}", file=sys.stderr)
            continue
        for ws in wb.worksheets:
            hrow, idx = find_table(ws)
            if not idx:
                continue
            scanned += 1
            i_narr = idx.get("transaction details", 1)
            i_dr, i_cr = idx["debit"], idx["credit"]
            for row in ws.iter_rows(min_row=hrow + 1, values_only=True):
                if len(row) <= max(i_dr, i_cr, i_narr):
                    continue
                narr, dr, cr = row[i_narr], row[i_dr], row[i_cr]
                if not narr or (not dr and not cr):
                    continue
                narr = str(narr)
                if _PAYMENT_RE.search(narr) or _BANK_CHARGE_RE.search(narr):
                    continue
                # The counterparty ledger is whichever side is NOT the card.
                dr_s, cr_s = str(dr or "").strip(), str(cr or "").strip()
                ledger = dr_s if "credit card" not in cr_s.lower() else cr_s
                if "credit card" in ledger.lower():
                    ledger = cr_s if "credit card" in dr_s.lower() else dr_s
                if not ledger or _norm_ledger(ledger) == "suspense":
                    continue
                for kt, key in extract_keys(narr):
                    observed[(kt, key)].append((order, ledger.strip(), narr))
                booked += 1

    seed, conflicts = [], []
    for (kt, key), rows in sorted(observed.items()):
        ledgers = {}
        for order, led, narr in rows:
            ledgers.setdefault(_norm_ledger(led), []).append((order, led, narr))
        if len(ledgers) > 1:
            latest = {n: max(v)[0] for n, v in ledgers.items()}
            conflicts.append((f"{kt}:{key}", sorted(latest.items(), key=lambda x: -x[1])))
        # winner = ledger seen in the most recent month
        best = max(rows, key=lambda r: r[0])
        seed.append({
            "key_type": kt,
            "key_value": key,
            "ledger": best[1],
            "months": sorted({r[0] for r in rows}),
            "observations": len(rows),
            "source_narration": best[2][:160],
        })

    with open(args.out, "w", encoding="utf-8") as fh:
        json.dump(seed, fh, indent=2, ensure_ascii=False)

    print(f"files={len(files)}  tables={scanned}  booked_rows={booked}")
    print(f"seed entries: {len(seed)}  ->  {args.out}")
    if conflicts:
        print(f"\nconflicts resolved to most-recent month ({len(conflicts)}):")
        for key, opts in conflicts:
            shown = ", ".join(f"{n!r}@m{o}" for n, o in opts)
            print(f"  {key:34s} {shown}")
    return 0


if __name__ == "__main__":
    sys.exit(main())
