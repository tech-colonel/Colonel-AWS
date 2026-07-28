"""
Universal Indian Bank Statement Classifier for Tally Ledger Entry Automation.

Supports any Indian bank (Canara, Kotak, HDFC, SBI, Axis, ICICI, Yes Bank, RBL, IDFC, etc.)
and any D2C brand. Only input required: a List of Ledgers Tally export + raw bank statement.

Usage:
    python classify.py --ledger <path> --bank <path> --output <path> [--brand <BrandName>]
"""

import os
import re
import collections
import sys
import pandas as pd
from thefuzz import process, fuzz
import openpyxl
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.utils import get_column_letter

# ---------------------------------------------------------------------------
# D2C Ecosystem keyword map — used to scan ledger master for matching ledgers.
# Keys are canonical brand names; values are narration search terms (lowercase).
# ---------------------------------------------------------------------------
D2C_KEYWORDS = {
    "amazon":        ["amazon", "amzn", "amazon seller", "amazon pay", "intermedier"],
    "blinkit":       ["blinkit", "grofers", "blinkcommerce", "blinkcommer"],
    "zepto":         ["zepto", "kiranakart"],
    "razorpay":      ["razorpay", "razor pay"],
    "shiprocket":    ["shiprocket", "shiprock", "bigfoot"],
    "facebook":      ["facebook", "meta ads", "fb ads", "facebook india"],
    "google":        ["google", "gpay", "google ads", "adwords"],
    "cashfree":      ["cashfree"],
    "payu":          ["payu", "pay u"],
    "ccavenue":      ["ccavenue", "ccavenuef"],
    "paytm":         ["paytm", "one97"],
    "phonepe":       ["phonepe", "phone pe"],
    "jiomart":       ["jiomart", "reliance retail", "reliance jio"],
    "flipkart":      ["flipkart", "fkrt"],
    "bigbasket":     ["bigbasket", "innovative retail"],
    "swiggy":        ["swiggy", "bundl technologies"],
    "zomato":        ["zomato"],
    "bluedart":      ["bluedart", "blue dart"],
    "dtdc":          ["dtdc"],
    "gati":          ["gati"],
    "safexpress":    ["safexpress", "safe express"],
    "myntra":        ["myntra"],
    "nykaa":         ["nykaa"],
    "ajio":          ["ajio"],
    "cred":          ["cred", "dreamplug"],
    "shipway":       ["shipway"],
    "easyecom":      ["easyecom"],
    "unicommerce":   ["unicommerce"],
    "eunimart":      ["eunimart"],
    "shopify":       ["shopify"],
    "instamojo":     ["instamojo"],
    "pine labs":     ["pine labs", "pinelabs"],
    "billdesk":      ["billdesk", "bill desk"],
    "delhivery":     ["delhivery"],
    "ekart":         ["ekart"],
    "xpressbees":    ["xpressbees"],
    "shadowfax":     ["shadowfax"],
    "nykaa fashion": ["nykaa fashion", "fsn ecommerce"],
    "tata cliq":     ["tata cliq", "tata digital"],
    "meesho":        ["meesho", "fashnear"],
    "ugvcl":         ["ugvcl", "uttar gujrat vij"],
    "torrent power": ["torrent power"],
    "tata power":    ["tata power"],
    "bescom":        ["bescom"],
    "msedcl":        ["msedcl"],
    "uppcl":         ["uppcl"],
}

# Statutory BDP service codes → fuzzy search terms against the ledger master
BDP_STATUTORY = {
    "TIN":  "tds payable",
    "GSTN": "gst payable",
    "ESI":  "esic payable",
    "EPF":  "provident fund payable",
    "PT":   "professional tax payable",
}

# Common Indian business name abbreviations → expanded forms.
# Normalizing before fuzzy matching bridges mismatches like "PVT LTD" vs "Private Limited".
_ABBREV = [
    (re.compile(r'\bPVT\.?\b'),   'PRIVATE'),
    (re.compile(r'\bLTD\.?\b'),   'LIMITED'),
    (re.compile(r'\bCO\.?\b'),    'COMPANY'),
    (re.compile(r'\bCORP\.?\b'),  'CORPORATION'),
    (re.compile(r'\bMFG\.?\b'),   'MANUFACTURING'),
    (re.compile(r'\bMKTG\.?\b'),  'MARKETING'),
    (re.compile(r'\bINTL\.?\b'),  'INTERNATIONAL'),
    (re.compile(r'\bINDUS\.?\b'), 'INDUSTRIES'),
    (re.compile(r'\bEXPRES\.?\b'),'EXPRESS'),
    # Truncation expansions common in bank 50-char limits
    (re.compile(r'\bLIMI\b'),     'LIMITED'),
    (re.compile(r'\bPRIV\b'),     'PRIVATE'),
]

# Words that can be inadvertently merged by bank systems (e.g. "TRADESPRIVATE").
# Split them back before fuzzy matching.
_DEMERGE_PAT = re.compile(
    r'(?<=[A-Z])(PRIVATE|LIMITED|VENTURES|INDUSTRIES|MARKETING|LOGISTICS|SERVICES|'
    r'SOLUTIONS|TECHNOLOGIES|ENTERPRISES|INTERNATIONAL|PRODUCTS|EXPORTS|IMPORTS)')


def _expand_abbrevs(text: str) -> str:
    """Expand common business abbreviations and de-merge accidentally merged words
    (e.g. 'TRADESPRIVATE' → 'TRADES PRIVATE') for more accurate fuzzy matching."""
    t = text.upper()
    t = _DEMERGE_PAT.sub(r' \1', t)   # split merged words first
    for pat, repl in _ABBREV:
        t = pat.sub(repl, t)
    return t


# Generic business-suffix words that carry no identity. Two ledgers sharing ONLY these
# (e.g. "UDM Enterprises" vs "M K Enterprises") must NOT be treated as a confident match —
# token_set_ratio inflates such pairs. Used by the generic-token guard in _fuzzy_match.
_GENERIC_BIZ_WORDS = {
    'enterprises', 'enterprise', 'services', 'service', 'foods', 'food', 'traders', 'trader',
    'private', 'limited', 'pvt', 'ltd', 'llp', 'and', 'co', 'company', 'the', 'industries',
    'industry', 'products', 'product', 'solutions', 'solution', 'india', 'indian', 'inc',
    'corporation', 'corp', 'logistics', 'logistic', 'technologies', 'technology', 'global',
    'retail', 'online', 'store', 'sons', 'son', 'bros', 'brothers', 'group', 'international',
    'intl', 'trading', 'exports', 'imports', 'distributors', 'distributor', 'agencies',
    'agency', 'marketing', 'sales', 'a/c', 'ac',
}


def _distinctive_tokens(s: str) -> list:
    """Identity-bearing tokens of a name: length>=3, not a generic business-suffix word."""
    s = re.sub(r'[^a-z0-9 ]', ' ', str(s).lower())
    return [w for w in s.split() if len(w) >= 3 and w not in _GENERIC_BIZ_WORDS]


# Words stripped when cleaning narrations for generic fuzzy matching
NOISE_WORDS = {
    "NEFT", "RTGS", "IMPS", "UPI", "MB", "TO", "FROM", "DR", "CR", "BANK",
    "TRANSFER", "REF", "REFERENCE", "UTR", "ID", "TXN", "BRANCH",
    "IB", "SC", "OTHER", "THAN", "SB", "IMB", "CNRBH", "ONLINE", "TRANSACTION",
    "BDP", "PAYMENT", "AGAINST", "INVOICE", "DUE", "CLEARED", "THROUGH",
    "BENEFICIARY", "ACCOUNT", "NSDL", "CHALLAN", "CHALLAN NO",
}


# ---------------------------------------------------------------------------
# Narration cleaning and parsing helpers
# ---------------------------------------------------------------------------

def clean_narration(narration: str) -> str:
    """Strip banking noise words, long numbers, and special chars for fuzzy matching."""
    if not narration or not isinstance(narration, str):
        return ""
    text = narration.upper()
    text = re.sub(r'\d{10,}', ' ', text)           # remove account numbers
    text = re.sub(r'[^A-Z0-9\s]', ' ', text)
    words = text.split()
    kept = [w for w in words
            if w not in NOISE_WORDS
            and not (any(c.isdigit() for c in w) and len(w) > 6)]
    return " ".join(kept)


def _parse_ib_neft(text: str) -> tuple:
    """
    Parse Canara Bank IB NEFT / SC NEFT / IB IFT narrations.
    Format: IB NEFT Dr CNRBH<UTR> <Entity Name>  <IFSC> <AccNo> <Remark>
    Entity = between UTR and IFSC code (separated by double-space).
    Returns (entity, remark).
    """
    orig = text.strip()
    ifsc_match = re.search(r'([A-Z]{4}0[A-Z0-9]{6})', orig)
    utr_match  = re.search(r'CNRBH[A-Z0-9]+', orig)

    entity = ""
    remark = ""

    if utr_match and ifsc_match and ifsc_match.start() > utr_match.end():
        entity = orig[utr_match.end():ifsc_match.start()].strip()
        remark = orig[ifsc_match.end():].strip()

    if not entity:
        # Fallback: strip prefix then split on double-space
        text_clean = re.sub(
            r'^(IB|SC)?\s*(NEFT|RTGS|IMPS|IFT|OAT|ITG)\s*(Dr|Cr)?\s*', '', orig, flags=re.I)
        text_clean = re.sub(r'^CNRBH[A-Z0-9]+\s+', '', text_clean, flags=re.I)
        parts = re.split(r'\s{2,}', text_clean)
        entity = parts[0].strip() if parts else text_clean
        remark = " ".join(parts[1:]).strip() if len(parts) > 1 else ""

    return entity, remark


def _parse_bdp(text: str) -> tuple:
    """
    Parse BDP payment gateway narrations.
    Format: BDP-<SERVICE>-<GATEWAY>-<txnid>-<clientref>
    Returns (service_code, gateway_code).
    """
    match = re.search(r'BDP-([A-Z0-9_]+)(?:-([A-Z0-9_]+))?', text)
    if match:
        return match.group(1), (match.group(2) or "")
    return "", ""


def _parse_upi(text: str) -> str:
    """Extract beneficiary name from UPI narration (4th slash-segment)."""
    parts = text.split('/')
    if len(parts) >= 4:
        name = parts[3].strip()
        name = re.sub(r'[*]{0,2}\d+@\w+', '', name)
        name = re.sub(r'PAY TO M.*', '', name, flags=re.I)
        return name.strip()
    return ""


def _parse_chq(text: str) -> str:
    """Extract payee from cheque / MICR inward clearing narration."""
    # Try explicit PAYEE- segment first
    match = re.search(r'PAYEE-([^-]+)', text, flags=re.I)
    if match:
        return match.group(1).strip()
    parts = text.split('-')
    if len(parts) >= 3:
        payee = parts[2].strip()
        # Don't return bank branch names
        if any(k in payee.upper() for k in ["STATE BANK", "CANARA", "HDFC", "ICICI", "KOTAK", "AXIS"]):
            return ""
        return payee
    return ""


def _parse_neft_return(text: str) -> str:
    """Extract original beneficiary from NEFT RETURN narration (3rd hyphen-segment)."""
    parts = text.split('-')
    return parts[2].strip() if len(parts) >= 3 else ""


def _parse_ift(text: str) -> str:
    """
    Parse Kotak IFT (Internal Fund Transfer) narrations.
    Format: IFT-<Entity Name>-<CODE>-<TransactionRef>
    E.g.: IFT-BVC TRADEPORT PRIVATE LIM-FCM-250428GDH9CV
    Returns entity name (segment between first and second dash, before the CODE segment).
    """
    # Strip leading "IFT-" then extract up to the alphanumeric code segment
    m = re.match(r'^IFT-([A-Z][A-Z0-9\s]+?)-[A-Z]{2,5}-', text.strip())
    if m:
        return m.group(1).strip()
    # Fallback: second dash-segment
    parts = text.split('-')
    if len(parts) >= 2:
        return parts[1].strip()
    return ""


def _parse_indusind_neft(text: str) -> str:
    """
    Parse IndusInd Bank NEFT debit narrations.
    Format: N/NNNNN/ENTITY NAME/INDBHXXXXXXX/
    Entity = 3rd slash-segment (bank truncates to ~14 chars).
    """
    parts = text.split('/')
    if len(parts) >= 3:
        entity = parts[2].strip()
        # Strip the leading INDBH reference if the entity segment looks like a ref number
        if re.match(r'^INDBH\d+$', entity):
            return ""
        return entity
    return ""


def _parse_indusind_rtgs_credit(text: str) -> str:
    """
    Parse IndusInd Bank RTGS credit narrations.
    Format: R/<UTR>/<IFSC>/<SENDER NAME>//<PURPOSE>//<UTR>/
    Sender = 4th slash-segment.
    """
    parts = text.split('/')
    if len(parts) >= 4:
        sender = parts[3].strip()
        # Discard if it looks like a bank reference, not a name
        if not sender or re.match(r'^[A-Z0-9]{10,}$', sender):
            return ""
        return sender
    return ""


def _parse_indusind_bill(text: str) -> str:
    """
    Parse IndusInd Bank bill/credit card payment narrations.
    Format: BILL/<INVOICE>/<TYPE>/<ENTITY PARTIAL>/
    Entity = 4th slash-segment.
    """
    parts = text.split('/')
    if len(parts) >= 4:
        return parts[3].strip()
    return ""


# ---------------------------------------------------------------------------
# Core classifier
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Payee-identity extraction — produces STABLE keys for the per-brand payee
# directory (learned ledger memory). The same payee changes payment rail and
# reference number every month, so we key on identity, not the raw narration:
#   exact narration (machine strings) → phone (UPI) → vpa → payee name.
# Validated on Zyden April+May: phone 98% purity / 82% coverage,
# NEFT-name 100% purity (40/41 agree across months).
# ---------------------------------------------------------------------------
DIRECTORY_SECTIONS = ('exact', 'phone', 'vpa', 'neft_name', 'name')
_PHONE_RE = re.compile(r'(?<!\d)([6-9]\d{9})(?!\d)')   # Indian mobile [6-9] prefix (not NEFT refs starting w/ digits)
_VPA_RE = re.compile(r'([A-Za-z0-9.\-]+@[A-Za-z]+)')


def _norm_key(s) -> str:
    """Lowercase, strip punctuation, collapse spaces — stable comparison key."""
    return ' '.join(re.sub(r'[^a-z0-9 ]', ' ', str(s).lower()).split())


# FLO slash-format NEFT/RTGS + IMPS-FROM payee extraction (Task 2.5a).
# A "junk token" is dropped from payee candidates: it is pure-numeric or
# contains any digit (an account number, or a bank-code+digits blob like
# "UTIB0001234"), or is an exact 4-letter bank/IFSC code (e.g. UTIB, HDFC).
# Shared by the IMPS-FROM tokenizer and the slash-NEFT/RTGS field filter below.
_BANKCODE_RE = re.compile(r'^[A-Z]{4}$')


def _is_junk_token(t: str) -> bool:
    """Pure-numeric, contains any digit, or an exact 4-letter bank/IFSC code."""
    return any(c.isdigit() for c in t) or bool(_BANKCODE_RE.match(t))


def _slash_neft_payee(u: str):
    """FLO slash format: (NEFT|RTGS)/<ref>/[<BANKCODE>/]<PAYEE>[/<BANK>/<num>].
    Drop the ref field, then drop any remaining field that is itself a single
    junk token (numeric / bank-code) -- multi-word fields are never dropped
    wholesale, so an alphanumeric abbreviation inside a real payee name (e.g.
    "P2P" in "NDX P2P PRIVATE LIMITED LENDER") is preserved. The remaining
    field with the most alphabetic characters is the payee. Never derive a
    key from the numeric reference itself."""
    m = re.match(r'^(?:NEFT|RTGS)/(.+)', u)
    if not m:
        return None
    parts = [p.strip() for p in m.group(1).split('/') if p.strip()]
    if not parts:
        return None
    parts = parts[1:]                                 # drop the ref field
    candidates = [p for p in parts if ' ' in p or not _is_junk_token(p)]
    if not candidates:
        return None
    best = max(candidates, key=lambda t: sum(c.isalpha() for c in t))
    return _norm_key(best)


def _imps_from_payee(u: str):
    """IMPS <ref> FROM <PAYEE>: tokenize the trailing text on whitespace,
    drop junk tokens (numeric / bank-code), and only report a payee if
    something with an actual letter survives (never an all-numeric key,
    which would otherwise leak the account number into neft_name)."""
    m = re.search(r'\bIMPS\s+\d+\s+FROM\s+(.+)', u)
    if not m:
        return None
    survivors = [t for t in m.group(1).split() if not _is_junk_token(t)]
    if not survivors:
        return None
    joined = ' '.join(survivors)
    if not any(c.isalpha() for c in joined):
        return None
    return _norm_key(joined)


# ── Payee-key patterns for the dash/slash narration formats (ICICI, Kotak, NACH) ──
# The original extractor only recognised FLO's slash-NEFT and IMPS-FROM shapes, so
# ICICI/Kotak statements yielded nothing but the 'exact' key -- and every 'exact' key
# embeds a one-time transaction reference, so nothing a brand learned could ever match
# again. These add the shapes those banks actually emit. Each pattern below is derived
# from a real narration; see tests/test_payee_keys.py for the fixtures.

def _fields_by_alpha(fields):
    """Non-junk fields, richest in letters first. A multi-word field is never dropped
    wholesale (an alphanumeric token inside a real payee name must survive)."""
    cand = [f.strip() for f in fields if f and f.strip()]
    cand = [f for f in cand if ' ' in f or not _is_junk_token(f)]
    return sorted(cand, key=lambda t: -sum(c.isalpha() for c in t))


# Filler words that can never identify a payee. Some banks put a literal placeholder in
# the payee position ("UPI/Bank Account XX/<ref>/Gaurav may"); keying on it would map every
# unrelated payee that shares the placeholder onto one ledger -- the worst kind of wrong.
_PLACEHOLDER_WORDS = {'bank', 'account', 'accounts', 'ac', 'a', 'self', 'other', 'misc'}

# PURPOSE words. Several rails put a free-text purpose in the field BEFORE the payee:
#   INF/NEFT/<ref>/<IFSC>/FEES   /AJMERAAV        -> payee is AJMERAAV, not FEES
#   INF/NEFT/<ref>/<IFSC>/SALARY /RAJESHEMPLOYEE  -> payee is RAJESHEMPLOYEE
# Keying on the purpose collapses every fee/salary payment to ANY party onto one ledger —
# the same collision class as 'bank account xx'. Skip these and keep scanning.
# Mandate/instruction words that occupy the payee slot in a UPI narration. The real
# counterparty is the NEXT segment.
_UPI_MANDATE_WORDS = {'autopay', 'mandate', 'si', 'ach', 'nach', 'ecs', 'emandate'}

_PURPOSE_WORDS = {
    'salary', 'salaries', 'wages', 'bonus', 'advance', 'fees', 'fee', 'rent',
    'payment', 'paid', 'inv', 'invoice', 'bill', 'reimbursement', 'incentive',
    'commission', 'refund', 'balance', 'settlement', 'expense', 'charges',
}

# Bare IFSC/bank prefixes. A 4-letter token cannot be told apart from a short real payee
# name ('AZAD') by shape alone, and positional rules only catch the code when a long digit
# run follows it. A bank code that slips through is the one kind of bad key that COLLIDES
# -- every narration mentioning that bank maps onto one ledger -- so the closed set of
# Indian bank prefixes is rejected by name. Mirrored in bankCorrectionsController.js.
_BANK_PREFIXES = {
    'hdfc', 'icic', 'sbin', 'utib', 'axis', 'kkbk', 'punb', 'barb', 'ubin', 'ioba',
    'cnrb', 'idib', 'yesb', 'indb', 'ratn', 'deut', 'citi', 'hsbc', 'scbl', 'idfb',
    'fdrl', 'karb', 'tmbl', 'jaka', 'mahb', 'orbc', 'ucba', 'psib', 'cbin', 'ibkl',
    'bkid', 'aubl', 'esfb', 'usfb', 'jsfb', 'fino', 'pytm', 'airp', 'kvbl', 'svcb',
}


def _is_placeholder(nm) -> bool:
    """True when every word is filler: a generic business suffix, a placeholder word, or a
    repeated-letter stub ('x', 'xx', 'xxx')."""
    words = [w for w in str(nm).lower().split() if w]
    if not words:
        return True
    return all(w in _GENERIC_BIZ_WORDS or w in _PLACEHOLDER_WORDS
               or w in _PURPOSE_WORDS or len(set(w)) == 1
               for w in words)


def _first_alpha_field(fields):
    """First non-junk field in positional order -- for rails where the payee precedes
    the bank name (UPI/<PAYEE>/<BANK>/..., MMT/IMPS/<ref>/<utr>/<PAYEE>/<BANK>).

    Only digit-bearing fields count as junk here, NOT bare 4-letter codes: in these
    rails the bank field always carries digits (SBIN0003, UTIB0CCH274) while the payee
    may legitimately be four letters ('AZAD'), and the bank-code rule would eat it.
    Any 4-letter code that does appear (UPI/<PAYEE>/ICIC/...) sits after the payee, so
    positional order already excludes it."""
    clean = [(f or '').strip() for f in fields]
    bank_fallback = None
    for i, f in enumerate(clean):
        if not f:
            continue
        if not (' ' in f or not any(c.isdigit() for c in f)):
            continue
        if not any(c.isalpha() for c in f):
            continue
        # Skip a placeholder in the payee slot and keep scanning -- the real payee is
        # usually a later field ("UPI/Bank Account XX/<ref>/Gaurav may").
        if _is_placeholder(_norm_key(f)):
            continue
        # A known bank prefix LOSES to a real payee that follows it, so "…/HDFC/SOMEBODY"
        # yields 'somebody'. But when the bank code is the ONLY thing in the payee slot the
        # transfer really is to that bank: Urban Plant's "INF/NEFT/<ref>/HDFC0000044/HDFC"
        # is a transfer to their own HDFC account, and dropping it outright cost 14
        # correctly-classified rows. So remember it, and use it only if nothing better
        # turns up.
        if _norm_key(f) in _BANK_PREFIXES:
            if bank_fallback is None:
                bank_fallback = f
            continue
        # A bare 4-letter code followed by a long digit run is an IFSC prefix
        # ("…/UBIN/708815530180/…"), not a payee. A genuinely short payee name ("AZAD")
        # is not followed by a reference number, so it survives this test.
        nxt = clean[i + 1] if i + 1 < len(clean) else ''
        if _BANKCODE_RE.match(f.upper()) and re.fullmatch(r'\d{6,}', nxt or ''):
            continue
        return f
    return bank_fallback


def _strip_ref_tail(name) -> str:
    """Drop everything from the first purpose word or number onward.

    'Rn inv 67 balance' -> 'Rn';  'Inky ponky inv 254 272 280' -> 'Inky ponky'.
    Manual-entry narrations name the counterparty and then trail an invoice/balance
    reference. Keeping the reference makes the key single-use — it embeds numbers that
    change every month — so the directory can never generalize.
    """
    out = []
    for w in str(name).split():
        lw = re.sub(r'[^a-z0-9]', '', w.lower())
        if not lw:
            continue
        if lw in _PURPOSE_WORDS or lw.isdigit():
            break
        out.append(w)
    return ' '.join(out)


_TAG_REF_RE = re.compile(r'^[A-Z][A-Z0-9]{0,5}\s*:\s*(.+)$')


def _tagged_ref_payee(u):
    """'MB:Inky ponky inv 254 272 280' -> 'inky ponky'.

    A short alpha tag, a colon, the counterparty, then a reference tail. Without this
    the only key produced is `exact`, which embeds the invoice numbers and therefore
    never matches again.
    """
    m = _TAG_REF_RE.match(u.strip())
    if not m:
        return None
    return _norm_key(_strip_ref_tail(m.group(1)))


_PCD_RE = re.compile(r'^PCD/\d+/([^/]+)/')


def _pcd_card_merchant(u):
    """'PCD/1073/IND*LINKEDIN/MUMBAI240626/20:23' -> 'linkedin'.

    Card-terminal narration: the third slash segment is the merchant. The acquirer tag
    ('IND*') and any parenthetical qualifier ('(PGSI)') are not identity. The rest of
    the narration is city + date + time, which is why the exact key is single-use.
    """
    m = _PCD_RE.match(u.strip())
    if not m:
        return None
    merch = re.sub(r'^[A-Z]{2,4}\*', '', m.group(1))     # IND*, SBI* acquirer tag
    merch = re.sub(r'\([^)]*\)', ' ', merch)             # (PGSI)
    return _norm_key(merch)


_ACQ_STAR_RE = re.compile(r'\b[A-Z]{2,4}\*([^/]+)')


def _acquirer_star_merchant(u):
    """'VISA-REFUND/250626/250626/IND*LINKEDIN' -> 'linkedin'.

    Card acquirers tag the merchant as '<ACQ>*<MERCHANT>'. _pcd_card_merchant only sees
    it in the PCD slash layout; the same tag also appears at the tail of refund and
    reversal narrations, where everything before it is dates.
    """
    m = _ACQ_STAR_RE.search(u.strip())
    if not m:
        return None
    return _norm_key(re.sub(r'\([^)]*\)', ' ', m.group(1)))


# Only trims a genuine bank name, never a payee that merely starts with such a word:
# the bank token must be followed by BANK (or be "BANK OF ..."), so "UNION TRADERS"
# survives while "... UNION BANK OF INDIA" is cut.
_BANK_TAIL_RE = re.compile(
    r'\s+(?:(?:UNION|STATE|CENTRAL|PUNJAB|CANARA|INDIAN|ORIENTAL|CORPORATION|SYNDICATE|'
    r'HDFC|ICICI|AXIS|KOTAK|YES|IDFC|IDBI|RBL|INDUSIND|BANDHAN|FEDERAL|KARNATAKA)\s+BANK\b.*'
    r'|BANK\s+OF\b.*)$', re.I)
_CLG_RE = re.compile(r'\bCLG\s+(?:TO|FROM)\s+(.+)$')


def _clg_payee(u):
    """'CLG TO PS WAREHOUSING ENTERPR UNION BANK OF INDIA' -> 'ps warehousing enterpr'.

    Cheque-clearing narrations append the counterparty's BANK after their name. The
    bank identifies the rails, not the payee, so it is trimmed.
    """
    m = _CLG_RE.search(u.strip())
    if not m:
        return None
    return _norm_key(_BANK_TAIL_RE.sub('', m.group(1)))


def _usable_key(nm) -> bool:
    """Reject keys that cannot identify a payee: too short, all digits, a bank code,
    a noise word, or nothing but generic business-suffix words ('bank account x')."""
    if not nm or len(nm) < 3:
        return False
    if nm.replace(' ', '').isdigit():
        return False
    # NOTE: deliberately NO bare 4-letter bank-code rejection here. Every producer above
    # either drops bank codes via _is_junk_token or picks the payee positionally, so a
    # 4-letter key reaching this point is a real short payee name ('AZAD'), not an IFSC.
    if nm.upper() in NOISE_WORDS:
        return False
    if _is_placeholder(nm):
        return False
    return True


def _dash_neft_payee(u: str):
    """ICICI: NEFT-<REF>-<PAYEE>-<digits>-<digits>  /  RTGS-<REF>-<PAYEE>-...
    Drop the leading rail word and the reference field, then take the field with the
    most letters (the trailing fields are numeric ids)."""
    m = re.match(r'^(?:NEFT|RTGS)-[A-Z0-9]+-(.+)$', u)
    if not m:
        return None
    best = _fields_by_alpha(m.group(1).split('-'))
    return _norm_key(best[0]) if best else None


def _space_neft_payee(u: str):
    """Kotak: 'NEFT <REF> <PAYEE...>' / 'RTGS-<REF>-<PAYEE>' with no trailing tail.
    Tokenize and drop junk tokens (the reference carries digits)."""
    m = re.match(r'^(?:NEFT|RTGS)\s+(.+)$', u)
    if not m:
        return None
    survivors = [t for t in m.group(1).split() if not _is_junk_token(t)]
    joined = ' '.join(survivors)
    if not any(c.isalpha() for c in joined):
        return None
    return _norm_key(joined)


def _slash_rail_payee(u: str):
    """Slash rails where the payee is the first non-junk field after the rail prefix:
      MMT/IMPS/<ref>/<utr>/<PAYEE>/<BANK>      (ICICI IMPS)
      INF/NEFT/<ref>/<IFSC>/<PAYEE>/<purpose>  (ICICI netbanking)
      BIL/ONL/<ref>/<VENDOR>                   (ICICI bill-pay)
      UPI/<PAYEE>/<BANK>/<ref>/<note>          (Kotak UPI)
      INF/INFT/<ref>/<code>/<PAYEE>            (ICICI internal transfer)"""
    m = re.match(r'^(?:MMT/IMPS|INF/(?:NEFT|RTGS|IMPS|INFT)|BIL/ONL|UPI)/(.+)$', u)
    if not m:
        return None
    best = _first_alpha_field(m.group(1).split('/'))
    return _norm_key(best) if best else None


# ── HDFC rails ───────────────────────────────────────────────────────────────
# HDFC statements use several shapes none of the rails above cover. Measured on a real
# 2026-06 Zaydn statement: only 63.9% of narrations produced a reusable key, and the gap
# was almost entirely NACH loan repayments — so every loan row had to be re-corrected by
# the accountant every month and could never be learned.

def _tpt_payee(u: str):
    """<account>-TPT-<ref>-<PAYEE>  (HDFC internal transfer). Payee is the last field."""
    m = re.match(r'^\d{6,}-TPT-[A-Z0-9]+-(.+)$', u)
    return _norm_key(m.group(1)) if m else None


def _ach_payee(u: str):
    """ACH D- <BILLER>-<mandate ref>  — NACH loan/mandate debit. The biller is the field
    after the rail word; the trailing field is a per-instalment reference."""
    m = re.match(r'^ACH\s+D-\s*(.+)$', u)
    if not m:
        return None
    fields = [f.strip() for f in m.group(1).split('-') if f.strip()]
    return _norm_key(fields[0]) if fields else None


def _ft_payee(u: str):
    """FT- <tag>-<account> - <PAYEE>  /  FT - DR - <account> - <PAYEE>. Payee is last."""
    m = re.match(r'^FT\s*-\s*(.+)$', u)
    if not m:
        return None
    fields = [f.strip() for f in m.group(1).split('-') if f.strip()]
    for f in reversed(fields):
        if any(c.isalpha() for c in f) and not any(c.isdigit() for c in f):
            return _norm_key(f)
    return None


def _imps_dash_payee(u: str):
    """IMPS-<ref>-<PAYEE>-<BANK>-<acct>-<note>. Payee is the field after the reference."""
    m = re.match(r'^IMPS-\d+-([^-]+)-', u)
    return _norm_key(m.group(1)) if m else None


def _neft_dash_name(u: str):
    """NEFT|RTGS (DR|CR)-<IFSC>-<PAYEE>-… without the '-NETBANK' tail the older pattern
    requires (e.g. '…-BLOCK POOL TECHNOLOGIES PRIVATE LIMITED-ZAYDN-IN426…')."""
    m = re.match(r'^(?:NEFT|RTGS)\s+(?:DR|CR)-[A-Z0-9]+-([^-]+)', u)
    return _norm_key(m.group(1)) if m else None


def _slash_merchant(u: str):
    """<gateway ref>/<MERCHANT-CODE> — HDFC biller/aggregator rail. The merchant code is
    the stable half ('DHDF91Y1LQ1SU2/BILLDKGOOGLEADS' -> billdkgoogleads); the leading
    reference changes every transaction, which is why the whole string was useless."""
    m = re.match(r'^[A-Z0-9]+/([A-Z][A-Z0-9]{5,})$', u)
    return _norm_key(m.group(1)) if m else None


def _prefix_merchant(u: str):
    """<MERCHANT>_<ref>_<ref> — e.g. 'BAJAJFINOTP_BFL15092537202_174550516'."""
    m = re.match(r'^([A-Z]{5,})_[A-Z0-9]+_', u)
    return _norm_key(m.group(1)) if m else None


def _label_rail(u: str):
    """Fixed-label rails where the narration IS the identity, not a payee:
    'ACH DEBIT RETURN CHARGES <date>-<ref>' and 'CBDT/BANK REFERENCE NO:…'."""
    if re.match(r'^ACH\s+DEBIT\s+RETURN\s+CHARGES', u):
        return 'ach debit return charges'
    if re.match(r'^CBDT/', u):
        return 'cbdt'
    return None


def _funds_transfer_payee(u: str):
    """Kotak plain-language rail: 'FUNDS TRANSFER TO <PAYEE>' / '... FROM <PAYEE>'."""
    m = re.search(r'\bFUNDS\s+TRANSFER\s+(?:TO|FROM)\s+(.+)$', u)
    if not m:
        return None
    survivors = [t for t in m.group(1).split() if not _is_junk_token(t)]
    joined = ' '.join(survivors)
    if not any(c.isalpha() for c in joined):
        return None
    return _norm_key(joined)


def _nach_counterparty(u: str):
    """NACH-<n>-<DR|CR>-<SPONSOR>-<COUNTERPARTY>. The sponsor is the collecting bank or
    aggregator (e.g. RAZORPAYSOFTWAREPRIV); the real counterparty is the LAST field.
    Keying on the sponsor is what made a Strategic Finvest EMI look like a Razorpay row."""
    m = re.match(r'^NACH-\d+-(?:DR|CR)-(.+)$', u)
    if not m:
        return None
    fields = [f.strip() for f in m.group(1).split('-') if f.strip()]
    for f in reversed(fields):
        if not any(c.isalpha() for c in f):
            continue
        # Banks often concatenate a per-instalment mandate reference onto the counterparty
        # with no separator: STRATEGICFT6FWSD8 / STRATEGICFT3UYYAU / STRATEGICFSXWJMCE are
        # the same payee on three different months. Such a field yields a different key
        # every time, so emit nothing and let the side rule's substring token handle it.
        if ' ' not in f and len(f) > 12:
            return None
        return _norm_key(f)
    return None


def extract_payee_keys(narration) -> dict:
    """Return the stable identity keys extractable from a bank narration.
    Any subset of: exact, phone, vpa, name (UPI payee), neft_name."""
    if not narration:
        return {}
    raw = str(narration).strip()
    u = raw.upper()
    keys = {'exact': ' '.join(u.split())}            # normalized full narration
    mph = _PHONE_RE.search(raw)
    if mph:
        keys['phone'] = mph.group(1)
    _name_is_merchant = False
    if 'UPI' in u:
        mv = _VPA_RE.search(raw)
        if mv:
            keys['vpa'] = mv.group(1).lower()
        # First segment after "UPI-" is normally the payee — but for a mandate debit it
        # is the MANDATE TYPE and the merchant follows it:
        #   UPI-AUTOPAY-APPLE MEDIA SERVICES-...   -> 'apple media services', not 'autopay'
        # Keying on 'autopay' made Apple, Airtel and Adobe look like one payee.
        mn = re.match(r'\s*UPI-(.+?)-', u)
        if mn and _norm_key(mn.group(1)) in _UPI_MANDATE_WORDS:
            mn = re.match(r'\s*UPI-[^-]+-(.+?)-', u)  # skip the mandate word
            _name_is_merchant = True                  # what follows a mandate is a business
        if mn:
            nm = _norm_key(mn.group(1))
            if nm and not nm.isdigit():
                keys['name'] = nm
    mnft = re.search(r'(?:NEFT|RTGS)\s+(?:DR|CR)-[A-Z0-9]+-(.+?)-(?:NETBANK|NB[ ,]|NB$)', u)
    if mnft:                                          # NEFT/RTGS payee, between IFSC and NETBANK tail
        nm = _norm_key(mnft.group(1))
        if nm:
            keys['neft_name'] = nm
    if 'neft_name' not in keys:                       # FLO slash-NEFT/RTGS + IMPS-FROM (Task 2.5a)
        nm = _slash_neft_payee(u) or _imps_from_payee(u)
        if nm:
            keys['neft_name'] = nm
    # Dash/slash rails (ICICI, Kotak, NACH). Tried only after the patterns above so no
    # brand that already produces keys can change behaviour.
    # Take the first candidate that is actually USABLE, not merely the first that is
    # non-None. A plain `a or b or c` chain short-circuits on a truthy-but-useless result
    # and never reaches a later rail that would have worked: 'NEFT CR-<IFSC>-BLOCK POOL
    # TECHNOLOGIES…' yielded 'technologies private' (all generic words -> rejected), which
    # then blocked the rail that correctly returns 'block pool technologies private limited'.
    def _first_usable(*candidates):
        for fn in candidates:
            nm = fn(u)
            if nm and _usable_key(nm):
                return nm
        return None

    if 'neft_name' not in keys:
        nm = _first_usable(_dash_neft_payee, _space_neft_payee, _neft_dash_name)
        if nm:
            keys['neft_name'] = nm
    if 'name' not in keys:
        # HDFC rails last: they are the most shape-specific, so anything an earlier,
        # broader pattern already resolved keeps the key it had.
        nm = _first_usable(_nach_counterparty, _slash_rail_payee, _funds_transfer_payee,
                           _tpt_payee, _ach_payee, _ft_payee, _imps_dash_payee,
                           _slash_merchant, _prefix_merchant, _label_rail,
                           # Appended last so no narration an existing rail already
                           # resolves can change key. These three shapes previously
                           # produced ONLY an `exact` key — which embeds invoice
                           # numbers, dates and times, so it matched once and never
                           # again (the M Brands regression, 2026-07-28).
                           _tagged_ref_payee, _pcd_card_merchant, _clg_payee,
                           _acquirer_star_merchant)
        if nm:
            keys['name'] = nm
    # A single-token first name is not an identity when we already hold a precise one.
    # Two different people called SHIVAM (VPAs golutomar284202@oksbi and 8826965849@ibl,
    # one a materials supplier, one reimbursed for travel) collapsed onto name|shivam and
    # would have been assigned the same ledger at High confidence. The vpa/phone keys
    # already distinguish them, so the bare name adds collision risk and no recall.
    # Kept when it is the ONLY identifier available, and multi-word names are untouched
    # ('mohammad sahil' is what unifies his @ptyes and @axl handles).
    # A single-token MERCHANT ('AIRTEL') is a reliable identity; a single-token PERSON
    # ('SHIVAM') is not. Merchants also change payment aggregator more often than people
    # change UPI app, so for them the name outlives the VPA and is the key worth keeping.
    if ('name' in keys and not _name_is_merchant
            and len(str(keys['name']).split()) == 1
            and (keys.get('vpa') or keys.get('phone'))):
        del keys['name']

    # Final guard: never emit an identity key that cannot identify anyone.
    for sec in ('name', 'neft_name'):
        if sec in keys and not _usable_key(keys[sec]):
            del keys[sec]
    return keys


class BankClassifier:
    def __init__(self, master_ledgers: list, corrections: dict | None = None,
                 brand_name: str = "", side_map: list | None = None):
        self.master_ledgers = [str(l).strip() for l in master_ledgers
                               if pd.notna(l) and str(l).strip()]
        # Per-brand learned payee directory. Accepts either the legacy flat
        # {normalized_narration → {"ledger","type"}} dict (treated as the 'exact'
        # section) or a keyed dict {section → {key → {"ledger","type"}}}.
        self.directory = self._normalize_directory(corrections or {})
        self.corrections = self.directory['exact']   # back-compat alias
        self._brand_name = brand_name
        # Optional per-brand SIDE-DEPENDENT ledger map (currently M-Brands only): a list
        # of {"tokens":[...], "credit": <ledger>, "debit": <ledger>}. When a narration
        # contains a token, the credit-side ledger is used for a Receipt and the debit-side
        # for a Payment. Checked ABOVE the payee directory. Empty for every other brand →
        # no behaviour change. Tokens are upper-cased once here for fast matching.
        self.side_map = []            # top-priority: checked ABOVE the directory
        self.side_map_fallback = []   # low-priority: checked LATE, just before Suspense, so a
                                      # broad rule (e.g. Salary→Salary Payable) never overrides a
                                      # specific per-payee/directory/fuzzy match.
        for e in (side_map or []):
            toks = [str(t).upper() for t in (e.get('tokens') or []) if str(t).strip()]
            if toks and e.get('credit') and e.get('debit'):
                entry = {'tokens': toks,
                         'credit': str(e['credit']).strip(),
                         'debit': str(e['debit']).strip(),
                         'type': (str(e['type']).strip() if e.get('type') else None)}
                (self.side_map_fallback if e.get('fallback') else self.side_map).append(entry)
        self._build_indices()

    def _coa_resolve(self, name):
        """Return the ledger's CANONICAL spelling as it appears in the COA, or None.

        Matching is case- and whitespace-insensitive, and the COA's spelling always wins.
        An accountant who retypes a ledger by hand will eventually vary the capitalisation
        ('Ria-Salary' for the COA's 'Ria-salary'); a case-sensitive check treats that as a
        ledger that does not exist and silently drops the correction, so the same row gets
        re-corrected every month and never learns. Snapping to the COA spelling also stops
        two casings of one ledger being stored as two separate learned entries.
        """
        if not name:
            return None
        name = str(name).strip()
        if name in self.master_ledgers:            # exact hit, the common case
            return name
        if not hasattr(self, '_coa_ci'):
            self._coa_ci = {}
            for l in self.master_ledgers:          # first spelling wins, stable ordering
                self._coa_ci.setdefault(' '.join(l.lower().split()), l)
        return self._coa_ci.get(' '.join(name.lower().split()))

    def _side_map_lookup(self, narration_upper: str, txn_type: str):
        """Side-dependent ledger override (per-brand, e.g. M-Brands marketplaces).
        If the narration contains a counterparty token, return its credit-side ledger
        for a Receipt / debit-side ledger for a Payment — but only if that ledger still
        exists in the brand COA (else skip so the row falls through to the directory)."""
        return self._match_side(self.side_map, narration_upper, txn_type, "Side Ledger (credit/debit)")

    def _match_side(self, entries, narration_upper, txn_type, rule):
        if not entries:
            return None
        for e in entries:
            if any(tok in narration_upper for tok in e['tokens']):
                ledger = self._coa_resolve(e['credit'] if txn_type == "Receipt" else e['debit'])
                if ledger:
                    # 'credit'/'debit' are carried through so the constrained side-verdict
                    # pass can offer Claude exactly these two ledgers and nothing else.
                    return {"ledger": ledger, "type": e.get('type') or txn_type, "confidence": "High",
                            "rule": rule, "entity": e['tokens'][0],
                            "credit": e['credit'], "debit": e['debit']}
        return None

    @staticmethod
    def _normalize_directory(corr: dict) -> dict:
        out = {s: {} for s in DIRECTORY_SECTIONS}
        if not isinstance(corr, dict) or not corr:
            return out
        keyed = all(k in DIRECTORY_SECTIONS and isinstance(v, dict) for k, v in corr.items())
        if keyed:
            for s in DIRECTORY_SECTIONS:
                out[s] = dict(corr.get(s, {}))
        else:
            out['exact'] = dict(corr)               # legacy flat → exact section
        # normalize the exact-section keys (uppercase, collapse whitespace)
        out['exact'] = {' '.join(str(k).upper().split()): v for k, v in out['exact'].items()}
        return out

    def _directory_lookup(self, narration: str, txn_type: str):
        """Look up the row's payee identity in the learned directory.
        Most-specific-first: exact → phone → vpa → neft_name → name. Only returns
        a ledger that still exists in the current COA (stale corrections are skipped)."""
        if not any(self.directory.values()):
            return None
        keys = extract_payee_keys(narration)
        for section in DIRECTORY_SECTIONS:
            kv = keys.get(section)
            if not kv:
                continue
            fix = self.directory.get(section, {}).get(kv)
            resolved = self._coa_resolve(fix.get('ledger')) if fix else None
            if resolved:
                return {
                    "ledger": resolved,
                    "type": fix.get('type') or txn_type,
                    "confidence": "High",
                    "rule": "Stored Correction" if section == 'exact' else f"Payee Directory ({section})",
                    "entity": kv,
                }
        return None

    def _build_indices(self):
        """Pre-compute filtered ledger subsets for fast routing."""
        ml = self.master_ledgers

        # Salary / payroll
        self.salary_ledgers = [l for l in ml if any(
            k in l.lower() for k in
            ['salary', 'stipend', 'wages', 'remuneration', 'allowance', 'director sitting']
        )]
        self.salary_ledgers_broad = [l for l in ml if any(
            k in l.lower() for k in
            ['salary', 'stipend', 'wages', 'employee', 'payable', 'staff', 'reimbursement']
        )]

        # Statutory
        self.tds_ledgers   = [l for l in ml if 'tds' in l.lower() or 'tax deducted' in l.lower()]
        self.gst_ledgers   = [l for l in ml if any(
            k in l.lower() for k in ['gst', 'goods & service', 'cgst', 'sgst', 'igst']
        )]
        self.pf_ledgers    = [l for l in ml if 'provident' in l.lower() or 'pf' in l.lower() or 'epf' in l.lower()]
        self.esic_ledgers  = [l for l in ml if 'esic' in l.lower() or 'esi' in l.lower()]
        self.pt_ledgers    = [l for l in ml if 'professional tax' in l.lower()]

        # Bank charges & interest — broader matching to catch naming variations across firms
        self.bank_charge_ledgers = [l for l in ml if any(
            k in l.lower() for k in [
                'bank charge', 'bank fees', 'bank commission', 'bank expense',
                'neft charge', 'rtgs charge', 'imps charge', 'dd charge',
                'service charge', 'transaction charge', 'amb charge', 'penal charge',
                'cheque return charge', 'chq return charge', 'ecs return', 'nach return',
            ]
        ) and not any(
            k in l.lower() for k in [
                'amazon', 'flipkart', 'razorpay', 'shiprocket', 'zepto',
                'blinkit', 'swiggy', 'zomato', 'myntra', 'nykaa', 'meesho',
                'paytm', 'phonepe', 'cashfree', 'payu', 'ccavenue',
            ]
        )]
        self.interest_exp_ledgers = [l for l in ml if any(
            k in l.lower() for k in [
                'interest exp', 'interest paid', 'finance cost', 'finance charge',
                'interest on od', 'interest on cc', 'interest on loan',
                'od interest', 'cc interest', 'bank interest exp',
                'interest payable', 'interest capitali',
            ]
        )]
        self.interest_inc_ledgers = [l for l in ml if any(
            k in l.lower() for k in [
                'interest inc', 'interest rec', 'interest earn', 'interest on fdr',
                'interest on fd', 'fd interest', 'interest on deposit',
                'interest on saving', 'interest credit', 'interest income',
            ]
        )]

        # Pre-compute suspense ledger. Convention (confirmed with the firm): unclassified
        # rows are labelled "Suspense A/c". Only reuse a COA ledger that is EXACTLY named
        # "Suspense A/c" / "Suspense" — do NOT borrow a substring match like "Suspense
        # Purchase" (a different, purchase-side ledger). Otherwise emit the literal label.
        self._suspense_ledger = (
            next((l for l in ml if l.strip().lower() == 'suspense a/c'), None)
            or next((l for l in ml if l.strip().lower() == 'suspense'), None)
            or "Suspense A/c"
        )

        # Own-brand guard: identify which COA ledger IS the brand's own company account.
        # Narrations like "INDIAIDEAS.COM LIMITED-PA ESCROW-ZAYDN SNEAKERS PVTLIM" embed
        # the brand's name as a beneficiary label, not the counterparty. We find this ledger
        # once and exclude it from fuzzy candidates so it can never win a contested match.
        self._own_brand_ledger = None
        if self._brand_name:
            _stop = {'private', 'limited', 'pvt', 'ltd', 'llp', 'and', 'co',
                     'company', 'india', 'the', 'of', 'for'}
            _brand_toks = {w for w in re.sub(r'[^a-z0-9 ]', ' ',
                           self._brand_name.lower()).split()
                           if len(w) >= 3 and w not in _stop}
            if _brand_toks:
                best_l, best_n = None, 0
                for l in self.master_ledgers:
                    l_toks = set(re.sub(r'[^a-z0-9 ]', ' ', l.lower()).split())
                    n = len(_brand_toks & l_toks)
                    if n > best_n and n >= max(2, len(_brand_toks) - 1):
                        best_n, best_l = n, l
                self._own_brand_ledger = best_l

        # Rent
        self.rent_ledgers = [l for l in ml
                             if ('rent' in l.lower() or 'godown' in l.lower() or 'lease' in l.lower())
                             and 'current' not in l.lower()]

        # Own-account bank ledgers (for Contra detection)
        self.own_account_ledgers = []
        self.own_account_numbers = set()
        skip_terms = {'charge', 'interest', 'commission', 'fees', 'expense',
                      'brand funded', 'marketing', 'received', 'income', 'tax', 'tds'}
        for l in ml:
            ll = l.lower()
            if any(k in ll for k in ['bank', 'cash', 'escrow', 'current a/c', 'od account', 'casa']):
                if not any(k in ll for k in skip_terms):
                    self.own_account_ledgers.append(l)
                    for num in re.findall(r'\d{8,}', l):
                        self.own_account_numbers.add(num)

        # D2C ledger subsets: for each brand keyword, find matching ledgers in master.
        # Use WORD-BOUNDARY matching for ledger lookup to avoid short terms like "cred"
        # matching "creditors", "ucredits", etc.
        self._d2c_ledger_map = {}
        for brand, terms in D2C_KEYWORDS.items():
            subset = [
                l for l in ml
                if any(re.search(r'\b' + re.escape(t) + r'\b', l.lower()) for t in terms)
            ]
            if subset:
                self._d2c_ledger_map[brand] = subset

    # ------------------------------------------------------------------
    def _find_master(self, keywords: list) -> str | None:
        """Find the first ledger in master that contains ANY of the given keywords (substring).

        This ensures returned ledger names always come from the user's actual chart of accounts,
        never hardcoded strings that might not exist in their master.
        """
        for kw in keywords:
            found = next((l for l in self.master_ledgers if kw.lower() in l.lower()), None)
            if found:
                return found
        return None

    # ------------------------------------------------------------------
    def _fuzzy_match(self, query: str, choices: list, threshold: int = 72,
                     guard_generic: bool = False,
                     exclude: str | None = None) -> tuple:
        """
        Returns (matched_ledger, confidence_band, score).
        confidence_band: 'High' (≥87), 'Medium' (72–86), 'Low' (<72).

        exclude: if set, this ledger is removed from candidates before scoring — used
        by the own-brand guard to prevent the brand's own company ledger from winning.

        Scoring strategy:
        - Expands common abbreviations (PVT→PRIVATE, LTD→LIMITED) before matching so
          bank narrations like "EKANEK NETWORKS PVT LTD" correctly match "Ekanek Network
          Private Limited" instead of a shorter ledger with a shared common word.
        - Uses a weighted blend: 0.6 × token_set_ratio + 0.4 × token_sort_ratio.
          Pure token_set_ratio inflates scores when a very short ledger shares one common
          token (e.g. "PF Charges" scoring 82 against a long narration containing "CHARGES").
          The blend preserves token_set's ability to handle word-order variation while
          giving enough weight to token_sort to penalise size mismatches.
        """
        if exclude:
            choices = [c for c in choices if c != exclude]
        if not choices or not query.strip():
            return "Suspense A/c", "Low", 0.0
        query_up = query.upper().strip()
        for c in choices:
            if c.upper().strip() == query_up:
                return c, "High", 100.0

        q_norm = _expand_abbrevs(query)

        def combined_score(choice):
            c_norm = _expand_abbrevs(choice)
            s1 = fuzz.token_set_ratio(q_norm, c_norm)
            s2 = fuzz.token_sort_ratio(q_norm, c_norm)
            # Weighted blend: token_set handles word-order freely, token_sort penalises
            # size mismatches (prevents short ledgers like "PF Charges" from scoring
            # artificially high against long narrations via a single shared token).
            # For very short queries (abbreviations like "VRL", "SBI", 3–8 chars),
            # token_sort_ratio is inherently low due to length disparity — use
            # token_set_ratio predominantly in those cases.
            if len(q_norm) <= 8:
                return round(0.85 * s1 + 0.15 * s2)
            return round(0.6 * s1 + 0.4 * s2)

        best = max(choices, key=combined_score)
        score = combined_score(best)
        conf = "High" if score >= 87 else ("Medium" if score >= 72 else "Low")

        # Generic-token guard (only for free-text matches against the full master).
        # If the query has identity-bearing tokens but shares NONE of them with the
        # winner — i.e. they overlap only on generic words like "Enterprises"/"Services" —
        # the score is inflated noise. Refuse it (return Low → caller falls to Suspense)
        # rather than emit a confident wrong vendor.
        if guard_generic and conf != "Low":
            q_dist = _distinctive_tokens(query)
            if q_dist:
                c_dist = _distinctive_tokens(best)
                shares = any(fuzz.ratio(a, b) >= 85 for a in q_dist for b in c_dist)
                if not shares:
                    return best, "Low", score

        # Downgrade to Medium only when runner-up is within 4 points — nearly tied match.
        # Gap of 8 was too aggressive; most real mismatches have a larger gap.
        if conf == "High" and len(choices) > 1:
            second = sorted([combined_score(c) for c in choices], reverse=True)
            if len(second) > 1 and (second[0] - second[1]) < 4:
                conf = "Medium"

        return best, conf, score

    def top_candidates(self, query: str, k: int = 15) -> list:
        """Candidate COA ledgers for the LLM fallback to choose from (all REAL ledgers).

        Combines two signals so the right ledger is reliably offered:
          1. Token expansion — every ledger that word-matches a distinctive (>=4 char) token
             of the query. Guarantees e.g. "NSDL New" surfaces "NSDL E-GOVERNANCE ... LIMITED",
             which pure fuzzy ranking misses (short alias vs long official name).
          2. Fuzzy top-k — nearest ledgers by the same blended score the matcher uses.
        """
        if not query or not str(query).strip():
            return []
        q = _expand_abbrevs(str(query))

        def sc(choice):
            cn = _expand_abbrevs(choice)
            s1 = fuzz.token_set_ratio(q, cn)
            s2 = fuzz.token_sort_ratio(q, cn)
            return 0.85 * s1 + 0.15 * s2 if len(q) <= 8 else 0.6 * s1 + 0.4 * s2

        fuzzy_top = sorted(self.master_ledgers, key=sc, reverse=True)[:k]

        tok_matches = []
        for t in (t for t in _distinctive_tokens(query) if len(t) >= 4):
            pat = re.compile(r'\b' + re.escape(t), re.IGNORECASE)
            tok_matches += [l for l in self.master_ledgers if pat.search(l)]

        seen, out = set(), []
        for l in tok_matches[:15] + fuzzy_top:   # token matches first (most relevant)
            if l not in seen:
                seen.add(l)
                out.append(l)
        return out[:20]

    def _frequent_ledgers(self, n: int = 14) -> list:
        """Ledgers this brand ACTUALLY uses, ranked by how often they appear in the learned
        directory. These are the accountant's working vocabulary — the category ledgers
        (Salary Payable, Raw Material, Courier & Shipping Expense) that a payee-name search
        can never surface, because a person's name shares no tokens with them."""
        if not hasattr(self, '_freq_cache'):
            cnt = collections.Counter()
            for section in self.directory.values():
                for fix in section.values():
                    led = (fix or {}).get('ledger')
                    if led:
                        cnt[led] += 1
            self._freq_cache = [l for l, _ in cnt.most_common()
                                if self._coa_resolve(l)][:n]
        return self._freq_cache

    def llm_candidates(self, query: str, narration: str = "", k: int = 15) -> list:
        """Candidate list for an LLM decision.

        top_candidates() alone scores purely on name similarity to the payee, so a
        CATEGORY ledger is never offered for a person-named narration. Measured on a real
        Zaydn statement: of 54 wrong Low/Medium rows, the correct ledger was missing from
        the 15 candidates in 34 of them — the model could not have got those right at any
        price. Widened with two signals that don't depend on the payee name:

          * ledgers the brand actually uses (learned-directory frequency)
          * COA ledgers word-matching any distinctive token of the FULL narration, not just
            the extracted entity ('…PACKAGING RAW MATE' -> 'Raw Material')

        Kept separate from top_candidates() so the arbitration pass's k=1 "best name match"
        check keeps its original, stricter meaning.

        NO CAP by default: the whole CoA is offered, RANKED — best name matches first, then
        narration-token hits, then the brand's most-used ledgers, then everything else. A
        shortlist is exactly what made the right answer unreachable, so the shortlist is
        gone; the ranking is what makes a long list usable rather than a wall of names.
        Set BANK_LLM_MAX_CANDIDATES to a positive integer to re-impose a ceiling (the
        ranking means a ceiling keeps the most plausible names).
        """
        out, seen = [], set()

        def add(items):
            for l in items:
                if l and l not in seen:
                    seen.add(l)
                    out.append(l)

        add(self.top_candidates(query, k))          # 1. best name matches

        if narration:                                # 2. narration-token hits
            for t in (t for t in _distinctive_tokens(narration) if len(t) >= 4):
                pat = re.compile(r'\b' + re.escape(t), re.IGNORECASE)
                add([l for l in self.master_ledgers if pat.search(l)])

        add(self._frequent_ledgers())                # 3. the brand's working vocabulary
        add(self.master_ledgers)                     # 4. everything else

        try:
            cap = int(os.environ.get('BANK_LLM_MAX_CANDIDATES', '0') or 0)
        except ValueError:
            cap = 0
        return out[:cap] if cap > 0 else out

    def _match_employee_ledger(self, entity: str, salary_ledgers: list) -> str:
        """
        Name-aware employee matching.
        Handles Indian name variations (order reversal, suffixes like 'bhai', 'kumar').
        Requires first-name overlap; prefers ledgers with 'salary' in the name.
        """
        noise = {'salary', 'payable', 'dr', 'cr', 'a/c', 'account', 'pvt', 'ltd', 'private', 'limited', 'factory'}
        suffixes = ['ben', 'bhai', 'kumar', 'kumari', 'lal', 'devi', 'prasad', 'singh', 'bai']

        def stem(word):
            for s in suffixes:
                if word.endswith(s) and len(word) > len(s) + 1:
                    return word[:-len(s)]
            return word

        entity_words = [stem(w) for w in entity.lower().replace('-', ' ').split()
                        if w not in noise and len(w) > 2]
        if not entity_words:
            return None

        best_ledger, best_score = None, 0
        for ledger in salary_ledgers:
            ledger_words = [stem(w) for w in ledger.lower().replace('-', ' ').split()
                            if w not in noise and len(w) > 2]
            if not ledger_words:
                continue
            # At least one word from the entity must match the first token of ledger (or vice versa)
            first_match = any(
                fuzz.ratio(ew, ledger_words[0]) >= 80 for ew in entity_words
            ) or any(
                fuzz.ratio(entity_words[0], lw) >= 80 for lw in ledger_words
            )
            if not first_match:
                continue
            overlaps = sum(
                1 for ew in entity_words
                if any(fuzz.ratio(ew, lw) >= 80 for lw in ledger_words)
            )
            if overlaps > 0:
                score = overlaps * 100 + fuzz.token_set_ratio(entity.lower(), ledger.lower())
                # Prefer "Salary A/c" or "Salary Payable" named ledgers
                if any(k in ledger.lower() for k in ['salary a/c', 'salary payable', 'salary - ']):
                    score += 50
                if score > best_score:
                    best_score = score
                    best_ledger = ledger

        # Threshold 200 requires at least 2 word overlaps (or 1 overlap + strong
        # token_set_ratio). Score of 150 allowed single-name matches like
        # "DIVYA ROHIT SA" → "Salary - Divya Pillai" (wrong) or
        # "PAWAR RAMESH G" → "Salary Payable-Ramesh" (wrong) to slip through.
        return best_ledger if best_score >= 200 else None

    # ------------------------------------------------------------------
    def classify(self, narration: str, debit: float, credit: float,
                 skip_side_map: bool = False) -> dict:
        orig  = str(narration).strip() if narration else ""
        orig_upper = orig.upper()
        is_credit  = credit > 0
        txn_type   = "Receipt" if is_credit else "Payment"

        # ------------------------------------------------------------------
        # STEP 0 — Per-brand learned payee directory (highest priority)
        # Matches the row's payee identity (phone / vpa / payee-name / exact
        # narration) against ledgers learned from corrected history. This is what
        # resolves UPI/NEFT payments to individuals (Amit→Salary Payable, etc.)
        # whose correct ledger is NOT inferable from the narration alone.
        # CoA validation: skips entries whose ledger no longer exists in master.
        # ------------------------------------------------------------------
        # STEP 0 (pre) — per-brand SIDE-DEPENDENT ledger map (M-Brands marketplaces),
        # checked ABOVE the directory: same counterparty → credit ledger on Receipt,
        # debit ledger on Payment. Empty for other brands, so a no-op there.
        # skip_side_map=True is used to re-classify a row after Claude answered
        # NOT-THIS-VENDOR: the token matched but the counterparty is someone else, so the
        # row must fall through to the directory / COA / LLM layers exactly as if no side
        # rule existed for it.
        if not skip_side_map:
            sm = self._side_map_lookup(orig_upper, txn_type)
            if sm:
                return sm

        hit = self._directory_lookup(orig, txn_type)
        if hit:
            return hit

        # ------------------------------------------------------------------
        # STEP 0.5 — PA-ESCROW aggregator counterparty routing
        # Narrations like "NEFT CR-CITI0100000-INDIAIDEAS.COM LIMITED-PA ESCROW-<BRAND>"
        # contain the brand's own legal name as the beneficiary label, NOT the counterparty.
        # India Ideas is the payment gateway — route to their COA ledger immediately.
        # This runs before entity extraction so the brand's own-name tokens never reach
        # the fuzzy matcher and never win a contested match.
        # ------------------------------------------------------------------
        if "INDIAIDEAS" in orig_upper or "INDIA IDEAS" in orig_upper:
            india_ideas = self._find_master(['india ideas'])
            if india_ideas:
                return {"ledger": india_ideas, "type": txn_type, "confidence": "High",
                        "rule": "PA-ESCROW India Ideas"}

        # ------------------------------------------------------------------
        # STEP 1 — Own-account transfers → Contra
        # ------------------------------------------------------------------
        for acc_num in self.own_account_numbers:
            if acc_num in orig_upper:
                matched = next((l for l in self.own_account_ledgers if acc_num in l), None)
                if matched:
                    return {"ledger": matched, "type": "Contra",
                            "confidence": "High", "rule": "Own Account Transfer"}

        if "IB OAT" in orig_upper or "DRAWDOWN FROM CASA" in orig_upper:
            matched = next((l for l in self.own_account_ledgers if "bank" in l.lower()), None)
            if matched:
                return {"ledger": matched, "type": "Contra",
                        "confidence": "High", "rule": "OAT/CASA Contra"}

        # IndusInd Bank sweep transfer between OD/FD and current account.
        # Format: "Sweep Trf From <ACCTNO>///"
        # Match the account number's last 4 digits against ledger names that contain
        # "XX<last4>" (e.g. account 301051314975 → last4=4975 → IndusInd FD-XX4975).
        # Search ALL master ledgers (not just own_account) because the sweep source
        # is often an FD sub-account, not a bank current account.
        if "SWEEP TRF" in orig_upper:
            acc_m = re.search(r'\b(\d{10,})\b', orig_upper)
            matched = None
            if acc_m:
                acc_no = acc_m.group(1)
                last4  = acc_no[-4:]
                matched = (
                    next((l for l in self.master_ledgers if acc_no in l), None)
                    or next((l for l in self.master_ledgers
                             if f'XX{last4}' in l.upper() or l.upper().endswith(last4)), None)
                    or next((l for l in self.own_account_ledgers), None)
                )
            if not matched:
                matched = next((l for l in self.own_account_ledgers), None)
            if matched:
                return {"ledger": matched, "type": "Contra",
                        "confidence": "High", "rule": "IndusInd Sweep Contra"}

        # "RATNR" is RBL Bank's RTGS prefix (Ratnakar→RBL). When it appears in an RTGS
        # narration alongside the company's own name, it signals an inter-bank self-transfer.
        if "RATNR" in orig_upper and ("RTGS" in orig_upper or "NEFT" in orig_upper):
            rbl = next((l for l in self.own_account_ledgers if 'rbl' in l.lower()), None)
            if rbl:
                return {"ledger": rbl, "type": "Contra", "confidence": "High",
                        "rule": "RATNR RBL Contra"}

        # IndusInd Bank "Repayment credit [<ACCTNO>//...]" narrations.
        # These are credits back FROM a sweep/FD sub-account into the current account.
        # The narration contains "TDS Recovery" as a REMARK, NOT the ledger type —
        # matching TDS keyword here would be wrong. Use account suffix to find the
        # source FD/sub-account ledger instead.
        if "REPAYMENT CREDIT" in orig_upper:
            acc_m = re.search(r'\[(\d{10,})', orig_upper)
            if acc_m:
                acc_no = acc_m.group(1)
                last4  = acc_no[-4:]
                matched = (
                    next((l for l in self.master_ledgers if acc_no in l), None)
                    or next((l for l in self.master_ledgers
                             if f'XX{last4}' in l.upper() or l.upper().endswith(last4)), None)
                )
                if matched:
                    return {"ledger": matched, "type": txn_type, "confidence": "High",
                            "rule": "IndusInd Repayment Credit"}

        # ------------------------------------------------------------------
        # STEP 2 — Statutory payments via BDP / ITG narration codes
        # ------------------------------------------------------------------
        if "BDP-" in orig_upper or "IB ITG" in orig_upper:
            service, _ = _parse_bdp(orig_upper)
            if service in BDP_STATUTORY:
                search = BDP_STATUTORY[service]
                pool = {
                    "TIN":  self.tds_ledgers,
                    "GSTN": self.gst_ledgers,
                    "EPF":  self.pf_ledgers,
                    "ESI":  self.esic_ledgers,
                    "PT":   self.pt_ledgers,
                }.get(service, [])
                if pool:
                    ledger, conf, _ = self._fuzzy_match(search, pool)
                    if conf != "Low":
                        return {"ledger": ledger, "type": txn_type,
                                "confidence": "High", "rule": f"BDP Statutory {service}"}
                # Fallback: first matching ledger by substring — always from master
                fallback = (
                    next((l for l in self.master_ledgers if search.split()[0] in l.lower()), None)
                    or self._fuzzy_match(search, self.master_ledgers)[0]
                )
                return {"ledger": fallback, "type": txn_type,
                        "confidence": "Medium", "rule": f"BDP Statutory {service} (fallback)"}

            # NOTE: non-statutory BDP services (e.g. BDP-CRED-CASHFREE for the CRED
            # credit-card app) are intentionally NOT hardcoded here. How they map is
            # brand-dependent — Drips books CRED → "Credit Card", STROOM books it →
            # Suspense. Hardcoding one convention creates confident-WRONG entries for the
            # other brand. These are handled by the per-brand learned corrections layer
            # (validated against the live COA), not a global rule.

        # ------------------------------------------------------------------
        # STEP 3 — GSTN challan / raw GST keyword in narration
        # GSTN can appear as standalone (GSTN Payable) or with digits (GSTN1082...-...)
        # ------------------------------------------------------------------
        if "GSTN" in orig_upper:
            # Prefer exact "GST Payable" > any payable > any GST ledger
            gst = (
                next((l for l in self.gst_ledgers if l.lower().strip() in ('gst payable', 'gst payble')), None)
                or next((l for l in self.gst_ledgers if 'payable' in l.lower() and 'input' not in l.lower() and 'inward' not in l.lower()), None)
                or next((l for l in self.gst_ledgers), None)
                or self._find_master(['gst payable', 'gst liability', 'output gst', 'cgst payable', 'sgst payable', 'igst payable'])
                or self._suspense_ledger
            )
            return {"ledger": gst, "type": txn_type, "confidence": "High", "rule": "GSTN Match"}

        # ------------------------------------------------------------------
        # STEP 4 — TDS keyword
        # CBDT = Central Board of Direct Taxes (IndusInd Bank ETAX payments)
        # IB ETAX CBDT narrations are income-tax / TDS challan payments.
        # ------------------------------------------------------------------
        if re.search(r'\bTDS\b', orig_upper) or re.search(r'\bTIN\b', orig_upper) \
                or "CBDT" in orig_upper or "ETAX" in orig_upper:
            query = "TDS Receivable" if is_credit else "TDS Payable"
            if self.tds_ledgers:
                ledger, conf, _ = self._fuzzy_match(query, self.tds_ledgers)
                if conf != "Low":
                    return {"ledger": ledger, "type": txn_type, "confidence": "High", "rule": "TDS Match"}
            fallback = (
                next((l for l in self.tds_ledgers), None)
                or self._find_master(['tds payable', 'tds receivable', 'tax deducted', 'income tax'])
                or self._suspense_ledger
            )
            return {"ledger": fallback, "type": txn_type, "confidence": "Medium", "rule": "TDS Fallback"}

        # ------------------------------------------------------------------
        # STEP 5 — EPF / ESI keywords
        # Note: Kotak writes "ETAX EPFONEW" (EPF followed by letters) not bare "EPF"
        # ------------------------------------------------------------------
        if re.search(r'\bEPF', orig_upper) or "PROVIDENT FUND" in orig_upper or "EPFONEW" in orig_upper:
            # Prefer "PF Payable" or "Provident Fund Payable" over generic "PF Account"
            ledger = (
                next((l for l in self.pf_ledgers if 'payable' in l.lower()), None)
                or next((l for l in self.pf_ledgers), None)
                or self._find_master(['provident fund', 'pf payable', 'epf payable'])
                or self._suspense_ledger
            )
            return {"ledger": ledger, "type": txn_type, "confidence": "High", "rule": "EPF Match"}

        if re.search(r'\bESIC?', orig_upper):
            ledger = (
                next((l for l in self.esic_ledgers if 'payable' in l.lower()), None)
                or next((l for l in self.esic_ledgers), None)
                or self._find_master(['esic payable', 'esi payable', 'employee state insurance'])
                or self._suspense_ledger
            )
            return {"ledger": ledger, "type": txn_type, "confidence": "High", "rule": "ESIC Match"}

        # ------------------------------------------------------------------
        # STEP 6 — Rent / godown  (must run BEFORE bank charges to prevent
        #           "Periodic Godown SC charges" hitting the SC CHARGES rule).
        # IMPORTANT: Skip RENT keyword check for structured bank transfer narrations
        # (IB NEFT / IB IFT) — in those, "Rent" appears in the remark field
        # and the real ledger is the payee entity (e.g., "Indra Hassija - Rent").
        # GODOWN always signals a godown charge so it's safe regardless of format.
        # ------------------------------------------------------------------
        is_transfer_narration = any(k in orig_upper for k in ["IB NEFT", "SC NEFT", "IB IFT", "IB RTGS"])
        # Also skip the pure-RENT rule when "RENT" appears after a dash in the narration,
        # e.g. "ANU TUSHIR-ADV RENT APR25" or "EKANEK NETWORKS PVT LTD- RENT MAR25".
        # In those cases the entity precedes the dash and should be matched directly;
        # the rent remark will be caught later in STEP 12 (Rent remark routing).
        rent_in_remark = bool(re.search(r'[A-Z][A-Z\s]{2,}-\s*(?:\w+\s+)?RENT\b', orig_upper))
        if "GODOWN" in orig_upper or (re.search(r'\bRENT\b', orig_upper) and not is_transfer_narration and not rent_in_remark):
            ledger = (
                next((l for l in self.rent_ledgers if re.search(r'\brent exp', l.lower())), None)
                or next((l for l in self.rent_ledgers if 'rent' in l.lower() and 'deposit' not in l.lower() and 'allowance' not in l.lower()), None)
                or next((l for l in self.rent_ledgers), None)
                or self._find_master(['rent expense', 'rent paid', 'office rent', 'godown rent'])
                or self._suspense_ledger
            )
            return {"ledger": ledger, "type": txn_type, "confidence": "High", "rule": "Rent Match"}

        # ------------------------------------------------------------------
        # STEP 7 — Bank charges
        # ------------------------------------------------------------------
        if any(k in orig_upper for k in [
            "OTHER THAN SB IMB", "SC CHARGES", "BANK CHARGES",
            "IMPS CHARGES", "SC NEFT", "SMS CHARGES",
            "ANNUAL FEE", "AMB CHARGES",
            "CHQ DEPOSITED AND RETURN", "CHQ RETURN", "CHEQUE RETURN",
        ]) or orig_upper.startswith("CHRG:"):
            # Prefer an explicitly named "Bank Charges" ledger; fall back to "Bank Charges" text
            ledger = (
                next((l for l in self.bank_charge_ledgers if 'bank charge' in l.lower()), None)
                or next((l for l in self.master_ledgers if l.lower() == 'bank charges'), None)
                or next((l for l in self.bank_charge_ledgers), None)
                or self._find_master(['bank charge', 'bank fees', 'service charge', 'transaction charge'])
                or self._suspense_ledger
            )
            return {"ledger": ledger, "type": "Payment", "confidence": "High", "rule": "Bank Charges"}

        # ------------------------------------------------------------------
        # STEP 8 — Interest
        # Note: Kotak writes "Int.Coll:" for interest collection (credit side)
        # ------------------------------------------------------------------
        if "INTEREST" in orig_upper or "INT.COLL" in orig_upper or "INT COLL" in orig_upper:
            if is_credit:
                ledger = (
                    next((l for l in self.interest_inc_ledgers), None)
                    or self._find_master(['interest income', 'interest earned', 'interest received', 'interest on fd'])
                    or self._suspense_ledger
                )
            else:
                ledger = (
                    next((l for l in self.interest_exp_ledgers), None)
                    or self._find_master(['interest expense', 'finance cost', 'interest paid', 'interest on od'])
                    or self._suspense_ledger
                )
            return {"ledger": ledger, "type": txn_type, "confidence": "High", "rule": "Interest Match"}

        # ------------------------------------------------------------------
        # STEP 9 — NEFT Return: extract original beneficiary
        # ------------------------------------------------------------------
        if "NEFT RETURN" in orig_upper:
            entity = _parse_neft_return(orig_upper)
            if entity:
                ledger, conf, _ = self._fuzzy_match(entity, self.master_ledgers)
                return {"ledger": ledger, "type": "Receipt", "confidence": conf, "rule": "NEFT Return"}

        # ------------------------------------------------------------------
        # STEP 10 — Extract entity from narration prefix
        # ------------------------------------------------------------------
        entity = ""
        remark = ""
        is_upi = "UPI/" in orig_upper

        if "CHQ PAID" in orig_upper or "MICR INWARD" in orig_upper:
            entity = _parse_chq(orig_upper)
        elif is_upi:
            entity = _parse_upi(orig_upper)
        elif "IB NEFT" in orig_upper or "SC NEFT" in orig_upper or "IB IFT" in orig_upper:
            entity, remark = _parse_ib_neft(orig_upper)
        elif orig_upper.startswith("IFT-"):
            # Kotak IFT format: IFT-<Entity>-<Code>-<Ref>
            entity = _parse_ift(orig_upper)
        elif "BDP-" in orig_upper:
            entity = _parse_bdp(orig_upper)[0]  # service name as entity for D2C fuzzy
        elif re.match(r'^N/\d+/', orig_upper):
            # IndusInd Bank NEFT debit: N/<seq>/<Entity Name>/<INDBH ref>/
            entity = _parse_indusind_neft(orig_upper)
        elif orig_upper.startswith("R/") and orig_upper.count('/') >= 3:
            # IndusInd Bank RTGS credit: R/<UTR>/<IFSC>/<Sender Name>//
            entity = _parse_indusind_rtgs_credit(orig_upper)
        elif orig_upper.startswith("BILL/") and orig_upper.count('/') >= 3:
            # IndusInd Bank bill/card payment: BILL/<inv>/<type>/<Entity>/
            # The 3rd segment is the payment type (CREDIT, DEBIT, etc.).
            # Append "CREDIT CARD" when type=CREDIT so the fuzzy matcher prefers
            # "HDFC Credit Card" over "HDFC Bank" or "Hdfc FD XX".
            _bparts = orig_upper.split('/')
            _bill_type   = _bparts[2].strip() if len(_bparts) >= 3 else ""
            _bill_entity = _bparts[3].strip() if len(_bparts) >= 4 else ""
            if "CREDIT" in _bill_type and _bill_entity:
                entity = _bill_entity + " CREDIT CARD"
            else:
                entity = _bill_entity
        elif re.match(r'^(?:NEFT|RTGS|IMPS)\s+[A-Z0-9]{8,}\s+', orig_upper):
            # Generic bank NEFT/RTGS format: "NEFT <UTR> <Beneficiary Name...>"
            # Handles formats from HSBC (HSBCN), ICICI (ICICINXXX), etc.
            # Strip anything from "(" onwards (value date annotations) and limit to 50 chars.
            m = re.match(r'^(?:NEFT|RTGS|IMPS)\s+[A-Z0-9]{8,}\s+(.+)', orig_upper)
            if m:
                raw = m.group(1).split('(')[0].strip()
                entity = raw[:50].strip()
        elif '-' in orig and not entity:
            # Fallback: "ENTITY NAME-DESCRIPTION" pattern used by many direct payments.
            # E.g. "EKANEK NETWORKS PVT LTD-APPX CHARGES" → entity = "EKANEK NETWORKS PVT LTD"
            # Only apply when the part before the first dash looks like a company/person name.
            _bank_prefixes = {'NEFT', 'RTGS', 'IMPS', 'UPI', 'IB', 'SC', 'CHQ', 'IFT',
                              'BDP', 'ETAX', 'GBO', 'INT', 'CHRG', 'JYOTI', 'VALUE'}
            dash_idx = orig_upper.index('-')
            possible = orig_upper[:dash_idx].strip()
            first_word = possible.split()[0] if possible.split() else ''
            if len(possible) > 5 and first_word not in _bank_prefixes and not possible[0].isdigit():
                entity = possible

        entity = entity.strip()

        # ------------------------------------------------------------------
        # STEP 10.5 — IndusInd bulk salary batch
        # Format: SALR<date><seq>/<batch>/NEFT/SALARY FOR <MONTH>/
        # This is a bulk payroll run — no individual employee name is present.
        # Map to the generic "Net Salary Payable" or "Salary Payable" ledger.
        # ------------------------------------------------------------------
        if re.match(r'^SALR\d+/', orig_upper) and "SALARY" in orig_upper:
            # Prefer the plain "Salary Payable" ledger (no employee suffix) —
            # this is a bulk payroll run, not a per-employee transfer.
            # "Net Salary Payable" is a secondary fallback.
            net_sal = (
                next((l for l in self.salary_ledgers
                      if l.lower().strip() == 'salary payable'), None)
                or next((l for l in self.master_ledgers if 'net salary' in l.lower()), None)
                or next((l for l in self.salary_ledgers), None)
                or self._find_master(['salary payable', 'net salary'])
                or self._suspense_ledger
            )
            return {"ledger": net_sal, "type": txn_type, "confidence": "High",
                    "rule": "IndusInd Salary Batch"}

        # ------------------------------------------------------------------
        # STEP 11 — Salary / allowance routing
        # Note: REIMBURSEMENT payments (REIMB) go to the person's reimbursement ledger,
        # not Salary Payable. We skip salary matching for those and let the anchored
        # entity match (STEP 12b) find the correct "Person Reimbursement Payable" ledger.
        # ------------------------------------------------------------------
        salary_keywords = ["SALARY", "ALLOWANCE", "WAGES", "STIPEND",
                           "FACTORY MILK", "LABOUR", "TA DA", "TA/DA"]
        is_salary_narration = any(k in orig_upper for k in salary_keywords) or \
                              any(k in remark.upper() for k in salary_keywords)

        # Detect reimbursement narrations (short-form "REIMB" or full "REIMBURSEMENT")
        is_reimb_narration = "REIMB" in orig_upper or "REIMBURSEMENT" in orig_upper

        # For bare reimbursement narrations like "VIKASH KUMAR- REIMB MAR25" (not IB NEFT),
        # entity won't have been extracted by STEP 10 parsers — extract it here.
        if is_reimb_narration and not entity:
            reimb_match = re.match(r'^([A-Z][A-Z\s]+?)\s*-\s*REIMB', orig_upper)
            if reimb_match:
                entity = reimb_match.group(1).strip()

        if is_salary_narration and not is_reimb_narration and entity:
            matched = self._match_employee_ledger(entity, self.salary_ledgers) \
                      or self._match_employee_ledger(entity, self.salary_ledgers_broad)
            if matched:
                return {"ledger": matched, "type": txn_type, "confidence": "High", "rule": "Salary Match"}

        # Compute once — used by STEP 11.5 and STEP 12b
        is_clean_neft = (
            any(k in orig_upper for k in ["IB NEFT", "SC NEFT", "IB IFT", "CHQ PAID", "IB RTGS"])
            or orig_upper.startswith("IFT-")
            or bool(re.match(r'^N/\d+/', orig_upper))   # IndusInd NEFT debit
            or orig_upper.startswith("R/")              # IndusInd RTGS credit
            or orig_upper.startswith("BILL/")           # IndusInd bill payment
            or bool(re.match(r'^(?:NEFT|RTGS|IMPS)\s+[A-Z0-9]{8,}\s+', orig_upper))  # Generic NEFT/RTGS (HSBC, ICICI, etc.)
        )

        # ------------------------------------------------------------------
        # STEP 11.5 — Employee match for clean NEFT/IFT without SALARY keyword
        # Many payroll transfers don't include the word "SALARY" in the narration —
        # the entity IS the employee name. Run name-aware matching here so these
        # don't fall through to the generic anchored/fuzzy steps which have no
        # suffix stemming or first-name anchoring logic.
        # Returns Medium (not High) — inferred without keyword confirmation.
        # ------------------------------------------------------------------
        if is_clean_neft and entity and not is_salary_narration and not is_reimb_narration:
            _emp = (self._match_employee_ledger(entity, self.salary_ledgers)
                    or self._match_employee_ledger(entity, self.salary_ledgers_broad))
            if _emp:
                return {"ledger": _emp, "type": txn_type, "confidence": "Medium",
                        "rule": "NEFT Employee Match (inferred)"}

        # ------------------------------------------------------------------
        # STEP 12 — Rent remark routing
        # Only checks the REMARK field (the decoded comment from IB NEFT narrations).
        # Narrations where "RENT" appears in the main body — like "ANU TUSHIR-ADV RENT APR25"
        # or "EKANEK NETWORKS PVT LTD- RENT MAR25" — are already handled upstream by
        # entity extraction + anchored fuzzy (the entity IS the payee, RENT is just context).
        # ------------------------------------------------------------------
        rent_keywords = ["RENT", "LEASE", "GODOWN CHARGES"]
        if any(k in remark.upper() for k in rent_keywords):
            if self.rent_ledgers:
                ledger, conf, _ = self._fuzzy_match(entity or remark, self.rent_ledgers)
                if conf != "Low":
                    return {"ledger": ledger, "type": txn_type, "confidence": conf, "rule": "Rent Remark"}

        # ------------------------------------------------------------------
        # STEP 11.9 — Marketplace settlement routing (MUST run before STEP 12b)
        # Marketplace payees (Zepto, Amazon, Blinkit, Flipkart, Swiggy, …) must book to the
        # brand's "<Marketplace> Settlement/Clearing A/c" — NOT a state sub-ledger
        # ("Zepto Private Limited Jaipur", "Amazon Delhi") or "…Debtors". Accountants can't
        # attribute a marketplace payout to a state. The first-word anchored match (STEP 12b)
        # would otherwise grab a state ledger, so we resolve the settlement ledger here first.
        # Only fires when such a ledger exists in THIS brand's COA — otherwise we fall through
        # untouched (non-marketplace vendors and settlement-less brands are unaffected).
        # BDP narrations scan only the SERVICE segment (not the gateway), mirroring STEP 15.
        # ------------------------------------------------------------------
        if "BDP-" in orig_upper:
            _mp_scan = _parse_bdp(orig_upper)[0].lower()
        else:
            _mp_scan = orig_upper.lower()
        for _brand, _terms in D2C_KEYWORDS.items():
            if not any(t in _mp_scan for t in _terms):
                continue
            _subset = self._d2c_ledger_map.get(_brand, [])
            if not _subset:
                continue
            if is_credit:
                # Receipt from a marketplace → collection / seller receipt / settlement.
                _pref = next((l for l in _subset if any(k in l.lower() for k in
                              ['collection', 'seller receipt', 'settlement', 'clearing'])), None)
            else:
                # Payment to a marketplace → settlement / clearing.
                _pref = next((l for l in _subset if 'settlement' in l.lower() or 'clearing' in l.lower()), None)
            if _pref:
                return {"ledger": _pref, "type": txn_type, "confidence": "High",
                        "rule": f"Marketplace Settlement ({_brand})"}
            # No settlement/clearing ledger for this brand in the COA → don't force a match here;
            # let STEP 12b / STEP 15 handle it (e.g. Creditors/Debtors per existing logic).

        # ------------------------------------------------------------------
        # STEP 12b — First-word-anchored entity match (before D2C scan)
        # When the entity is cleanly extracted from a structured NEFT/CHQ/IFT narration,
        # restrict candidates to ledgers containing the entity's FIRST WORD before fuzzy
        # scoring. This prevents false matches like "Sudhir Balgovind" → "Vijay Balgovind"
        # or "Alpino HealthFoods" → "WLDD Private Limited" due to common token overlap.
        # If no anchor candidates exist, fall through to D2C and general fuzzy.
        # ------------------------------------------------------------------
        # STEP 12b — First-word-anchored entity match (before D2C scan)
        # is_clean_neft was computed before STEP 11.5 (above).
        # Generic NEFT/RTGS now included — anchor match runs first-word filter.
        # If no anchored ledgers are found, falls through to D2C scan and general fuzzy.
        # Also run anchored match for reimbursement narrations extracted above
        if entity and len(entity) > 4 and (is_clean_neft or is_reimb_narration):
            first_word = entity.split()[0].lower() if entity.split() else ""
            # Word-boundary match + minimum length guard.
            # Substring "in l.lower()" causes false anchoring: first_word="int" would
            # pull in "interest expense", "interior", etc. Also skip anchoring for very
            # short/generic tokens (< 4 chars) — they match too many unrelated ledgers.
            if first_word and len(first_word) >= 4:
                _fw_pat = re.compile(r'\b' + re.escape(first_word) + r'\b', re.IGNORECASE)
                anchored = [l for l in self.master_ledgers if _fw_pat.search(l)]
            else:
                anchored = []
            # Own-brand guard: remove the brand's own company ledger from anchored
            # candidates so it can't win when the narration embeds the brand's name
            # as a beneficiary/escrow label rather than as the actual counterparty.
            if self._own_brand_ledger:
                anchored = [c for c in anchored if c != self._own_brand_ledger]
            if anchored:
                # Use a slightly lower threshold (68) for anchored matches — the first-word
                # constraint already filters out unrelated ledgers, so a score of 68+ within
                # the anchored set is more reliable than a 87 score in the full master.
                remark_upper = remark.upper()
                # Catch both "REIMB" (short form, e.g. "VIKASH KUMAR-REIMB MAR25")
                # and "REIMBURS" / "REIMBURSEMENT" (full form in remark)
                is_reimbursement = ("REIMB" in orig_upper or "REIMBURS" in remark_upper)

                # For expense reimbursements: prefer person/vendor ledger over salary ledgers.
                # NOTE: We exclude salary/wages but NOT "payable" — we want to find
                # "Vikash Kumar Reimbursement Payable", not exclude it.
                if is_reimbursement:
                    non_salary = [c for c in anchored
                                  if 'salary' not in c.lower() and 'wages' not in c.lower()]
                    anchored_effective = non_salary if non_salary else anchored
                else:
                    anchored_effective = anchored

                def combined(c):
                    # Expand abbreviations before scoring — mirrors _fuzzy_match exactly.
                    # Without this, "ALPINO HEALTHFOODS PVT LIM" scores ~73 against
                    # "Alpino Health Foods Private Limited" instead of ~90.
                    e_norm = _expand_abbrevs(entity)
                    c_norm = _expand_abbrevs(c)
                    s1 = fuzz.token_set_ratio(e_norm, c_norm)
                    s2 = fuzz.token_sort_ratio(e_norm, c_norm)
                    s = round(0.85 * s1 + 0.15 * s2) if len(e_norm) <= 8 else round(0.6 * s1 + 0.4 * s2)
                    # Boost ledgers whose name echoes a remark keyword (tolerates typos)
                    if is_reimbursement and fuzz.partial_ratio("reimburse", c.lower()) >= 70:
                        s += 20
                    # For credit (receipt) IFT transactions, strongly prefer Collection
                    # sub-ledgers over the plain company/3PL ledger.
                    # E.g.: IFT-BVC TRADEPORT... (receipt) → "BVC Tradeport Collection (Shopify)"
                    # not "BVC TRADEPORT PRIVATE LIMITED (Vamaship)".
                    ift_credit = is_credit and orig_upper.startswith("IFT-")
                    if ift_credit and any(k in c.lower() for k in ['collection', 'seller receipt']):
                        s += 30
                    elif is_credit and any(k in c.lower() for k in ['collection', 'seller receipt']):
                        s += 10
                    return s
                best_a = max(anchored_effective, key=combined)
                score_a = combined(best_a)
                # Raise bar when multiple similar ledgers compete — a score of 69
                # is not reliable enough to commit when two vendors share a first word.
                _anchor_threshold = 76 if len(anchored_effective) > 1 else 68
                if score_a >= _anchor_threshold:
                    conf_a = "High" if score_a >= 87 else "Medium"
                    return {"ledger": best_a, "type": txn_type, "confidence": conf_a,
                            "rule": "Anchored Entity Match"}

        # ------------------------------------------------------------------
        # STEP 13 — Customer Refund early exit (before D2C scan)
        # Narrations like "JYOTI K- SWIGGY CUSTOMER REFUND" are outgoing refunds to customers,
        # not D2C marketplace transactions. Detect this before D2C scan.
        # ------------------------------------------------------------------
        if "CUSTOMER REFUND" in orig_upper or "CUST REFUND" in orig_upper:
            refund_ledger = next(
                (l for l in self.master_ledgers if 'customer refund' in l.lower()), None
            ) or next(
                (l for l in self.master_ledgers if 'refund' in l.lower() and 'customer' in l.lower()), None
            )
            if refund_ledger:
                return {"ledger": refund_ledger, "type": txn_type, "confidence": "High",
                        "rule": "Customer Refund"}

        # ------------------------------------------------------------------
        # STEP 15 — D2C marketplace keyword scan
        #
        # For BDP narrations, ONLY scan the SERVICE segment (first BDP- token).
        # Scanning the full narration would match the gateway name (RAZORPAY, CCAVENUEF)
        # instead of the actual payee (e.g., BDP-SHIPROCK-RAZORPAY → should match
        # Shiprocket, not Razorpay).
        # ------------------------------------------------------------------
        if "BDP-" in orig_upper:
            bdp_service_name, _ = _parse_bdp(orig_upper)
            d2c_scan_text = bdp_service_name.lower()
        else:
            d2c_scan_text = orig_upper.lower()

        is_bdp_payment = "BDP-" in orig_upper and not is_credit
        for brand, terms in D2C_KEYWORDS.items():
            if any(t in d2c_scan_text for t in terms):
                subset = self._d2c_ledger_map.get(brand, [])
                if subset:
                    # Direction-aware D2C ledger preference:
                    # Outgoing (debit/payment) → prefer Settlement Clearing A/c or Creditors
                    # Incoming (credit/receipt) → prefer Collection, Seller Receipt, or Settlement
                    # This prevents receipts going to "Debtors" and payments going to "Debtors".
                    if is_bdp_payment or not is_credit:
                        preferred = (
                            next((l for l in subset if 'settlement' in l.lower() or 'clearing' in l.lower()), None)
                            or next((l for l in subset if 'creditor' in l.lower()), None)
                        )
                    else:
                        # Credit/receipt: prefer collection or seller receipt ledgers
                        preferred = (
                            next((l for l in subset if any(k in l.lower() for k in
                                  ['collection', 'seller receipt', 'settlement', 'clearing'])), None)
                        )
                    if preferred:
                        return {"ledger": preferred, "type": txn_type, "confidence": "High",
                                "rule": f"D2C Settlement Match ({brand})"}
                    ledger, conf, _ = self._fuzzy_match(brand, subset)
                    if conf != "Low":
                        return {"ledger": ledger, "type": txn_type, "confidence": conf,
                                "rule": f"D2C Match ({brand})"}
                # No ledger in master for this brand — try full master fuzzy with brand name
                max_term = max(terms, key=len)
                ledger, conf, _ = self._fuzzy_match(max_term, self.master_ledgers)
                if conf != "Low":
                    return {"ledger": ledger, "type": txn_type, "confidence": conf,
                            "rule": f"D2C Broad Match ({brand})"}

        # ------------------------------------------------------------------
        # STEP 14 — General fuzzy: try entity first, then full cleaned narration
        # ------------------------------------------------------------------
        if entity and len(entity) > 2:
            ledger, conf, _ = self._fuzzy_match(entity, self.master_ledgers, guard_generic=True,
                                                 exclude=self._own_brand_ledger)
            if conf != "Low":
                return {"ledger": ledger, "type": txn_type, "confidence": conf,
                        "rule": "Fuzzy Entity Match"}

        generic = clean_narration(orig)
        if generic and len(generic) > 3:
            ledger, conf, _ = self._fuzzy_match(generic, self.master_ledgers, guard_generic=True,
                                                 exclude=self._own_brand_ledger)
            if conf != "Low":
                return {"ledger": ledger, "type": txn_type, "confidence": conf,
                        "rule": "Fuzzy Generic Match"}

        # Per-brand LOW-PRIORITY side-map fallback (e.g. Urban Plant Salary/Stipend →
        # Salary Payable): only reached when nothing specific matched, so it never
        # overrides a real per-payee/directory/fuzzy match. Empty for other brands.
        smf = self._match_side(self.side_map_fallback, orig_upper, txn_type, "Side Ledger (fallback)")
        if smf:
            return smf

        # ------------------------------------------------------------------
        # FALLBACK — Suspense A/c (pre-computed from master during init)
        # ------------------------------------------------------------------
        return {"ledger": self._suspense_ledger, "type": txn_type, "confidence": "Low",
                "rule": "Suspense Fallback", "entity": entity}


# ---------------------------------------------------------------------------
# File loaders
# ---------------------------------------------------------------------------

def load_ledger_master(filepath: str) -> list:
    """
    Load ledger names from a Tally / CoA export.
    Auto-detects the right sheet (by name or by scanning for the most ledger-like content)
    and auto-detects the right column (the one with the most text-heavy, non-numeric entries).
    Filters out header rows, totals, and group summary lines.
    """
    xl = pd.ExcelFile(filepath)

    # Prefer a sheet whose name signals a chart of accounts / ledger list.
    # Falls back to scanning all sheets for the one with the most text rows.
    _coa_keywords = [
        "ledger", "list", "master", "account", "chart", "coa", "gl",
        "tally", "voucher", "accounts master", "list of ledger",
    ]
    target = next(
        (s for s in xl.sheet_names
         if any(k in s.lower() for k in _coa_keywords)),
        None,
    )
    if target is None:
        # No keyword match — pick the sheet with the most non-numeric rows in column A
        best_sheet, best_count = xl.sheet_names[0], 0
        for s in xl.sheet_names:
            try:
                _df = pd.read_excel(filepath, sheet_name=s, header=None, nrows=200)
                count = _df[0].dropna().apply(
                    lambda v: bool(re.search(r'[A-Za-z]{3,}', str(v)))
                ).sum()
                if count > best_count:
                    best_count, best_sheet = count, s
            except Exception:
                pass
        target = best_sheet

    df = pd.read_excel(filepath, sheet_name=target, header=None)

    # Auto-detect the column containing ledger names.
    #
    # Scored by DISTINCT text values, not raw count. A Tally "List of Ledgers" export has
    # both a "Name of Ledger" column and an "Under" (group) column; the group column often
    # has one or two MORE non-empty cells, so a raw-count heuristic silently picks it — and
    # because group names repeat, the COA collapses to a handful of entries. Seen for real:
    # Urban Plant's 862-row export loaded as 96 "ledgers", which would have failed COA
    # validation for every side rule and every learned key. Ledger names are unique;
    # group names repeat, so cardinality separates them cleanly (856 vs 96).
    #
    # An explicit header ("Name of Ledger" / "Ledger Name" / "Particulars") wins outright.
    _col_idx = 0
    best_score = -1
    _HEADER_RE = re.compile(r'name of ledger|ledger name|particulars|account name', re.I)
    header_col = None
    for col in df.columns:
        for v in df[col].dropna().head(15):
            if _HEADER_RE.fullmatch(str(v).strip()):
                header_col = col
                break
        if header_col is not None:
            break

    if header_col is not None:
        _col_idx = header_col
    else:
        for col in df.columns:
            vals = df[col].dropna().apply(str)
            text = vals[vals.apply(
                lambda v: bool(re.search(r'[A-Za-z]{3,}', v)) and not v.strip().isdigit()
            )]
            score = text.str.strip().str.lower().nunique()
            if score > best_score:
                best_score, _col_idx = score, col

    raw = df[_col_idx].dropna().astype(str).tolist()

    # Standard Tally group names that appear in hierarchical Chart of Accounts exports —
    # these are not ledgers and should be excluded from the master list.
    _TALLY_GROUPS = {
        "capital account", "reserves & surplus", "current liabilities",
        "loans (liability)", "branch / divisions", "suspense a/c",
        "fixed assets", "investments", "current assets", "misc. expenses (asset)",
        "profit & loss a/c", "purchase accounts", "sales accounts",
        "direct incomes", "indirect incomes", "direct expenses", "indirect expenses",
        "sundry debtors", "sundry creditors", "bank accounts", "bank od a/c",
        "cash-in-hand", "loans & advances (asset)", "deposits (asset)",
        "duties & taxes", "provisions",
        "stock-in-hand", "secured loans", "unsecured loans",
        # One-word accounting terms that are never standalone ledger names
        "assets", "liabilities", "income", "expenses", "equity",
    }

    # Header/label cells that sit in the same column as the ledger names and would
    # otherwise be imported as a ledger (seen: "Name of Ledger" arriving as a new ledger).
    skip_exact = {"total", "grand total", "list of ledgers", "ledger name", "name",
                  "name of ledger", "particulars", "account name", "under",
                  "opening balance", "closing balance", "sl. no.", "sl no", "s. no.",
                  "sr. no.", "sr no", "amount"}

    # Patterns that identify company header / metadata rows — not ledger names.
    # E.g. "Plot - C/59 Platina Bandra Kurla Complex G-Block Near City Union Bank"
    #      "1-Apr-24 to 31-Mar-26"   "45 Group(s) and 261 Ledger(s)"
    _METADATA_PATTERNS = [
        re.compile(r'\bPlot\b', re.IGNORECASE),           # address line
        re.compile(r'\d+\s*-\s*[A-Za-z]{3}\s*-\s*\d{2}'), # date range "1-Apr-24"
        re.compile(r'\d+\s+Group', re.IGNORECASE),         # "45 Group(s)"
        re.compile(r'\d+\s+Ledger', re.IGNORECASE),        # "261 Ledger(s)"
        re.compile(r'\bto\s+\d{2}-\w{3}-\d{2,4}\b', re.IGNORECASE),  # "to 31-Mar-26"
    ]

    cleaned = []
    for l in raw:
        l = l.strip()
        if not l:
            continue
        if l.lower() in skip_exact:
            continue
        if l.lower() in _TALLY_GROUPS:
            continue
        if l.isdigit():
            continue
        # Skip rows longer than 100 chars — Tally ledger names are never this long;
        # these are almost always company address lines from the CoA header.
        if len(l) > 100:
            continue
        # Skip rows matching known metadata patterns
        if any(p.search(l) for p in _METADATA_PATTERNS):
            continue
        cleaned.append(l)

    # ------------------------------------------------------------------
    # COA INTEGRITY GUARD (P1) — fail loudly instead of matching against garbage.
    #
    # The single biggest source of mass mis-classification ("50% wrong") was the app
    # feeding a *bank statement* in as the COA (e.g. a sheet mis-named "List of Ledger"
    # that actually held transactions). A bank statement's text column is full of
    # transaction narrations (IB NEFT / SC NEFT / UPI/ / IMPS / RTGS / IFT- / BDP- ...),
    # which a real Chart of Accounts never contains. Detect that signature and refuse,
    # so the run errors with a clear message rather than silently producing nonsense.
    # ------------------------------------------------------------------
    _TXN_SIG = re.compile(
        r'\b(IB\s+NEFT|SC\s+NEFT|IB\s+ITG|IB\s+IFT|IB\s+OAT|IB\s+RTGS|NEFT\s+Dr|NEFT\s+Cr'
        r'|RTGS|IMPS|UPI/|IFT-|BDP-|CHQ\s+PAID|MICR\s+INWARD|OTH-Transfer'
        r'|OTHER\s+THAN\s+SB)\b',
        re.IGNORECASE,
    )
    # Date-dominated column → the auto-detector picked a Date / Value-Date column from a
    # transaction sheet (e.g. "01 May 2026", "1-Apr-24", "01-05-2026", "2026-06-03").
    # A real Chart of Accounts column holds names, never dates.
    _DATE_SIG = re.compile(
        r'^\s*(\d{1,2}[\s\-/](Jan|Feb|Mar|Apr|May|Jun|Jul|Aug|Sep|Oct|Nov|Dec)[a-z]*[\s\-/]\d{2,4}'
        r'|\d{1,2}[-/]\d{1,2}[-/]\d{2,4}'
        r'|\d{4}-\d{2}-\d{2})\s*$',
        re.IGNORECASE,
    )
    if cleaned:
        txn_like = sum(1 for l in cleaned if _TXN_SIG.search(l))
        date_like = sum(1 for l in cleaned if _DATE_SIG.search(l))
        bad_frac = (txn_like + date_like) / len(cleaned)
        if bad_frac > 0.20:
            kind = "BANK STATEMENT (transaction narrations)" if txn_like >= date_like \
                else "transaction sheet (date column)"
            raise ValueError(
                f"COA integrity check failed: the selected sheet/column in "
                f"'{os.path.basename(filepath)}' (sheet '{target}', column {_col_idx}) looks like a "
                f"{kind}, not a Chart of Accounts — {txn_like + date_like}/{len(cleaned)} "
                f"({bad_frac:.0%}) of entries are transactions/dates. "
                f"Upload the brand's Ledger Master / List of Ledgers instead."
            )

    if not cleaned:
        raise ValueError(
            f"COA integrity check failed: no ledger names could be read from "
            f"'{os.path.basename(filepath)}' (sheet '{target}', column {_col_idx})."
        )

    # Surface what was loaded so a wrong/tiny COA is visible in logs (never silent).
    if len(cleaned) < 50:
        print(f"⚠️  COA WARNING: only {len(cleaned)} ledgers loaded from "
              f"'{os.path.basename(filepath)}' (sheet '{target}') — verify this is the full "
              f"Chart of Accounts, not a partial/derived list.", file=sys.stderr)

    return cleaned


def load_bank_statement(filepath: str) -> tuple:
    """
    Auto-detect the bank statement sheet and header row.
    Returns (DataFrame, col_map dict, sheet_name).
    col_map keys: txn_date, description, debit, credit, balance.
    Supports .xlsx, .xls, and .csv formats.
    """
    ext = str(filepath).rsplit('.', 1)[-1].lower()

    if ext == 'csv':
        # CSV path: bank-statement CSVs are RAGGED — metadata rows (account/address)
        # have a different column count than the transaction table, which crashes
        # pandas' C parser ("Expected N fields ... saw M"). Read with the csv module
        # (handles any per-row width + quoting), pad every row to the widest, then build
        # the frame — reproducing pd.read_csv(header=None, dtype=str) but ragged-safe.
        import csv as _csv
        rows = None
        for enc in ('utf-8-sig', 'utf-8', 'latin-1', 'cp1252'):
            try:
                with open(filepath, newline='', encoding=enc) as _fh:
                    rows = list(_csv.reader(_fh))
                break
            except UnicodeDecodeError:
                continue
        if rows is None:
            raise ValueError(f"Cannot read CSV {filepath} — tried utf-8, latin-1, cp1252")
        rows = [r for r in rows if any((c or '').strip() for c in r)]   # drop fully-blank rows
        if not rows:
            raise ValueError(f"CSV {filepath} is empty")
        width = max(len(r) for r in rows)
        rows = [r + [''] * (width - len(r)) for r in rows]              # pad ragged → rectangle
        df_raw = pd.DataFrame(rows).replace('', float('nan'))           # match read_csv(dtype=str)
        target = 'CSV'
    else:
        # Excel path (unchanged)
        xl = pd.ExcelFile(filepath)
        # Prefer sheet with 'od acc', 'raw', 'statement', 'bank', 'transactions' in name
        target = next(
            (s for s in xl.sheet_names
             if any(k in s.lower() for k in ["od acc", "raw", "statement", "bank", "transactions", "account"])),
            xl.sheet_names[0]
        )
        df_raw = pd.read_excel(filepath, sheet_name=target, header=None)

    header_idx = None
    for i, row in df_raw.iterrows():
        if i > 30:
            break
        vals = [str(v).lower().strip() for v in row.values if pd.notna(v)]
        has_date  = any("date" in v or "txn" in v for v in vals)
        has_desc  = any("desc" in v or "narration" in v or "particulars" in v or "details" in v for v in vals)
        has_amt   = any(any(k in v for k in ["debit", "credit", "withdrawal", "withdrawl", "deposit", "amount"]) for v in vals)
        if (has_date and has_desc) or (has_desc and has_amt):
            header_idx = i
            break

    if header_idx is None:
        header_idx = 0

    df_raw.columns = df_raw.iloc[header_idx]
    # Deduplicate column names (e.g. Kotak CSV has two "Dr / Cr" columns)
    _seen = {}
    _deduped = []
    for _c in df_raw.columns:
        _k = str(_c).strip()
        if _k in _seen:
            _seen[_k] += 1
            _deduped.append(f"{_k}.{_seen[_k]}")
        else:
            _seen[_k] = 0
            _deduped.append(_k)
    df_raw.columns = _deduped
    df = df_raw.iloc[header_idx + 1:].reset_index(drop=True)
    cols = [str(c).strip() for c in df.columns]

    col_map = {}
    for c in cols:
        cl = c.lower()
        if "txn_date" not in col_map and any(k in cl for k in ["txn date", "transaction date", "value date", "date"]):
            col_map["txn_date"] = c
        elif "description" not in col_map and any(k in cl for k in ["description", "narration", "particulars", "details"]):
            col_map["description"] = c
        elif "debit" not in col_map and any(k in cl for k in ["withdrawal", "withdrawl", "debit", "dr amount"]):
            col_map["debit"] = c
        elif "credit" not in col_map and any(k in cl for k in ["deposit", "credit", "cr amount"]):
            col_map["credit"] = c
        elif "balance" not in col_map and "balance" in cl:
            col_map["balance"] = c
        elif "chq_ref" not in col_map and any(k in cl for k in
                ["chq", "cheque", "ref no", "reference no", "chq/ref", "chq / ref", "instrument no", "utr/ref"]):
            col_map["chq_ref"] = c

    # Handle combined Amount + direction-indicator column (e.g. Kotak: "Amount" + "Dr / Cr")
    if "debit" not in col_map and "credit" not in col_map:
        amt_col = next(
            (c for c in df.columns if str(c).strip().lower() in ('amount', 'amt')),
            None
        ) or next(
            # Fallback: a column whose header *contains* amount / amt (e.g. ICICI
            # "Transaction Amount(INR)"), excluding the running balance column.
            (c for c in df.columns
             if ('amount' in str(c).strip().lower() or re.search(r'\bamt\b', str(c).strip().lower()))
             and 'balance' not in str(c).strip().lower()),
            None
        )
        dir_col = next(
            (c for c in df.columns
             if str(c).strip().lower() in ('dr / cr', 'dr/cr', 'type', 'txn type', 'dr.cr', 'cr/dr')),
            None
        )
        if amt_col and dir_col:
            def _to_num(v):
                try:
                    return float(str(v).replace(',', '').strip()) if pd.notna(v) else 0.0
                except (ValueError, TypeError):
                    return 0.0
            df['__debit'] = df.apply(
                lambda r: _to_num(r[amt_col]) if str(r[dir_col]).strip().upper().startswith('DR') else 0.0,
                axis=1
            )
            df['__credit'] = df.apply(
                lambda r: _to_num(r[amt_col]) if str(r[dir_col]).strip().upper().startswith('CR') else 0.0,
                axis=1
            )
            col_map['debit'] = '__debit'
            col_map['credit'] = '__credit'

    return df, col_map, target


# ---------------------------------------------------------------------------
# Output writer
# ---------------------------------------------------------------------------

def write_output(rows: list, summary: dict, brand: str, output_path: str):
    """
    Write classified rows to a color-coded Excel workbook.
    Sheet 1: Bank Statement (7 columns + Confidence)
    Sheet 2: Summary metrics
    """
    wb = openpyxl.Workbook()
    ws = wb.active
    ws.title = "Bank Statement"

    # Fills
    GREEN  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
    YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
    RED    = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
    HEADER = PatternFill(start_color="263238", end_color="263238", fill_type="solid")

    WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    BOLD       = Font(name="Calibri", size=11, bold=True)
    REGULAR    = Font(name="Calibri", size=11)
    THIN       = Border(
        left=Side(style='thin', color='CCCCCC'),   right=Side(style='thin', color='CCCCCC'),
        top=Side(style='thin', color='CCCCCC'),    bottom=Side(style='thin', color='CCCCCC'),
    )

    # "Source" is APPENDED as column 10 on purpose: recoController reads this workbook by
    # fixed column indices 1-9, so a trailing column is invisible to it while giving the
    # accountant the provenance of every row. Without it "High" is unfalsifiable from their
    # seat — they cannot tell a learned mapping from a fuzzy guess without re-checking all
    # 261 rows by hand, and a misbehaving layer stays invisible for months.
    headers = ["Txn Date", "Description", "Chq / Ref No.", "Debit", "Credit", "Balance", "Type", "Ledger Name", "Confidence", "Source"]
    ws.append(headers)
    ws.row_dimensions[1].height = 25
    for ci, _ in enumerate(headers, 1):
        cell = ws.cell(row=1, column=ci)
        cell.fill, cell.font, cell.border = HEADER, WHITE_BOLD, THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")

    def _source_label(rule):
        """Collapse the internal rule string into a provenance label the accountant can act
        on: which LAYER decided this row."""
        rule = str(rule or "")
        claude = " + Claude" if "Claude" in rule else ""
        for prefix, label in (
            ("Side Ledger (flipped)", "Side Rule (flipped)"),
            ("Side Ledger (unconfirmed)", "Side Rule (unconfirmed)"),
            ("Side Ledger", "Side Rule"),
            ("Stored Correction", "Stored"),
            ("Payee Directory", "Directory"),
            ("Own Account", "Own Account"), ("OAT", "Own Account"), ("Sweep", "Own Account"),
            ("Contra", "Contra"),
            ("BDP Statutory", "Statutory"), ("GSTN", "Statutory"), ("TDS", "Statutory"),
            ("EPF", "Statutory"), ("ESIC", "Statutory"), ("PT ", "Statutory"),
            ("Bank Charges", "Bank Charges"), ("Interest", "Interest"),
        ):
            if rule.startswith(prefix):
                return label + claude
        if not rule:
            return "Claude" if claude else "Suspense"
        return (rule.split("(")[0].strip() or "Rule") + claude

    for r in rows:
        ws.append([
            r.get("txn_date", ""),
            r.get("description", ""),
            r.get("chq_ref", ""),
            r.get("debit") or "",
            r.get("credit") or "",
            r.get("balance") or "",
            r.get("predicted_type", ""),
            r.get("predicted_ledger", ""),
            r.get("confidence", ""),
            _source_label(r.get("rule", "")),
        ])
        row_num = ws.max_row
        ws.row_dimensions[row_num].height = 20
        conf   = r.get("confidence", "Low")
        fill   = GREEN if conf == "High" else (YELLOW if conf == "Medium" else RED)

        for ci in range(1, len(headers) + 1):
            cell = ws.cell(row=row_num, column=ci)
            cell.font, cell.border = REGULAR, THIN
            if ci in (4, 5, 6):
                cell.number_format = '#,##0.00'
                cell.alignment = Alignment(horizontal="right", vertical="center")
            elif ci in (1, 7, 9):
                cell.alignment = Alignment(horizontal="center", vertical="center")
            else:
                cell.alignment = Alignment(horizontal="left", vertical="center")
            if ci in (8, 9):
                cell.fill = fill

    # Auto-width
    for col in ws.columns:
        ltr = get_column_letter(col[0].column)
        ws.column_dimensions[ltr].width = min(max(max(len(str(c.value or '')) for c in col) + 3, 10), 55)

    # --- Summary sheet ---
    ws2 = wb.create_sheet("Summary")
    ws2.append([f"Bank Statement Classification — {brand}"])
    ws2.cell(1, 1).font = Font(name="Calibri", size=16, bold=True)
    ws2.row_dimensions[1].height = 30
    ws2.append([])

    hdrs2 = ["Metric", "Count", "Percentage"]
    ws2.append(hdrs2)
    ws2.row_dimensions[3].height = 22
    for ci, _ in enumerate(hdrs2, 1):
        c = ws2.cell(row=3, column=ci)
        c.fill, c.font, c.border = HEADER, WHITE_BOLD, THIN
        c.alignment = Alignment(horizontal="center", vertical="center")

    total = sum(summary.values())
    for lvl, fill in [("High", GREEN), ("Medium", YELLOW), ("Low", RED)]:
        cnt = summary.get(lvl, 0)
        pct = f"{cnt/total*100:.1f}%" if total else "0%"
        ws2.append([f"{lvl} Confidence", cnt, pct])
        rn = ws2.max_row
        ws2.row_dimensions[rn].height = 20
        for ci, v in enumerate([fill, None, None], 1):
            c = ws2.cell(row=rn, column=ci)
            c.font, c.border = (BOLD, THIN)
            c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")
            if ci == 1:
                c.fill = fill

    ws2.append(["Total Transactions", total, "100%"])
    rn = ws2.max_row
    ws2.row_dimensions[rn].height = 22
    for ci in range(1, 4):
        c = ws2.cell(row=rn, column=ci)
        c.font, c.border = BOLD, THIN
        c.alignment = Alignment(horizontal="center" if ci > 1 else "left", vertical="center")

    for col in ws2.columns:
        ltr = get_column_letter(col[0].column)
        ws2.column_dimensions[ltr].width = max(max(len(str(c.value or '')) for c in col) + 5, 20)

    wb.save(output_path)


# ---------------------------------------------------------------------------
# CLI entry point
# ---------------------------------------------------------------------------

# ---------------------------------------------------------------------------
# Gemini LLM fallback — resolves rows the rule engine could not match (Low).
# SAFETY: the model may ONLY return a ledger that exists in `candidates` (all of
# which are real COA ledgers). Anything else — abstention, an invented name, a
# timeout, an API error — yields None, and the caller keeps "Suspense A/c".
# It can never produce a wrong/invented ledger.
# ---------------------------------------------------------------------------
def gemini_classify(narration, candidates, api_key, model="gemini-2.5-flash", timeout=30):
    import urllib.request
    import json as _json
    if not api_key or not candidates or not str(narration).strip():
        return None
    cand_block = "\n".join(candidates)
    prompt = (
        "You are an expert Indian bank-reconciliation assistant mapping a bank transaction "
        "to a Tally ledger. From the CANDIDATES list, choose the ONE ledger that best matches "
        "the transaction narration (consider payee/vendor name, UPI/NEFT counterparty, and "
        "purpose). You MUST copy a candidate EXACTLY as written. If none is a clear match, "
        "reply exactly: SUSPENSE. Reply with ONLY the chosen ledger name or SUSPENSE — no "
        "explanation, no extra text.\n\n"
        f"TRANSACTION NARRATION:\n{narration}\n\nCANDIDATES:\n{cand_block}"
    )
    body = _json.dumps({
        "contents": [{"parts": [{"text": prompt}]}],
        "generationConfig": {"temperature": 0, "maxOutputTokens": 2000},
    }).encode("utf-8")
    url = (f"https://generativelanguage.googleapis.com/v1beta/models/"
           f"{model}:generateContent?key={api_key}")
    # Build an SSL context backed by certifi's CA bundle. Many environments (notably
    # macOS Python) lack a system CA bundle, which makes urllib fail every HTTPS call
    # with CERTIFICATE_VERIFY_FAILED — silently disabling Gemini. certifi fixes that.
    import ssl
    try:
        import certifi
        _ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        _ctx = ssl.create_default_context()
    try:
        req = urllib.request.Request(url, data=body,
                                     headers={"Content-Type": "application/json"})
        with urllib.request.urlopen(req, timeout=timeout, context=_ctx) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
        text = data["candidates"][0]["content"]["parts"][0]["text"].strip()
    except Exception as _e:
        # Safe fallback (caller keeps Suspense A/c) but log so failures aren't invisible.
        print(f"      → Gemini call failed: {type(_e).__name__}: {str(_e)[:160]}", file=sys.stderr)
        return None
    if not text or text.strip().upper() == "SUSPENSE":
        return None
    # Strict validation: the reply MUST exactly (normalised) equal a candidate.
    def _n(x):
        return ' '.join(re.sub(r'[^a-z0-9 ]', ' ', str(x).lower()).split())
    nt = _n(text)
    for c in candidates:
        if _n(c) == nt:
            return c
    return None  # model returned something not in the COA candidate list → reject


# ---------------------------------------------------------------------------
# Claude (Anthropic) LLM fallback — same contract as gemini_classify: the model
# may ONLY return a ledger that exists in `candidates`; anything else (abstention,
# invented name, timeout, API/refusal error) yields None and the caller keeps
# "Suspense A/c". Raw HTTPS (urllib) to mirror gemini_classify — no SDK dependency.
# ---------------------------------------------------------------------------
def _post_via(router, api_key, model, ask, max_tokens, timeout, base_url=None,
              cached_block=None):
    """Send through the router when one exists (so failover applies), else direct."""
    if router is not None:
        return router.post(model, ask, max_tokens, timeout, cached_block)
    return _llm_post(api_key, model, ask, max_tokens, timeout, base_url, cached_block)


def anthropic_side_verdict(narration, credit_ledger, debit_ledger, assigned,
                           debit, credit, api_key, model="claude-haiku-4-5", timeout=30,
                           base_url=None, router=None):
    """Constrained verdict for a row a per-brand side rule already claimed.

    Claude may ONLY answer with one of the rule's own two ledgers, ABSTAIN, or
    NOT-THIS-VENDOR. It can never write an unrelated COA name onto a side-rule row --
    that containment is the whole point, and is what stops the arbitration failure from
    recurring in a new form.

    Returns (verdict, ledger):
      ('confirm',  <assigned>)      Claude agrees with the amount-implied side
      ('flip',     <other side>)    genuine refund / reversal -> use the other side
      ('abstain',  <assigned>)      ambiguous -> keep the ledger, caller demotes to Medium
      ('reject',   None)            token matched but the counterparty is someone else
      (None,       None)            call failed -> caller keeps the rule's answer as-is
    """
    import urllib.request, json as _json, ssl
    if not api_key or not str(narration).strip():
        return (None, None)
    other = debit_ledger if assigned == credit_ledger else credit_ledger
    # The bank statement itself is the authority on direction -- state it explicitly.
    # Without this the model infers direction from narration wording, which carries no
    # reliable cue on most rails, and it flips rows essentially at random.
    if credit and credit > 0:
        direction = f"money was RECEIVED into the account (credit {credit})"
    else:
        direction = f"money was PAID out of the account (debit {debit})"
    prompt = (
        "You are an expert Indian bank-reconciliation assistant. A per-brand rule matched "
        "a vendor token in this bank narration and assigned a Tally ledger based on which "
        "side the money moved.\n\n"
        f"NARRATION:\n{narration}\n\n"
        f"BANK STATEMENT FACT: {direction}. This is authoritative -- do not second-guess "
        "the direction of the money.\n\n"
        f"CREDIT-SIDE LEDGER (used when money is received): {credit_ledger}\n"
        f"DEBIT-SIDE LEDGER  (used when money is paid):     {debit_ledger}\n"
        f"RULE ASSIGNED: {assigned}\n\n"
        "The assigned ledger is correct in the overwhelming majority of cases. Answer "
        "OTHER-SIDE ONLY if the narration gives POSITIVE evidence of a refund, reversal, "
        "chargeback or return -- not merely because the wording is ambiguous.\n\n"
        "Reply with EXACTLY ONE of these four tokens and nothing else:\n"
        "  CONFIRM          - the assigned ledger is correct\n"
        "  OTHER-SIDE       - explicit refund/reversal evidence; the other ledger is correct\n"
        "  NOT-THIS-VENDOR  - the token matched but the real counterparty is a different "
        "party (e.g. a NACH mandate collected by an aggregator on behalf of someone else)\n"
        "  UNSURE           - genuinely ambiguous\n"
    )
    try:
        _txt, _ = _post_via(router, api_key, model, prompt, max_tokens=16,
                            timeout=timeout, base_url=base_url)
        data = {"content": [{"type": "text", "text": _txt}]}
    except Exception as _e:
        print(f"      → Claude side-verdict failed: {type(_e).__name__}: {str(_e)[:140]}",
              file=sys.stderr)
        return (None, None)
    text = ""
    for block in data.get("content", []):
        if block.get("type") == "text":
            text = (block.get("text") or "").strip().upper()
            break
    if "NOT-THIS-VENDOR" in text:
        return ("reject", None)
    if "OTHER-SIDE" in text:
        return ("flip", other)
    if "CONFIRM" in text:
        return ("confirm", assigned)
    return ("abstain", assigned)          # UNSURE, empty, or anything unrecognised


# ── LLM transport ─────────────────────────────────────────────────────────────
# The SAME Claude models are reachable over two different wire formats:
#   * Anthropic native   POST https://api.anthropic.com/v1/messages
#                        header  x-api-key + anthropic-version
#                        reply   data["content"][0]["text"]
#   * OpenAI-compatible  POST {base}/chat/completions      (GenSpark's llm_proxy)
#                        header  Authorization: Bearer
#                        reply   data["choices"][0]["message"]["content"]
#
# GenSpark's proxy is the OpenAI shape but forwards to Anthropic underneath, so an
# Anthropic-style `cache_control` block in the system message IS honoured — verified
# 2026-07-28: 8,017 tokens written to cache on the first call and read back on the
# second. That matters because it means the batching + prompt-caching cost design
# works identically on either transport; we are not trading cost for availability.
LLM_UA = "curl/8.7.1"   # urllib's default User-Agent trips Cloudflare 1010 on GenSpark


def _llm_ssl_ctx():
    import ssl                                   # imported locally, as elsewhere in this file
    try:
        import certifi
        return ssl.create_default_context(cafile=certifi.where())
    except Exception:
        return ssl.create_default_context()


def _credential_is_finished(exc):
    """Does this error mean the credential will not work again for the rest of the run?

    Returns (True, reason) only for conditions that are permanent for this key:
      * 401 / 403        — revoked, wrong, or not permitted
      * 400 + a billing/quota message — Anthropic reports an exhausted balance as a
        400 ("Your credit balance is too low"), NOT a 402, so the status alone is not
        enough; the body has to be read.

    Deliberately NOT failover-worthy: timeouts, connection resets, 429 (rate limit) and
    529 (overloaded). Those are transient on a WORKING credential — treating them as
    fatal would abandon the primary for the whole run because one call was unlucky, and
    would throw away the prompt cache already built up on it.
    """
    import urllib.error
    if not isinstance(exc, urllib.error.HTTPError):
        return False, ""                      # timeout / DNS / reset -> transient
    try:
        body = exc.read().decode("utf-8", "replace")[:400]
    except Exception:
        body = ""
    if exc.code in (401, 403):
        return True, f"HTTP {exc.code}"
    if exc.code == 400 and re.search(
            r"credit balance|billing|insufficient|quota|exceeded your", body, re.I):
        return True, f"HTTP 400 — {' '.join(body.split())[:100]}"
    return False, ""


class _LLMRouter:
    """One primary transport with a sticky, one-way failover to a secondary.

    The switch happens at most ONCE per run. After it, every later call goes straight to
    the secondary — the alternative (retrying the primary each time) would pay a failed
    round trip on every single call once the primary is dry.

    Note the cache consequence: prompt caches are per-endpoint, so the first call after a
    switch writes a fresh cache on the secondary instead of reading the primary's. That
    is a one-off cost of the failover, not a per-call one.
    """

    def __init__(self, primary, secondary=None):
        import threading
        self.active = primary                 # (label, api_key, base_url)
        self.standby = secondary
        self.switched = False
        self._lock = threading.Lock()

    @property
    def label(self):
        return self.active[0]

    def post(self, model, ask, max_tokens, timeout, cached_block=None):
        # Layer 3.4 runs six threads in parallel, so several calls are usually in flight
        # when the primary dies. Each of those threads gets the same fatal error. Only
        # ONE performs the switch; the others must RETRY on the new transport rather than
        # give up — otherwise a single dead credential fails every request that happened
        # to be concurrent with the discovery (measured: 66 of 135 rows lost that way).
        for _ in range(2):
            label, key, base = self.active
            try:
                return _llm_post(key, model, ask, max_tokens, timeout, base, cached_block)
            except Exception as exc:
                finished, why = _credential_is_finished(exc)
                if not finished:
                    raise                       # transient — caller decides
                with self._lock:
                    if self.active != (label, key, base):
                        continue                # another thread already switched: retry
                    if not self.standby:
                        raise                   # nothing left to fall back to
                    nlabel = self.standby[0]
                    print(f"[llm] {label} credential is finished ({why}) — failing over "
                          f"to {nlabel} for the rest of this run", file=sys.stderr)
                    self.active, self.standby, self.switched = self.standby, None, True
                continue                        # retry once on the new transport
        raise RuntimeError("LLM transport exhausted after failover")


def _llm_post(api_key, model, ask, max_tokens, timeout, base_url=None,
              cached_block=None):
    """One completion request over whichever wire format `base_url` implies.

    `cached_block` is the large, byte-stable prefix (the chart of accounts) and is sent
    in its own content block marked for caching. Returns (text, usage_dict).
    """
    import urllib.request, json as _json
    payload_msgs = []
    if base_url:                                    # OpenAI-compatible
        url = f"{base_url.rstrip('/')}/chat/completions"
        headers = {"content-type": "application/json",
                   "authorization": f"Bearer {api_key}",
                   "user-agent": LLM_UA}
        if cached_block:
            payload_msgs.append({"role": "system", "content": [
                {"type": "text", "text": cached_block,
                 "cache_control": {"type": "ephemeral"}}]})
        payload_msgs.append({"role": "user", "content": ask})
    else:                                           # Anthropic native
        url = "https://api.anthropic.com/v1/messages"
        headers = {"content-type": "application/json", "x-api-key": api_key,
                   "anthropic-version": "2023-06-01"}
        content = []
        if cached_block:
            content.append({"type": "text", "text": cached_block,
                            "cache_control": {"type": "ephemeral"}})
        content.append({"type": "text", "text": ask})
        payload_msgs.append({"role": "user", "content": content})

    body = _json.dumps({"model": model, "max_tokens": max_tokens,
                        "messages": payload_msgs}).encode("utf-8")
    req = urllib.request.Request(url, data=body, headers=headers)
    with urllib.request.urlopen(req, timeout=timeout, context=_llm_ssl_ctx()) as resp:
        data = _json.loads(resp.read().decode("utf-8"))
    if base_url:
        text = ((data.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
    else:
        text = next((b.get("text", "") for b in data.get("content", [])
                     if b.get("type") == "text"), "")
    return text, (data.get("usage") or {})


def anthropic_classify_batch(items, coa, api_key, model="claude-haiku-4-5",
                             timeout=120, chunk=25, base_url=None, router=None):
    """Classify MANY rows against ONE copy of the chart of accounts.

    WHY: the per-row call sent the whole CoA every time. On FLO (3,968 ledgers) a single
    statement with ~60 unresolved rows meant ~2.14M input tokens, because the same ledger
    list was re-sent 60 times. The CoA is identical for every row of a brand, so it should
    be sent ONCE.

    Two savings, stacked:
      * BATCHING — rows are numbered and answered in one response. ~20x fewer tokens.
      * PROMPT CACHING — the CoA sits in its own content block marked `cache_control`,
        in a STABLE (sorted) order so the prefix is byte-identical across chunks, retries
        and re-runs. Anthropic bills a cache hit at ~10%. Another ~2.5x.
    Together roughly 50x: FLO ~2.14M -> ~43k input tokens per run.

    Per-row name ranking is kept as a short "closest matches" hint inside each row, so no
    signal is lost by sorting the shared block.

    `items` = [{'idx': int, 'narration': str, 'hints': [str, ...]}]
    Returns {idx: ledger} for rows the model resolved; absent/unmatched rows are simply
    not returned, and the caller leaves them untouched.
    """
    import urllib.request, json as _json, ssl
    if not api_key or not items or not coa:
        return {}

    # Stable order => cacheable prefix. Never reorder this per row.
    coa_sorted = sorted(set(coa))
    coa_block = "\n".join(coa_sorted)
    valid = {' '.join(re.sub(r'[^a-z0-9 ]', ' ', l.lower()).split()): l for l in coa_sorted}

    try:
        import certifi
        ctx = ssl.create_default_context(cafile=certifi.where())
    except Exception:
        ctx = ssl.create_default_context()

    out = {}
    for start in range(0, len(items), chunk):
        batch = items[start:start + chunk]
        lines = []
        for it in batch:
            hint = ", ".join(it.get('hints') or [])
            lines.append(f"{it['idx']}. {it['narration']}"
                         + (f"\n   closest name matches: {hint}" if hint else ""))
        ask = (
            "For EACH numbered bank transaction below, choose the ONE ledger from the "
            "LEDGER LIST that best matches it (consider payee/vendor name, UPI/NEFT "
            "counterparty, and purpose). You MUST copy a ledger EXACTLY as written in the "
            "list. If none is a clear match for a row, answer SUSPENSE for that row.\n\n"
            "Reply with one line per transaction, formatted exactly as:\n"
            "<number>: <ledger name>\n"
            "No other text.\n\n"
            "TRANSACTIONS:\n" + "\n".join(lines)
        )
        # Cached prefix — byte-identical across every chunk, every retry and every run
        # for a brand, which is what makes the cache hit.
        cached_block = ("You are an expert Indian bank-reconciliation assistant.\n\n"
                        "LEDGER LIST (the brand's full chart of accounts):\n" + coa_block)
        try:
            text, usage = _post_via(router, api_key, model, ask,
                                    max_tokens=32 * len(batch) + 256,
                                    timeout=timeout, base_url=base_url,
                                    cached_block=cached_block)
        except Exception as _e:
            _detail = getattr(_e, 'read', lambda: b'')()
            print(f"      → LLM batch failed ({type(_e).__name__}: {str(_e)[:110]} "
                  f"{_detail[:120]}); rows in this chunk keep their rule answer",
                  file=sys.stderr)
            continue

        if usage:
            # prompt_tokens on the OpenAI shape, input_tokens on the native one.
            print(f"        chunk {start//chunk + 1}: "
                  f"in={usage.get('input_tokens') or usage.get('prompt_tokens')} "
                  f"cache_write={usage.get('cache_creation_input_tokens')} "
                  f"cache_read={usage.get('cache_read_input_tokens')}")
        for line in text.splitlines():
            mm = re.match(r'\s*(\d+)\s*[:.\)]\s*(.+?)\s*$', line)
            if not mm:
                continue
            idx, pick = int(mm.group(1)), mm.group(2).strip()
            if pick.upper() == "SUSPENSE":
                continue
            # Only a ledger that really exists is accepted — never free text.
            hit = valid.get(' '.join(re.sub(r'[^a-z0-9 ]', ' ', pick.lower()).split()))
            if hit:
                out[idx] = hit
    return out


def anthropic_classify(narration, candidates, api_key, model="claude-haiku-4-5", timeout=30,
                       base_url=None, router=None):
    import urllib.request
    import json as _json
    if not api_key or not candidates or not str(narration).strip():
        return None
    cand_block = "\n".join(candidates)
    prompt = (
        "You are an expert Indian bank-reconciliation assistant mapping a bank transaction "
        "to a Tally ledger. From the CANDIDATES list, choose the ONE ledger that best matches "
        "the transaction narration (consider payee/vendor name, UPI/NEFT counterparty, and "
        "purpose). You MUST copy a candidate EXACTLY as written. If none is a clear match, "
        "reply exactly: SUSPENSE. Reply with ONLY the chosen ledger name or SUSPENSE — no "
        "explanation, no extra text.\n\n"
        f"TRANSACTION NARRATION:\n{narration}\n\nCANDIDATES:\n{cand_block}"
    )
    try:
        _txt, _ = _post_via(router, api_key, model, prompt, max_tokens=100,
                            timeout=timeout, base_url=base_url)
        data = {"content": [{"type": "text", "text": _txt}]}
    except Exception as _e:
        print(f"      → Claude call failed: {type(_e).__name__}: {str(_e)[:160]}", file=sys.stderr)
        return None
    if data.get("stop_reason") == "refusal":
        return None
    text = ""
    for block in data.get("content", []):
        if block.get("type") == "text":
            text = (block.get("text") or "").strip()
            break
    if not text or text.strip().upper() == "SUSPENSE":
        return None
    def _n(x):
        return ' '.join(re.sub(r'[^a-z0-9 ]', ' ', str(x).lower()).split())
    nt = _n(text)
    for c in candidates:
        if _n(c) == nt:
            return c
    return None  # model returned something not in the COA candidate list → reject


def main():
    import argparse
    parser = argparse.ArgumentParser(description="Universal Tally Bank Statement Classifier")
    parser.add_argument("--ledger", required=True,
                        help="Path to List of Ledgers Tally export (Excel)")
    parser.add_argument("--bank",   required=False,
                        help="Path to raw Bank Statement file (Excel or CSV)")
    parser.add_argument("--output", required=False,
                        help="Path to save the classified output Excel file")
    parser.add_argument("--brand",  default="Brand",
                        help="Brand name for the output filename/summary (default: Brand)")
    parser.add_argument("--corrections", default=None,
                        help="Path to corrections JSON file: {NARRATION_KEY: {ledger, type}}")
    parser.add_argument("--side-map", default=None,
                        help="Path to a per-brand side-dependent ledger map JSON "
                             "({counterparties:[{tokens,credit,debit}]}). Credit ledger for "
                             "Receipts, debit ledger for Payments. Checked above the directory.")
    parser.add_argument("--gemini-key", default=None,
                        help="Gemini API key for the LLM fallback (else env GEMINI_API_KEY)")
    parser.add_argument("--gemini-model", default="gemini-2.5-flash",
                        help="Gemini model for the fallback (default: gemini-2.5-flash)")
    parser.add_argument("--anthropic-key", default=None,
                        help="Anthropic (Claude) API key for the LLM fallback (else env "
                             "ANTHROPIC_API_KEY). Preferred over Gemini when present.")
    parser.add_argument("--llm-base-url", default=None,
                        help="OpenAI-compatible base URL (e.g. GenSpark's llm_proxy). When "
                             "set, the LLM key is sent as 'Authorization: Bearer' to "
                             "{base}/chat/completions instead of Anthropic's /v1/messages. "
                             "Defaults to $GSK_BASE_URL when only a GenSpark key is available.")
    parser.add_argument("--anthropic-model", default="claude-haiku-4-5",
                        help="Claude model for the fallback (default: claude-haiku-4-5)")
    parser.add_argument("--list-ledgers", action="store_true",
                        help="COA ingest mode: load + validate --ledger and print the cleaned "
                             "ledger names as a JSON array to stdout, then exit. Applies the same "
                             "Tally-group/metadata filtering and COA-integrity guard as a real run, "
                             "so a bank statement uploaded as a COA is rejected here with a clear error.")
    args = parser.parse_args()

    # ------------------------------------------------------------------
    # COA INGEST MODE (--list-ledgers): used by the backend to validate + extract a
    # brand's Chart of Accounts before storing it in the ledger_master DB table.
    # Reuses load_ledger_master() so cleaning + the integrity guard are identical to a
    # classification run. On a bad file (bank statement / date column) load_ledger_master
    # raises ValueError → we print it to stderr and exit 2 so the caller can surface it.
    # ------------------------------------------------------------------
    if args.list_ledgers:
        import json as _json
        try:
            _ledgers = load_ledger_master(args.ledger)
        except ValueError as _e:
            print(str(_e), file=sys.stderr)
            sys.exit(2)
        sys.stdout.write(_json.dumps(_ledgers, ensure_ascii=False))
        sys.stdout.flush()
        sys.exit(0)

    if not args.bank or not args.output:
        parser.error("--bank and --output are required unless --list-ledgers is set")

    print(f"[1/4] Loading ledger master: {args.ledger}")
    ledgers = load_ledger_master(args.ledger)
    print(f"      → {len(ledgers)} ledgers loaded.")

    print(f"[2/4] Loading bank statement: {args.bank}")
    df_bank, col_map, sheet = load_bank_statement(args.bank)
    print(f"      → Sheet: '{sheet}' | Columns: {col_map}")

    if not col_map.get("description"):
        print("ERROR: Could not detect a Description/Narration column. "
              "Check that the bank file has a clear header row within the first 30 rows.", file=sys.stderr)
        sys.exit(1)

    # Load per-brand corrections (Layer 0 — highest priority, checked before CoA/fuzzy).
    # Prefer --corrections <path> if provided by caller (Node.js writes this from DB).
    # Fall back to sidecar convention: corrections/<brand-slug>_corrections.json
    # Format: {"NORMALIZED NARRATION KEY": {"ledger": "Ledger Name", "type": "Payment"}}
    import json as _json
    corrections = {}
    _corr_path = None
    if args.corrections and os.path.exists(args.corrections):
        _corr_path = args.corrections
    else:
        _slug = re.sub(r'[^A-Za-z0-9_-]', '-', args.brand.strip()).lower()
        _sidecar = os.path.join(os.path.dirname(args.output), '..', 'corrections', f'{_slug}_corrections.json')
        _sidecar = os.path.normpath(_sidecar)
        if os.path.exists(_sidecar):
            _corr_path = _sidecar
    if _corr_path:
        try:
            with open(_corr_path, 'r', encoding='utf-8') as _f:
                corrections = _json.load(_f)
            print(f"      → Loaded {len(corrections)} stored corrections")
        except Exception as _e:
            print(f"      → Warning: could not load corrections file: {_e}")

    # Optional per-brand side-dependent ledger map (M-Brands marketplaces). Only passed
    # by the backend for the specific brand; absent → empty → no behaviour change.
    side_map = []
    if args.side_map and os.path.exists(args.side_map):
        try:
            with open(args.side_map, 'r', encoding='utf-8') as _f:
                _sm = _json.load(_f)
            side_map = _sm.get('counterparties', []) if isinstance(_sm, dict) else (_sm or [])
            print(f"      → Loaded side-ledger map: {len(side_map)} counterparties")
        except Exception as _e:
            print(f"      → Warning: could not load side-map file: {_e}")

    print("[3/4] Classifying transactions …")
    classifier = BankClassifier(ledgers, corrections=corrections, brand_name=args.brand,
                                side_map=side_map)
    rows, summary = [], {"High": 0, "Medium": 0, "Low": 0}

    date_col  = col_map.get("txn_date")
    desc_col  = col_map.get("description")
    debit_col = col_map.get("debit")
    credit_col = col_map.get("credit")
    bal_col   = col_map.get("balance")
    chq_col   = col_map.get("chq_ref")

    _JUNK_RE = re.compile(
        r'^[*.\-=~_\s]+$'                        # separator lines: ***, ..., ---
        r'|^-+\s*end\s+of\s+statement\s*-+$'     # --- End Of Statement ---
        r'|statement\s+summary'                   # STATEMENT SUMMARY :-
        r'|opening\s+balance'                     # Opening Balance
        r'|closing\s+balance'                     # Closing Balance (footer)
        r'|generated\s+on'                        # Generated On: ...
        r'|state\s+account\s+branch'              # State account branch GSTIN
        r'|bank\s+gstin\s+number'                 # Bank GSTIN number details...
        r'|lower\s+parel'                         # FC Bank House, ...Lower Parel
        r'|senapati\s+bapat'                      # bank address line
        r'|www\.'                                 # URL rows
        r'|payment.*goods.*service.*tax',         # GSTIN URL path fragment
        re.IGNORECASE,
    )

    for _, row in df_bank.iterrows():
        txn_date = row.get(date_col, "") if date_col else ""
        desc     = row.get(desc_col)
        if pd.isna(desc) and pd.isna(txn_date):
            continue

        desc_str_check = str(desc).strip() if pd.notna(desc) else ""
        if not desc_str_check or _JUNK_RE.search(desc_str_check):
            continue

        # Skip rows with no monetary value — they are metadata/summary lines
        # (separators, "Generated On", GSTIN lines, bank address lines, etc.)
        # Real transactions always have a non-zero debit OR credit.
        def _to_float(v):
            try: return float(v) if pd.notna(v) and str(v).strip() not in ('', '-') else 0.0
            except (ValueError, TypeError): return 0.0
        if _to_float(row.get(debit_col) if debit_col else None) == 0.0 and \
           _to_float(row.get(credit_col) if credit_col else None) == 0.0:
            continue

        def format_excel_date(val):
            import datetime
            if pd.isna(val) or val == "":
                return ""
            if isinstance(val, (datetime.datetime, datetime.date, pd.Timestamp)):
                return val.strftime("%d-%m-%Y")
            try:
                num = float(val)
                if 30000 <= num <= 60000:
                    base = datetime.datetime(1899, 12, 30)
                    delta = datetime.timedelta(days=num)
                    return (base + delta).strftime("%d-%m-%Y")
            except Exception:
                pass
            val_str = str(val).strip()
            for fmt in ["%Y-%m-%d %H:%M:%S", "%Y-%m-%d", "%d-%m-%Y", "%d/%m/%Y", "%Y/%m/%d"]:
                try:
                    dt = datetime.datetime.strptime(val_str.split(".")[0], fmt)
                    return dt.strftime("%d-%m-%Y")
                except Exception:
                    continue
            if " 00:00:00" in val_str:
                val_str = val_str.replace(" 00:00:00", "")
            return val_str

        formatted_date = format_excel_date(txn_date)

        def safe_float(v):
            try:
                return float(v) if pd.notna(v) and v != "" else 0.0
            except (ValueError, TypeError):
                return 0.0

        debit   = safe_float(row.get(debit_col))  if debit_col  else 0.0
        credit  = safe_float(row.get(credit_col)) if credit_col else 0.0
        balance = safe_float(row.get(bal_col))    if bal_col    else 0.0
        desc_str = str(desc).strip() if pd.notna(desc) else ""

        chq_ref_val = ""
        if chq_col is not None:
            _cv = row.get(chq_col)
            if pd.notna(_cv):
                chq_ref_val = str(_cv).strip()
                if chq_ref_val.endswith(".0"):
                    chq_ref_val = chq_ref_val[:-2]

        result = classifier.classify(desc_str, debit, credit)
        summary[result["confidence"]] += 1
        rows.append({
            "txn_date":       formatted_date,
            "description":    desc_str,
            "chq_ref":        chq_ref_val,
            "debit":          debit,
            "credit":         credit,
            "balance":        balance,
            "predicted_type": result["type"],
            "predicted_ledger": result["ledger"],
            "confidence":     result["confidence"],
            "entity":         result.get("entity", ""),
            "rule":           result.get("rule", ""),
            # only populated for side-rule rows; consumed by the side-verdict pass below
            "side_credit":    result.get("credit", ""),
            "side_debit":     result.get("debit", ""),
        })

    total = len(rows)
    print(f"      → {total} rows processed.")
    print(f"         High:   {summary['High']}  ({summary['High']/total*100:.1f}%)")
    print(f"         Medium: {summary['Medium']}  ({summary['Medium']/total*100:.1f}%)")
    print(f"         Low:    {summary['Low']}  ({summary['Low']/total*100:.1f}%)")

    # ------------------------------------------------------------------
    # LLM fallback — fires on Low AND Medium rows (Medium may be uncertain; the
    # LLM can confirm or correct). High rows are never touched. On any failure the
    # row stays unchanged. Claude is preferred when its key is present; else Gemini.
    # ------------------------------------------------------------------
    anthropic_key = args.anthropic_key or os.environ.get("ANTHROPIC_API_KEY")
    # Transport selection. An Anthropic key talks to api.anthropic.com directly. With no
    # Anthropic key we fall back to a GenSpark proxy key if one is configured: same Claude
    # models, OpenAI-shaped wire format, and prompt caching still works through it, so the
    # batching/caching cost design is unchanged. Explicit --llm-base-url overrides both.
    llm_base_url = args.llm_base_url or None
    # LLM_PROVIDER=genspark forces the proxy even when ANTHROPIC_API_KEY is present.
    # Needed because an EXHAUSTED Anthropic key is still a set key: it fails with 400
    # "credit balance is too low", which an "is the key missing?" test cannot see. To go
    # back to Anthropic, drop LLM_PROVIDER — nothing else changes.
    _prefer_gsk = os.environ.get("LLM_PROVIDER", "").strip().lower() in ("genspark", "gsk")
    _gsk = os.environ.get("GSK_API_KEY")
    _gsk_base = llm_base_url or os.environ.get("GSK_BASE_URL")

    # Build an ordered transport chain. Anthropic is preferred when its key is present,
    # with GenSpark on standby; LLM_PROVIDER=genspark inverts that. Failover is sticky
    # and fires only on a permanently-dead credential (see _credential_is_finished), so
    # an exhausted key costs ONE failed call per run rather than one per request.
    _anthropic_leg = ("Anthropic", anthropic_key, None) if anthropic_key else None
    _gsk_leg = ("GenSpark", _gsk, _gsk_base) if (_gsk and _gsk_base) else None
    _chain = ([_gsk_leg, _anthropic_leg] if _prefer_gsk
              else [_anthropic_leg, _gsk_leg])
    _chain = [leg for leg in _chain if leg]
    llm_router = None
    if _chain:
        llm_router = _LLMRouter(_chain[0], _chain[1] if len(_chain) > 1 else None)
        anthropic_key = _chain[0][1]          # keeps the "is an LLM available?" checks true
        llm_base_url = _chain[0][2]
        _standby = f", standby {_chain[1][0]}" if len(_chain) > 1 else ", no standby"
        print(f"[llm] primary {_chain[0][0]}"
              + (f" ({_chain[0][2]})" if _chain[0][2] else "") + _standby)
    elif _prefer_gsk:
        print("[llm] LLM_PROVIDER=genspark but GSK_API_KEY/GSK_BASE_URL missing",
              file=sys.stderr)
    gemini_key = args.gemini_key or os.environ.get("GEMINI_API_KEY")
    llm_key = anthropic_key or gemini_key
    llm_name = (f"Claude ({args.anthropic_model})" if anthropic_key
                else f"Gemini ({args.gemini_model})")

    def llm_pick(desc, cands):
        if anthropic_key:
            return anthropic_classify(desc, cands, anthropic_key, args.anthropic_model,
                                      base_url=llm_base_url, router=llm_router)
        # Was `return llm_pick(...)` — an infinite self-call. Only reachable with a Gemini
        # key and no Anthropic/GenSpark key, which is why it survived: every run so far had
        # one of the other two. It would have hit RecursionError, not a graceful skip.
        return gemini_classify(desc, cands, gemini_key, args.gemini_model)

    # ------------------------------------------------------------------
    # STEP 3.4 — constrained verdict on per-brand SIDE-RULE rows.
    # Runs BEFORE the Low/Medium fallback so a NOT-THIS-VENDOR row is re-classified in
    # time to be picked up by it. Claude is offered only the rule's own two ledgers, so
    # it can confirm, flip the side, or disown the vendor -- never invent a third ledger.
    # Verdicts cache on (entity, side, payee-key), so ~150 side rows cost ~10 calls.
    # ------------------------------------------------------------------
    if anthropic_key and total:
        side_idx = [i for i, r in enumerate(rows)
                    if str(r.get("rule", "")).startswith("Side Ledger (credit/debit)")
                    and r.get("side_credit") and r.get("side_debit")]
        if side_idx:
            print(f"[3.4/4] Claude side-rule verdict on {len(side_idx)} rows …")
            from concurrent.futures import ThreadPoolExecutor as _STPE
            _scache = {}

            def _verdict(i):
                r = rows[i]
                pk = extract_payee_keys(r["description"])
                ck = (r.get("entity", ""), r["predicted_ledger"],
                      pk.get("name") or pk.get("neft_name") or pk.get("exact", ""))
                if ck in _scache:
                    return i, _scache[ck]
                out = anthropic_side_verdict(
                    r["description"], r["side_credit"], r["side_debit"],
                    r["predicted_ledger"], r["debit"], r["credit"],
                    anthropic_key, args.anthropic_model, base_url=llm_base_url,
                    router=llm_router)
                _scache[ck] = out
                return i, out

            n_conf = n_flip = n_abst = n_rej = n_err = 0
            try:
                with _STPE(max_workers=6) as ex:
                    for i, (verdict, ledger) in ex.map(_verdict, side_idx):
                        if verdict is None:
                            # Fail OPEN: the rule is accountant-authored and was right ~95%
                            # of the time unaided. An API outage must not dump every side
                            # row into the review queue.
                            n_err += 1
                            continue
                        if verdict == "confirm":
                            n_conf += 1
                        elif verdict == "flip":
                            rows[i]["predicted_ledger"] = ledger
                            rows[i]["rule"] = "Side Ledger (flipped)"
                            n_flip += 1
                        elif verdict == "abstain":
                            if rows[i]["confidence"] == "High":
                                rows[i]["confidence"] = "Medium"
                                summary["High"] -= 1; summary["Medium"] += 1
                            rows[i]["rule"] = "Side Ledger (unconfirmed)"
                            n_abst += 1
                        elif verdict == "reject":
                            res = classifier.classify(rows[i]["description"],
                                                      rows[i]["debit"], rows[i]["credit"],
                                                      skip_side_map=True)
                            old = rows[i]["confidence"]
                            rows[i].update({
                                "predicted_ledger": res["ledger"],
                                "predicted_type":   res["type"],
                                "confidence":       res["confidence"],
                                "entity":           res.get("entity", ""),
                                "rule":             res.get("rule", "") + " (side rule disowned)",
                            })
                            if old != res["confidence"]:
                                summary[old] -= 1; summary[res["confidence"]] += 1
                            n_rej += 1
            except Exception as _e:
                print(f"        → side-verdict pass aborted ({_e}); rows kept as-is")
            print(f"        → confirmed {n_conf}, flipped {n_flip}, unconfirmed {n_abst}, "
                  f"disowned {n_rej}, call-failed {n_err} (kept High)")

    if llm_key and total:
        low_idx = [i for i, r in enumerate(rows) if r["confidence"] in ("Low", "Medium")]
        if low_idx:
            print(f"[3.5/4] {llm_name} fallback on {len(low_idx)} low/medium-confidence rows …")
            suspense_label = classifier._suspense_ledger

            # De-duplicate by narration BEFORE calling: identical narrations get one
            # decision, applied to every row that shares it.
            by_key = {}
            for i in low_idx:
                by_key.setdefault(' '.join(str(rows[i]["description"]).upper().split()),
                                  []).append(i)

            picks = {}
            if anthropic_key:
                # Batched: ONE copy of the CoA for the whole run instead of one per row.
                # Per-row name ranking survives as a short hint, so sorting the shared
                # (cacheable) ledger block costs nothing.
                items = []
                for n, (key, idxs) in enumerate(by_key.items()):
                    i0 = idxs[0]
                    ent = rows[i0].get("entity", "") or rows[i0]["description"]
                    items.append({"idx": n, "narration": rows[i0]["description"],
                                  "hints": classifier.top_candidates(ent, k=8)})
                answers = anthropic_classify_batch(items, classifier.master_ledgers,
                                                   anthropic_key, args.anthropic_model,
                                                   base_url=llm_base_url,
                                                   router=llm_router)
                for n, key in enumerate(by_key):
                    if n in answers:
                        picks[key] = answers[n]
            else:
                # Gemini path unchanged: per-row, candidate-constrained.
                from concurrent.futures import ThreadPoolExecutor
                def _resolve(key):
                    i0 = by_key[key][0]
                    ent = rows[i0].get("entity", "") or rows[i0]["description"]
                    cands = classifier.llm_candidates(ent, rows[i0]["description"], k=15)
                    if suspense_label not in cands:
                        cands = cands + [suspense_label]
                    return key, llm_pick(rows[i0]["description"], cands)
                try:
                    with ThreadPoolExecutor(max_workers=6) as ex:
                        for key, pick in ex.map(_resolve, list(by_key)):
                            if pick:
                                picks[key] = pick
                except Exception as _e:
                    print(f"        → {llm_name} fallback aborted ({_e}); rows kept as-is")

            resolved = 0
            for key, idxs in by_key.items():
                pick = picks.get(key)
                if not pick or pick == suspense_label:
                    continue
                for i in idxs:
                    old_conf = rows[i]["confidence"]
                    rows[i]["predicted_ledger"] = pick
                    # Confidence contract: High means a deterministic rule fired AND
                    # (where applicable) Claude agreed.
                    #   Medium in → a rule DID fire and Claude confirmed it → High
                    #   Low in    → NO rule fired; an unverified pick stops at Medium and
                    #               stays in the review queue. Once the accountant confirms
                    #               it, the payee directory answers it as High next month.
                    new_conf = "High" if old_conf == "Medium" else "Medium"
                    rows[i]["rule"] = (rows[i].get("rule") or "") + " + Claude"
                    if old_conf != new_conf:
                        rows[i]["confidence"] = new_conf
                        summary[old_conf] = summary.get(old_conf, 0) - 1
                        summary[new_conf] = summary.get(new_conf, 0) + 1
                    resolved += 1
            print(f"        → {llm_name} resolved {resolved}/{len(low_idx)} rows "
                  f"via {len(by_key)} distinct narration(s) "
                  f"(unresolved stay Suspense A/c / Medium)")

        # ── Gemini ARBITRATION on generic-rule High rows ───────────────────
        # If a rule marked a row High but its ledger DISAGREES with the best COA
        # match for the narration (e.g. "EKANEK ... RENT" → generic "6. Rent
        # Expenses" while the COA has the vendor "Ekanek Networks Private Limited"),
        # let Gemini arbitrate. Authoritative rules (statutory / contra / learned
        # corrections) are NEVER second-guessed. Gemini confident → keep High with
        # its pick; Gemini abstains → demote to Medium so the accountant reviews it.
        from concurrent.futures import ThreadPoolExecutor as _TPE
        suspense_label = classifier._suspense_ledger
        # 'Side Ledger (credit/debit)' is accountant-authored and side-aware; arbitration
        # cannot know the credit-side/debit-side convention, so left arbitrable it silently
        # rewrote 89 of 148 correct rows on the 2026-06 Urban Plant statement. The
        # 'Side Ledger (fallback)' tier is deliberately NOT listed -- it is a broad
        # catch-all (Salary/Stipend → Salary Payable) that Claude should still refine into
        # named ledgers. Side-rule rows get their own constrained verdict pass instead.
        AUTHORITATIVE = ('Stored Correction', 'Payee Directory', 'Side Ledger (credit/debit)',
                         'Own Account', 'OAT',
                         'Sweep', 'Contra', 'BDP Statutory', 'GSTN', 'TDS', 'EPF',
                         'ESIC', 'PT ', 'NEFT Return', 'Bank Charges', 'Interest')
        def _authoritative(rule):
            rule = str(rule or '')
            return any(rule.startswith(a) for a in AUTHORITATIVE)
        arb_idx = []
        for i, r in enumerate(rows):
            if r["confidence"] != "High" or _authoritative(r.get("rule", "")):
                continue
            entity = r.get("entity", "") or r["description"]
            assigned = r["predicted_ledger"]
            top = classifier.top_candidates(entity, k=1) or classifier.top_candidates(r["description"], k=1)
            if top and top[0] and top[0] != assigned:
                arb_idx.append(i)
        if arb_idx:
            print(f"[3.6/4] {llm_name} arbitration on {len(arb_idx)} generic-rule High rows "
                  f"with a competing COA match …")
            def _arb(i):
                r = rows[i]; desc = r["description"]; entity = r.get("entity", "") or desc
                cands = classifier.llm_candidates(entity, desc, k=15)
                assigned = r["predicted_ledger"]
                if assigned and assigned not in cands:
                    cands = [assigned] + cands
                if suspense_label not in cands:
                    cands = cands + [suspense_label]
                return i, llm_pick(desc, cands)
            corrected = 0; flagged = 0
            try:
                with _TPE(max_workers=6) as ex:
                    for i, pick in ex.map(_arb, arb_idx):
                        if pick and pick != suspense_label:
                            if pick != rows[i]["predicted_ledger"]:
                                rows[i]["predicted_ledger"] = pick
                                corrected += 1
                            # Gemini confident → row stays High
                        else:
                            # Gemini abstained → genuinely ambiguous → flag for review
                            if rows[i]["confidence"] == "High":
                                rows[i]["confidence"] = "Medium"
                                summary["High"] -= 1; summary["Medium"] += 1
                            flagged += 1
            except Exception as _e:
                print(f"        → {llm_name} arbitration aborted ({_e}); rows kept as-is")
            print(f"        → arbitration: {corrected} ledger(s) corrected, "
                  f"{flagged} flagged Medium for review")

    print(f"[4/4] Writing output: {args.output}")
    write_output(rows, summary, args.brand, args.output)
    print(f"\n✅ Done! Output saved to: {args.output}")
    print(f"   Open the file and review any RED (Low confidence) rows manually.")


if __name__ == "__main__":
    main()
