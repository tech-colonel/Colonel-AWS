import os
import re
from io import BytesIO
import pandas as pd

try:
    from thefuzz import fuzz
except ImportError:
    from fuzzywuzzy import fuzz

_OLE2 = b"\xd0\xcf\x11\xe0\xa1\xb1\x1a\xe1"

def _read_any(path):
    """Read .xls/.xlsx into a sheet-name -> DataFrame(header=None) dict, dynamically."""
    with open(path, "rb") as f:
        data = f.read()
    # Try xlrd first for .xls OLE2 format, fallback to openpyxl for xlsx
    try:
        if data[:8] == _OLE2 or path.lower().endswith(".xls"):
            xls = pd.ExcelFile(BytesIO(data), engine="xlrd")
        else:
            xls = pd.ExcelFile(BytesIO(data), engine="openpyxl")
    except Exception:
        # Fallback: try openpyxl
        xls = pd.ExcelFile(BytesIO(data), engine="openpyxl")
    return {s: xls.parse(s, header=None) for s in xls.sheet_names}

def _find_header_row(df, must_have):
    """Return the row index whose cells contain all `must_have` tokens (case-insensitive)."""
    for i in range(min(40, len(df))):
        cells = [str(x).strip().lower() for x in df.iloc[i].tolist()]
        joined = " | ".join(cells)
        if all(tok in joined for tok in must_have):
            return i
    return None

def _num(x):
    if x is None: return 0.0
    s = str(x).replace(",", "").strip()
    if s in ("", "nan", "None"): return 0.0
    try: return float(s)
    except ValueError: return 0.0

def _clean_str(x):
    """Free-text string cleaner: NaN/None/'nan'/'NaT'/'none' all render as empty string.
    (float('nan') is truthy, so `str(x or "")` alone lets it through as the literal "nan".)"""
    if x is None: return ""
    s = str(x).strip()
    return "" if s.lower() in ("nan", "nat", "none") else s

def _colmap(header_cells):
    """Map logical fields to column indices by fuzzy header text."""
    m = {}
    for idx, c in enumerate(header_cells):
        cl = str(c).strip().lower()
        if cl == "date" and "date" not in m: m["date"] = idx
        elif "particular" in cl: m["particulars"] = idx
        elif "narration" in cl: m["narration"] = idx
        elif "vch type" in cl or "voucher type" in cl: m["vch_type"] = idx
        elif "vch no" in cl or "voucher no" in cl: m["vch_no"] = idx
        elif cl == "debit": m["debit"] = idx
        elif cl == "credit": m["credit"] = idx
    return m

def _tally_frame(path):
    sheets = _read_any(path)
    # pick the sheet containing a Date/Particulars/Debit/Credit header
    for name, df in sheets.items():
        h = _find_header_row(df, ["date", "particulars", "debit", "credit"])
        if h is not None:
            return df, h
    raise ValueError("No Tally daybook header row found (Date/Particulars/Debit/Credit)")

def _is_as_per_details(party):
    """True if the ledger cell is Tally's composite-voucher placeholder '(as per details)' --
    the header of a multi-party voucher whose real party breakdown is on following sub-rows."""
    return "as per details" in normalize_party(party)


def parse_tally(path):
    """
    Parse a Tally bank-ledger daybook into flat transaction rows.

    Composite vouchers: when one voucher pays/receives against several parties, Tally prints a
    header row with party = '(as per details)' and the *total* in Debit/Credit, then one sub-row
    per real party where that party's amount sits in the **Vch No. column** (NOT Debit/Credit).
    We explode such vouchers into one matchable row per party (amount pulled from the Vch No.
    cell, direction/date/vch_no inherited from the header) so each party line can reconcile
    against its own bank line. If a '(as per details)' header has no readable sub-rows, we keep
    the lump as a fallback row so no money is lost and the closing-balance totals never drift.
    """
    df, h = _tally_frame(path)
    cm = _colmap(df.iloc[h].tolist())
    rows = []
    pending = None  # active composite-voucher header buffering its sub-rows
    SPLIT_TOL = 0.05  # explode a composite only if its sub-rows sum to the lump (float-noise
    # tolerance only) -- keeps the Tally total byte-identical; rounding/complex vouchers stay lumps

    def _flush_pending():
        # Decide a buffered composite: explode into per-party rows ONLY if the sub-rows sum to
        # the header lump (a provably-clean split, e.g. one payment across several vendors);
        # otherwise keep the lump untouched so the total debit/credit never drifts (complex/
        # mixed vouchers whose sub-rows don't reconcile to the lump, and lumps with no sub-rows).
        if pending is None:
            return
        lump = pending["credit"] or pending["debit"]
        subs = pending["subs"]
        subsum = sum(a for _, a in subs)
        if subs and lump and abs(subsum - lump) <= SPLIT_TOL:
            d = pending["direction"]
            for p_party, amt in subs:
                rows.append({
                    "date": pending["date"], "party": p_party, "narration": pending["narration"],
                    "vch_type": pending["vch_type"], "vch_no": pending["vch_no"],
                    "debit": amt if d == "in" else 0.0,
                    "credit": amt if d == "out" else 0.0,
                    "direction": d, "row": pending["row"],
                })
        elif lump:
            rows.append({
                "date": pending["date"], "party": pending["party"], "narration": pending["narration"],
                "vch_type": pending["vch_type"], "vch_no": pending["vch_no"],
                "debit": pending["debit"], "credit": pending["credit"],
                "direction": "in" if pending["debit"] > 0 else "out", "row": pending["row"],
            })

    for i in range(h + 1, len(df)):
        r = df.iloc[i].tolist()

        def cell(key):
            j = cm.get(key)
            return r[j] if j is not None and j < len(r) else None

        date = cell("date")
        # Tally puts Cr/Dr flag in Particulars col and the ledger name in the next col
        pcol = cm.get("particulars", 1)
        party = ""
        for cand in (pcol + 1, pcol, pcol + 2):
            if cand < len(r) and str(r[cand]).strip() not in ("", "nan", "Cr", "Dr"):
                party = str(r[cand]).strip(); break
        debit = _num(cell("debit"))
        credit = _num(cell("credit"))
        narr = str(cell("narration")).strip() if cell("narration") is not None else ""
        if narr == "nan": narr = ""
        vch_type = str(cell("vch_type")).strip() if cell("vch_type") is not None else ""
        if vch_type.lower() == "nan": vch_type = ""
        vch_no_raw = cell("vch_no")
        vch_no = str(vch_no_raw).strip() if vch_no_raw is not None else ""
        if vch_no.lower() == "nan": vch_no = ""
        parsed_date = pd.to_datetime(date, errors="coerce") if date is not None else None
        party_l = party.lower()

        # A voucher header carries either a real date or a Vch Type; sub-rows carry neither.
        is_header = (not pd.isna(parsed_date)) or bool(vch_type)

        # Skip labeled summary/footer rows (opening/closing balance, grand total) — these are
        # daybook bookkeeping rows, not transactions, even though they contain alphabetic text.
        if "opening balance" in party_l or "closing balance" in party_l or "grand total" in party_l:
            _flush_pending(); pending = None
            continue

        # --- Composite '(as per details)' header: buffer it; decide explode-vs-keep at flush. ---
        if is_header and _is_as_per_details(party) and (debit or credit):
            _flush_pending()
            pending = {"date": parsed_date, "party": party, "narration": narr,
                       "vch_type": vch_type, "vch_no": vch_no,
                       "debit": debit, "credit": credit,
                       "direction": "in" if debit > 0 else "out",
                       "row": i, "subs": []}
            continue

        # --- Sub-row of a pending composite header. Tally writes the sub-amount EITHER in the
        # Vch No. column (Busybees case) OR in the Debit/Credit column (Round Off, Penalty,
        # Discount, and some party lines). Any non-header party row inside a composite block is a
        # breakdown line — absorb it (never let it fall through to become a standalone, date-less
        # row) and buffer its amount for the explode-or-keep-lump decision at flush. ---
        if pending is not None and not is_header and party:
            sub_amt = _num(vch_no_raw) or debit or credit
            if sub_amt > 0:
                pending["subs"].append((party, sub_amt))
            continue

        # Any real voucher header ends the current composite block.
        if is_header:
            _flush_pending(); pending = None

        if debit == 0 and credit == 0: continue
        if not party: continue
        # Skip footer/totals rows: date is NaT/unparseable AND party has no alphabetic chars
        if pd.isna(parsed_date) and not any(c.isalpha() for c in party):
            continue
        rows.append({
            "date": parsed_date,
            "party": party,
            "narration": narr,
            "vch_type": vch_type,
            "vch_no": vch_no,
            "debit": debit, "credit": credit,
            "direction": "in" if debit > 0 else "out",
            "row": i,
        })
    _flush_pending()  # tail composite block at EOF
    return rows

def parse_tally_account(path):
    """
    Return the bank ledger name printed in the Tally daybook's own header block, e.g. the
    daybook prints a row like "RBL Bank - 2588  Book" above the Date/Particulars/Debit/Credit
    transaction-header row -- that IS the account this daybook is for. Scan the rows above the
    header row for one whose text (case-insensitively) ends with "Book", strip the trailing
    "Book" and collapse whitespace. Returns None if no such row is found.
    """
    df, h = _tally_frame(path)
    for i in range(h):
        cells = [str(x).strip() for x in df.iloc[i].tolist() if str(x).strip() not in ("", "nan")]
        for cell in cells:
            if cell.strip().lower().endswith("book"):
                stripped = re.sub(r"(?i)\s*book\s*$", "", cell).strip()
                stripped = re.sub(r"\s+", " ", stripped)
                if stripped:
                    return stripped
    return None


def parse_tally_opening(path):
    df, h = _tally_frame(path)
    cm = _colmap(df.iloc[h].tolist())
    for i in range(h + 1, min(h + 6, len(df))):
        r = df.iloc[i].tolist()
        joined = " ".join(str(x).lower() for x in r)
        if "opening balance" in joined:
            for x in r:
                v = _num(x)
                if v > 0: return v
    return None

def _bank_colmap(header_cells):
    """Map logical fields to column indices by fuzzy header text (Universal Bank output)."""
    m = {}
    for idx, c in enumerate(header_cells):
        cl = str(c).strip().lower().replace(" ", "")
        if "txndate" in cl or cl == "date": m.setdefault("txn_date", idx)
        elif "description" in cl or "narration" in cl: m["description"] = idx
        elif "chq" in cl or "ref" in cl: m["chq_ref"] = idx
        elif cl == "debit": m["debit"] = idx
        elif cl == "credit": m["credit"] = idx
        elif "balance" in cl: m["balance"] = idx
        elif cl == "type": m["type"] = idx
        elif "ledgername" in cl or (cl == "ledger"): m["ledger"] = idx
        elif "confidence" in cl: m["confidence"] = idx
    return m

def parse_bank_output(path):
    """Parse Universal Bank output format: sheet 'Bank Statement', headers include 'Description' and 'Ledger'."""
    sheets = _read_any(path)
    df = None
    for name, d in sheets.items():
        if name.strip().lower() in ("bank statement", "sheet1") or _find_header_row(d, ["description", "ledger"]) is not None:
            df = d; break
    if df is None:
        df = next(iter(sheets.values()))
    h = _find_header_row(df, ["description"]) or 0
    cm = _bank_colmap(df.iloc[h].tolist())
    rows = []
    for i in range(h + 1, len(df)):
        r = df.iloc[i].tolist()
        def g(k):
            j = cm.get(k)
            return r[j] if j is not None and j < len(r) else None
        debit, credit = _num(g("debit")), _num(g("credit"))
        typ = _clean_str(g("type"))
        description = _clean_str(g("description"))
        if debit == 0 and credit == 0 and not description:
            continue
        rows.append({
            "txn_date": pd.to_datetime(g("txn_date"), dayfirst=True, errors="coerce"),
            "description": description,
            "chq_ref": _clean_str(g("chq_ref")),
            "debit": debit, "credit": credit, "balance": _num(g("balance")),
            "type": typ, "ledger": _clean_str(g("ledger")),
            "confidence": _clean_str(g("confidence")),
            "direction": "in" if (credit > 0 or typ.lower() == "receipt") else "out",
            "row": i,
        })
    return rows

# Party normalization for fuzzy matching
_SUFFIXES = [" pvt ltd", " pvt.ltd.", " private limited", " pvt. ltd.", " limited", " llp", " (dr)", " (cr)"]
_LOC_SUFFIX = re.compile(r"[-(]\s*(delhi|telangana|bangalore|blr|hyd|gurgaon|mumbai|vasai|chennai|kolkata|pune|noida|factory)\s*\)?\s*$", re.I)

def normalize_party(name):
    """Normalize party name by removing suffixes, locations, punctuation, and collapsing whitespace."""
    s = str(name or "").lower().strip()
    s = s.replace("\n", " ")
    # Drop the classifier's '⟨add state⟩' marker so a branch-blanked ledger normalizes to its base
    # and matches ANY branch of that vendor (we don't know the branch — the Tally match decides it).
    s = s.replace(ADD_STATE_MARKER.lower(), " ")
    s = _LOC_SUFFIX.sub("", s).strip()
    for suf in _SUFFIXES:
        if s.endswith(suf): s = s[: -len(suf)].strip()
    s = re.sub(r"[^a-z0-9 ]", " ", s)
    s = re.sub(r"\s+", " ", s).strip()
    return s

_PAREN_RE = re.compile(r"\(([^()]*)\)")
# Markers that, if that's ALL a parenthetical contains, mean it's not a second-company alias —
# it's a Dr/Cr flag, a location tag, or a pure number (voucher/cheque no. etc). Skip those.
_PURE_MARKER_RE = re.compile(r"^\s*(dr|cr|cod|delhi|telangana|bangalore|blr|hyd|gurgaon|mumbai|vasai|chennai|kolkata|pune|noida|factory|[0-9,.\s]*)\s*$", re.I)

def _match_keys(name):
    """
    Derive the set of normalized keys under which `name` should be found for matching.
    A bracketed ledger name like "Aurorax Private Limited (BharatX)" is ONE ledger identity —
    we never split it into two parties. We just derive EXTRA keys (full name, name without the
    parenthetical, each parenthetical's inner content) so any of those spellings match.
    """
    keys = set()
    full = normalize_party(name)
    if full:
        keys.add(full)

    raw = str(name or "")
    # name with all parentheticals stripped out
    stripped = normalize_party(_PAREN_RE.sub(" ", raw))
    if stripped:
        keys.add(stripped)

    # each parenthetical's inner content, unless it's a pure Dr/Cr/COD/location/numeric marker
    for inner in _PAREN_RE.findall(raw):
        if _PURE_MARKER_RE.match(inner.strip()):
            continue
        inner_norm = normalize_party(inner)
        if inner_norm:
            keys.add(inner_norm)

    # de-spaced variant of every key so far (so "BharatX" and "Bharat X" both -> "bharatx")
    for k in list(keys):
        despaced = k.replace(" ", "")
        if despaced:
            keys.add(despaced)

    return keys

# Editable alias groups: sets of normalized substrings that denote the SAME ledger identity
# even though the literal names differ (e.g. a courier/marketplace known by multiple trade names
# across Tally vs bank narration). Extend this list as new aliases are discovered on real data.
ALIAS_GROUPS = [
    {"amazon"},
    {"e kart", "ekart", "instakart"},
    {"delhivery"},
    {"flipkart"},
    {"snapmint"},
]

def alias_same(a, b):
    """True if `a` and `b` both contain a substring from the SAME alias group."""
    na, nb = normalize_party(a), normalize_party(b)
    if not na or not nb:
        return False
    for group in ALIAS_GROUPS:
        if any(sub in na for sub in group) and any(sub in nb for sub in group):
            return True
    return False

# Markers that, inside a parenthetical, mean "not a distinguishing branch/location tag" --
# just a Dr/Cr flag, a COD marker, or a pure number (voucher/cheque no. etc). Deliberately NOT
# a fixed city list here (unlike _PURE_MARKER_RE above): any other parenthetical content is
# treated as a distinguishing qualifier, so we don't miss unlisted city/branch names (e.g. the
# full word "hyderabad" was never in the old location list -- only the abbreviation "hyd" was).
_QUALIFIER_MARKER_RE = re.compile(r"^\s*(dr|cr|cod|[0-9,.\s]*)\s*$", re.I)

def _qualifiers(name):
    """
    The set of distinguishing branch/location tags in `name`: each parenthetical's normalized
    inner content that isn't a pure Dr/Cr/COD/numeric marker, plus any trailing location/branch
    suffix that `normalize_party`'s `_LOC_SUFFIX` would strip (e.g. "-Delhi", "-Telangana").
    Two names that BOTH carry qualifiers, but share none in common, are different branches/
    ledgers of the same parent (e.g. "Flo Sleep Solutions (Vasai)" vs "...(Hyderabad)") and must
    never be treated as an exact-string full match just because their un-tagged base collapses
    to the same string.
    """
    raw = str(name or "")
    quals = set()
    for inner in _PAREN_RE.findall(raw):
        inner = inner.strip()
        if not inner or _QUALIFIER_MARKER_RE.match(inner):
            continue
        q = normalize_party(inner)
        if q:
            quals.add(q)
    m = _LOC_SUFFIX.search(raw.lower())
    if m:
        quals.add(m.group(1).strip().lower())
    return quals

def _qualifiers_overlap(qa, qb, threshold=85):
    """True if any qualifier in qa is the same branch/location as any qualifier in qb
    (equal, alias-group same, or fuzzy match >= threshold -- e.g. "hyd" ~ "hyderabad")."""
    for x in qa:
        for y in qb:
            if x == y or alias_same(x, y) or fuzz.token_sort_ratio(x, y) >= threshold:
                return True
    return False

def party_matches(a, b, threshold=85):
    """Check if two party names match within the fuzzy threshold (default 85)."""
    qa, qb = _qualifiers(a), _qualifiers(b)
    if qa and qb and not _qualifiers_overlap(qa, qb, threshold):
        # Both names carry a distinguishing branch/location tag and they don't agree --
        # different branches of the same parent. Never full-match these.
        return False
    na, nb = normalize_party(a), normalize_party(b)
    if not na or not nb: return False
    if na == nb: return True
    if fuzz.token_sort_ratio(na, nb) >= threshold:
        return True
    if alias_same(a, b):
        return True
    # bracket / de-space match keys: any key of a equals or fuzz-matches any key of b
    keys_a, keys_b = _match_keys(a), _match_keys(b)
    for ka in keys_a:
        for kb in keys_b:
            if ka == kb or fuzz.token_sort_ratio(ka, kb) >= threshold:
                return True
    return False

# Marker the Universal classifier writes when it declines to guess a multi-state vendor's branch
# (kept identical to classify.py's ADD_STATE_MARKER). bank_reco resolves it from the matched Tally
# party's branch where the match proves one, otherwise leaves it for a human to complete.
ADD_STATE_MARKER = "⟨add state⟩"


def _resolved_ledger(bank_ledger, tally_party=None):
    """Display ledger for a bank row. If the classifier left a '⟨add state⟩' marker and this row
    fully matched a Tally party that carries a branch/location tag, show that Tally party (the
    branch is now proven and the name exists in the COA). Otherwise keep the ledger as-is (marker
    stays when there's no match, or the matched Tally party is itself branch-less)."""
    led = bank_ledger or ""
    if ADD_STATE_MARKER not in led:
        return led
    if tally_party and _qualifiers(tally_party):
        return tally_party
    return led


def _amt_in(trow):
    """Tally money-in magnitude (debit)."""
    return trow["debit"]

def _amt_out(trow):
    """Tally money-out magnitude (credit)."""
    return trow["credit"]

def _bank_amt(brow):
    """Bank amount based on direction: credit for in, debit for out."""
    return brow["credit"] if brow["direction"] == "in" else brow["debit"]

# Aggressive company-ROOT key for the dense-party aggregate reco: collapse every branch/unit
# variant of one party to a single root so "Flo Sleep Solutions (HYD) Factory", "…(Medchal)",
# "…- ⟨add state⟩" and "FLO SLEEP SOLUTIONS LLP(HIPL)" all bucket together (bank and Tally name
# their branches differently, so only the root is common). Strips the state marker, ALL
# parentheticals (branch/city/code tags), punctuation, and trailing generic unit/company words.
_AGG_DROP_WORDS = {"factory", "llp", "pvt", "ltd", "limited", "private", "unit", "units",
                   "godown", "warehouse", "branch", "division", "dept", "department", "co", "company"}

def _agg_base_key(name):
    s = str(name or "").lower().replace(ADD_STATE_MARKER.lower(), " ")
    s = re.sub(r"\([^)]*\)", " ", s)      # drop all parentheticals (branch/city/code tags)
    s = re.sub(r"[^a-z0-9 ]", " ", s)     # punctuation -> space
    toks = s.split()
    while toks and toks[-1] in _AGG_DROP_WORDS:   # drop trailing generic unit/company words
        toks.pop()
    return " ".join(toks).strip()


# Default salary/payroll keywords (dynamic — extended per brand by the learned aggregate config).
SALARY_KEYWORDS_DEFAULT = ("salary", "wages", "payroll", "stipend")

def _is_salary_ledger(name, keywords=SALARY_KEYWORDS_DEFAULT):
    """True if a ledger is a salary/payroll ledger (keyword match, no hardcoded ledger names).
    Salary is paid to many workers individually in the bank but booked in aggregate in Tally
    with a month/period timing lag, so it can't match line-by-line or per-month like vendors --
    it's reconciled as its own category by monthly total (see the aggregate-reco phase)."""
    n = str(name or "").lower()
    return any(k in n for k in keywords)

def _d(x):
    """Date-only component of a datetime, or None for a missing/unparseable value.
    NOTE: pandas `NaT.date()` returns NaT (not None and not a raise), so an explicit NaN/NaT
    check is required — otherwise callers that test `_d(x) is None` silently mis-handle NaT
    dates (they'd treat an undated row as dated). Guards the same-date/nearest-date matcher."""
    try:
        if pd.isna(x):
            return None
        return x.date()
    except Exception:
        return None

def reconcile(tally_rows, bank_rows, amt_tol=1.0, agg_config=None):
    """
    Reconcile Tally and bank rows using greedy 1:1 matching.

    Two passes:
    1. Full match — same direction + amount within amt_tol + party_matches True. Among all such
       candidates the NEAREST-dated bank row is chosen (same date wins), so a recurring
       same-amount payment (salary/rent) pairs with the correct month instead of whichever bank
       row happens to appear first in the file.
    2. Partial match (soft "review" tier) — over rows still unmatched after pass 1,
       greedily pair same direction + amount within amt_tol + same date (date-only)
       where party_matches is False. These are flagged for human review, not auto-merged.
       Any rows still unmatched after both passes fall to bank_only/tally_only.

    Returns a dict with:
    - "matched": list of full-match pairs with date reconciliation
    - "partial": list of {"bank","tally","reason"} soft-matched pairs (name differs)
    - "bank_only": bank rows with no match at all
    - "tally_only": Tally rows with no match at all
    - "counts": {"matched", "date_updated", "partial", "bank_only", "tally_only"}
    """
    bank_used = [False] * len(bank_rows)

    # Amount index: bucket bank rows by rounded magnitude so each Tally row only compares against
    # the handful of bank rows near its amount. Without this, scanning every bank row per Tally
    # row (to find the nearest date) would run party_matches millions of times on 10k+ rows.
    radius = int(amt_tol) + 1
    bank_by_amt = {}
    for j, b in enumerate(bank_rows):
        bank_by_amt.setdefault(int(round(_bank_amt(b))), []).append(j)

    def _candidates(t_amt, direction):
        """Unused bank indices with the same direction and amount within amt_tol of t_amt."""
        base = int(round(t_amt))
        out = []
        for k in range(base - radius, base + radius + 1):
            for j in bank_by_amt.get(k, ()):
                if bank_used[j]:
                    continue
                b = bank_rows[j]
                if b["direction"] == direction and abs(_bank_amt(b) - t_amt) <= amt_tol:
                    out.append(j)
        return out

    def _pick_nearest(cands, t_date):
        """Choose the candidate whose txn date is nearest t_date (same date = gap 0); unknown
        dates sort last; ties broken by file row order for determinism."""
        best_j, best_key = None, None
        for j in cands:
            bd = _d(bank_rows[j]["txn_date"])
            gap = abs((bd - t_date).days) if (bd and t_date) else 10 ** 9
            key = (gap, bank_rows[j]["row"])
            if best_key is None or key < best_key:
                best_j, best_key = j, key
        return best_j

    # --- Pass 1: full match (direction + amount + party). Two phases so an obvious same-date
    # pair is never stolen by a far-date one competing for the same bank row:
    #   1a. claim EXACT same-date pairs first,
    #   1b. then match each remaining Tally row to its NEAREST available date. ---
    matched, tally_only = [], []
    t_done = [False] * len(tally_rows)

    def _record(t, j):
        bank_used[j] = True
        b = bank_rows[j]
        od, nd = t["date"], b["txn_date"]
        matched.append({"bank": b, "tally": t, "old_date": od, "new_date": nd,
                        "date_changed": bool(_d(od) and _d(nd) and _d(od) != _d(nd))})

    for idx, t in enumerate(tally_rows):          # 1a — exact same-date
        t_date = _d(t["date"])
        if t_date is None:
            continue
        t_amt = _amt_in(t) if t["direction"] == "in" else _amt_out(t)
        for j in _candidates(t_amt, t["direction"]):
            if _d(bank_rows[j]["txn_date"]) == t_date and party_matches(t["party"], bank_rows[j]["ledger"]):
                _record(t, j); t_done[idx] = True
                break

    for idx, t in enumerate(tally_rows):          # 1b — nearest available date
        if t_done[idx]:
            continue
        t_amt = _amt_in(t) if t["direction"] == "in" else _amt_out(t)
        cands = [j for j in _candidates(t_amt, t["direction"])
                 if party_matches(t["party"], bank_rows[j]["ledger"])]
        found = _pick_nearest(cands, _d(t["date"]))
        if found is None:
            tally_only.append(t)
            continue
        _record(t, found)

    # Partial tier: over the still-unmatched rows, greedily pair same direction + amount +
    # same date-only where names differ (party_matches False -- if it matched, pass 1 would
    # have already claimed it). This is a soft "please review" flag, not an auto-merge.
    partial = []
    still_unmatched_tally = []
    for t in tally_only:
        t_amt = _amt_in(t) if t["direction"] == "in" else _amt_out(t)
        t_date = _d(t["date"])
        found = None
        if t_date is not None:
            for j in _candidates(t_amt, t["direction"]):
                if _d(bank_rows[j]["txn_date"]) == t_date:
                    found = j
                    break

        if found is None:
            still_unmatched_tally.append(t)
            continue

        bank_used[found] = True
        b = bank_rows[found]
        partial.append({"bank": b, "tally": t, "reason": "amount+date agree, name differs"})

    # --- Pass 3: AGGREGATE match (N bank lines -> 1 Tally voucher). Accountants often book ONE
    # Tally voucher for several individual bank payments to the SAME party in the SAME month.
    # Over the leftover unmatched rows (salary EXCLUDED -- reconciled as a separate category),
    # bucket by party-base + direction + month and, for each Tally voucher (largest first), find
    # the smallest subset of bank lines that sums to it. Guardrails against coincidental grouping:
    # same party required, capped bucket + group size, ₹amt_tol. ---
    from itertools import combinations as _combos

    def _agg_key(name, direction, mk):
        return (normalize_party(name), direction, mk)

    # Learnable aggregate config (from prior runs, see Phase 2 + the DB learning loop):
    #   salary_keywords: extra keywords that mark a salary/payroll ledger
    #   parties: normalized party-base keys already known to be DENSE aggregate parties (Flo, etc.)
    #            -- recalled so even a single-month file treats them by total, not line matching.
    agg_config = agg_config or {}
    _sal_kw = tuple(set(SALARY_KEYWORDS_DEFAULT) | set(agg_config.get("salary_keywords") or []))
    _learned_parties = set(agg_config.get("parties") or [])
    def _is_sal(name):
        return _is_salary_ledger(name, _sal_kw)

    AGG_BUCKET_CAP = 26   # a party-month with more unmatched lines than this is too dense to
                          # line-partition safely (Flo Sleep Solutions, etc.) -> left for the
                          # category/total reconciliation, not force-grouped here.
    AGG_GROUP_MAX = 6     # a Tally voucher is built from at most this many bank lines

    bank_buckets = {}
    for j, b in enumerate(bank_rows):
        if bank_used[j] or _is_sal(b["ledger"]):
            continue
        mk = _month_key(b["txn_date"])
        if mk is not None:
            bank_buckets.setdefault(_agg_key(b["ledger"], b["direction"], mk), []).append(j)

    tally_buckets = {}
    for t in still_unmatched_tally:
        if _is_sal(t["party"]):
            continue
        mk = _month_key(t["date"])
        t_amt = _amt_in(t) if t["direction"] == "in" else _amt_out(t)
        if mk is not None and t_amt > 0:
            tally_buckets.setdefault(_agg_key(t["party"], t["direction"], mk), []).append((t, t_amt))

    aggregated = []
    agg_tally_ids = set()
    for key, aggs in tally_buckets.items():
        pool = [j for j in bank_buckets.get(key, []) if not bank_used[j]]
        if len(pool) < 2 or len(pool) > AGG_BUCKET_CAP:
            continue
        for t, t_amt in sorted(aggs, key=lambda x: -x[1]):   # biggest voucher claims its lines first
            avail = [(j, _bank_amt(bank_rows[j])) for j in pool if not bank_used[j]]
            if len(avail) < 2:
                break
            avail.sort(key=lambda ja: ja[1])   # ascending -> first exact match is the smallest group
            found = None
            for rsz in range(2, min(len(avail), AGG_GROUP_MAX) + 1):
                for combo in _combos(range(len(avail)), rsz):
                    if abs(sum(avail[i][1] for i in combo) - t_amt) <= amt_tol:
                        found = [avail[i][0] for i in combo]
                        break
                if found:
                    break
            if found:
                for j in found:
                    bank_used[j] = True
                aggregated.append({"tally": t, "banks": [bank_rows[j] for j in found]})
                agg_tally_ids.add(id(t))
    still_unmatched_tally = [t for t in still_unmatched_tally if id(t) not in agg_tally_ids]

    # --- Phase 2: CATEGORY / TOTAL reconciliation for DENSE aggregate parties (salary + Flo-like
    # parties). Some parties have so many individual bank lines booked as few Tally aggregates,
    # with a month/period timing lag, that safe line-matching is impossible. For these we reconcile
    # by MONTHLY TOTAL: per party-month, sum bank vs sum Tally and show the gap, collapsing hundreds
    # of noise lines into a summary. Dynamic + learnable: salary is keyword-detected; other parties
    # are detected THIS run by volume (>= DENSE_MIN unmatched bank lines) AND recalled from
    # `agg_config['parties']` learned in prior runs (so a single-month file still catches Flo/etc.).
    from collections import defaultdict as _dd
    DENSE_MIN = 30       # this-run volume threshold to flag a (non-salary) party as a dense aggregate
    SALARY_KEY = "__salary__"   # ALL payroll ledgers reconcile as ONE category (sub-ledgers like
                                # 'Salary Payable' vs 'Salary Payable -Staff' must not split apart)

    def _party_key(name):
        # Salary -> one category; every other party -> its company ROOT so all branch/unit variants
        # (bank's "(HYD) Factory" vs Tally's "(Medchal)") fold into one dense-party bucket.
        return SALARY_KEY if _is_sal(name) else _agg_base_key(name)

    unbank_by_party = _dd(list)
    for j, b in enumerate(bank_rows):
        if not bank_used[j] and _month_key(b["txn_date"]) is not None:
            unbank_by_party[_party_key(b["ledger"])].append(j)
    untally_by_party = _dd(list)
    for t in still_unmatched_tally:
        if _month_key(t["date"]) is not None:
            untally_by_party[_party_key(t["party"])].append(t)

    # Decide which party-bases are aggregate parties: salary (one bucket) OR dense-this-run OR learned.
    agg_bases = set(); detected_parties = set()
    for pb, js in unbank_by_party.items():
        is_sal = pb == SALARY_KEY
        if is_sal or pb in _learned_parties or (len(js) >= DENSE_MIN and untally_by_party.get(pb)):
            # Only reconcile parties present on BOTH sides (else it's a genuine unbooked gap).
            if is_sal or untally_by_party.get(pb):
                agg_bases.add(pb)
                if not is_sal:
                    detected_parties.add(pb)   # learn dense non-salary parties (Flo, etc.)

    # For a dense/salary party present on BOTH sides, reconcile over the whole PERIOD: fold ALL its
    # unmatched bank + Tally lines (a party-month with only one side is timing lag, not a per-month
    # gap -- summed across months the monthly gaps net to the TRUE period gap). One summary row per
    # (party, month) shows that month's bank vs Tally; the period gap = sum of the monthly gaps.
    aggregate_reco = []
    agg_recon_tally_ids = set()
    for pb in agg_bases:
        is_sal = pb == SALARY_KEY
        bmonth = _dd(list); tmonth = _dd(list)
        for j in unbank_by_party.get(pb, []):
            if not bank_used[j]:
                bmonth[_month_key(bank_rows[j]["txn_date"])].append(j)
        for t in untally_by_party.get(pb, []):
            tmonth[_month_key(t["date"])].append(t)
        if is_sal:
            label = "Salary (all payroll ledgers)"
        else:
            label = next((t["party"] for t in untally_by_party.get(pb, [])), None) \
                or next((bank_rows[j]["ledger"] for j in unbank_by_party.get(pb, [])), pb)
        for mk in sorted(m for m in (set(bmonth) | set(tmonth)) if m):
            js = bmonth.get(mk, []); ts = tmonth.get(mk, [])
            bt = sum(_bank_amt(bank_rows[j]) for j in js)
            tt = sum((_amt_in(t) if t["direction"] == "in" else _amt_out(t)) for t in ts)
            for j in js:
                bank_used[j] = True
            for t in ts:
                agg_recon_tally_ids.add(id(t))
            aggregate_reco.append({
                "party": label, "party_base": pb, "month": mk,
                "kind": "salary" if is_sal else "party",
                "bank_total": bt, "tally_total": tt, "gap": bt - tt,
                "bank_lines": len(js), "tally_vouchers": len(ts),
                "banks": [bank_rows[j] for j in js], "tallys": ts,
            })
    still_unmatched_tally = [t for t in still_unmatched_tally if id(t) not in agg_recon_tally_ids]

    tally_only = still_unmatched_tally
    bank_only = [b for j, b in enumerate(bank_rows) if not bank_used[j]]
    counts = {
        "matched": len(matched),
        "date_updated": sum(1 for m in matched if m["date_changed"]),
        "partial": len(partial),
        "aggregated": sum(len(a["banks"]) for a in aggregated),
        "aggregated_vouchers": len(aggregated),
        "aggregate_reco": sum(a["bank_lines"] for a in aggregate_reco),
        "aggregate_reco_groups": len(aggregate_reco),
        "bank_only": len(bank_only),
        "tally_only": len(tally_only),
    }
    return {"matched": matched, "partial": partial, "aggregated": aggregated,
            "aggregate_reco": aggregate_reco, "detected_aggregate_parties": sorted(detected_parties),
            "bank_only": bank_only, "tally_only": tally_only, "counts": counts}


# ---------------------------------------------------------------------------
# Styled 9-sheet workbook writer + CLI
# ---------------------------------------------------------------------------

from openpyxl import Workbook, load_workbook
from openpyxl.styles import PatternFill, Font, Border, Side, Alignment
from openpyxl.utils import get_column_letter

GREEN  = PatternFill(start_color="C8E6C9", end_color="C8E6C9", fill_type="solid")
YELLOW = PatternFill(start_color="FFF9C4", end_color="FFF9C4", fill_type="solid")
ORANGE = PatternFill(start_color="FFE0B2", end_color="FFE0B2", fill_type="solid")   # partial match
RED    = PatternFill(start_color="FFCDD2", end_color="FFCDD2", fill_type="solid")
BLUE   = PatternFill(start_color="BBDEFB", end_color="BBDEFB", fill_type="solid")   # aggregated in Tally
GREY   = PatternFill(start_color="ECEFF1", end_color="ECEFF1", fill_type="solid")
HEADER = PatternFill(start_color="263238", end_color="263238", fill_type="solid")
WHITE_BOLD = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
THIN = Border(*(Side(style='thin', color='CCCCCC'),)*4)
MONEY = '#,##0.00'


def _sheet(wb, title, headers):
    ws = wb.create_sheet(title)
    for c, h in enumerate(headers, 1):
        cell = ws.cell(row=1, column=c, value=h)
        cell.fill = HEADER; cell.font = WHITE_BOLD; cell.border = THIN
        cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.freeze_panes = "A2"
    return ws


def _autowidth(ws):
    for col in ws.columns:
        width = max((len(str(c.value)) for c in col if c.value is not None), default=10)
        ws.column_dimensions[get_column_letter(col[0].column)].width = min(max(width + 2, 10), 48)


def _fmt_date(x):
    """NaT/None-safe dd-mm-yyyy formatter shared by every sheet."""
    if x is None:
        return ""
    try:
        if pd.isna(x):
            return ""
    except (TypeError, ValueError):
        pass
    try:
        return x.strftime("%d-%m-%Y")
    except Exception:
        return ""


def _month_key(x):
    """NaT/None-safe YYYY-MM month bucket key."""
    if x is None:
        return None
    try:
        if pd.isna(x):
            return None
    except (TypeError, ValueError):
        pass
    try:
        return x.strftime("%Y-%m")
    except Exception:
        return None


def _fill_row(ws, row, ncols, fill):
    for c in range(1, ncols + 1):
        cell = ws.cell(row=row, column=c)
        cell.fill = fill
        cell.border = THIN


def _write_summary(wb, result, ctx):
    """Sheet 1: Summary — two-column key/value."""
    ws = _sheet(wb, "Summary", ["Field", "Value"])
    bank_rows = ctx["bank_rows"]
    dates = [b["txn_date"] for b in bank_rows if _month_key(b["txn_date"]) is not None]
    period = ("%s to %s" % (_fmt_date(min(dates)), _fmt_date(max(dates)))) if dates else ""
    total_in = sum(b["credit"] or 0 for b in bank_rows)
    total_out = sum(b["debit"] or 0 for b in bank_rows)
    c = result["counts"]
    reconciled = c["bank_only"] == 0 and c["tally_only"] == 0

    money_fields = {"Total money-in", "Total money-out"}
    rows = [
        ("Brand", ctx.get("brand", "")),
        ("Bank", os.path.basename(ctx.get("bank_path") or "")),
        ("Period", period),
        ("Matched", c["matched"]),
        ("Date-updated", c["date_updated"]),
        ("Aggregated in Tally", "%d lines → %d vouchers" % (c.get("aggregated", 0), c.get("aggregated_vouchers", 0))),
        ("Aggregate reconciled (salary/dense)", "%d lines → %d party-months" % (c.get("aggregate_reco", 0), c.get("aggregate_reco_groups", 0))),
        ("Partially matched", c["partial"]),
        ("Bank-only", c["bank_only"]),
        ("Tally-only", c["tally_only"]),
        ("Total money-in", total_in),
        ("Total money-out", total_out),
        ("Reconciled?", "Yes" if reconciled else "No"),
    ]
    for field, value in rows:
        ws.append([field, value])
        rr = ws.max_row
        if field in money_fields:
            ws.cell(row=rr, column=2).number_format = MONEY
        if field == "Reconciled?":
            _fill_row(ws, rr, 2, GREEN if reconciled else RED)
        else:
            ws.cell(row=rr, column=1).border = THIN
            ws.cell(row=rr, column=2).border = THIN
    _autowidth(ws)


def _write_date_updates(wb, result):
    """Sheet 3: Date Updates (old -> new) — matched rows where date_changed."""
    ws = _sheet(wb, "Date Updates", ["Party", "Amount", "Vch No.", "Old Date", "New Date"])
    for m in result["matched"]:
        if not m["date_changed"]:
            continue
        t = m["tally"]
        amt = t["debit"] if t["direction"] == "in" else t["credit"]
        vals = [t["party"], amt, t["vch_no"], _fmt_date(m["old_date"]), _fmt_date(m["new_date"])]
        ws.append(vals)
        rr = ws.max_row
        ws.cell(row=rr, column=2).number_format = MONEY
        _fill_row(ws, rr, len(vals), YELLOW)
    _autowidth(ws)


def _write_ready_to_paste(wb, result, ctx):
    """Sheet 4: Add to Tally (Ready-to-Paste) — one row per bank-only row."""
    ws = _sheet(wb, "Add to Tally", ["Date", "Dr Ledger", "Cr Ledger", "Amount", "Narration"])
    bank_ledger = ctx.get("bank_ledger") or "Bank Account"
    for b in result["bank_only"]:
        amt = b["debit"] if b["direction"] == "out" else b["credit"]
        ledger = b["ledger"] or "Suspense"
        if b["direction"] == "out":
            dr, cr = ledger, bank_ledger
        else:
            dr, cr = bank_ledger, ledger
        vals = [_fmt_date(b["txn_date"]), dr, cr, amt, b["description"]]
        ws.append(vals)
        rr = ws.max_row
        ws.cell(row=rr, column=4).number_format = MONEY
        _fill_row(ws, rr, len(vals), RED)
    _autowidth(ws)


def _write_tally_only(wb, result):
    """Sheet 5: Tally-only (Check) — tally rows with no bank match at all."""
    ws = _sheet(wb, "Tally-only", ["Date", "Party", "Debit", "Credit", "Vch No.", "Remark"])
    remark = "In Tally, not in bank — check duplicate / wrong-brand entry"
    for t in result["tally_only"]:
        vals = [_fmt_date(t["date"]), t["party"], t["debit"] or "", t["credit"] or "", t["vch_no"], remark]
        ws.append(vals)
        rr = ws.max_row
        for c in (3, 4):
            ws.cell(row=rr, column=c).number_format = MONEY
        _fill_row(ws, rr, len(vals), GREY)
    _autowidth(ws)


def _write_aggregate_reco(wb, result):
    """Sheet 6: Aggregate Reco — dense/salary parties reconciled by monthly total (not line-by-line
    -- the bank shows many individual payments that Tally books as few aggregate vouchers). One row
    per (party, month): bank total, Tally total, and the gap; a green period subtotal per party.
    The gap is the real 'needs review / booking' amount -- everything else is reconciled."""
    reco = result.get("aggregate_reco", [])
    if not reco:
        return
    ws = _sheet(wb, "Aggregate Reco",
                ["Party / Category", "Month", "Bank total", "Tally total", "Gap (Bank − Tally)",
                 "Bank lines", "Tally vouchers"])
    # group rows by party, print monthly rows then a period subtotal
    from collections import OrderedDict
    by_party = OrderedDict()
    for a in reco:
        by_party.setdefault(a["party"], []).append(a)
    for party, rowsp in by_party.items():
        rowsp.sort(key=lambda a: a["month"])
        b_sum = t_sum = 0.0; bl = tv = 0
        for a in rowsp:
            ws.append([a["party"], a["month"], a["bank_total"], a["tally_total"], a["gap"],
                       a["bank_lines"], a["tally_vouchers"]])
            rr = ws.max_row
            for c in (3, 4, 5):
                ws.cell(row=rr, column=c).number_format = MONEY
            _fill_row(ws, rr, 7, BLUE if abs(a["gap"]) <= 1 else YELLOW)
            b_sum += a["bank_total"]; t_sum += a["tally_total"]; bl += a["bank_lines"]; tv += a["tally_vouchers"]
        ws.append(["%s — PERIOD TOTAL" % party, "", b_sum, t_sum, b_sum - t_sum, bl, tv])
        rr = ws.max_row
        for c in (3, 4, 5):
            ws.cell(row=rr, column=c).number_format = MONEY
        _fill_row(ws, rr, 7, GREEN if abs(b_sum - t_sum) <= max(1.0, 0.02 * max(b_sum, t_sum)) else RED)
        for c in range(1, 8):
            ws.cell(row=rr, column=c).font = Font(bold=True)
    _autowidth(ws)


def _write_pivot(wb, ctx):
    """Sheet 6: Pivot — bank rows grouped by month, summed debit/credit."""
    ws = _sheet(wb, "Pivot", ["Month", "Withdrawal", "Deposit"])
    agg = {}
    for b in ctx["bank_rows"]:
        mk = _month_key(b["txn_date"])
        if mk is None:
            continue
        a = agg.setdefault(mk, {"debit": 0.0, "credit": 0.0})
        a["debit"] += b["debit"] or 0
        a["credit"] += b["credit"] or 0
    for mk in sorted(agg.keys()):
        a = agg[mk]
        ws.append([mk, a["debit"], a["credit"]])
        rr = ws.max_row
        ws.cell(row=rr, column=2).number_format = MONEY
        ws.cell(row=rr, column=3).number_format = MONEY
        for c in range(1, 4):
            ws.cell(row=rr, column=c).border = THIN
    _autowidth(ws)


def _write_query(wb, ctx):
    """Sheet 7: Query — bank rows grouped by (ledger, month), dynamic, no hardcoded vendors."""
    ws = _sheet(wb, "Query", ["Ledger", "Month", "Debit", "Credit"])
    agg = {}
    for b in ctx["bank_rows"]:
        mk = _month_key(b["txn_date"]) or ""
        key = (b["ledger"] or "", mk)
        a = agg.setdefault(key, {"debit": 0.0, "credit": 0.0})
        a["debit"] += b["debit"] or 0
        a["credit"] += b["credit"] or 0
    for key in sorted(agg.keys()):
        ledger, mk = key
        a = agg[key]
        ws.append([ledger, mk, a["debit"], a["credit"]])
        rr = ws.max_row
        ws.cell(row=rr, column=3).number_format = MONEY
        ws.cell(row=rr, column=4).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=rr, column=c).border = THIN
    _autowidth(ws)


def _closing_series(ctx):
    """Shared aggregation behind BOTH the 'Bank vs Tally' sheet AND analytics.monthly:
    Tally closing = opening + cumulative(debit-credit) by month; Bank closing = last
    balance seen in that month (chronological within the month, not file order).
    Returns a chronological list of (month_key, tally_closing, bank_closing_or_None).
    """
    tally_rows = ctx["tally_rows"]
    bank_rows = ctx["bank_rows"]
    opening = ctx.get("opening") or 0.0

    t_by_month = {}
    for t in tally_rows:
        mk = _month_key(t["date"])
        if mk is None:
            continue
        a = t_by_month.setdefault(mk, {"debit": 0.0, "credit": 0.0})
        a["debit"] += t["debit"] or 0
        a["credit"] += t["credit"] or 0

    # Last balance seen per month, walking bank rows in chronological (date, row) order so
    # "last" means last-by-date-in-month, not last-in-file-order.
    b_by_month_last = {}
    bank_sorted = sorted(
        (b for b in bank_rows if _month_key(b["txn_date"]) is not None),
        key=lambda b: (b["txn_date"], b["row"]),
    )
    for b in bank_sorted:
        b_by_month_last[_month_key(b["txn_date"])] = b["balance"]

    months = sorted(set(t_by_month.keys()) | set(b_by_month_last.keys()))
    out = []
    cum = opening
    for mk in months:
        a = t_by_month.get(mk, {"debit": 0.0, "credit": 0.0})
        cum += (a["debit"] - a["credit"])
        bank_closing = b_by_month_last.get(mk)  # None if no bank txns that month
        out.append((mk, cum, bank_closing))
    return out


def _write_closing(wb, ctx):
    """Sheet 8: Bank vs Tally (Closing) — DIFF = Tally - Bank. Values come from
    `_closing_series`, shared with analytics.monthly so the sheet and the dashboard
    chart can never disagree."""
    ws = _sheet(wb, "Bank vs Tally", ["Month", "As per Tally", "As per Bank", "DIFF"])
    for mk, cum, bank_closing in _closing_series(ctx):
        bank_val = bank_closing if isinstance(bank_closing, (int, float)) else ""
        diff = (cum - bank_closing) if isinstance(bank_closing, (int, float)) else ""
        ws.append([mk, cum, bank_val, diff])
        rr = ws.max_row
        for c in (2, 3, 4):
            ws.cell(row=rr, column=c).number_format = MONEY
        for c in range(1, 5):
            ws.cell(row=rr, column=c).border = THIN
    _autowidth(ws)


def _top_ledgers(ctx, n=10):
    """Top `n` bank-side ledgers by absolute total amount (sum of debit+credit per
    `ledger_name`) — same `ctx["bank_rows"]` source the Query/Pivot sheets aggregate,
    just grouped by ledger alone instead of (ledger, month)."""
    agg = {}
    for b in ctx["bank_rows"]:
        name = (b.get("ledger") or "").strip()
        if not name:
            continue
        agg[name] = agg.get(name, 0.0) + (b["debit"] or 0) + (b["credit"] or 0)
    top = sorted(agg.items(), key=lambda kv: abs(kv[1]), reverse=True)[:n]
    return [{"ledger": name, "amount": amount} for name, amount in top]


def _build_analytics(result, ctx):
    """Analytics for the dashboard's charts. Reuses `result["counts"]` (already computed
    by `reconcile`) and the closing-balance / ledger aggregations shared with the
    workbook's 'Bank vs Tally (Closing)' sheet — never re-runs the reconciliation itself.
    """
    counts = result["counts"]
    monthly = []
    for mk, tally_closing, bank_closing in _closing_series(ctx):
        if isinstance(bank_closing, (int, float)):
            bc = float(bank_closing)
            diff = float(tally_closing) - bc
        else:
            bc = None
            diff = None
        monthly.append({
            "month": mk,
            "tally_closing": float(tally_closing),
            "bank_closing": bc,
            "diff": diff,
        })
    buckets = [
        {"key": "matched", "label": "Matched", "count": counts["matched"]},
        {"key": "date_updated", "label": "Date-updated", "count": counts["date_updated"]},
        {"key": "aggregated", "label": "Aggregated in Tally", "count": counts.get("aggregated", 0)},
        {"key": "partial", "label": "Partially matched", "count": counts["partial"]},
        {"key": "bank_only", "label": "Bank-only", "count": counts["bank_only"]},
        {"key": "tally_only", "label": "Tally-only", "count": counts["tally_only"]},
    ]
    return {
        "monthly": monthly,
        "top_ledgers": _top_ledgers(ctx),
        "buckets": buckets,
    }


def _write_universal(wb, ctx):
    """Sheet 9: Universal Output — verbatim copy of the raw Universal 'Bank Statement' sheet."""
    ws = wb.create_sheet("Universal Output")
    src_wb = load_workbook(ctx["bank_path"], data_only=True)
    src_ws = None
    for name in src_wb.sheetnames:
        if name.strip().lower() == "bank statement":
            src_ws = src_wb[name]
            break
    if src_ws is None:
        src_ws = src_wb[src_wb.sheetnames[0]]
    for row in src_ws.iter_rows():
        ws.append([c.value for c in row])
    if ws.max_row >= 1:
        ws.freeze_panes = "A2"
    _autowidth(ws)


def write_workbook(result, ctx, output_path):
    wb = Workbook(); wb.remove(wb.active)
    # --- Sheet 2: Reconciliation (main) ---
    ws = _sheet(wb, "Reconciliation", ["Txn Date","Description","Chq / Ref No.","Debit","Credit",
        "Balance","Type","Ledger Name","Reco Status","Matched Tally Party","Tally Date","Date Flag"])
    matched_by_bankrow = {id(m["bank"]): m for m in result["matched"]}
    partial_by_bankrow = {id(p["bank"]): p for p in result.get("partial", [])}
    # aggregated: each bank line -> (its Tally voucher, how many bank lines fed that voucher)
    agg_by_bankrow = {}
    for a in result.get("aggregated", []):
        for bk in a["banks"]:
            agg_by_bankrow[id(bk)] = (a["tally"], len(a["banks"]))
    # aggregate-reco (dense/salary category): each bank line -> its (party label, month)
    aggreco_by_bankrow = {}
    for a in result.get("aggregate_reco", []):
        for bk in a["banks"]:
            aggreco_by_bankrow[id(bk)] = a
    for b in ctx["bank_rows"]:
        m = matched_by_bankrow.get(id(b))
        p = partial_by_bankrow.get(id(b))
        ag = agg_by_bankrow.get(id(b))
        ar = aggreco_by_bankrow.get(id(b))
        if m:
            status = "Already in Tally"
            party = m["tally"]["party"]
            tdate = _fmt_date(m["old_date"])
            flag = ("Date updated → %s" % _fmt_date(m["new_date"])) if m["date_changed"] else "OK"
            fill = YELLOW if m["date_changed"] else GREEN
            disp_ledger = _resolved_ledger(b["ledger"], party)  # fill branch back from Tally
        elif ag:
            tv, n = ag
            status = "Aggregated in Tally"
            party = tv["party"]
            tdate = _fmt_date(tv["date"])
            flag = "%d bank lines → 1 voucher (vch %s)" % (n, tv.get("vch_no") or "")
            fill = BLUE
            disp_ledger = _resolved_ledger(b["ledger"], party)
        elif ar:
            status = "Aggregate reconciled"
            party = ar["party"]
            tdate = ar["month"]
            flag = ("Salary — reconciled in aggregate" if ar["kind"] == "salary"
                    else "Booked in aggregate under %s" % ar["party"])
            fill = BLUE
            disp_ledger = b["ledger"]
        elif p:
            status = "Partially matched — verify name"
            party = p["tally"]["party"]
            tdate = _fmt_date(p["tally"]["date"])
            flag = "Amount+date agree, name differs"
            fill = ORANGE
            disp_ledger = b["ledger"]  # name differs on a partial — don't borrow the branch
        else:
            status, party, tdate, flag, fill = "In bank, not in Tally — add", "", "", "", RED
            disp_ledger = b["ledger"]  # unmatched — keep the marker for a human
        vals = [_fmt_date(b["txn_date"]),
                b["description"], b["chq_ref"], b["debit"] or "", b["credit"] or "",
                b["balance"] or "", b["type"], disp_ledger, status, party, tdate, flag]
        ws.append(vals)
        rr = ws.max_row
        for c in (4,5,6):
            ws.cell(row=rr, column=c).number_format = MONEY
        for c in range(1, len(vals)+1):
            ws.cell(row=rr, column=c).fill = fill; ws.cell(row=rr, column=c).border = THIN
    _autowidth(ws)
    # --- Sheets 1,3,4,5,6,7,8,9 ---
    _write_summary(wb, result, ctx)          # sheet 1
    _write_date_updates(wb, result)          # sheet 3
    _write_ready_to_paste(wb, result, ctx)   # sheet 4
    _write_tally_only(wb, result)            # sheet 5
    _write_aggregate_reco(wb, result)        # sheet 6 (only if any)
    _write_pivot(wb, ctx)                    # sheet 7
    _write_query(wb, ctx)                    # sheet 8
    _write_closing(wb, ctx)                  # sheet 9
    _write_universal(wb, ctx)                # sheet 10
    # reorder tabs to spec order
    order = ["Summary","Reconciliation","Date Updates","Add to Tally","Tally-only","Aggregate Reco","Pivot","Query","Bank vs Tally","Universal Output"]
    wb._sheets.sort(key=lambda s: order.index(s.title) if s.title in order else 99)
    wb.save(output_path)


def _fair_caps(sizes, total):
    """Water-filling allocation: split `total` slots across buckets in `sizes`
    (name -> count) so every non-empty bucket gets a share, redistributing any
    leftover from buckets smaller than their share to the buckets that still
    want more. Returns {name: cap}, summing to min(total, sum(sizes.values())).
    """
    remaining = dict(sizes)
    caps = {k: 0 for k in sizes}
    budget = total
    active = [k for k, v in sizes.items() if v > 0]
    while budget > 0 and active:
        share = max(1, budget // len(active))
        progressed = False
        for k in list(active):
            take = min(share, remaining[k], budget)
            if take <= 0:
                continue
            caps[k] += take
            remaining[k] -= take
            budget -= take
            progressed = True
            if remaining[k] <= 0:
                active.remove(k)
            if budget <= 0:
                break
        if not progressed:
            break
    return caps


def _build_results(result, max_rows=1500):
    """Flatten all reco buckets into a list of flat row dicts for the UI results table.
    Capped at `max_rows` total (the UI only shows the first ~200 anyway), fairly
    split across buckets via `_fair_caps` so a huge "matched" bucket never crowds
    out partial/bank_only/tally_only rows entirely. Keys: txn_date, description,
    chq_ref, debit, credit, balance, type, ledger_name, reco_status, tally_party,
    date_flag -- dates dd-mm-yyyy, money as numbers ("" if blank).
    """
    rows = []

    def _money(v):
        return v if v else ""

    partial = result.get("partial", [])
    aggregated = result.get("aggregated", [])
    aggregate_reco = result.get("aggregate_reco", [])
    caps = _fair_caps({
        "matched": len(result["matched"]),
        "partial": len(partial),
        "aggregated": sum(len(a["banks"]) for a in aggregated),
        "aggregate_reco": sum(a["bank_lines"] for a in aggregate_reco),
        "bank_only": len(result["bank_only"]),
        "tally_only": len(result["tally_only"]),
    }, max_rows)

    for m in result["matched"][:caps["matched"]]:
        b = m["bank"]
        date_changed = m["date_changed"]
        rows.append({
            "txn_date": _fmt_date(b["txn_date"]),
            "description": b["description"],
            "chq_ref": b["chq_ref"],
            "debit": _money(b["debit"]),
            "credit": _money(b["credit"]),
            "balance": _money(b["balance"]),
            "type": b["type"],
            "ledger_name": _resolved_ledger(b["ledger"], m["tally"]["party"]),
            "reco_status": "Date updated" if date_changed else "Already in Tally",
            "tally_party": m["tally"]["party"],
            "date_flag": ("Date updated → %s" % _fmt_date(m["new_date"])) if date_changed else "OK",
        })

    for p in partial[:caps["partial"]]:
        b = p["bank"]
        rows.append({
            "txn_date": _fmt_date(b["txn_date"]),
            "description": b["description"],
            "chq_ref": b["chq_ref"],
            "debit": _money(b["debit"]),
            "credit": _money(b["credit"]),
            "balance": _money(b["balance"]),
            "type": b["type"],
            "ledger_name": b["ledger"],
            "reco_status": "Partially matched",
            "tally_party": p["tally"]["party"],
            "date_flag": "amount+date agree, name differs",
        })

    agg_flat = [(bk, a["tally"], len(a["banks"])) for a in aggregated for bk in a["banks"]]
    for b, tv, n in agg_flat[:caps["aggregated"]]:
        rows.append({
            "txn_date": _fmt_date(b["txn_date"]),
            "description": b["description"],
            "chq_ref": b["chq_ref"],
            "debit": _money(b["debit"]),
            "credit": _money(b["credit"]),
            "balance": _money(b["balance"]),
            "type": b["type"],
            "ledger_name": _resolved_ledger(b["ledger"], tv["party"]),
            "reco_status": "Aggregated in Tally",
            "tally_party": tv["party"],
            "date_flag": "%d bank lines → 1 voucher" % n,
        })

    aggreco_flat = [(bk, a) for a in aggregate_reco for bk in a["banks"]]
    for b, a in aggreco_flat[:caps["aggregate_reco"]]:
        rows.append({
            "txn_date": _fmt_date(b["txn_date"]),
            "description": b["description"],
            "chq_ref": b["chq_ref"],
            "debit": _money(b["debit"]),
            "credit": _money(b["credit"]),
            "balance": _money(b["balance"]),
            "type": b["type"],
            "ledger_name": b["ledger"],
            "reco_status": "Aggregate reconciled",
            "tally_party": a["party"],
            "date_flag": ("Salary — reconciled in aggregate" if a["kind"] == "salary"
                          else "Booked in aggregate (%s)" % a["month"]),
        })

    for b in result["bank_only"][:caps["bank_only"]]:
        rows.append({
            "txn_date": _fmt_date(b["txn_date"]),
            "description": b["description"],
            "chq_ref": b["chq_ref"],
            "debit": _money(b["debit"]),
            "credit": _money(b["credit"]),
            "balance": _money(b["balance"]),
            "type": b["type"],
            "ledger_name": b["ledger"],
            "reco_status": "Bank-only — add",
            "tally_party": "",
            "date_flag": "",
        })

    for t in result["tally_only"][:caps["tally_only"]]:
        rows.append({
            "txn_date": _fmt_date(t["date"]),
            "description": t["party"],
            "chq_ref": "",
            "debit": _money(t["debit"]),
            "credit": _money(t["credit"]),
            "balance": "",
            "type": "",
            "ledger_name": t["party"],
            "reco_status": "Tally-only — check",
            "tally_party": t["party"],
            "date_flag": "",
        })

    return rows


def _llm_confirm_aggregate_parties(candidates, api_key, model, base_url, timeout=30):
    """LLM gate (GenSpark/Haiku) before LEARNING a detected dense party: confirm it is genuinely an
    aggregate/batch-booked ledger (many individual bank payments booked as fewer Tally vouchers),
    so a party that is merely dense-by-coincidence is never persisted per brand. `candidates`:
    [{party, party_base, bank_lines, tally_vouchers, bank_total, tally_total}]. Returns the list of
    confirmed party_base keys, or None on missing key / any error (caller then keeps the
    deterministic detected set unchanged). One small call per run — not per row."""
    if not api_key or not candidates:
        return None
    import urllib.request, json as _json, ssl, re as _re
    rows = "\n".join(
        'key="%s" name="%s" bank_lines=%d tally_vouchers=%d bank_total=%d tally_total=%d'
        % (c["party_base"], c["party"], c["bank_lines"], c["tally_vouchers"],
           round(c["bank_total"]), round(c["tally_total"]))
        for c in candidates)
    ask = (
        "You reconcile Indian bank statements against Tally books. A DENSE AGGREGATE PARTY is a "
        "ledger where MANY individual bank payments are booked in Tally as FEWER aggregate vouchers "
        "(salary batches, or a vendor/branch account the accountant lumps into periodic entries) — "
        "its bank and Tally totals are roughly comparable and it has many bank lines vs few Tally "
        "vouchers. For each candidate, decide if it is genuinely such a party. Reply with ONLY a "
        "compact JSON object mapping each key to true or false.\n\n" + rows)
    try:
        if base_url:
            url = "%s/chat/completions" % base_url.rstrip("/")
            headers = {"content-type": "application/json", "authorization": "Bearer %s" % api_key}
        else:
            url = "https://api.anthropic.com/v1/messages"
            headers = {"content-type": "application/json", "x-api-key": api_key,
                       "anthropic-version": "2023-06-01"}
        body = _json.dumps({"model": model, "max_tokens": 400,
                            "messages": [{"role": "user", "content": ask}]}).encode("utf-8")
        req = urllib.request.Request(url, data=body, headers=headers)
        with urllib.request.urlopen(req, timeout=timeout, context=ssl.create_default_context()) as resp:
            data = _json.loads(resp.read().decode("utf-8"))
        if base_url:
            text = ((data.get("choices") or [{}])[0].get("message") or {}).get("content") or ""
        else:
            text = next((b.get("text", "") for b in data.get("content", []) if b.get("type") == "text"), "")
        m = _re.search(r"\{.*\}", text, _re.S)
        verdict = _json.loads(m.group(0)) if m else {}
        return [c["party_base"] for c in candidates
                if verdict.get(c["party_base"]) is True or verdict.get(c["party"]) is True]
    except Exception as e:
        sys.stderr.write("[bank_reco] LLM confirm skipped: %s\n" % e)
        return None


import argparse, json, sys

def main():
    ap = argparse.ArgumentParser()
    ap.add_argument("--tally", required=True)
    ap.add_argument("--bank", required=True)
    ap.add_argument("--coa", default=None)
    ap.add_argument("--output", required=True)
    ap.add_argument("--brand", default="Brand")
    ap.add_argument("--bank-ledger", default=None,
                     help="Override the bank ledger name used on 'Add to Tally' rows. "
                          "Defaults to the ledger parsed from the Tally daybook's own title "
                          "(the '<Ledger Name>  Book' header row).")
    ap.add_argument("--tolerance", type=float, default=1.0,
                     help="Amount-match tolerance in ₹ (bank vs Tally). Amounts within this "
                          "range are treated as equal. Defaults to 1.0 (matches the UI default).")
    ap.add_argument("--aggregate-config", default=None,
                     help="Path to a per-BRAND learned aggregate-reco config JSON: "
                          '{"parties": [<normalized dense-party keys>], "salary_keywords": [...]}. '
                          "Recalls parties learned in prior runs so even a single-month file "
                          "catches them. The run prints the parties it detected under "
                          "'detected_aggregate_parties' for the controller to persist per brand.")
    ap.add_argument("--llm-key", default=None,
                     help="Optional LLM key (e.g. GenSpark GSK_API_KEY) to CONFIRM a detected dense "
                          "party before it is learned. One small call per run; absent = no gate.")
    ap.add_argument("--llm-base-url", default=None,
                     help="OpenAI-compatible base URL for the LLM (e.g. GenSpark's proxy). "
                          "Absent = Anthropic native.")
    ap.add_argument("--llm-model", default="claude-haiku-4-5",
                     help="LLM model for the confirm gate (default claude-haiku-4-5).")
    a = ap.parse_args()
    agg_config = {}
    if a.aggregate_config and os.path.exists(a.aggregate_config):
        try:
            with open(a.aggregate_config) as f:
                agg_config = json.load(f) or {}
        except Exception:
            agg_config = {}
    tally = parse_tally(a.tally)
    bank = parse_bank_output(a.bank)
    opening = parse_tally_opening(a.tally)
    res = reconcile(tally, bank, amt_tol=max(a.tolerance, 0.0), agg_config=agg_config)
    # Resolve the bank-account ledger name for "Add to Tally" rows:
    # 1. explicit --bank-ledger CLI override
    # 2. the ledger parsed straight from the daybook's own title row (authoritative --
    #    the daybook literally IS that ledger's book)
    # 3. SAFE fallback: first Tally party containing "bank" but NOT "charge(s)" -- guards
    #    against an expense ledger like "Bank Charges" appearing before the real bank
    #    account and silently becoming the ledger name on every ready-to-paste row.
    # 4. last-resort generic literal
    bank_ledger = (
        a.bank_ledger
        or parse_tally_account(a.tally)
        or next((r["party"] for r in tally
                 if "bank" in r["party"].lower() and "charge" not in r["party"].lower()), None)
        or "Bank Account"
    )
    ctx = {"brand": a.brand, "bank_rows": bank, "tally_rows": tally,
           "opening": opening, "bank_path": a.bank,
           "bank_ledger": bank_ledger}
    write_workbook(res, ctx, a.output)
    results = _build_results(res)
    analytics = _build_analytics(res, ctx)
    # aggregate-reco summary for the controller/UI + the parties detected this run (to be persisted
    # per brand so future single-month files recall them).
    aggregate_reco_summary = [
        {"party": a2["party"], "party_base": a2["party_base"], "month": a2["month"],
         "kind": a2["kind"], "bank_total": a2["bank_total"], "tally_total": a2["tally_total"],
         "gap": a2["gap"], "bank_lines": a2["bank_lines"], "tally_vouchers": a2["tally_vouchers"]}
        for a2 in res.get("aggregate_reco", [])
    ]
    # LLM gate: only LEARN detected dense parties the model confirms are genuine aggregate ledgers.
    # Falls back to the full detected set if no key / any error, so learning still works without it.
    detected = res.get("detected_aggregate_parties", [])
    confirmed = detected
    if a.llm_key and detected:
        stats = {}
        for ar2 in res.get("aggregate_reco", []):
            pb = ar2["party_base"]
            if pb in detected:
                s = stats.setdefault(pb, {"party": ar2["party"], "party_base": pb,
                                          "bank_lines": 0, "tally_vouchers": 0,
                                          "bank_total": 0.0, "tally_total": 0.0})
                s["bank_lines"] += ar2["bank_lines"]; s["tally_vouchers"] += ar2["tally_vouchers"]
                s["bank_total"] += ar2["bank_total"]; s["tally_total"] += ar2["tally_total"]
        conf = _llm_confirm_aggregate_parties(list(stats.values()), a.llm_key, a.llm_model, a.llm_base_url)
        if conf is not None:
            confirmed = conf
    print(json.dumps({"counts": res["counts"],
                      "summary": {"brand": a.brand, "total_bank_rows": len(bank), "total_tally_rows": len(tally)},
                      "results": results,
                      "analytics": analytics,
                      "aggregate_reco": aggregate_reco_summary,
                      "detected_aggregate_parties": detected,
                      "confirmed_aggregate_parties": confirmed}))

if __name__ == "__main__":
    main()
