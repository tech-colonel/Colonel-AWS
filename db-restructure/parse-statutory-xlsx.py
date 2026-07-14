#!/usr/bin/env python3
"""Parse the monthly-workflow Excel trackers into JSON for the Statutory import.
Each sheet: a title row, a header row (Period|Task|Data Source|Frequency|Remarks|Field),
then alternating period-range header rows ('1st – 5th') and numbered task rows.
Output: { SHUMEE:[...], MBRANDS:[...], URBAN:[...] } — each task:
  { window, num, task, source, freq, remarks, field }
"""
import openpyxl, json, sys, re

SHUMEE = "/Users/dhavalchauhan/Downloads/Shumee_Finance_Team_Work.xlsx"
MU     = "/Users/dhavalchauhan/Downloads/M Brands & Urban Plant.xlsx"

def norm(v): return "" if v is None else str(v).strip()

def parse_sheet(ws):
    rows = [[norm(c) for c in r] for r in ws.iter_rows(values_only=True)]
    # find header row + column map
    hdr_i, cols = None, {}
    for i, r in enumerate(rows):
        low = [c.lower() for c in r]
        if 'period' in low and 'task' in low and 'field' in low:
            hdr_i = i
            for j, c in enumerate(low):
                if c in ('period', 'task', 'data source', 'frequency', 'remarks', 'field'):
                    cols[c] = j
            break
    if hdr_i is None:
        return []
    P, T = cols['period'], cols['task']
    DS, FR, RM, FD = cols.get('data source'), cols.get('frequency'), cols.get('remarks'), cols['field']
    out, window = [], ""
    for r in rows[hdr_i + 1:]:
        def g(j): return r[j] if (j is not None and j < len(r)) else ""
        pcell, tcell = g(P), g(T)
        if not tcell and pcell:          # period-range header row
            window = pcell
            continue
        if not tcell:                    # blank / spacer
            continue
        out.append({
            "window": window,
            "num": pcell,
            "task": tcell,
            "source": g(DS),
            "freq": g(FR),
            "remarks": g(RM),
            "field": g(FD) or "General",
        })
    return out

wb1 = openpyxl.load_workbook(SHUMEE, data_only=True)
wb2 = openpyxl.load_workbook(MU, data_only=True)

def find_sheet(wb, needle):
    for ws in wb.worksheets:
        if needle.lower() in ws.title.lower():
            return ws
    return wb.worksheets[0]

data = {
    "SHUMEE":  parse_sheet(wb1.worksheets[0]),
    "MBRANDS": parse_sheet(find_sheet(wb2, "m brand")),
    "URBAN":   parse_sheet(find_sheet(wb2, "urban")),
}

# report
for k, tasks in data.items():
    fields = {}
    for t in tasks:
        fields[t["field"]] = fields.get(t["field"], 0) + 1
    # duplicate (field,task) within the sheet?
    seen, dups = set(), []
    for t in tasks:
        key = (t["field"], t["task"])
        if key in seen: dups.append(key)
        seen.add(key)
    print(f"== {k}: {len(tasks)} tasks | fields={fields}")
    if dups:
        print(f"   !! {len(dups)} duplicate (field,task) pairs:")
        for d in dups[:20]: print("     ", d)

out = "/private/tmp/claude-501/-Users-dhavalchauhan-Colonel-Full/1d5248a9-e459-4980-a1d2-9af7c726ba1d/scratchpad/statutory_import.json"
with open(out, "w") as f:
    json.dump(data, f, ensure_ascii=False, indent=1)
print("\nWrote", out)
